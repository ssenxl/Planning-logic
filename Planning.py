import argparse
import os
import pandas as pd
import io
import math
import re
import sys
from datetime import date, datetime
import configparser as _cp
from pathlib import Path
from Calendar import load_calendar, calendar_week_map
import WorkDay  # ต้อง import ก่อน _base_work_hours ถูกเรียกที่ระดับ module (WorkDay.init เรียกทีหลังตอน master_mc พร้อม)


def _base_work_hours(mc_group, gauge=None) -> float:
    """ชั่วโมงฐานของกลุ่มสำหรับแปลง CAP (ฐาน 24 ชม.) → กำลังผลิตจริง
    ชั่วโมงของกลุ่มมาจากชีท Work Day (หน้า WorkDayPanel) เท่านั้น · ไม่ตั้ง = 24 ชม.
    ไม่ใช้ Working Hours. ของ MasterMC แล้ว
    หมายเหตุ: Item Special ยังชนะเสมอ — จัดการที่ adjust_daily_cap_for_item_special แยกต่างหาก"""
    return WorkDay.get_working_hours(mc_group, gauge)


# Set UTF-8 encoding for Windows console

if sys.platform == "win32":
    # ห้าม detach buffer — ตอนรันเป็น exe ทุก step อยู่ใน process เดียวกัน
    # ถ้า detach แล้ว input() ของ run_all จะพังหลัง Planning จบ
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except (AttributeError, ValueError):
        pass
# =========================
# CONFIG
# =========================
BASE_DIR = Path(sys.executable).parent if getattr(sys, 'frozen', False) else Path(__file__).parent
_cfg = _cp.ConfigParser(interpolation=None)  # ปิด interpolation ให้ใช้ %USERPROFILE% ใน path ได้
if not _cfg.read(BASE_DIR / "config.ini", encoding="utf-8"):
    raise FileNotFoundError(f"ไม่พบ config.ini ที่ {BASE_DIR / 'config.ini'} — กรุณาสร้างไฟล์ config.ini ก่อนรัน")
TODAY = pd.Timestamp.today().normalize()
DATA_PLAN_DIR = BASE_DIR / "data_plan"
DATA_DIR = BASE_DIR / "data"
ITEMCORE_DIR = DATA_DIR / "Itemcore"
ORDER_FILE = DATA_PLAN_DIR / "order_ready.xlsx"
def _latest_booking_final():
    """หาไฟล์ booking_final_ready ที่ใหม่สุด (ชื่อมีวันที่ พ.ศ.) ; fallback ชื่อเก่า"""
    cands = sorted(DATA_PLAN_DIR.glob("booking_final_ready_*.xlsx"),
                   key=lambda p: p.stat().st_mtime, reverse=True)
    if cands:
        return cands[0]
    return DATA_PLAN_DIR / "booking_final_ready25.xlsx"  # backward-compat
MC_REMAIN_FILE = _latest_booking_final()


def _load_outsource_split() -> dict:
    """อ่านการแบ่งงานไปจ้างทอที่ user ยืนยันจากหน้าเว็บ
    → { ITEM_CODE: {"outsource_qty": float, "start_week": int|None} }

    ไฟล์ค้างอยู่จนกว่า user จะลบเอง (งานที่ส่งไปจ้างทอแล้ว แผนรอบถัดไปต้องไม่ดึงกลับมาทอในบ้าน)
    รายการที่ item ไม่มีของค้างแล้วจะถูกข้ามเองตอน apply"""
    import json as _json
    if not OUTSOURCE_SPLIT_FILE.exists():
        return {}
    try:
        data = _json.loads(OUTSOURCE_SPLIT_FILE.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"[OUTSOURCE] อ่าน {OUTSOURCE_SPLIT_FILE.name} ไม่ได้: {e}")
        return {}
    out = {}
    for k, v in (data.get("items") or {}).items():
        if not isinstance(v, dict):
            continue
        try:
            qty = float(v.get("outsource_qty") or 0)
        except (TypeError, ValueError):
            continue
        if qty <= 0:
            continue
        try:
            wk = int(v.get("start_week")) if v.get("start_week") not in (None, "") else None
        except (TypeError, ValueError):
            wk = None
        out[str(k).strip().upper()] = {"outsource_qty": qty, "start_week": wk}
    return out
ITEMCORE_FILE = ITEMCORE_DIR / "Itemcore.xlsx"
# การแบ่งงานไปจ้างทอที่ user ยืนยันจากหน้าเว็บ (หน้า "จ้างทอ (AI)")
OUTSOURCE_SPLIT_FILE = DATA_PLAN_DIR / "outsource_split.json"
CALENDAR_FILE = "https://nanyangtextilegroup.sharepoint.com/:x:/s/SCM_Cloud/IQCXP4jH73zhQozDNvw1XF8OAY5m4p-UFv35Tcpza6v8mJo?e=43ffCc"
BOOKING_DIR = BASE_DIR / "Booking"
today = datetime.now()
OUTPUT_FILE = DATA_PLAN_DIR / f"production_plan_{today.day}-{today.month}-{today.year+543}_{today.hour:02d}{today.minute:02d}.xlsx"
SETUP_DAYS = 3
# โรงอ้อมน้อย (OMNOI): ทุก item ใช้ setup เท่ากันหมด ไม่ดู prefix ITEM / POLY / DTY
OMNOI_SETUP_DAYS = 3
SETUP_GAP_WEEK = 3
# Week ที่ไม่ต้องการวางแผน (เช่น สัปดาห์หยุด/ปิดโรงงาน)
SKIP_WEEKS = {}
# Allow carryover even when SC/SO changes (user option)
ALLOW_CARRYOVER_ACROSS_SO = False
# อนุญาต carry เฉพาะ FG ถัดไปของ item เดียวกันภายในสัปดาห์เดียวกัน
ALLOW_SAME_ITEM_WEEK_CARRY = True
# MC_GROUP redirect: เมื่อเลือก MC_GROUP+GAUGE นี้ให้เปลี่ยนไปใช้ MC_GROUP+GAUGE ใหม่แทนเสมอ
# key = (mc_group, gauge_str), value = (new_mc_group, new_gauge_str)
MC_GROUP_REDIRECT = {
    ("SKP", "20"): ("FA", "20"),  # SKP G20 → FA G20 (อ้อมน้อย) รับผลิตแทน
}

# Progressive machine reduction: เริ่มต้นด้วยเครื่องเยอะ แล้วค่อยๆ ลดลงให้ทัน target

USE_PROGRESSIVE_REDUCTION = True
# MAX_SETUP_MC แบบ static ถูกยกเลิก → ใช้ _dynamic_setup_limit() แทน (dynamic ตาม urgency RDD)
# จำนวนเครื่อง *ใหม่* (new setup) สูงสุดต่อ item/week — carry-over ไม่นับ
MAX_NEW_SETUP_MC = 2
# Objective: ใช้เครื่องให้มากที่สุด โดยยังจบตรง TARGET_KNIT
PREFER_FULL_MACHINE_TO_TARGET = True
# fallback: ถ้าจบตรง TARGET ไม่ได้ อนุญาต TARGET-1
ALLOW_TARGET_MINUS_ONE = False

# TARGET_KNIT offset (สัปดาห์ที่หักจาก FG Week)
# ใช้ร่วมกันทุกจุดที่คำนวณ TARGET_KNIT (วางแผนจริง / sort orders / overlay แผนเก่า)
#
# ลำดับกฎ:
#   1. FOB_TYPE ในตาราง FOB_TARGET_OFFSET_WEEKS → ใช้ค่านั้น (ชนะทุกกฎ)
#   2. MC_CAT = DOUBLE-30 และเกจ 16/18/19        → -3
#   3. Type ขึ้นต้น SINGLE (รวม SINGLE (TUBE)) และ yarn เส้นเดียว (ไม่มี '+') → -3
#   4. อื่น ๆ                                     → -4
# (LAB-DIP ไม่ผ่านทางนี้ — ใช้ TODAY_IDX + 2 แยกต่างหาก)
FOB_TARGET_OFFSET_WEEKS = {
    "PILOT_RUN": 1,
    "Salesman": 1,
    "Salesman-PO": 1,
    "Sample": 1,
    # FOB_TYPE อื่น (Replacement SO / Make to Order / RESERVOIR-GF ฯลฯ) ไม่ระบุที่นี่
    # → ตกไปตัดสินด้วยกฎเครื่อง/เส้นด้ายด้านล่าง (-3 หรือ -4)
}
DEFAULT_TARGET_OFFSET_WEEKS = 4  # กรณีทั่วไป → N-4
SHORT_TARGET_OFFSET_WEEKS = 3  # DOUBLE-30 เกจ 16/18/19 และ SINGLE yarn เส้นเดียว → N-3
DOUBLE30_SHORT_GAUGES = {"16", "18", "19"}


def target_offset_weeks(
    fob_type, mc_group="", gauge="", yarn_used="", item_code=""
) -> int:
    """คืนจำนวนสัปดาห์ที่ต้องหักจาก FG Week เพื่อได้ TARGET_KNIT"""
    _fob = FOB_TARGET_OFFSET_WEEKS.get(str(fob_type).strip())
    if _fob is not None:
        return _fob

    _mc_u = str(mc_group).strip().upper() if mc_group else ""
    if _mc_u and _mc_u not in ("NAN", "NONE"):
        _g = _normalize_gauge(gauge)
        _cat = str(_mc_to_type1(_mc_u, gauge)).strip().upper()

        # DOUBLE-30 เกจ 16/18/19
        if _cat == "DOUBLE-30" and _g in DOUBLE30_SHORT_GAUGES:
            return SHORT_TARGET_OFFSET_WEEKS

        # SINGLE ทุกเกจ/ทุกโรง (รวม SINGLE (TUBE) และสะกด SINGEL) ที่ใช้ yarn เส้นเดียว
        _type = str(MC_TYPE_MAP.get(_mc_u, "")).strip().upper().replace("SINGEL", "SINGLE")
        _is_single = _type.startswith("SINGLE") or _cat.replace("SINGEL", "SINGLE").startswith("SINGLE")
        if _is_single:
            _yarn = str(yarn_used).strip()
            if not _yarn or _yarn.upper() in ("NAN", "NONE"):
                _yarn = str(
                    _yarn_used_lookup.get(str(item_code).strip().upper(), "")
                ).strip()
            if _yarn and _yarn.upper() not in ("NAN", "NONE") and "+" not in _yarn:
                return SHORT_TARGET_OFFSET_WEEKS

    return DEFAULT_TARGET_OFFSET_WEEKS


def target_offset_weeks_for_row(row) -> int:
    """เรียก target_offset_weeks() จากแถว order (ดึงฟิลด์ให้ครบในที่เดียว)"""
    return target_offset_weeks(
        row.get("FOB_TYPE", ""),
        mc_group=row.get("MC GROUP", "") or row.get("MC_GROUP", ""),
        gauge=row.get("MC_GUAGE", "") or row.get("GUAGE", ""),
        yarn_used=(
            row.get("YARN-USED", "")
            or row.get("YARN_USED", "")
            or row.get("YARN_ITEM", "")
        ),
        item_code=row.get("Item Code", "") or row.get("ITEM_CODE", ""),
    )
# Load Balancing Configuration
USE_LOAD_BALANCING = True  # เปิดใช้งาน Load Balancing สำหรับการกระจายงาน
LOAD_BALANCING_WEIGHT = 0.1  # น้ำหนักสำหรับการพิจารณาเครื่องว่างใน load balancing score
LOAD_BALANCING_THRESHOLD = 0.2  # เกณฑ์ที่ใช้พิจารณาว่าโหลดไม่สมดุล (20% จากค่าเฉลี่ย)



# =========================
# SHARED MACHINE POOL
# กลุ่ม MC ที่ใช้เครื่องร่วมกัน — ต้องตรงกับ AVA_MC.py
# =========================
# โหลด SHARED_POOL_MAP แบบ dynamic จาก MasterMC
_MASTER_MC_PATH = os.path.expandvars(_cfg["paths"]["master_mc"])



# =========================
# LOAD DATA
orders = pd.read_excel(ORDER_FILE)
# Map column names from order_ready.xlsx to expected names
column_mapping = {
    'SC_NO': 'SC/SO NO',
    'FABRIC_ITEM': 'Item Code', 
    'ORDERS_QTY': 'Orders.Qty',
    'PLAN_QTY': 'Plan Qty',
    'PENDING_PLAN': 'Pending Plan',
    'FGWEEK': 'FG Week',
    'CUSTOMER_NAME': 'Customer',
    'ORDER_TYPE': 'Orders Type',
    'NAY_COLOR': 'NAY_COLOR',
    'COLOR_DESC': 'COLOR_DESC'
}

orders = orders.rename(columns=column_mapping)

# แถวที่ Order.py กรองออกก่อนวางแผน (ORDER_TYPE/MC_GROUP ต้องห้าม) — ไม่เคยเข้า planning loop
# ต้องเอามาโชว์ในชีท UNPLANNED ไม่งั้นงานพวกนี้หายเงียบทั้งที่ยังมีของค้าง
# rename ด้วย column_mapping ชุดเดียวกับ orders เพื่อให้คอลัมน์ตรงกันตอนรวมเป็น DataFrame เดียว
_EXCLUDED_ORDER_FILE = DATA_PLAN_DIR / "order_excluded.xlsx"
_excluded_orders = pd.DataFrame()
try:
    if _EXCLUDED_ORDER_FILE.exists():
        _excluded_orders = pd.read_excel(_EXCLUDED_ORDER_FILE)
        _excluded_orders.columns = _excluded_orders.columns.str.strip()
        _excluded_orders = _excluded_orders.rename(columns=column_mapping)
        print(f"📄 order ที่ถูกกรองออกก่อนวางแผน: {len(_excluded_orders)} แถว")
except Exception as _e_exc:
    print(f"⚠️ อ่าน {_EXCLUDED_ORDER_FILE.name} ไม่ได้: {_e_exc}")
    _excluded_orders = pd.DataFrame()
summary_mc = pd.read_excel(MC_REMAIN_FILE, sheet_name="SUMMARY_MC_REMAIN")
detail_mc = pd.read_excel(MC_REMAIN_FILE, sheet_name="DETAIL")  # โหลด DETAIL
fd6_check = detail_mc[detail_mc["ITEM_CODE"].astype(str).str.upper().str.strip() == "FD6GNTLG27/58A0"]
if not fd6_check.empty:

    print(f"[DEBUG] FD6GNTLG27/58A0 found in detail_mc: {len(fd6_check)} rows")
    print(fd6_check[["ITEM_CODE", "WEEK", "MC_GROUP", "GUAGE", "MC_USE_CEIL"]])
else:

    print(f"[DEBUG] FD6GNTLG27/58A0 NOT found in detail_mc")
item_cap_data = pd.DataFrame(columns=["ITEM_CODE", "MC_GROUP", "CAP ทอ", "REVOLUTION/WEIGHT", "GUAGE"])
master_mc = pd.read_excel(_MASTER_MC_PATH)
# โหลด Itemcore สำหรับเช็ค RTS items
# อ่านทะเบียน Core Item จากไฟล์ Target_Stock (คอลัมน์ Item code + Customer) ซึ่ง user แก้บน server ได้
# → ตัดปัญหาไฟล์ Itemcore.xlsx แยกที่ค้างของเก่า (data drift). ถ้าอ่านไม่ได้ค่อย fallback ไฟล์เดิม
try:

    try:
        _itemcore_src = os.path.expandvars(_cfg["paths"]["target_stock"])
    except KeyError:
        _itemcore_src = None

    if _itemcore_src and Path(_itemcore_src).exists():
        itemcore_df = pd.read_excel(_itemcore_src)
    else:
        itemcore_df = pd.read_excel(ITEMCORE_FILE)

    itemcore_df.columns = itemcore_df.columns.str.strip()

except Exception as e:

    print(f"⚠️ ไม่สามารถโหลด Itemcore: {e}")

    itemcore_df = pd.DataFrame()

calendar = load_calendar(CALENDAR_FILE, sheet_name="Sheet1")

calendar_week = calendar_week_map(calendar)

orders.columns = orders.columns.str.strip()

summary_mc.columns = summary_mc.columns.str.strip().str.upper()

if "MC_CAT" in summary_mc.columns and "TYPE_1" not in summary_mc.columns:
    summary_mc = summary_mc.rename(columns={"MC_CAT": "TYPE_1"})

summary_mc["TYPE_1"] = summary_mc["TYPE_1"].astype(str).str.strip().str.upper()

calendar_week.columns = calendar_week.columns.str.strip().str.upper()

detail_mc.columns = detail_mc.columns.str.strip().str.upper()  # เพิ่ม detail_mc

master_mc.columns = master_mc.columns.str.strip()

if "MC_CAT" in master_mc.columns and "Type_1" not in master_mc.columns:
    master_mc = master_mc.rename(columns={"MC_CAT": "Type_1"})



# MC → Type_1 lookup: (MC_upper, guage_str) → Type_1

_MC_TYPE1_MAP: dict = {}

for _, _mrow in master_mc.iterrows():

    _mmc = str(_mrow.get("MC", "")).strip().upper()

    _mg = str(_mrow.get("Guage", "")).strip()

    _mt1 = str(_mrow.get("Type_1", "")).strip().upper()

    if _mmc and _mt1:

        _MC_TYPE1_MAP[(_mmc, _mg)] = _mt1

        if (_mmc, "") not in _MC_TYPE1_MAP:

            _MC_TYPE1_MAP[(_mmc, "")] = _mt1

print(f"✅ MC→Type_1 map: {len(_MC_TYPE1_MAP)} entries")

# MC → Factory map: (mc_upper, guage_str) → factory_upper
_mc_factory_map: dict = {}
for _, _mrow in master_mc.iterrows():
    _mmc = str(_mrow.get("MC", "")).strip().upper()
    _mg = str(_mrow.get("Guage", "")).strip()
    _mf = str(_mrow.get("Factory", "")).strip().upper()
    if _mmc and _mf:
        _mc_factory_map[(_mmc, _mg)] = _mf
        if (_mmc, "") not in _mc_factory_map:
            _mc_factory_map[(_mmc, "")] = _mf
print(f"✅ MC→Factory map: {len(_mc_factory_map)} entries")

# MC → Setup time map: mc_upper → setup_days (จาก MasterMC column "Set up time")
_mc_setup_time_map: dict = {}
for _, _mrow in master_mc.iterrows():
    _mmc = str(_mrow.get("MC", "")).strip().upper()
    _st = _mrow.get("Set up time", None)
    if _mmc and _st is not None and str(_st).strip() not in ("", "nan", "NAN", "NaT", "None"):
        try:
            _mc_setup_time_map[_mmc] = int(float(_st))
        except (ValueError, TypeError):
            pass
print(f"✅ MC→SetupTime map: {len(_mc_setup_time_map)} entries")

# Item prefix → Setup days map (จากชีท "Item Setup" ใน MasterMC) — ใช้เฉพาะ Type SINGLE
# format: prefix_upper → setup_days  (เช่น "FD1" → 2)  จับคู่แบบ longest-prefix match
# ช่อง ITEM อาจมีหลาย prefix คั่นด้วย comma (เช่น "F1 , FD1") → แยกเป็นหลาย key
_item_prefix_setup_map: dict = {}
try:
    _isetup_df = pd.read_excel(_MASTER_MC_PATH, sheet_name="Item Setup")
    _isetup_df.columns = _isetup_df.columns.str.strip()
    for _, _isrow in _isetup_df.iterrows():
        _iprefix_raw = str(_isrow.get("ITEM", "")).strip()
        _iday = _isrow.get("Set up (day)", None)
        if not _iprefix_raw or _iprefix_raw.upper() in ("", "NAN", "NONE"):
            continue
        if _iday is None or str(_iday).strip() in ("", "nan", "NAN", "NaT", "None"):
            continue  # prefix ว่าง (เช่น F9) → ข้าม ให้ตกไปใช้ default
        try:
            _iday_int = int(float(_iday))
        except (ValueError, TypeError):
            continue
        for _pfx in _iprefix_raw.split(","):
            _pfx = _pfx.strip().upper()
            if _pfx:
                _item_prefix_setup_map[_pfx] = _iday_int
    print(f"✅ Item prefix→Setup map: {len(_item_prefix_setup_map)} entries")
except Exception as _e_isetup:
    print(f"⚠️ ไม่สามารถโหลดชีท Item Setup: {_e_isetup}")

# Item → Team map (จากชีท "Program" ใน MasterMC) — user กำหนดเองว่า item ไหนอยู่ทีมไหน
# ใช้เติมคอลัมน์ TEAM ในไฟล์แผน เพื่อให้หน้าเว็บไฮไลต์ (ITEM+TEAM ต้องตรงคู่กัน) ได้
# item เดียวมีได้หลายทีม → เก็บรวมคั่นด้วย " , " (หน้าเว็บแยกเทียบทีละทีม)
_item_program_map: dict = {}
try:
    _prog_df = pd.read_excel(_MASTER_MC_PATH, sheet_name="Program")
    _prog_df.columns = _prog_df.columns.str.strip()
    for _, _prow in _prog_df.iterrows():
        _pitem = str(_prow.get("ITEM_CODE", "")).strip().upper()
        _pteam = str(_prow.get("TEAM_NAME", "")).strip()
        if not _pitem or _pitem in ("NAN", "NONE"):
            continue
        if not _pteam or _pteam.upper() in ("NAN", "NONE"):
            continue
        _teams = _item_program_map.setdefault(_pitem, [])
        if _pteam not in _teams:
            _teams.append(_pteam)
    print(f"✅ Item→Program team map: {len(_item_program_map)} items")
except Exception as _e_prog:
    print(f"⚠️ ไม่สามารถโหลดชีท Program: {_e_prog}")


def _lookup_item_team(item_code) -> str:
    """คืนทีม (จากชีท Program) ของ item — ไม่มีในชีท = คืนค่าว่าง"""
    _ic = str(item_code).strip().upper()
    return " , ".join(_item_program_map.get(_ic, []))


def _lookup_item_prefix_setup(item_code: str):
    """คืน setup days จาก prefix ของ item (longest-prefix match) หรือ None ถ้าไม่ match"""
    _ic = str(item_code).strip().upper()
    if not _ic or not _item_prefix_setup_map:
        return None
    # longest prefix ก่อน (เช่น "FDC" ต้องมาก่อน "FC"/"FD")
    for _pfx in sorted(_item_prefix_setup_map, key=len, reverse=True):
        if _ic.startswith(_pfx):
            return _item_prefix_setup_map[_pfx]
    return None

# Spare cylinder map: (factory_upper, mc_cat_upper, guage_str) → total_spare
_spare_cylinder_map: dict = {}
try:
    _spare_df = pd.read_excel(_MASTER_MC_PATH, sheet_name="Spare part")
    _spare_df.columns = _spare_df.columns.str.strip()
    for _, _srow in _spare_df.iterrows():
        _sf = str(_srow.get("Factory", "")).strip().upper()
        _scat_full = str(_srow.get("MC_CAT", "")).strip().upper()  # e.g. "SINGLE-32", "DOUBLE-30"
        _sg_raw = str(_srow.get("Guage", "")).strip()
        # normalize gauge: ลบ "G"/"GAUGE" prefix (ตรงกับ _normalize_gauge)
        _sg = _sg_raw.upper().replace("GAUGE", "").replace("G", "").strip()
        _st = int(_srow.get("Total Spare", 0) or 0)
        if not (_sf and _scat_full and _sg):
            continue
        # เก็บ key เต็ม: ('PHET', 'SINGLE-32', '36')
        _spare_cylinder_map[(_sf, _scat_full, _sg)] = _st
        # เก็บ key ย่อ (ตัด gauge suffix ออก): ('PHET', 'SINGLE', '36')
        # เพราะ TYPE_1 ใน MasterMC/summary_mc อาจใช้แค่ SINGLE/DOUBLE ไม่มี -NN suffix
        _scat_base = _scat_full.split("-")[0].strip()  # "SINGLE-32" → "SINGLE"
        if _scat_base != _scat_full:
            _base_key = (_sf, _scat_base, _sg)
            # รวม spare (ถ้ามีหลาย gauge class ใน factory เดียวกัน ใช้ค่าสูงสุด ไม่ใช่ sum)
            _spare_cylinder_map[_base_key] = max(_spare_cylinder_map.get(_base_key, 0), _st)
    print(f"✅ Spare cylinder map: {len(_spare_cylinder_map)} entries")
    for _k, _v in _spare_cylinder_map.items():
        print(f"   Spare: {_k} = {_v}")
except Exception as _e_spare:
    print(f"⚠️ ไม่สามารถโหลด Spare part: {_e_spare}")

# Job sheet: cylinder changes ที่เกิดขึ้นจริงแล้วในแต่ละ week (จาก MasterMC)
# format: {week_int: {group: count}}  group = 'OMNOI' / 'PHET_SINGLE' / 'PHET_DOUBLE'
_job_cylinder_done: dict = {}
try:
    _job_df = pd.read_excel(_MASTER_MC_PATH, sheet_name="Job ", header=None)
    _job_header = _job_df.iloc[0]
    for _, _jrow in _job_df.iloc[1:].iterrows():
        _jgroup = str(_jrow.iloc[0]).strip()
        if not _jgroup or _jgroup == "nan":
            continue
        for _jcol in range(1, len(_job_header)):
            _jweek_val = _job_header.iloc[_jcol]
            _jcount = _jrow.iloc[_jcol]
            if pd.isna(_jweek_val) or pd.isna(_jcount):
                continue
            _jweek = int(_jweek_val)
            _job_cylinder_done.setdefault(_jweek, {})[_jgroup] = int(_jcount)
    print(f"✅ Job cylinder done: {_job_cylinder_done}")
except Exception as _e_job:
    print(f"⚠️ ไม่สามารถโหลด Job sheet: {_e_job}")

# Lock_MC: เครื่องที่ถูก lock ไม่ให้ใช้วางแผน per (week, mc_cat, gauge)
# Source: MasterMC.xlsx sheet "Lock_MC"
# - _LOCK_MC_MAP: หักจำนวนเครื่องออกจาก pool (กันคนอื่นใช้) — ทำงานเสมอไม่ว่าจะระบุ item หรือไม่
# - _LOCK_ITEM_WARM: ถ้าระบุคอลัมน์ ITEM = "กันเครื่องให้ item นี้ + คงความอุ่น"
#   ทำให้พอ item กลับมาผลิตหลังเว้นช่วง ระบบถือว่าเครื่องยังอุ่น → ไม่ setup ใหม่
#   ⚠️ ต้องใช้กติกาเดียวกับ _LOCK_WARM ใน AVA_MC.py (ฝั่ง OLD) ไม่งั้น NEW/OLD ให้ผลคนละแบบ:
#     - จำนวนที่กัน = จำนวนเครื่องที่อุ่นได้ (กัน 3 → อุ่น 3, ตัวที่ 4 เป็นเครื่องใหม่ต้อง setup)
#     - คอลัมน์ MC: ระบุ = อุ่นเฉพาะกลุ่มเครื่องนั้น, เว้นว่าง = อุ่นทุกกลุ่มใน MC_CAT
_LOCK_MC_MAP: dict = {}
_LOCK_ITEM_WARM: dict = {}  # key=(item_upper, cat_upper, gauge_norm, mc_upper|"") → {week: locked_count}
try:
    _lock_mc_df = pd.read_excel(_MASTER_MC_PATH, sheet_name="Lock_MC")
    _lock_mc_df.columns = _lock_mc_df.columns.str.strip()
    def _safe_int(v):
        # ทนค่าว่าง/NaN/ทศนิยม/สตริง — คืน None ถ้าแปลงไม่ได้ (กันทั้งชีทพังเพราะแถวเดียว)
        try:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return None
            s = str(v).strip()
            if s == "" or s.lower() == "nan":
                return None
            return int(float(s))
        except (ValueError, TypeError):
            return None
    for _, _lrow in _lock_mc_df.iterrows():
        _lweek = _safe_int(_lrow.get("Week"))
        _lcat = str(_lrow.get("MC_CAT", "")).strip().upper()
        _lg_raw = str(_lrow.get("Guage", "")).strip()
        _lg = _lg_raw.upper().replace("GAUGE", "").replace("G", "").strip()
        _lcount = _safe_int(_lrow.get("MC Lock")) or 0
        if _lweek is not None and _lcat and _lg and _lcount > 0:
            _k = (_lweek, _lcat, _lg)
            _LOCK_MC_MAP[_k] = _LOCK_MC_MAP.get(_k, 0) + _lcount
        # ITEM (ไม่บังคับ): กันเครื่องให้ item นี้คงความอุ่นในสัปดาห์ที่ระบุ
        _litem = str(_lrow.get("ITEM", _lrow.get("ITEM_CODE", "")) or "").strip().upper()
        if _litem and _litem != "NAN" and _lweek is not None and _lcount > 0:
            _lmc = str(_lrow.get("MC", "") or "").strip().upper().replace("NAN", "")
            _lk = (_litem, _lcat, _lg, _lmc)
            _lslot = _LOCK_ITEM_WARM.setdefault(_lk, {})
            # คีย์ซ้ำ (สัปดาห์เดียวกัน) → บวกรวม ให้ตรงกับที่หัก pool และที่การ์ดในเว็บเตือนไว้
            _lslot[_lweek] = _lslot.get(_lweek, 0) + _lcount
    print(f"✅ Lock_MC map: {_LOCK_MC_MAP}")
    if _LOCK_ITEM_WARM:
        print(f"✅ Lock_MC item-warm: {_LOCK_ITEM_WARM}")
except Exception as _e_lock:
    print(f"⚠️ ไม่สามารถโหลด Lock_MC sheet: {_e_lock}")


# MC_Total: จำนวนเครื่องทั้งหมดของ (สัปดาห์ + MC_CAT + เกจ) ที่ user กำหนดเอง
# Source: MasterMC.xlsx sheet "MC_Total" — AVA_MC เขียนค่านี้ลง SUMMARY_MC_REMAIN ให้แล้ว
# ที่นี่ใช้เฉพาะ "เส้นทาง fallback" ของ _get_available_mc (สัปดาห์ที่ไม่มีแถวใน summary
# → เดิมรวม Total MC จาก MasterMC ซึ่งไม่รู้จัก override) ไม่งั้นสัปดาห์นั้นจะได้ยอดเก่า
_MC_TOTAL_OVERRIDE: dict = {}   # (week, cat_upper, gauge_norm) → total (รวมทุกโรงงาน)
try:
    _mt_df_plan = pd.read_excel(_MASTER_MC_PATH, sheet_name="MC_Total")
    _mt_df_plan.columns = _mt_df_plan.columns.str.strip()

    def _mt_int_plan(v):
        # ทนค่าว่าง/NaN/ทศนิยม/สตริง (นิยามซ้ำ ไม่พึ่ง _safe_int เผื่อบล็อก Lock_MC พังก่อน)
        try:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return None
            s = str(v).strip()
            if not s or s.lower() == "nan":
                return None
            return int(float(s))
        except (ValueError, TypeError):
            return None

    _mt_by_fac: dict = {}   # (week, cat, gauge) → ผลรวมของแถวที่ระบุโรงงาน
    for _, _mt_r in _mt_df_plan.iterrows():
        _mt_w = _mt_int_plan(_mt_r.get("Week"))
        _mt_tot = _mt_int_plan(_mt_r.get("Total MC"))
        _mt_c = str(_mt_r.get("MC_CAT", "") or "").strip().upper()
        _mt_g_raw = str(_mt_r.get("Guage", "") or "").strip()
        _mt_g = _mt_g_raw.upper().replace("GAUGE", "").replace("G", "").strip()
        if _mt_w is None or _mt_tot is None or _mt_tot < 0 or not _mt_c or not _mt_g:
            continue
        _mt_fac = str(_mt_r.get("Factory", "") or "").strip().upper().replace("NAN", "")
        # Week To (ไม่บังคับ): ระบุช่วงสัปดาห์ที่ใช้ยอดเดียวกัน — ไม่มี/น้อยกว่า Week = สัปดาห์เดียว
        _mt_w_to = _mt_int_plan(_mt_r.get("Week To"))
        _mt_w_end = _mt_w_to if (_mt_w_to is not None and _mt_w_to >= _mt_w) else _mt_w
        for _mt_wk in range(_mt_w, _mt_w_end + 1):
            _mt_k = (_mt_wk, _mt_c, _mt_g)
            if _mt_fac:
                # ระบุโรงงาน = ยอดของโรงงานนั้น → บวกรวมกันเป็นยอดของ CAT+เกจ
                _mt_by_fac[_mt_k] = _mt_by_fac.get(_mt_k, 0) + _mt_tot
            else:
                # ไม่ระบุ = ยอดรวมทั้ง CAT+เกจ อยู่แล้ว → ใช้ค่านี้เลย (ชนะแถวที่ระบุโรงงาน)
                _MC_TOTAL_OVERRIDE[_mt_k] = _mt_tot
    for _mt_k, _mt_v in _mt_by_fac.items():
        _MC_TOTAL_OVERRIDE.setdefault(_mt_k, _mt_v)
    if _MC_TOTAL_OVERRIDE:
        print(f"✅ MC_Total override: {_MC_TOTAL_OVERRIDE}")
except Exception as _e_mt_plan:
    print(f"⚠️ ไม่สามารถโหลด MC_Total sheet: {_e_mt_plan}")


def _lock_warm_weeks_plan(item, mc_group, gauge) -> dict:
    """{week: จำนวนเครื่องที่ lock ไว้ให้ item นี้} — รวม lock ที่ระบุ MC ตรงกับที่เว้นว่าง

    กติกาเดียวกับ _lock_warm_weeks() ใน AVA_MC.py (เว้นว่าง = ครอบคลุมทุกกลุ่มใน MC_CAT
    → ถ้ามีทั้งสองแบบใช้ค่าที่มากกว่า ไม่บวกกัน เพราะอ้างถึงเครื่องชุดเดียวกัน)
    """
    if not _LOCK_ITEM_WARM:
        return {}
    _item = str(item).strip().upper()
    _mc = str(mc_group).strip().upper()
    _cat = _mc_to_type1(_mc, gauge)
    _g = _normalize_gauge(gauge)
    out: dict = {}
    for _k in ((_item, _cat, _g, _mc), (_item, _cat, _g, "")):
        for _w, _c in _LOCK_ITEM_WARM.get(_k, {}).items():
            out[_w] = max(out.get(_w, 0), _c)
    return out


def _bridge_locks(key, current_idx):
    """เลื่อน last_production[key] มาที่สัปดาห์ที่ Lock กันเครื่องให้ item นี้ (ถ้ามี)
    เพื่อให้ทุกจุดเช็ค gap (setup) ถือว่าเครื่องยังอุ่น → ไม่ setup ใหม่ตอน item กลับมาผลิต

    เงื่อนไข: item ต้องเคยผลิตมาก่อน (key อยู่ใน last_production) จึงจะ "อุ่นต่อ" ได้
    (ครั้งแรกสุดยังต้อง setup ตามปกติ) และเลื่อนได้เฉพาะสัปดาห์ lock ที่อยู่ระหว่าง
    สัปดาห์ผลิตล่าสุด..สัปดาห์ปัจจุบัน — ไม่ล้ำอนาคต

    เดินเป็นทอด ๆ ทีละ lock โดยแต่ละช่วงห้ามเกิน SETUP_GAP_WEEK (เครื่องเย็นระหว่างทางไม่ได้)
    — เดิมกระโดดไป lock ที่ไกลสุดเลยโดยไม่เช็คว่าช่วงก่อนหน้าขาดตอนหรือเปล่า

    คืนค่า: จำนวนเครื่องที่อุ่นได้ (= จำนวนที่กันไว้ แคบสุดตลอดเส้นทาง) หรือ None ถ้าไม่มี
    lock มาเกี่ยว → ผู้เรียกต้องเอาไปจำกัดจำนวนเครื่อง carry-over ด้วย
    """
    if not _LOCK_ITEM_WARM or current_idx is None:
        return None
    if not isinstance(key, tuple) or not key or key not in last_production:
        return None
    weeks = _lock_warm_weeks_plan(
        key[0],
        key[1] if len(key) > 1 else "",
        key[2] if len(key) > 2 else "",
    )
    if not weeks:
        return None
    # แปลงสัปดาห์ → index ก่อน (last_production เก็บเป็น index ไม่ใช่เลขสัปดาห์)
    _by_idx: dict = {}
    for _w, _c in weeks.items():
        _wi = week_index(_w)
        if _wi is not None:
            _by_idx[_wi] = max(_by_idx.get(_wi, 0), _c)
    cur = last_production[key]
    start = cur
    cap = None
    for _wi in sorted(_by_idx):
        if cur < _wi < current_idx and (_wi - cur) <= SETUP_GAP_WEEK:
            cur = _wi
            cap = _by_idx[_wi] if cap is None else min(cap, _by_idx[_wi])
    if cur > start:
        last_production[key] = cur
        print(
            f"[LOCK WARM] {key[0]}: คงเครื่องอุ่นถึง idx {cur} (จาก Lock_MC, อุ่นได้ {cap} เครื่อง)"
            f" → ไม่ setup ใหม่ที่ idx {current_idx}"
        )
    return cap


def _same_week_lock_mc(key, prod_idx) -> int:
    """จำนวนเครื่องที่ Lock_MC กันไว้ให้ item ใน "สัปดาห์ที่ item ผลิตเอง" (idx = prod_idx)

    ต่างจาก _bridge_locks ที่ใช้ล็อกเชื่อมช่วงที่ item หยุดผลิต — ล็อกที่ทับสัปดาห์ผลิต
    เป็นเครื่องอีกชุดที่กันไว้เพิ่มจากเครื่องที่รันอยู่ (_LOCK_MC_MAP หักออกจาก pool
    ไปแล้วโดยไม่สนว่าสัปดาห์นั้นผลิตหรือไม่) → เครื่องชุดนี้ยังอุ่น ต้องบวกเข้ากับ
    carry-over ไม่ใช่ปล่อยหายแล้วให้สัปดาห์ถัดไป setup ใหม่

    กติกาเดียวกับคอลัมน์ _lock_same_week_mc ใน AVA_MC.py (ฝั่ง OLD)
    """
    if not _LOCK_ITEM_WARM or prod_idx is None:
        return 0
    if not isinstance(key, tuple) or not key:
        return 0
    weeks = _lock_warm_weeks_plan(
        key[0],
        key[1] if len(key) > 1 else "",
        key[2] if len(key) > 2 else "",
    )
    for _w, _c in weeks.items():
        if week_index(_w) == prod_idx:
            return int(_c or 0)
    return 0

# MC Special: per-(Factory, MC_CAT, MC, Gauge) → POLY/COTTON machine count
# Source: MasterMC.xlsx sheet "MC Special"
_PLANS_REF: list = []  # อ้างถึง list แถวแผนที่กำลังสร้างใน _run_planning_loop() (ใช้นับเครื่อง sub-pool จริง)
_MC_SPECIAL_PLAN: dict = {}  # key=(factory_upper, mc_cat_upper, mc_upper, gauge_str) → {"POLY": int, "COTTON": int}
_mc_special_weekly_usage: dict = {}  # (factory, mc_cat, gauge, week, "COTTON"/"POLY") → machines used by new plan
try:
    _mcs_plan_df = pd.read_excel(_MASTER_MC_PATH, sheet_name="MC Special")
    _mcs_plan_df.columns = _mcs_plan_df.columns.str.strip()
    for _, _mcs_row in _mcs_plan_df.iterrows():
        _mcs_f = str(_mcs_row.get("Factory", "")).strip().upper()
        _mcs_cat = str(_mcs_row.get("MC_CAT", "")).strip().upper()
        _mcs_mc_raw = str(_mcs_row.get("MC", "")).strip().upper()
        _mcs_g = str(_mcs_row.get("Guage", "")).strip().upper().replace("GAUGE", "").replace("G", "").strip()
        # ตัด ".0" ท้าย ให้ตรงกับ gauge_str ที่ผ่าน _normalize_gauge (float 36.0 → "36")
        if _mcs_g.endswith(".0"):
            _mcs_g = _mcs_g[:-2]
        if not _mcs_f or not _mcs_cat or not _mcs_g:
            continue
        _mcs_poly_raw = _mcs_row.get("POLY")
        _mcs_cotton_raw = _mcs_row.get("COTTON")
        _mcs_poly = int(float(_mcs_poly_raw)) if pd.notna(_mcs_poly_raw) and str(_mcs_poly_raw).strip() not in ("", "-", "NAN") else 0
        _mcs_cotton = int(float(_mcs_cotton_raw)) if pd.notna(_mcs_cotton_raw) and str(_mcs_cotton_raw).strip() not in ("", "-", "NAN") else 0
        if _mcs_poly <= 0 and _mcs_cotton <= 0:
            continue
        # MC column อาจมีหลายค่าคั่นด้วย comma เช่น "SKPTA,SKPLE"
        _mcs_mc_parts = [p.strip() for p in _mcs_mc_raw.split(",") if p.strip() and p.strip() not in ("NAN", "")]
        if not _mcs_mc_parts:
            _mcs_mc_parts = [""]  # general match (ไม่ระบุ MC เฉพาะ)
        for _mcs_mc in _mcs_mc_parts:
            _MC_SPECIAL_PLAN[(_mcs_f, _mcs_cat, _mcs_mc, _mcs_g)] = {"POLY": _mcs_poly, "COTTON": _mcs_cotton}
    print(f"✅ MC Special (Plan): {len(_MC_SPECIAL_PLAN)} entries loaded")
    for _k, _v in _MC_SPECIAL_PLAN.items():
        print(f"   {_k} → {_v}")
except Exception as _e_mcs:
    print(f"⚠️ ไม่สามารถโหลด MC Special: {_e_mcs}")
    _MC_SPECIAL_PLAN = {}

# Booking ที่กินเครื่อง POLY/COTTON sub-pool ไปแล้ว ต่อ (cat, gauge, week) — อ่านจาก DETAIL
# ของ booking_final (POOL_TYPE ลงท้าย :POLY/:COTTON) เพื่อหักจาก capacity ก่อนวางงานใหม่
# ไม่งั้นจะวางงาน POLY/COTTON ทับเครื่องที่ booking จองเต็มแล้ว (over-plan)
_MC_SPECIAL_BOOKING_USED: dict = {}  # (cat_upper, gauge_str, week, "POLY"/"COTTON") → machines
try:
    if "POOL_TYPE" in detail_mc.columns:
        def _bku_gauge(g):
            try:
                return str(int(float(g)))
            except (TypeError, ValueError):
                return str(g).strip()
        for _, _bk_r in detail_mc.iterrows():
            _bk_pt = str(_bk_r.get("POOL_TYPE", "")).strip()
            if _bk_pt.endswith(":POLY"):
                _bk_type = "POLY"
            elif _bk_pt.endswith(":COTTON"):
                _bk_type = "COTTON"
            else:
                continue
            _bk_cat = str(_bk_r.get("MC_CAT", "")).strip().upper()
            _bk_g = _bku_gauge(_bk_r.get("GUAGE", ""))
            try:
                _bk_wk = int(_bk_r.get("WEEK"))
            except (TypeError, ValueError):
                continue
            _bk_key = (_bk_cat, _bk_g, _bk_wk, _bk_type)
            _MC_SPECIAL_BOOKING_USED[_bk_key] = (
                _MC_SPECIAL_BOOKING_USED.get(_bk_key, 0) + int(_bk_r.get("MC_USE_CEIL", 0) or 0)
            )
    print(f"✅ MC Special booking used: {len(_MC_SPECIAL_BOOKING_USED)} (cat,gauge,week,type)")
except Exception as _e_bku:
    print(f"⚠️ ไม่สามารถคำนวณ MC Special booking used: {_e_bku}")
    _MC_SPECIAL_BOOKING_USED = {}

# TYPE_SPECIAL quota (BABY FRENCH / SINGLE JACQUARD / TWILL)
def _is_description_special_type_plan(desc: str, keywords: list) -> bool:
    """FRENCH TERRY → False เสมอ; ถ้า match keyword ใดๆ → True"""
    d = str(desc).strip().upper()
    if "FRENCH TERRY" in d:
        return False
    return any(kw.lstrip("$") in d for kw in keywords)

_TYPE_DESC_RULES_PLAN: dict = {}  # (factory_upper, type_upper) → {max_mc, keywords, mc_cat}
try:
    _tdr_df = pd.read_excel(_MASTER_MC_PATH, sheet_name="MC Special")
    _tdr_df.columns = _tdr_df.columns.str.strip()
    _kw_col_p = "Unnamed: 8" if "Unnamed: 8" in _tdr_df.columns else None
    _cur_key_p, _cur_kws_p, _cur_max_p, _cur_cat_p = None, [], 0, ""
    for _, _r in _tdr_df.iterrows():
        _fac2 = str(_r.get("Factory", "")).strip().upper()
        _typ2 = str(_r.get("Type",    "")).strip().upper()
        _cat2 = str(_r.get("MC_CAT",  "")).strip().upper()
        if _cat2 == "NAN": _cat2 = ""
        _desc2 = _r.get("DESCRIPTION")
        _kw2   = str(_r.get(_kw_col_p, "") if _kw_col_p else "").strip().upper()
        _has_g2 = pd.notna(_r.get("Guage")) and str(_r.get("Guage", "")).strip() not in ("", "-", "NAN")
        if _fac2 not in ("", "NAN") and _typ2 not in ("", "NAN") and not _has_g2:
            if _cur_key_p and _cur_kws_p:
                _TYPE_DESC_RULES_PLAN[_cur_key_p] = {"max_mc": _cur_max_p, "keywords": _cur_kws_p[:], "mc_cat": _cur_cat_p}
            _cur_key_p = (_fac2, _typ2)
            _cur_cat_p = _cat2
            _cur_max_p = int(_desc2) if pd.notna(_desc2) and str(_desc2).strip() not in ("", "-") else 0
            _cur_kws_p = [_kw2] if _kw2 and _kw2 != "NAN" else []
        elif _fac2 in ("", "NAN") and _typ2 in ("", "NAN") and _kw2 not in ("", "NAN"):
            if _cur_key_p is not None:
                _cur_kws_p.append(_kw2)
    if _cur_key_p and _cur_kws_p:
        _TYPE_DESC_RULES_PLAN[_cur_key_p] = {"max_mc": _cur_max_p, "keywords": _cur_kws_p[:], "mc_cat": _cur_cat_p}
    print(f"✅ TYPE_DESC_RULES (Plan): {len(_TYPE_DESC_RULES_PLAN)} entries")
except Exception as _e_tdrp:
    print(f"⚠️ ไม่สามารถโหลด TYPE_DESC_RULES (Plan): {_e_tdrp}")
    _TYPE_DESC_RULES_PLAN = {}

# (MC_upper, gauge_str) → Type raw (SINGLE / DOUBLE / ...) จาก master_mc
_mc_to_type_raw_plan: dict = {
    (str(_r.get("MC", "")).strip().upper(), str(_r.get("Guage", "")).strip()): str(_r.get("Type", "")).strip().upper()
    for _, _r in master_mc.iterrows()
    if str(_r.get("MC", "")).strip()
}

# item_code → description จาก detail_mc (booking)
_item_desc_map_plan: dict = {}
if "ITEM_CODE" in detail_mc.columns and "DESCRIPTION" in detail_mc.columns:
    for _ic, _dsc in zip(detail_mc["ITEM_CODE"].astype(str).str.strip().str.upper(),
                         detail_mc["DESCRIPTION"].astype(str).str.strip()):
        if _ic and _dsc.upper() not in ("", "NAN"):
            _item_desc_map_plan[_ic] = _dsc

# booking usage ที่ consume TYPE_SPECIAL quota ไปแล้ว: (factory, type, week) → mc_used
_type_special_booking_usage: dict = {}
if "DESC_POOL_TYPE" in detail_mc.columns and "MC_USE_CEIL" in detail_mc.columns:
    _ts_bk_rows = detail_mc[
        detail_mc["DESC_POOL_TYPE"].astype(str).str.strip().str.endswith(":TYPE_SPECIAL", na=False)
    ]
    for _, _br in _ts_bk_rows.iterrows():
        _bp = str(_br["DESC_POOL_TYPE"]).strip()      # "PHET|SINGLE:TYPE_SPECIAL"
        _bparts = _bp.replace(":TYPE_SPECIAL", "").split("|")
        _bf = _bparts[0] if len(_bparts) >= 2 else ""
        _bt = _bparts[1] if len(_bparts) >= 2 else _bparts[0]
        _bw = int(_br.get("WEEK", 0) or 0)
        _bu = int(_br.get("MC_USE_CEIL", 0) or 0)
        if _bw and _bu > 0:
            _type_special_booking_usage[(_bf, _bt, _bw)] = (
                _type_special_booking_usage.get((_bf, _bt, _bw), 0) + _bu
            )

# new-plan usage ต่อ (factory, type, week)
_type_special_weekly_usage: dict = {}

# Cylinder change tracking
# ปิดการเปลี่ยนกระบอกอัตโนมัติ: Planning ไม่ยืมเครื่องข้ามเกจให้เอง
# เพราะการเปลี่ยนกระบอกเป็นการตัดสินใจของคน (มีต้นทุน/เวลา) — จะให้ AI เสนอแนะแยกต่างหาก
# ผลข้างเคียงที่ตั้งใจ: cylinder_adjustments ว่างเสมอ → เครื่องว่างที่ carry gate เห็น
# เป็นเครื่องจริงในพูลล้วนๆ ไม่มีเครื่องที่ "จะได้ถ้าเปลี่ยนกระบอก" ปนมา
# (Job sheet = กระบอกที่เปลี่ยนจริงไปแล้ว ยังโหลดตามปกติ)
CYLINDER_CHANGE_ENABLED = False
CYLINDER_CHANGE_LIMIT = 2
cylinder_change_count: dict = {}
cylinder_adjustments: dict = {}
_cylinder_change_for_item: dict = {}
_cylinder_change_start_map: dict = {}  # (factory, mc_cat, tgt_gauge) → (initiation_week, src_g, tgt_g)
_cylinder_change_mc_count: dict = {}  # (factory, mc_cat, tgt_gauge) → จำนวนเครื่องที่เปลี่ยนแล้วสะสม (ห้ามเกิน Total Spare)
_cylinder_change_done: set = set()  # (factory, mc_cat, src_gauge, tgt_gauge, week) — tracking only (not used for blocking)
_carry_cyl_pending: dict = {}  # {(week_int, item_code, mc_cat, tgt_gauge) → count} — per-item, no pool double-count


def _get_spare_info(factory: str, mc_cat: str, tgt_g: str):
    """คืน (spare_count, matched_key) — รองรับ exact match และ prefix match (SINGLE → SINGLE-32)
    เพราะ MasterMC Type_1 อาจเป็น 'SINGLE' แต่ Spare sheet ใช้ 'SINGLE-32', 'SINGLE-36' เป็นต้น
    ถ้าตรง MC_CAT เดียวกัน (prefix) ก็เปลี่ยนได้ — ใช้ key ที่มี Total Spare สูงสุด
    """
    # exact match ก่อน
    exact_val = _spare_cylinder_map.get((factory, mc_cat, tgt_g), -1)
    if exact_val >= 0:
        return exact_val, (factory, mc_cat, tgt_g)
    # prefix match: 'SINGLE' ตรงกับ 'SINGLE-32', 'SINGLE-36' ฯลฯ
    best_val = 0
    best_key = None
    prefix = mc_cat.split("-")[0]  # 'SINGLE-32' → 'SINGLE', 'SINGLE' → 'SINGLE'
    for _k, _v in _spare_cylinder_map.items():
        if _k[0] == factory and _k[2] == tgt_g:
            _k_prefix = _k[1].split("-")[0]
            if _k_prefix == prefix and _v > best_val:
                best_val = _v
                best_key = _k
    return best_val, best_key
_current_order_rdd_idx = None  # rdd_idx ของ order ที่กำลัง plan อยู่ (ใช้ตรวจ JIT timing)



def _mc_to_type1(mc_group: str, gauge=None) -> str:

    """แปลง MC_GROUP + gauge → Type_1 จาก MasterMC"""

    mc_u = str(mc_group).strip().upper() if mc_group else ""

    g_u = _normalize_gauge(gauge)

    t1 = _MC_TYPE1_MAP.get((mc_u, g_u))

    if t1 is None and g_u:

        t1 = _MC_TYPE1_MAP.get((mc_u, ""))

    return t1 if t1 else mc_u


def _mc_to_factory(mc_group: str, gauge=None) -> str:
    """แปลง MC_GROUP + gauge → Factory จาก MasterMC"""
    mc_u = str(mc_group).strip().upper() if mc_group else ""
    g_u = _normalize_gauge(gauge)
    f = _mc_factory_map.get((mc_u, g_u))
    if f is None and g_u:
        f = _mc_factory_map.get((mc_u, ""))
    return f if f else ""


def _cylinder_quota_group(factory: str, mc_cat: str):
    """คืน quota group: 'OMNOI', 'PHET_DOUBLE', 'PHET_SINGLE' หรือ None"""
    f = str(factory).strip().upper()
    cat = str(mc_cat).strip().upper()
    if f == "OMNOI":
        return "OMNOI"
    elif f == "PHET":
        if "DOUBLE" in cat:
            return "PHET_DOUBLE"
        else:
            return "PHET_SINGLE"
    return None


def _find_source_gauge_for_cylinder(factory: str, mc_cat: str, target_gauge: str, week, mc_group: str = "", debug: bool = False) -> str:
    """หา gauge ที่จะเปลี่ยน cylinder มาจาก: AVA > 0 ใน week นั้น และมี TOTAL_MC มากที่สุด
    พูลแยก (SKP vs SKPTA/SKPLE): พิจารณาเฉพาะเครื่องพูลเดียวกับ mc_group เท่านั้น (ไม่มองรวมพูลอื่น)"""
    f_upper = str(factory).strip().upper()
    cat_upper = str(mc_cat).strip().upper()
    if "FACTORY" not in summary_mc.columns:
        if debug: print(f"[SRC GAUGE] ไม่มีคอลัมน์ FACTORY ใน summary_mc")
        return None

    _mc_u = str(mc_group).strip().upper()
    _has_pool_col = "MC_POOL" in summary_mc.columns

    def _split_pool_of(_g):
        # คืน pool label ที่ต้องกรอง ถ้า (cat,g) แยกพูลและรู้ mc_group; ไม่งั้น None (พฤติกรรมเดิม)
        if _mc_u and _has_pool_col and (cat_upper, _g) in _SPLIT_CG:
            return _mc_to_pool_label(_mc_u, _g)
        return None

    def _used_at(_wk, _g):
        _pl = _split_pool_of(_g)
        if _pl:
            return weekly_new_plan_pool_usage.get(_wk, {}).get(_pl, 0)
        return weekly_new_plan_usage.get(_wk, {}).get((cat_upper, _g), 0)
    candidates = summary_mc[
        (summary_mc["FACTORY"].astype(str).str.strip().str.upper() == f_upper)
        & (summary_mc["TYPE_1"] == cat_upper)
        & (summary_mc["WEEK"] == week)
        & (summary_mc["GUAGE"].apply(_normalize_gauge) != target_gauge)
    ]
    # ห้ามใช้ SINGLE-32 Gauge 20 เป็น source cylinder change ไปให้ gauge อื่นเด็ดขาด
    if "SINGLE" in cat_upper and "32" in cat_upper:
        candidates = candidates[candidates["GUAGE"].apply(_normalize_gauge) != "20"]
    if debug:
        all_weeks = summary_mc[
            (summary_mc["FACTORY"].astype(str).str.strip().str.upper() == f_upper)
            & (summary_mc["TYPE_1"] == cat_upper)
        ]["WEEK"].unique().tolist()
        print(f"[SRC GAUGE] {f_upper}/{cat_upper} W{week}: candidates={len(candidates)}, available weeks in summary_mc={sorted(all_weeks)}")
    # pre-build future weeks list (trigger_week+1 onwards) for forward-check
    _fw_list = sorted(int(w) for w in summary_mc["WEEK"].unique() if int(w) >= int(week) + 1)
    # pre-index summary_mc rows per (factory, mc_cat, week, gauge) for fast lookup
    _src_summary_idx = {}
    for _, _sr in summary_mc[
        (summary_mc["FACTORY"].astype(str).str.strip().str.upper() == f_upper)
        & (summary_mc["TYPE_1"] == cat_upper)
    ].iterrows():
        _sw = int(_sr["WEEK"])
        _sg = _normalize_gauge(_sr["GUAGE"])
        _wp = _split_pool_of(_sg)   # พูลแยก: เก็บเฉพาะพูลเดียวกับ mc_group
        if _wp and str(_sr.get("MC_POOL", "")).strip() != _wp:
            continue
        _src_summary_idx[(_sw, _sg)] = int(_sr.get("TOTAL_MC_REMAIN", 0) or 0)

    best_gauge = None
    best_total_mc = -1
    for _, row in candidates.iterrows():
        g = _normalize_gauge(row["GUAGE"])
        # พูลแยก: ข้ามแถวพูลอื่น (เช่น ถ้า mc_group=SKPTA ห้ามเอาเครื่องพูล SKP มาเปลี่ยน cylinder)
        _want_pool = _split_pool_of(g)
        if _want_pool and str(row.get("MC_POOL", "")).strip() != _want_pool:
            continue
        base = int(row.get("TOTAL_MC_REMAIN", 0) or 0)
        used = _used_at(week, g)
        adj = cylinder_adjustments.get((week, f_upper, cat_upper, g), 0)
        # 🔧 FIX: ใช้แค่ physical machines + negative adj เท่านั้น
        # ไม่นับ positive adj (เครื่องที่ถูก cyl change เข้ามา) เป็น source
        # เพราะเครื่อง "virtual" เหล่านั้นไม่ควรถูกเปลี่ยน cylinder ซ้ำอีกครั้ง
        ava = base - used + min(0, adj)
        if debug: print(f"[SRC GAUGE]   G{g}: TOTAL_MC_REMAIN={base}, used={used}, adj={adj}, ava_physical={ava}")
        if g == "26":  # DEBUG: trace G26 source gauge selection
            _mc_grp_dbg = row.get("MC_GROUP", row.get("MC_CAT", row.get("GROUP", "?")))
            _total_mc_dbg = row.get("TOTAL_MC", "?")
            print(f"[DEBUG G26 SRC] W{week} {f_upper}/{cat_upper} G26 MC_GROUP={_mc_grp_dbg} TOTAL_MC={_total_mc_dbg}: base={base}, used={used}, adj={adj}, ava_physical={ava}")
        if ava > 0:
            # 🔧 FIX: Pool rows มี TOTAL_MC = pool total (ทุก gauge รวมกัน) ไม่ใช่ G gauge เฉพาะ
            # ตรวจ non-pool row (FACTORY="") เพื่อดูจำนวนเครื่อง G{g} จริงๆ
            _nonpool_g_rows = summary_mc[
                (summary_mc["FACTORY"].astype(str).str.strip() == "")
                & (summary_mc["TYPE_1"] == cat_upper)
                & (summary_mc["WEEK"] == week)
                & (summary_mc["GUAGE"].apply(_normalize_gauge) == g)
            ]
            if not _nonpool_g_rows.empty:
                _np_total = int(_nonpool_g_rows["TOTAL_MC"].sum())
                _np_remain = int(_nonpool_g_rows["TOTAL_MC_REMAIN"].sum())
                _np_ava = _np_remain - used + min(0, adj)
                if debug or g == "26":
                    print(f"[SRC GAUGE PER-GAUGE]   G{g}: np_total={_np_total}, np_remain={_np_remain}, used={used}, np_ava={_np_ava}")
                if _np_total > 0 and _np_ava <= 0:
                    # เครื่อง G{g} จริงใช้หมดแล้ว — ห้ามเปลี่ยน cylinder จาก gauge นี้
                    if debug: print(f"[SRC GAUGE]   G{g}: BLOCKED (per-gauge physical machines exhausted: np_remain={_np_remain}, np_ava={_np_ava})")
                    continue
            # ตรวจ future weeks: การหัก 1 เครื่องจาก src gauge จะทำให้ week ไหนติดลบไหม
            _future_ok = True
            for _fw in _fw_list:
                _base_fw = _src_summary_idx.get((_fw, g), 0)
                _used_fw = _used_at(_fw, g)
                _adj_fw = cylinder_adjustments.get((_fw, f_upper, cat_upper, g), 0)
                # ใช้ min(0, adj) เช่นเดียวกัน — ไม่นับ incoming virtual machines
                _ava_fw = _base_fw - _used_fw + min(0, _adj_fw) - 1  # -1 สำหรับ cylinder change นี้
                if _ava_fw < 0:
                    if debug: print(f"[SRC GAUGE]   G{g}: BLOCKED — W{_fw} ava_after={_ava_fw} (base={_base_fw}, used={_used_fw}, adj={_adj_fw}) → ติดลบ")
                    _future_ok = False
                    break
            if not _future_ok:
                continue
            total_mc = int(row.get("TOTAL_MC", 0) or 0)
            if total_mc > best_total_mc:
                best_total_mc = total_mc
                best_gauge = g
    if debug: print(f"[SRC GAUGE] result: src_gauge={best_gauge}")
    return best_gauge


def _try_cylinder_change(mc_cat: str, factory: str, target_gauge: str, week, item_code: str = "", mc_group: str = "", debug: bool = False, jit_override: bool = False) -> bool:
    """พยายาม cylinder change (gauge เดิม → target_gauge ภายใน MC_CAT เดิม) คืน True ถ้าสำเร็จ
    jit_override=True: ข้าม JIT 2-week window (ใช้เฉพาะ carryover path ที่ pool หมดและต้องการเครื่องแน่นอน)
    """
    _dbg = debug or bool(item_code)
    # ปิดทั้งระบบ: ไม่เปลี่ยนกระบอกอัตโนมัติในแผน (ดู CYLINDER_CHANGE_ENABLED)
    if not CYLINDER_CHANGE_ENABLED:
        if _dbg: print(f"[CYL OFF] {item_code} W{week}: cylinder change ปิดอยู่ → ไม่เปลี่ยนกระบอกในแผน")
        return False
    # บล็อค late orders (rdd_idx=None) — ไม่ใช้ spare cylinder กับ order ที่เลย deadline ไปแล้ว
    if _current_order_rdd_idx is None:
        if _dbg: print(f"[CYL BLOCKED] {item_code} W{week}: rdd_idx=None (late/no-target order)")
        return False
    _pw_idx = week_index(week)
    if _pw_idx is None:
        if _dbg: print(f"[CYL BLOCKED] {item_code} W{week}: week ไม่อยู่ใน calendar")
        return False
    if not jit_override:
        # JIT check: trigger เฉพาะเมื่อ plan_week-1 อยู่ภายใน 2 week ของ target (rdd_idx)
        if _pw_idx + 2 < _current_order_rdd_idx:
            if _dbg: print(f"[CYL BLOCKED] {item_code} W{week}: JIT fail (trigger_idx={_pw_idx}+2={_pw_idx+2} < rdd={_current_order_rdd_idx})")
            return False
    # ห้ามเปลี่ยน cylinder เร็วกว่า today+2 weeks (trigger_week ต้อง >= TODAY_IDX+2)
    if _pw_idx < TODAY_IDX + 2:
        if _dbg: print(f"[CYL BLOCKED] {item_code} W{week}: too early (trigger_idx={_pw_idx} < TODAY+2={TODAY_IDX+2})")
        return False
    f_upper = str(factory).strip().upper()
    cat_upper = str(mc_cat).strip().upper()
    tgt_g = _normalize_gauge(target_gauge)
    if not f_upper or not cat_upper or not tgt_g:
        if _dbg: print(f"[CYL BLOCKED] {item_code} W{week}: Factory หรือ MC_CAT หรือ Gauge ว่าง (Factory={f_upper}, MC_CAT={cat_upper}, Gauge={tgt_g})")
        return False
    # ห้าม cylinder change ทุกทิศทางที่เกี่ยวกับ SINGLE-32 Gauge 20 เด็ดขาด
    # (ห้ามเปลี่ยน G20→อื่น และห้ามเปลี่ยน อื่น→G20)
    if "SINGLE" in cat_upper and "32" in cat_upper and tgt_g == "20":
        if _dbg: print(f"[CYL BLOCKED] {item_code} W{week}: SINGLE-32 G20 ห้าม cylinder change ทุกกรณี (target=G20)")
        return False
    if _dbg: print(f"[CYL CHECK] {item_code} W{week}: Factory={f_upper}, MC_CAT={cat_upper}, Gauge={tgt_g} → ค้นหา spare ใน MasterMC[Spare part]...")
    spare, _spare_key = _get_spare_info(f_upper, cat_upper, tgt_g)
    if spare <= 0:
        if _dbg:
            _avail_keys = [k for k in _spare_cylinder_map if k[0] == f_upper]
            print(f"[CYL BLOCKED] {item_code} W{week}: ไม่มี spare สำหรับ Factory={f_upper}, MC_CAT={cat_upper}, Gauge={tgt_g}")
            print(f"[CYL BLOCKED]   spare ที่มีสำหรับ Factory={f_upper}: {_avail_keys if _avail_keys else 'ไม่มีเลย'}")
        return False
    # ใช้ key จริงที่ match (อาจเป็น 'SINGLE-32' แม้ cat_upper='SINGLE')
    _matched_mc_cat = _spare_key[1] if _spare_key else cat_upper
    # ตรวจ Total Spare limit: จำนวนที่เปลี่ยนแล้วสะสมต้องไม่เกิน Total Spare
    _used_so_far = _cylinder_change_mc_count.get((f_upper, _matched_mc_cat, tgt_g), 0)
    if _used_so_far >= spare:
        if _dbg: print(f"[CYL BLOCKED] {item_code} W{week}: ใช้ spare ครบแล้ว {_used_so_far}/{spare} สำหรับ {f_upper}/{_matched_mc_cat}/G{tgt_g}")
        return False
    if _dbg: print(f"[CYL CHECK] {item_code} W{week}: spare match → {_spare_key} total={spare}, used={_used_so_far}, remaining={spare-_used_so_far}")
    group = _cylinder_quota_group(f_upper, cat_upper)
    if group is None:
        if _dbg: print(f"[CYL BLOCKED] {item_code} W{week}: factory/cat ไม่อยู่ใน quota group ({f_upper}, {cat_upper})")
        return False
    current_count = cylinder_change_count.get(week, {}).get(group, 0)
    if current_count >= CYLINDER_CHANGE_LIMIT:
        if _dbg: print(f"[CYL BLOCKED] {item_code} W{week}: quota เต็ม {group}={current_count}/{CYLINDER_CHANGE_LIMIT}")
        return False
    # ใช้ _matched_mc_cat ใน find source เพื่อ filter summary_mc ได้ถูกต้อง
    _src_cat = _matched_mc_cat
    src_g = _find_source_gauge_for_cylinder(f_upper, _src_cat, tgt_g, week, mc_group=mc_group, debug=_dbg)
    if src_g is None:
        if _dbg: print(f"[CYL BLOCKED] {item_code} W{week}: ไม่มี source gauge ที่ว่างสำหรับ {f_upper} {_src_cat} G{tgt_g} ใน W{week}")
        return False
    _done_key = (f_upper, cat_upper, src_g, tgt_g, week)
    if week not in cylinder_change_count:
        cylinder_change_count[week] = {}
    cylinder_change_count[week][group] = current_count + 1
    # Apply adjustment week+1 เป็นต้นไปจนสิ้นสุด horizon (cylinder เปลี่ยนถาวรจนกว่าจะเปลี่ยนอีกครั้ง)
    # ใช้ _src_cat (matched MC_CAT จาก spare sheet) เพื่อ track adjustment ให้ถูก pool
    _future_weeks = sorted(int(w) for w in summary_mc["WEEK"].unique() if int(w) >= int(week) + 1)
    for _fw in _future_weeks:
        _src_key = (_fw, f_upper, _src_cat, src_g)
        _tgt_key = (_fw, f_upper, _src_cat, tgt_g)
        cylinder_adjustments[_src_key] = cylinder_adjustments.get(_src_key, 0) - 1
        cylinder_adjustments[_tgt_key] = cylinder_adjustments.get(_tgt_key, 0) + 1
    _cylinder_change_done.add(_done_key)
    # บันทึก start week ครั้งแรกเท่านั้น
    if (f_upper, _matched_mc_cat, tgt_g) not in _cylinder_change_start_map:
        _cylinder_change_start_map[(f_upper, _matched_mc_cat, tgt_g)] = (int(week), src_g, tgt_g)
    # นับจำนวนเครื่องที่เปลี่ยนสะสม — ใช้ _matched_mc_cat เพื่อ enforce Total Spare ถูก key
    _cylinder_change_mc_count[(f_upper, _matched_mc_cat, tgt_g)] = _cylinder_change_mc_count.get((f_upper, _matched_mc_cat, tgt_g), 0) + 1
    item_key = (week, str(item_code).strip().upper(), str(mc_group).strip().upper())
    _cylinder_change_for_item[item_key] = (src_g, tgt_g)
    print(f"🔄 CYLINDER CHANGE W{week}: {f_upper} {_matched_mc_cat} G{src_g}→G{tgt_g} [{group} {current_count+1}/{CYLINDER_CHANGE_LIMIT}] spare used {_used_so_far+1}/{spare} ({item_code}) → apply weeks {_future_weeks[:3]}...")
    return True


# สร้าง lookup dictionary สำหรับ Itemcore: {item_code: customer}

itemcore_lookup = {}

if not itemcore_df.empty:

    for _, row in itemcore_df.iterrows():

        item = str(row.get('Item code', row.get('Item code ', ''))).strip().upper()

        customer = str(row.get('Customer', '')).strip()

        if item and item not in ('NAN', 'NONE'):

            itemcore_lookup[item] = customer



# Stock Data for Core Item trigger week calculation

_FILTERED_STOCK_FILE = DATA_PLAN_DIR / "filtered_stock_data.xlsx"

stock_inventory_lookup = {}

try:

    _fstock_df = pd.read_excel(_FILTERED_STOCK_FILE)

    _fstock_df.columns = _fstock_df.columns.str.strip()

    for _, _frow in _fstock_df.iterrows():

        _fitem = str(_frow.get('ITEM_CODE', '')).strip().upper()

        _finv = float(_frow.get('Inventory', 0) or 0)

        _fsm = float(_frow.get('STOCK_MIN', 0) or 0)

        _fs5 = float(_frow.get('Stock 5 Week', 0) or 0)

        _fteam = str(_frow.get('TEAM_NAME', '')).strip().upper()

        if _fitem and _fsm > 0:

            if _fitem in stock_inventory_lookup:

                _existing_team = stock_inventory_lookup[_fitem].get('_team', '')

                if _existing_team != 'RTS' and _fteam == 'RTS':

                    continue

            stock_inventory_lookup[_fitem] = {'Inventory': _finv, 'STOCK_MIN': _fsm, 'Stock_5_Week': _fs5, '_team': _fteam}

    print(f"Stock Data: {len(stock_inventory_lookup)} items")

except Exception as _e_stock:

    print(f"Cannot load Stock Data: {_e_stock}")

# Gauge lookup: (ITEM_CODE, MC_GROUP) → GUAGE string

# ใช้เป็น fallback เมื่อ gauge ไม่ได้มาจาก data source โดยตรง

_item_mc_to_gauge = {}





def _normalize_gauge(gauge) -> str:

    """Normalize gauge string for comparison"""

    if pd.isna(gauge):

        return ""

    gauge_str = str(gauge).strip().upper()

    # Remove common suffixes/prefixes

    gauge_str = gauge_str.replace("G", "").replace("GAUGE", "")

    # ตัด ".0" ท้าย: gauge จาก float (เช่น 36.0) ให้ตรงกับ gauge จาก string ("36")
    # ไม่งั้นคีย์ lookup ไม่ตรง (เช่น MC Special ใน MasterMC เป็น float แต่ DETAIL เป็น string)
    if gauge_str.endswith(".0"):
        gauge_str = gauge_str[:-2]

    return gauge_str





def _get_item_cotton_poly(item_code: str) -> str:
    """คืน 'COTTON' ถ้า prefix FD5/F5, 'POLY' ถ้า prefix FD4/F4, '' ถ้าไม่ตรง"""
    item = str(item_code).strip().upper()
    if item.startswith("FD5") or item.startswith("F5"):
        return "COTTON"
    if item.startswith("FD4") or item.startswith("F4"):
        return "POLY"
    return ""


def _get_subgroup_by_item_prefix(mc_group: str, gauge_str: str, item_code: str) -> str:
    """คืน 'COTTON' หรือ 'POLY' ถ้า (mc_group, gauge) มีใน MC Special และ item prefix ตรง
    อ่านจากชีท MC Special ใน MasterMC.xlsx แทนการ hardcode
    คืน None ถ้าไม่มี sub-pool พิเศษสำหรับ item นี้
    """
    item_type = _get_item_cotton_poly(item_code)
    if not item_type:
        return None

    mc_upper = str(mc_group).strip().upper()
    g_norm = _normalize_gauge(gauge_str)
    factory = _mc_to_factory(mc_upper, g_norm)
    mc_cat = _mc_to_type1(mc_upper, g_norm)

    # ค้นหาใน MC Special: specific MC ก่อน แล้วค่อย general (MC="")
    ms_entry = (
        _MC_SPECIAL_PLAN.get((factory, mc_cat, mc_upper, g_norm))
        or _MC_SPECIAL_PLAN.get((factory, mc_cat, "", g_norm))
    )
    if not ms_entry:
        return None

    if ms_entry.get(item_type, 0) > 0:
        return item_type
    return None





def _apply_mc_redirect(mc_group: str, gauge) -> tuple:
    """ถ้า (mc_group, gauge) อยู่ใน MC_GROUP_REDIRECT ให้เปลี่ยนไปใช้ค่าใหม่แทนเสมอ
    เช่น SKP gauge-20 → FA gauge-20 (อ้อมน้อย)
    คืนค่า (new_mc_group, new_gauge_str)
    """
    gauge_str = _normalize_gauge(gauge)
    key = (str(mc_group).strip().upper() if mc_group else "", gauge_str)
    if key in MC_GROUP_REDIRECT:
        return MC_GROUP_REDIRECT[key]
    return mc_group, gauge_str
for _, _r in item_cap_data.iterrows():
    _ic = str(_r.get("ITEM_CODE", "")).strip().upper()
    _mc = str(_r.get("MC_GROUP", "")).strip().upper()
    _gg = _r.get("GUAGE")
    _gs = _normalize_gauge(_gg)
    if _ic and _mc and _gs and _gs.lower() != "nan":
        _item_mc_to_gauge[(_ic, _mc)] = _gs



# =========================
# ITEM SPECIAL: per-(Item, MC, Guage) override for Working day and Working hour
# Source: MasterMC.xlsx sheet "Item Special"
# =========================
ITEM_SPECIAL_LOOKUP: dict = {}  # key=(item_upper, mc_upper, gauge_str), value=(working_day, working_hour)
try:
    _is_df = pd.read_excel(_MASTER_MC_PATH, sheet_name="Item Special")
    _is_df.columns = _is_df.columns.str.strip()
    for _, _is_row in _is_df.iterrows():
        _is_mc = str(_is_row.get("MC", "")).strip().upper()
        _is_guage = _normalize_gauge(_is_row.get("Guage", ""))
        _is_item = str(_is_row.get("Item", "")).strip().upper()
        _is_wd = int(_is_row.get("Working day", 6) or 6)
        _is_wh = int(_is_row.get("Working hour", 20) or 20)
        if _is_item and _is_mc:
            ITEM_SPECIAL_LOOKUP[(_is_item, _is_mc, _is_guage)] = (_is_wd, _is_wh)
    print(f"Item Special: {len(ITEM_SPECIAL_LOOKUP)} entries loaded from MasterMC.xlsx (sheet: Item Special)")
except Exception as _e_is:
    print(f"Cannot load Item Special ({_e_is}) -- using MasterMC defaults")
    ITEM_SPECIAL_LOOKUP = {}


# S9 Logic: โหลด S9 Only + MC S9
# ⚠️ เลิกใช้ S9 Eligible (fallback เมื่อเครื่องปกติไม่ทัน) แล้ว — เหลือเฉพาะ "S9 Only"
# ที่ต้อง route ไป S9 pool เสมอ จึงไม่อ่านชีท "Item S9" อีก
_s9_only_items: set = set()       # items ที่ใช้ S9 เสมอ
_mc_s9_df: "pd.DataFrame" = pd.DataFrame()
_s9_weekly_usage: dict = {}  # {(week, s9_mc_upper, gauge_norm): machines_allocated}
try:
    _s9only_df = pd.read_excel(_MASTER_MC_PATH, sheet_name="S9 Only")
    _s9only_df.columns = _s9only_df.columns.str.strip()
    _s9_only_items = set(str(v).strip().upper() for v in _s9only_df["ITEM_CODE"].dropna())
    _mc_s9_df = pd.read_excel(_MASTER_MC_PATH, sheet_name="MC S9")
    _mc_s9_df.columns = _mc_s9_df.columns.str.strip()
    print(f"[S9] Loaded {len(_s9_only_items)} S9-only items, {len(_mc_s9_df)} MC S9 entries")
except Exception as _s9_err:
    print(f"[WARN] Cannot load S9 data from MasterMC: {_s9_err}")


def get_item_special(item_code, mc_group, gauge=None):
    """Return (working_day, working_hour) from Item Special for (item, MC, gauge), or None if not found."""
    if not item_code or not mc_group:
        return None
    item_u = str(item_code).strip().upper()
    mc_u = str(mc_group).strip().upper()
    g_u = _normalize_gauge(gauge)
    result = ITEM_SPECIAL_LOOKUP.get((item_u, mc_u, g_u))
    if result is None and g_u:
        result = ITEM_SPECIAL_LOOKUP.get((item_u, mc_u, ""))
    return result


def _calc_s9_required_machines(qty, plan_week, rdd_idx, gauge, mc_cat="", setup_days=SETUP_DAYS, material_content=""):
    """คำนวณ required machines จาก S9 MC pool โดย match MC_CAT + gauge
    คืน (s9_mc_group, s9_cap_per_day, req_mc, feasible, gauge)
    ถ้า gauge ไม่ตรงกับ MC S9 → คืน (None, None, None, False, gauge)
    """
    if _mc_s9_df.empty or gauge is None:
        return None, None, None, False, gauge
    g = _normalize_gauge(gauge)
    if not g:
        return None, None, None, False, gauge
    df = _mc_s9_df.copy()
    df["_g"] = df["Guage"].apply(_normalize_gauge)
    matches = df[df["_g"] == g]
    if mc_cat:
        mc_cat_u = str(mc_cat).strip().upper()
        _cat_m = matches[matches["MC_CAT"].str.strip().str.upper() == mc_cat_u]
        if not _cat_m.empty:
            matches = _cat_m
    if matches.empty:
        return None, None, None, False, gauge
    # เลือก pool ตาม material: POLY → Poly Only, อื่นๆ → non-Poly-Only
    _is_poly = "POLY" in str(material_content).strip().upper() if material_content else False
    _remark_col = "Remark" if "Remark" in matches.columns else None
    if _remark_col:
        _poly_flag = matches[_remark_col].fillna("").str.strip().str.upper().str.contains("POLY")
        if _is_poly:
            _poly_rows = matches[_poly_flag]
            if not _poly_rows.empty:
                matches = _poly_rows
        else:
            _non_poly_rows = matches[~_poly_flag]
            if not _non_poly_rows.empty:
                matches = _non_poly_rows
    matches = matches.sort_values("Total MC", ascending=False)
    row = matches.iloc[0]
    s9_mc = str(row["MC Group"]).strip()
    s9_cap = float(row["Cap/Day"])
    s9_total = int(row["Total MC"])
    s9_wd = int(row["Working day"])
    if s9_cap <= 0:
        return None, None, None, False, gauge
    g_norm = _normalize_gauge(g)
    s9_mc_upper = str(s9_mc).strip().upper()
    s9_used = _s9_weekly_usage.get((plan_week, s9_mc_upper, g_norm), 0)
    s9_remain = max(0, s9_total - s9_used)
    # คำนวณ req_mc ตาม qty และ weeks เหลือจนถึง RDD
    _pw_idx = week_index(plan_week)
    weeks_left = max(1, (rdd_idx - _pw_idx)) if rdd_idx is not None and _pw_idx is not None else 1
    week_cap_1mc = s9_cap * s9_wd
    if week_cap_1mc > 0:
        import math as _math
        req_mc = max(1, _math.ceil(qty / (weeks_left * week_cap_1mc)))
    else:
        req_mc = 1
    req_mc = min(req_mc, s9_remain) if s9_remain > 0 else req_mc
    feasible = s9_remain > 0
    # S9: ใช้เครื่องที่ว่างทั้งหมด — ไม่คำนวณเพื่อจบใน 1 สัปดาห์ (แผนปกติยังจบตาม target)
    print(f"[S9 CALC] gauge={g}, MC_CAT={mc_cat}, material={material_content} → MC={s9_mc}, cap={s9_cap}, total={s9_total}, used={s9_used}, remain={s9_remain}, req_mc={req_mc}, feasible={feasible}")
    return s9_mc, s9_cap, req_mc, feasible, gauge


def _ck(item, mc_group, gauge=None):
    """สร้าง carryover key: (item, mc_group, gauge) — match ITEM+MC_GROUP+GUAGE"""
    item_norm = str(item).strip().upper() if item is not None else ""
    mc_norm = str(mc_group).strip().upper() if mc_group is not None else ""
    g = _normalize_gauge(gauge)
    if not g:
        g = _normalize_gauge(
            _item_mc_to_gauge.get(
                (item_norm, mc_norm), ""
            )
        )
    return (item_norm, mc_norm, g)


def _combine_carry_with_booking(booking_key, carry_mc, prev_week_idx, cur_week_idx):
    """รวมเครื่อง carry ของแผน กับเครื่อง booking ของสัปดาห์ปัจจุบัน — ห้ามนับเครื่อง booking ซ้ำ

    machines_in_use[key] เก็บ "เครื่องรวม" ของสัปดาห์ก่อน (เครื่องแผน + เครื่อง booking)
    ดังนั้นเครื่องที่เป็นของแผนเองจริงๆ = carry_mc − booking(สัปดาห์ก่อน)
    เครื่องรวมสัปดาห์นี้ = เครื่องของแผนเอง + booking(สัปดาห์นี้)

    เดิมใช้ carry_mc + booking(สัปดาห์นี้) ตรงๆ → เครื่อง booking ตัวเดิมถูกบวกเข้าไปใหม่
    ทุกสัปดาห์ ทำให้จำนวนเครื่องโตขึ้นเรื่อยๆ (เช่น 1 → 2 → 3) ทั้งที่เป็นเครื่องตัวเดียวกัน

    คืน (เครื่องรวมสัปดาห์นี้, เครื่องของแผนเอง, booking สัปดาห์ก่อน, booking สัปดาห์นี้)
    """
    _bk = booking_mc_by_week.get(booking_key, {})
    bk_prev = int(_bk.get(prev_week_idx, 0) or 0) if prev_week_idx is not None else 0
    bk_now = int(_bk.get(cur_week_idx, 0) or 0) if cur_week_idx is not None else 0
    plan_own = max(0, int(carry_mc or 0) - bk_prev)
    return plan_own + bk_now, plan_own, bk_prev, bk_now


def _build_plan_remark(mc_group, carryover_mc, new_mc, setup_days, booking_mc=0):
    """สร้างข้อความ REMARK อธิบายที่มาของเครื่องในแถวแผน

    เพื่อให้คนอ่านแผนแยกออกว่าเครื่องที่เห็นเป็น "เครื่องใหม่ที่ต้อง setup" หรือ
    "เครื่องเดิมที่วิ่งอยู่แล้ว (carry / เติม order เข้าไป)" ซึ่งไม่เสียเวลา setup และไม่กินเครื่องเพิ่ม
    """
    parts = []
    if new_mc > 0:
        parts.append(f"setup เครื่องใหม่ {new_mc} ตัว ({setup_days} วัน)")
    if carryover_mc > 0:
        if booking_mc > 0:
            _own = max(0, int(carryover_mc) - int(booking_mc))
            if _own > 0:
                parts.append(f"carry เครื่องเดิม {_own} ตัว + เติมเข้าเครื่อง booking ที่ถักไอเทมนี้อยู่ {booking_mc} ตัว")
            else:
                parts.append(f"เติมเข้าเครื่อง booking ที่ถักไอเทมนี้อยู่แล้ว {booking_mc} ตัว (ไม่ใช้เครื่องเพิ่ม)")
        else:
            parts.append(f"carry เครื่องเดิม {carryover_mc} ตัวจากสัปดาห์ก่อน (ไม่ต้อง setup)")
    if not parts:
        parts.append("ไม่ใช้เครื่อง")
    return f"{mc_group}: " + " + ".join(parts)


def _resolve_carry_key(item, mc_group, gauge=None):
    """คืน key ที่เหมาะสุดสำหรับ carryover โดย fallback เป็น item+mc แม้ gauge ไม่ตรง"""
    base_key = _ck(item, mc_group, gauge)
    if (
        base_key in last_production
        or base_key in machines_in_use
        or base_key in last_sc_so_no
    ):
        return base_key

    item_norm, mc_norm, _ = base_key
    best_key = None
    best_score = (-1, -1)
    for cand in set(list(last_production.keys()) + list(machines_in_use.keys())):
        if len(cand) < 2:
            continue
        if cand[0] != item_norm or cand[1] != mc_norm:
            continue
        cand_last = int(last_production.get(cand, -1))
        cand_mc = int(machines_in_use.get(cand, 0) or 0)
        score = (cand_last, cand_mc)
        if score > best_score:
            best_score = score
            best_key = cand

    return best_key if best_key is not None else base_key

def _has_item_mc_key(key_set, key):
    """เช็คว่ามี key ของ item+mc อยู่ใน set แล้วหรือไม่ (ไม่สน gauge)"""
    if not key or len(key) < 2:
        return False
    item_norm, mc_norm = key[0], key[1]
    for cand in key_set:
        if len(cand) < 2:
            continue
        if cand[0] == item_norm and cand[1] == mc_norm:
            return True
    return False


# เฉพาะ item ที่จะวางแผน (จาก orders)
_plan_items = set(orders["Item Code"].astype(str).str.strip().str.upper().dropna().unique())
_detail_mc_plan = detail_mc[detail_mc["ITEM_CODE"].astype(str).str.strip().str.upper().isin(_plan_items)] if "ITEM_CODE" in detail_mc.columns else detail_mc


# สร้าง YARN_USED lookup จาก detail_mc (ITEM_CODE → YARN_USED)
_yarn_used_lookup = {}
_yarn_col_mc = next((c for c in _detail_mc_plan.columns if c.strip().upper() in ("YARN-USED", "YARN_USED")), None)
if _yarn_col_mc and "ITEM_CODE" in _detail_mc_plan.columns:
    for _, _row in (
        _detail_mc_plan[["ITEM_CODE", _yarn_col_mc]]
        .dropna()
        .drop_duplicates("ITEM_CODE")
        .iterrows()
    ):
        _yarn_used_lookup[str(_row["ITEM_CODE"]).strip().upper()] = str(
            _row[_yarn_col_mc]
        ).strip()



# สร้าง MATERIAL_CONTENT lookup จาก detail_mc (ITEM_CODE → MATERIAL_CONTENT)

_material_content_lookup = {}

if "MATERIAL_CONTENT" in _detail_mc_plan.columns and "ITEM_CODE" in _detail_mc_plan.columns:

    for _, _row in (

        _detail_mc_plan[["ITEM_CODE", "MATERIAL_CONTENT"]]

        .dropna()

        .drop_duplicates("ITEM_CODE")

        .iterrows()

    ):

        _v = str(_row["MATERIAL_CONTENT"]).strip()

        if _v and _v.upper() != "NAN":

            _material_content_lookup[str(_row["ITEM_CODE"]).strip().upper()] = _v



# Fallback: เติม item ที่ไม่มีใน detail_mc จาก order_ready (YARN_ITEM)

try:

    _order_ready_df = pd.read_excel(ORDER_FILE)

    _order_ready_df.columns = _order_ready_df.columns.str.strip()

    _yi_col = next((c for c in _order_ready_df.columns if c.strip().upper() == "YARN_ITEM"), None)

    _ic_col = next((c for c in _order_ready_df.columns if c.strip().upper() in ("FABRIC_ITEM", "ITEM_CODE")), None)

    if _yi_col and _ic_col:

        for _, _or in _order_ready_df[[_ic_col, _yi_col]].dropna().drop_duplicates(_ic_col).iterrows():

            _ic = str(_or[_ic_col]).strip().upper()

            _yi = str(_or[_yi_col]).strip()

            if _ic not in _yarn_used_lookup and _yi and _yi.upper() != "NAN":

                _yarn_used_lookup[_ic] = _yi

except Exception as _e:

    print(f"⚠️ ไม่สามารถโหลด order_ready สำหรับ YARN-USED fallback: {_e}")



# =========================

# LT_YARN LEAD TIME LOOKUP

# =========================

_LT_YARN_FILE = DATA_DIR / "LT_Yarn" / "fmit_yarn_leadtime.xlsx"

try:

    _lt_yarn_df = pd.read_excel(_LT_YARN_FILE)

    _lt_yarn_df.columns = _lt_yarn_df.columns.str.strip()

    _lt_yarn_df["Yarn Item"] = _lt_yarn_df["Yarn Item"].astype(str).str.strip().str.upper()

    _lt_yarn_df["POIN LT"] = pd.to_numeric(_lt_yarn_df["POIN LT"], errors="coerce")

    _lt_yarn_df["Piority"] = pd.to_numeric(_lt_yarn_df["Piority"], errors="coerce")

    print(f"[LT_YARN] โหลด fmit_yarn_leadtime สำเร็จ: {len(_lt_yarn_df)} rows")

except Exception as _e:

    print(f"⚠️ ไม่สามารถโหลด LT_Yarn: {_e}")

    _lt_yarn_df = pd.DataFrame()





def _resolve_yarn_code(yarn_code: str) -> str:

    """
    แปลง yarn code ที่อาจไม่ตรงกับ LT file ให้ตรงขึ้น
    Rules (ลองตามลำดับ จนกว่าจะ match):
    1. ชื่อเดิม (exact)
    2. Strip YD ก่อนตัวเลข: CMYD→CM, DTYYD→DTY, BCIYD→BCI ฯลฯ
    3. Strip trailing single uppercase color suffix (D/S/L/W/M/H)
    4. ทั้ง rule 2 และ 3 รวมกัน
    5. แปลง RPEM → RPE ท้ายชื่อ
    """
    if _lt_yarn_df.empty:
        return yarn_code
    lt_set = set(_lt_yarn_df["Yarn Item"].dropna().unique())
    if yarn_code in lt_set:
        return yarn_code
    # rule 2: strip YD before digit
    s2 = re.sub(r"YD(?=\d)", "", yarn_code)
    if s2 != yarn_code and s2 in lt_set:
        return s2
    # rule 3: strip trailing single uppercase letter
    s3 = re.sub(r"[A-Z]$", "", yarn_code)
    if s3 != yarn_code and s3 in lt_set:
        return s3
    # rule 4: rule2 + rule3
    s4 = re.sub(r"[A-Z]$", "", s2)
    if s4 != s2 and s4 in lt_set:
        return s4
    # rule 5: RPEM → RPE
    s5 = re.sub(r"RPEM$", "RPE", yarn_code)
    if s5 != yarn_code and s5 in lt_set:
        return s5
    s5b = re.sub(r"RPEM$", "RPE", s2)
    if s5b != s2 and s5b in lt_set:
        return s5b
    # no match found — return original (will result in empty lookup)
    return yarn_code

def get_yarn_lt_days(item_code: str) -> int:
    """
    หา Lead Time (วัน) สูงสุดจาก YARN-USED ของ item นั้น
    Logic:
    - แยก YARN-USED ด้วย '+' → ได้ yarn codes A, B, C, …
    - resolve ชื่อ yarn ให้ตรงกับ LT file
    - สำหรับแต่ละ yarn code:
        * ถ้ามี Piority=1 rows → ใช้ max POIN LT จาก Piority=1 เท่านั้น
        * ถ้าไม่มี Piority=1 → fallback ใช้ max POIN LT จากทุก row
    - เปรียบเทียบค่าตัวแทนของแต่ละ yarn แล้วเลือกค่ามากสุด
    - คืน 0 ถ้าหาไม่พบ
    """

    if _lt_yarn_df.empty:

        return 0

    yarn_used = _yarn_used_lookup.get(str(item_code).strip().upper(), "")

    if not yarn_used or str(yarn_used).strip().upper() in ("", "NAN"):

        return 0

    parts = [p.strip().upper() for p in str(yarn_used).split("+") if p.strip()]

    if not parts:

        return 0

    resolved = [_resolve_yarn_code(p) for p in parts]

    representative_lts = []

    for yarn in resolved:

        yarn_rows = _lt_yarn_df[_lt_yarn_df["Yarn Item"] == yarn]

        if yarn_rows.empty:

            continue

        p1_rows = yarn_rows[yarn_rows["Piority"] == 1]

        if not p1_rows.empty:

            rep_lt = p1_rows["POIN LT"].max()

        else:

            rep_lt = yarn_rows["POIN LT"].max()

        if pd.notna(rep_lt):

            representative_lts.append(rep_lt)

    if not representative_lts:

        return 0

    return int(max(representative_lts))





def get_yarn_lt_earliest_week(item_code: str, date_in=None):

    """

    คืนค่า week number ที่เริ่มวางแผนได้เร็วสุด

    สูตร: max(DATE_IN + POIN_LT, TODAY + 2 weeks)

    คืน None ถ้าคำนวณไม่ได้

    """

    min_start_idx = get_yarn_lt_min_start_idx(item_code, date_in=date_in)

    if min_start_idx < len(calendar_week):

        return int(calendar_week.iloc[min_start_idx]["WEEK"])

    return None





def get_yarn_lt_min_start_idx(item_code: str, date_in=None) -> int:

    """

    คำนวณ minimum start index (row index ใน calendar_week) 

    🔧 DISABLED: ไม่พิจารณา Yarn LT แล้ว - ใช้ TODAY+2 weeks เสมอ

    """

    min_two_weeks_idx = TODAY_IDX + 2

    print(f"[YARN LT] {item_code}: DISABLED - using TODAY+2 weeks (idx {min_two_weeks_idx})")

    return min_two_weeks_idx





def get_setup_days_for_item(material_content: str = "", yarn_used: str = "", mc_group: str = "", item_code: str = "") -> int:

    """

    คำนวณ setup days



    Logic:

    - Factory OMNOI (อ้อมน้อย) → OMNOI_SETUP_DAYS ทุก item เสมอ (ไม่ดู prefix / POLY / DTY)

    - Type SINGLE (เช็คจาก MC_TYPE_MAP[mc_group]) → ใช้ prefix ของ ITEM code จับคู่แบบ

      longest-prefix match จากชีท "Item Setup" (ไม่ดูตรรกะ POLY/DTY เดิม)

      ถ้าไม่ match prefix ใดเลย (รวม prefix ที่เว้นว่าง เช่น F9) → default SETUP_DAYS

    - ไม่ใช่ SINGLE → ตรรกะเดิม:

      0. ถ้า mc_group มีค่าใน _mc_setup_time_map → ใช้ค่านั้น

      1. ถ้า MATERIAL_CONTENT เป็น POLY → 5 วัน

      2. ถ้า YARN_USED มี DTY → 5 วัน (ทุก material รวม COTTON)

      3. default → 3 วัน

    """

    _mc_u = str(mc_group).strip().upper() if mc_group else ""

    # โรงอ้อมน้อย → setup คงที่ทุก item (ด่านแรก ชนะทุกกฎอื่น)
    # MasterMC ไม่มี MC group ที่อยู่ 2 โรงงาน → ระบุ factory จาก mc_group ได้เลย ไม่ต้องใช้ gauge

    if _mc_u and _mc_factory_map.get((_mc_u, ""), "") in ("OMNOI", "OM"):

        return OMNOI_SETUP_DAYS

    # เฉพาะ Type SINGLE → ใช้ prefix ของ ITEM (ไม่ดูตรรกะเดิม)
    # ครอบคลุม "SINGLE", "SINGLE (TUBE)" และเผื่อสะกด "SINGEL"

    _mc_type = MC_TYPE_MAP.get(_mc_u, "").replace("SINGEL", "SINGLE")

    if _mc_u and _mc_type.startswith("SINGLE"):

        _pfx_days = _lookup_item_prefix_setup(item_code)

        return _pfx_days if _pfx_days is not None else SETUP_DAYS

    if mc_group:

        if _mc_u in _mc_setup_time_map:

            return _mc_setup_time_map[_mc_u]

    mat = str(material_content).strip().upper() if material_content else ""

    yarn = str(yarn_used).strip().upper() if yarn_used else ""



    if "POLY" in mat and "COTTON" not in mat:

        return 5

    if "DTY" in yarn:

        return 5

    return SETUP_DAYS





def get_fiber_type_for_item(item_code: str) -> str:

    """หา FIBER_TYPE ของ item โดยดึง YARN-USED จาก detail_mc แล้วแยก '+' เช็คแต่ละ code"""

    yarn_used = _yarn_used_lookup.get(str(item_code).strip().upper(), "")

    if not yarn_used:

        return "None POLY"



    parts = [p.strip() for p in yarn_used.split("+") if p.strip()]

    for part in parts:

        if _fiber_lookup.get(part, "None POLY") == "POLY":

            return "POLY"



    return "None POLY"





_fiber_lookup = {}

# yarn ITEM_CODE → ITEM_DESC (uppercase) สำหรับ material detection

_yarn_desc_lookup = {}





def _get_material_content_from_yarn(item_code: str) -> str:

    """ดึง MATERIAL_CONTENT โดยดูจาก YARN-USED ของ item นั้น → เช็ค ITEM_DESC ของแต่ละ yarn code

    ผลลัพธ์: COTTON / POLY / CD / COTTON/CD / TC / หรือ '' ถ้าหาไม่ได้

    """

    yarn_used = _yarn_used_lookup.get(str(item_code).strip().upper(), "")

    if not yarn_used or str(yarn_used).strip().upper() in ("", "NAN"):

        return ""



    has_cotton = False

    has_poly = False

    has_cd = False



    parts = [p.strip() for p in str(yarn_used).split("+") if p.strip()]

    for part in parts:

        desc = _yarn_desc_lookup.get(part, "")

        if "COTTON" in desc:

            has_cotton = True

        if "POLYESTER" in desc or "DTY" in desc or "FDY" in desc:

            has_poly = True

        if "CD" in desc or "CATIONIC" in desc:

            has_cd = True



    if has_cd and has_cotton:

        return "COTTON/CD"

    elif has_cd:

        return "CD"

    elif has_cotton and has_poly:

        return "TC"

    elif has_cotton:

        return "COTTON"

    elif has_poly:

        return "POLY"

    return ""





def _normalize_capacity(item_code: str, mc_group: str, original_cap: float, gauge: str = None) -> float:

    """

    Normalize capacity to ensure consistency across MC_GROUPS

    แปลง CAP ทอ จากฐาน 24 ชั่วโมง → ชั่วโมงทำงานจริงของกลุ่ม

    ปกติ 20 ชม./วัน · ถ้าตั้ง WORK_HOUR override ในชีท Work Day ใช้ค่านั้น

    """

    return original_cap * (_base_work_hours(mc_group, gauge) / 24)





def get_load_balanced_machine(

    item_code,

    plan_week,

    last_production,

    required_machines_info=None,

    urgent_mode=False,

    past_rdd=False,

    force_max_mc=False,

    qty_left=0,

    daily_capacity=0,

    progressive_plan=None,

    current_machines=1,  # จำนวนเครื่องปัจจุบันสำหรับกรณี carryover

    qty_left_current_fg=0,  # qty_left เฉพาะ FG ปัจจุบัน (ไม่รวม FG ถัดไป) สำหรับ bypass check

):

    """

    Load Balancing แบบ Gradual Increase พร้อม Validate:

    1. ใช้เครื่องที่วิ่งจริงเป็นฐาน (carryover baseline)

    2. เพิ่มเครื่อง setup ได้สูงสุด +2 จากฐาน

    3. Validate ทุกครั้งก่อนจองเครื่อง

    """

    # ถ้าปิด Load Balancing ให้ใช้ฟังก์ชันเดิม

    if not USE_LOAD_BALANCING:

        return get_best_machine_for_item(

            item_code,

            plan_week,

            last_production,

            required_machines_info,

            urgent_mode,

            past_rdd,

            force_max_mc,

        )

    # ใช้ logic เดิมในการหาเครื่องที่เหมาะสม

    mc_group, daily_cap, setup_needed, available_machines, item_gauge = get_best_machine_for_item(

        item_code,

        plan_week,

        last_production,

        required_machines_info,

        urgent_mode,

        past_rdd,

        force_max_mc,

    )

    

    if mc_group is None:

        return None, None, None, None, None

    # ถ้า get_best_machine_for_item ทำ cylinder change เพิ่งสำเร็จ → bypass load balancing ทั้งหมด
    # เพราะเครื่อง cylinder change เป็นเครื่องใหม่ที่สร้างมาเพื่อ item นี้โดยเฉพาะ
    _cyl_trigger_for_check = int(plan_week) - 1
    _cyl_lb_key = (_cyl_trigger_for_check, str(item_code).strip().upper(), str(mc_group).strip().upper())
    if _cyl_lb_key in _cylinder_change_for_item:
        return mc_group, daily_cap, setup_needed, available_machines, item_gauge


    # ฐานจำนวนเครื่องที่ใช้งานจริงในสัปดาห์นี้ (carryover) เพื่อใช้เป็น baseline +2

    # ALWAYS ใช้ current_machines จริง ถึงแม้ว่า setup_needed=True

    # เพราะ carry machines คือเครื่องที่วิ่งจริง ไม่ว่าจะมี gap หรือไม่

    carryover_base_machines = 0

    try:

        if current_machines is not None:

            carryover_base_machines = max(0, int(current_machines))

    except (ValueError, TypeError):

        carryover_base_machines = 0

    

    # ค่า qty สำหรับใช้ตรวจสอบ weeks_needed: ใช้ current FG qty (ไม่ inflate จาก FG อื่น)

    # เพื่อป้องกันการอนุมัติเพิ่มเครื่องเมื่อ FG ปัจจุบันมี qty เหลือน้อย

    _actual_fg_qty = qty_left_current_fg if qty_left_current_fg > 0 else qty_left



    # 🔧 Bypass Load Balancing restrictions when required_machines_info specifies higher machine count

    # This ensures that when the system calculates more machines are needed, it uses that value

    if required_machines_info and len(required_machines_info) > 2:

        required_mc = required_machines_info[2]  # index 2 = required_machines

        if required_mc > available_machines:

            print(f"   ⚠️ Required machines {required_mc} > available {available_machines}, using available")

            required_mc = available_machines

        if required_mc > carryover_base_machines:

            # 🔧 FIX: ตรวจสอบว่า current FG qty เพียงพอให้ required_mc ทำงานต่อเนื่องใน W+1 หรือไม่

            # ถ้า qty_left_current_fg < required_mc × 5 days × daily_cap → W33 จะผลิตจบหมดแทบหมดใน week เดียว

            # แปลว่า W34 จะเหลือ qty น้อยมาก → ไม่ควรเพิ่มเครื่องใน W33 (setup cost ไม่คุ้ม)

            _can_bypass = True

            if _actual_fg_qty > 0 and daily_cap > 0:

                try:

                    _bp_wk = int(plan_week[1:]) if isinstance(plan_week, str) and plan_week.startswith('W') else int(plan_week)

                    _est_this_wk = required_mc * 5 * daily_cap

                    _est_remaining_next = max(0.0, _actual_fg_qty - _est_this_wk)

                    _est_next_mc = int(_est_remaining_next / (5 * daily_cap)) + (1 if _est_remaining_next % (5 * daily_cap) > 0 else 0)

                    if _est_next_mc < required_mc:

                        _can_bypass = False

                        print(f"   ⚠️ Bypass blocked: current FG qty={_actual_fg_qty:.0f}, after W{_bp_wk} est. remaining={_est_remaining_next:.0f} → W+1 needs {_est_next_mc} < {required_mc} mc → ไม่เพิ่มเครื่อง (setup cost ไม่คุ้ม)")

                except Exception:

                    pass

            if _can_bypass:

                # week แรก (carryover=0): ห้าม setup เกิน 2 เครื่อง
                _bypass_new_setup_cap = carryover_base_machines + 2
                if required_mc > _bypass_new_setup_cap:
                    print(f"   🔒 Bypass new-setup cap: {required_mc} → {_bypass_new_setup_cap} (carryover {carryover_base_machines} + 2)")
                    required_mc = _bypass_new_setup_cap

                print(f"   ✅ Bypass Load Balancing: Using required_machines_info {required_mc} machines (instead of gradual increase)")

                return mc_group, daily_cap, setup_needed, required_mc, item_gauge



    # Gradual Load Balancing Logic

    # แปลง plan_week เป็นตัวเลข (ถ้าเป็น "23" ให้ได้ 23, ถ้าเป็น "W23" ให้ได้ 23)

    if isinstance(plan_week, str):

        if plan_week.startswith('W'):

            current_week_num = int(plan_week[1:])

        else:

            current_week_num = int(plan_week)

    else:

        current_week_num = int(plan_week)

    

    # หาสัปดาห์ล่าสุดที่ item นี้ใช้งาน MC_GROUP นี้ (ไม่ใช่ MC_GROUP ทั้งหมด)

    # เช็คจาก plans ที่ append ไปแล้ว

    item_last_week = -1

    for plan_entry in plans:
        plan_item = plan_entry.get('ITEM_CODE', '')
        plan_mc = plan_entry.get('MC_GROUP', '')
        plan_week_value = plan_entry.get('PLAN_WEEK', '')

        if plan_item == item_code and plan_mc == mc_group:
            try:
                if isinstance(plan_week_value, str):
                    if plan_week_value.startswith('W'):
                        plan_week_num = int(plan_week_value[1:])
                    else:
                        plan_week_num = int(plan_week_value)
                else:
                    plan_week_num = int(plan_week_value)

                if plan_week_num < current_week_num and plan_week_num > item_last_week:
                    item_last_week = plan_week_num
            except (ValueError, TypeError):
                continue

    # เช็คจาก progressive_plan ด้วย (สำหรับ carryover items)
    # progressive_plan มี format: {week_str: machines}
    if progressive_plan:
        for week_str in progressive_plan.keys():
            try:
                if isinstance(week_str, str):
                    if week_str.startswith('W'):
                        prog_week_num = int(week_str[1:])
                    else:
                        prog_week_num = int(week_str)
                else:
                    prog_week_num = int(week_str)
            
                if prog_week_num < current_week_num and prog_week_num > item_last_week:
                    item_last_week = prog_week_num

            except (ValueError, TypeError):

                continue

    

    # หาจำนวนเครื่องรวมที่ใช้จริงในสัปดาห์ก่อนหน้าสำหรับ MC Group เดียวกัน

    # ใช้ weekly_mc_usage ที่เก็บ total machines per (week, mc_group)

    previous_week_mc_usage = {}

    for (week_idx, mc_grp), total_machines in weekly_mc_usage.items():

        # แปลง week index เป็นตัวเลขก่อนเปรียบเทียบ

        try:

            week_num = int(week_idx) if isinstance(week_idx, str) else week_idx

            

            if mc_grp == mc_group and week_num < current_week_num:

                # หาสัปดาห์ล่าสุดที่ใช้ MC Group นี้

                if week_num not in previous_week_mc_usage or week_num > previous_week_mc_usage.get('latest_week', -1):

                    previous_week_mc_usage['latest_week'] = week_num

                    previous_week_mc_usage['machines_used'] = total_machines

        except (ValueError, TypeError):

            # ข้ามถ้าไม่สามารถแปลง week index เป็นตัวเลขได้

            continue

    

    # คำนวณจำนวนเครื่องที่ควรใช้ในสัปดาห์นี้ (Gradual Increase)

    max_allowed_machines = available_machines

    is_first_time_setup = False

    first_time_cap = 0

    

    print(f"🔍 Load Balancing Debug for {mc_group}:")

    print(f"   - Available machines: {available_machines}")

    print(f"   - Current week: {current_week_num} ({plan_week})")

    print(f"   - Item last week: {item_last_week if item_last_week > 0 else 'None'}")

    print(f"   - Carryover base machines: {carryover_base_machines}")

    print(f"   - Previous week usage (MC_GROUP): {previous_week_mc_usage}")

    

    # 🔧 เช็คจำนวนเครื่องรวมที่ใช้ในสัปดาห์ปัจจุบันแล้ว (จาก SC อื่นที่ประมวลผลไปแล้ว)

    current_week_key = (current_week_num, mc_group)

    current_week_total_machines = weekly_mc_usage.get(current_week_key, 0)

    print(f"   - Current week total machines (so far): {current_week_total_machines}")

    

    # ถ้าสัปดาห์ปัจจุบันมีเครื่องใช้แล้ว ต้องนับรวมด้วย

    if current_week_total_machines > 0:

        print(f"   ⚠️ สัปดาห์ปัจจุบันมี SC อื่นใช้ไปแล้ว {current_week_total_machines} เครื่อง")

    

    # 🔧 ใช้ item_last_week แทน previous_week_mc_usage สำหรับการตรวจสอบ gap

    # เพราะต้องเช็ค gap จากการใช้งานของ item นี้โดยเฉพาะ ไม่ใช่ MC_GROUP ทั้งหมด

    if item_last_week > 0:

        # Item นี้เคยใช้งาน MC_GROUP นี้มาก่อน

        previous_week_num = item_last_week

        previous_week_str = f"W{previous_week_num:02d}"

        

        # 🔧 เช็คว่า previous week ห่างจาก current week มากกว่า 1 สัปดาห์หรือไม่

        # ถ้าห่างมากกว่า 1 สัปดาห์ → ถือว่าเป็น "first time use" ใหม่

        week_gap = current_week_num - previous_week_num

        

        # 🔧 หาจำนวนเครื่องที่ ITEM นี้ใช้จริง (รวมทุก SC ของ item เดียวกัน)

        # ดูย้อนหลังสูงสุด 3 สัปดาห์จาก item_last_week เพื่อหา peak

        # ป้องกันกรณี SC เปลี่ยน ทำให้สัปดาห์สุดท้ายเครื่องน้อยกว่าปกติ

        item_prev_mc = 0

        _lookback_detail = {}

        for lookback in range(0, 3):

            check_week = item_last_week - lookback

            if check_week <= 0:

                break

            week_sum = 0

            for pe in plans:

                pi = pe.get('ITEM_CODE', '')

                pm = pe.get('MC_GROUP', '')

                pw = pe.get('PLAN_WEEK', '')

                try:

                    if isinstance(pw, str):

                        pwn = int(pw[1:]) if pw.startswith('W') else int(pw)

                    else:

                        pwn = int(pw)

                    if pi == item_code and pm == mc_group and pwn == check_week:

                        week_sum += pe.get('ACTUAL_MC', 0)

                except (ValueError, TypeError):

                    continue

            if week_sum > 0:

                _lookback_detail[check_week] = week_sum

                item_prev_mc = max(item_prev_mc, week_sum)

            else:

                break  # หยุดถ้าไม่มีข้อมูลในสัปดาห์นี้ (gap)

        # fallback: ลองหาจาก progressive_plan

        if item_prev_mc == 0 and progressive_plan:

            for k in [str(item_last_week), f"W{item_last_week}"]:

                if k in progressive_plan:

                    item_prev_mc = progressive_plan[k]

                    break

        if item_prev_mc == 0:

            item_prev_mc = 1

        previous_machines = item_prev_mc

        print(f"   - Item's own machines (peak of recent weeks): {previous_machines} {_lookback_detail}")

        

        # ประกาศตัวแปรที่จะใช้ใน loop

        next_week_idx = current_week_num + 1

        next_week_str = f"W{next_week_idx:02d}"

        next_week_machines_planned = 0

        next_week_actual_remain = 0

        order_will_continue = False

        next_week_continuity_limit = 0

        try:

            next_week_actual_remain = get_actual_mc_remain(mc_group, next_week_idx, gauge=item_gauge, item_code=item_code)

        except Exception:

            next_week_actual_remain = 0

        print(f"   - {next_week_str} machine remain จริง = {next_week_actual_remain}")



        _next_week_prog_key = None

        if progressive_plan and next_week_str in progressive_plan:

            _next_week_prog_key = next_week_str

        elif progressive_plan and str(next_week_idx) in progressive_plan:

            _next_week_prog_key = str(next_week_idx)

        if _next_week_prog_key is not None:

            order_will_continue = True

            next_week_machines_planned = progressive_plan[_next_week_prog_key]

            try:

                next_week_continuity_limit = int(next_week_machines_planned)

            except (ValueError, TypeError):

                next_week_continuity_limit = 0

            print(f"   - {next_week_str} order นี้จะใช้ต่อ {next_week_machines_planned} เครื่อง (จาก progressive_plan:{_next_week_prog_key})")

        else:

            next_week_machines_planned = 0

            for plan_entry in plans:

                plan_week_value = plan_entry.get('PLAN_WEEK', '')

                plan_mc_group = plan_entry.get('MC_GROUP', '')



                try:

                    plan_week_str = str(plan_week_value)

                    if plan_week_str.startswith('W'):

                        plan_week_num = int(plan_week_str[1:])

                    else:

                        plan_week_num = int(plan_week_value)



                    if plan_mc_group == mc_group and plan_week_num == next_week_idx:

                        next_week_machines_planned += plan_entry.get('ACTUAL_MC', 0)

                except (ValueError, TypeError):

                    continue

            print(f"   - {next_week_str} จะใช้ {next_week_machines_planned} เครื่อง (จาก plans)")

        

        if week_gap > 1:

            print(f"   ⚠️ Previous week {previous_week_str} ห่างจาก current week {week_gap} สัปดาห์ → ถือว่าเป็น first time use ใหม่")

            baseline = carryover_base_machines

            gap_start_cap = min(baseline + 2, available_machines)

            print(f"📝 First time using {mc_group} (after gap) - baseline {baseline} + setup <= 2 → max {gap_start_cap} machines")

            # กรณี after-gap ให้ยึดเครื่องที่วิ่งจริงเป็น baseline และเพิ่ม setup ได้สูงสุด +2

            max_allowed_machines = gap_start_cap

            # ข้ามการลองเพิ่มเครื่อง เพราะเป็นสัปดาห์แรก

        else:

            # week_gap == 1: item ใช้งานต่อเนื่อง → เพิ่มได้สูงสุด +2 จาก item's own previous

            print(f"   - Previous week {previous_week_str}: item ใช้ {previous_machines} เครื่อง → max +2 = {previous_machines + 2}")

            

            # ลองเพิ่มทีละ 1 เครื่อง: ตรวจสอบว่าควรเพิ่ม +1 หรือ +2

            max_allowed_machines = previous_machines  # เริ่มต้นที่จำนวนเดิม

        

        # สำหรับกรณี carryover ที่มีเครื่องว่างและ order ยังเหลือเยอะ → ให้เพิ่มเครื่องได้

        if current_machines > 1 and _actual_fg_qty > 0:

            print(f"   - Carryover mode: มี {current_machines} เครื่องอยู่แล้ว, order เหลือ {_actual_fg_qty:.0f} units (current FG)")

            print(f"   - Daily capacity: {daily_capacity}, Available machines: {available_machines}")

            # เริ่มจากเครื่องที่มีอยู่แล้วและตรวจสอบการเพิ่ม

            max_allowed_machines = current_machines

            

            # ถ้า order เหลือเยอะและมีเครื่องว่าง → ลองเพิ่มเครื่องทันที

            if daily_capacity > 0 and _actual_fg_qty > daily_capacity * 5:  # เหลือมากกว่า 1 สัปดาห์

                for increment in [1, 2]:

                    test_machines = max_allowed_machines + increment

                    if test_machines <= available_machines:

                        if next_week_actual_remain < test_machines:

                            print(f"   ❌ Hard gate (W+1): {next_week_str} machine remain {next_week_actual_remain} < {test_machines} เครื่อง → ไม่อนุมัติเพิ่มใน W{current_week_num}")

                            break

                        if order_will_continue and next_week_machines_planned < test_machines:

                            print(f"   ❌ Hard gate (W+1): {next_week_str} ต่อเนื่องได้แค่ {next_week_machines_planned} < {test_machines} เครื่อง → ไม่อนุมัติเพิ่มใน W{current_week_num}")

                            break

                        if next_week_continuity_limit > 0 and next_week_continuity_limit < test_machines:

                            print(f"   ❌ Hard gate (W+1): {next_week_str} continuity limit {next_week_continuity_limit} < {test_machines} เครื่อง → ไม่อนุมัติเพิ่มใน W{current_week_num}")

                            break

                        weeks_needed = _actual_fg_qty / (test_machines * daily_capacity * 5) if test_machines > 0 else 999

                        if weeks_needed >= 2:  

                            print(f"   ✅ อนุมัติ (carryover): FG qty เหลือ {_actual_fg_qty:.0f} units (~{weeks_needed:.1f} weeks ที่ {test_machines} เครื่อง) → เพิ่มเป็น {test_machines} เครื่อง")

                            max_allowed_machines = test_machines

                            break

                        else:

                            print(f"   ❌ ไม่อนุมัติ (carryover): FG qty เหลือ {_actual_fg_qty:.0f} units (~{weeks_needed:.1f} weeks ที่ {test_machines} เครื่อง) ต้องใช้อย่างน้อย 2 สัปดาห์")

            else:

                print(f"   - ไม่เพิ่มเครื่อง: daily_capacity={daily_capacity}, qty_left={qty_left}, threshold={daily_capacity * 5 if daily_capacity > 0 else 0}")

        

        for increment in [1, 2]:

            test_machines = max_allowed_machines + increment

            if test_machines > available_machines:

                break  # เกินเครื่องที่มี

            if next_week_actual_remain < test_machines:

                print(f"   ❌ Hard gate (W+1): {next_week_str} machine remain {next_week_actual_remain} < {test_machines} เครื่อง → ไม่อนุมัติเพิ่มใน W{current_week_num}")

                break

            if order_will_continue and next_week_machines_planned < test_machines:

                print(f"   ❌ Hard gate (W+1): {next_week_str} ต่อเนื่องได้แค่ {next_week_machines_planned} < {test_machines} เครื่อง → ไม่อนุมัติเพิ่มใน W{current_week_num}")

                break

            if next_week_continuity_limit > 0 and next_week_continuity_limit < test_machines:

                print(f"   ❌ Hard gate (W+1): {next_week_str} continuity limit {next_week_continuity_limit} < {test_machines} เครื่อง → ไม่อนุมัติเพิ่มใน W{current_week_num}")

                break

            

            print(f"   🔍 ลองเพิ่มเป็น {test_machines} เครื่อง (+{increment}): {next_week_str} ใช้ {next_week_machines_planned} เครื่อง")

            

            # เงื่อนไขการอนุมัติ:

            # 1. Week ถัดไปจะใช้เครื่องเต็มตามจำนวนที่เพิ่ม (มี orders ใหม่)

            # 2. หรือ week ถัดไปยังเป็น order เดียวกันที่ carryover ต่อ (ใช้เครื่องเพิ่มได้)

            # 3. หรือ order ปัจจุบันยังเหลือ qty มากพอที่จะใช้เครื่องเพิ่มได้

            

            # 🔧 กฎใหม่: ตรวจสอบว่าจะใช้เครื่องครบอย่างน้อย 2 สัปดาห์หรือไม่

            # คำนวณจำนวนสัปดาห์ที่จะใช้เครื่องจำนวนนี้

            weeks_will_use = 0

            if _actual_fg_qty > 0 and daily_capacity > 0:

                weeks_will_use = _actual_fg_qty / (test_machines * daily_capacity * 5) if test_machines > 0 else 0

            

            # เช็คว่า week ถัดไปจะใช้เครื่องเต็มหรือไม่

            if next_week_machines_planned >= test_machines:

                # ต้องเช็คว่าจะใช้อย่างน้อย 2 สัปดาห์หรือไม่

                if weeks_will_use >= 2:

                    print(f"   ✅ อนุมัติ: {next_week_str} จะใช้เต็ม {test_machines} เครื่อง (จะใช้ ~{weeks_will_use:.1f} weeks)")

                    max_allowed_machines = test_machines

                else:

                    print(f"   ❌ ไม่อนุมัติ: จะใช้แค่ {weeks_will_use:.1f} weeks ที่ {test_machines} เครื่อง ต้องใช้อย่างน้อย 2 สัปดาห์")

                    break

            elif next_week_machines_planned > 0 and next_week_machines_planned >= max_allowed_machines:

                # Week ถัดไปยังใช้เครื่องอยู่ (carryover) และไม่น้อยกว่าเดิม

                # แสดงว่า order ยังไม่เสร็จ สามารถเพิ่มเครื่องเพื่อทำให้เสร็จเร็วขึ้นได้

                # ต้องเช็คว่าจะใช้อย่างน้อย 2 สัปดาห์หรือไม่

                if weeks_will_use >= 2:

                    print(f"   ✅ อนุมัติ: {next_week_str} ยังมี order ต่อเนื่อง ({next_week_machines_planned} เครื่อง) → เพิ่มเป็น {test_machines} เพื่อเร่งผลิต (จะใช้ ~{weeks_will_use:.1f} weeks)")

                    max_allowed_machines = test_machines

                else:

                    print(f"   ❌ ไม่อนุมัติ: จะใช้แค่ {weeks_will_use:.1f} weeks ต้องใช้อย่างน้อย 2 สัปดาห์")

                    break

            elif _actual_fg_qty > 0 and daily_capacity > 0:

                # เช็คว่า order ยังเหลือ qty มากพอที่จะใช้เครื่องเพิ่มได้หรือไม่

                # ถ้า order ยังไม่เสร็จและมีเครื่องว่าง → อนุมัติให้เพิ่ม

                # 🔧 กฎใหม่: ต้องใช้เครื่องครบอย่างน้อย 2 สัปดาห์ติดต่อกัน (ป้องกัน 2-3-2)

                weeks_needed = _actual_fg_qty / (test_machines * daily_capacity * 5) if test_machines > 0 else 999

                if weeks_needed >= 2:  # ต้องใช้อย่างน้อย 2 สัปดาห์เต็ม

                    print(f"   ✅ อนุมัติ: FG qty เหลือ {_actual_fg_qty:.0f} units (~{weeks_needed:.1f} weeks ที่ {test_machines} เครื่อง) → เพิ่มเป็น {test_machines} เพื่อเร่งผลิต")

                    max_allowed_machines = test_machines

                else:

                    print(f"   ❌ ไม่อนุมัติ: FG qty เหลือ {_actual_fg_qty:.0f} units (~{weeks_needed:.1f} weeks ที่ {test_machines} เครื่อง) ต้องใช้อย่างน้อย 2 สัปดาห์")

                    break

            else:

                print(f"   ❌ ไม่อนุมัติ: {next_week_str} จะใช้แค่ {next_week_machines_planned}/{test_machines} เครื่อง")

                break  # หยุดลองเพิ่ม

        

        # 🔧 Hard cap: บังคับทุก item ห้ามเพิ่มเกินกฎ

        if week_gap > 1:

            # First time after gap → baseline (เครื่องที่วิ่งจริง) + setup สูงสุด 2

            hard_cap = min(carryover_base_machines + 2, available_machines)

            if max_allowed_machines > hard_cap:

                print(f"   ⚠️ Hard cap (gap>{week_gap}): {max_allowed_machines} → {hard_cap} (carryover {carryover_base_machines} + 2)")

                max_allowed_machines = hard_cap

        else:

            # week_gap == 1 → ห้ามเพิ่มเกิน +2 จากฐานเครื่องที่ใช้งานจริง

            growth_base = carryover_base_machines if carryover_base_machines > 0 else previous_machines

            hard_cap = min(growth_base + 2, available_machines)

            if max_allowed_machines > hard_cap:

                print(f"   ⚠️ Hard cap: {max_allowed_machines} → {hard_cap} (base {growth_base} + 2)")

                max_allowed_machines = hard_cap

        

        print(f"   - Approved machines: {max_allowed_machines}")

    else:

        is_first_time_setup = True

        baseline = carryover_base_machines

        first_time_cap = min(baseline + 2, available_machines)

        next_week_idx = current_week_num + 1

        try:

            next_week_actual_remain = get_actual_mc_remain(mc_group, next_week_idx, gauge=item_gauge, item_code=item_code)

        except Exception:
            next_week_actual_remain = 0

        if first_time_cap > next_week_actual_remain:

            print(f"   ❌ Hard gate (W+1): W{next_week_idx:02d} machine remain {next_week_actual_remain} < {first_time_cap} เครื่อง → จำกัดเครื่องสัปดาห์นี้")

            first_time_cap = next_week_actual_remain

        print(f"📝 First time using {mc_group} - baseline {baseline} + setup <= 2 → max {first_time_cap} machines")

        # สัปดาห์แรกของ item นี้: อนุญาต setup เพิ่มได้สูงสุด +2 จากเครื่องที่วิ่งจริง

        max_allowed_machines = first_time_cap

    

    print(f"   - Final max allowed machines: {max_allowed_machines}")

    

    # ตรวจสอบว่าจะทำงานได้เต็มที่หรือไม่

    if required_machines_info and len(required_machines_info) > 0:

        # required_machines_info is a tuple: (mc_group, daily_cap, required_machines, feasible, gauge)

        # ใช้ required_machines เป็น target สำหรับ Gradual Increase

        required_mc = required_machines_info[2]  # index 2 = required_machines

        if required_mc > max_allowed_machines:

            # target เยอะกว่าที่มี → ใช้ required_mc เป็น target (เพิ่มเครื่องให้ได้)

            max_allowed_machines = required_mc

            print(f"[DEBUG LB] Using required_machines_info: {required_mc} machines as target (increase from {max_allowed_machines})")

        else:

            # target น้อยกว่าที่มี → จำกัดที่ required_mc

            max_allowed_machines = required_mc

            print(f"[DEBUG LB] Using required_machines_info: {required_mc} machines as target")

    # Hard cap สัปดาห์แรก: ห้าม setup เกิน first_time_cap (baseline + 2)
    if is_first_time_setup and first_time_cap > 0 and max_allowed_machines > first_time_cap:
        print(f"   🔒 First-week hard cap: {max_allowed_machines} → {first_time_cap} (new setup ห้ามเกิน 2 เครื่อง)")
        max_allowed_machines = first_time_cap

    # Validate ก่อนจองเครื่อง

    actual_remain = get_actual_mc_remain(mc_group, plan_week, gauge=item_gauge, item_code=item_code)

    

    # ตรวจสอบว่ามีเครื่องว่างพอตามที่กำหนดหรือไม่ — ใช้ actual_remain เสมอ

    if actual_remain < max_allowed_machines:

        print(f"⚠️ Load Balancing Validation: {mc_group} มีเครื่องว่างแค่ {actual_remain} แต่ต้องการ {max_allowed_machines}")

        max_allowed_machines = actual_remain



    # ตรวจสอบ job capacity limit เสมอ

    type_used = get_type_used_jobs(plan_week, mc_group)

    final_available = check_job_capacity_limit(

        mc_group, max_allowed_machines, urgent_mode, type_used

    )

    

    if final_available > 0:

        print(f"✅ Load Balancing: {mc_group} ใช้ {final_available} เครื่อง (Gradual Increase)")

        print("   ℹ️ Deferred weekly_mc_usage update until plan commit to avoid duplicate-count drift")

        

        return mc_group, daily_cap, setup_needed, final_available, item_gauge

    

    return None, None, None, None, None





def distribute_load_across_machines(

    item_code,

    plan_week,

    required_machines,

    available_machines_list,

    last_production,

    item_gauge=None,

):

    """

    กระจายงานไปยังเครื่องจักรหลายๆ เครื่องอย่างสมดุล

    สำหรับกรณีที่ต้องการใช้เครื่องจักรมากกว่า 1 เครื่อง

    """

    if required_machines <= 1 or len(available_machines_list) <= 1:

        # ถ้าต้องการเครื่องเดียวหรือมีเครื่องว่างเครื่องเดียว ให้ใช้เครื่องแรก

        return available_machines_list[:1]

    

    # รวบรวมข้อมูลการใช้เครื่องจักรปัจจุบัน

    machine_usage_history = {}

    for key, machines in machines_in_use.items():

        mc_group = key[1]

        if mc_group not in machine_usage_history:

            machine_usage_history[mc_group] = 0

        machine_usage_history[mc_group] += machines

    

    # จัดเรียงเครื่องจักรตาม load balancing score

    scored_machines = []

    for mc_group in available_machines_list:

        usage_score = machine_usage_history.get(mc_group, 0)

        actual_remain = get_actual_mc_remain(mc_group, plan_week, gauge=item_gauge, item_code=item_code)

        

        if actual_remain > 0:

            # คำนวณ load score: น้อยกว่า = โหลดน้อยกว่า = ดีกว่า

            load_score = usage_score

            scored_machines.append((mc_group, load_score, actual_remain))

    

    # เรียงตาม load score (น้อยไปมาก)

    scored_machines.sort(key=lambda x: x[1])

    

    # กระจายงานไปยังเครื่องที่โหลดน้อยที่สุดก่อน

    selected_machines = []

    remaining_machines_needed = required_machines

    

    for mc_group, load_score, available_count in scored_machines:

        if remaining_machines_needed <= 0:

            break

        

        # จัดสรรเครื่องตามที่มีว่าง แต่ไม่เกินที่ต้องการ

        machines_to_allocate = min(available_count, remaining_machines_needed)

        if machines_to_allocate > 0:

            selected_machines.extend([mc_group] * machines_to_allocate)

            remaining_machines_needed -= machines_to_allocate

    

    return selected_machines





def _analyze_and_balance_load(plan_week, current_plan):

    """

    วิเคราะห์ความสมดุลของ ACTUAL_MC ในสัปดาห์ที่กำหนดและปรับจำนวนเครื่องให้สมดุล

    โดยการปรับจำนวนเครื่องที่ใช้จริงในแต่ละ MC_GROUP

    """

    if not USE_LOAD_BALANCING:

        return current_plan

    

    print(f"🔍 Analyzing and balancing ACTUAL_MC for week {plan_week}...")

    

    # สร้างข้อมูลการใช้งานเครื่องจักรปัจจุบันจาก current_plan

    mc_load = {}

    mc_actual = {}

    for row in current_plan:

        mc_group = row['MC_GROUP']

        # ใช้ REQUIRED_MC แทน ALLOCATED_MC ตามข้อมูลจริง

        allocated_machines = row.get('REQUIRED_MC', row.get('ALLOCATED_MC', 1))

        if isinstance(allocated_machines, str):

            allocated_machines = float(allocated_machines.replace(',', ''))

        allocated_machines = float(allocated_machines)

        

        if mc_group not in mc_load:

            mc_load[mc_group] = 0

            mc_actual[mc_group] = []

        mc_load[mc_group] += allocated_machines

        mc_actual[mc_group].append(row)



    # หาค่าเฉลี่ยโหลดต่อเครื่องจักรในสัปดาห์นั้นๆ

    if not mc_load:

        print(f"ℹ️ No machine load to balance for week {plan_week}.")

        return current_plan



    total_allocated_mc = sum(mc_load.values())

    avg_load_per_mc = total_allocated_mc / len(mc_load)

    print(f"  Average ACTUAL_MC load for week {plan_week}: {avg_load_per_mc:.2f}")

    print(f"  Current ACTUAL_MC distribution: {mc_load}")



    # ระบุเครื่องจักรที่มีโหลดสูงและต่ำ

    hot_spots = {mc: load for mc, load in mc_load.items() 

                 if load > avg_load_per_mc * (1 + LOAD_BALANCING_THRESHOLD)}

    cold_spots = {mc: load for mc, load in mc_load.items() 

                  if load < avg_load_per_mc * (1 - LOAD_BALANCING_THRESHOLD)}



    if not hot_spots and not cold_spots:

        print(f"✅ ACTUAL_MC load for week {plan_week} is already balanced within {LOAD_BALANCING_THRESHOLD*100:.0f}% threshold.")

        return current_plan



    print(f"  Hot spots (> {avg_load_per_mc * (1 + LOAD_BALANCING_THRESHOLD):.2f}): {hot_spots}")

    print(f"  Cold spots (< {avg_load_per_mc * (1 - LOAD_BALANCING_THRESHOLD):.2f}): {cold_spots}")



    # ปรับจำนวน ACTUAL_MC ให้สมดุล

    new_plan = list(current_plan)  # ทำสำเนาเพื่อแก้ไข

    changes_made = False



    # คำนวณจำนวนเครื่องที่ควรมีในแต่ละ MC_GROUP เพื่อความสมดุล

    target_mc_per_group = avg_load_per_mc

    total_adjustment = 0



    # ปรับ hot spots ให้ลดลง

    for hot_mc, hot_load in sorted(hot_spots.items(), key=lambda x: x[1], reverse=True):

        if hot_mc in mc_actual:

            excess_load = hot_load - target_mc_per_group

            if excess_load > 0.1:  # ปรับเฉพาะที่เกินเกินมากพอ

                # กระจาย excess_load ไปยัง cold spots

                excess_to_distribute = excess_load * 0.5  # กระจาย 50% ของ excess

                

                # หา cold spots ที่สามารถรับเพิ่มได้

                available_cold_spots = []

                for cold_mc, cold_load in cold_spots.items():

                    if cold_mc in mc_actual:

                        capacity_for_increase = target_mc_per_group - cold_load

                        if capacity_for_increase > 0.1:

                            available_cold_spots.append((cold_mc, capacity_for_increase))

                

                if available_cold_spots:

                    # กระจาย excess ไปยัง cold spots ตามสัดส่วน

                    total_capacity = sum(cap for _, cap in available_cold_spots)

                    

                    for cold_mc, capacity in available_cold_spots:

                        portion = (capacity / total_capacity) * excess_to_distribute

                        if portion > 0.01:  # ปรับเฉพาะที่มีความเปลี่ยนแปลงมากพอ

                            # ปรับจำนวนเครื่องใน cold_mc

                            cold_rows = mc_actual[cold_mc]

                            total_current_cold = sum(float(r.get('REQUIRED_MC', r.get('ALLOCATED_MC', 1))) for r in cold_rows)

                            

                            # กระจายการเพิ่มไปยัง rows ต่างๆ ตามสัดส่วน

                            for row in cold_rows:

                                current_mc = float(row.get('REQUIRED_MC', row.get('ALLOCATED_MC', 1)))

                                if total_current_cold > 0:

                                    increase_ratio = portion / total_current_cold

                                    new_mc = current_mc * (1 + increase_ratio)

                                    row['REQUIRED_MC'] = new_mc

                            

                            print(f"  📈 Increased ACTUAL_MC in {cold_mc} by {portion:.2f} machines")

                            changes_made = True

                    

                    # ปรับจำนวนเครื่องใน hot_mc ให้ลดลง

                    hot_rows = mc_actual[hot_mc]

                    total_current_hot = sum(float(r.get('REQUIRED_MC', r.get('ALLOCATED_MC', 1))) for r in hot_rows)

                    

                    for row in hot_rows:

                        current_mc = float(row.get('REQUIRED_MC', row.get('ALLOCATED_MC', 1)))

                        if total_current_hot > 0:

                            decrease_ratio = excess_to_distribute / total_current_hot

                            new_mc = current_mc * (1 - decrease_ratio)

                            row['REQUIRED_MC'] = max(0.1, new_mc)  # ไม่ให้น้อยกว่า 0.1

                    

                    print(f"  📉 Decreased ACTUAL_MC in {hot_mc} by {excess_to_distribute:.2f} machines")

                    changes_made = True

    

    if changes_made:

        print(f"✅ ACTUAL_MC balancing completed for week {plan_week}")

        

        # แสดงผลลัพธ์หลังการปรับ

        final_mc_load = {}

        for row in new_plan:

            mc_group = row['MC_GROUP']

            allocated_machines = float(row.get('REQUIRED_MC', row.get('ALLOCATED_MC', 1)))

            if mc_group not in final_mc_load:

                final_mc_load[mc_group] = 0

            final_mc_load[mc_group] += allocated_machines

        

        print(f"  Final ACTUAL_MC distribution: {final_mc_load}")

        new_avg = sum(final_mc_load.values()) / len(final_mc_load)

        print(f"  New average: {new_avg:.2f}")

    else:

        print(f"ℹ️ No ACTUAL_MC balancing adjustments needed for week {plan_week}")

    

    return new_plan





# ชั่วโมงทำงานอ่านจากแผง Work Day ผ่าน _base_work_hours() เท่านั้น
# (WORKING_HOURS_MAP / _get_working_hours_for_mc ที่ hard-code 20 ชม. ถูกลบแล้ว — ไม่มีใครเรียก
#  และขัดกับแหล่งเดียว)



def _convert_cap_per_day(cap_per_day: float, mc_group: str, gauge: str = None) -> float:

    """

    Convert CAP_PER_DAY to usable capacity based on working hours

    - CAP_PER_DAY อยู่บนฐาน 24 ชม. → แปลงเป็นชั่วโมงทำงานจริงของกลุ่ม

    - ปกติ 20 ชม. · ถ้าตั้ง WORK_HOUR override ในชีท Work Day ใช้ค่านั้น

    """

    return cap_per_day * (_base_work_hours(mc_group, gauge) / 24)



def _get_capacity_for_mc_group(item_code: str, mc_group: str, gauge: str = None) -> float:

    """Get daily capacity for a specific MC_GROUP (capacity ดูต่อ MC, availability ดูต่อ Type_1+Gauge pool)"""

    item_rows = item_cap_data[

        (item_cap_data["ITEM_CODE"] == item_code) &

        (item_cap_data["MC_GROUP"] == mc_group)

    ]

    if not item_rows.empty:

        return _normalize_capacity(item_code, mc_group, item_rows["CAP ทอ"].iloc[0], gauge)



    # Fallback: ใช้ capacity ต่ำสุดของ item นี้

    all_item_caps = item_cap_data[item_cap_data["ITEM_CODE"] == item_code]["CAP ทอ"]

    if not all_item_caps.empty:

        return _normalize_capacity(item_code, mc_group, all_item_caps.min(), gauge)



    return 0



# =========================

# BOOKING RAW DATA LOADER

# =========================





def load_all_booking_data() -> pd.DataFrame:

    """โหลดข้อมูล booking ทั้งหมดจาก Booking/ directory (ประวัติการผลิตจริง)"""

    if not BOOKING_DIR.exists():

        return pd.DataFrame()



    all_files = [

        f for f in BOOKING_DIR.iterdir() if f.suffix.lower() in (".xlsx", ".xls")

    ]

    if not all_files:

        return pd.DataFrame()



    dfs = []

    for f in all_files:

        try:

            raw = f.read_bytes()

            is_zip = raw[:2] == b"PK"

            is_biff = raw[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"

            df = None

            if is_zip:

                df = pd.read_excel(io.BytesIO(raw), engine="openpyxl")

            elif is_biff:

                df = pd.read_excel(io.BytesIO(raw), engine="xlrd")

            else:

                for enc in ("cp874", "utf-8-sig", "latin-1"):

                    try:

                        text = raw.decode(enc, errors="replace")

                        df = pd.read_csv(

                            io.StringIO(text), sep="\t", on_bad_lines="skip"

                        )

                        if df.shape[1] > 3:

                            break



                    except Exception:

                        continue



            if df is not None:

                df.columns = df.columns.str.strip().str.upper()

                dfs.append(df)

        except Exception as e:

            print(f"⚠️ ไม่สามารถโหลด booking {f.name}: {e}")

    if not dfs:

        return pd.DataFrame()



    return pd.concat(dfs, ignore_index=True)



orders["Date"] = pd.to_datetime(orders["DATE_IN"], errors="coerce")

orders["DYE_END_DATE"] = pd.to_datetime(

    orders.get("DYE_END_DATE", orders.get("YARN_DYE_FINISH_DATE", orders.get("วันที่ย้อมเส้นด้ายจบ"))), errors="coerce"

)

orders["Item Code"] = orders["Item Code"].astype(str).str.upper().str.strip()

orders["Orders Type"] = orders["Orders Type"].astype(str).str.upper().str.strip()

orders["MC GROUP"] = orders["MC_GROUP"].astype(str).str.upper().str.strip()

orders["Orders.Qty"] = pd.to_numeric(orders["Orders.Qty"], errors="coerce").fillna(0)

orders["Plan Qty"] = pd.to_numeric(orders["Plan Qty"], errors="coerce").fillna(0)

orders["FG Week"] = orders["FG Week"]  # Already mapped

orders["SC/SO NO"] = orders["SC/SO NO"]  # Already mapped

orders["Pending Plan"] = pd.to_numeric(orders["Pending Plan"], errors="coerce").fillna(0)

orders["Confirm"] = pd.to_numeric(orders["CONFIRM_KNIT_PLAN_SC"], errors="coerce").fillna(0)

summary_mc["WEEK"] = summary_mc["WEEK"].astype(int)

calendar_week["WEEK"] = calendar_week["WEEK"].astype(int)

# YW_INT = คีย์สัปดาห์แบบ composite YYYYWW (เช่น 2027 W5 -> 202705) — unique ข้ามปี
# ใช้แก้ปัญหาเลข week ซ้ำข้ามปี (bare week 1-53 ไม่ unique เมื่อปฏิทินมีหลายปี)
calendar_week["YW_INT"] = calendar_week["YEAR"].astype(int) * 100 + calendar_week["WEEK"].astype(int)



# =========================

# BUILD item_cap_data จาก CAP_PER_DAY ทุก row ใน orders

# dedup ด้วย (ITEM_CODE, MC_GROUP) — ใช้ค่าแรกที่พบสำหรับแต่ละ pair

# =========================

_no_cap_items = set()

_no_cap_order_rows = []  # เก็บ order rows ที่ไม่มี CAP_PER_DAY สำหรับสร้าง sheet แยก

_seen_item_mc: set = set()  # dedup key: (ITEM_CODE, MC_GROUP)

_existing_cap_items: set = set()  # items ที่มี cap อย่างน้อย 1 MC_GROUP



for _, _ord_row in orders.iterrows():

    _ord_item = str(_ord_row.get("Item Code", "")).strip().upper()

    if not _ord_item:

        continue



    _ord_mc = str(_ord_row.get("MC GROUP", "")).strip().upper()

    _ord_sc = str(_ord_row.get("SC/SO NO", "")).strip()

    _ord_cap_per_day = _ord_row.get("CAP_PER_DAY", 0)



    # ตรวจสอบว่ามี CAP_PER_DAY หรือไม่

    if pd.isna(_ord_cap_per_day) or float(_ord_cap_per_day) == 0:

        if _ord_item not in _existing_cap_items:

            if _ord_item not in _no_cap_items:

                print(

                    f"[SKIP] {_ord_item} (SC {_ord_sc}): ไม่มี CAP_PER_DAY — ไม่วางแผนผลิต"

                )

            _no_cap_items.add(_ord_item)

            _no_cap_order_rows.append(_ord_row)

        continue



    raw_cap = float(_ord_cap_per_day)

    _pair = (_ord_item, _ord_mc)



    # เพิ่มเข้า item_cap_data เฉพาะ pair ใหม่ (ไม่ซ้ำ)

    if _pair not in _seen_item_mc:

        _seen_item_mc.add(_pair)

        new_cap_row = {

            "ITEM_CODE": _ord_item,

            "MC_GROUP": _ord_mc,

            "CAP ทอ": raw_cap,

            "REVOLUTION/WEIGHT": 1.0,

            "GUAGE": str(_ord_row.get("MC_GUAGE", "")).strip().upper(),

        }

        item_cap_data = pd.concat(

            [item_cap_data, pd.DataFrame([new_cap_row])], ignore_index=True

        )

        converted_cap = raw_cap * (_base_work_hours(_ord_mc, _ord_row.get("MC_GUAGE")) / 24)

        print(

            f"[ADD CAP] {_ord_item} (SC {_ord_sc}) MC={_ord_mc}: "

            f"CAP_PER_DAY {_ord_cap_per_day} (24hr) → {converted_cap:.2f} (20hr)"

        )



    _existing_cap_items.add(_ord_item)



_no_cap_df = pd.DataFrame(_no_cap_order_rows) if _no_cap_order_rows else pd.DataFrame()



# Rebuild gauge lookup หลัง CAP_PER_DAY ถูก populate เข้า item_cap_data แล้ว

for _, _r in item_cap_data.iterrows():

    _ic = str(_r.get("ITEM_CODE", "")).strip().upper()

    _mc = str(_r.get("MC_GROUP", "")).strip().upper()

    _gg = _r.get("GUAGE")

    _gs = _normalize_gauge(_gg)

    if _ic and _mc and _gs and _gs.lower() != "nan":

        _item_mc_to_gauge[(_ic, _mc)] = _gs



# =========================

# TRACK ITEMS WITH MULTIPLE CAPACITIES (SINGLE MC_TYPE and OM/OMNOI)



# =========================

_multi_cap_items = set()

_multi_cap_order_rows = []  # เก็บ order rows ที่มี multiple capacities



def _check_multiple_capacities(item_code: str, mc_group: str) -> bool:

    """ตรวจสอบว่า MC_GROUP นั้นๆ มีหลายค่า capacity ที่ไม่เหมือนกัน สำหรับ SINGLE MC_TYPE หรือ OM/OMNOI หรือไม่"""

    # หาข้อมูล MC_GROUP จาก Master_MC_5

    mc_info = master_mc[master_mc["MC"] == mc_group]

    if mc_info.empty:

        return False

    

    factory = str(mc_info.iloc[0]["Factory"]).strip().upper()

    mc_type = str(mc_info.iloc[0].get("Type", "")).strip().upper()

    

    # Debug: แสดงข้อมูล MC_GROUP

    if factory in ("OM", "OMNOI") or mc_type == "SINGLE":

        print(f"[DEBUG MC_INFO] {item_code} in {mc_group}: Factory={factory}, Type={mc_type}")

    

    # ตรวจสอบสำหรับ SINGLE MC_TYPE หรือ OM/OMNOI factories

    if mc_type == "SINGLE" or factory in ("OM", "OMNOI"):

        # ตรวจสอบว่า MC_GROUP นั้นๆ มีหลายค่า capacity หรือไม่

        mc_group_caps = item_cap_data[

            (item_cap_data["ITEM_CODE"] == item_code) &

            (item_cap_data["MC_GROUP"] == mc_group)

        ]

        print(f"[DEBUG CAP_COUNT] {item_code} in {mc_group}: {len(mc_group_caps)} entries found")

        

        if len(mc_group_caps) > 1:

            # ตรวจสอบว่าค่า capacity ไม่เหมือนกันหรือไม่

            unique_caps = set(mc_group_caps["CAP ทอ"].tolist())

            if len(unique_caps) > 1:

                # Debug: แสดงข้อมูล capacity ที่ไม่เหมือนกัน

                caps_list = mc_group_caps["CAP ทอ"].tolist()

                print(f"[DEBUG MULTI_CAP] {item_code} ใน {mc_group} มี {len(mc_group_caps)} entries ค่าต่างกัน: {caps_list}")

                return True

            else:

                # มีหลาย entries แต่ค่าเหมือนกัน - เป็นข้อมูลซ้ำ

                caps_list = mc_group_caps["CAP ทอ"].tolist()

                print(f"[DEBUG DUPLICATE] {item_code} ใน {mc_group} มี {len(mc_group_caps)} entries ค่าเหมือนกัน: {caps_list}")

    

    return False



# ตรวจสอบ items ที่มี multiple capacities

print("🔍 Starting MULTI_CAP detection...")

debug_count = 0

for _, _ord_row in orders.iterrows():

    _ord_item = str(_ord_row.get("Item Code", "")).strip().upper()

    _ord_mc = str(_ord_row.get("MC GROUP", "")).strip().upper()

    

    if _ord_item and _ord_mc:

        debug_count += 1

        if debug_count <= 5:  # Debug แค่ 5 rows แรก

            print(f"[DEBUG CHECKING] {_ord_item} with {_ord_mc}")

        

        if _check_multiple_capacities(_ord_item, _ord_mc):

            if _ord_item not in _multi_cap_items:

                print(f"[MULTI_CAP] {_ord_item} (MC {_ord_mc}): มีหลาย capacities - SINGLE MC_TYPE หรือ OM/OMNOI")

            _multi_cap_items.add(_ord_item)

            _multi_cap_order_rows.append(_ord_row)



print(f"🔍 Checked {debug_count} order rows for MULTI_CAP detection")



_multi_cap_df = pd.DataFrame(_multi_cap_order_rows) if _multi_cap_order_rows else pd.DataFrame()



# =========================

# TRACK RESERVOIR-GF ORDERS (Default N-3 Offset)



# =========================

_reservoir_items = set()

_reservoir_order_rows = []  # เก็บ order rows สำหรับ RESERVOIR-GF





for _, _ord_row in orders.iterrows():

    _ord_fob_type = str(_ord_row.get("FOB_TYPE", "")).strip()

    

    if _ord_fob_type == "RESERVOIR-GF":

        _ord_item = str(_ord_row.get("Item Code", "")).strip().upper()

        _ord_sc = str(_ord_row.get("SC/SO NO", "")).strip()

        

        _reservoir_items.add(_ord_item)

        _reservoir_order_rows.append(_ord_row)



_reservoir_df = pd.DataFrame(_reservoir_order_rows) if _reservoir_order_rows else pd.DataFrame()



# =========================

# FACTORY TYPE CONFIGURATION



# =========================

# สร้าง Factory Type mapping จาก Master_MC_5.xlsx

FACTORY_TYPE_MAP = {}

MC_TYPE_MAP = {}  # mc_group → Type (DOUBLE / SINGLE / etc.)

# วันทำงานย้ายไปอ่านจากชีท Work Day (Calendar.xlsx) ผ่าน WorkDay.get_working_days()
# — ไม่ใช้ Working Day ของ MasterMC และไม่มีกฎพิเศษ Week 17/32 อีกต่อไป

for _, row in master_mc.iterrows():

    mc_name = str(row["MC"]).strip().upper()  # ใช้คอลัมน์ MC

    factory_type = str(row["Factory"]).strip().upper()

    mc_type = str(row.get("Type", "")).strip().upper()  # คอลัมน์ Type

    # ข้าม OUTSOURCE เพราะเป็นการจ้างงานภายนอก

    if factory_type == "OUTSOURCE":

        continue

    # ใช้ MC name โดยตรง

    main_mc_group = mc_name

    FACTORY_TYPE_MAP[main_mc_group] = factory_type

    MC_TYPE_MAP[main_mc_group] = mc_type


# วันทำงานรายสัปดาห์ (แหล่งเดียว: ชีท Work Day / Week Merge ใน Calendar.xlsx)
# ส่ง lambda นับวันทำงานตามปฏิทิน (get_working_days_in_week นิยามด้านล่าง — resolve ตอนเรียกใช้)
# หมายเหตุ: import WorkDay + def _base_work_hours ย้ายไปไว้ต้นไฟล์แล้ว (ถูกเรียกที่ระดับ module ก่อนถึงจุดนี้)
WorkDay.init(master_mc, lambda w: len(get_working_days_in_week(w)))
# สอนคู่ (เครื่อง, เกจ) → กลุ่มในแผง จาก booking (คีย์แผงคือ FACTORY·MC_CAT·เกจ)
# ครอบคลุมเครื่องที่ไม่มีใน MasterMC เช่น SKPHF/SKPSL/SKPAO — ไม่งั้นได้ 0 วันทั้งที่ตั้งค่าไว้แล้ว
_wd_alias_cnt = WorkDay.register_aliases(detail_mc)
if _wd_alias_cnt:
    print(f"✅ จับคู่เครื่อง→กลุ่มในแผงวันทำงานเพิ่ม {_wd_alias_cnt} คู่ (จาก FACTORY+MC_CAT+เกจ)")


# set คู่ (MC_GROUP, gauge_normalized) ที่มีจริงใน MasterMC — ใช้ skip order ที่ไม่มีใน Master
_VALID_MC_GAUGE_SET: set = set()
for _, _vmrow in master_mc.iterrows():
    _vmc = str(_vmrow.get("MC", "")).strip().upper()
    _vg = _normalize_gauge(_vmrow.get("Guage", ""))
    if _vmc:
        _VALID_MC_GAUGE_SET.add((_vmc, _vg))
        _VALID_MC_GAUGE_SET.add((_vmc, ""))  # fallback ไม่เช็ค gauge
print(f"✅ Valid MC+Gauge set: {len(_VALID_MC_GAUGE_SET)} entries")

# =========================

# TODAY (ห้ามวางย้อนหลัง)



# =========================





def get_week_from_date(date):

    if pd.isna(date):

        return None



    row = calendar_week[

        (calendar_week["WEEK_START"] <= date) & (calendar_week["WEEK_END"] >= date)

    ]

    return None if row.empty else int(row.iloc[0]["WEEK"])





def wk_year(week):
    """คืนปีของค่าสัปดาห์: composite YYYYWW -> YYYY, bare -> None (ไม่รู้ปี)"""
    if week is None:
        return None
    try:
        w = int(week)
    except (TypeError, ValueError):
        return None
    return w // 100 if w >= 100000 else None


def wk_num(week):
    """คืนเลขสัปดาห์ 1-53: composite YYYYWW -> WW, bare -> คงเดิม"""
    if week is None:
        return None
    try:
        w = int(week)
    except (TypeError, ValueError):
        return None
    return w % 100 if w >= 100000 else w


_week_anchor_idx = None  # ตำแหน่งที่แผนกำลังเดินอยู่ (None = ใช้ TODAY_IDX) — ดู week_index()

# bare week ที่อ้างถึงระหว่างวางแผนย้อนหลังจาก anchor ได้ไม่เกินกี่สัปดาห์
# ต่ำกว่านี้ = ตีความเป็น occurrence ของปีถัดไปแทน (กันจับผิดปีย้อนอดีต)
# 8 สัปดาห์เผื่อไว้สำหรับ cross-SC fill / สัปดาห์ที่เพิ่งผลิต ซึ่งอยู่ก่อน anchor เล็กน้อย
WEEK_LOOKBACK_LIMIT = 8


def set_week_anchor(idx):
    """ตั้งจุดอ้างอิงของ week_index ให้ตามตำแหน่งที่แผนกำลังเดินอยู่

    ต้องเลื่อนตาม plan_week เสมอ เพื่อให้ bare week ที่อ้างถึงระหว่างวางแผน
    ถูกตีความเป็นปีที่ถูกต้อง แม้แผนจะยืดข้ามปีไปไกลจาก TODAY
    ส่ง None เพื่อรีเซ็ตกลับไปยึด TODAY_IDX (ใช้ตอนจบ order)
    """
    global _week_anchor_idx
    _week_anchor_idx = int(idx) if idx is not None else None


def week_index(week, anchor=None):
    """แปลงค่าสัปดาห์ -> row index ใน calendar_week (ตำแหน่งบน timeline)

    รองรับ 2 รูปแบบ:
      - composite YYYYWW (>= 100000) : จับคู่ YW_INT แบบ unique -> ถูกต้องข้ามปีเสมอ
      - bare week 1-53   (< 100000)  : เลือก occurrence ที่ใกล้ TODAY_IDX ที่สุด
                                        (กันเลข week ซ้ำข้ามปีจับผิดปี; ถูกทั้งอดีตใกล้ ๆ และอนาคตใกล้ ๆ)
    หมายเหตุ: ตอน bootstrap (คำนวณ TODAY_IDX ครั้งแรก) TODAY_IDX ยังไม่มี -> fallback แถวแรกสุด
    """
    if week is None:
        return None
    try:
        w = int(week)
    except (TypeError, ValueError):
        return None

    if w >= 100000:  # composite YYYYWW
        idx = calendar_week.index[calendar_week["YW_INT"] == w]
        return None if len(idx) == 0 else int(idx[0])

    idx = list(calendar_week.index[calendar_week["WEEK"] == w])
    if not idx:
        return None
    # anchor = จุดอ้างอิงสำหรับเลือก occurrence เมื่อ week number ซ้ำข้ามปี
    #   1) anchor ที่ caller ส่งมา (ชัดเจนที่สุด)
    #   2) _week_anchor_idx = ตำแหน่งที่แผน "กำลังเดินอยู่" (เลื่อนตาม plan_week)
    #   3) TODAY_IDX (ค่าเริ่มต้นตอน bootstrap / นอกบริบทการวางแผน)
    # ทำไมต้องมี (2): anchor ที่ตรึงไว้ที่ TODAY จะพังเมื่อแผนยาวเกิน ~26 สัปดาห์
    # เช่น แผนเดินถึง W48 แล้วอ้าง W20 (ตั้งใจหมายถึงปีหน้า) — วัดจาก TODAY จะได้
    # W20 ปีปัจจุบัน (ย้อนอดีต) แต่วัดจากตำแหน่งแผนจริงจะได้ปีหน้าถูกต้อง
    base = anchor
    if base is None:
        base = globals().get("_week_anchor_idx")
    if base is None:
        base = globals().get("TODAY_IDX")
    if base is None:
        return int(idx[0])
    # กันการ "ย้อนอดีตข้ามปี": ถ้าเลือกแบบใกล้สุดล้วน ๆ สัปดาห์ที่อยู่เกินครึ่งปี
    # ข้างหน้า anchor จะถูกจับเป็น occurrence ของปีก่อนแทน (เช่น anchor W48 ปีนี้
    # อ้าง W25 → ได้ W25 ปีนี้ซึ่งผ่านไปแล้ว ทั้งที่ต้องหมายถึง W25 ปีหน้า)
    # จึงคัดเฉพาะ occurrence ที่ไม่ย้อนหลังเกิน WEEK_LOOKBACK_LIMIT ก่อน
    # (เผื่อไว้สำหรับ cross-SC fill / สัปดาห์ที่เพิ่งผลิต ซึ่งอยู่ก่อน anchar ไม่กี่สัปดาห์)
    _fwd = [i for i in idx if int(i) >= base - WEEK_LOOKBACK_LIMIT]
    if _fwd:
        idx = _fwd
    # เลือกแถวที่ใกล้ anchor ที่สุด (tie -> แถวก่อน)
    return int(min(idx, key=lambda i: (abs(int(i) - base), int(i))))





def get_revolution_weight_from_orders(item_code, mc_group):

    """ค้นหา REVOLUTION_WEIGHT ของ item จาก orders DataFrame (order_ready)"""

    # หาข้อมูลของ item นี้จาก orders DataFrame

    item_rows = orders[orders['Item Code'] == item_code]

    if not item_rows.empty:

        # ใช้ REVOLUTION_WEIGHT จาก order_ready โดยตรง

        rev_weight = item_rows.iloc[0].get('REVOLUTION_WEIGHT', 0)

        return rev_weight if rev_weight and rev_weight > 0 else 0

    return 0





def get_revolution_weight(item_code, mc_group, plan_week):

    """ค้นหา REVOLUTION/WEIGHT ของ item จาก item_cap_data ที่โหลดไว้"""

    # หาข้อมูลของ item นี้จาก item_cap_data

    item_rows = item_cap_data[item_cap_data["ITEM_CODE"] == item_code]

    if not item_rows.empty:

        mc_rows = item_rows[item_rows["MC_GROUP"] == mc_group]

        if not mc_rows.empty:

            rev_weight = mc_rows.iloc[0].get("REVOLUTION/WEIGHT", 0)

            return rev_weight



        else:

            # ถ้าไม่เจอ MC_GROUP ตรงๆ ให้ใช้ค่าแรกของ item

            rev_weight = item_rows.iloc[0].get("REVOLUTION/WEIGHT", 0)

            return rev_weight



    return None





def get_working_days_by_factory(mc_group, available_machines_count=None, week=None, item_code=None, gauge=None):

    """คืนวันทำงานของ (mc_group, gauge) ในสัปดาห์ที่ระบุ จากชีท Work Day (แหล่งเดียว)

    ค่านี้เป็น "ผลลัพธ์สุดท้าย" แล้ว — ใช้ค่าที่กรอกตรง ๆ ไม่หักวันหยุดต่อ,
    รวม logic ยุบสัปดาห์ + สัปดาห์ที่ปฏิทินหยุดทั้งก้อน (=0) ไว้ให้แล้ว

    week=None → คืนค่ามาตรฐานของกลุ่ม (ไม่อิงสัปดาห์)
    (available_machines_count / item_code เก็บพารามิเตอร์ไว้เพื่อความเข้ากันได้กับ caller เดิม)
    """
    return WorkDay.get_working_days(mc_group, gauge=gauge, week=week)





def adjust_daily_cap_for_item_special(daily_cap, item_code, mc_group, gauge=None, base_working_hour=None):

    """ปรับ daily_cap ตาม Working hour จาก Item Special (Item Special ชนะเสมอ)

    daily_cap เข้ามาถูกแปลงด้วยชั่วโมงของกลุ่มจากแผง Work Day แล้ว (base)
    ถ้า item อยู่ใน Item Special และ Working hour ต่างจาก base → rescale เป็นชั่วโมงของ Item Special
    → ผลลัพธ์ = cap(24hr) * (item_special_wh / 24) เสมอ ไม่ว่ากลุ่มจะตั้งไว้เท่าใด
    """

    if not item_code or not daily_cap:

        return daily_cap

    _is = get_item_special(item_code, mc_group, gauge)

    if _is is not None:

        base = base_working_hour if base_working_hour is not None else _base_work_hours(mc_group, gauge)

        _wh = _is[1]  # working_hour

        if base > 0 and _wh != base:

            return daily_cap * (_wh / base)

    return daily_cap





def _dynamic_setup_limit(

    plan_week: int, rdd_idx, required_mc: int, remaining_job_slots: int

) -> int:

    """คืนจำนวน new machines สูงสุดที่ควร setup ใน week นี้ ตาม urgency ของ RDD

    - ห่าง RDD >= 2 week : ใช้แค่ required_mc  (ประหยัด job slot ไว้ให้ order อื่น)

    - ห่าง RDD == 1 week : ใช้แค่ required_mc (จำกัดตาม required_mc)

    - plan_week >= RDD   : ใช้แค่ required_mc (จำกัดตาม required_mc)

    ทุก case ยังต้องผ่าน check_job_capacity_limit อีกรอบเสมอ

    rdd_idx = row index ใน calendar_week (ใช้แทน fg_week_int เพื่อรองรับข้ามปี)"""

    fallback = required_mc  # เปลี่ยนจาก remaining_job_slots เป็น required_mc

    if not required_mc:

        required_mc = fallback

    if rdd_idx is None:

        # ไม่มี RDD → conservative = required_mc เท่านั้น

        return required_mc



    plan_idx = week_index(plan_week)

    if plan_idx is None:

        return required_mc



    weeks_to_rdd = rdd_idx - plan_idx

    if weeks_to_rdd <= 0:

        # urgent / เลยกำหนดแล้ว → ใช้ required_mc

        return required_mc



    elif weeks_to_rdd == 1:

        # สัปดาห์สุดท้ายก่อน RDD → ใช้ required_mc

        return required_mc



    else:

        # ยังเหลือเวลา → ใช้แค่เท่าที่จำเป็นเพื่อทัน RDD

        return required_mc





def check_job_capacity_limit(

    mc_group,

    available_machines_count,

    urgent_mode=False,

    current_week_jobs=None,

    committed_carryover=0,

):

    """ตรวจสอบว่าจำนวนเครื่องไม่เกิน job/week capacity

    committed_carryover: จำนวนเครื่อง carry-over ที่ผูกพันแล้ว (ต้องผ่านเสมอ, ห้าม cap)

    """

    # หาข้อมูล MC_GROUP จาก Master_MC_5

    mc_info = master_mc[master_mc["MC"] == mc_group]

    if mc_info.empty:

        # ถ้าไม่เจอใน Master_MC_5 ใช้ค่า default

        factory = "PHET"

        mc_type = "DOUBLE"

    else:

        # ดูว่า MC_GROUP นี้อยู่ Factory ไหน และเป็น Type อะไร

        factory = str(mc_info.iloc[0]["Factory"]).strip().upper()

        mc_type = str(mc_info.iloc[0].get("Type", "DOUBLE")).strip().upper()

    # กำหนด job/week capacity ตาม FAC และ Type

    # ตัด (TUBE) ออก — "SINGLE (TUBE)" ถือเป็น SINGLE เหมือนกัน

    mc_type_clean = mc_type.replace("(TUBE)", "").strip()

    if factory == "PHET":

        if mc_type_clean == "DOUBLE":

            max_jobs = 33

        elif mc_type_clean == "SINGLE":

            max_jobs = 44

        else:

            max_jobs = 33  # default PHET

    elif factory in ("OM", "OMNOI"):

        max_jobs = 13

    else:

        # OUTSOURCE หรือ factory อื่นๆ ไม่มี job cap → ผ่านเสมอ

        return available_machines_count



    # ห้ามเกิน cap เด็ดขาด (urgent mode ก็ใช้ cap เดิม)

    max_jobs_effective = max_jobs

    # Normal/urgent: ห้ามเกิน cap เด็ดขาด

    if current_week_jobs is not None:

        remaining_jobs = max(0, max_jobs_effective - current_week_jobs)

        if committed_carryover > 0:

            # Carryover ไม่กิน job slot เลย — cap เฉพาะ new machines เท่านั้น

            # remaining_jobs = slots ที่ยังว่างสำหรับ new setups (carryover ไม่นับ)

            new_mc = max(0, available_machines_count - committed_carryover)

            allowed_new = min(new_mc, remaining_jobs)  # ต้องไม่เกิน slot ที่เหลือ

            result = committed_carryover + allowed_new

            return result



        return min(available_machines_count, remaining_jobs)



    # ถ้าไม่มีข้อมูล current_week_jobs ให้จำกัดตาม max_jobs_effective

    return min(available_machines_count, max_jobs_effective)





# ยกเลิกกฎพิเศษราย week ของ "ปฏิทิน" — วันทำงานอ่านจากชีท Work Day อย่างเดียว
WEEK_DAYS_OVERRIDE = {}





def get_working_days_in_week(week):

    """Get working days for a specific week from calendar (กรองวันหยุดออก)

    นับวันทำงานตามปฏิทินจริง (is_working_day == 1) ทุกวัน รวมวันศุกร์ด้วย

    ถ้า calendar บอก status=1 (ทำงาน) จะนับเป็นวันทำงานไม่ว่าจะเป็นวันใด

    """

    # ตรวจสอบ override ก่อน

    if week in WEEK_DAYS_OVERRIDE:

        override_days = WEEK_DAYS_OVERRIDE[week]

        week_data = calendar_week[calendar_week["WEEK"] == week]

        if not week_data.empty:

            week_start = week_data.iloc[0]["WEEK_START"]

            week_end = week_data.iloc[0]["WEEK_END"]

            mask = (

                (calendar["DATE"] >= week_start)

                & (calendar["DATE"] <= week_end)

                & (calendar["is_working_day"] == 1)

            )

            base_days = calendar.loc[mask, "DATE"].tolist()

        else:

            base_days = []

        # เติมให้ครบตาม override (pad ด้วย None สำหรับวันพิเศษ)

        while len(base_days) < override_days:

            base_days.append(None)

        return base_days[:override_days]



    week_data = calendar_week[calendar_week["WEEK"] == week]

    if week_data.empty:

        return []



    week_start = week_data.iloc[0]["WEEK_START"]

    week_end = week_data.iloc[0]["WEEK_END"]

    # กรองเฉพาะวันที่ is_working_day == 1 จาก daily calendar

    mask = (

        (calendar["DATE"] >= week_start)

        & (calendar["DATE"] <= week_end)

        & (calendar["is_working_day"] == 1)

    )

    working_days = calendar.loc[mask, "DATE"].tolist()

    return working_days





def _sp_used_from_plans(factory, cat, gauge_str, week, sp_type):
    """นับเครื่อง sub-pool (POLY/COTTON) ที่แถวแผน "วางไปแล้วจริง" ใช้อยู่ในสัปดาห์นั้น

    ตัวนับ incremental (_mc_special_weekly_usage) ข้ามแถว FG-split และแถว SC อื่นของ item
    เดียวกันในสัปดาห์เดียวกัน (ถือว่าใช้เครื่องร่วมกัน) แต่บางแถวใช้เครื่องคนละชุดจริง
    → ตัวนับต่ำกว่าความจริง hard cap เลยปล่อยงานเกินโควตา
    ที่นี่นับจาก plans ตรงๆ: dedupe เฉพาะแถวที่ใช้เครื่องชุดเดียวกัน (item+SC+เครื่อง+เกจ)
    โดยเอาค่าสูงสุด — แถวที่ใช้เครื่องร่วมกันจริงจะมี ACTUAL_MC = 0 อยู่แล้ว
    """
    if not _PLANS_REF:
        return 0
    seen = {}
    for _r in _PLANS_REF:
        if _r.get("PLAN_WEEK") != week:
            continue
        _it = str(_r.get("ITEM_CODE", "")).strip().upper()
        if _get_item_cotton_poly(_it) != sp_type:
            continue
        _mc = str(_r.get("MC_GROUP", "")).strip().upper()
        _g = _normalize_gauge(_r.get("MC_GUAGE"))
        if _g != gauge_str:
            continue
        if _mc_to_factory(_mc, _g) != factory or _mc_to_type1(_mc, _g) != cat:
            continue
        _k = (_it, str(_r.get("SC_SO_NO", "")).strip(), _mc, _g)
        _v = int(_r.get("ACTUAL_MC", 0) or 0)
        seen[_k] = max(seen.get(_k, 0), _v)
    return sum(seen.values())


# =========================
# Pool split (SINGLE-32): SKP vs SKPTA/SKPLE เป็นคนละพูล (นิยามตามคอลัมน์ Group ใน MasterMC)
# อ่าน MC_POOL_MAP + MC_POOL จาก booking_final (สร้างโดย AVA_MC) เพื่อให้ label ตรงกันเป๊ะ
# get_actual_mc_remain จะคิดเครื่องเฉพาะพูลของเครื่องนั้น ไม่ยืมข้ามพูล
# =========================
_MC_POOL_LABEL: dict = {}   # (mc_upper, gauge_str) → pool label เช่น "PHET|SKPLE , SKPTA 26G"
_SPLIT_CG: set = set()      # (type_1_upper, gauge_str) ที่แยกพูล → คิดเครื่องต่อพูล
try:
    _pm_df = pd.read_excel(MC_REMAIN_FILE, sheet_name="MC_POOL_MAP")
    _pm_df.columns = _pm_df.columns.str.strip().str.upper()
    for _, _pm in _pm_df.iterrows():
        _pm_mc = str(_pm.get("MC_GROUP", "")).strip().upper()
        _pm_g = _normalize_gauge(_pm.get("GUAGE", ""))
        _pm_pool = str(_pm.get("MC_POOL", "")).strip()
        if _pm_mc and _pm_g and _pm_pool:
            _MC_POOL_LABEL[(_pm_mc, _pm_g)] = _pm_pool
    if "MC_POOL" in summary_mc.columns:
        _pool_by_cg: dict = {}
        for _, _sr in summary_mc.iterrows():
            _cat = str(_sr["TYPE_1"]).strip().upper()
            # แยกพูลเฉพาะ SINGLE-32 (SKP vs SKPTA/SKPLE) — cat อื่นคงพฤติกรรมเดิม (รวมทั้ง cat|gauge)
            if _cat != "SINGLE-32":
                continue
            _sp = str(_sr.get("MC_POOL", "")).strip()
            if not _sp:
                continue
            _cg = (_cat, _normalize_gauge(_sr["GUAGE"]))
            _pool_by_cg.setdefault(_cg, set()).add(_sp)
        _SPLIT_CG = {cg for cg, pools in _pool_by_cg.items() if len(pools) > 1}
    print(f"✅ Pool split: {len(_MC_POOL_LABEL)} MC labels, {len(_SPLIT_CG)} split (cat,gauge): {sorted(_SPLIT_CG)}")
except Exception as _e_pm:
    print(f"⚠️ ไม่มีชีท MC_POOL_MAP (ข้าม pool split, ใช้พฤติกรรมเดิม): {_e_pm}")


def _mc_to_pool_label(mc_group, gauge):
    """pool label ต่อเครื่อง (เช่น 'PHET|SKPLE , SKPTA 26G') สำหรับกลุ่มที่แยกพูล
    คืน None ถ้าเครื่องไม่ได้อยู่ในกลุ่มที่แยกพูล → get_actual_mc_remain ใช้พฤติกรรมเดิม"""
    return _MC_POOL_LABEL.get((str(mc_group).strip().upper(), _normalize_gauge(gauge)))


def _mc_special_remain(mc_group, week, gauge, item_code):
    """คืนจำนวนเครื่อง POLY/COTTON sub-pool ที่ยัง setup ใหม่ได้ = กันไว้ − booking − แผน(new)
    คืน None ถ้า item นี้ไม่อยู่ในกลุ่มที่กันเครื่อง (ไม่ต้อง hard cap)
    ใช้ hard-block: setup ใหม่ห้ามเกินเครื่องที่กันไว้ (carryover เครื่องเดิมไม่นับตรงนี้)"""
    if not _MC_SPECIAL_PLAN or not item_code:
        return None
    sp_type = _get_item_cotton_poly(item_code)
    if not sp_type:
        return None
    gauge_str = _normalize_gauge(gauge)
    mc_u = str(mc_group).strip().upper()
    factory = _mc_to_factory(mc_u, gauge_str)
    cat = _mc_to_type1(mc_u, gauge_str)
    entry = (
        _MC_SPECIAL_PLAN.get((factory, cat, mc_u, gauge_str))
        or _MC_SPECIAL_PLAN.get((factory, cat, "", gauge_str))
    )
    if not entry:
        return None
    reserved = entry.get(sp_type, 0)
    if reserved <= 0:
        return None
    used = max(
        _mc_special_weekly_usage.get((factory, cat, gauge_str, week, sp_type), 0),
        _sp_used_from_plans(factory, cat, gauge_str, week, sp_type),
    )
    bk = _MC_SPECIAL_BOOKING_USED.get((cat, gauge_str, week, sp_type), 0)
    return max(0, reserved - bk - used)


def get_actual_mc_remain(mc_group, week, gauge, item_code=None, allow_negative=False):

    """คืนค่าจำนวนเครื่องว่างจริงของ pool = TOTAL_MC_REMAIN จาก summary_mc

    ดูจาก Type_1 + GUAGE (รวมเครื่องทุก MC ใน Type_1+Gauge เดียวกัน)

    allow_negative=True → คืนค่าติดลบได้เมื่อ booking จองเกินเครื่องจริง (จองเกิน 5 → -5)
    ค่าปกติ (False) ตัดที่ 0 ทำให้ "ว่างพอดี 0" กับ "จองเกิน 5" แยกกันไม่ออก
    ใช้ True เฉพาะจุดที่ต้องรู้ว่าพูลเต็มแค่ไหน เช่น carry gate / look-ahead
    """

    gauge_str = _normalize_gauge(gauge)

    if not gauge_str:

        return 0



    # MC_GROUP_REDIRECT ยังคงใช้ได้ (redirect ก่อนแปลง Type_1)

    mc_group, gauge_str = _apply_mc_redirect(mc_group, gauge_str)



    # แปลง MC_GROUP → Type_1 เพื่อดูเครื่องระดับ pool

    type_1 = _mc_to_type1(mc_group, gauge_str)

    # พูลแยก (เช่น SINGLE-32: SKP vs SKPTA/SKPLE เป็นคนละชุดเครื่อง ใช้แทนกันไม่ได้)
    # → คิดเครื่องเฉพาะพูลของเครื่องนี้ ไม่รวมพูลอื่นใน Type_1+Gauge เดียวกัน
    _pool_label = _mc_to_pool_label(mc_group, gauge_str)
    _is_split = (
        _pool_label is not None
        and (type_1, gauge_str) in _SPLIT_CG
        and "MC_POOL" in summary_mc.columns
    )

    mc_rows = summary_mc[

        (summary_mc["WEEK"] == week)

        & (summary_mc["TYPE_1"] == type_1)

        & (summary_mc["GUAGE"].apply(_normalize_gauge) == gauge_str)

    ]

    if mc_rows.empty:

        # Fallback: รวม Total MC จาก master_mc ทุก MC ใน Type_1+Gauge
        # (MC_Total ของสัปดาห์นี้ ถ้ามี ให้ชนะค่าจาก master — ไม่งั้นสัปดาห์ที่ไม่มีแถวใน
        #  summary จะได้จำนวนเครื่องเก่า ขัดกับที่ user กำหนดไว้)
        _mt_ov = _MC_TOTAL_OVERRIDE.get((int(week), type_1, gauge_str))

        _mm_filter = (

            (master_mc["Type_1"].astype(str).str.strip() == type_1)

            & (master_mc["Guage"].apply(_normalize_gauge) == gauge_str)

        )

        _mm_rows = master_mc[_mm_filter]

        if _mm_rows.empty and _mt_ov is None:
            # gauge นี้ไม่มีใน master (เช่น gauge ใหม่จาก cylinder change)
            # ต้องตรวจ cylinder adjustments ก่อน return 0
            _cyl_adj_only = sum(
                v for (w, _f, cat, g), v in cylinder_adjustments.items()
                if w == week and cat == type_1 and g == gauge_str
            )
            return max(0, _cyl_adj_only)

        base_remain = (
            int(_mt_ov) if _mt_ov is not None
            else int(pd.to_numeric(_mm_rows["Total MC"], errors="coerce").fillna(0).sum())
        )

    else:

        # พูลแยก → คิดเฉพาะพูลของเครื่องนี้ (SKP กับ SKPTA/SKPLE เป็นคนละชุด ใช้แทนกันไม่ได้)
        # กลุ่มที่ไม่แยก → รวมทุกพูลย่อยของ Type_1+Gauge เหมือนเดิม (พูลติดลบ = จองเกิน ดึงยอดรวมลง)
        _pool_rows = (
            mc_rows[mc_rows["MC_POOL"].astype(str).str.strip() == _pool_label]
            if _is_split else mc_rows
        )
        base_remain = int(_pool_rows["TOTAL_MC_REMAIN"].sum())



    # key สำหรับ pool tracking — พูลแยกใช้ pool label, อื่นๆ ใช้ (type_1, gauge)
    if _is_split:
        already_used = weekly_new_plan_pool_usage.get(week, {}).get(_pool_label, 0)
    else:
        already_used = weekly_new_plan_usage.get(week, {}).get((type_1, gauge_str), 0)

    # รวม cylinder adjustments (factory-agnostic: รวม delta ทุก factory ที่มี mc_cat+gauge เดียวกัน)
    _cyl_adj = sum(
        v for (w, _f, cat, g), v in cylinder_adjustments.items()
        if w == week and cat == type_1 and g == gauge_str
    )

    # _floor: ตัดที่ 0 ตามปกติ แต่ปล่อยค่าติดลบผ่านเมื่อ allow_negative
    _floor = (lambda v: v) if allow_negative else (lambda v: max(0, v))

    result = _floor(base_remain - already_used + _cyl_adj)

    # หักเครื่อง Lock_MC (เครื่องที่ถูก lock ไม่ให้วางแผน)
    _lock_count = _LOCK_MC_MAP.get((int(week), type_1, gauge_str), 0)
    if _lock_count > 0:
        result = _floor(result - _lock_count)

    # MC Special cap: แบ่ง COTTON/POLY sub-pool ตามชีท MC Special
    if _MC_SPECIAL_PLAN:
        _sp_mc_u = str(mc_group).strip().upper()
        _sp_factory = _mc_to_factory(_sp_mc_u, gauge_str)
        _sp_cat = type_1
        _ms_entry = (
            _MC_SPECIAL_PLAN.get((_sp_factory, _sp_cat, _sp_mc_u, gauge_str))
            or _MC_SPECIAL_PLAN.get((_sp_factory, _sp_cat, "", gauge_str))
        )
        if _ms_entry:
            _sp_type = _get_item_cotton_poly(item_code) if item_code else ""
            _cotton_reserved = _ms_entry.get("COTTON", 0)
            _poly_reserved = _ms_entry.get("POLY", 0)
            # โมเดลเดียวกับ AVA_MC: 1 item อยู่ได้พูลเดียว ห้ามล้นข้ามพูล
            # base_remain (SUMMARY_MC_REMAIN) = พูลทั่วไปที่หัก reserved ออกให้แล้ว
            # → item ทั่วไปใช้ base_remain ตรงๆ, COTTON/POLY ใช้เฉพาะ sub-pool ของตัวเอง
            if _sp_type == "COTTON" and _cotton_reserved > 0:
                _sp_used = max(
                    _mc_special_weekly_usage.get((_sp_factory, _sp_cat, gauge_str, week, "COTTON"), 0),
                    _sp_used_from_plans(_sp_factory, _sp_cat, gauge_str, week, "COTTON"),
                )
                _sp_bk = _MC_SPECIAL_BOOKING_USED.get((_sp_cat, gauge_str, week, "COTTON"), 0)
                result = max(0, _cotton_reserved - _sp_bk - _sp_used)
            elif _sp_type == "POLY" and _poly_reserved > 0:
                _sp_used = max(
                    _mc_special_weekly_usage.get((_sp_factory, _sp_cat, gauge_str, week, "POLY"), 0),
                    _sp_used_from_plans(_sp_factory, _sp_cat, gauge_str, week, "POLY"),
                )
                _sp_bk = _MC_SPECIAL_BOOKING_USED.get((_sp_cat, gauge_str, week, "POLY"), 0)
                result = max(0, _poly_reserved - _sp_bk - _sp_used)

    # TYPE_SPECIAL quota check (BABY FRENCH / SINGLE JACQUARD / TWILL)
    if _TYPE_DESC_RULES_PLAN and item_code:
        _ts_mc_u = str(mc_group).strip().upper()
        _ts_fac  = _mc_to_factory(_ts_mc_u, gauge_str)
        _ts_type = _mc_to_type_raw_plan.get((_ts_mc_u, gauge_str), "").strip().upper()
        _ts_rule_key = (_ts_fac.upper(), _ts_type)
        if _ts_rule_key in _TYPE_DESC_RULES_PLAN:
            _ts_rule   = _TYPE_DESC_RULES_PLAN[_ts_rule_key]
            _ts_mc_cat = _ts_rule.get("mc_cat", "")
            _ts_t1     = _mc_to_type1(_ts_mc_u, gauge_str)
            if not ((_ts_mc_cat and _ts_t1 != _ts_mc_cat) or gauge_str == "20"):
                _ts_desc = _item_desc_map_plan.get(str(item_code).strip().upper(), "")
                if _is_description_special_type_plan(_ts_desc, _ts_rule["keywords"]):
                    _ts_max      = _ts_rule["max_mc"]
                    _ts_bk_used  = _type_special_booking_usage.get((_ts_fac, _ts_type, week), 0)
                    _ts_new_used = _type_special_weekly_usage.get((_ts_fac, _ts_type, week), 0)
                    result = min(result, max(0, _ts_max - _ts_bk_used - _ts_new_used))

    return result





def _carry_blocked_by_gap(item_code, mc_key, mc_group, item_gauge, prev_week_idx, current_week_idx):

    """

    ตรวจสอบว่า carryover ถูกบล็อคโดย intermediate week ที่ remaining = 0 จาก item อื่น

    Rule:

      - ถ้า intermediate week (ระหว่าง prev_week_idx+1 ถึง current_week_idx-1) มี remaining = 0

        AND item นี้เองไม่ได้ผลิตใน week นั้น → ตัด carry (return True)

      - ถ้า item นี้เองผลิตอยู่ (remaining = 0 เพราะ item ใช้เครื่องหมด) → ไม่บล็อค

    Returns True ถ้า carry ไม่ได้ (blocked by other item)

    """

    if prev_week_idx is None or current_week_idx is None:

        return False

    if current_week_idx - prev_week_idx <= 1:

        return False  # ติดกัน หรือ same week → ไม่มี intermediate weeks



    for w_idx in range(prev_week_idx + 1, current_week_idx):

        if w_idx >= len(calendar_week):

            break

        w_num = int(calendar_week.iloc[w_idx]["WEEK"])

        remaining = get_actual_mc_remain(mc_group, w_num, item_gauge, item_code=item_code)

        if remaining > 0:

            continue  # ยังมีเครื่องว่าง → ไม่บล็อค week นี้



        # remaining = 0 → ตรวจว่า item นี้เองกำลังผลิตอยู่ใน week นี้ไหม

        item_was_producing = False

        # เช็คจาก booking data

        if booking_mc_by_week.get(mc_key, {}).get(w_idx, 0) > 0:

            item_was_producing = True

        if not item_was_producing and mc_key in booking_active_week_set:

            if w_idx in booking_active_week_set[mc_key]:

                item_was_producing = True

        # เช็คจาก new plans ที่วางไปแล้ว

        if not item_was_producing:

            for pe in plans:

                if pe.get("ITEM_CODE") == item_code and pe.get("MC_GROUP") == mc_group:

                    pe_w_idx = week_index(pe.get("PLAN_WEEK"))

                    if pe_w_idx == w_idx:

                        item_was_producing = True

                        break



        if not item_was_producing:

            prev_w_num = int(calendar_week.iloc[prev_week_idx]["WEEK"]) if prev_week_idx < len(calendar_week) else "?"

            print(f"[CARRY BLOCKED GAP] {item_code} W{w_num}: remaining=0 จาก item อื่น → ตัด carry จาก W{prev_w_num} (machines ถูก item อื่นยึดครบ)")

            return True



    return False





def get_next_fg_orders_for_item(item_code, current_sc_so, current_fg_week, orders_df):

    """ตรวจสอบว่ามี FG ถัดไปของ item เดียวกันหรือไม่ และคืนค่า ORDER_QTY รวม

    Args:

        item_code: รหัส item ที่ต้องการตรวจสอบ

        current_sc_so: SC/SO NO ปัจจุบัน

        current_fg_week: FG Week ปัจจุบัน (YYYYWW format)

        orders_df: DataFrame ของ orders ทั้งหมด

    Returns:

        total_next_qty: ปริมาณรวมของ FG ถัดไปที่ต้องผลิต

    """

    if orders_df is None or orders_df.empty:

        return 0.0

    

    # กรอง order ของ item เดียวกันที่มี FG Week มากกว่า current_fg_week

    same_item_orders = orders_df[orders_df["Item Code"] == item_code].copy()

    

    if same_item_orders.empty:

        return 0.0

    

    total_next_qty = 0.0

    

    for _, next_order in same_item_orders.iterrows():

        next_sc_so = str(next_order.get("SO_NO", next_order.get("SC/SO NO", ""))).strip()

        next_fg_week = next_order.get("FG Week")

        

        # ข้าม row ที่เป็น SC+FG เดียวกันกับปัจจุบัน (exact same row)

        # ไม่ skip same-SC ที่มี FG ต่างกัน เพราะต้องนับ FG ถัดไปของ SC เดียวกัน

        try:

            _curr_fg_int = int(current_fg_week) if pd.notna(current_fg_week) else 0

            _next_fg_int = int(next_fg_week) if pd.notna(next_fg_week) else 0

        except (ValueError, TypeError):

            _curr_fg_int = 0

            _next_fg_int = 0

        if next_sc_so == current_sc_so and _next_fg_int == _curr_fg_int:

            continue

        

        # ตรวจสอบว่า FG Week ถัดไปมากกว่า current หรือไม่

        if pd.notna(next_fg_week):

            try:

                next_fg_int = int(next_fg_week)

                current_fg_int = int(current_fg_week) if pd.notna(current_fg_week) else 0

                

                # ถ้า FG Week ถัดไปมากกว่า current (ผลิตหลัง)

                if next_fg_int > current_fg_int:

                    pending_plan = pd.to_numeric(next_order.get("Pending Plan", 0), errors="coerce")

                    if not pd.isna(pending_plan) and pending_plan > 0:

                        total_next_qty += float(pending_plan)

            except (ValueError, TypeError):

                continue

    

    return total_next_qty





def get_total_pending_qty_for_item(item_code, current_sc_so, current_fg_week, orders_df, same_sc_only=False):

    """รวม Pending Plan ของทุก FG สำหรับ item เดียวกัน (รวมทุก SC/SO)

    ใช้สำหรับคำนวณ machine allocation จาก total demand แทนที่จะดูแค่ FG เดียว

    

    Args:

        same_sc_only: ถ้า True จะรวมเฉพาะ SC/SO เดียวกัน (ใช้สำหรับ carry optimization)

    

    คืนค่า: total_pending_qty รวมของ FG ปัจจุบัน + FG ถัดไปทั้งหมด

    (เฉพาะ FG Week >= current_fg_week เพื่อไม่นับ FG ที่ผ่านไปแล้ว)

    """

    if orders_df is None or orders_df.empty:

        return 0.0

    

    same_item_orders = orders_df[orders_df["Item Code"] == item_code]

    if same_item_orders.empty:

        return 0.0

    

    try:

        current_fg_int = int(current_fg_week) if pd.notna(current_fg_week) else 0

    except (ValueError, TypeError):

        current_fg_int = 0

    

    total_qty = 0.0

    for _, row in same_item_orders.iterrows():

        # ถ้า same_sc_only=True → กรองเฉพาะ SC/SO เดียวกัน

        if same_sc_only:

            row_sc = str(row.get("SO_NO", row.get("SC/SO NO", ""))).strip()

            if row_sc != current_sc_so:

                continue

        

        fg_week = row.get("FG Week")

        try:

            fg_int = int(fg_week) if pd.notna(fg_week) else 0

        except (ValueError, TypeError):

            continue

        

        # รวมเฉพาะ FG Week >= current (FG ปัจจุบัน + อนาคต)

        if fg_int >= current_fg_int:

            pending = pd.to_numeric(row.get("Pending Plan", 0), errors="coerce")

            if not pd.isna(pending) and pending > 0:

                total_qty += float(pending)

    

    return total_qty





# =========================
# FOLD ROUNDING — จำนวนพับ = qty ÷ REVOLUTION_WEIGHT
# ทอครึ่งพับไม่ได้ → ทุกกลุ่มปัดลงเป็นจำนวนเต็มพับ
# กลุ่ม IRMT/SJT ปัดที่ระดับ "ต่อเครื่อง" (เครื่องละกี่พับก็ได้ ขอให้เต็มพับ) แล้วคูณจำนวนเครื่อง
#
# เงื่อนไข "เป็นพับคู่ (หาร 2 ลงตัว)" เป็นเรื่องของ "ยอดรวมทั้ง order" ไม่ใช่ของแต่ละสัปดาห์
# → ตรวจที่ ORDERS_QTY ด้วย fold_check_order() แล้วเตือน (ดู FOLD_WARN ในชีท PLAN)
#   สัปดาห์ไหนจะทอกี่พับก็ได้ ขอแค่รวมกันครบยอด order
# =========================
FOLD_ROUND_MC_GROUPS = {"IRMT", "SJT"}
FOLD_MULTIPLE = 2


def fold_round(total_cap, rev_weight, mc_group, per_mc_cap=None, mc_count=None):
    """ปัด capacity ลงตามจำนวนพับ (จำนวนเต็มพับ — สัปดาห์ละกี่พับก็ได้)
    - กลุ่มอื่น: (total_cap // rev_weight) * rev_weight
    - IRMT/SJT (ต้องส่ง per_mc_cap + mc_count): ปัดจำนวนพับ "ต่อเครื่อง" ลงเป็นจำนวนเต็ม
      แล้วคูณจำนวนเครื่อง
    """
    if not rev_weight or rev_weight <= 0:
        return total_cap
    if (
        per_mc_cap is not None
        and mc_count
        and str(mc_group).strip().upper() in FOLD_ROUND_MC_GROUPS
    ):
        folds = int(per_mc_cap // rev_weight)
        return folds * rev_weight * int(mc_count)
    return int(total_cap // rev_weight) * rev_weight


def fold_round_mixed(
    total_cap, rev_weight, mc_group,
    carry_cap_per_mc, carry_mc, new_cap_per_mc, new_mc,
):
    """เหมือน fold_round แต่รองรับเครื่อง 2 กลุ่มที่ cap/เครื่องไม่เท่ากัน
    (carry = วิ่งเต็มสัปดาห์, new = ถูกหักวัน setup)
    - IRMT/SJT: ปัดจำนวนพับต่อเครื่องเป็นจำนวนเต็ม แยกกลุ่ม carry/new แล้วรวมกัน
    - กลุ่มอื่น: ปัดยอดรวมทั้งก้อนเป็นจำนวนพับเต็ม
    """
    if not rev_weight or rev_weight <= 0:
        return total_cap
    if str(mc_group).strip().upper() in FOLD_ROUND_MC_GROUPS:
        return (
            fold_round(carry_cap_per_mc * carry_mc, rev_weight, mc_group, carry_cap_per_mc, carry_mc)
            + fold_round(new_cap_per_mc * new_mc, rev_weight, mc_group, new_cap_per_mc, new_mc)
        )
    return int(total_cap // rev_weight) * rev_weight


def fold_check_order(orders_qty, rev_weight, mc_groups=None):
    """ตรวจยอดที่ลูกค้าเปิดมาว่าแบ่งเป็นพับแล้วเป็นพับคู่ (หาร 2 ลงตัว) หรือไม่

    ⚠️ กฎ "พับคู่" ใช้เฉพาะกลุ่ม IRMT / SJT เท่านั้น — กลุ่มอื่นไม่มีข้อบังคับนี้
    (mc_groups = MC_GROUP ทั้งหมดที่ order นี้ใช้; ไม่ส่ง = ไม่เช็ค)

    คืน (folds, remainder, warn)
      folds     = จำนวนพับของยอด order (ปัดทศนิยม 2 ตำแหน่ง — เศษพับก็ต้องเห็น)
      remainder = เศษพับที่ไม่เป็นคู่ (0 = เป็นพับคู่พอดี)
      warn      = True เมื่อเป็น IRMT/SJT แล้วหารไม่ลงตัว
    แผนยังวางเต็มตามยอดจริงเสมอ — ค่านี้ใช้เตือน planner ให้ไปแก้ที่ต้นทาง (ยอดเปิด order)
    """
    if not any(
        str(_m).strip().upper() in FOLD_ROUND_MC_GROUPS for _m in (mc_groups or [])
    ):
        return (0.0, 0.0, False)
    try:
        _qty = float(orders_qty or 0)
        _rw = float(rev_weight or 0)
    except (TypeError, ValueError):
        return (0.0, 0.0, False)
    if _qty <= 0 or _rw <= 0:
        return (0.0, 0.0, False)
    folds = round(_qty / _rw, 2)
    remainder = round(folds % FOLD_MULTIPLE, 2)
    return (folds, remainder, remainder != 0)


def compute_backward_start_idx(
    item_code, order_qty, rdd_idx, mc_group, daily_cap, item_gauge, n_machines,
    setup_days=SETUP_DAYS, rev_weight=None, floor_idx=0
):
    """Backward scheduling: หา week index ที่ "ช้าที่สุด" ที่ยังผลิตจบทัน rdd_idx

    ไล่ถอยหลังจาก rdd_idx ลงมาถึง floor_idx สะสม capacity สัปดาห์ต่อสัปดาห์
    จนกว่าจะพอกับ order_qty → คืน index นั้น (= JIT start)

    - target = rdd_idx ตรง ๆ (ไม่มี safety buffer)
    - setup_days หักเฉพาะสัปดาห์แรกที่เริ่มผลิต
    - floor_idx = พื้นแข็ง (TODAY_IDX+2) ห้ามถอยเลยไปกว่านี้
    - คืน None ถ้าไม่ทันแม้เริ่มที่ floor_idx → caller ใช้ ASAP แทน
    """
    if rdd_idx is None or not mc_group or not n_machines or not daily_cap:
        return None
    try:
        _qty_need = float(order_qty or 0)
    except (TypeError, ValueError):
        return None
    if _qty_need <= 0:
        return None

    floor_idx = max(0, int(floor_idx))
    rdd_idx = int(rdd_idx)
    if rdd_idx < floor_idx or rdd_idx >= len(calendar_week):
        return None

    # cache ข้อมูลรายสัปดาห์ (avail / working days) ช่วง floor_idx..rdd_idx ครั้งเดียว
    _wk_info = {}
    for _i in range(floor_idx, rdd_idx + 1):
        _wk = int(calendar_week.iloc[_i]["WEEK"])
        if _wk in SKIP_WEEKS:
            _wk_info[_i] = {"week": _wk, "avail": 0, "wd": 0}
            continue
        _wk_info[_i] = {
            "week": _wk,
            "avail": get_actual_mc_remain(mc_group, _wk, gauge=item_gauge, item_code=item_code),
            "wd": get_working_days_by_factory(mc_group, 1, week=_wk, item_code=item_code, gauge=item_gauge),
        }

    # ไล่ถอยหลัง: หา _start ที่ช้าที่สุดที่ capacity สะสม (_start..rdd_idx) >= qty
    for _start in range(rdd_idx, floor_idx - 1, -1):
        _total_cap = 0.0
        for _i in range(_start, rdd_idx + 1):
            _w = _wk_info[_i]
            # setup หักเฉพาะสัปดาห์แรกที่เริ่มผลิตจริง
            _pd = max(0, _w["wd"] - setup_days) if _i == _start else _w["wd"]
            if _pd <= 0:
                continue
            _use_mc = min(int(n_machines), _w["avail"])
            if _use_mc <= 0:
                continue
            _total_cap += fold_round(
                _use_mc * _pd * daily_cap, rev_weight, mc_group,
                per_mc_cap=_pd * daily_cap, mc_count=_use_mc,
            )
        if _total_cap >= _qty_need:
            return _start

    return None  # ไม่ทันแม้เริ่มที่ floor_idx → ต้อง ASAP


def calculate_progressive_reduction(

    item_code, order_qty, start_week, fg_week, mc_group, daily_cap, item_gauge,

    setup_days=SETUP_DAYS, rev_weight=None

):

    """คำนวณจำนวนเครื่องแต่ละ week แบบประหยัดที่สุด แต่ให้เสร็จพอดี target week

    Strategy:

    1. หาจำนวนเครื่องน้อยที่สุดที่ทัน target (fixed machines)

    2. ถ้าเสร็จเร็วกว่า target → กระจายเครื่องให้น้อยลงในแต่ละ week

    3. เริ่มต้นด้วยเครื่องมากกว่า แล้วค่อยๆ ลดลงให้เสร็จพอดี target week

    Returns: list of (week, machines) หรือ None ถ้าไม่สามารถทันได้

    """

    weeks_until_target = []

    current_week = start_week

    target_week_index = fg_week  # TARGET_KNIT



    # สร้าง list ของ weeks จาก start_week ถึง TARGET_KNIT โดยเด็ดขาด
    # target_week_index เป็น None ได้ (RDD นอกช่วง calendar) → ข้าม loop ให้คืน None ตาม logic เดิม

    _wut_guard = 0

    while current_week is not None and target_week_index is not None and week_index(current_week) <= target_week_index:

        _wut_guard += 1  # SAFETY: กัน infinite loop ตอนข้ามปลายปี (week_index จับผิดปี → ไม่เดินหน้า)

        if _wut_guard > 300:

            break

        weeks_until_target.append(current_week)

        current_week = next_week(current_week)

    if not weeks_until_target:

        return None



    # เก็บ availability และ working days ของแต่ละ week

    week_info = []

    for week in weeks_until_target:

        actual_remain = get_actual_mc_remain(mc_group, week, gauge=item_gauge, item_code=item_code)

        # วันทำงานจากชีท Work Day เป็นค่าสุดท้ายแล้ว (ไม่หักวันหยุด/ไม่ cap ด้วยปฏิทิน)
        actual_wd = get_working_days_by_factory(mc_group, 1, week=week, item_code=item_code, gauge=item_gauge)

        week_info.append({

            'week': week,

            'avail': actual_remain,

            'wd': actual_wd

        })

    # Step 1: หาจำนวนเครื่องน้อยที่สุดที่ทัน (fixed machines ทุก week)

    min_machines = None

    for try_mc in range(1, max(w['avail'] for w in week_info) + 1):

        qty_left = order_qty

        for i, w in enumerate(week_info):

            if qty_left <= 0:

                break



            prod_days = max(0, w['wd'] - setup_days) if i == 0 else w['wd']

            if prod_days <= 0:

                continue



            use_mc = min(try_mc, w['avail'])

            prod = fold_round(
                use_mc * prod_days * daily_cap, rev_weight, mc_group,
                per_mc_cap=prod_days * daily_cap, mc_count=use_mc,
            )

            qty_left -= prod

        if qty_left <= 0:

            min_machines = try_mc

            break



    if min_machines is None:

        return None  # ไม่ทันแม้ใช้เครื่องเต็มที่



    # Step 1b: หา optimal_start_idx — เลื่อน start week ให้ช้าที่สุดที่ยังจบทัน TARGET_KNIT

    # ผลลัพธ์: carryover-first จะผลิตตั้งแต่ optimal_start_idx ถึง TARGET_KNIT พอดี

    target_idx = len(week_info) - 1

    optimal_start_idx = 0  # default: เริ่มจาก week แรก

    for _start in range(target_idx, -1, -1):

        _total_cap = 0

        for _i in range(_start, target_idx + 1):

            _pd_i = max(0, week_info[_i]['wd'] - setup_days) if _i == _start else week_info[_i]['wd']

            _use_i = min(min_machines, week_info[_i]['avail'])

            _p = fold_round(
                _use_i * _pd_i * daily_cap, rev_weight, mc_group,
                per_mc_cap=_pd_i * daily_cap, mc_count=_use_i,
            )

            _total_cap += _p

        if _total_cap >= order_qty:

            optimal_start_idx = _start

            break  # พบ start ช้าที่สุดที่ยังทัน TARGET_KNIT

    if optimal_start_idx > 0:

        print(f"[TARGET_KNIT START] {item_code}: เลื่อนเริ่มจาก W{week_info[0]['week']} → W{week_info[optimal_start_idx]['week']} (TARGET W{week_info[target_idx]['week']})")



    # Step 2: Carryover-first strategy - ใช้เครื่องเต็มจำนวนจนกว่า qty จะเหลือน้อย

    # ลดเครื่องเฉพาะเมื่อ qty ที่เหลือน้อยกว่า capacity ของ 1 สัปดาห์

    # เริ่มผลิตจาก optimal_start_idx (ไม่ใช่ week 0) เพื่อให้จบที่ TARGET_KNIT

    result = []

    qty_left = order_qty

    num_weeks = len(week_info)

    prev_week_mc = min_machines  # เริ่มต้นด้วย min_machines

    

    for i, w in enumerate(week_info):

        if i < optimal_start_idx:

            # ยังไม่ถึง optimal start → ไม่ผลิต

            result.append((w['week'], 0))

            continue



        if qty_left <= 0:

            # เสร็จแล้ว แต่ยังมี week เหลือ → ไม่ผลิต

            result.append((w['week'], 0))

            continue



        # Setup days ใช้เฉพาะ week แรกที่ผลิตจริง (optimal_start_idx)

        prod_days = max(0, w['wd'] - setup_days) if i == optimal_start_idx else w['wd']

        if prod_days <= 0 or w['avail'] <= 0:

            result.append((w['week'], 0))

            continue



        # Strategy: ใช้เครื่องเต็มจำนวน (min_machines) จนกว่า qty จะเหลือน้อย

        # ลดเครื่องเฉพาะเมื่อ qty ที่เหลือผลิตไม่เต็ม capacity ของเครื่องทั้งหมด

        weeks_remaining = num_weeks - i

        

        # คำนวณ capacity ต่อสัปดาห์ของเครื่องทั้งหมด

        full_week_capacity = fold_round(
            min_machines * prod_days * daily_cap, rev_weight, mc_group,
            per_mc_cap=prod_days * daily_cap, mc_count=min_machines,
        )

        

        if qty_left >= full_week_capacity:

            # qty เหลือมาก → ใช้เครื่องเต็มจำนวน

            needed_mc = max(1, int(qty_left / (prod_days * daily_cap)) + 1)

            use_mc = min(min_machines, needed_mc, w['avail'])

            if use_mc < min_machines:

                print(f"[PROGRESSIVE REDUCTION FIX] Week {w['week']}: min_machines={min_machines} > needed_mc={needed_mc} → use {use_mc}")

        else:

            # qty เหลือน้อย → คำนวณเครื่องที่ต้องการจริงๆ

            needed_mc = max(1, int(qty_left / (prod_days * daily_cap)) + 1)

            use_mc = min(needed_mc, w['avail'], min_machines)

            

            # Gradual Reduction: ลดได้สูงสุด 2 เครื่องต่อสัปดาห์

            if i > optimal_start_idx and use_mc < prev_week_mc:

                max_reduction = max(1, prev_week_mc - 2)

                if use_mc < max_reduction:

                    use_mc = max_reduction

        

        # คำนวณ production จริง

        # เครื่องใหม่ที่เพิ่มกลางแผน (ไม่ใช่สัปดาห์แรก) ต้องหัก setup_days ของตัวเอง

        if i > optimal_start_idx and use_mc > prev_week_mc:

            carry_in_wk = prev_week_mc

            added_in_wk = use_mc - carry_in_wk

            prod = fold_round_mixed(
                (carry_in_wk * prod_days + added_in_wk * max(0, prod_days - setup_days)) * daily_cap,
                rev_weight, mc_group,
                prod_days * daily_cap, carry_in_wk,
                max(0, prod_days - setup_days) * daily_cap, added_in_wk,
            )

        else:

            prod = fold_round(
                use_mc * prod_days * daily_cap, rev_weight, mc_group,
                per_mc_cap=prod_days * daily_cap, mc_count=use_mc,
            )

        result.append((w['week'], use_mc))

        qty_left -= prod

        prev_week_mc = use_mc  # บันทึกเครื่องของ week นี้

    # ถ้ายังเหลือ qty หลังจาก loop ครบ → ไม่ทัน (แต่ไม่น่าเกิดเพราะ min_machines ทันแล้ว)

    if qty_left > 0:

        return None



    return result





def calculate_required_machines(

    item_code, order_qty, start_week, fg_week, setup_days=SETUP_DAYS, only_mc_group=None,

    order_type="", sub_color="", dye_end_date=None,

):

    print(f"[DEBUG CALC] calculate_required_machines called: {item_code}, qty={order_qty}, start={start_week}, target={fg_week}")

    """คำนวณจำนวนเครื่องขั้นต่ำที่ต้องการเพื่อทัน RDD

    หลักการ: ใช้เครื่องน้อยแต่ผลิตหลาย week ดีกว่าใช้เครื่องเยอะแค่ 1 week

    - setup เป็น per-machine: 3mc setup = เสีย 3×3=9 mc-days

    - week 2+ ไม่ต้อง setup → ได้ผลิตเต็มที่

    - simulate per-week ด้วยเครื่องว่างจริงของแต่ละ week (cap at n_mc)

    ตัวอย่าง order 3277.5, cap=163, factory 7d, เครื่องว่าง [6, 1, 5]:

      6mc×3wk: wk1=6×4×163=3912, wk2=1×7×163=1141, wk3=5×7×163=5705 → setup_waste=18

      2mc×3wk: wk1=2×4×163=1304, wk2=1×7×163=1141, wk3=2×7×163=2282 → setup_waste=6

    """

    # fg_week (RDD index) เป็น None ได้เมื่อ RDD อยู่นอกช่วง calendar
    # → ใช้ start_week + 3 เป็น target แทน (เร่งจบ เหมือนกรณี past RDD)
    if fg_week is None:
        _sw_idx_norm = week_index(start_week)
        if _sw_idx_norm is None:
            return None, None, None, None, None
        fg_week = _sw_idx_norm + 3
        print(f"[DEBUG CALC] {item_code}: fg_week=None (RDD นอกช่วง calendar) → ใช้ start+3 = {fg_week}")

    # หา MC_GROUP ที่สามารถผลิต item นี้ได้

    available_machines = item_cap_data[item_cap_data["ITEM_CODE"] == item_code]

    if available_machines.empty:

        return None, None, None, None, None



    # ใช้ความจุตาม logic ใหม่: สำหรับ SHARED_POOL_MAP ให้เลือกจาก cap ที่น้อยที่สุดในกลุ่มเดียวกัน

    # และ FA 20 มี priority สูงกว่า SKP 20

    # ใช้ความจุของ MC_GROUP แรกที่พบเป็นค่าเริ่มต้น และจะปรับเมื่อวนลูปแต่ละ MC_GROUP

    daily_cap = None

    # เรียงตาม MC_GROUP ที่มีเครื่องเหลือมากที่สุดก่อน (start_week)

    available_machines = available_machines.copy()

    available_machines["_mc_remain"] = available_machines.apply(

        lambda r: get_actual_mc_remain(r["MC_GROUP"], start_week, gauge=r.get("GUAGE"), item_code=item_code),

        axis=1,

    )

    available_machines = available_machines.sort_values("_mc_remain", ascending=False)

    # ถ้ามี only_mc_group → บังคับใช้ MC_GROUP นั้น (lock สำหรับ SC/SO+Item เดิม)

    if only_mc_group is not None:

        _filt = available_machines[available_machines["MC_GROUP"] == only_mc_group]

        if not _filt.empty:

            available_machines = _filt

    # คำนวณจำนวนสัปดาห์ที่เหลือถึง TARGET_KNIT (บังคับให้จบตรง TARGET_KNIT)

    # fg_week คือ TARGET_KNIT index (row index ใน calendar_week) เพื่อรองรับข้ามปีได้

    weeks_until_target = []

    current_week = start_week

    target_week_index = fg_week  # TARGET_KNIT

    start_week_idx = week_index(start_week)

    

    # สร้าง list ของ weeks จาก start_week ถึง TARGET_KNIT โดยเด็ดขาด
    # target_week_index เป็น None ได้ (RDD นอกช่วง calendar) → ข้าม loop ให้เข้า branch past RDD ด้านล่าง

    # เดินด้วย row index ตรง ๆ (idx เพิ่มขึ้นเสมอ → จบแน่นอน)
    # กัน infinite loop ที่เกิดจาก week_index(bare week) จับผิดปีตอนข้ามปลายปี
    _iter_idx = start_week_idx
    while (_iter_idx is not None and target_week_index is not None
           and _iter_idx <= target_week_index and _iter_idx < len(calendar_week)):
        weeks_until_target.append(int(calendar_week.iloc[_iter_idx]["WEEK"]))
        _iter_idx += 1
        while _iter_idx < len(calendar_week) and int(calendar_week.iloc[_iter_idx]["WEEK"]) in SKIP_WEEKS:
            _iter_idx += 1



    # ถ้า target อยู่ในอดีต (past RDD) → ใช้ weeks จาก start_week ถึง plan_week+3 เพื่อให้รีบเสร็จ

    # เพื่อให้คำนวณ required machines สูงขึ้นและเพิ่มเครื่องค่อยๆ (Gradual Increase)

    if not weeks_until_target:

        print(f"[DEBUG PAST RDD] Target {target_week_index} < Start {start_week_idx} → using weeks to plan_week+3 for faster completion")

        target_for_past_rdd = start_week_idx + 3  # จำกัด 3 weeks เพื่อให้รีบเสร็จ → คำนวณเครื่องสูงขึ้น

        # เดินด้วย row index ตรง ๆ (กัน infinite loop จาก week_index จับผิดปี)
        _iter_idx = start_week_idx
        while (_iter_idx is not None and _iter_idx <= target_for_past_rdd
               and _iter_idx < len(calendar_week)):
            weeks_until_target.append(int(calendar_week.iloc[_iter_idx]["WEEK"]))
            _iter_idx += 1
            while _iter_idx < len(calendar_week) and int(calendar_week.iloc[_iter_idx]["WEEK"]) in SKIP_WEEKS:
                _iter_idx += 1

    

    # ตรวจสอบว่า week สุดท้ายคือ TARGET_KNIT จริงๆ

    last_week_index = week_index(weeks_until_target[-1])

    if last_week_index != target_week_index:

        print(f"⚠️  TARGET_KNIT mismatch: last week {last_week_index} != target {target_week_index}")

    

    num_weeks = len(weeks_until_target)

    # ลองแต่ละ MC_GROUP ที่สามารถผลิตได้ (เรียงตาม cap น้อยไปมาก — ใช้ cap ต่ำสุดในการคำนวณ)

    for _, machine_row in available_machines.iterrows():

        mc_group = machine_row["MC_GROUP"]

        # ใช้ความจุตาม logic ใหม่สำหรับ MC_GROUP นี้

        item_gauge = machine_row["GUAGE"] if "GUAGE" in machine_row else None

        daily_cap = _get_capacity_for_mc_group(item_code, mc_group, item_gauge)

        daily_cap = adjust_daily_cap_for_item_special(daily_cap, item_code, mc_group, item_gauge)

        if daily_cap <= 0:

            continue  # ข้ามถ้าไม่มีความจุ

        # เก็บจำนวนเครื่องว่างจริงของแต่ละ week

        avail_per_week = []

        has_any_machine = False

        for week in weeks_until_target:

            actual_remain = get_actual_mc_remain(mc_group, week, gauge=item_gauge, item_code=item_code)

            avail_per_week.append(actual_remain)

            if actual_remain > 0:

                has_any_machine = True

        # เครื่องที่วิ่งอยู่แล้ว (carry-over) ถือว่า "มี" เครื่องพร้อมผลิตโดยไม่ต้องดู actual_remain

        _key_check = _resolve_carry_key(item_code, mc_group, item_gauge)

        if not has_any_machine and machines_in_use.get(_key_check, 0) <= 0:

            continue



        # ---- Setup-aware: ตรวจสอบว่าต้อง setup หรือไม่ ----

        key = _resolve_carry_key(item_code, mc_group, item_gauge)

        setup_needed = True

        start_week_idx = week_index(start_week)

        _bridge_locks(key, start_week_idx)

        if key in last_production:

            last_week_idx = last_production[key]

            if start_week_idx - last_week_idx <= SETUP_GAP_WEEK:

                setup_needed = False

        # เครื่องที่วิ่งอยู่แล้ว (carry-over จาก booking/old plan)

        # ถ้า setup_needed=False = เครื่องยังอุ่นอยู่ → ใช้เป็น committed_mc ตั้งต้น

        carryover_start = machines_in_use.get(key, 0) if not setup_needed else 0

        # YD-ORDERS: ถ้า วันนัดย้อม เปลี่ยน → carryover ต้อง setup เพิ่ม 1 วัน (เฉพาะ week แรก)

        _yd_color_setup_days = 0

        if order_type == "YD-ORDERS" and carryover_start > 0:

            _dye_end_date = dye_end_date

            if pd.notna(_dye_end_date):

                _prev_dye_date = last_dye_end_date.get(key, None)

                if _prev_dye_date and _prev_dye_date != _dye_end_date:

                    _yd_color_setup_days = 1

        factory_wd = get_working_days_by_factory(mc_group, 1, week=start_week, item_code=item_code, gauge=item_gauge)

        # หาจำนวนเครื่องสูงสุดที่สามารถลองได้ (จาก week ที่มีเครื่องมากที่สุด)

        max_possible = max(avail_per_week)

        # จำกัดตาม job/week capacity (รวม type ทั้งหมด ไม่ใช่แค่ MC_GROUP เดียว)

        type_used_start = get_type_used_jobs(start_week, mc_group)

        max_try = check_job_capacity_limit(

            mc_group,

            int(max_possible),

            urgent_mode=False,

            current_week_jobs=type_used_start,

        )

        # carry-over machines ไม่ต้อง setup ไม่ต้องนับเป็น new job

        # ดังนั้น max_try ต้องอย่างน้อย = carryover_start

        if carryover_start > 0 and max_try < carryover_start:

            max_try = carryover_start



        # ไม่ใช้ tolerance แบบ 1 batch

        # ต้องผลิตให้ครบจริง ถ้าไม่พอต้องเพิ่มเครื่อง

        _rw_tol = 0.0



        # ---- เปรียบเทียบทุก option ด้วย per-week simulation ----

        best_option = None  # (n_machines, weeks_needed, setup_waste, efficiency)

        

        # คำนวณเครื่องที่เหมาะสมเพื่อจบตรง TARGET_KNIT

        weeks_available = target_week_index - week_index(weeks_until_target[0]) + 1

        # ถ้า weeks_available เป็นค่าลบ (past RDD) → ใช้จำนวน weeks จริงที่มี

        if weeks_available <= 0:

            weeks_available = len(weeks_until_target)

        if weeks_available > 0:

            # คำนวณเครื่องที่ต้องการเพื่อกระจายผลิตจนถึง TARGET_KNIT

            total_production_needed = order_qty

            total_days_available = weeks_available * factory_wd

            optimal_mc = max(1, int(total_production_needed / (total_days_available * daily_cap) * 1.1))  # +10% buffer

            print(f"[DEBUG TARGET] Optimal machines for {item_code}: {optimal_mc} (target {target_week_index}, weeks {weeks_available})")

        

        for n_mc in range(1, int(max_try) + 1):

            # Simulate: ต้องการ n_mc เครื่อง แต่ละ week อาจได้ไม่ครบตาม availability

            # เครื่องที่เพิ่มใหม่ต้อง setup, เครื่องที่ carry-over ไม่ต้อง setup

            qty_remaining = order_qty

            weeks_needed = 0

            # เริ่มต้น simulation ด้วยเครื่องที่วิ่งอยู่แล้ว (ถ้า setup_needed=False)

            committed_mc = min(carryover_start, n_mc)  # ไม่เกิน target n_mc

            total_setup_mc_days = 0

            actual_use_list = []

            actual_wd_list = []  # เก็บ actual working days ของแต่ละ week

            for w_idx, week in enumerate(weeks_until_target):

                if qty_remaining <= _rw_tol:

                    break



                # วันทำงานจากชีท Work Day เป็นค่าสุดท้าย (จัดการยุบสัปดาห์ + สัปดาห์ปิดทั้งก้อน=0 แล้ว)

                actual_wd = factory_wd

                # ถ้า summary_mc ไม่มีข้อมูลในสัปดาห์นี้ แต่เครื่องกำลังวิ่งอยู่ (carry-over)

                # ให้เครื่องเดิมยังคงผลิตต่อได้ (ไม่ต้องมีข้อมูลใน summary_mc)

                avail_this_week = avail_per_week[w_idx]

                if avail_this_week <= 0 and committed_mc > 0:

                    avail_this_week = committed_mc  # carry-over เท่านั้น ไม่เพิ่มเครื่องใหม่

                # จำนวนเครื่องที่ต้องการใน week นี้ (ไม่เกิน availability)

                want_mc = min(n_mc, avail_this_week)

                if want_mc <= 0:

                    actual_use_list.append(0)

                    actual_wd_list.append(actual_wd)

                    continue



                # แยก carry-over vs ใหม่

                carryover = min(committed_mc, want_mc)

                new_added = (

                    want_mc - carryover

                )  # ไม่มี MAX_SETUP_MC → job/week cap ควบคุมแทน

                want_mc = carryover + new_added

                if committed_mc == 0 and setup_needed:

                    # week แรกที่เริ่มผลิต (cold start): ทุกเครื่องต้อง setup

                    setup_mc = want_mc

                    want_mc = setup_mc

                    prod_days_carry = 0

                    prod_days_new = max(0, actual_wd - setup_days)

                elif new_added > 0 and (setup_needed or committed_mc > 0):

                    # มีเครื่องเพิ่มใหม่นอกเหนือจาก carryover → เฉพาะเครื่องใหม่ต้อง setup

                    setup_mc = new_added if (setup_needed or committed_mc > 0) else 0

                    prod_days_carry = actual_wd

                    prod_days_new = (

                        max(0, actual_wd - setup_days) if setup_mc > 0 else actual_wd

                    )

                else:

                    # carry-over ล้วน หรือ warm start (setup_needed=False, committed_mc=0)

                    setup_mc = 0

                    prod_days_carry = actual_wd

                    prod_days_new = actual_wd  # warm → ผลิตเต็มสัปดาห์ที่เปิดจริง

                # YD-ORDERS: SUB_COLOR เปลี่ยน → carryover หัก 1 วัน (เฉพาะ week แรก)

                if _yd_color_setup_days > 0 and w_idx == 0 and carryover > 0:

                    prod_days_carry = max(0, prod_days_carry - _yd_color_setup_days)

                total_setup_mc_days += setup_mc * setup_days

                committed_mc = want_mc  # อัปเดตเครื่องที่ใช้จริง

                weeks_needed += 1

                actual_use_list.append(want_mc)

                actual_wd_list.append(actual_wd)

                week_production = (

                    carryover * prod_days_carry + new_added * prod_days_new

                ) * daily_cap

                qty_remaining -= week_production

            finished = qty_remaining <= _rw_tol



            setup_waste = total_setup_mc_days  # mc-days ที่เสียไปกับ setup

            # คำนวณ efficiency (ใช้ actual working days ของแต่ละ week)

            total_machine_days = sum(

                mc * wd for mc, wd in zip(actual_use_list, actual_wd_list) if mc > 0
            )
            productive_days = max(0, total_machine_days - setup_waste)
            efficiency = (
                (productive_days / total_machine_days * 100)
                if total_machine_days > 0
                else 0
            )

            if finished:

                # ตรวจสอบว่าจบตรง TARGET_KNIT หรือ TARGET_KNIT-1 หรือไม่

                last_week_produced = weeks_until_target[w_idx]

                last_week_index = week_index(last_week_produced)

                target_week_index = fg_week  # TARGET_KNIT

                

                print(f"[DEBUG TARGET] {item_code}: finished at week {last_week_index}, target {target_week_index}")

                

                if last_week_index == target_week_index or last_week_index == target_week_index - 1:

                    # จบตรง TARGET_KNIT หรือ TARGET_KNIT-1 → ยอมรับ option นี้

                    print(f"[DEBUG TARGET] ✅ Acceptable match: {item_code} finishes at {last_week_index} (target {target_week_index}, tolerance -1)")

                    if best_option is None or n_mc < best_option[0]:

                        best_option = (n_mc, weeks_needed, total_setup_mc_days, efficiency)

                    continue

                elif last_week_index < target_week_index - 1:

                    # จบก่อน TARGET_KNIT-1 → ต้องใช้เครื่องน้อยลงเพื่อกระจายไปจนถึง TARGET_KNIT

                    print(f"[DEBUG TARGET] ❌ Too early: {item_code} finishes at {last_week_index}, target {target_week_index}")

                    # คำนวณว่าต้องลดเครื่องกี่เครื่องให้จบตรง TARGET_KNIT

                    weeks_available = target_week_index - week_index(weeks_until_target[0]) + 1

                    if weeks_available > 0:

                        # คำนวณเครื่องที่ต้องการเพื่อกระจายผลิตจนถึง TARGET_KNIT

                        avg_daily_needed = order_qty / (weeks_available * factory_wd * daily_cap)

                        suggested_mc = max(1, int(avg_daily_needed * 0.8))  # ใช้ 80% เพื่อความปลอดภัย

                        if suggested_mc < n_mc:

                            # ลองใหม่ด้วยเครื่องน้อยลง

                            continue

                    continue

                else:

                    # เลย TARGET_KNIT → ไม่ใช่ option นี้

                    print(f"[DEBUG TARGET] ❌ Too late: {item_code} finishes at {last_week_index}, target {target_week_index}")

                    continue



        if best_option:

            required_machines = best_option[0]

            return mc_group, daily_cap, required_machines, True, item_gauge  # feasible



        else:

            # ไม่มี option ที่จบตรง TARGET_KNIT → หา option ที่ใกล้เคียงที่สุด

            print(f"[DEBUG TARGET] ⚠️ No perfect match for {item_code}, finding closest option...")

            

            # คำนวณ option ที่ใกล้เคียงที่สุด

            best_option = None

            min_distance = float('inf')

            

            for n_mc in range(1, int(max_try) + 1):

                # Simulate อีกครั้งเพื่อหา option ที่ใกล้เคียงที่สุด

                qty_remaining = order_qty

                weeks_needed = 0

                committed_mc = min(carryover_start, n_mc)

                total_setup_mc_days = 0

                

                for w_idx, week in enumerate(weeks_until_target):

                    if qty_remaining <= _rw_tol:

                        break

                    

                    actual_wd = factory_wd  # ค่าสุดท้ายจากชีท Work Day

                    avail_this_week = avail_per_week[w_idx]

                    if avail_this_week <= 0 and committed_mc > 0:

                        avail_this_week = committed_mc

                    

                    want_mc = min(n_mc, avail_this_week)

                    if want_mc <= 0:

                        continue

                    

                    # คำนวณ production และ setup

                    new_added = max(0, want_mc - committed_mc)

                    if new_added > 0:

                        setup_mc_days = new_added * setup_days

                        if setup_mc_days > actual_wd * want_mc:

                            continue

                        total_setup_mc_days += setup_mc_days

                        committed_mc = want_mc

                    

                    prod_days_new = actual_wd - setup_days

                    prod_days_carry = actual_wd

                    # YD-ORDERS: SUB_COLOR เปลี่ยน → carryover หัก 1 วัน (เฉพาะ week แรก)

                    if _yd_color_setup_days > 0 and w_idx == 0 and committed_mc > 0:

                        prod_days_carry = max(0, prod_days_carry - _yd_color_setup_days)

                    week_production = (

                        committed_mc * prod_days_carry + new_added * prod_days_new

                    ) * daily_cap

                    qty_remaining -= week_production

                    weeks_needed += 1

                

                if qty_remaining <= _rw_tol:

                    last_week_index = week_index(weeks_until_target[min(w_idx, len(weeks_until_target)-1)])

                    distance = abs(last_week_index - target_week_index)

                    

                    if distance < min_distance:

                        min_distance = distance

                        best_option = (n_mc, weeks_needed, total_setup_mc_days, 0)

            

            if best_option:

                print(f"[DEBUG TARGET] 🎯 Using closest option for {item_code}: distance {min_distance} weeks")

                required_machines = best_option[0]

                return mc_group, daily_cap, required_machines, True, item_gauge  # feasible (closest)

            else:

                # ไม่ทันทุก option → ใช้ optimal_mc ที่คำนวณไว้แล้ว + บอก caller ว่า NOT feasible

                print(f"[DEBUG TARGET] ❌ No feasible option for {item_code}, using optimal machines")

                required_machines = max(1, int(optimal_mc)) if 'optimal_mc' in locals() else int(max_try)

                return (

                    mc_group,

                    daily_cap,

                    required_machines,

                    False,

                    item_gauge,

                )  # not feasible

    return None, None, None, None, None





def get_best_machine_for_item(

    item_code,

    plan_week,

    last_production,

    required_machines_info=None,

    urgent_mode=False,

    past_rdd=False,

    force_max_mc=False,

):

    print(f"[DEBUG BEST MC] Processing {item_code} for week {plan_week}")

    print(f"[DEBUG BEST MC] Total items processed: {len([x for x in globals().get('_processed_items', [])])}")

    if '_processed_items' not in globals():

        globals()['_processed_items'] = []

    globals()['_processed_items'].append(f"{item_code}_{plan_week}")

    

    # ถ้ามีการคำนวณจำนวนเครื่องที่ต้องการมาแล้ว ให้ใช้ค่านั้น

    if required_machines_info is not None:

        mc_group, daily_cap, required_machines, *_ = required_machines_info

        print(f"[DEBUG BEST MC] Using pre-calculated: {mc_group}, {required_machines} machines")

        # ดึง gauge จาก required_machines_info (ตำแหน่งที่ 4 ถ้ามี)

        _rmi_gauge = (

            required_machines_info[4] if len(required_machines_info) > 4 else None

        )

        if mc_group and required_machines > 0:

            # หา GUAGE ของ item นี้

            item_machine_info = item_cap_data[

                (item_cap_data["ITEM_CODE"] == item_code)

                & (item_cap_data["MC_GROUP"] == mc_group)

            ]

            item_gauge = (

                item_machine_info.iloc[0]["GUAGE"]

                if not item_machine_info.empty

                else _rmi_gauge

            )

            # ดูเครื่องว่างจริง (หักที่จองไปแล้ว)

            actual_remain = get_actual_mc_remain(mc_group, plan_week, gauge=item_gauge, item_code=item_code)

            # ตรวจสอบว่าเคยผลิต item นี้ใน week ก่อน (= เครื่องเดิม carry over)

            key = _resolve_carry_key(item_code, mc_group, item_gauge)

            setup_needed = True

            current_week_idx = week_index(plan_week)

            is_continuing = False  # order เดิมกำลังผลิตต่อจาก week ก่อน

            _bridge_locks(key, current_week_idx)

            if key in last_production:

                last_week_idx = last_production[key]

                if current_week_idx - last_week_idx <= SETUP_GAP_WEEK:

                    setup_needed = False

                if current_week_idx - last_week_idx == 1:

                    is_continuing = True  # week ติดกัน = เครื่องเดิม carry over

                # carry-over จาก old plan: SC/SO NO เดิม และยังไม่ได้เริ่มผลิตใน new plan

                same_sc = last_sc_so_no.get(key) == sc_so_no

                if same_sc:  # SC/SO เดิม → carry over เสมอ (รวมกรณี FG Week ต่างกัน)

                    is_continuing = True

                    setup_needed = False

                else:

                    # Optionally allow carryover across different SC/SO if configured

                    if ALLOW_CARRYOVER_ACROSS_SO:

                        prev_m = machines_in_use.get(key, 0)

                        last_idx = last_production.get(key)

                        if (

                            prev_m > 0

                            and last_idx is not None

                            and current_week_idx - last_idx <= SETUP_GAP_WEEK

                        ):

                            is_continuing = True

                            setup_needed = False

            if is_continuing:

                # เครื่องเดิมจาก week ก่อน carry over โดยไม่ต้องเช็ค actual_remain

                # ใช้ fallback=0 ตรงกับ main loop เพื่อป้องกัน committed_carryover ผิด

                prev_mc = machines_in_use.get(key, 0)

                carryover = prev_mc  # เครื่องทั้งหมดจาก week ก่อนวิ่งต่อได้เลย

                # ถ้า feasible=True: cap ที่ required_machines (คำนวณมาแล้วว่า N เครื่องพอตั้งแต่ต้น)

                # ถ้า feasible=False: ใช้เต็มที่ (ไม่ทันด้วย N เครื่อง เปิดทุกสล็อที่มี)

                _is_feasible = (

                    required_machines_info[3]

                    if required_machines_info and len(required_machines_info) > 3

                    else True

                )

                extra_avail = max(0, actual_remain)

                if _is_feasible:

                    can_add = max(0, required_machines - carryover)

                    new_additions = min(extra_avail, can_add, MAX_NEW_SETUP_MC)

                else:

                    new_additions = min(extra_avail, MAX_NEW_SETUP_MC)

                available_machines_count = carryover + new_additions

                type_used = get_type_used_jobs(plan_week, mc_group)

                # ส่ง committed_carryover=carryover เพื่อให้ carryover ผ่าน cap เสมอ

                available_machines_count = check_job_capacity_limit(

                    mc_group,

                    available_machines_count,

                    urgent_mode,

                    type_used,

                    committed_carryover=carryover,

                )

                return (



                    mc_group,

                    daily_cap,

                    setup_needed,

                    available_machines_count,

                    item_gauge,

                )

            if actual_remain > 0:

                _is_feasible = (

                    required_machines_info[3]

                    if required_machines_info and len(required_machines_info) > 3

                    else True

                )

                if _is_feasible:

                    # feasible: cap ที่ required_machines เพื่อไม่ใช้เครื่องเกินความจำเป็น

                    available_machines_count = min(required_machines, actual_remain)

                else:

                    # not feasible: ใช้เต็มที่ + main loop จะใช้ _forward_sim หา minimum

                    available_machines_count = actual_remain

                type_used = get_type_used_jobs(plan_week, mc_group)
                available_machines_count = check_job_capacity_limit(

                    mc_group, available_machines_count, urgent_mode, type_used

                )

                if available_machines_count <= 0:

                    # MC_GROUP จาก required_machines_info เต็ม → ลอง MC_GROUP อื่นที่ item มี CAP

                    # (fallthrough ไปใช้ logic ด้านล่างที่ลองทุก MC_GROUP)

                    pass



                else:

                    return (



                        mc_group,

                        daily_cap,

                        setup_needed,

                        available_machines_count,

                        item_gauge,

                    )

    # เครื่องสำรอง: ใช้ logic เดิมถ้าไม่มีการคำนวณล่วงหน้า

    available_machines = item_cap_data[item_cap_data["ITEM_CODE"] == item_code]

    if available_machines.empty:

        return None, None, None, None, None



    # ใช้ความจุตาม logic ใหม่: สำหรับ SHARED_POOL_MAP ให้เลือกจาก cap ที่น้อยที่สุดในกลุ่มเดียวกัน

    # และ FA 20 มี priority สูงกว่า SKP 20

    # เรียงตาม MC_GROUP ที่มีเครื่องเหลือมากที่สุดก่อน (plan_week)

    available_machines = available_machines.copy()

    available_machines["_mc_remain"] = available_machines.apply(

        lambda r: get_actual_mc_remain(r["MC_GROUP"], plan_week, gauge=r.get("GUAGE"), item_code=item_code),

        axis=1,

    )

    available_machines = available_machines.sort_values("_mc_remain", ascending=False)

    current_week_idx = week_index(plan_week)

    # 1. ลองเครื่องที่ว่างในสัปดาห์นี้ก่อน

    for _, machine_row in available_machines.iterrows():

        mc_group = machine_row["MC_GROUP"]

        # ใช้ความจุตาม logic ใหม่สำหรับ MC_GROUP นี้

        item_gauge = machine_row["GUAGE"] if "GUAGE" in machine_row else None

        daily_cap = _get_capacity_for_mc_group(item_code, mc_group, item_gauge)

        daily_cap = adjust_daily_cap_for_item_special(daily_cap, item_code, mc_group, item_gauge)

        if daily_cap <= 0:

            continue  # ข้ามถ้าไม่มีความจุ

        # ดูเครื่องว่างจริง (หักที่จองไปแล้ว)

        actual_remain = get_actual_mc_remain(mc_group, plan_week, gauge=item_gauge, item_code=item_code)

        # 🔧 FIX: เช็ค booking_mc_by_week (old booking) เพื่อป้องกันเกิน capacity

        # ถ้า old booking ใช้เครื่องไปแล้วใน week นี้ → หักออกจาก actual_remain

        mc_key = _resolve_carry_key(item_code, mc_group, item_gauge)

        _current_week_idx = week_index(plan_week)

        _booking_mc_used = 0

        if mc_key in booking_mc_by_week and _current_week_idx is not None:

            _booking_mc_used = booking_mc_by_week.get(mc_key, {}).get(_current_week_idx, 0)

        # หัก booking_mc_used จาก actual_remain

        actual_remain_before = actual_remain

        actual_remain = max(0, actual_remain - _booking_mc_used)

        # Cylinder change: ถ้า pool ว่าง ลองเปลี่ยน cylinder จาก gauge อื่นใน MC_CAT เดียวกัน
        _cylinder_changed = False
        if actual_remain == 0:
            # ตรวจว่า item มี booking active สัปดาห์นี้หรือไม่
            # ถ้ามี → booking machines ถูกหักจาก TOTAL_MC_REMAIN ไปแล้ว จึงไม่ต้อง cylinder change
            _has_booking_this_week = (
                _booking_mc_used > 0
                and mc_key in booking_active_week_set
                and _current_week_idx is not None
                and _current_week_idx in booking_active_week_set.get(mc_key, set())
            )
            if _has_booking_this_week:
                # booking machines คือเครื่องที่มีอยู่แล้ว (ถูกหักจาก TOTAL_MC_REMAIN แล้ว)
                # ไม่ต้อง cylinder change — restore actual_remain กลับเป็น booking count
                actual_remain = _booking_mc_used
                print(f"[BOOKING RESTORE] {item_code} W{plan_week}: actual_remain=0 เพราะ double-subtract booking → restore to {_booking_mc_used}")
            elif str(item_code).strip().upper() in _s9_only_items:
                # S9 Only: ข้าม cylinder change เสมอ (ทั้ง Pass 1 และ Pass 2) — ใช้ MC จาก S9 pool เท่านั้น
                pass
            else:
                _cyl_mc_cat = _mc_to_type1(mc_group, item_gauge)
                _cyl_factory = _mc_to_factory(mc_group, item_gauge)
                _cyl_tgt_g = _normalize_gauge(item_gauge)
                _cyl_trigger_week = int(plan_week) - 1  # สั่งเปลี่ยน 1 week ก่อนผลิต → machine พร้อมทัน week ที่ผลิต
                # ตรวจว่าถ้ารอ JIT window ปกติ (rdd-2) จะยังมีเวลาผลิตพอไหม
                # ถ้าไม่พอ → override JIT เพื่อเริ่มผลิตทันเวลา
                _pw_idx_p0 = week_index(plan_week)
                _req_mc_p0 = required_machines_info[2] if required_machines_info and len(required_machines_info) > 2 else 1
                _weeks_to_rdd_p0 = (
                    (_current_order_rdd_idx - _pw_idx_p0)
                    if (_pw_idx_p0 is not None and _current_order_rdd_idx is not None)
                    else 99
                )
                # ถ้า deadline อยู่ภายใน (required_machines + 2) สัปดาห์ → ต้องเริ่มตอนนี้ → override JIT
                # _weeks_to_rdd > 0: ป้องกัน PAST_RDD (negative weeks) ที่ทำให้ override เสมอ
                _cyl_jit_override_p0 = 0 < _weeks_to_rdd_p0 <= _req_mc_p0 + 2
                if _cyl_jit_override_p0:
                    print(f"[CYL JIT OVERRIDE] {item_code} W{plan_week}: pool=0, weeks_to_rdd={_weeks_to_rdd_p0} <= req_mc={_req_mc_p0}+2 → jit_override=True")
                if _try_cylinder_change(_cyl_mc_cat, _cyl_factory, _cyl_tgt_g, _cyl_trigger_week, item_code, mc_group, jit_override=_cyl_jit_override_p0):
                    # cylinder change สำเร็จ → เครื่องใหม่ 1 ตัวพร้อมใช้ (ไม่ขึ้นกับ pool ที่อาจถูกใช้ไปแล้ว)
                    actual_remain = 1
                    _cylinder_changed = True  # เครื่องที่เพิ่งเปลี่ยน cylinder ต้อง setup ใหม่เสมอ
        if actual_remain > 0:
            if _cylinder_changed:
                # cylinder เพิ่งถูกเปลี่ยนมาเพื่อ item นี้โดยเฉพาะ → bypass job cap
                available_machines_count = actual_remain
            else:
                type_used = get_type_used_jobs(plan_week, mc_group)

                available_machines_count = check_job_capacity_limit(

                    mc_group, actual_remain, urgent_mode, type_used

                )

                if available_machines_count <= 0:

                    continue  # ลอง MC_GROUP ถัดไป



            key = _resolve_carry_key(item_code, mc_group, item_gauge)

            setup_needed = True

            _bridge_locks(key, current_week_idx)
            _prev_cyl_key = (int(plan_week) - 1, str(item_code).strip().upper(), str(mc_group).strip().upper())
            if not _cylinder_changed and (_prev_cyl_key not in _cylinder_change_for_item) and key in last_production:

                last_week_idx = last_production[key]

                if current_week_idx - last_week_idx <= SETUP_GAP_WEEK:

                    setup_needed = False

            return (



                mc_group,

                daily_cap,

                setup_needed,

                available_machines_count,

                item_gauge,

            )

    # 2. ถ้าไม่มีเครื่องว่าง ลอง MC ที่เคยผลิต item เดียวกัน

    previous_mcs = [key[1] for key in last_production if key[0] == item_code]

    for prev_mc in previous_mcs:

        prev_mc_row = available_machines[available_machines["MC_GROUP"] == prev_mc]

        if not prev_mc_row.empty:

            mc_group = prev_mc

            # ใช้ความจุตาม logic ใหม่สำหรับ MC_GROUP นี้

            item_gauge = (

                prev_mc_row.iloc[0]["GUAGE"] if "GUAGE" in prev_mc_row.iloc[0] else None

            )

            daily_cap = _get_capacity_for_mc_group(item_code, mc_group, item_gauge)

            daily_cap = adjust_daily_cap_for_item_special(daily_cap, item_code, mc_group, item_gauge)

            if daily_cap <= 0:

                continue  # ข้ามถ้าไม่มีความจุ

            actual_remain = get_actual_mc_remain(mc_group, plan_week, gauge=item_gauge, item_code=item_code)

            if actual_remain > 0:

                type_used = get_type_used_jobs(plan_week, mc_group)

                available_machines_count = check_job_capacity_limit(

                    mc_group, actual_remain, urgent_mode, type_used

                )

                if available_machines_count <= 0:

                    continue



                setup_needed = False

                return (



                    mc_group,

                    daily_cap,

                    setup_needed,

                    available_machines_count,

                    item_gauge,

                )

    return None, None, None, None, None





def next_week(week):

    idx = week_index(week)

    if idx is None:

        return None



    idx += 1

    # ข้าม SKIP_WEEKS

    while idx < len(calendar_week):

        w = int(calendar_week.iloc[idx]["WEEK"])

        if w not in SKIP_WEEKS:

            return w



        idx += 1

    return None



TODAY_WEEK = get_week_from_date(TODAY)

# TODAY_IDX: หา row บน timeline ที่ครอบ TODAY โดยตรงจากวันที่ (กันยึดผิดปี)
# เดิมใช้ week_index(TODAY_WEEK) แต่ TODAY_WEEK เป็น bare week และตอน bootstrap
# TODAY_IDX ยังไม่มี → week_index คืน idx[0] = ปีแรกสุดในปฏิทิน (เช่นปี 2026 แทนที่จะเป็น 2027)
_today_rows = calendar_week.index[
    (calendar_week["WEEK_START"] <= TODAY) & (calendar_week["WEEK_END"] >= TODAY)
]
if len(_today_rows) > 0:
    TODAY_IDX = int(_today_rows[0])
else:
    # TODAY อยู่นอกช่วงปฏิทิน → ใช้ row สุดท้ายที่เริ่มก่อน/ตรงกับ TODAY (fallback = แถวแรก)
    _before = calendar_week.index[calendar_week["WEEK_START"] <= TODAY]
    TODAY_IDX = int(_before[-1]) if len(_before) > 0 else week_index(TODAY_WEEK)





def _make_type_key(factory: str, mc_type: str) -> str:

    """สร้าง type_key: OM/OMNOI ไม่มี Type ใช้ชื่อ factory อย่างเดียว

    SINGLE (TUBE) รวมกับ SINGLE เป็นตัวเดียวกัน"""

    if factory in ("OM", "OMNOI"):

        return "OM"

    

    # ตัด (TUBE) ออก - SINGLE (TUBE) = SINGLE

    mc_type_clean = mc_type.replace(" (TUBE)", "").strip() if mc_type else ""

    

    return f"{factory}_{mc_type_clean}" if mc_type_clean else factory





def _get_type_key_for_mc(mc_group: str) -> str:

    """คืน type_key ของ MC_GROUP จาก master_mc"""

    _info = master_mc[master_mc["MC"] == mc_group]

    if _info.empty:

        return "PHET_DOUBLE"



    _fac = str(_info.iloc[0]["Factory"]).strip().upper()

    _raw = _info.iloc[0].get("Type", "")

    _typ = "" if pd.isna(_raw) else str(_raw).strip().upper()

    return _make_type_key(_fac, _typ)





def get_type_used_jobs(plan_week: int, mc_group: str) -> int:

    """คืนจำนวน jobs ที่ใช้ไปแล้วใน week นั้น รวมทุก MC_GROUP ใน factory type เดียวกัน cap PHET_DOUBLE=33, PHET_SINGLE=44, OM=13 นับรวม factory-wide ทุก MC_GROUP ใน type นั้น"""

    _target_type = _get_type_key_for_mc(mc_group)

    _week_usage = weekly_job_usage.get(plan_week, {})

    _total = 0

    for _mc, _jobs in _week_usage.items():

        if _get_type_key_for_mc(_mc) == _target_type:

            _total += _jobs

    return _total





def get_remaining_job_slots(plan_week: int, mc_group: str) -> int:

    """คืน job slots ที่เหลืออยู่สำหรับ factory type ของ mc_group ใน week นั้น"""

    mc_info = master_mc[master_mc["MC"] == mc_group]

    if mc_info.empty:

        factory, mc_type = "PHET", "DOUBLE"

    else:

        factory = str(mc_info.iloc[0]["Factory"]).strip().upper()

        mc_type = str(mc_info.iloc[0].get("Type", "DOUBLE")).strip().upper()

    if factory == "PHET":

        max_jobs = 33 if mc_type == "DOUBLE" else 44

    elif factory in ("OM", "OMNOI"):

        max_jobs = 13

    else:

        return 9999  # OUTSOURCE → unlimited



    used = get_type_used_jobs(plan_week, mc_group)

    return max(0, max_jobs - used)





def detect_and_fill_unused_capacity(plans_list, orders_df, summary_mc):

    """

    Detect unused machine capacity and fill with same-item orders from different SCs.

    Optimizes machine utilization by identifying week-item-mc-gauge combinations 

    with unused capacity and filling them with pending orders of the same item.

    """

    if not plans_list:

        return plans_list



    print("🔍 Detecting and filling unused capacity...")

    # Convert plans to DataFrame for easier analysis

    plan_df = pd.DataFrame(plans_list)

    

    # 🔧 FIX: Include OLD bookings (detail_mc) in current_usage calculation

    # to prevent filling capacity already used by old bookings

    old_booking_usage = []

    if not detail_mc.empty:

        for _, row in detail_mc.iterrows():

            item_code = str(row.get("ITEM_CODE", "")).strip().upper()

            mc_group = str(row.get("MC_GROUP", "")).strip().upper()

            gauge = str(row.get("GUAGE", "")).strip()

            week = row.get("WEEK")

            kp_weight = row.get("KP_WEIGHT", 0)

            cap_tor = row.get("CAP ทอ", 0)

            

            if not item_code or not mc_group or pd.isna(week) or pd.isna(kp_weight):

                continue

            

            try:

                week = int(week)

                kp_weight = float(kp_weight)

                cap_tor = float(cap_tor) if not pd.isna(cap_tor) else 0

            except (ValueError, TypeError):

                continue

            

            if kp_weight <= 0:

                continue

            

            # Calculate daily capacity from CAP ทอ (แปลงฐาน 24 ชม. → ชั่วโมงจริงของกลุ่ม)

            daily_cap = cap_tor * (_base_work_hours(mc_group, gauge) / 24) if cap_tor > 0 else 0

            

            old_booking_usage.append({

                'PLAN_WEEK': week,

                'ITEM_CODE': item_code,

                'MC_GROUP': mc_group,

                'MC_GUAGE': gauge,

                'PRODUCE_QTY': kp_weight,

                'REQUIRED_MC': 0,  # Will be calculated from grouping

                'DAILY_CAPACITY': daily_cap

            })

    

    # Combine NEW plans and OLD bookings

    if old_booking_usage:

        old_booking_df = pd.DataFrame(old_booking_usage)

        combined_df = pd.concat([plan_df, old_booking_df], ignore_index=True)

    else:

        combined_df = plan_df.copy()

    

    # Group by week, item, mc_group, gauge to find current usage (NEW + OLD)

    current_usage = combined_df.groupby(['PLAN_WEEK', 'ITEM_CODE', 'MC_GROUP', 'MC_GUAGE']).agg({

        'PRODUCE_QTY': 'sum',

        'REQUIRED_MC': 'max',

        'DAILY_CAPACITY': 'first'

    }).reset_index()

    # Calculate theoretical full capacity per week-item-mc-gauge

    # Use actual available machines in that week, not just REQUIRED_MC

    def get_available_machines_for_week_mc(week, mc_group, gauge=None):

        """Get actual available machines for specific week, Type_1+Gauge pool"""

        type_1 = _mc_to_type1(mc_group, gauge)

        gauge_str = _normalize_gauge(gauge) if gauge else ""

        mask = (summary_mc["WEEK"] == week) & (summary_mc["TYPE_1"] == type_1)

        if gauge_str:

            mask = mask & (summary_mc["GUAGE"].apply(_normalize_gauge) == gauge_str)

        mc_rows = summary_mc[mask]

        if mc_rows.empty:

            return 0

        # พูลแยก (SKP/SKPTA·SKPLE): คิดเฉพาะพูลของเครื่องนี้ ไม่รวมพูลอื่น
        _pool_label = _mc_to_pool_label(mc_group, gauge)
        if (_pool_label and (type_1, gauge_str) in _SPLIT_CG
                and "MC_POOL" in summary_mc.columns):
            mc_rows = mc_rows[mc_rows["MC_POOL"].astype(str).str.strip() == _pool_label]

        return int(mc_rows["TOTAL_MC_REMAIN"].sum())

    

    # Add available machines column

    current_usage['AVAILABLE_MC'] = current_usage.apply(

        lambda row: get_available_machines_for_week_mc(row['PLAN_WEEK'], row['MC_GROUP'], row.get('MC_GUAGE')),

        axis=1

    )

    

    current_usage['FULL_CAPACITY'] = (

        current_usage['AVAILABLE_MC'] * 

        current_usage['DAILY_CAPACITY'] * 

        7  # Assuming 7 working days for full capacity calculation

    )

    # Find combinations with unused capacity (less than 95% utilization)

    unused_capacity = current_usage[

        current_usage['PRODUCE_QTY'] < (current_usage['FULL_CAPACITY'] * 0.95)

    ].copy()

    if unused_capacity.empty:

        print("✅ No unused capacity detected")

        return plans_list

    
    # 🔧 FIX: Filter out weeks where OLD bookings exist for same item/MC
    # to prevent filling capacity already used by old bookings
    unused_capacity_filtered = []
    for _, row in unused_capacity.iterrows():
        week = row['PLAN_WEEK']
        item = row['ITEM_CODE']
        mc_group = row['MC_GROUP']
        # Check if OLD booking exists for this item/MC/week in detail_mc
        old_check = detail_mc[
            (detail_mc["ITEM_CODE"].astype(str).str.upper() == item.upper()) &
            (detail_mc["MC_GROUP"].astype(str).str.upper() == mc_group.upper()) &
            (detail_mc["WEEK"] == week)
        ]
        if not old_check.empty:

            old_mc_ceil = old_check.iloc[0].get("MC_USE_CEIL", 0)

            if old_mc_ceil > 0:

                print(f"  🔧 SKIP CAPACITY OPT: {item}+{week}+{mc_group}: OLD booking using {old_mc_ceil} machines, skip to avoid exceeding capacity")

                continue

        unused_capacity_filtered.append(row)

    

    if not unused_capacity_filtered:

        print("✅ No unused capacity after filtering OLD booking weeks")

        return plans_list

    

    unused_capacity = pd.DataFrame(unused_capacity_filtered)



    print(f"📊 Found {len(unused_capacity)} week-item-mc combinations with unused capacity")

    # Get pending orders that could fill the capacity

    pending_orders = orders_df[

        (orders_df['Pending Plan'] > 0) &

        (~orders_df['SC/SO NO'].isin(plan_df['SC_SO_NO'].unique()))

    ].copy()

    

    # 🔧 FIX: สร้าง set ของ Item+Week+MC_GROUP ที่มีอยู่แล้ว เพียงครั้งเดียว

    existing_item_week_mc = set()

    for existing_plan in plans_list:

        existing_item = existing_plan.get('ITEM_CODE', '')

        existing_week = existing_plan.get('PLAN_WEEK', 0)

        existing_mc = existing_plan.get('MC_GROUP', '')

        if existing_item and existing_week and existing_mc:

            existing_item_week_mc.add((existing_item, existing_week, existing_mc))

    

    additional_plans = []

    for _, usage_row in unused_capacity.iterrows():

        week = usage_row['PLAN_WEEK']

        item = usage_row['ITEM_CODE']

        mc_group = usage_row['MC_GROUP']

        gauge = usage_row['MC_GUAGE']

        unused_qty = usage_row['FULL_CAPACITY'] - usage_row['PRODUCE_QTY']

        if unused_qty <= 0:

            continue

        

        # 🔧 FIX: Skip weeks <= TODAY+2 to prevent planning before yarn arrives

        if week <= TODAY_WEEK + 2:

            print(f"  🔧 SKIP CAPACITY OPT: {item}+{week}+{mc_group}: week <= TODAY+2 (W{TODAY_WEEK+2}), skip")

            continue



        # Find pending orders of same item (different SC)

        matching_orders = pending_orders[

            (pending_orders['Item Code'] == item) &

            (pending_orders['MC GROUP'] == mc_group)

        ]

        if matching_orders.empty:

            continue



        # Sort by RDD (urgent first)

        matching_orders = matching_orders.copy()

        matching_orders['FG_WEEK_NUM'] = matching_orders['FG Week'].astype(str).str[-2:].astype(int)

        matching_orders = matching_orders.sort_values('FG_WEEK_NUM')

        for _, order in matching_orders.iterrows():

            if unused_qty <= 0:

                break



            sc_so_no = str(order.get('SO_NO', order.get('SC/SO NO', ''))).strip()

            pending_qty = order['Pending Plan']

            # Calculate how much we can produce in remaining capacity

            produce_qty = min(unused_qty, pending_qty)

            if produce_qty > 0:

                # 🔧 FIX: ตรวจสอบว่า Item+Week+MC_GROUP นี้มีการ plan ไปแล้วหรือไม่

                # เพื่อป้องกันการซ้ำซ้อนกับ plans ที่มีอยู่แล้ว

                current_item_week_mc = (item, week, mc_group)

                if current_item_week_mc in existing_item_week_mc:

                    print(f"  ⚠️  SKIPPED capacity optimization: {item}+{week}+{mc_group} already planned!")

                    break

                

                # Create additional plan entry

                additional_plan = {

                    'ITEM_CODE': item,

                    'SC_SO_NO': sc_so_no,

                    'MC_GROUP': mc_group,

                    'MC_GUAGE': gauge,

                    'FACTORY_TYPE': FACTORY_TYPE_MAP.get(mc_group, "UNKNOWN"),

                    'PLAN_WEEK': week,

                    'PRODUCE_QTY': produce_qty,

                    'SETUP_DAYS': 0,  # No setup needed - same item/mc/gauge

                    'REQUIRED_MC': usage_row['REQUIRED_MC'],

                    'ACTUAL_MC': usage_row['REQUIRED_MC'],

                    'CARRYOVER_MC': usage_row['REQUIRED_MC'],

                    'NEW_MC': 0,

                    'FACTORY_WORKING_DAYS': get_working_days_by_factory(mc_group, usage_row['REQUIRED_MC'], week=week, gauge=gauge),

                    'CALENDAR_WORKING_DAYS': len(get_working_days_in_week(week)),

                    'ACTUAL_WORKING_DAYS': get_working_days_by_factory(mc_group, usage_row['REQUIRED_MC'], week=week, gauge=gauge),

                    'DAILY_CAPACITY': usage_row['DAILY_CAPACITY'],

                    'REVOLUTION_WEIGHT': get_revolution_weight_from_orders(item, mc_group),

                    'AVAILABLE_DAYS': get_working_days_by_factory(mc_group, usage_row['REQUIRED_MC'], week=week, gauge=gauge),

                    'ORDERS_QTY': order['Orders.Qty'],

                    'PENDING_PLAN': pending_qty - produce_qty,

                    'PLAN_QTY': pending_qty - produce_qty,

                    'ORDER_TYPE': order['Orders Type'],

                    'ORDER_DATE': order['Date'],

                    'FG_WEEK': order['FG Week'],

                    'TARGET_KNIT': order['FG Week'],  # Simplified

                    'MATERIAL_CONTENT': str(order.get('MATERIAL_CONTENT', '')).strip(),

                    'IS_CORE_ITEM': '',

                    'CUSTOMER': str(order.get('Customer', '')).strip(),

                    'PLAN_SOURCE': 'NEW',

                    'LT_YARN': order.get('DYE_END_DATE') if str(order.get('Orders Type', '')).strip() == 'YD-ORDERS' else get_yarn_lt_days(item),

                    'YARN_USED': _yarn_used_lookup.get(str(item).strip().upper(), ''),

                    'DATE_IN': order.get('DATE_IN'),

                    'EARLIEST_PLAN_WEEK': order.get('DYE_END_DATE') if str(order.get('Orders Type', '')).strip() == 'YD-ORDERS' else get_yarn_lt_earliest_week(item, date_in=order.get('DATE_IN')),

                    'SUB_COLOR': str(order.get('SUB_COLOR', '')).strip(),

                    'PO_NO': str(order.get('PO_NO', '')).strip(),

                    'YARNPO': str(order.get('YARNPO', '') or '').strip(),

                    'RDD_WEEK': order.get('FG Week'),

                    'SC_LINE_ID': str(order.get('SC_LINE_ID', '')).strip(),

                }

                additional_plan['NAY_COLOR'] = str(order.get('NAY_COLOR', '')).strip()

                additional_plan['COLOR_DESC'] = str(order.get('COLOR_DESC', '')).strip()

                additional_plan['LINE_REMARK'] = str(order.get('LINE_REMARK', '')).strip()

                additional_plans.append(additional_plan)

                unused_qty -= produce_qty

                print(f"  📈 Added {produce_qty:.2f} units of {item} (SC: {sc_so_no}) to week {week}")

    if additional_plans:

        print(f"✅ Added {len(additional_plans)} capacity optimization plans")

        plans_list.extend(additional_plans)

    else:

        print("ℹ️  No suitable pending orders found for capacity optimization")

    return plans_list



# =========================

# LOAD OLD PRODUCTION PLAN FOR VALIDATION

# =========================

OLD_PLAN_FILE = DATA_PLAN_DIR / "committed_plan.xlsx"

# ❌ DISABLED: old-plan carryover จาก committed_plan.xlsx
# เดิมโค้ดตรงนี้อ่าน committed_plan.xlsx (= ผลลัพธ์รอบก่อน) มาเป็น "old plan carryover"
# แล้วเอาไป pre-load เครื่อง/หัก qty/ล็อกเครื่องเดิม → ทำให้ "รันแผนใหม่" ยังหยิบแผนเก่ามาปน
# เกิด feedback loop: ผลลัพธ์ไม่นิ่ง (booking/order เดิมเป๊ะ ๆ แต่รันได้คนละสัปดาห์ W34↔W36)
# บังคับ old_plan ให้ว่างเสมอ → รันแผนใหม่คำนวณสดล้วนจาก booking + order (deterministic)
# หมายเหตุ: ฝั่ง "เขียน" committed_plan ก็ถูกปิดแล้ว (ดู block ปิด auto-backup ท้ายไฟล์)
#   สำคัญบน server (deploy): แม้ committed_plan.xlsx เก่าจะค้างอยู่ ก็จะไม่ถูกอ่านมาใช้อีก
#   ถ้าอนาคตจะทำ carryover จาก "แผนที่ user กด commit จริง" ค่อยเปิดอ่านเฉพาะไฟล์นั้น
old_plan_df = pd.DataFrame()

# สร้าง dict: {sc_so_no_upper: total_produce_qty} จาก old plan (NEW) ที่วาง week >= TODAY

# ใช้หัก qty_left เพื่อไม่วางซ้ำ (carry-over qty)

old_plan_produced_qty = {}

if (

    not old_plan_df.empty

    and "PRODUCE_QTY" in old_plan_df.columns

    and "PLAN_WEEK" in old_plan_df.columns

):

    _sc_col = "SC_SO_NO" if "SC_SO_NO" in old_plan_df.columns else None

    if _sc_col:

        for _, _opr in old_plan_df.iterrows():

            _opw = _opr.get("PLAN_WEEK")

            if pd.isna(_opw):

                continue



            _opw_idx = week_index(int(_opw))

            if _opw_idx is None or _opw_idx < TODAY_IDX:

                continue  # เฉพาะ week อนาคต



            _op_sc = str(_opr[_sc_col]).strip().upper()

            _op_qty = pd.to_numeric(_opr.get("PRODUCE_QTY", 0), errors="coerce")

            if pd.isna(_op_qty):

                _op_qty = 0

            if _op_sc and _op_sc != "NAN":

                # handle comma-separated SC/SO (e.g. "715033,S716032")

                for _sc_part in _op_sc.split(","):

                    _sc_part = _sc_part.strip().upper().lstrip("S")

                    if _sc_part:

                        old_plan_produced_qty[_sc_part] = old_plan_produced_qty.get(

                            _sc_part, 0

                        ) + float(_op_qty)

    print(

        f"📋 Old plan carry-over qty (week>={TODAY_WEEK}): {len(old_plan_produced_qty)} SOs"

    )



# =========================

# LOAD BOOKING DATA (ประวัติการผลิตจริง)

# =========================

_BOOKING_EXCLUDE_MC = {

    "CL-NP",

    "CL-OM",

    "COMKN",

    "F-CL",

    "CL",

    "FQCCL-NP",

    "FQCCL-OM",

    "FQC-OMNOI",

    "FQC-PHET",

    "FQC",

    "F-TSD",

}

# กลุ่มที่ "ไม่ใช่เครื่องถัก" → ห้ามนับเป็น setup job ต่อสัปดาห์
# IMPORT = ผ้านำเข้า ไม่ได้ถักเอง (ไม่มี Factory/TOTAL_MC ใน MasterMC) แต่ MC_USE_CEIL
# ถูกคำนวณจากน้ำหนักผ้าล้วนๆ จนได้เลขหลักร้อย แล้วไปกินโควตา job ของ PHET_DOUBLE (default)
_JOB_EXCLUDE_MC = {"IMPORT"}

booking_last_production = {}  # {(item, mc_group): week_index} — week สุดท้ายที่ผลิตจริง

booking_last_so = {}  # {(item, mc_group): so_no_normalized} — SO หมายเลขสุดท้าย

booking_produced_qty = {}  # {so_no_upper: total_knit_weight} — ผลิตไปแล้วทั้งหมด

booking_raw = load_all_booking_data()

if not booking_raw.empty:

    for col in ["ITEM_CODE", "MC_GROUP", "SO_NO", "TYPE", "KP_NO"]:

        if col in booking_raw.columns:

            booking_raw[col] = booking_raw[col].astype(str).str.strip().str.upper()

    booking_raw["WEEK"] = pd.to_numeric(booking_raw.get("WEEK"), errors="coerce")

    booking_raw["YEAR"] = pd.to_numeric(booking_raw.get("YEAR"), errors="coerce")

    if "KNIT WEIGHT" not in booking_raw.columns:

        booking_raw["KNIT WEIGHT"] = 0

    else:

        booking_raw["KNIT WEIGHT"] = pd.to_numeric(

            booking_raw["KNIT WEIGHT"], errors="coerce"

        ).fillna(0)

    if "MC_GROUP" in booking_raw.columns:

        booking_raw = booking_raw[~booking_raw["MC_GROUP"].isin(_BOOKING_EXCLUDE_MC)]

    if "TYPE" in booking_raw.columns:

        booking_raw = booking_raw[booking_raw["TYPE"] != "COLLAR"]

    # เฉพาะแถวที่ผลิตจริง (KNIT WEIGHT > 0) และ YEAR 2025-2026

    _produced = booking_raw[

        (booking_raw["KNIT WEIGHT"] > 0) & (booking_raw["YEAR"].isin([2025, 2026]))

    ].copy()

    # สร้าง last_production จาก booking

    for _, _row in _produced.iterrows():

        _bi = str(_row.get("ITEM_CODE", "")).strip().upper()

        _bm = str(_row.get("MC_GROUP", "")).strip().upper()

        _bw = _row.get("WEEK")

        _bs = str(_row.get("SO_NO", "")).strip().upper()

        if not _bi or not _bm or pd.isna(_bw):

            continue



        _bw = int(_bw)

        _wi = week_index(_bw)

        if _wi is None:

            continue



        _key = _ck(_bi, _bm)

        if _key not in booking_last_production or _wi > booking_last_production[_key]:

            booking_last_production[_key] = _wi

            booking_last_so[_key] = _bs

    # สร้าง produced_qty per SO — ใช้ KP_NO เพื่อหลีกเลี่ยงการนับซ้ำรายสัปดาห์

    if "KP_NO" in _produced.columns:

        _kp_latest = (

            _produced.groupby(["KP_NO", "SO_NO"])["KNIT WEIGHT"].max().reset_index()

        )

        for _, _r in _kp_latest.iterrows():

            _so = str(_r["SO_NO"]).strip().upper()

            if _so and _so != "NAN":

                booking_produced_qty[_so] = (

                    booking_produced_qty.get(_so, 0) + _r["KNIT WEIGHT"]

                )

    else:

        if "SO_NO" in _produced.columns:

            for _so, _grp in _produced.groupby("SO_NO"):

                _so_key = str(_so).strip().upper()

                if _so_key and _so_key != "NAN":

                    booking_produced_qty[_so_key] = _grp["KNIT WEIGHT"].sum()

    print(

        f"📚 Booking history: {len(booking_last_production)} (item,mc) records, {len(booking_produced_qty)} unique SOs"

    )

else:

    print("📚 ไม่พบข้อมูล booking history")



# =========================

# TRACK LAST PRODUCTION

# =========================

last_production = {}

machines_in_use = {}  # {(item, mc_group): จำนวนเครื่องที่ใช้จริงใน week ล่าสุด}

last_sc_machines = {}  # {(item, mc_group, gauge, sc_so_no): จำนวนเครื่องล่าสุดต่อ SC — ป้องกัน SC อื่นทับข้อมูล}

last_sc_week = {}  # {(item, mc_group, gauge, sc_so_no): week_index ล่าสุดต่อ SC — ใช้แทน last_sc_so_no ที่ถูก SC อื่นทับ

weekly_mc_usage = {}  # {(week, mc_group): total machines} สำหรับ gradual increase

last_sc_so_no = (

    {}

)  # {(item, mc_group): SC/SO NO ของ order ที่ผลิตล่าสุด — ป้องกัน carry-over ข้าม color/order}

last_dye_end_date = {}  # {carry_key: DYE_END_DATE} — สำหรับ YD-ORDERS: ถ้า วันนัดย้อม เปลี่ยน ต้อง setup เพิ่ม 1 วัน

last_sub_color = {}  # {carry_key: sub_color} — สีล่าสุดที่ผลิตต่อ YD-ORDERS item

_yd_week_color_setups = {}  # {(carry_key, plan_week): cumulative color setup days} — สะสมวัน setup สีใน week เดียวกัน

booking_production_keys = set()  # keys ที่มาจาก detail_mc (booking จริง) — ใช้อนุญาต carryover จาก booking แม้ยังไม่มีใน new plan

booking_active_week_set = {}  # {(item, mc, gauge): set of w_idx ที่ booking active} — ใช้เช็ค carryover เมื่อ booking ครอบหลาย week

booking_mc_by_week = {}  # {(item, mc, gauge): {w_idx: mc_used}} — จำนวนเครื่องจาก booking รายสัปดาห์

# งานที่วางไม่ได้เพราะพูลเครื่องเต็ม (booking จองเกิน) — สรุปให้ผู้ใช้ท้ายการรัน
# ไม่งั้นงานจะหายจากแผนเงียบๆ โดยไม่รู้ว่าติดสัปดาห์ไหน
no_mc_skips = []  # [{"item", "sc", "mc_group", "gauge", "setup_week", "blocked_week", "remain", "need"}]

# Pre-populate last_production จาก detail_mc (DETAIL sheet ของ booking_final_ready25)

# ใช้ข้อมูลจริงจาก booking เพื่อรู้ว่า item นี้ถูก book ถึง week ไหน

for _, row in detail_mc.iterrows():

    item_code = str(row.get("ITEM_CODE", "")).strip().upper()

    mc_group = str(row.get("MC_GROUP", "")).strip().upper()

    plan_week = row.get("WEEK")

    mc_used = row.get("MC_USE_CEIL", 0)

    if not item_code or not mc_group or pd.isna(plan_week) or pd.isna(mc_used):

        continue



    if int(mc_used) == 0:

        continue



    plan_week = int(plan_week)

    # ใช้ index เปรียบเทียบแทน raw week number เพื่อรองรับข้ามปี

    # เช่น booking week 50 ปี 2025 vs TODAY_WEEK=10 ปี 2026: 50 > 10 → ผิด ถ้าใช้ raw number

    w_idx = week_index(plan_week)

    if w_idx is None:

        continue  # ข้ามเฉพาะ week ที่ไม่มีใน calendar (ไม่กรองตาม TODAY อีกต่อไป → ให้ดู last booking week จริงๆ)



    _det_gauge = str(row.get("GUAGE", "")).strip()

    key = _ck(item_code, mc_group, _det_gauge)

    booking_production_keys.add(key)  # track ว่า key นี้มาจาก booking จริง

    # 🔧 NEW: เก็บทุก week ที่ booking active เพื่อให้ carryover ตรวจได้ครอบคลุม

    # (last_production เก็บเฉพาะ week ล่าสุด ไม่พอถ้า plan_week อยู่ก่อน last booking week)

    booking_active_week_set.setdefault(key, set()).add(w_idx)

    _week_mc_map = booking_mc_by_week.setdefault(key, {})

    try:

        _mc_used_int = int(mc_used)

    except (ValueError, TypeError):

        _mc_used_int = 0

    # 🔧 FIX: ไม่ปรับ machines_in_use ตาม working days

    # MC_USE คือจำนวนเครื่องจริงที่ใช้ ไม่ใช่ capacity calculation

    # ถ้า booking บันทึก MC_USE=1 แปลว่าใช้เครื่อง 1 เครื่องจริงๆ ไม่ว่า working days จะเท่าไหร่

    if _mc_used_int > _week_mc_map.get(w_idx, 0):

        _week_mc_map[w_idx] = _mc_used_int



    if key not in last_production or w_idx > last_production[key]:

        last_production[key] = w_idx

        machines_in_use[key] = _mc_used_int

print(

    f"📋 โหลด last_production จาก detail_mc (booking_final_ready25): {len(last_production)} รายการ"

)

# Merge old plan → ใช้เป็น fallback สำหรับ carryover ถ้า detail_mc/booking ไม่มี

if not old_plan_df.empty:

    for _, _row in old_plan_df.iterrows():

        # หาชื่อคอลัมน์ที่เป็นไปได้

        item_code = (

            _row.get("ITEM")

            or _row.get("Item")

            or _row.get("ITEM_CODE")

            or _row.get("Item Code")

        )

        mc_group = _row.get("MC_GROUP") or _row.get("MC GROUP") or _row.get("MC")

        plan_week = (

            _row.get("PLAN_WEEK") or _row.get("PLAN WEEK") or _row.get("PLAN_WEEK")

        )

        sc_no = (

            _row.get("SC_SO_NO")

            or _row.get("SC/SO NO")

            or _row.get("SC SO NO")

            or _row.get("SC/SO")

            or _row.get("SC")

        )

        if pd.isna(item_code) or pd.isna(mc_group) or pd.isna(plan_week):

            continue



        try:

            item_code = str(item_code).strip().upper()

            mc_group = str(mc_group).strip().upper()

            plan_week = int(plan_week)

        except Exception:

            continue



        # ใช้ index เปรียบเทียบแทน raw week number เพื่อรองรับข้ามปี

        w_idx = week_index(plan_week)

        if w_idx is None:

            continue



        # ใช้ ACTUAL_MC ก่อน (จำนวนเครื่องจริงที่ผลิต) — ถ้า = 0 แปลว่า week นั้นไม่ได้ผลิต → ข้าม

        _old_actual_mc = _row.get("ACTUAL_MC", None)

        try:

            _old_actual_mc_int = int(_old_actual_mc) if _old_actual_mc is not None and not pd.isna(_old_actual_mc) else 0

        except Exception:

            _old_actual_mc_int = 0

        if _old_actual_mc_int == 0:

            continue



        _old_gauge = _row.get("MC_GUAGE") or _row.get("MC GUAGE") or _row.get("GUAGE")

        _old_gauge_str = (

            str(_old_gauge).strip() if _old_gauge and not pd.isna(_old_gauge) else None

        )

        key = _ck(item_code, mc_group, _old_gauge_str)

        # old plan ใช้เป็น fallback เท่านั้น: ห้าม override baseline จาก detail_mc/booking

        # หา week ล่าสุดที่ ACTUAL_MC > 0 (ไม่ใช่แค่แถวแรก) และไม่ override ถ้า key มีข้อมูล booking จริงอยู่แล้ว

        if key not in booking_production_keys and (key not in last_production or w_idx > last_production[key]):

            last_production[key] = w_idx

            machines_in_use[key] = _old_actual_mc_int

            # Normalize SC/SO NO เล็กน้อย

            if sc_no and not pd.isna(sc_no):

                s = str(sc_no).strip().upper()

                if s.startswith("S") and s[1:].isdigit():

                    s = s[1:]

                # ตั้งค่าแค่ถ้ายังไม่มี

                if key not in last_sc_so_no:

                    last_sc_so_no[key] = s

    print(

        f"📋 เติม last_production จาก old_plan: {len([k for k in last_production])} รายการ (รวม)"

    )

# Merge booking history → last_production (booking ข้อมูลจริงแทนถ้า recent กว่า)

for _bk_key, _bk_widx in booking_last_production.items():

    if _bk_key not in last_production or _bk_widx > last_production[_bk_key]:

        last_production[_bk_key] = _bk_widx

        _raw_so = booking_last_so.get(_bk_key, "")

        # Normalize booking SO_NO: "S717492" → "717492" เพื่อให้ตรงกับ order SC/SO NO

        if _raw_so.startswith("S") and _raw_so[1:].isdigit():

            last_sc_so_no[_bk_key] = _raw_so[1:]

        else:

            last_sc_so_no[_bk_key] = _raw_so

    # เสมอ: ถ้ายังไม่มี last_sc_so_no ให้เซ็ตจาก booking (detail_mc ไม่มี SO info)

    if _bk_key not in last_sc_so_no or not last_sc_so_no.get(_bk_key):

        _raw_so = booking_last_so.get(_bk_key, "")

        if _raw_so.startswith("S") and _raw_so[1:].isdigit():

            last_sc_so_no[_bk_key] = _raw_so[1:]

        elif _raw_so:

            last_sc_so_no[_bk_key] = _raw_so

print(f"📚 last_production หลัง merge booking: {len(last_production)} รายการรวม")



# =========================

# TRACK WEEKLY JOB USAGE

# =========================

weekly_job_usage = {}  # {week: {mc_group: jobs_used}}

# Pre-populate weekly_job_usage จาก booking_final_ready25 (DETAIL sheet)
# ใช้ _IS_NEW_SETUP / _MC_INCREASE ที่ AVA_MC.py คำนวณถูกต้องแล้ว
# new_jobs ต่อ row:
#   _is_new_setup=True  → MC_USE_CEIL (ทุกเครื่องต้อง setup ใหม่)
#   _is_new_setup=False → _mc_increase (เฉพาะส่วนที่เพิ่ม)

if (
    not detail_mc.empty
    and "WEEK" in detail_mc.columns
    and "ITEM_CODE" in detail_mc.columns
    and "MC_USE_CEIL" in detail_mc.columns
    and "MC_GROUP" in detail_mc.columns
):
    _det = detail_mc.copy()
    _det["WEEK"] = pd.to_numeric(_det["WEEK"], errors="coerce")
    _det["MC_USE_CEIL"] = pd.to_numeric(_det["MC_USE_CEIL"], errors="coerce").fillna(0).astype(int)
    _det = _det.dropna(subset=["WEEK", "ITEM_CODE", "MC_GROUP"])
    _det["WEEK"] = _det["WEEK"].astype(int)
    _det["ITEM_CODE"] = _det["ITEM_CODE"].astype(str).str.strip().str.upper()
    _det["MC_GROUP"] = _det["MC_GROUP"].astype(str).str.strip().str.upper()

    _has_ava_flags = "_IS_NEW_SETUP" in _det.columns and "_MC_INCREASE" in _det.columns
    if _has_ava_flags:
        _det["_IS_NEW_SETUP"] = _det["_IS_NEW_SETUP"].fillna(True).astype(bool)
        _det["_MC_INCREASE"] = pd.to_numeric(_det["_MC_INCREASE"], errors="coerce").fillna(0).astype(int)
        print("✅ Pre-populate weekly_job_usage: ใช้ _IS_NEW_SETUP / _MC_INCREASE จาก AVA_MC")
    else:
        print("⚠️  _IS_NEW_SETUP ไม่อยู่ใน DETAIL — fallback: นับ MC_USE_CEIL ทั้งหมด (อาจ overcount)")

    _det_active = _det[_det["MC_USE_CEIL"] > 0].copy()
    # ตัดกลุ่มที่ไม่ใช่เครื่องถัก (IMPORT = ผ้านำเข้า) ออกก่อนนับ job — ไม่งั้นโควตา setup เพี้ยน
    _det_active = _det_active[~_det_active["MC_GROUP"].isin(_JOB_EXCLUDE_MC)]

    for _, _drow in _det_active.iterrows():
        _wk = int(_drow["WEEK"])
        _wk_idx = week_index(_wk)
        if _wk_idx is None or _wk_idx < TODAY_IDX:
            continue

        _mc_key = str(_drow["MC_GROUP"]).strip().upper()
        _mc_ceil = int(_drow["MC_USE_CEIL"])
        _item = str(_drow["ITEM_CODE"]).strip().upper()

        if _has_ava_flags:
            _is_new = bool(_drow["_IS_NEW_SETUP"])
            _mc_inc = int(_drow["_MC_INCREASE"])
            _new_jobs = _mc_ceil if _is_new else _mc_inc
        else:
            _new_jobs = _mc_ceil

        if _item == "FD6GNTLG27/58A0":
            print(f"[DEBUG] FD6GNTLG27/58A0 W{_wk}: mc_ceil={_mc_ceil}, is_new={_is_new if _has_ava_flags else 'N/A'}, new_jobs={_new_jobs}")

        if _new_jobs > 0:
            if _wk not in weekly_job_usage:
                weekly_job_usage[_wk] = {}
            weekly_job_usage[_wk][_mc_key] = weekly_job_usage[_wk].get(_mc_key, 0) + _new_jobs

total_booked = sum(sum(v.values()) for v in weekly_job_usage.values())

print(

    f"📋 Pre-loaded weekly_job_usage จาก booking_final_ready25 DETAIL"

    f" (new setup + เพิ่มเครื่อง, week>={TODAY_WEEK}): {total_booked} jobs"

)

# Snapshot ค่า OLD ก่อนเริ่ม loop ใหม่ (deep copy)

weekly_job_usage_old = {wk: dict(mc_dict) for wk, mc_dict in weekly_job_usage.items()}

# weekly_new_plan_usage: เฉพาะงานที่วางแผนใหม่ในรอบนี้ (ใช้กับ get_actual_mc_remain)

# แยกจาก weekly_job_usage ที่รวม booking เก่าด้วย (TOTAL_MC_REMAIN หักเก่าไปแล้ว)


# =========================
# ชีท UNPLANNED — คอลัมน์ชุดเดิม + DATE_IN เท่านั้น (ไม่เอาทุกคอลัมน์ของ order)
# =========================
UNPLANNED_COLS = ["SC_SO_NO", "ITEM_CODE", "MC_GROUP", "GAUGE", "ORDERS_QTY",
                  "PENDING_PLAN", "FG_WEEK", "DATE_IN", "UNPLANNED_QTY", "REASON"]


def _unplanned_row(order, reason, qty=None, gauge=None) -> dict:
    """สร้างแถวชีท UNPLANNED จาก order (pandas Series) — ดึงเฉพาะคอลัมน์ใน UNPLANNED_COLS

    qty   = จำนวนที่ยังไม่ได้วาง (ไม่ส่ง = ใช้ Pending Plan ทั้งก้อน)
    gauge = เกจที่ต้องการระบุเอง (ไม่ส่ง = ใช้ MC_GUAGE ของ order)
    """
    _pend = pd.to_numeric(order.get("Pending Plan", 0), errors="coerce")
    _pend = 0.0 if pd.isna(_pend) else float(_pend)
    # 'MC GROUP' (เว้นวรรค) เป็นคอลัมน์ที่ Planning สร้างเพิ่มให้ orders เท่านั้น
    # แถวที่มาจาก order_excluded.xlsx เป็นข้อมูลดิบ มีแค่ 'MC_GROUP' (ขีดล่าง) → ต้อง fallback
    _mc = order.get("MC GROUP", "")
    if pd.isna(_mc) or str(_mc).strip().lower() in ("", "nan", "none"):
        _mc = order.get("MC_GROUP", "")
    if pd.isna(_mc):
        _mc = ""
    return {
        "SC_SO_NO": str(order.get("SO_NO", order.get("SC/SO NO", ""))).strip(),
        "ITEM_CODE": order.get("Item Code", ""),
        "MC_GROUP": str(_mc).strip().upper(),
        "GAUGE": order.get("MC_GUAGE", "") if gauge is None else gauge,
        "ORDERS_QTY": order.get("Orders.Qty", ""),
        "PENDING_PLAN": _pend,
        "FG_WEEK": order.get("FG Week", ""),
        "DATE_IN": order.get("DATE_IN", ""),
        "UNPLANNED_QTY": _pend if qty is None else qty,
        "REASON": reason,
    }


# =========================
# PLANNING LOOP (รันรอบเดียว — S9 เหลือเฉพาะ routing ของ item ใน "S9 Only")
# =========================
def _run_planning_loop() -> list:
    global weekly_new_plan_usage, weekly_new_plan_pool_usage, remaining_week_cap, remaining_week_cap_owner
    global _type_special_weekly_usage
    global cylinder_change_count, cylinder_adjustments, _cylinder_change_for_item
    global _cylinder_change_start_map, _cylinder_change_mc_count, _cylinder_change_done
    global _carry_cyl_pending, _current_order_rdd_idx, _s9_weekly_usage
    global orders, sc_so_no, plans, _skip_no_cap, _skip_no_mc_group, _skip_no_factory, _skip_not_in_master, _skip_qty_short

    # Reset state ที่ถูก init ก่อน line 6737 (ต้อง reset ทุก pass)
    _type_special_weekly_usage = {}
    cylinder_change_count = {}
    # Pre-load จำนวน cylinder ที่เปลี่ยนจริงไปแล้วจาก Job sheet (MasterMC)
    for _jw, _jgroups in _job_cylinder_done.items():
        for _jg, _jc in _jgroups.items():
            cylinder_change_count.setdefault(_jw, {})[_jg] = _jc
    cylinder_adjustments = {}
    _cylinder_change_for_item = {}
    _cylinder_change_start_map = {}
    _cylinder_change_mc_count = {}
    _cylinder_change_done = set()
    _carry_cyl_pending = {}
    _current_order_rdd_idx = None
    _s9_weekly_usage = {}
    # ล้างสี YD ค้างจาก pass ก่อน (กัน color-change setup ผิดข้าม pass)
    # last_sub_color ไม่ได้ pre-load จาก booking จึงล้างทิ้งได้ปลอดภัย (ต่างจาก last_production/machines_in_use)
    last_sub_color.clear()

    weekly_new_plan_usage = {}  # {week: {(type_1,gauge): new_plan_machines}}
    weekly_new_plan_pool_usage = {}  # {week: {pool_label: new_plan_machines}} — สำหรับพูลแยก (SKP/SKPTA·SKPLE)

    # Pre-populate weekly_new_plan_usage จาก old plan (NEW plans รอบก่อน)

    # เพื่อให้ get_actual_mc_remain หัก capacity ที่ old plan ใช้ไปแล้ว

    if not old_plan_df.empty:

        _old_plan_preloaded = 0

        for _, _row in old_plan_df.iterrows():

            # Skip OLD plan rows if PLAN_SOURCE exists (safety check)

            # 🔧 FIX: Pre-load เฉพาะ NEW plan เพื่อป้องกัน double-count

            # summary_mc.TOTAL_MC_REMAIN ถูกหัก MC_USE_CEIL (booking เก่า) อยู่แล้ว

            # ถ้า pre-load OLD plan เข้า weekly_new_plan_usage → get_actual_mc_remain() หักซ้ำ

            if "PLAN_SOURCE" in old_plan_df.columns:

                if str(_row.get("PLAN_SOURCE", "")).strip().upper() != "NEW":
                    continue
            _op_week = _row.get("PLAN_WEEK") or _row.get("PLAN WEEK")
            _op_mc = _row.get("MC_GROUP") or _row.get("MC GROUP") or _row.get("MC")
            _op_machines = (
                _row.get("ACTUAL_MC")
                or _row.get("AVAILABLE_MACHINES")
                or _row.get("REQUIRED_MC")
                or _row.get("AVAILABLE_MACHINES")
            )

            _op_gauge = (

                _row.get("MC_GUAGE")

                or _row.get("MC GUAGE")

                or _row.get("GUAGE")

                or ""

            )

            if pd.isna(_op_week) or pd.isna(_op_mc) or pd.isna(_op_machines):

                continue

            try:

                _op_week = int(_op_week)

                _op_mc = str(_op_mc).strip().upper()

                _op_machines = int(float(_op_machines))

                _op_gauge = str(_op_gauge).strip() if _op_gauge and not pd.isna(_op_gauge) else ""

            except Exception:

                continue

            # Apply same redirect logic as main loop for pool deduction

            _redirected_mc, _redirected_gauge = _apply_mc_redirect(_op_mc, _op_gauge)

            _wpu_key = (_mc_to_type1(_redirected_mc, _redirected_gauge), _redirected_gauge)

            if _op_week not in weekly_new_plan_usage:

                weekly_new_plan_usage[_op_week] = {}

            weekly_new_plan_usage[_op_week][_wpu_key] = (

                weekly_new_plan_usage[_op_week].get(_wpu_key, 0) + _op_machines

            )

            # พูลแยก (SKP/SKPTA·SKPLE): นับต่อ pool label ด้วย (ให้ get_actual_mc_remain หักถูกพูล)
            _op_pool = _mc_to_pool_label(_redirected_mc, _redirected_gauge)
            if _op_pool:
                weekly_new_plan_pool_usage.setdefault(_op_week, {})[_op_pool] = (
                    weekly_new_plan_pool_usage.setdefault(_op_week, {}).get(_op_pool, 0) + _op_machines
                )

            _old_plan_preloaded += _op_machines

        print(

            f"📋 Pre-loaded weekly_new_plan_usage จาก old plan (NEW): "

            f"{_old_plan_preloaded} machines across {len(weekly_new_plan_usage)} weeks"

        )

    # cap ที่เหลือในสัปดาห์เมื่อ order จบก่อนใช้สุด — ใช้ผลิต FG ถัดไป (item+machine เดียวกัน)

    remaining_week_cap = {}  # {(week, item_code, mc_group): remaining_capacity_units}

    remaining_week_cap_owner = {}  # {(week, item_code, mc_group): owner marker (item-level carry)}



    # =========================

    # MERGE SAME SC + SAME ITEM (+ FG Week เดียวกัน)

    # =========================

    # ถ้า SC/SO NO เหมือนกัน + SO_NO เหมือนกัน + Item Code เหมือนกัน + FG Week เดียวกัน → รวมเป็น 1 row ผลิตทีเดียว

    # ต้องใช้ SO_NO ด้วยเพราะ SC เดียวกันอาจมีหลาย SO (คนละ order จริง) → ห้ามรวม qty

    # ถ้า FG Week ต่างกัน → คง row แยกไว้ (deadline ต่างกัน → plan แยก)

    _so_no_col = "SO_NO" if "SO_NO" in orders.columns else None
    _po_no_col = "PO_NO" if "PO_NO" in orders.columns else None

    _grp_keys = ["SC/SO NO"] + (["SO_NO"] if _so_no_col else []) + (["PO_NO"] if _po_no_col else []) + ["Item Code", "MC GROUP", "MC_GUAGE"]

    if "FG Week" in orders.columns:

        _grp_keys = _grp_keys + ["FG Week"]

    _sum_cols = [

        c

        for c in ["Orders.Qty", "Plan Qty", "Pending Plan", "Confirm"]

        if c in orders.columns

    ]

    _min_cols = [c for c in ["DYE_END_DATE"] if c in orders.columns]

    _first_cols = [c for c in orders.columns if c not in _grp_keys + _sum_cols + _min_cols]

    _agg_dict = {}

    _agg_dict.update({c: "sum" for c in _sum_cols})

    _agg_dict.update({c: "min" for c in _min_cols})

    _agg_dict.update({c: "first" for c in _first_cols})

    _orders_before = len(orders)

    orders = orders.groupby(_grp_keys, sort=False).agg(_agg_dict).reset_index()

    print(

        f"✅ รวม orders same SC+Item: {_orders_before} → {len(orders)} rows (merged {_orders_before - len(orders)} rows)"

    )



    # =========================

    # MAIN PLANNING

    # =========================

    plans = []
    # ให้ _sp_used_from_plans() เห็นแถวที่วางไปแล้ว (ใช้ list เดียวกัน ไม่ใช่ copy)
    global _PLANS_REF
    _PLANS_REF = plans

    _skip_no_cap = []  # เก็บ item ที่ไม่มี cap เพื่อแสดงรวมท้ายสุด

    _skip_qty_short = []  # เก็บ order ที่เข้า loop วางแผนแล้วแต่วางไม่ครบ (เครื่องไม่พอ/หมดสัปดาห์/ชน safety guard)

    _skip_no_mc_group = []  # เก็บ order ที่ไม่มี MC GROUP → ไม่วางแผน

    _skip_no_factory = []  # เก็บ order ที่ MC GROUP ไม่มี FACTORY_TYPE → ไม่วางแผน

    _skip_not_in_master = []  # เก็บ order ที่ MC+Gauge ไม่มีใน MasterMC → ไม่วางแผน

    new_plan_started_items = set()  # ติดตาม (item, mc_group) ที่เริ่มการผลิตใน new plan แล้ว

    # YD-ORDERS: ล็อกจำนวนเครื่องต่อ (item, mc_group, plan_week)

    # ภายใน week เดียวกัน ห้ามเพิ่ม/ลดเครื่อง — SO ถัดไปใช้เครื่องเท่าเดิม

    _yd_week_locked_mc: dict = {}  # {(item, mc_group, plan_week): (actual_mc, carryover_mc, new_mc)}

    locked_mc_group_for: dict = (

        {}

    )  # ล็อก MC_GROUP (highest-cap) ต่อ (sc_so_no, item) ให้ FG Week ต่างๆ ใช้ร่วมกัน

    # ติดตาม last plan week index ต่อ (sc_so_no, item) เพื่อบังคับให้ FG_WEEK ถัดไป

    # เริ่มหลัง FG_WEEK ก่อนหน้าจบ (ไม่ผลิตซ้อนกัน)

    _last_fg_plan_idx: dict = {}  # {(sc_so_no, item): last_week_index}

    # เรียง orders ตาม TARGET_KNIT (rdd_idx) จริง ไม่ใช่ FG Week

    # เพราะ order type ต่างกัน TARGET_KNIT ต่างกัน (LAB-DIP = FG-1, SC-ORDERS = FG-offset ตาม FOB_TYPE)





    def _order_rdd_idx(row):

        """คำนวณ rdd_idx (TARGET_KNIT index) สำหรับ sort เท่านั้น"""

        try:

            fg_w = row.get("FG Week")

            o_type = str(row.get("Orders Type", "")).strip()

            if pd.isna(fg_w) or fg_w is None:

                return 99999



            fg_w_str = str(int(fg_w)).strip()

            if len(fg_w_str) >= 6:

                _yr = int(fg_w_str[:4])

                _wk = int(fg_w_str[4:])

            elif len(fg_w_str) == 5:

                _yr = int(fg_w_str[:4])

                _wk = int(fg_w_str[4:])

            else:

                _yr = TODAY.year

                _wk = int(fg_w_str)

            _row = calendar_week[

                (calendar_week["YEAR"] == _yr) & (calendar_week["WEEK"] == _wk)

            ]

            if _row.empty:

                return 99999



            _raw_idx = _row.index[0]

            _rdd = max(0, _raw_idx - target_offset_weeks_for_row(row))

            if o_type == "LAB-DIP":

                # LAB-DIP: sort เรียงตาม TODAY_IDX + 2 (เริ่มและเสร็จใน week +2)

                _rdd = min(len(calendar_week) - 1, TODAY_IDX + 2)

            return _rdd



        except Exception:

            return 99999



    orders["_sort_rdd_idx"] = orders.apply(_order_rdd_idx, axis=1)

    # Priority: orders with NAY_COLOR or COLOR_DESC data, or ORDER_TYPE LAB-DIP/YD-ORDERS should be planned first

    def _has_color_data(row):

        nay_color = str(row.get("NAY_COLOR", "")).strip()

        color_desc = str(row.get("COLOR_DESC", "")).strip()

        order_type = str(row.get("Orders Type", "")).strip().upper()

        return 0 if (nay_color and nay_color != "nan") or (color_desc and color_desc != "nan") or order_type in ("LAB-DIP", "YD-ORDERS") else 1

    orders["_sort_color_priority"] = orders.apply(_has_color_data, axis=1)

    # tiebreaker 1: FG Week raw (น้อยกว่า = ก่อน) → หาก rdd_idx เท่ากัน ให้ FG เร็วกว่าก่อน

    # tiebreaker 2: PENDING_PLAN (น้อยกว่า = ก่อน) → ของน้อยจบไว ปล่อยเครื่องให้ item เดียวกัน carry ต่อ

    _fg_week_col = "FG Week" if "FG Week" in orders.columns else None

    _pending_col = "Pending Plan" if "Pending Plan" in orders.columns else None

    _sort_cols = ["_sort_color_priority", "_sort_rdd_idx"]

    if _fg_week_col:

        orders["_sort_fg_week"] = pd.to_numeric(orders[_fg_week_col], errors="coerce").fillna(99999999)

        _sort_cols.append("_sort_fg_week")

    if _pending_col:

        orders["_sort_pending"] = pd.to_numeric(orders[_pending_col], errors="coerce").fillna(0)

        _sort_cols.append("_sort_pending")

    orders_sorted = orders.sort_values(_sort_cols, na_position="last")

    _drop_cols = [c for c in ["_sort_color_priority", "_sort_rdd_idx", "_sort_fg_week", "_sort_pending"] if c in orders_sorted.columns]

    orders_sorted = orders_sorted.drop(columns=_drop_cols)

    orders = orders.drop(columns=[c for c in _drop_cols if c in orders.columns])

    print(f"[DEBUG ORDERS] Total orders: {len(orders)}")

    # ---------- จ้างทอ (Outsource split) ----------
    # user แบ่ง qty ของ item ไปจ้างทอจากหน้าเว็บ → data_plan/outsource_split.json
    # ตัด Pending Plan ออกจาก order ของ item นั้น (ใบ FG/RDD เร็วสุดก่อน) แล้ว clone เป็น order ใหม่
    # ที่ติดธง OUTSOURCE_FORCE → planning loop จะ route ก้อนนี้เข้า S9 pool (COMKN) ซึ่งเป็นเครื่อง
    # จ้างทอ แยกพูลจากเครื่องในบ้าน ส่วนที่เหลือ (pending ที่ถูกตัดแล้ว) วางแผนในบ้านตาม logic เดิม
    _outsource_split = _load_outsource_split()
    if _outsource_split:
        orders_sorted = orders_sorted.copy()
        orders_sorted["OUTSOURCE_FORCE"] = False
        orders_sorted["OUTSOURCE_WEEK"] = None
        _os_clones = []
        _os_item_col = orders_sorted["Item Code"].astype(str).str.strip().str.upper()
        for _os_item, _os_cfg in _outsource_split.items():
            _os_qty = float(_os_cfg["outsource_qty"])
            _os_week = _os_cfg.get("start_week")
            _os_mask = (_os_item_col == _os_item) & (
                pd.to_numeric(orders_sorted["Pending Plan"], errors="coerce").fillna(0) > 0
            )
            _os_rows = orders_sorted[_os_mask]
            if _os_rows.empty:
                print(f"[OUTSOURCE] {_os_item}: ไม่มีของค้างในแผนแล้ว → ข้ามการจ้างทอ")
                continue
            # ตัดจากใบที่ FG/RDD เร็วที่สุดก่อน (งานด่วนได้จ้างทอช่วยก่อน)
            if "FG Week" in _os_rows.columns:
                _os_rows = _os_rows.assign(
                    _os_fg=pd.to_numeric(_os_rows["FG Week"], errors="coerce").fillna(99999999)
                ).sort_values("_os_fg")
            _left = _os_qty
            for _oi in _os_rows.index:
                if _left <= 0:
                    break
                _pend = pd.to_numeric(orders_sorted.at[_oi, "Pending Plan"], errors="coerce")
                _pend = 0.0 if pd.isna(_pend) else float(_pend)
                _cut = min(_pend, _left)
                if _cut <= 0:
                    continue
                orders_sorted.at[_oi, "Pending Plan"] = _pend - _cut
                _clone = orders_sorted.loc[_oi].copy()
                _clone["Pending Plan"] = _cut
                _clone["OUTSOURCE_FORCE"] = True
                _clone["OUTSOURCE_WEEK"] = _os_week
                _os_clones.append(_clone)
                _left -= _cut
            _done = _os_qty - _left
            print(f"[OUTSOURCE] {_os_item}: จ้างทอ {_done:.0f} กก. "
                  f"(ขอ {_os_qty:.0f}) W{_os_week} → ที่เหลือวางแผนในบ้านตามปกติ")
            if _left > 0:
                print(f"[OUTSOURCE] {_os_item}: ของค้างไม่พอ — จ้างทอได้แค่ {_done:.0f} จาก {_os_qty:.0f} กก.")
        if _os_clones:
            orders_sorted = pd.concat(
                [orders_sorted, pd.DataFrame(_os_clones)], ignore_index=True
            )
            print(f"[OUTSOURCE] สร้าง order จ้างทอ {len(_os_clones)} ใบ")



    # 🔧 FIX: ตรวจสอบว่า Item+Week+MC_GROUP นี้มีการ plan ไปแล้วหรือไม่

    # เพื่อป้องกันการซ้ำซ้อนของ orders ที่มี Item+MC_GROUP เดียวกันใน week เดียวกัน

    # โดยไม่สนใจว่าจะเป็น SC หรือ TARGET_KNIT อะไร (เพราะอาจมีการ merge SC)

    _existing_item_week_mc = set()

    # 🔧 FIX: ติดตาม qty ที่วางแผนไปแล้วสำหรับแต่ละ item (สะสมข้าม order/SC/FG)

    # ใช้สำหรับคำนวณ next_fg_qty ที่แม่นยำ — ไม่นับ demand จาก order ที่ plan ไปแล้ว

    _item_cumulative_planned = {}  # {item_code: total_qty_planned_so_far}

    print(f"[DEBUG DUPLICATE] Starting with {len(plans)} existing plans")



    for _, order in orders_sorted.iterrows():

        item = order["Item Code"]

        order_qty = order["Orders.Qty"]  # ปริมาณที่สั่งทั้งหมด

        plan_qty = order["Plan Qty"]  # ปริมาณที่วางแผนไปแล้ว (รอ approve)

        pending_plan = pd.to_numeric(order.get("Pending Plan", 0), errors="coerce")

        pending_plan = 0.0 if pd.isna(pending_plan) else float(pending_plan)

        sc_so_no = str(order.get("SO_NO", order.get("SC/SO NO", ""))).strip()  # ใช้แยก order ต่างสี

        sub_color = str(order.get("SUB_COLOR", "")).strip()  # SUB_COLOR สำหรับ YD-ORDERS carry logic

        # ถ้าไม่มี MC GROUP → เก็บไว้แจ้ง ไม่วางแผน

        _ord_mc_grp = str(order.get("MC GROUP", "")).strip().upper()

        if not _ord_mc_grp or _ord_mc_grp in ("", "NAN", "NONE"):

            _skip_no_mc_group.append(_unplanned_row(order, "ไม่มี MC GROUP"))

            continue

        # ถ้า MC GROUP ไม่มีใน FACTORY_TYPE_MAP → เก็บไว้แจ้ง ไม่วางแผน

        _ord_factory = FACTORY_TYPE_MAP.get(_ord_mc_grp, "")

        if not _ord_factory or _ord_factory in ("", "UNKNOWN"):

            _skip_no_factory.append(
                _unplanned_row(order, f"MC GROUP '{_ord_mc_grp}' ไม่มี FACTORY_TYPE")
            )

            continue

        # ถ้า (MC GROUP, Gauge) ไม่มีใน MasterMC → เก็บไว้แจ้ง ไม่วางแผน
        _ord_gauge_norm = _normalize_gauge(order.get("MC_GUAGE", ""))
        if (_ord_mc_grp, _ord_gauge_norm) not in _VALID_MC_GAUGE_SET:
            _skip_not_in_master.append(_unplanned_row(
                order,
                f"MC GROUP '{_ord_mc_grp}' Gauge '{_ord_gauge_norm}' ไม่มีใน MasterMC",
                gauge=_ord_gauge_norm,
            ))
            continue

        # ถ้า Pending Plan = 0 แสดงว่า order นี้วางแผนครบแล้ว ไม่ต้องวางแผนซ้ำ

        print(f"[DEBUG ORDERS] Order {sc_so_no}: item={item}, pending_plan={pending_plan}")

        if pending_plan <= 0:

            print(f"[DEBUG ORDERS] Skipping order {sc_so_no} - pending_plan <= 0")

            continue



        order_type = order["Orders Type"]

        fg_week = order.get("FG Week")

        # ตรวจสอบว่า SO นี้ผลิตไปแล้วบางส่วนใน booking จริงหรือไม่

        _so_try = ["S" + sc_so_no.lstrip("S"), sc_so_no.lstrip("S")]

        already_made = 0.0

        for _s in _so_try:

            _s_up = _s.upper()

            if _s_up in booking_produced_qty:

                already_made = booking_produced_qty[_s_up]

                break



        # qty_left = Pending Plan (ยังไม่ได้วางแผน) หักส่วนที่ผลิตจริงไปแล้วจาก booking

        qty_left = max(0.0, pending_plan - already_made)

        # ก้อนจ้างทอ (clone จาก order เดิม) — pending คือยอดที่ user แบ่งไว้แล้ว
        # ไม่ต้องหัก already_made ซ้ำ เพราะถูกหักที่ order ต้นฉบับ (ก้อนในบ้าน) ไปแล้ว
        _outsource_force = bool(order.get("OUTSOURCE_FORCE", False))
        _outsource_week = order.get("OUTSOURCE_WEEK")
        if _outsource_force:
            qty_left = max(0.0, pending_plan)

        # Special rule: เช็คจาก Itemcore และ Customer (CORE ITEM)

        # ถ้า item อยู่ใน Itemcore และ customer เป็น "ที คัลเจอร์ บจ." หรือ "CENTER DOMESTIC" -> ผลิตต่อท้าย

        # ถ้า item อยู่ใน Itemcore แต่ customer ไม่ตรง -> ผลิตตาม target ปกติ

        CORE_CUSTOMERS = {"ที คัลเจอร์ บจ.", "CENTER DOMESTIC"}

        rts_local_force = None

        is_core_item = False  # flag สำหรับ output column

        # เช็คว่า item อยู่ใน Itemcore หรือไม่

        item_upper = str(item).strip().upper()

        item_in_itemcore = item_upper in itemcore_lookup

        if item_in_itemcore:

            # Item อยู่ใน Itemcore - เช็คว่า customer เป็น core customer หรือไม่

            actual_customer = str(order.get("Customer", "")).strip()

            customer_match = (actual_customer in CORE_CUSTOMERS)

            if customer_match:

                # Item อยู่ใน Itemcore + Customer เป็น core -> ใช้กฎ CORE ITEM (ผลิตต่อท้าย)

                is_core_item = True

                try:

                    dm = detail_mc[

                        detail_mc["ITEM_CODE"].astype(str).str.upper().str.strip()

                        == str(item).upper()

                    ]

                    if not dm.empty:

                        last_w = int(dm["WEEK"].dropna().astype(int).max())

                        row_last = dm[dm["WEEK"] == last_w].iloc[-1]

                        sel_mc = str(row_last.get("MC_GROUP", "")).strip().upper()

                        sel_mc_used = int(row_last.get("MC_USE_CEIL", 0) or 0)

                        start_after = next_week(last_w)

                        # Build maps per MC_GROUP: last booked week and machines used

                        last_old_by_mc = {}

                        machines_by_mc = {}

                        daily_cap_by_mc = {}

                        booking_remain_cap_by_mc = {}  # remaining cap ในสัปดาห์สุดท้าย booking

                        for mc, grp in dm.groupby("MC_GROUP"):

                            try:

                                w = int(grp["WEEK"].dropna().astype(int).max())

                            except Exception:

                                continue



                            last_old_by_mc[str(mc).strip().upper()] = w

                            # get MC_USE_CEIL from the last week that had > 0 machines

                            # (week order may have 0 at the end e.g. paused week → skip those)

                            grp_active = grp[

                                pd.to_numeric(grp["MC_USE_CEIL"], errors="coerce").fillna(0)

                                > 0

                            ]

                            if not grp_active.empty:

                                w_active = int(

                                    grp_active["WEEK"].dropna().astype(int).max()

                                )

                                last_active_row = grp_active[

                                    grp_active["WEEK"] == w_active

                                ].iloc[-1]

                            else:

                                last_active_row = grp[grp["WEEK"] == w].iloc[-1]

                            try:

                                machines_by_mc[str(mc).strip().upper()] = int(

                                    last_active_row.get("MC_USE_CEIL", 0) or 0

                                )

                            except Exception:

                                machines_by_mc[str(mc).strip().upper()] = 0

                            # try to get daily cap from item_cap_data per mc

                            # ใช้ cap น้อยที่สุดของ item นี้ในการคำนวณ

                            try:

                                _all_cap_for_item = item_cap_data[

                                    item_cap_data["ITEM_CODE"] == item

                                ]

                                if not _all_cap_for_item.empty:

                                    daily_cap_by_mc[str(mc).strip().upper()] = float(

                                        _all_cap_for_item["CAP ทอ"].min()

                                    )

                                else:

                                    cap_row2 = item_cap_data[

                                        item_cap_data["MC_GROUP"] == str(mc).strip().upper()

                                    ]

                                    if not cap_row2.empty:

                                        daily_cap_by_mc[str(mc).strip().upper()] = (

                                            cap_row2.iloc[0].get("CAP ทอ", None)

                                        )

                            except Exception:

                                daily_cap_by_mc[str(mc).strip().upper()] = None

                            # คำนวณ remaining capacity ในสัปดาห์สุดท้ายของ booking

                            # เพื่อให้ new SO สามารถ carry ต่อในสัปดาห์เดียวกันได้

                            # โดยจำกัด qty ตาม old item ที่ใช้เครื่องร่วมกัน (MC_USE fraction)

                            try:

                                _last_bk_row = grp[grp["WEEK"] == w]

                                if not _last_bk_row.empty:

                                    _mc_use_ceil = float(_last_bk_row.iloc[0].get("MC_USE_CEIL", 0) or 0)

                                    _bk_kp_weight = float(_last_bk_row.iloc[0].get("KP_WEIGHT", 0) or 0)

                                    # WORKING_DAY มาจากชีท Work Day แล้ว — ใช้ตรง ๆ (float: 5.5 วันต้องไม่ปัดเป็น 5)
                                    _bk_wd = float(_last_bk_row.iloc[0].get("WORKING_DAY", 0) or 0)

                                    _bk_gauge = str(_last_bk_row.iloc[0].get("GUAGE", "")).strip()

                                    _bk_plan_cap = _get_capacity_for_mc_group(

                                        item, str(mc).strip().upper(), _bk_gauge

                                    )

                                    # remaining = total machine capacity − qty ที่ old item ผลิตจริง (KP_WEIGHT)

                                    # สูตร: (MC_USE_CEIL × wd × daily_cap) − KP_WEIGHT

                                    # เช่น FD4DRTPC88A0 w17: (2×8×103.06)−1143 = 1648−1143 = 505 units

                                    # 🔧 FIX: ตรวจสอบว่า booking week นี้มีการ setup เพิ่มเครื่องหรือไม่

                                    # โดยเปรียบเทียบ MC_USE_CEIL กับ week ก่อนหน้าใน grp

                                    _sorted_bk_weeks = sorted(grp["WEEK"].dropna().astype(int).tolist())

                                    _w_pos_bk = _sorted_bk_weeks.index(w) if w in _sorted_bk_weeks else -1

                                    _prev_bk_mc = 0

                                    _prev_w_bk_val = None

                                    if _w_pos_bk > 0:

                                        _prev_w_bk = _sorted_bk_weeks[_w_pos_bk - 1]

                                        _prev_w_bk_val = _prev_w_bk

                                        _prev_bk_row = grp[grp["WEEK"] == _prev_w_bk]

                                        if not _prev_bk_row.empty:

                                            _prev_bk_mc = float(_prev_bk_row.iloc[0].get("MC_USE_CEIL", 0) or 0)

                                    # 🔧 FIX v2: ถ้ามี gap > 1 week ระหว่าง prev booking week และ current week

                                    # → เครื่องทุกตัวต้อง setup ใหม่ (แม้ MC count จะเท่ากัน)

                                    _bk_week_gap = (int(w) - int(_prev_w_bk_val)) if _prev_w_bk_val is not None else 999

                                    if _bk_week_gap > 1:

                                        _bk_new_mc = float(_mc_use_ceil)  # all machines are new setup

                                        _bk_carry_mc = 0.0

                                    else:

                                        _bk_new_mc = max(0.0, _mc_use_ceil - _prev_bk_mc)

                                        _bk_carry_mc = _mc_use_ceil - _bk_new_mc

                                    # คำนวณ total_cap โดยหัก setup days สำหรับเครื่องใหม่

                                    _total_cap = (_bk_carry_mc * _bk_wd + _bk_new_mc * max(0, _bk_wd - SETUP_DAYS)) * _bk_plan_cap

                                    if _bk_new_mc > 0:

                                        print(f"[BOOKING SETUP DETECT] {item} W{w} MC={mc}: new_mc={_bk_new_mc:.0f} carry_mc={_bk_carry_mc:.0f} → adjusted total_cap={_total_cap:.2f} (naive={_mc_use_ceil*_bk_wd*_bk_plan_cap:.2f})")

                                    if _bk_plan_cap > 0 and _mc_use_ceil > 0:

                                        _bk_rem = max(0.0, _total_cap - _bk_kp_weight)

                                        if _bk_rem > 0:

                                            booking_remain_cap_by_mc[str(mc).strip().upper()] = _bk_rem

                            except Exception:

                                pass

                        rts_local_force = {

                            "last_old_by_mc": last_old_by_mc,

                            "machines_by_mc": machines_by_mc,

                            "daily_cap_by_mc": daily_cap_by_mc,

                            "booking_remain_by_mc": booking_remain_cap_by_mc,

                        }

                except Exception:

                    rts_local_force = None



        # ----------------------

        # RDD Check and Urgent Planning

        # ----------------------

        # rdd_idx = row index ใน calendar_week ของ RDD จริง (ใช้แทน week number

        # เพื่อรองรับการเปรียบเทียบข้ามปีได้ถูกต้อง เช่น order FG ปี 2027)

        rdd_idx = None  # row index ใน calendar_week สำหรับ comparison

        fg_week_int = None  # week number (1-53) สำหรับ output/display เท่านั้น

        fg_week_num = None  # FG week number (1-53) สำหรับ FG constraint check

        if pd.notna(fg_week):

            fg_week_str = str(int(fg_week))

            if len(fg_week_str) >= 6:

                fg_year = int(fg_week_str[:4])

                fg_week_num = int(fg_week_str[4:])

            elif len(fg_week_str) == 5:

                fg_year = int(fg_week_str[:4])

                fg_week_num = int(fg_week_str[4:])

            elif len(fg_week_str) <= 2:  # รูปแบบ WW (เช่น 13) → ใช้ปีปัจจุบัน

                fg_year = TODAY.year

                fg_week_num = int(fg_week_str)

            else:

                fg_year = TODAY.year

                fg_week_num = int(fg_week)

            # หา row index ใน calendar_week ด้วย YEAR + WEEK (รองรับข้ามปี)

            _fg_row = calendar_week[

                (calendar_week["YEAR"] == fg_year) & (calendar_week["WEEK"] == fg_week_num)

            ]

            _fg_raw_idx = None

            if not _fg_row.empty:

                _fg_raw_idx = int(_fg_row.index[0])

            else:

                # FALLBACK: FG week ไม่มีในปฏิทิน เพราะปฏิทินบริษัท "ยุบ" สัปดาห์หยุดยาว
                # รวมเข้าสัปดาห์ถัดไป (เช่น ปี 2026 ไม่มี W16 → รวมใน W17 สงกรานต์,
                # ไม่มี W31 → รวมใน W32 หยุดบริษัท) แต่ฝ่ายขายยังกรอก FG = 202631 ได้
                # เดิมเงียบ ๆ ปล่อย rdd_idx = None → order หลุด logic RDD ทั้งชุด
                # (ไม่คำนวณเครื่อง, ไม่ทำ backward, ใช้ spare cylinder ไม่ได้)
                # วิธีแก้: แปลง (year, week) กลับเป็นวันที่จริง แล้วหาสัปดาห์ที่ครอบวันนั้น
                # สัปดาห์ของบริษัทเริ่มวันศุกร์ = ISO Monday ของสัปดาห์นั้น - 3 วัน
                try:

                    _fg_anchor = pd.Timestamp(date.fromisocalendar(fg_year, fg_week_num, 1)) - pd.Timedelta(days=3)

                    _fg_hit = calendar_week.index[
                        (calendar_week["WEEK_START"] <= _fg_anchor)
                        & (calendar_week["WEEK_END"] >= _fg_anchor)
                    ]

                    if len(_fg_hit) > 0:

                        _fg_raw_idx = int(_fg_hit[0])

                        print(
                            f"[FG WEEK FALLBACK] {item} SC {sc_so_no}: FG {fg_year}W{fg_week_num} "
                            f"ไม่มีในปฏิทิน (สัปดาห์ถูกยุบรวม) → ใช้สัปดาห์ที่ครอบ "
                            f"{_fg_anchor.date()} = idx {_fg_raw_idx} "
                            f"(W{int(calendar_week.iloc[_fg_raw_idx]['WEEK'])})"
                        )

                except (ValueError, TypeError):

                    pass  # ISO ไม่มีสัปดาห์นี้จริง (เช่น W53 ของปีที่มี 52 สัปดาห์) → คง None

                if _fg_raw_idx is None:

                    print(
                        f"⚠️  [FG WEEK MISS] {item} SC {sc_so_no}: FG {fg_week} หาสัปดาห์บนปฏิทินไม่ได้ "
                        f"→ rdd_idx=None (order นี้จะวางแผนโดยไม่รู้ deadline)"
                    )

            if _fg_raw_idx is not None:

                # คำนวณ offset (FOB_TYPE / DOUBLE-30 เกจ 16-19 / SINGLE yarn เส้นเดียว)

                offset_weeks = target_offset_weeks_for_row(order)



                # หัก offset สัปดาห์ (RDD = FG Week - offset ตาม FOB_TYPE) ด้วย index arithmetic

                rdd_idx = max(0, _fg_raw_idx - offset_weeks)

                fg_week_int = int(calendar_week.iloc[rdd_idx]["WEEK"])  # สำหรับ display

        # LAB-DIP: deadline = TODAY + 2 weeks (ต้องเสร็จภายใน week ที่เริ่มผลิต)

        if order_type == "LAB-DIP":

            rdd_idx = min(len(calendar_week) - 1, TODAY_IDX + 2)

            fg_week_int = int(calendar_week.iloc[rdd_idx]["WEEK"])  # อัพเดท display

        if rdd_idx is not None and rdd_idx < TODAY_IDX:

            # RDD ผ่านไปแล้ว = URGENT!

            # สำหรับ urgent order ต้องใช้ความสามารถสูงสุด

            # อาจจะต้องเพิ่มเครื่อง แต่ต้องไม่เกิน job/day capacity

            # urgent_mode = True  # DISABLED

            urgent_mode = False

        else:

            urgent_mode = False



        # ----------------------

        # determine order week based on order type



        # ----------------------

        if order_type == "LAB-DIP":

            # LAB-DIP: +2 weeks from planning date (TODAY) และต้องเสร็จภายใน week นั้น

            if TODAY_IDX + 2 < len(calendar_week):

                order_week = calendar_week.iloc[TODAY_IDX + 2]["WEEK"]

            else:

                _skip_qty_short.append(_unplanned_row(
                    order, "ปฏิทินไม่มีข้อมูลถึง TODAY+2 สัปดาห์ (LAB-DIP)", qty=qty_left
                ))

                continue



        elif order_type in ("SC-ORDERS", "ผ้าเติม"):

            # SC-ORDERS / ผ้าเติม: เริ่มวางแผนที่ TODAY+2 โดยปกติ

            # แต่ถ้า RDD ผ่านไปแล้ว → เริ่มผลิตทันทีที่ TODAY_WEEK (urgent)

            if rdd_idx is not None and rdd_idx < TODAY_IDX:

                # RDD ผ่านไปแล้ว = URGENT! เริ่มผลิตทันที

                order_week = calendar_week.iloc[TODAY_IDX]["WEEK"]

                urgent_mode = True  # เปิด urgent mode เพื่อใช้ความจุสูงสุด

                print(f"[URGENT RDD] {item} SC {sc_so_no}: RDD passed (Week {calendar_week.iloc[rdd_idx]['WEEK']}) → start TODAY Week {order_week}")

            elif TODAY_IDX + 2 < len(calendar_week):

                order_week = calendar_week.iloc[TODAY_IDX + 2]["WEEK"]

            else:

                _skip_qty_short.append(_unplanned_row(
                    order, "ปฏิทินไม่มีข้อมูลถึง TODAY+2 สัปดาห์ (SC-ORDERS)", qty=qty_left
                ))

                continue



        elif order_type == "YD-ORDERS":

            yd_week = get_week_from_date(order["DYE_END_DATE"])

            if yd_week is not None:

                order_week = next_week(yd_week)  # +1 week หลังวันย้อมเสร็จ

            else:

                # ถ้าไม่มี DYE_END_DATE ให้ใช้ FG Week แทน (fallback)

                if pd.notna(fg_week) and fg_week is not None:

                    fg_week_str = str(int(fg_week)).strip()

                    if len(fg_week_str) >= 6:

                        year = int(fg_week_str[:4])

                        week = int(fg_week_str[4:])

                    elif len(fg_week_str) == 5:

                        year = int(fg_week_str[:4])

                        week = int(fg_week_str[4:])

                    else:

                        year = TODAY.year

                        week = int(fg_week_str)

                    idx = calendar_week[

                        (calendar_week["YEAR"] == year) & (calendar_week["WEEK"] == week)

                    ].index

                    if len(idx) > 0 and idx[0] >= 3:

                        order_week = int(calendar_week.iloc[idx[0] - 3]["WEEK"])

                    else:

                        order_week = None

                else:

                    order_week = None

        else:

            _skip_qty_short.append(_unplanned_row(
                order, f"ORDER_TYPE '{order_type}' ไม่รองรับ (routing ไม่ครอบคลุม)", qty=qty_left
            ))

            continue



        if order_week is None:

            _skip_qty_short.append(_unplanned_row(
                order, "คำนวณสัปดาห์เริ่มผลิตไม่ได้ (order_week=None)", qty=qty_left
            ))

            continue



        # ❗ ห้ามวางย้อนหลัง - เริ่มเร็วสุดที่ TODAY+2

        start_idx = TODAY_IDX + 2

        # รีเซ็ต anchor ของ week_index ทุก order ใหม่ (กันค่าค้างจาก order ก่อนหน้า
        # ที่แผนอาจเดินไปไกลข้ามปี) — จากนี้ anchor จะเลื่อนตาม plan_week เอง
        set_week_anchor(start_idx)

        plan_week = int(calendar_week.iloc[start_idx]["WEEK"])



        # ข้าม SKIP_WEEKS สำหรับ plan_week เริ่มต้น

        while plan_week in SKIP_WEEKS:

            plan_week = next_week(plan_week)

            if plan_week is None:

                break



        if plan_week is None:

            _skip_qty_short.append(_unplanned_row(
                order, "ไม่มีสัปดาห์ทำงานเหลือให้เริ่มผลิต (ปฏิทิน/SKIP_WEEKS)", qty=qty_left
            ))

            continue



        # YD-ORDERS: ห้ามวางก่อนย้อมเสร็จ — plan_week ต้องไม่ก่อน order_week (next_week หลัง DYE_END_DATE)
        if order_type == "YD-ORDERS" and order_week is not None:
            _yd_start_idx = week_index(order_week)
            if _yd_start_idx is not None and _yd_start_idx > start_idx:
                start_idx = _yd_start_idx
                plan_week = order_week



        # warm-carry flag: True เมื่อ plan_week ถูก "ดึงกลับ" มาเริ่มเร็วขึ้นเพื่อใช้เครื่องที่ยังอุ่น
        # (WARM GAP FILL / FG CONTINUATION) → backward/JIT ต้องไม่ shift ทับ เพราะ warm-carry ชนะ
        _warm_carry_applied = False

        # ❗ ถ้า booking ของ item+mc_group นี้ยังวิ่งถึง week ≥ plan_week → จัดการ 2 กรณี:

        # กรณี 1: plan_week == last booking week → plan ลงได้ใน week เดียวกัน แต่จำกัด qty ตาม remaining cap

        #          (remaining_week_cap ถูก seed ไว้แล้วจาก booking_remain_by_mc)
        #          ไม่ต้อง push plan_week ออก เพราะ cross-SC fill จะจัดการ qty เอง
        # กรณี 2: last booking week > plan_week → push ออกไปหลัง booking (ถ้ายังทัน RDD)
        _bk_mc_grp = str(order.get("MC GROUP", "")).strip().upper()
        if _bk_mc_grp:
            _bk_last_idx = last_production.get(_ck(item, _bk_mc_grp))
            if _bk_last_idx is not None and _bk_last_idx >= start_idx:
                _bk_last_week = int(calendar_week.iloc[_bk_last_idx]["WEEK"])
                _after_bk_idx = _bk_last_idx + 1
                if plan_week == _bk_last_week:
                    # plan_week == last booking week: ปล่อยให้ plan_week อยู่ตรงนั้น
                    # cross-SC fill จะใช้ remaining_week_cap ที่ seed ไว้ (idle capacity หลัง old item)
                    pass
                elif _after_bk_idx < len(calendar_week):
                    # plan_week < last booking week: push ออกไปหลัง booking ถ้าทัน RDD
                    if rdd_idx is None or _after_bk_idx <= rdd_idx:
                        start_idx = _after_bk_idx
                        plan_week = int(calendar_week.iloc[start_idx]["WEEK"])
                        # ข้าม SKIP_WEEKS หลัง booking check
                        while plan_week in SKIP_WEEKS:
                            plan_week = next_week(plan_week)
                            if plan_week is None:
                                break

                    # ถ้าขยับแล้วไม่ทัน RDD → ใช้ start_idx เดิม (ผลิตซ้อน booking ได้)
            elif _bk_last_idx is not None:
                # _bk_last_idx < start_idx: booking จบก่อน plan_week
                # ถ้าเครื่องยังอุ่น (gap ≤ SETUP_GAP_WEEK) และกำลังวิ่งอยู่ → pull plan_week กลับมา carry ต่อ
                _bk_warm_key = _ck(item, _bk_mc_grp)
                _bk_mc_running = machines_in_use.get(_bk_warm_key, 0)
                _bk_pull_plan_idx = week_index(plan_week)
                if _bk_pull_plan_idx is not None:
                    _bk_warm_gap = _bk_pull_plan_idx - _bk_last_idx
                    if (
                        1 < _bk_warm_gap <= SETUP_GAP_WEEK
                        and _bk_mc_running > 0
                    ):
                        # ❗ floor: ห้ามดึงกลับเร็วกว่า TODAY+2 (สัปดาห์ที่ผลิตจริงได้)
                        _bk_pull_idx = max(_bk_last_idx + 1, TODAY_IDX + 2)
                        if _bk_pull_idx < len(calendar_week) and _bk_pull_idx < _bk_pull_plan_idx:
                            _bk_pull_week = int(calendar_week.iloc[_bk_pull_idx]["WEEK"])
                            if _bk_pull_week >= TODAY_WEEK:
                                print(
                                    f"[WARM GAP FILL] {item}: machine last active W{int(calendar_week.iloc[_bk_last_idx]['WEEK'])} "
                                    f"→ pull plan_week W{plan_week}→W{_bk_pull_week} "
                                    f"(gap={_bk_warm_gap}, mc_running={_bk_mc_running}, no setup needed)"
                                )
                                start_idx = _bk_pull_idx
                                plan_week = _bk_pull_week
                                _warm_carry_applied = True  # warm-carry ชนะ backward

        # 🔧 FIX: Skip weeks where OLD bookings already exist for same item/MC/Gauge
        # to prevent exceeding capacity that old bookings already use
        _plan_idx = week_index(plan_week)
        if _plan_idx is not None and _bk_mc_grp:
            # Get gauge for this item/MC_GROUP
            _cap_row_check = item_cap_data[
                (item_cap_data["ITEM_CODE"] == item) &
                (item_cap_data["MC_GROUP"] == _bk_mc_grp)
            ]
            _plan_gauge = None
            if not _cap_row_check.empty:
                _plan_gauge = str(_cap_row_check.iloc[0].get("GUAGE", "")).strip()
            _plan_ck = _ck(item, _bk_mc_grp, _plan_gauge) if _plan_gauge else None

            # Check if OLD booking exists in plan_week and skip to next week
            if _plan_ck and _plan_ck in booking_mc_by_week and booking_mc_by_week[_plan_ck].get(_plan_idx, 0) > 0:
                print(f"[SKIP OLD BOOKING] {item}+{plan_week}+{_bk_mc_grp}: OLD booking already using machines, skip to next week")
                # Skip to next week (max 10 weeks to prevent infinite loop)
                _max_skip = 10
                _skipped = 0
                while (_plan_idx is not None and 
                       _plan_ck in booking_mc_by_week and 
                       booking_mc_by_week[_plan_ck].get(_plan_idx, 0) > 0 and
                       _skipped < _max_skip):
                    plan_week = next_week(plan_week)
                    if plan_week is None:
                        break
                    _plan_idx = week_index(plan_week)
                    _skipped += 1

                if plan_week is None:
                    continue
        # ❗ ถ้า SC/SO+Item เดิมเคยวาง FG_WEEK ก่อนหน้าแล้ว → ให้เริ่มจาก week ที่ FG แรกเสร็จ
        # เพื่อให้ FG ที่ 2 สามารถ carry over ใช้ capacity ที่เหลือจาก FG แรกได้
        # ยกเว้น: ถ้า item เดียวกัน มี cap เหลือจาก week นั้น → เริ่มใน week เดิมได้
        # ⚠️ แต่ถ้าต่อท้ายแล้วไม่ทัน RDD → อนุญาตให้ผลิตซ้อน FG ก่อนหน้าได้
        # 🔧 FIX: อนุญาตให้ FG ที่ 2 carry over ไปใช้ capacity ที่เหลือจาก FG แรก
        _prev_fg_idx = _last_fg_plan_idx.get((sc_so_no, item))
        if _prev_fg_idx is not None:
            _prev_fg_week = int(calendar_week.iloc[_prev_fg_idx]["WEEK"])
            # ตรวจว่ามี remaining cap ใน week นั้นสำหรับ item นี้หรือไม่
            # หมายเหตุ: remaining_week_cap จะถูกสร้างขึ้นหลังจาก FG แรกผลิตเสร็จ
            # ดังนั้น FG ที่ 2 สามารถใช้ capacity ที่เหลือได้ผ่าน cross-SC carryover logic (line 2802-2907)
            _item_mcs = set(
                str(r).strip().upper()
                for r in item_cap_data.loc[
                    item_cap_data["ITEM_CODE"] == str(item).strip().upper(), "MC_GROUP"
                ]
            )
            # ตรวจ remaining cap ใน ทุก week ตั้งแต่ start_idx จนถึง _prev_fg_idx
            # (ไม่ใช่แค่ week สุดท้ายของ FG แรก เพราะ FG แรกอาจทิ้ง cap ไว้ใน week กลางๆ)
            # 🔧 FIX: รวม _prev_fg_idx ในการตรวจด้วย แม้ _prev_fg_idx < start_idx
            # เพราะ target_start_idx = rdd_idx + 1 อาจข้าม week ที่ FG ก่อนหน้าจบ (มี remaining cap)
            _rem_check_start = min(start_idx, _prev_fg_idx)
            _has_rem_cap = any(
                remaining_week_cap.get((int(calendar_week.iloc[_wi]["WEEK"]), item, _mc), 0) > 0
                for _wi in range(_rem_check_start, _prev_fg_idx + 1)
                if _wi < len(calendar_week)
                for _mc in _item_mcs
            )
            if _has_rem_cap:
                # 🔧 FIX: เริ่ม FG ถัดไปจาก week สุดท้ายของ FG ก่อนหน้า (_prev_fg_idx)
                # ห้ามเริ่มจาก week กลางๆ ที่มี remaining cap จาก rev_weight rounding
                # เพราะ FG ก่อนหน้ายังผลิตไม่จบ → ต้องรอให้จบก่อนค่อย carry ต่อ
                _prev_fg_w = int(calendar_week.iloc[_prev_fg_idx]["WEEK"])
                _has_rem_at_last = any(
                    remaining_week_cap.get((_prev_fg_w, item, _mc), 0) > 0
                    for _mc in _item_mcs
                )
                if _has_rem_at_last:
                    # FG ก่อนหน้าจบใน week นี้และมี cap เหลือ → เริ่ม FG ถัดไปที่นี่
                    start_idx = _prev_fg_idx
                    plan_week = _prev_fg_w
                    _warm_carry_applied = True  # warm-carry ชนะ backward
                    print(
                        f"[FG CONTINUATION] {item}: FG ถัดไป start at W{_prev_fg_w} "
                        f"(remaining cap from previous FG last week)"
                    )
                else:
                    # FG ก่อนหน้าใช้ cap หมดใน week สุดท้าย → เริ่ม week ถัดไป
                    _after_prev = _prev_fg_idx + 1
                    if _after_prev < len(calendar_week):
                        start_idx = _after_prev
                        plan_week = int(calendar_week.iloc[_after_prev]["WEEK"])
                        _warm_carry_applied = True  # warm-carry ชนะ backward
                        print(
                            f"[FG CONTINUATION] {item}: FG ถัดไป start at W{plan_week} "
                            f"(next week after previous FG finished at W{_prev_fg_w})"
                        )
            else:
                if _prev_fg_idx >= start_idx:
                    # FG แรกเสร็จหลัง start_idx ของ FG ที่ 2
                    # สำหรับ SC-ORDERS ให้เริ่มจาก earliest feasible week เดิม
                    # ไม่บังคับรอ FG ก่อนหน้าจบทั้งก้อน เพื่อให้ทอได้ไวขึ้น (เช่นเริ่มที่ W19)
                    if order_type in ("SC-ORDERS", "ผ้าเติม"):
                        print(
                            f"[KEEP EARLIEST START] {item}: keep start week {int(calendar_week.iloc[start_idx]['WEEK'])} "
                            f"instead of pushing to previous FG week {_prev_fg_week}"
                        )
                    # non SC-ORDERS ยังใช้กฎเดิม
                    elif rdd_idx is not None and _prev_fg_idx > rdd_idx:
                        pass  # ไม่ push — ใช้ start_idx เดิม (อนุญาต overlap กับ FG แรก)
                    else:
                        start_idx = _prev_fg_idx
                        plan_week = _prev_fg_week
                # else: ใช้ start_idx เดิม (FG ที่ 2 เริ่มก่อน FG แรกเสร็จ → ผลิตซ้อนกันได้)


        # YD-ORDERS: re-clamp ห้ามวางก่อนวันย้อมเสร็จ (order_week = next_week หลัง DYE_END_DATE)
        # gate แรก (บรรทัด ~7826) อาจถูก override โดย pull-back ทีหลัง:
        # WARM GAP FILL / FG CONTINUATION / booking pull-back → ดึง plan_week กลับมาก่อน order_week
        # โดยไม่เช็ควันนัดย้อม ทำให้แผนวางก่อนย้อมเสร็จ จึงต้อง clamp ซ้ำครั้งสุดท้าย
        if order_type == "YD-ORDERS" and order_week is not None:
            _yd_floor_idx = week_index(order_week)
            _yd_cur_idx = week_index(plan_week)
            if _yd_floor_idx is not None and _yd_cur_idx is not None and _yd_cur_idx < _yd_floor_idx:
                start_idx = _yd_floor_idx
                plan_week = order_week

        # ----------------------
        # weekly allocation with best machine selection
        # ----------------------
        # คำนวณจำนวนเครื่องที่ต้องการตั้งแต่แรก (ถ้าทัน RDD)
        required_machines_info = None
        # S9 Logic: per-order flags
        _s9_active = False   # True เมื่อ S9 routing ถูก activate สำหรับ order นี้
        _s9_split_done = False  # True = order ถูก split แล้ว → normal loop ห้าม activate S9 ซ้ำ
        _s9_no_cap_weeks = set()  # weeks ที่ S9 ลองแล้วได้ 0 เครื่อง → ห้าม S9 reset กลับ week เหล่านี้
        _s9_week_locked = False  # True = plan_week ถูก set ไป TODAY+3 แล้ว → ครั้งถัดไปไม่ reset กลับ
        # คำนวณ setup days ล่วงหน้า (ใช้ใน calculate_required_machines ด้วย)
        order_material_content = str(order.get("MATERIAL_CONTENT", "")).strip() or _material_content_lookup.get(str(item).strip().upper(), "")
        _order_yarn_used = str(order.get("YARN-USED", "") or order.get("YARN_USED", "") or order.get("YARN_ITEM", "") or _yarn_used_lookup.get(str(item).strip().upper(), "")).strip()
        order_setup_days = get_setup_days_for_item(order_material_content, _order_yarn_used, item_code=item)
        # ❗ ตรวจสอบว่า item นี้มี cap data หรือไม่ — ถ้าไม่มีให้ข้ามทันที
        _item_cap_rows = item_cap_data[item_cap_data["ITEM_CODE"] == str(item).strip().upper()]
        if _item_cap_rows.empty:
            _skip_no_cap.append(
                _unplanned_row(order, "ไม่พบ CAP data (item_cap2025.xlsx)", qty=qty_left)
            )
            print(f"⚠️  ไม่พบ CAP data สำหรับ item '{item}' (SC/SO:{sc_so_no}) → ข้ามการวางแผน")
            continue


        # คำนวณ machine allocation ล่วงหน้า
        progressive_plan = None  # {week: machines} สำหรับแต่ละ week
        # สัปดาห์แรกที่ progressive_plan "ผลิตจริง" (mc > 0) = จุดเริ่ม backward/JIT
        # ห้ามใช้ min(progressive_plan.keys()) เพราะ dict มี key ของสัปดาห์ที่ mc=0 อยู่ด้วย
        # (จะได้ plan_week เดิมเสมอ) และ min() บน week number พังเมื่อแผนคร่อมปี
        _prog_start_week = None
        # Initialize core production schedule variable before use
        _core_production_schedule = []
        # 🔧 FIX: รวม Pending Plan ของทุก FG สำหรับ item เดียวกัน (ทุก SC/SO)
        # เพื่อให้คำนวณจำนวนเครื่องจาก total demand ไม่ใช่แค่ FG ปัจจุบัน
        # ป้องกันการลดเครื่องเมื่อ FG หนึ่งจบ ทั้งที่ยังมี FG ถัดไปรอผลิต
        _total_qty_all_fg = get_total_pending_qty_for_item(item, sc_so_no, fg_week, orders_sorted)
        _qty_for_machine_calc = max(qty_left, _total_qty_all_fg)
        _initial_qty_for_mc_calc = qty_left  # qty_left ตอนเริ่ม order — ใช้คำนวณ adjusted demand
        # 🔧 FIX: คำนวณ demand เฉพาะ SC ปัจจุบัน (ไม่รวม SC อื่นของ item เดียวกัน)
        # ใช้สำหรับ carry optimization เพื่อป้องกันการ hold เครื่องเกินจำเป็นจาก cross-SC demand
        _total_qty_same_sc = get_total_pending_qty_for_item(item, sc_so_no, fg_week, orders_sorted, same_sc_only=True)
        _qty_for_current_sc = max(qty_left, _total_qty_same_sc)
        if _total_qty_all_fg > qty_left:
            print(f"[TOTAL FG QTY] {item} SC {sc_so_no}: qty_left={qty_left:.0f}, total_all_fg={_total_qty_all_fg:.0f}, same_sc={_total_qty_same_sc:.0f} → ใช้ {_qty_for_machine_calc:.0f} สำหรับคำนวณเครื่อง")
        if (rdd_idx is not None and rdd_idx >= week_index(plan_week)) or _core_production_schedule:
            _locked_mc = locked_mc_group_for.get((sc_so_no, item))
            if _locked_mc:
                _locked_mc_setup = get_setup_days_for_item(order_material_content, _order_yarn_used, mc_group=_locked_mc, item_code=item)
                if _locked_mc_setup != order_setup_days:
                    order_setup_days = _locked_mc_setup
            print(f"[DEBUG START] Processing order: {sc_so_no}, item: {item}, qty: {order_qty}, FG week: {fg_week}")


            # คำนวณจำนวนเครื่องที่ต้องการสำหรับ item นี้ (เพื่อจบก่อน TARGET_KNIT/RDD)
            # ใช้ rdd_idx เป็น target เพราะต้องทอเสร็จก่อน RDD (FG Week - offset ตาม FOB_TYPE)
            target_knit_idx = rdd_idx
            # CORE ITEM: ไม่สน RDD — ใช้ plan_week+5 เป็น target
            if _core_production_schedule:
                target_knit_idx = week_index(plan_week) + 5
            mc_group_calc, daily_cap_calc, required_machines, feasible_calc, _gauge_calc = (
                calculate_required_machines(
                    item,
                    _qty_for_machine_calc,
                    plan_week,
                    target_knit_idx,  # TARGET_KNIT/RDD
                    setup_days=order_setup_days,
                    only_mc_group=_locked_mc,
                    order_type=order_type,
                    sub_color=sub_color,
                    dye_end_date=order.get("DYE_END_DATE"),
                )
            )
            print(f"[DEBUG CALC] {item}: mc_group={mc_group_calc}, required_machines={required_machines}, feasible={feasible_calc}")
            if required_machines:
                required_machines_info = (
                    mc_group_calc,
                    daily_cap_calc,
                    required_machines,
                    feasible_calc,
                    _gauge_calc,
                )
                # ล็อก MC_GROUP highest-cap นี้ไว้ให้ FG Week ถัดไปของ SC/SO+Item เดิมใช้ต่อ
                if mc_group_calc and (sc_so_no, item) not in locked_mc_group_for:
                    locked_mc_group_for[(sc_so_no, item)] = mc_group_calc
                # ถ้าเปิดใช้ progressive reduction → คำนวณเครื่องทุก week ล่วงหน้า
                if USE_PROGRESSIVE_REDUCTION and mc_group_calc and feasible_calc:
                    prog_result = calculate_progressive_reduction(
                        item,
                        _qty_for_machine_calc,
                        plan_week,
                        rdd_idx,
                        mc_group_calc,
                        daily_cap_calc,
                        _gauge_calc,
                        setup_days=order_setup_days,
                        rev_weight=get_revolution_weight_from_orders(item, mc_group_calc)
                    )
                    if prog_result:
                        progressive_plan = {wk: mc for wk, mc in prog_result}
                        # prog_result เรียงตาม week index อยู่แล้ว → tuple แรกที่ mc>0 คือ start จริง
                        _prog_start_week = next((wk for wk, mc in prog_result if mc > 0), None)

        _produced_week = None  # init สำหรับ track FG_WEEK sequential
        _earliest_backshift_done = False  # ป้องกันการย้อนกลับ EARLIEST_PLAN_WEEK ซ้ำใน order เดียวกัน
        # [FORCE START EARLIEST] ถูกลบออก → วางแผนตาม TARGET_KNIT (JIT) แทน
        # ============================================================
        # BACKWARD SCHEDULING (JIT START) — ใช้กับทุก order
        # ============================================================
        # หลักการ: เริ่มผลิต "ช้าที่สุดเท่าที่ยังทอเสร็จทัน RDD พอดี" (ไม่กันเผื่อ)
        #   start_idx = max(backward_idx, TODAY_IDX+2, yd_floor)
        # ลำดับความสำคัญของแหล่งที่มา backward_idx:
        #   1) progressive_plan (ถ้าคำนวณได้) → min(week) คือ optimal start ที่ผ่าน Step 1b มาแล้ว
        #   2) compute_backward_start_idx() → fallback สำหรับ order ที่ progressive_plan คำนวณไม่ได้
        # ข้อยกเว้น: ถ้า warm-carry (WARM GAP FILL / FG CONTINUATION) ดึง plan_week มาแล้ว
        #            → ข้าม backward ทั้งหมด เพราะเครื่องยังอุ่น ไม่ต้อง setup ใหม่ (คุ้มกว่า JIT)
        if _warm_carry_applied:
            if progressive_plan:
                print(f"[BACKWARD SKIP] {item}: warm-carry ใช้งานอยู่ → คง plan_week W{plan_week} (ไม่ shift JIT)")
        elif plan_week is not None:
            _bw_idx = None
            _bw_src = ""
            if progressive_plan and _prog_start_week is not None:
                _bw_idx = week_index(_prog_start_week)
                _bw_src = "progressive"
            elif required_machines_info and rdd_idx is not None:
                _bw_mc_grp, _bw_daily_cap, _bw_req_mc, _bw_feasible, _bw_gauge = required_machines_info
                _bw_idx = compute_backward_start_idx(
                    item,
                    _qty_for_machine_calc,
                    rdd_idx,
                    _bw_mc_grp,
                    _bw_daily_cap,
                    _bw_gauge,
                    _bw_req_mc,
                    setup_days=order_setup_days,
                    rev_weight=get_revolution_weight_from_orders(item, _bw_mc_grp),
                    floor_idx=TODAY_IDX + 2,
                )
                _bw_src = "backward"

            if _bw_idx is not None:
                # พื้นแข็ง: ห้ามถอยเร็วกว่า TODAY+2 และ (YD) ห้ามก่อนย้อมเสร็จ
                _bw_floor = TODAY_IDX + 2
                if order_type == "YD-ORDERS" and order_week is not None:
                    _bw_yd_idx = week_index(order_week)
                    if _bw_yd_idx is not None:
                        _bw_floor = max(_bw_floor, _bw_yd_idx)
                _bw_idx = max(_bw_idx, _bw_floor)

                _plan_cur_idx = week_index(plan_week)
                # เลื่อนได้ทางเดียว: ไปข้างหน้าเท่านั้น (ห้ามดึงกลับ — จะทับ push ของ booking/RDD gate)
                if _plan_cur_idx is not None and _bw_idx > _plan_cur_idx and _bw_idx < len(calendar_week):
                    _bw_week = int(calendar_week.iloc[_bw_idx]["WEEK"])
                    print(f"[JIT START/{_bw_src}] {item}: shift plan_week W{plan_week}→W{_bw_week} (TARGET_KNIT W{fg_week_int})")
                    plan_week = _bw_week
                    start_idx = _bw_idx
                    set_week_anchor(_bw_idx)  # backward กระโดดได้หลายสัปดาห์ → ย้าย anchor ตาม
        # Seed remaining_week_cap จาก booking สำหรับ carry-over ในสัปดาห์สุดท้ายของ old booking
        # ให้ new SO สามารถใช้ capacity ที่เหลืออยู่บน machine ของ old item ในสัปดาห์นั้นได้
        # สูตร: (MC_USE_CEIL - MC_USE) × working_days × daily_cap
        # = เครื่องที่ idle × วันทำงาน × cap ต่อวัน → qty ที่ new item สามารถผลิตได้ใน week เดียวกัน
        # --- Core Item trigger week simulation & batch schedule ---
        _core_trigger_week = None
        _core_production_schedule = []
        if is_core_item and item_upper in stock_inventory_lookup:
            _csd = stock_inventory_lookup[item_upper]
            _csm = _csd.get('STOCK_MIN', 0)
            _cs5 = _csd.get('Stock_5_Week', 0)
            if _csm > 0 and _cs5 > 0:
                _core_so_set = set()
                for _, _co in orders_sorted.iterrows():
                    if str(_co.get("Item Code","")).strip().upper() == item_upper:
                        if str(_co.get("Customer","")).strip() in CORE_CUSTOMERS:
                            _core_so_set.add("S" + str(_co.get("SC/SO NO","")).strip().lstrip("S"))
                            _core_so_set.add(str(_co.get("SC/SO NO","")).strip().lstrip("S"))
                _cinv = _csd.get('Inventory', 0)
                # ❗ ห้ามผลิต week ที่น้อยกว่า TODAY_WEEK + 2 ทุกกรณี
                _ccw = int(TODAY_WEEK)
                _ccw = next_week(_ccw)  # TODAY_WEEK + 1
                if _ccw is not None:
                    _ccw = next_week(_ccw)  # TODAY_WEEK + 2
                if _ccw is None:
                    _ccw = int(TODAY_WEEK) + 2  # Fallback
                _remaining_order = qty_left
                for _sim_i in range(104):
                    _ckp_rows = detail_mc[
                        (detail_mc["ITEM_CODE"].astype(str).str.upper().str.strip() == item_upper)
                        & (detail_mc["WEEK"].astype(int) == int(_ccw))
                    ]
                    for _, _kr in _ckp_rows.iterrows():
                        _so_str = str(_kr.get("SO_NO", ""))
                        _so_list = [s.strip() for s in _so_str.split(",") if s.strip()]
                        _total_so = max(len(_so_list), 1)
                        _match_so = sum(1 for s in _so_list if s in _core_so_set)
                        if _match_so > 0:
                            _cinv += float(_kr.get("KP_WEIGHT", 0)) * _match_so / _total_so
                    # ลด STOCK_MIN ทุก week (เริ่มจาก TODAY_WEEK + 2 แล้ว)
                    _cinv -= _csm
                    _cover = _cinv / _csm if _csm > 0 else 999
                    if _cover < 2 and _remaining_order > 0:
                        if _core_trigger_week is None:
                            _core_trigger_week = _ccw
                        _batch = min(_remaining_order, _cs5)
                        _core_production_schedule.append((_ccw, _batch))
                        _remaining_order -= _batch
                        _cinv += _batch
                        print(f"[CORE BATCH] {item}: W{_ccw} inv={_cinv-_batch:.0f} produce={_batch:.0f} remain_order={_remaining_order:.0f}")
                        if _remaining_order <= 0:
                            break
                    _ccw = next_week(_ccw)
                    if _ccw is None:
                        break
        if rts_local_force:
            _bkr_data = rts_local_force.get("booking_remain_by_mc", {})
            _last_old = rts_local_force.get("last_old_by_mc", {})
            for _bkr_mc, _bkr_cap in _bkr_data.items():
                _bkr_w = _last_old.get(_bkr_mc)
                if _bkr_w is not None and _bkr_cap > 0:
                    # ❗ ห้าม seed week เก่ากว่า TODAY_WEEK + 2 ทุกกรณี
                    if _bkr_w < int(TODAY_WEEK) + 2:
                        continue
                    # CORE ITEM: skip seed if booking week < trigger week
                    if _core_trigger_week is not None and _bkr_w < _core_trigger_week:
                        print(f"[CORE SKIP SEED] {item}: W{_bkr_w} < trigger W{_core_trigger_week}")
                        continue
                    # ห้าม seed ถ้า gap ระหว่าง booking week กับ plan_week เกิน SETUP_GAP_WEEK
                    if plan_week is not None:
                        _bkr_plan_idx = week_index(plan_week)
                        _bkr_seed_idx = week_index(_bkr_w)
                        if _bkr_plan_idx is not None and _bkr_seed_idx is not None and (_bkr_plan_idx - _bkr_seed_idx) > SETUP_GAP_WEEK:
                            print(f"[BOOKING CARRY SKIP] {item} booking W{_bkr_w} → plan W{plan_week}: gap {_bkr_plan_idx - _bkr_seed_idx} > {SETUP_GAP_WEEK} weeks → ไม่ seed")
                            continue
                    if (item, _bkr_w, _bkr_mc) not in _existing_item_week_mc:
                        _bkr_key = (_bkr_w, item, _bkr_mc)
                        if _bkr_key not in remaining_week_cap:
                            remaining_week_cap[_bkr_key] = _bkr_cap
                            remaining_week_cap_owner[_bkr_key] = None
                            print(f"[BOOKING CARRY] Seeded remaining_week_cap[({_bkr_w}, {item}, {_bkr_mc})] = {_bkr_cap:.2f} from booking")
        else:
            # Non-CORE items: seed จาก detail_mc โดยตรง (ไม่ผ่าน rts_local_force)
            if _bk_mc_grp:
                _dm_item = detail_mc[
                    detail_mc["ITEM_CODE"].astype(str).str.upper().str.strip() == item_upper
                ]
                if not _dm_item.empty:
                    for _dm_mc, _dm_grp in _dm_item.groupby("MC_GROUP"):
                        _dm_mc_str = str(_dm_mc).strip().upper()
                        # หา last active week
                        _dm_active = _dm_grp[
                            pd.to_numeric(_dm_grp["MC_USE_CEIL"], errors="coerce").fillna(0) > 0
                        ]
                        if _dm_active.empty:
                            continue
                        # สร้าง sorted week list สำหรับตรวจสอบ setup (เปรียบเทียบกับ week ก่อนหน้า)
                        _dm_sorted_weeks = sorted(_dm_grp["WEEK"].dropna().astype(int).tolist())
                        for _, _dm_last_row in _dm_active.iterrows():
                            _dm_last_w = int(_dm_last_row.get("WEEK", 0) or 0)
                            if _dm_last_w <= 0:
                                continue
                            _dm_mc_ceil = float(_dm_last_row.get("MC_USE_CEIL", 0) or 0)
                            _dm_kp_weight = float(_dm_last_row.get("KP_WEIGHT", 0) or 0)
                            # ใช้ค่าจากชีท Work Day ตรง ๆ (float: 5.5 วันต้องไม่ปัดเป็น 5)
                            _dm_wd = float(_dm_last_row.get("WORKING_DAY", 0) or 0)
                            _dm_gauge = str(_dm_last_row.get("GUAGE", "")).strip()
                            _dm_cap = _get_capacity_for_mc_group(item, _dm_mc_str, _dm_gauge)
                            # 🔧 FIX: ตรวจสอบว่า booking week นี้มีการ setup เพิ่มเครื่องหรือไม่
                            # โดยเปรียบเทียบ MC_USE_CEIL กับ week ก่อนหน้าใน _dm_grp
                            _dm_w_pos = _dm_sorted_weeks.index(_dm_last_w) if _dm_last_w in _dm_sorted_weeks else -1
                            _dm_prev_mc = 0
                            _dm_prev_w_val = None
                            if _dm_w_pos > 0:
                                _dm_prev_w = _dm_sorted_weeks[_dm_w_pos - 1]
                                _dm_prev_w_val = _dm_prev_w
                                _dm_prev_row = _dm_grp[_dm_grp["WEEK"] == _dm_prev_w]
                                if not _dm_prev_row.empty:
                                    _dm_prev_mc = float(_dm_prev_row.iloc[0].get("MC_USE_CEIL", 0) or 0)
                            # 🔧 FIX v2: ถ้ามี gap > 1 week ระหว่าง prev booking week และ current week
                            # → เครื่องทุกตัวต้อง setup ใหม่ (แม้ MC count จะเท่ากัน)
                            _dm_week_gap = (int(_dm_last_w) - int(_dm_prev_w_val)) if _dm_prev_w_val is not None else 999
                            if _dm_week_gap > 1:
                                _dm_new_mc = float(_dm_mc_ceil)  # all machines are new setup
                                _dm_carry_mc = 0.0
                            else:
                                _dm_new_mc = max(0.0, _dm_mc_ceil - _dm_prev_mc)
                                _dm_carry_mc = _dm_mc_ceil - _dm_new_mc
                            _dm_setup = order_setup_days  # setup days สำหรับ item นี้
                            # remaining = (MC_USE_CEIL - MC_USE) × working_days × daily_cap
                            # = เครื่องที่ idle × วันทำงาน × cap ต่อวัน → qty ที่ new item สามารถผลิตได้ใน week เดียวกัน
                            _dm_total_cap = (_dm_carry_mc * _dm_wd + _dm_new_mc * max(0, _dm_wd - _dm_setup)) * _dm_cap
                            if _dm_new_mc > 0 and _dm_cap > 0:
                                print(f"[BOOKING SETUP DETECT] {item} W{_dm_last_w} MC={_dm_mc_str}: new_mc={_dm_new_mc:.0f} carry_mc={_dm_carry_mc:.0f} → adjusted total_cap={_dm_total_cap:.2f} (naive={_dm_mc_ceil*_dm_wd*_dm_cap:.2f})")
                            if _dm_cap > 0 and _dm_mc_ceil > 0:
                                _dm_rem = max(0.0, _dm_total_cap - _dm_kp_weight)
                                if _dm_rem > 0 and (item, _dm_last_w, _dm_mc_str) not in _existing_item_week_mc:
                                    # ❗ ห้ามใช้ week เก่ากว่า TODAY_WEEK + 2 ทุกกรณี
                                    if _dm_last_w < int(TODAY_WEEK) + 2:
                                        continue
                                    if plan_week is not None:
                                        _dm_plan_idx = week_index(plan_week)
                                        _dm_seed_idx = week_index(_dm_last_w)
                                        if _dm_plan_idx is not None and _dm_seed_idx is not None and abs(_dm_plan_idx - _dm_seed_idx) > SETUP_GAP_WEEK:
                                            continue
                                    _dm_key = (_dm_last_w, item, _dm_mc_str)
                                    if _dm_key not in remaining_week_cap:
                                        remaining_week_cap[_dm_key] = _dm_rem
                                        remaining_week_cap_owner[_dm_key] = None
                                        print(f"[BOOKING CARRY] Seeded remaining_week_cap[({_dm_last_w}, {item}, {_dm_mc_str})] = {_dm_rem:.2f} from booking (non-CORE)")


        # CORE ITEM: batch production — cap qty_left to first batch size
        _core_real_qty = qty_left
        _core_batch_idx = 0
        if _core_production_schedule:
            _bw0, _bq0 = _core_production_schedule[0]
            qty_left = min(qty_left, _bq0)
            if plan_week is not None:
                _tw_idx = week_index(_bw0)
                _pw_idx = week_index(plan_week)
                if _tw_idx is not None and _pw_idx is not None and _tw_idx > _pw_idx:
                    plan_week = int(_bw0)
            print(f"[CORE BATCH INIT] {item}: {len(_core_production_schedule)} batches, B1 W{_bw0} qty={qty_left:.0f} (total={_core_real_qty:.0f}), plan_week=W{plan_week}")

        _mainloop_guard = 0
        while (qty_left > 0 or (_core_batch_idx + 1 < len(_core_production_schedule))) and plan_week is not None:
            # SAFETY: กัน infinite loop ตอนข้ามปลายปี (next_week คืน bare week แล้ว week_index
            # อาจ resolve ถอยปีทำให้ plan_week ไม่เดินหน้า) — เกินจำนวนสัปดาห์*fill ที่เป็นไปได้จริง → หยุด
            _mainloop_guard += 1
            if _mainloop_guard > 3000:
                print(f"[SAFETY YEAR-END] {item} SC {sc_so_no}: main placement loop เกิน 3000 รอบ → break (qty เหลือจะเข้า UNPLANNED)")
                break
            # FIX: กัน floating-point residual qty ที่น้อยเกินไปจาก booking deduction
            # qty ที่น้อยกว่า 1.0 ไม่สามารถผลิตได้จริง และจะทำลาย booking carryover seed ของ order อื่น
            if qty_left < 1.0:
                # CORE ITEM: batch continuation — check for next batch
                if _core_production_schedule:
                    _produced_this_batch = _core_production_schedule[_core_batch_idx][1] - max(0, qty_left)
                    _core_real_qty -= _produced_this_batch
                    _core_batch_idx += 1
                    if _core_batch_idx < len(_core_production_schedule) and _core_real_qty > 1.0:
                        _nbw, _nbq = _core_production_schedule[_core_batch_idx]
                        plan_week = int(_nbw)
                        qty_left = min(_core_real_qty, _nbq)
                        required_machines_info = None
                        print(f"[CORE NEXT BATCH] {item}: W{_nbw} batch={qty_left:.0f} total_remaining={_core_real_qty:.0f}")
                        continue
                    else:
                        qty_left = _core_real_qty
                break

            # ถ้า FG ใหม่ (SC/SO ใหม่) เริ่มใน week เดิมและมี cap เหลือ ให้ผลิตใน week เดิมจน cap หมดก่อนข้ามไป week ถัดไป
            _fill_last_week = None  # track week สุดท้ายที่ fill cross-SC
            _already_filled_this_sc = set()  # track (item, week, mc) ที่ SC นี้ fill ไปแล้ว (ป้องกัน duplicate ของ SC เดียวกัน)
            # ก้อนจ้างทอ (OUTSOURCE_FORCE) ห้ามเข้า loop นี้ — remaining cap เป็นของ "เครื่องในบ้าน"
            # ที่แถวปกติของ item เดียวกันเปิดไว้ ถ้าเติมตรงนี้ก้อนจ้างทอจะไปกินเครื่องในบ้าน
            while qty_left > 0 and ALLOW_SAME_ITEM_WEEK_CARRY and not _outsource_force:
                # ค้นหา remaining capacity สำหรับ ITEM เดียวกัน ในทุก week (เรียง week น้อยสุดก่อน)
                _found_rem_mc = None
                _found_rem_cap = 0
                _found_rem_week = None
                for _rk, _rv in sorted(remaining_week_cap.items(), key=lambda x: x[0][0]):
                    if _rk[1] == item and _rv > 0:
                        # 🔧 FIX: ไม่ดึง cap ของ "future week" มาใช้ก่อนเวลา
                        # ให้เริ่มผลิตตาม plan_week ปัจจุบันก่อน แล้วค่อยใช้ carry เมื่อถึง week นั้นจริง
                        if plan_week is not None and _rk[0] > plan_week:
                            continue
                        # ❗ ห้ามใช้ week เก่ากว่า TODAY_WEEK + 2 ทุกกรณี
                        if _rk[0] < int(TODAY_WEEK) + 2:
                            remaining_week_cap.pop(_rk, None)
                            remaining_week_cap_owner.pop(_rk, None)
                            continue
                        # CORE ITEM: skip fill if week < trigger week
                        if _core_trigger_week is not None and _rk[0] < _core_trigger_week:
                            continue
                        _found_rem_week = _rk[0]
                        _found_rem_mc = _rk[2]
                        _found_rem_cap = _rv
                        break

                if _found_rem_mc is None or _found_rem_cap <= 0:
                    break

                # ถ้า SC ปัจจุบันเคย fill item+week+mc นี้ไปแล้ว → หยุดเพื่อป้องกัน duplicate ของ SC เดียวกัน
                # ไม่ใช้ _existing_item_week_mc เพราะ set นั้นมี SC อื่น (เช่น SC ก่อนหน้า) อยู่ด้วย
                # ซึ่งจะ block SC ปัจจุบันไม่ให้ใช้ remaining cap ที่ SC ก่อนหน้าทิ้งไว้
                if (_found_rem_week, item, _found_rem_mc) in _already_filled_this_sc:
                    print(
                        f"[DEBUG DUPLICATE] SC {sc_so_no} already filled {item}+{_found_rem_week}+{_found_rem_mc}, stop"
                    )
                    break


                # ตั้งค่า mc_group และตัวแปรที่เกี่ยวข้องจาก remaining capacity ที่พบ
                _fill_mc_group = _found_rem_mc
                _fill_week = _found_rem_week
                _rem_cap_key = (_fill_week, item, _fill_mc_group)
                _cap_row = item_cap_data[
                    (item_cap_data["ITEM_CODE"] == item)
                    & (item_cap_data["MC_GROUP"] == _fill_mc_group)
                ]
                if _cap_row.empty:
                    remaining_week_cap.pop(_rem_cap_key, None)
                    remaining_week_cap_owner.pop(_rem_cap_key, None)
                    break

                # ใช้ capacity ตาม logic ใหม่: สำหรับ SHARED_POOL_MAP ให้เลือกจาก cap ที่น้อยที่สุดในกลุ่มเดียวกัน
                _fill_gauge = _cap_row.iloc[0].get("GUAGE")
                _fill_daily_cap = _get_capacity_for_mc_group(item, _fill_mc_group, _fill_gauge)
                if _fill_daily_cap <= 0:
                    # Fallback: use original capacity if new logic returns 0 (normalize 24→20hr)
                    _fill_daily_cap = _normalize_capacity(item, _fill_mc_group, float(_cap_row.iloc[0]["CAP ทอ"]))
                _fill_daily_cap = adjust_daily_cap_for_item_special(_fill_daily_cap, item, _fill_mc_group, _fill_gauge)
                _fill_rev_weight = get_revolution_weight_from_orders(item, _fill_mc_group)
                _fill_ck = _ck(item, _fill_mc_group, _fill_gauge)
                _fill_avail_mc = machines_in_use.get(_fill_ck, 1)
                # 🔧 FIX: ถ้า new plan มี carry machine จาก week ก่อนหน้าติดกัน (gap=1)
                # และ fill week นี้มี old booking machines → รวม cap ทั้งสองชุด
                # เช่น W23 setup 1 mc (new plan) + W24 old booking 3 mc → fill row W24 = 4 mc
                _fill_w_idx = week_index(_fill_week)
                _prev_lp_at_fill = last_production.get(_fill_ck)
                _new_carry_at_fill = machines_in_use.get(_fill_ck, 0)
                _bk_mc_at_fill_week = booking_mc_by_week.get(_fill_ck, {}).get(_fill_w_idx, 0)
                _is_new_plan_carry_fill = (
                    _fill_ck in new_plan_started_items
                    and _prev_lp_at_fill is not None
                    and _fill_w_idx is not None
                    and _fill_w_idx - _prev_lp_at_fill == 1
                    and _new_carry_at_fill > 0
                    and _bk_mc_at_fill_week > 0
                )
                if _is_new_plan_carry_fill:
                    _fill_avail_mc, _fill_own_mc, _fill_bk_prev, _bk_mc_at_fill_week = _combine_carry_with_booking(
                        _fill_ck, _new_carry_at_fill, _prev_lp_at_fill, _fill_w_idx
                    )
                    machines_in_use[_fill_ck] = _fill_avail_mc  # อัปเดตก่อน implied_mc cap
                    print(f"[CARRY+BK FILL] {item} W{_fill_week}: plan-own={_fill_own_mc} (carry {_new_carry_at_fill} − booking W ก่อน {_fill_bk_prev}) + booking={_bk_mc_at_fill_week} → combined={_fill_avail_mc} mc")
                _fill_actual_wd = get_working_days_by_factory(
                    _fill_mc_group, _fill_avail_mc, week=_fill_week, gauge=_fill_gauge
                )
                # 🔧 FIX: ถ้า remaining capacity มากกว่า 1 เครื่อง → คำนวณเครื่องจาก capacity
                # เพื่อรองรับกรณี freed capacity จากเครื่องที่ปล่อย (tail end)
                # ⚠️ ห้าม implied_mc เกินกว่าเครื่องที่ setup จริง (machines_in_use) เพราะ remaining cap
                # อาจมาจาก freed capacity ซึ่งยังนับเป็นเครื่องที่ active อยู่ ไม่ใช่เครื่องใหม่
                if _fill_daily_cap > 0 and _fill_actual_wd > 0:
                    _mc_one_cap = _fill_actual_wd * _fill_daily_cap
                    if _mc_one_cap > 0:
                        _implied_mc = int(_found_rem_cap / _mc_one_cap) + (1 if _found_rem_cap % _mc_one_cap > 0 else 0)
                        # cap ที่ machines_in_use จริง (ห้ามเกิน)
                        _actual_mc_in_use = machines_in_use.get(_fill_ck, _fill_avail_mc)
                        _implied_mc = min(_implied_mc, _actual_mc_in_use)
                        if _implied_mc > _fill_avail_mc:
                            _fill_avail_mc = _implied_mc
                _rem_cap = _found_rem_cap
                # เพิ่ม cap จาก new plan carry machine (full week capacity) เข้า _rem_cap
                # ⚠️ ใช้ _fill_own_mc (เครื่องของแผนเอง) ไม่ใช่ _new_carry_at_fill (รวมเครื่อง booking)
                # เพราะ capacity ที่เหลือของเครื่อง booking อยู่ใน _found_rem_cap อยู่แล้ว → บวกซ้ำไม่ได้
                if _is_new_plan_carry_fill and _fill_actual_wd > 0 and _fill_daily_cap > 0 and _fill_own_mc > 0:
                    _extra_carry_cap = _fill_own_mc * _fill_daily_cap * _fill_actual_wd
                    _rem_cap += _extra_carry_cap
                    print(f"[CARRY+BK FILL] {item} W{_fill_week}: +carry cap {_fill_own_mc}mc × {_fill_daily_cap:.2f} × {_fill_actual_wd}วัน = +{_extra_carry_cap:.2f} → total _rem_cap={_rem_cap:.2f}")
                # YD-ORDERS: ถ้า SUB_COLOR เปลี่ยน → หัก capacity ออก 1 วัน (สะสมในสัปดาห์เดียวกัน)
                if order_type == "YD-ORDERS" and sub_color:
                    _fill_prev_color = last_sub_color.get(_fill_ck, "")
                    if _fill_prev_color and _fill_prev_color != sub_color:
                        _wcs_fill_key = (_fill_ck, _fill_week)
                        _yd_week_color_setups[_wcs_fill_key] = _yd_week_color_setups.get(_wcs_fill_key, 0) + 1
                        _color_deduct = _fill_avail_mc * 1 * _fill_daily_cap  # 1 วัน × เครื่อง × cap (หักเต็มวัน ไม่ round ลง rev_weight)
                        _rem_cap = max(0, _rem_cap - _color_deduct)
                        print(f"[YD SUB_COLOR FILL] {item} W{_fill_week}: SUB_COLOR {_fill_prev_color}→{sub_color} หัก {_color_deduct:.0f} units (color setup #{_yd_week_color_setups[_wcs_fill_key]})")
                        # 🔧 FIX: ถ้าหัก color setup จน cap หมด → ลูปในจะไม่ผลิต ทำให้ last_sub_color/remaining_week_cap
                        # ไม่ถูกอัปเดต และ qty_left ไม่ลด → ลูปนอก while qty_left>0 วนซ้ำ key เดิมไม่รู้จบ (infinite loop)
                        # จึงต้องอัปเดต last_sub_color, ล้าง remaining cap ก้อนนี้ แล้ว break ออก
                        if _rem_cap <= 0:
                            last_sub_color[_fill_ck] = sub_color
                            remaining_week_cap.pop(_rem_cap_key, None)
                            remaining_week_cap_owner.pop(_rem_cap_key, None)
                            print(f"[YD SUB_COLOR FILL] {item} W{_fill_week}: color setup หัก cap จนหมด → ข้าม fill week นี้ (กัน infinite loop)")
                            break
                # เครื่องของ fill row นี้ใช้ร่วมกับแถวที่ครองเครื่องอยู่แล้วใน (item, mc, week) เดียวกันหรือไม่
                # ถ้าใช่ → แถวนี้ต้องรายงาน ACTUAL_MC = 0 มิฉะนั้นผลรวมเครื่องต่อสัปดาห์จะนับซ้ำ
                # (เช่น SO A 3 เครื่อง + SO B 3 เครื่อง = 6 ทั้งที่เป็นเครื่องชุดเดียวกัน 3 ตัว)
                _fill_shared_owner = next(
                    (
                        str(p.get("SC_SO_NO", "")).strip()
                        for p in plans
                        if p.get("ITEM_CODE") == item
                        and p.get("MC_GROUP") == _fill_mc_group
                        and p.get("PLAN_WEEK") == _fill_week
                        and int(p.get("ACTUAL_MC", 0) or 0) > 0
                    ),
                    None,
                )
                while qty_left > 0 and _rem_cap > 0:
                    if _fill_rev_weight and _fill_rev_weight > 0:
                        _fill_mc_cnt = max(1, int(_fill_avail_mc or 1))
                        produce = min(
                            qty_left,
                            fold_round(
                                _rem_cap, _fill_rev_weight, _fill_mc_group,
                                per_mc_cap=_rem_cap / _fill_mc_cnt, mc_count=_fill_mc_cnt,
                            ),
                        )
                    else:
                        produce = min(qty_left, _rem_cap)

                    if produce > 0:
                        plans.append({
                            "ITEM_CODE": item,
                            "SC_SO_NO": str(order.get("SO_NO", order.get("SC/SO NO", ""))).strip(),
                            "MC_GROUP": _fill_mc_group,
                            "MC_GUAGE": order["MC_GUAGE"],
                            "FACTORY_TYPE": FACTORY_TYPE_MAP.get(_fill_mc_group, "UNKNOWN"),
                            "PLAN_WEEK": _fill_week,
                            # เก็บ row index บน timeline ไว้ด้วย — PLAN_WEEK เป็น bare week
                            # ที่ระบุปีไม่ได้เมื่อแผนคร่อมปี (ใช้เติม PLAN_YEAR ตอน export)
                            "PLAN_IDX": week_index(_fill_week),
                            "PRODUCE_QTY": produce,
                            "SETUP_DAYS": 0,
                            "REQUIRED_MC": _fill_avail_mc,
                            "ACTUAL_MC": 0 if _fill_shared_owner else _fill_avail_mc,
                            "CARRYOVER_MC": 0 if _fill_shared_owner else _fill_avail_mc,
                            "NEW_MC": 0,
                            "MC_SHARED": _fill_avail_mc if _fill_shared_owner else 0,
                            "MC_BOOKING": _bk_mc_at_fill_week,  # เครื่องที่ booking ถักไอเทมนี้อยู่แล้ว
                            "REMARK": (
                                f"เติมเข้าเครื่องเดิม — ใช้เครื่อง {_fill_avail_mc} ตัวร่วมกับ SO {_fill_shared_owner} ({_fill_mc_group}) ไม่ใช้เครื่องเพิ่ม ไม่ต้อง setup"
                                if _fill_shared_owner
                                else f"เติมเข้าเครื่องที่วิ่งอยู่ {_fill_mc_group} {_fill_avail_mc} เครื่อง (ใช้ capacity ที่เหลือ) ไม่ต้อง setup"
                            ),
                            "FACTORY_WORKING_DAYS": get_working_days_by_factory(_fill_mc_group, _fill_avail_mc, week=_fill_week, gauge=_fill_gauge),
                            "WD_BASE": get_working_days_by_factory(_fill_mc_group, _fill_avail_mc, week=0, item_code=item, gauge=_fill_gauge),
                            "WD_W32": get_working_days_by_factory(_fill_mc_group, _fill_avail_mc, week=32, gauge=_fill_gauge),
                            "CALENDAR_WORKING_DAYS": len(get_working_days_in_week(_fill_week)),
                            "ACTUAL_WORKING_DAYS": _fill_actual_wd,
                            "DAILY_CAPACITY": _fill_daily_cap,
                            "REVOLUTION_WEIGHT": _fill_rev_weight if _fill_rev_weight is not None else 0,
                            "AVAILABLE_DAYS": _fill_actual_wd,
                            "ORDERS_QTY": order_qty,
                            "PENDING_PLAN": pending_plan,
                            "PLAN_QTY": qty_left - produce,
                            "ORDER_TYPE": order_type,
                            "ORDER_DATE": order["Date"],
                            "FG_WEEK": fg_week,
                            "TARGET_KNIT": fg_week_int,
                            "MATERIAL_CONTENT": str(order.get("MATERIAL_CONTENT", "")).strip(),
                            "IS_CORE_ITEM": "CORE ITEM" if is_core_item else "",
                            "CUSTOMER": str(order.get("Customer", "")).strip(),
                            "PLAN_SOURCE": "NEW",
                            "LT_YARN": order.get("DYE_END_DATE") if order_type == "YD-ORDERS" else get_yarn_lt_days(item),
                            "YARN_USED": _yarn_used_lookup.get(str(item).strip().upper(), ""),
                            "DATE_IN": order.get("DATE_IN"),
                            "EARLIEST_PLAN_WEEK": order.get("DYE_END_DATE") if order_type == "YD-ORDERS" else get_yarn_lt_earliest_week(item, date_in=order.get("DATE_IN")),
                            "SUB_COLOR": sub_color,
                            "PO_NO": str(order.get("PO_NO", "")).strip(),
                            "YARNPO": str(order.get("YARNPO", "") or "").strip(),
                            "RDD_WEEK": fg_week,
                            "SC_LINE_ID": str(order.get("SC_LINE_ID", "")).strip(),
                        })
                        plans[-1]["NAY_COLOR"] = str(order.get("NAY_COLOR", "")).strip()
                        plans[-1]["COLOR_DESC"] = str(order.get("COLOR_DESC", "")).strip()
                        plans[-1]["LINE_REMARK"] = str(order.get("LINE_REMARK", "")).strip()
                        # แถวแรกประกาศเครื่องไปแล้ว → แถวถัดไปใน week เดียวกันถือว่าใช้เครื่องร่วม
                        if not _fill_shared_owner:
                            _fill_shared_owner = str(order.get("SO_NO", order.get("SC/SO NO", ""))).strip()
                        qty_left -= produce
                        _rem_cap -= produce
                        # อัปเดต remaining cap หลังใช้งาน
                        _new_rem = max(0, _rem_cap)
                        if _new_rem > 0:
                            remaining_week_cap[_rem_cap_key] = _new_rem
                            # 🔧 FIX: ถ้า owner เดิมเป็น None (จาก booking seed) ให้คงไว้
                            # เพื่อไม่ให้ main loop คิดว่าเป็น new plan cap แล้ว override
                            if remaining_week_cap_owner.get(_rem_cap_key) is not None:
                                remaining_week_cap_owner[_rem_cap_key] = sc_so_no
                        else:
                            remaining_week_cap.pop(_rem_cap_key, None)
                            remaining_week_cap_owner.pop(_rem_cap_key, None)
                        # อัปเดต tracking สำหรับ cross-SC carryover
                        last_production[_fill_ck] = week_index(_fill_week)
                        machines_in_use[_fill_ck] = _fill_avail_mc
                        last_sc_so_no[_fill_ck] = sc_so_no
                        if order_type == "YD-ORDERS" and sub_color:
                            last_sub_color[_fill_ck] = sub_color
                        new_plan_started_items.add(_fill_ck)
                        # บันทึก cross-SC fill — ทั้งใน _existing_item_week_mc และ _already_filled_this_sc
                        _existing_item_week_mc.add((item, _fill_week, _fill_mc_group))
                        _already_filled_this_sc.add((_fill_week, item, _fill_mc_group))
                        _produced_week = _fill_week
                        _fill_last_week = _fill_week
                    else:
                        # cap เหลือน้อยเกินไป (rev_weight rounding) → mark week เต็ม (=0)
                        remaining_week_cap[_rem_cap_key] = 0
                        # 🔧 FIX: ถ้า owner เดิมเป็น None (จาก booking seed) ให้คงไว้
                        # เพื่อไม่ให้ main loop คิดว่าเป็น new plan cap แล้ว override ด้วย full capacity
                        if remaining_week_cap_owner.get(_rem_cap_key) is not None:
                            remaining_week_cap_owner[_rem_cap_key] = sc_so_no
                        break

                if qty_left <= 0:
                    break

            # หลัง item-carry fill: ถ้า fill เกิดที่ plan_week เดิม → ขยับ plan_week ไปสัปดาห์ถัดไป
            # เพื่อให้ main loop ไม่ setup เครื่องใหม่ทับในสัปดาห์ที่ item อื่นยังผลิตอยู่
            if qty_left > 0 and _fill_last_week is not None and _fill_last_week == plan_week:
                _next_after_fill = next_week(_fill_last_week)
                if _next_after_fill is not None:
                    print(f"[ITEM CARRY] {item}: advance plan_week W{plan_week} → W{_next_after_fill} (filled same week)")
                    plan_week = _next_after_fill
            # ถ้า fill เกิดที่สัปดาห์ก่อนหน้า plan_week → ย้อนกลับมา plan ต่อที่สัปดาห์นั้น (เดิม)
            elif qty_left > 0 and _fill_last_week is not None and not _earliest_backshift_done:
                _fl_idx = week_index(_fill_last_week)
                _pw_idx = week_index(plan_week)
                # ❗ floor: ห้าม backshift ย้อนเร็วกว่า TODAY+2 (สัปดาห์ที่ผลิตจริงได้)
                if _fl_idx is not None and _pw_idx is not None and _fl_idx < _pw_idx and _fl_idx >= TODAY_IDX + 2:
                    print(
                        f"[DEBUG BACKSHIFT] {item}: cross-SC filled W{_fill_last_week} while plan_week=W{plan_week} -> backshift to W{_fill_last_week}"
                    )
                    plan_week = _fill_last_week
                    _earliest_backshift_done = True

            # 🔧 REMOVED: Aggressive week-skipping when cross-SC fill marked week as full
            # This was preventing main planning loop from scheduling items in that week
            # with different mc_groups. Cross-SC fill should not block main planning.
            # if qty_left > 0:
            #     _items_in_week = [
            #         (_rk, _rv) for _rk, _rv in remaining_week_cap.items()
            #         if _rk[0] == plan_week and _rk[1] == item
            #     ]
            #     if _items_in_week and all(_rv == 0 for _, _rv in _items_in_week):
            #         # week full → skip to next week
            #         plan_week = next_week(plan_week)
            #         if plan_week is None:
            #             break
            #         continue

            # ⚠️ ตรวจสอบ RDD ก่อนว่าทันหรือไม่
            _current_order_rdd_idx = rdd_idx
            _plan_idx = week_index(plan_week)
            # เลื่อน anchor ตามตำแหน่งที่แผนเดินไปจริง — ทำให้ bare week ที่อ้างถึง
            # หลังจากนี้ (ทั้งในลูปและ helper) ถูกตีความเป็นปีที่ถูกต้องเมื่อแผนข้ามปี
            if _plan_idx is not None:
                set_week_anchor(_plan_idx)
            past_rdd = bool(
                rdd_idx is not None and _plan_idx is not None and _plan_idx >= rdd_idx
            )
            # CORE ITEM: ไม่สน RDD
            if _core_production_schedule:
                past_rdd = False
            if rdd_idx is not None and _plan_idx is not None and _plan_idx > rdd_idx:
                # urgent_mode = True  # DISABLED
                urgent_mode = False
            # ถ้ายังไม่ได้คำนวณ required_machines (เพราะตอนแรก avail=0 ทุก week)
            # ให้ลองคำนวณใหม่ด้วย plan_week ปัจจุบันที่มีเครื่องว่างจริง
            # 🔧 แก้ไข: คำนวณ required machines เสมอแม้ past RDD เพื่อให้มีข้อมูลสำหรับ planning
            print(f"[DEBUG CALC CHECK] required_machines_info={required_machines_info}, past_rdd={past_rdd}, rdd_idx={rdd_idx}, _plan_idx={_plan_idx}")
            if rdd_idx is not None and _plan_idx is not None:
                print(f"[DEBUG CALC CHECK] _plan_idx <= rdd_idx: {_plan_idx <= rdd_idx}")
            if (
                (rdd_idx is not None or _core_production_schedule)
                and _plan_idx is not None
            ):
                _locked_mc2 = locked_mc_group_for.get((sc_so_no, item))
                # CORE ITEM: ไม่สน RDD — ใช้ plan_week+5
                if _core_production_schedule:
                    target_knit_idx2 = _plan_idx + 5
                else:
                    target_knit_idx2 = rdd_idx
                _already_planned_item_qty = _item_cumulative_planned.get(item, 0)
                _dynamic_qty_for_calc = max(qty_left, _qty_for_machine_calc - _already_planned_item_qty)
                _mc_r, _cap_r, _req_r, _feas_r, _gauge_r = calculate_required_machines(
                    item,
                    _dynamic_qty_for_calc,
                    plan_week,
                    target_knit_idx2,  # TARGET_KNIT
                    setup_days=order_setup_days,
                    only_mc_group=_locked_mc2,
                    order_type=order_type,
                    sub_color=sub_color,
                    dye_end_date=order.get("DYE_END_DATE"),
                )
                if _req_r:
                    required_machines_info = (_mc_r, _cap_r, _req_r, _feas_r, _gauge_r)
                    if _mc_r and (sc_so_no, item) not in locked_mc_group_for:
                        locked_mc_group_for[(sc_so_no, item)] = _mc_r
                # S9 Logic: reset flag ต่อ iteration แล้วตรวจสอบใหม่
                _s9_active = False
                # S9 trigger ทุก iteration — เฉพาะ item ใน "S9 Only" เท่านั้น
                # (เลิกใช้ S9 Eligible/fallback แล้ว: งานที่เครื่องปกติไม่ทันจะไม่ถูกโยนไป S9 อีก)
                if required_machines_info:
                    _item_u_s9 = str(item).strip().upper()
                    # trigger: item ใน "S9 Only" (MasterMC บังคับ) หรือก้อนที่ user สั่งจ้างทอ
                    _s9_trigger = _item_u_s9 in _s9_only_items or _outsource_force
                    if _s9_trigger and _gauge_r is not None and _mc_r is not None:
                        _s9_mc_cat_t = _mc_to_type1(_mc_r, _gauge_r)
                        _s9r = _calc_s9_required_machines(
                            _dynamic_qty_for_calc, plan_week, target_knit_idx2,
                            _gauge_r, _s9_mc_cat_t, order_setup_days, order_material_content
                        )
                        if _s9r[0] is not None:
                            _is_s9_only = _item_u_s9 in _s9_only_items
                            _s9_mc_g, _s9_cap_g, _s9_req_g, _s9_feas_g, _s9_gauge_g = _s9r
                            required_machines_info = (_s9_mc_g, _s9_cap_g, _s9_req_g, _s9_feas_g, _s9_gauge_g)
                            locked_mc_group_for[(sc_so_no, item)] = _s9_mc_g
                            _s9_active = True
                            # Week setting:
                            # ครั้งแรก → set plan_week = TODAY+3 เสมอ (แม้ plan_week > TODAY+3)
                            # ครั้งถัดไป → ไม่ reset กลับ (advance ตาม loop ปกติ)
                            _s9_min_idx = TODAY_IDX + 3
                            # ก้อนจ้างทอ: user เป็นคนกำหนดสัปดาห์เอง → เริ่มที่สัปดาห์ที่กรอก
                            # (ถ้า S9 pool สัปดาห์นั้นเครื่องไม่พอ จะไหลต่อสัปดาห์ถัดไปตาม logic เดิม)
                            if _outsource_force and _outsource_week is not None:
                                _os_idx = week_index(int(_outsource_week))
                                if _os_idx is not None:
                                    _s9_min_idx = _os_idx
                            if _s9_min_idx < len(calendar_week):
                                _s9_target_week = int(calendar_week.iloc[_s9_min_idx]["WEEK"])
                                if not _s9_week_locked:
                                    if _s9_target_week not in _s9_no_cap_weeks:
                                        plan_week = _s9_target_week
                                        print(f"[S9] {item}: plan_week set to W{plan_week} (S9 first → today+3w)")
                                    _s9_week_locked = True
                                else:
                                    _s9_pw_idx = week_index(plan_week) if plan_week else None
                                    _s9_tgt_idx = week_index(_s9_target_week) if _s9_target_week else None
                                    if _s9_pw_idx is not None and _s9_tgt_idx is not None and _s9_pw_idx < _s9_tgt_idx:
                                        if _s9_target_week not in _s9_no_cap_weeks:
                                            plan_week = _s9_target_week
                                            print(f"[S9] {item}: plan_week set to W{plan_week} (S9 advance to today+3w)")
                            _s9_tag = "S9_ONLY" if _is_s9_only else ("ELIGIBLE+LATE" if past_rdd else "ELIGIBLE")
                            print(f"[S9] {item}: S9 activated ({_s9_tag}) → MC={_s9_mc_g}, cap={_s9_cap_g}, req={_s9_req_g}")
            # เลือกเครื่องที่เหมาะสมที่สุดสำหรับ item นี้
            # ถ้าเป็นกรณี RTS+LOCAL ให้บังคับใช้ MC เดิมและเริ่มหลัง old สุดท้าย
            mc_group = daily_capacity = setup_needed = available_machines = _sel_gauge = (
                None
            )
            # RTS: บังคับ MC_GROUP ตาม booking เดิม (detail_mc)
            # เพราะ get_best_machine_for_item อาจเลือก MC_GROUP อื่น (เช่น SKPTA→SKP)
            if rts_local_force:
                print(f"[DEBUG RTS] Processing RTS for {item} in week {plan_week}")
                _rts_old_mc = rts_local_force.get("last_old_by_mc", {})
                if _rts_old_mc:
                    # เลือก MC_GROUP ที่มี last week ล่าสุด
                    _rts_mc = max(_rts_old_mc, key=_rts_old_mc.get)
                    _rts_machines = rts_local_force.get("machines_by_mc", {}).get(
                        _rts_mc, 0
                    )
                    _rts_cap = rts_local_force.get("daily_cap_by_mc", {}).get(_rts_mc)
                    print(f"[DEBUG RTS] RTS MC: {_rts_mc}, machines: {_rts_machines}")
                    if _rts_machines > 0:
                        mc_group = _rts_mc
                        available_machines = _rts_machines
                        # 🔧 FIX: RTS ต้องเช็ค gap ก่อน set setup_needed=False
                        # ถ้าห่างจาก last_production > SETUP_GAP_WEEK → ต้อง setup ใหม่
                        _rts_ck = _ck(item, _rts_mc, _item_mc_to_gauge.get(
                            (str(item).strip().upper(), str(_rts_mc).strip().upper()), None
                        ))
                        _rts_last_widx = last_production.get(_rts_ck)
                        _rts_cur_widx = week_index(plan_week)
                        _rts_gap = (
                            (_rts_cur_widx - _rts_last_widx)
                            if _rts_last_widx is not None and _rts_cur_widx is not None
                            else None
                        )
                        if _rts_gap is not None and _rts_gap <= SETUP_GAP_WEEK:
                            setup_needed = False  # gap เล็ก → carry ได้
                        else:
                            setup_needed = True  # gap ใหญ่หรือไม่มี last_production → ต้อง setup
                        _sel_gauge = _item_mc_to_gauge.get(
                            (str(item).strip().upper(), str(_rts_mc).strip().upper()), None
                        )
                        # daily_capacity: ใช้ capacity ตาม logic ใหม่: สำหรับ SHARED_POOL_MAP ให้เลือกจาก cap ที่น้อยที่สุดในกลุ่มเดียวกัน
                        daily_capacity = _get_capacity_for_mc_group(item, _rts_mc, _sel_gauge)
                        if daily_capacity <= 0:
                            # Fallback: use original logic if new logic returns 0
                            _all_cap_rts = item_cap_data[item_cap_data["ITEM_CODE"] == item]
                            if not _all_cap_rts.empty:
                                daily_capacity = _normalize_capacity(item, _rts_mc, float(_all_cap_rts["CAP ทอ"].min()))
                            elif _rts_cap and not pd.isna(_rts_cap) and float(_rts_cap) > 0:
                                daily_capacity = _normalize_capacity(item, _rts_mc, float(_rts_cap))
                            else:
                                _rts_cap_row = item_cap_data[
                                    (item_cap_data["ITEM_CODE"] == item)
                                    & (item_cap_data["MC_GROUP"] == _rts_mc)
                                ]
                                if not _rts_cap_row.empty:
                                    daily_capacity = _normalize_capacity(
                                        item, _rts_mc,
                                        float(_rts_cap_row.iloc[0].get("CAP ทอ", 0) or 0)
                                    )
                        daily_capacity = adjust_daily_cap_for_item_special(daily_capacity, item, _rts_mc, _sel_gauge)
                        #  FIX: Recalculate required_machines_info for current week in RTS logic
                        #  This prevents using cached values from previous weeks (e.g., week 26 values in week 27)
                        _locked_mc2 = locked_mc_group_for.get((sc_so_no, item))
                        target_knit_idx2 = rdd_idx if not _core_production_schedule else (week_index(plan_week) + 5)
                        _already_planned_item_qty = _item_cumulative_planned.get(item, 0)
                        _dynamic_qty_for_calc = max(qty_left, _qty_for_machine_calc - _already_planned_item_qty)
                        _mc_r, _cap_r, _req_r, _feas_r, _gauge_r = calculate_required_machines(
                            item,
                            _dynamic_qty_for_calc,
                            plan_week,
                            target_knit_idx2,
                            setup_days=order_setup_days,
                            only_mc_group=_locked_mc2,
                            order_type=order_type,
                            sub_color=sub_color,
                            dye_end_date=order.get("DYE_END_DATE"),
                        )
                        if _req_r:
                            required_machines_info = (_mc_r, _cap_r, _req_r, _feas_r, _gauge_r)
                            print(f"[DEBUG RTS RECALC] Week {plan_week}: updated required_machines_info=({_mc_r}, {_cap_r}, {_req_r}, {_feas_r}, {_gauge_r})")
            # คำนวณ _req_feasible ก่อนเพื่อส่งเข้า get_best_machine_for_item
            _req_feasible = (
                required_machines_info[3]
                if required_machines_info and len(required_machines_info) > 3
                else True
            )
            # 🔧 แก้ไข: ให้ Load Balancing ทำงานทุกครั้ง ไม่ใช่แค่ครั้งแรก
            # ถ้ามี required_machines_info ให้ใช้ Load Balancing ทุกครั้งเพื่อจัดการ Gradual Increase
            print(f"[DEBUG LB CHECK] {item} week {plan_week}: required_machines_info={required_machines_info}, USE_LOAD_BALANCING={USE_LOAD_BALANCING}")
            print(f"[DEBUG LB CHECK] {item} week {plan_week}: mc_group={mc_group}, rts_local_force={bool(rts_local_force)}")

            # ⚠️ ตรวจสอบว่าเป็น carryover หรือไม่ก่อนเรียก Load Balancing
            # ถ้าเป็น carryover → ใช้เครื่องเดิมโดยไม่ต้องเช็ค available_machines
            # ต้องหา mc_group จาก last_production ก่อน เพราะ mc_group อาจยัง None
            # 🔧 FIX: validate ว่า _prev_mc_group ต้องตรงกับ MC_GROUP จาก order หรือมีใน item_cap_data
            #   เพื่อป้องกัน carryover จาก MC_GROUP ผิดที่ค้างอยู่ใน old plan (เช่น SKP แทน SKPTA)
            _valid_mc_groups_for_item = set(
                str(r).strip().upper()
                for r in item_cap_data.loc[
                    item_cap_data["ITEM_CODE"] == item, "MC_GROUP"
                ]
            )
            _prev_mc_key = None
            _prev_mc_group = None
            for _key, _week_idx in last_production.items():
                if _key[0] == item and _week_idx is not None:
                    # 🔧 FIX: ข้าม MC_GROUP ที่ไม่ตรงกับ order และไม่มีใน item_cap_data
                    _candidate_mc = _key[1]
                    if _candidate_mc != _ord_mc_grp and _candidate_mc not in _valid_mc_groups_for_item:
                        continue
                    current_week_idx = week_index(plan_week)
                    _gap_week = (
                        current_week_idx - _week_idx
                        if current_week_idx is not None
                        else None
                    )
                    # ต้องอยู่ภายใน SETUP_GAP_WEEK ของ plan_week
                    # AND ต้องไม่เก่ากว่า TODAY_WEEK − SETUP_GAP_WEEK (ไม่ carry จาก booking เก่าเกินไป)
                    # AND ต้องเป็นสัปดาห์ก่อนหน้าเท่านั้น (ห้ามดึงสัปดาห์อนาคตมาเป็น carry)
                    if (
                        current_week_idx is not None
                        and _gap_week is not None
                        and _gap_week >= 0
                        and _gap_week <= SETUP_GAP_WEEK
                        and _week_idx >= TODAY_IDX
                    ):
                        _prev_mc_key = _key
                        _prev_mc_group = _key[1]
                        break
                    # 🔧 NEW: ถ้า last_production เป็นสัปดาห์อนาคต (gap < 0) แต่ booking active
                    # ครอบคลุม plan_week ด้วย (booking หลาย week ต่อเนื่อง) → อนุญาต carry
                    # เพราะเครื่องกำลังวิ่งอยู่ใน plan_week แล้วจาก booking เดิม
                    if (
                        current_week_idx is not None
                        and _gap_week is not None
                        and _gap_week < 0
                        and _key in booking_active_week_set
                        and current_week_idx in booking_active_week_set[_key]
                    ):
                        _prev_mc_key = _key
                        _prev_mc_group = _key[1]
                        print(f"[CARRY BOOKING SPAN] {item} week {plan_week}: booking active ครอบ plan_week (last_production=W{int(calendar_week.iloc[_week_idx]['WEEK'])}, booking weeks={sorted(booking_active_week_set[_key])}) → อนุญาต carry")
                        break

            _is_carryover = _prev_mc_key is not None and _prev_mc_group is not None
            # CORE ITEM: always fresh setup — no carry history
            if _core_production_schedule:
                _is_carryover = False

            # S9 จ้างทอ: ข้ามการเลือกเครื่องปกติทั้งหมด → ใช้ข้อมูล S9 pool โดยตรง
            if _s9_active and required_machines_info and required_machines_info[0] and required_machines_info[2] > 0:
                mc_group = required_machines_info[0]
                daily_capacity = required_machines_info[1]
                available_machines = int(required_machines_info[2])
                _sel_gauge = required_machines_info[4]
                setup_needed = False
                _is_carryover = False  # S9 จ้างทอ: เครื่องแยกจากแผนปกติ ไม่ carry
                new_mc = available_machines  # S9: ใช้เต็ม pool ที่เหลือ
                carryover_mc = 0
                print(f"[S9 MACHINE] {item} W{plan_week}: MC={mc_group}, cap={daily_capacity}, avail={available_machines}, gauge={_sel_gauge}")

            # ก้อนจ้างทอ: ถ้าสัปดาห์นี้ S9 pool ไม่ว่าง (S9 ไม่ activate) → เลื่อนไปสัปดาห์ถัดไป
            # ห้าม fallback ไปเลือกเครื่องในบ้านเด็ดขาด (user แบ่งก้อนนี้ออกไปทอข้างนอกแล้ว)
            if _outsource_force and not _s9_active:
                print(f"[OUTSOURCE] {item} W{plan_week}: เครื่องจ้างทอ (S9) ไม่ว่าง → เลื่อนไปสัปดาห์ถัดไป")
                plan_week = next_week(plan_week)
                if plan_week is None:
                    break
                continue

            # ถ้าเป็น carryover แต่มีเครื่องว่างมากและ order ยังเหลือเยอะ → ให้ Load Balancing ทำงาน
            # ถ้าเป็น carryover ปกติ → ใช้เครื่องเดิมต่อไป
            # ตรวจสอบว่ามีเครื่องว่างพอที่จะเพิ่มหรือไม่
            _has_capacity_for_increase = False
            if _is_carryover and _prev_mc_group:
                # หา gauge จาก last_production key
                _carry_gauge = _prev_mc_key[2] if len(_prev_mc_key) > 2 else None
                _actual_remain = get_actual_mc_remain(_prev_mc_group, plan_week, gauge=_carry_gauge, item_code=item)
                _current_machines = machines_in_use.get(_prev_mc_key, 1)
                _has_capacity_for_increase = _actual_remain > 0
                if item == "FD6PRTPG99A0" or item == "FD4DRTIT49/08A0" or item == "FD3GNTPE54/14A0":
                    print(f"[DEBUG CAPACITY] {_prev_mc_group} week {plan_week}: gauge={_carry_gauge}, actual_remain={_actual_remain}, current_machines={_current_machines}, can_increase={_has_capacity_for_increase}")
            # ตรวจสอบว่า order เหลือเยอะพอที่จะเพิ่มเครื่องหรือไม่
            _needs_increase = False
            _total_item_demand = max(qty_left, _qty_for_machine_calc - (order_qty - qty_left))

            if _is_carryover and qty_left > 0:
                # ประมาณว่าถ้าเหลือมากกว่า 2 สัปดาห์ อาจต้องการเพิ่มเครื่อง
                _cap_row = item_cap_data[
                    (item_cap_data["ITEM_CODE"] == item)
                    & (item_cap_data["MC_GROUP"] == _prev_mc_group)
                ]
                if not _cap_row.empty:
                    _ni_gauge = _cap_row.iloc[0].get("GUAGE")
                    _daily_cap = _get_capacity_for_mc_group(item, _prev_mc_group, _ni_gauge)
                    if _daily_cap <= 0:
                        _daily_cap = _normalize_capacity(item, _prev_mc_group, float(_cap_row.iloc[0]["CAP ทอ"]))
                    if _daily_cap > 0:
                        # 🔧 FIX: หัก production ที่ผลิตไปแล้วออกจาก _qty_for_machine_calc
                        # รวมทั้ง SC อื่นของ item เดียวกันที่วางแผนเสร็จแล้ว (_item_cumulative_planned)
                        # เพื่อไม่ให้นับ demand ซ้ำจาก SC ที่ planned ไปแล้ว
                        _already_produced_this_sc = order_qty - qty_left
                        _adjusted_machine_calc = max(0, _qty_for_machine_calc - _already_produced_this_sc)
                        _total_item_demand = max(qty_left, _adjusted_machine_calc)
                        # ใช้ effective machines = min(เครื่องเดิม, cap ที่เหลือ) ในการคำนวณว่าทันไหม
                        _effective_machines = min(_current_machines, _actual_remain) if _actual_remain > 0 else 0
                        if _effective_machines > 0:
                            _weeks_remaining = _total_item_demand / (_effective_machines * _daily_cap * 5)
                            _needs_increase = _weeks_remaining > 2.0  # เหลือมากกว่า 2 สัปดาห์
                        else:
                            _needs_increase = True  # ไม่มี cap เหลือ → ยังไม่ทันแน่นอน
            # 🔧 แก้ไข: แยก 3 กรณี
            # 1. Carryover ที่มีเครื่องว่างและต้องการเพิ่ม → ส่งไป Load Balancing
            # 2. Carryover ที่ไม่มีเครื่องว่างหรือไม่ต้องการเพิ่ม → ใช้เครื่องเดิมต่อไป
            # 3. ไม่ใช่ carryover → ส่งไป Load Balancing ปกติ

            should_check_increase = _is_carryover and _has_capacity_for_increase and _needs_increase

            # [SETUP EARLY PRIORITY] และ [SETUP EARLY] ถูกลบออก → วางแผนตาม TARGET_KNIT (JIT) แทน
            # 🔧 สำคัญ: ถ้าเป็น carryover ให้จัดการก่อน mc_group check
            # กรณีที่ 1 & 2: Carryover ทุกกรณีที่ไม่ต้องการเพิ่มเครื่อง (ไม่มีเครื่องว่างหรือไม่จำเป็น)
            if _is_carryover and not should_check_increase:
                # Carryover แบบไม่ต้องเพิ่มเครื่อง: ใช้เครื่องเดิม แต่ต้องเช็ค actual_remain
                # ถ้า old plan จองเครื่องไปหมดแล้ว (คนละ SC) → ไม่ควร carry เพิ่ม
                mc_group = _prev_mc_group
                # ใช้ค่าจริงจาก plan ก่อนหน้าสำหรับ carryover
                _requested_machines = machines_in_use.get(_prev_mc_key, 1)
                # required_machines_info ใช้เฉพาะสำหรับ target gradual increase ไม่ใช้ค่า carryover จริง
                if required_machines_info and len(required_machines_info) > 0:
                    _required_mc = required_machines_info[2]
                    if _required_mc > 3:
                        print(f"[DEBUG CARRY] Actual carry from previous week: {_requested_machines} machines (target for gradual increase: {_required_mc})")
                _carry_gauge = _prev_mc_key[2] if len(_prev_mc_key) > 2 else None
                _actual_remain = get_actual_mc_remain(_prev_mc_group, plan_week, gauge=_carry_gauge, item_code=item)
                # Same SC ใช้เครื่องเดิมได้เต็มจำนวน (เครื่องเป็นของ SC นั้น) แต่ต้องเช็ค actual_remain ด้วย
                # เพราะ actual_remain คือ cap จริงที่เหลือหลังหัก old booking + new plan อื่นๆ
                _carry_key_for_sc = _resolve_carry_key(item, _prev_mc_group, _carry_gauge)
                _last_sc = last_sc_so_no.get(_carry_key_for_sc)
                _same_sc_carry = (_last_sc == sc_so_no)
                if _requested_machines > _actual_remain:
                    print(
                        f"[CARRY CLAMP] {item} SC {sc_so_no} W{plan_week}: "
                        f"requested {_requested_machines} > actual_remain {_actual_remain} "
                        f"(same_sc={_same_sc_carry}, last SC {_last_sc}) → clamp to {max(0, _actual_remain)}"
                    )
                    _requested_machines = max(0, int(_actual_remain))
                available_machines = _requested_machines
                setup_needed = False
                # ใช้ daily_capacity เดิม
                _cap_row = item_cap_data[
                    (item_cap_data["ITEM_CODE"] == item)
                    & (item_cap_data["MC_GROUP"] == mc_group)
                ]
                if not _cap_row.empty:
                    _sel_gauge = _cap_row.iloc[0].get("GUAGE")
                    daily_capacity = _get_capacity_for_mc_group(item, mc_group, _sel_gauge)
                    if daily_capacity <= 0:
                        daily_capacity = _normalize_capacity(item, mc_group, float(_cap_row.iloc[0]["CAP ทอ"]))
                else:
                    # Fallback: หา capacity จาก item ทั่วไป
                    _all_cap = item_cap_data[item_cap_data["ITEM_CODE"] == item]
                    if not _all_cap.empty:
                        _sel_gauge = _all_cap.iloc[0].get("GUAGE")
                        daily_capacity = _get_capacity_for_mc_group(item, mc_group, _sel_gauge)
                        if daily_capacity is None or daily_capacity <= 0:
                            daily_capacity = _normalize_capacity(item, mc_group, float(_all_cap["CAP ทอ"].min()))
                    else:
                        # ไม่มี capacity เลย → ข้ามไป
                        print(f"[WARNING] No capacity found for {item} in {mc_group}, skipping to next week")
                        plan_week = next_week(plan_week)
                        continue
                daily_capacity = adjust_daily_cap_for_item_special(daily_capacity, item, mc_group, _sel_gauge)

                if item == "FD1BASFZ15/1A0" or item == "FD3GNTPE54/14A0":
                    print(f"[DEBUG CARRY] Week {plan_week}: Using carryover - mc_group={mc_group}, machines={available_machines}, daily_cap={daily_capacity}, actual_remain={_actual_remain}, same_sc={_same_sc_carry}")
            elif (USE_LOAD_BALANCING or should_check_increase) and not _s9_active:
                print(f"[DEBUG LB] Using Load Balancing for {item} in week {plan_week}")
                # ดึง daily_capacity จาก required_machines_info ถ้ามี
                _input_daily_cap = required_machines_info[1] if required_machines_info and len(required_machines_info) > 1 else 0
                # ส่งข้อมูลเครื่องปัจจุบันสำหรับกรณี carryover ที่ต้องการเพิ่ม
                _current_machines = machines_in_use.get(_prev_mc_key, 1) if _prev_mc_key else 1

                # 🔧 ตรวจสอบว่าเป็นการใช้ครั้งแรกของ MC_GROUP นี้หรือไม่
                is_first_time_use = True
                current_week_idx = week_index(plan_week)
                current_week_num = int(current_week_idx) if isinstance(current_week_idx, str) else current_week_idx
                _candidate_mc_for_first_use = (
                    _prev_mc_group
                    if (_is_carryover and _prev_mc_group)
                    else (
                        required_machines_info[0]
                        if required_machines_info and len(required_machines_info) > 0
                        else None
                    )
                )
                for (week_idx, mc_grp), total_machines in weekly_mc_usage.items():
                    try:
                        week_num = int(week_idx) if isinstance(week_idx, str) else week_idx
                        if _candidate_mc_for_first_use and mc_grp != _candidate_mc_for_first_use:
                            continue
                        if week_num < current_week_num:
                            # มีการใช้ MC_GROUP นี้มาก่อนในสัปดาห์ที่ผ่านมา
                            is_first_time_use = False
                            break
                    except (ValueError, TypeError):
                        continue
                # ถ้าเป็น carryover ให้ใช้ daily_capacity จากเครื่องปัจจุบัน
                if _is_carryover and _input_daily_cap == 0:
                    _carry_cap_row = item_cap_data[
                        (item_cap_data["ITEM_CODE"] == item)
                        & (item_cap_data["MC_GROUP"] == _prev_mc_group)
                    ]
                    if not _carry_cap_row.empty:
                        _carry_gauge = _carry_cap_row.iloc[0].get("GUAGE")
                        _input_daily_cap = _get_capacity_for_mc_group(item, _prev_mc_group, _carry_gauge)
                        if _input_daily_cap <= 0:
                            _input_daily_cap = _normalize_capacity(item, _prev_mc_group, float(_carry_cap_row.iloc[0]["CAP ทอ"]))
                    _input_daily_cap = adjust_daily_cap_for_item_special(_input_daily_cap, item, _prev_mc_group, _carry_gauge)
                    print(f"[DEBUG CARRY] Using daily_capacity from carryover: {_input_daily_cap}")


                # 🔧 สำคัญ: สำหรับ carryover ที่ต้องการเพิ่มเครื่อง ต้องรักษา mc_group เดิม
                if _is_carryover and should_check_increase:
                    # บังคับให้ใช้ mc_group เดิม ไม่ให้ Load Balancing เปลี่ยน
                    print(f"[DEBUG CARRY] Force using original mc_group: {_prev_mc_group} for carryover increase")
                    # ส่ง mc_group เดิมไปให้ Load Balancing และบอกว่าเป็น carryover
                    # 🔧 FIX: ใช้ _total_item_demand แทน qty_left เพื่อรวม demand จาก FG ถัดไปของ item เดียวกัน
                    # ป้องกันกรณีที่ FG ปัจจุบันเหลือน้อย แต่ยังมี FG อื่นรอผลิต → ควรเพิ่มเครื่องได้
                    _already_produced_this_sc = order_qty - qty_left
                    _already_produced_other_sc = _item_cumulative_planned.get(item, 0)
                    _already_produced = _already_produced_this_sc + _already_produced_other_sc
                    _adjusted_machine_calc = max(0, _qty_for_machine_calc - _already_produced)
                    _total_item_demand = max(qty_left, _adjusted_machine_calc)
                    mc_group_result, daily_capacity, setup_needed, available_machines, _sel_gauge = (
                        get_load_balanced_machine(
                            item,
                            plan_week,
                            last_production,
                            required_machines_info,
                            urgent_mode,
                            past_rdd,
                            force_max_mc=(not _req_feasible and not past_rdd and not is_first_time_use),
                            qty_left=_total_item_demand,
                            daily_capacity=_input_daily_cap,
                            progressive_plan=progressive_plan,
                            current_machines=_current_machines,
                            qty_left_current_fg=qty_left,
                        )
                    )
                    # รักษา mc_group เดิมเสมอ
                    mc_group = _prev_mc_group
                    # ถ้า Load Balancing คืนค่า machines มา ให้ใช้ค่านั้น
                    if available_machines is not None and available_machines > _current_machines:
                        print(f"[DEBUG CARRY] Load Balancing approved increase: {_current_machines} → {available_machines} machines")
                    else:
                        print(f"[DEBUG CARRY] Load Balancing no increase: keep {_current_machines} machines")
                        available_machines = _current_machines
                else:
                    mc_group, daily_capacity, setup_needed, available_machines, _sel_gauge = (
                        get_load_balanced_machine(
                            item,
                            plan_week,
                            last_production,
                            required_machines_info,
                            urgent_mode,
                            past_rdd,
                            force_max_mc=(not _req_feasible and not past_rdd and not is_first_time_use),
                            qty_left=_total_item_demand,
                            daily_capacity=_input_daily_cap,
                            progressive_plan=progressive_plan,
                            current_machines=_current_machines,
                        )
                    )
            elif mc_group is None:
                # ดึง daily_capacity จาก required_machines_info ถ้ามี
                _input_daily_cap = required_machines_info[1] if required_machines_info and len(required_machines_info) > 1 else 0
                _current_machines = machines_in_use.get(_prev_mc_key, 1) if _prev_mc_key else 1
                mc_group, daily_capacity, setup_needed, available_machines, _sel_gauge = (
                    get_load_balanced_machine(
                        item,
                        plan_week,
                        last_production,
                        required_machines_info,
                        urgent_mode,
                        past_rdd,
                        force_max_mc=(not _req_feasible and not past_rdd and not is_first_time_use),
                        qty_left=_total_item_demand,
                        daily_capacity=_input_daily_cap,
                        progressive_plan=progressive_plan,
                        current_machines=_current_machines,
                    )
                )
            if mc_group is None:
                plan_week = next_week(plan_week)
                continue

            # ถ้า mc_group+gauge อยู่ใน MC_GROUP_REDIRECT → เปลี่ยนไปใช้ target แทนเสมอ
            # เช่น SKP 20 → FA 20 (อ้อมน้อย) เพื่อให้วางแผนและหักเครื่องถูกที่
            _redir_mc, _redir_gauge = _apply_mc_redirect(mc_group, _sel_gauge)
            if _redir_mc != mc_group:
                print(f"[MC_REDIRECT] {mc_group}/{_sel_gauge} → {_redir_mc}/{_redir_gauge} for item {item} week {plan_week}")
                mc_group = _redir_mc
                _sel_gauge = _redir_gauge

            # Fallback: ถ้า daily_capacity ยัง None ให้หาจาก item_cap_data
            if daily_capacity is None or daily_capacity <= 0:
                _fallback_cap_row = item_cap_data[
                    (item_cap_data["ITEM_CODE"] == item)
                    & (item_cap_data["MC_GROUP"] == mc_group)
                ]
                if not _fallback_cap_row.empty:
                    _sel_gauge = _fallback_cap_row.iloc[0].get("GUAGE")
                    daily_capacity = _get_capacity_for_mc_group(item, mc_group, _sel_gauge)
                    if daily_capacity is None or daily_capacity <= 0:
                        daily_capacity = _normalize_capacity(item, mc_group, float(_fallback_cap_row.iloc[0]["CAP ทอ"]))
                else:
                    # หาจาก item ทั่วไป
                    _fallback_all = item_cap_data[item_cap_data["ITEM_CODE"] == item]
                    if not _fallback_all.empty:
                        daily_capacity = _normalize_capacity(item, mc_group, float(_fallback_all["CAP ทอ"].min()))
                    else:
                        print(f"[WARNING] No capacity found for {item} in {mc_group}, skipping to next week")
                        plan_week = next_week(plan_week)
                        continue
                daily_capacity = adjust_daily_cap_for_item_special(daily_capacity, item, mc_group, _sel_gauge)
            # ถ้ามี progressive_plan → ใช้จำนวนเครื่องที่คำนวณไว้ล่วงหน้า
            # S9: ไม่ใช้ progressive_plan — S9 pool เป็น independent จาก normal MC plan
            if progressive_plan and plan_week in progressive_plan and not _s9_active:
                available_machines = progressive_plan[plan_week]
            elif required_machines_info and not past_rdd and _req_feasible and not rts_local_force:
                # Cap available_machines ตาม required_mc เมื่อ feasible=True
                # (คำนวณมาแล้วว่า N เครื่องพอตั้งแต่ต้น ไม่ต้องเพิ่มบนกลางคัน)
                req_mc = required_machines_info[2]
                # 🔧 FIX: ถ้า SC เดียวกันมีเครื่องวิ่งอยู่แล้ว (last_sc_machines) อย่าลดต่ำกว่านั้น
                # ป้องกันกรณี FG ถัดไปของ SC เดียวกันใน week เดียวกัน ถูก cap ลงทั้งที่เครื่องยังวิ่งอยู่
                _sc_key_cap = (item, mc_group, _sel_gauge, sc_so_no)
                _sc_existing_mc = last_sc_machines.get(_sc_key_cap, 0)
                if _sc_existing_mc > 0 and req_mc < _sc_existing_mc:
                    print(f"[SC CAP PREVENT] {item} SC {sc_so_no} W{plan_week}: req_mc={req_mc} < existing={_sc_existing_mc} → keep {_sc_existing_mc}")
                    req_mc = _sc_existing_mc
    # ?? REMOVED:             if available_machines > req_mc:
    # ?? REMOVED:                 available_machines = req_mc

            # ถ้า plan_week เกิน target แล้ว → ยังสามารถเพิ่มเครื่องได้ (ไม่ cap ที่ required_mc)
            # 🔧 แก้ไข: บังคับกฎ "สัปดาห์แรกไม่เกิน 2 เครื่อง" เสมอ
            # ตรวจสอบว่าเป็นการใช้ครั้งแรกของ MC_GROUP นี้หรือไม่
            _is_first_week = True
            _current_week_idx = week_index(plan_week)
            _current_week_num = int(_current_week_idx) if isinstance(_current_week_idx, str) else _current_week_idx
            for (week_idx, mc_grp), total_machines in weekly_mc_usage.items():
                try:
                    week_num = int(week_idx) if isinstance(week_idx, str) else week_idx
                    if mc_grp == mc_group and week_num < _current_week_num:
                        _is_first_week = False
                        break
                except (ValueError, TypeError):
                    continue

            # ถ้าเป็นสัปดาห์แรก -> ใช้เครื่องจริง (carry) + ที่ setup ได้สูงสุด 2 เครื่อง
            # S9 จ้างทอ: ไม่มี gradual ramp-up — ใช้เครื่องได้เต็มที่ทันที
            if _is_first_week and not _s9_active:
                # 🔧 FIX: นับ carry เฉพาะจาก item ที่เคยผลิตใน new plan บน MC_GROUP นี้เท่านั้น
                # booking เก่าไม่นับเป็น carry (ต้อง setup ใหม่)
                _fwl_key = _resolve_carry_key(item, mc_group, _sel_gauge)
                if _has_item_mc_key(new_plan_started_items, _fwl_key):
                    carry_machines = machines_in_use.get(_fwl_key, 0)
                else:
                    carry_machines = 0
                carry_machines = max(0, int(carry_machines) if carry_machines else 0)
                # สูงสุด = carry + 2 setup (Gradual Increase)
                first_week_limit = carry_machines + 2
                if available_machines > first_week_limit:
                    print(f"🔒 ENFORCING First week limit for {mc_group} in week {plan_week}: carry={carry_machines} + setup<=2 = {first_week_limit}, reducing from {available_machines} to {first_week_limit} machines")
                    available_machines = first_week_limit
            # Calculate available capacity considering setup days and factory type
            working_days = get_working_days_in_week(plan_week)
            factory_working_days = get_working_days_by_factory(mc_group, available_machines, week=plan_week, gauge=_sel_gauge)
            # วันทำงานจากชีท Work Day เป็นค่าสุดท้าย (ไม่ cap/หักวันหยุด)
            actual_working_days = factory_working_days
            # หา REVOLUTION/WEIGHT ที่มากที่สุด
            rev_weight = get_revolution_weight_from_orders(item, mc_group)
            # กำหนด setup days จาก MasterMC "Set up time" ก่อน ถ้า blank ใช้ MATERIAL_CONTENT/YARN_ITEM
            item_material_content = str(order.get("MATERIAL_CONTENT", "")).strip() or _material_content_lookup.get(str(item).strip().upper(), "")
            _item_yarn_used = str(order.get("YARN-USED", "") or order.get("YARN_USED", "") or order.get("YARN_ITEM", "") or _yarn_used_lookup.get(str(item).strip().upper(), "")).strip()
            item_setup_days = get_setup_days_for_item(item_material_content, _item_yarn_used, mc_group=mc_group, item_code=item)
            # หมายเหตุ: เดิมมีบล็อก urgent mode ตรงนี้เพื่อ "เพิ่มวันทำงาน" ให้ order ที่เลย RDD
            # แต่ถูกยกเลิกถาวรเพราะขัดความเป็นจริง — โรงงานเปิดกี่วันก็ผลิตได้เท่านั้น
            # เร่งด้วยการเสกวันทำงานเพิ่มไม่ได้ (ดู urgent_mode ที่เหลือเป็น no-op ทั้งหมด)
            # การเร่งงานค้างทำผ่าน PAST RDD window (plan_week+3) ใน calculate_required_machines แทน

            # ตรวจสอบว่าสัปดาห์นี้เคยใช้ setup ไปแล้วหรือไม่
            week_key = (plan_week, mc_group)
            factory_working_days = get_working_days_by_factory(mc_group, available_machines, week=plan_week, gauge=_sel_gauge)
            # แยกเครื่อง carry-over (ไม่ต้อง setup) vs เครื่องใหม่ (ต้อง setup)
            mc_key = _resolve_carry_key(item, mc_group, _sel_gauge)
            _sc_key_init = (item, mc_group, _sel_gauge, sc_so_no)
            prev_machines = last_sc_machines.get(_sc_key_init, machines_in_use.get(mc_key, 0))
            # If RTS+LOCAL rule applies and the selected mc_group matches, force carryover-only
            if rts_local_force:
                last_old_by_mc = rts_local_force.get("last_old_by_mc", {})
                machines_by_mc = rts_local_force.get("machines_by_mc", {})
                if str(mc_group).strip().upper() in last_old_by_mc:
                    last_w = last_old_by_mc.get(str(mc_group).strip().upper())
                    start_after = next_week(last_w)
                    if plan_week is None or plan_week < start_after:
                        plan_week = start_after
                        # Recalculate working days for updated plan_week
                        working_days = get_working_days_in_week(plan_week)
                        factory_working_days = get_working_days_by_factory(mc_group, available_machines, week=plan_week, gauge=_sel_gauge)
                        actual_working_days = factory_working_days
                    # ใช้ machines_by_mc (จาก booking_final_ready25 ทุก week) เป็น primary
                    # เพราะ machines_in_use มีแค่ week <= TODAY_WEEK ซึ่งอาจเป็น SO เก่า/เครื่องมากกว่าจริง
                    _bk_mc = machines_by_mc.get(str(mc_group).strip().upper())
                    if _bk_mc is not None and _bk_mc > 0:
                        forced_m = _bk_mc
                    else:
                        forced_m = machines_in_use.get(mc_key, 0)
                    prev_machines = int(forced_m or 0)
            current_week_idx = week_index(plan_week)
            # อ่าน last_production ก่อน _bridge_locks เพราะ _bridge_locks จะเลื่อนค่าไปที่สัปดาห์ lock
            _same_lock = _same_week_lock_mc(mc_key, last_production.get(mc_key))
            _warm_cap = _bridge_locks(mc_key, current_week_idx)
            # ล็อกที่ทับสัปดาห์ผลิต = เครื่องอีกชุดที่ยังอุ่น → บวกเพิ่มก่อนค่อยไปเจอเพดาน
            if _same_lock > 0:
                print(
                    f"[LOCK WARM] {item} W{plan_week}: carry-over {prev_machines}"
                    f" + {_same_lock} เครื่องที่กันไว้สัปดาห์ผลิตก่อนหน้า = {prev_machines + _same_lock}"
                )
                prev_machines += _same_lock
            # กันไว้กี่เครื่อง = อุ่นได้แค่นั้น — ที่เกินต้องนับเป็นเครื่องใหม่ (setup)
            # กติกาเดียวกับฝั่ง OLD ใน AVA_MC.py (_solve_carry_group)
            if _warm_cap is not None and prev_machines > _warm_cap:
                print(
                    f"[LOCK WARM] {item} W{plan_week}: carry-over {prev_machines} → {_warm_cap} เครื่อง"
                    f" (Lock_MC กันไว้ {_warm_cap})"
                )
                prev_machines = _warm_cap
            prev_week_idx = last_production.get(mc_key)
            _booking_week_mc = (
                booking_mc_by_week.get(mc_key, {}).get(current_week_idx, 0)
                if current_week_idx is not None
                else 0
            )
            # ดึง TOTAL_MC_REMAIN จาก booking เพื่อเช็ค remaining capacity
            _booking_total_mc_remain = 0
            if mc_key in booking_mc_by_week and current_week_idx in booking_mc_by_week[mc_key]:
                for _, row in detail_mc.iterrows():
                    if (str(row.get("ITEM_CODE", "")).strip().upper() == item and
                        str(row.get("MC_GROUP", "")).strip().upper() == mc_group and
                        str(row.get("GUAGE", "")).strip() == _sel_gauge and
                        week_index(int(row.get("WEEK"))) == current_week_idx):
                        _booking_total_mc_remain = int(row.get("TOTAL_MC_REMAIN", 0))
                        break
            if _booking_week_mc > 0:
                # ตรวจสอบ remaining capacity ก่อนบังคับใช้ booking machines
                if _booking_total_mc_remain > 0:
                    # มี remaining capacity → ใช้ booking machines
                    prev_machines = max(prev_machines, _booking_week_mc)
                    if available_machines < _booking_week_mc:
                        available_machines = _booking_week_mc
                        print(f"[CARRY BOOKING MC] {item} W{plan_week}: use booking machines={_booking_week_mc} (remaining cap={_booking_total_mc_remain})")
                else:
                    # ไม่มี remaining capacity → ไม่บังคับใช้ booking machines (ให้ไป week ถัดไป)
                    print(f"[NO REMAINING CAP] {item} W{plan_week}: booking {_booking_week_mc} mc but no remaining cap (TOTAL_MC_REMAIN={_booking_total_mc_remain}) → skip forcing, will try next week")
            # is_continuing = week ติดกัน AND เป็น SC/SO NO เดียวกัน (ต่างสี = เริ่มใหม่)
            # For RTS_LOCAL rule we must continue from old regardless of SC/SO, so
            # treat same_order as True when rts_local_force applies for this mc_group.
            # ถ้าเป็นกรณี RTS+LOCAL และ mc_group นี้มี old booking ให้บังคับต่อจาก old (ไม่สน SC/SO)
            if rts_local_force and str(mc_group).strip().upper() in rts_local_force.get(
                "last_old_by_mc", {}
            ):
                same_order = True
            else:
                # เปรียบเทียบ SC/SO
                same_order = last_sc_so_no.get(mc_key) == sc_so_no
                # 🔧 FIX: ใช้ last_sc_week (per-SC) แทน last_sc_so_no (global) เพื่อป้องกัน SC อื่นทับ
                # ถ้า SC นี้เคยผลิตเอง (last_sc_week มีข้อมูล) และ gap ไม่เกิน → same_order = True
                _sc_key_same = (item, mc_group, _sel_gauge, sc_so_no)
                _same_sc_last_week_idx = last_sc_week.get(_sc_key_same)
                if not same_order and _same_sc_last_week_idx is not None and current_week_idx is not None:
                    if (current_week_idx - _same_sc_last_week_idx) <= SETUP_GAP_WEEK:
                        print(f"[SAME SC CARRY] {item} SC {sc_so_no}: last SC week idx {_same_sc_last_week_idx} → current {current_week_idx} (gap {current_week_idx - _same_sc_last_week_idx} ≤ {SETUP_GAP_WEEK}) → same_order=True")
                        same_order = True
                # ถ้าอนุญาต carryover ข้าม SO ให้ตรวจเงื่อนไขเพิ่มเติม
                if not same_order and ALLOW_CARRYOVER_ACROSS_SO:
                    prev_m = machines_in_use.get(mc_key, 0)
                    prev_week_idx = last_production.get(mc_key)
                    if (
                        prev_m > 0
                        and prev_week_idx is not None
                        and current_week_idx is not None
                        and (current_week_idx - prev_week_idx) <= SETUP_GAP_WEEK
                    ):
                        same_order = True
                    # เพิ่มเติม: ถ้า ALLOW_CARRYOVER_ACROSS_SO ให้ same_order = True เสมอ (ข้าม SC/SO ได้)
                    if ALLOW_CARRYOVER_ACROSS_SO:
                        same_order = True
                # Item carry: FG ถัดไปของ item เดียวกัน (คนละ SC/FG) → carry เครื่องต่อ ไม่ต้อง setup ใหม่
                # ทำให้ SC ที่มี TARGET_KNIT ช้ากว่า รอ carry ต่อจาก SC ที่เร็วกว่า
                if not same_order:
                    _item_prev_m = machines_in_use.get(mc_key, 0)
                    _item_prev_widx = last_production.get(mc_key)
                    if (
                        _item_prev_m > 0
                        and _item_prev_widx is not None
                        and current_week_idx is not None
                        and (current_week_idx - _item_prev_widx) <= SETUP_GAP_WEEK
                    ):
                        same_order = True  # item-level carry: inherit machines จาก FG ก่อนหน้า
            _is_same_sc = (last_sc_so_no.get(mc_key) == sc_so_no) or (_same_sc_last_week_idx is not None)
            # Item carry: ถ้า item+mc กำลังวิ่งอยู่ใน future week (prev_week_idx > current_week_idx)
            # หมายความว่า SC อื่นของ item เดียวกันยังผลิตอยู่ → ดัน plan_week ไปต่อท้าย แล้ว carry
            if (
                not _s9_active  # S9 ใช้เครื่อง commission knitting คนละ pool — ไม่ต้องรอ normal machine carryover
                and prev_week_idx is not None
                and current_week_idx is not None
                and prev_week_idx > current_week_idx
                and _has_item_mc_key(new_plan_started_items, mc_key)
            ):
                # เครื่องยังวิ่งอยู่ใน future → ข้ามไป week ถัดจาก week สุดท้ายที่ผลิต
                _after_prev = None
                _prev_cal_week = int(calendar_week.iloc[prev_week_idx]["WEEK"]) if prev_week_idx < len(calendar_week) else None
                if _prev_cal_week is not None:
                    _after_prev = _prev_cal_week  # เริ่มที่ week เดียวกับที่ SC ก่อนหน้าจบ → carry เข้าได้เลย ไม่ลดเครื่อง
                if _after_prev is not None:
                    print(f"[ITEM CARRY WAIT] {item} SC {sc_so_no}: machine busy until W{_prev_cal_week} → advance plan_week W{plan_week}→W{_after_prev} (carry in same week)")
                    plan_week = _after_prev
                    current_week_idx = week_index(plan_week)
                    # 🔧 FIX: After advancing plan_week, this is now a carryover situation.
                    # available_machines was set by LB for the ORIGINAL plan_week (possibly first-time use)
                    # but after ITEM CARRY WAIT it should respect existing carry machines.
                    # ป้องกัน LB สั่งเพิ่มเครื่องทั้งที่ carry เดิมเพียงพอผลิตได้ใน 1-2 weeks ถัดไป
                    _sc_key_wait = (item, mc_group, _sel_gauge, sc_so_no)
                    _carry_mc_after_wait = last_sc_machines.get(_sc_key_wait, machines_in_use.get(mc_key, 0))
                    if _carry_mc_after_wait > 0 and available_machines > _carry_mc_after_wait:
                        print(f"[ITEM CARRY WAIT] Capping available_machines {available_machines} → {_carry_mc_after_wait} (existing carry machines)")
                        available_machines = _carry_mc_after_wait
                same_order = True  # carry จาก week สุดท้ายของ SC ก่อนหน้า

            # ถ้าเงื่อนไข carryover อนุญาต (same_order=True) ให้ same-week ต่อได้
            _week_ok = (
                current_week_idx is not None
                and prev_week_idx is not None
                and (current_week_idx >= prev_week_idx)
            )
            is_continuing = (
                prev_week_idx is not None
                and current_week_idx is not None
                and same_order
                and _week_ok
                and (
                    current_week_idx - prev_week_idx <= 1  # week ติดกัน หรือ same week
                    or (current_week_idx - prev_week_idx <= SETUP_GAP_WEEK)
                )
            )  # ต้องไม่ห่างเกิน gap
            # 🔧 NEW: ถ้า booking active ครอบ plan_week (แต่ last_production เป็นสัปดาห์อนาคต)
            # → ถือว่า is_continuing เพราะเครื่องกำลังวิ่งอยู่ใน plan_week จาก booking เดิม
            if (
                not is_continuing
                and current_week_idx is not None
                and mc_key in booking_active_week_set
                and current_week_idx in booking_active_week_set[mc_key]
            ):
                is_continuing = True
                # ยก prev_machines ให้ตรงกับ booking machine count จริง เพื่อให้ carryover_mc ถูกต้อง
                if _booking_week_mc > prev_machines:
                    print(f"[IS_CONTINUING BOOKING SPAN] {item} W{plan_week}: booking active → is_continuing=True, raise prev_machines {prev_machines}→{_booking_week_mc}")
                    prev_machines = _booking_week_mc
                else:
                    print(f"[IS_CONTINUING BOOKING SPAN] {item} W{plan_week}: booking active ครอบ plan_week → is_continuing=True")
            # ❗ ถ้า item+mc นี้ยังไม่เคยผลิตใน new plan → บังคับ setup (ไม่อ้าง old plan)
            # ยกเว้น: ถ้ามีข้อมูลจาก booking จริง (detail_mc) → อนุญาต carryover จาก booking ได้
            if not _has_item_mc_key(new_plan_started_items, mc_key) and not _has_item_mc_key(booking_production_keys, mc_key):
                is_continuing = False

            # 🔧 NEW: ตรวจ intermediate weeks — ถ้ามี week ที่เครื่องว่าง=0 จาก item อื่น → ตัด carry
            # Rule: remaining = 0 ใน gap + item นี้ไม่ได้ผลิตใน week นั้น = เครื่องถูก item อื่นยึดครบ
            # Exception: ถ้า item เดิมผลิตอยู่เองและเครื่องเหลือ 0 → ยังคง carry ได้
            if is_continuing and prev_week_idx is not None and current_week_idx is not None:
                if _carry_blocked_by_gap(item, mc_key, mc_group, _sel_gauge, prev_week_idx, current_week_idx):
                    is_continuing = False
                    setup_needed = True

            # S9: ใช้เครื่อง commission knitting แยก pool — ไม่ carry จาก normal
            if _s9_active:
                is_continuing = False
            _cyl_pending_cnt = 0  # จำนวนเครื่องที่ถูก cylinder change และพร้อมใช้ week นี้ (committed)
            if is_continuing:
                # เครื่อง carry = เครื่องที่วิ่งอยู่แล้วจากสัปดาห์ก่อน จึงไม่ต้อง setup ใหม่
                # แต่เครื่องนั้นต้อง "ยังอยู่ในพูล" ของสัปดาห์นี้ด้วย — ถ้า booking ของสัปดาห์นี้
                # โตขึ้นจนจองเกินเครื่องจริง (remain ติดลบ) เครื่องที่แผนถืออยู่ถูกยึดไปแล้ว
                # → carry ต่อไม่ได้ (เดิมยก available_machines = prev_machines ทับ [CARRY CLAMP])
                # เครื่องของ booking ที่ถักไอเทมนี้อยู่ ถูกหักจาก remain ไปแล้ว → บวกกลับก่อนเทียบ
                _carry_pool_remain = get_actual_mc_remain(
                    mc_group, plan_week, gauge=_sel_gauge, item_code=item, allow_negative=True
                ) + int(_booking_week_mc or 0)
                if _carry_pool_remain < prev_machines:
                    _carry_cap = max(0, min(prev_machines, _carry_pool_remain))
                    print(
                        f"[CARRY POOL FULL] {item} SC {sc_so_no} W{plan_week}: "
                        f"pool remain={_carry_pool_remain} < prev={prev_machines} "
                        f"→ carry ได้ {_carry_cap} เครื่อง (booking จองเกินพูล)"
                    )
                    available_machines = min(available_machines, _carry_cap)
                # 🔧 FIX: ถ้า SC เดียวกันกำลัง carry ต่อ อย่าลดเครื่องต่ำกว่าที่วิ่งอยู่แล้ว
                # ป้องกันกรณี FG ถัดไปของ SC เดียวกัน ถูก cap ลงทั้งที่เครื่องยังวิ่งอยู่
                elif prev_machines > available_machines:
                    print(f"[SC CONTINUING RAISE] {item} SC {sc_so_no} W{plan_week}: prev={prev_machines} > avail={available_machines} → raise to {prev_machines}")
                    available_machines = prev_machines
                # Cyl-changed machines ready THIS week (triggered in carry path of previous week)
                # เครื่องเหล่านี้เป็น NEW (ต้อง setup) ไม่ใช่ carry — เพิ่ม available แบบ additive
                if _sel_gauge:
                    _cyl_t1_k = _mc_to_type1(mc_group, _sel_gauge)
                    _cyl_g_k = _normalize_gauge(_sel_gauge)
                    _item_cyl_k = str(item).strip().upper()
                    _cyl_pending_cnt = _carry_cyl_pending.get((int(plan_week), _item_cyl_k, _cyl_t1_k, _cyl_g_k), 0)
                    if _cyl_pending_cnt > 0:
                        available_machines = max(available_machines, prev_machines) + _cyl_pending_cnt
                        print(f"[CARRY CYL READY] {item} W{plan_week}: +{_cyl_pending_cnt} cyl machine(s) ready → available={available_machines}")
                # ถ้า required_mc > available_machines (carry) → trigger cylinder change เท่าที่ทำได้ใน week นี้
                # machine พร้อมใน NEXT week — จำนวนที่ trigger ขึ้นกับ quota (CYLINDER_CHANGE_LIMIT)
                _req_mc_carry = required_machines_info[2] if required_machines_info and len(required_machines_info) > 2 else 0
                _cyl_this_wk_cap = (available_machines * daily_capacity * get_working_days_by_factory(mc_group, 1, week=plan_week, gauge=_sel_gauge)) if daily_capacity else 0
                # เช็คว่า native gauge pool ยังมีเครื่องเหลืออยู่ไหม — ถ้ายังมี ไม่ต้อง cylinder change
                # (available_machines อาจต่ำกว่า required เพราะ MAX_NEW_SETUP_MC cap ไม่ใช่เพราะ pool หมด)
                _cyl_native_pool = get_actual_mc_remain(mc_group, plan_week, gauge=_sel_gauge, item_code=item) if _sel_gauge else 0
                if _req_mc_carry > available_machines and _sel_gauge and not past_rdd and qty_left > _cyl_this_wk_cap and _cyl_native_pool <= 0:
                    _cyl_cat_carry = _mc_to_type1(mc_group, _sel_gauge)
                    _cyl_fact_carry = _mc_to_factory(mc_group, _sel_gauge)
                    _cyl_tgt_g_carry = _normalize_gauge(_sel_gauge)
                    _cyl_base_trigger = int(plan_week)  # trigger week นี้ → machine พร้อม week ถัดไป
                    print(f"[CARRY CYL ATTEMPT] {item} W{plan_week}: carry={available_machines} < required={_req_mc_carry}, trigger W{_cyl_base_trigger} ({_cyl_fact_carry}/{_cyl_cat_carry}/G{_cyl_tgt_g_carry}) [jit_override=True]")
                    _cyl_pending_added = 0
                    while _req_mc_carry > available_machines + _cyl_pending_added:
                        if _try_cylinder_change(_cyl_cat_carry, _cyl_fact_carry, _cyl_tgt_g_carry, _cyl_base_trigger, item, mc_group, debug=True, jit_override=True):
                            _cyl_pending_added += 1
                            _next_cyl_w = next_week(plan_week)
                            if _next_cyl_w is not None:
                                _pk = (int(_next_cyl_w), str(item).strip().upper(), _cyl_cat_carry, _cyl_tgt_g_carry)
                                _carry_cyl_pending[_pk] = _carry_cyl_pending.get(_pk, 0) + 1
                                _cyl_f_up_c = str(_cyl_fact_carry).strip().upper()
                                for _fw_u_c in sorted(int(w) for w in summary_mc["WEEK"].unique() if int(w) >= int(_next_cyl_w)):
                                    _undo_k_c = (_fw_u_c, _cyl_f_up_c, _cyl_cat_carry, _cyl_tgt_g_carry)
                                    if cylinder_adjustments.get(_undo_k_c, 0) > 0:
                                        cylinder_adjustments[_undo_k_c] -= 1
                            print(f"[CARRY CYL CHANGE] {item} W{plan_week}: trigger W{_cyl_base_trigger} → +{_cyl_pending_added} machine(s) ready W{_next_cyl_w}")
                        else:
                            break  # quota เต็มสำหรับ week นี้
                carryover_mc = min(prev_machines, available_machines)

                # DEBUG: เช็คค่าสำหรับ FD5PRTJJ20/37A0
                if "FD5PRTJJ20" in str(item) and plan_week == 20:
                    print(f"[DEBUG CARRYOVER] {item} W{plan_week}: prev_machines={prev_machines}, available_machines={available_machines}, carryover_mc={carryover_mc}")

                # 🔧 FIX: ถ้า qty_left น้อย → ลด carryover_mc ให้พอเหมาะกับ qty_left
                # ป้องกันกรณี carryover เครื่องเยอะแต่ qty_left น้อย (เช่น FD1BASMZ26B0 W23)
                if daily_capacity and qty_left > 0 and carryover_mc > 1:
                    # คำนวณ weekly capacity สำหรับ 1 เครื่อง
                    factory_wd = get_working_days_by_factory(mc_group, 1, week=plan_week, gauge=_sel_gauge)
                    weekly_cap_per_mc = daily_capacity * factory_wd
                    # 🔧 FIX: ใช้ _qty_for_machine_calc (รวมทุก FG ของ item) ไม่ใช่แค่ qty_left (FG ปัจจุบัน)
                    # ป้องกันลดเครื่องทั้งที่ item ยังมี FG ถัดไปรอผลิตอยู่
                    _reduce_qty = max(qty_left, _qty_for_machine_calc)
                    # คำนวณจำนวนเครื่องที่ต้องการจริงๆ ตาม working days จริง
                    needed_mc = max(1, int(_reduce_qty / (factory_wd * daily_capacity)) + 1)
                    # ถ้า needed_mc < carryover_mc → ลด carryover_mc
                    if needed_mc * 2 < carryover_mc:
                        print(f"[CARRYOVER REDUCE] {item} W{plan_week}: carryover_mc {carryover_mc} → {needed_mc} (qty_left={qty_left:.0f}, factory_wd={factory_wd}, weekly_cap_per_mc={weekly_cap_per_mc:.0f})")
                        carryover_mc = needed_mc
                        # อัปเดต available_machines ตาม carryover_mc ที่ลดลง
                        # 🔧 FIX: รักษา _cyl_pending_cnt — เครื่องที่ถูก cylinder change ไปแล้วต้องใช้เสมอ
                        available_machines = min(available_machines, carryover_mc + _cyl_pending_cnt)

                # เครื่องใหม่ = available_machines - carryover
                # 🔧 FIX: บังคับให้ available_machines >= carryover_mc เพื่อป้องกัน new_mc ติดลบ
                available_machines = max(available_machines, carryover_mc)
                new_mc = max(0, available_machines - carryover_mc)
            else:
                carryover_mc = 0
                new_mc = available_machines  # ทุกเครื่องต้อง setup
                # 🔧 FIX: ถ้า is_continuing=False → ต้อง setup ใหม่เสมอ
                # ป้องกัน RTS/LB set setup_needed=False ทั้งที่ gap > SETUP_GAP_WEEK
                if not _is_carryover:
                    setup_needed = True
                # Trigger cylinder change ใน setup week — trigger เท่าที่ทำได้เพื่อทัน target
                # machine พร้อมใน NEXT week — จำนวนขึ้นกับ quota (CYLINDER_CHANGE_LIMIT)
                _req_mc_setup = required_machines_info[2] if required_machines_info and len(required_machines_info) > 2 else 0
                _cyl_this_wk_cap_s = (available_machines * daily_capacity * get_working_days_by_factory(mc_group, 1, week=plan_week, gauge=_sel_gauge)) if daily_capacity else 0
                # เช็คว่า native gauge pool ยังมีเครื่องเหลืออยู่ไหม — ถ้ายังมี ไม่ต้อง cylinder change
                # (new_mc อาจต่ำกว่า required เพราะ MAX_NEW_SETUP_MC cap ไม่ใช่เพราะ pool หมด)
                _cyl_native_pool_s = get_actual_mc_remain(mc_group, plan_week, gauge=_sel_gauge, item_code=item) if _sel_gauge else 0
                if _req_mc_setup > new_mc and _sel_gauge and not past_rdd and qty_left > _cyl_this_wk_cap_s and _cyl_native_pool_s <= 0:
                    _cyl_cat_s = _mc_to_type1(mc_group, _sel_gauge)
                    _cyl_fact_s = _mc_to_factory(mc_group, _sel_gauge)
                    _cyl_tgt_g_s = _normalize_gauge(_sel_gauge)
                    _cyl_base_s = int(plan_week)
                    _cyl_pending_added_s = 0
                    while _req_mc_setup > new_mc + _cyl_pending_added_s:
                        if _try_cylinder_change(_cyl_cat_s, _cyl_fact_s, _cyl_tgt_g_s, _cyl_base_s, item, mc_group, debug=True, jit_override=True):
                            _cyl_pending_added_s += 1
                            _next_cyl_w_s = next_week(plan_week)
                            if _next_cyl_w_s is not None:
                                _pk_s = (int(_next_cyl_w_s), str(item).strip().upper(), _cyl_cat_s, _cyl_tgt_g_s)
                                _carry_cyl_pending[_pk_s] = _carry_cyl_pending.get(_pk_s, 0) + 1
                                _cyl_f_up_s = str(_cyl_fact_s).strip().upper()
                                for _fw_u_s in sorted(int(w) for w in summary_mc["WEEK"].unique() if int(w) >= int(_next_cyl_w_s)):
                                    _undo_k_s = (_fw_u_s, _cyl_f_up_s, _cyl_cat_s, _cyl_tgt_g_s)
                                    if cylinder_adjustments.get(_undo_k_s, 0) > 0:
                                        cylinder_adjustments[_undo_k_s] -= 1
                            print(f"[SETUP CYL CHANGE] {item} W{plan_week}: setup trigger W{_cyl_base_s} → +{_cyl_pending_added_s} machine(s) ready W{_next_cyl_w_s}")
                        else:
                            break  # quota เต็มสำหรับ week นี้
            # Hard cap POLY/COTTON: เครื่องที่แผนใช้ (carryover + new) ห้ามเกินที่กันไว้
            # remain = กันไว้ − booking − แผนที่ใช้ไปแล้วในสัปดาห์นั้น
            # ถ้าไม่มีเครื่องเหลือ → carry ก็ทำไม่ได้ (ไม่มีเครื่องให้ carry) ตามที่ user ยืนยัน
            # ตัด new ก่อน แล้วค่อยตัด carry (รักษาเครื่องที่วิ่งอยู่ไว้ก่อน)
            _sp_remain = _mc_special_remain(mc_group, plan_week, _sel_gauge, item)
            if _sp_remain is not None and (carryover_mc + new_mc) > _sp_remain:
                _sp_old = (carryover_mc, new_mc)
                carryover_mc = min(carryover_mc, _sp_remain)
                new_mc = max(0, min(new_mc, _sp_remain - carryover_mc))
                available_machines = carryover_mc + new_mc
                print(f"[POLY/COTTON CAP] {item} W{plan_week}: carry/new {_sp_old} → ({carryover_mc}, {new_mc}) remain={_sp_remain} (เครื่องกันไว้เต็ม)")

            # Enforce RTS+LOCAL: use existing carryover machines only (no new setup)
            # prev_machines comes from machines_in_use (last active week, MC_USE_CEIL>0)
            # 🔧 FIX: เช็ค is_continuing ก่อน — ถ้า gap > SETUP_GAP_WEEK ต้อง setup ใหม่
            if rts_local_force and str(mc_group).strip().upper() in rts_local_force.get(
                "last_old_by_mc", {}
            ) and is_continuing:
                # 🔧 FIX: จำกัด carryover ไม่ให้เกิน available_machines
                carryover_mc = min(int(prev_machines or 0), available_machines)
                new_mc = 0
                available_machines = carryover_mc
                setup_needed = False
            # ===== Carryover-first: ตรวจว่า carryover เพียงพอทัน rdd ไหม =====
            # Simulate production จาก plan_week ถึง rdd_idx ด้วย carry เครื่อง
            # Week 1: carry ผลิตเต็ม, new ผลิตหัก setup  |  Week 2+: ทุกเครื่องเป็น carry
            # YD-ORDERS: คำนวณ SUB_COLOR setup days ก่อน forward sim
            _yd_sub_color_setup = 0
            if order_type == "YD-ORDERS" and carryover_mc > 0:
                _prev_sub_color = last_sub_color.get(mc_key, "")
                if _prev_sub_color and sub_color and _prev_sub_color != sub_color:
                    _wcs_key = (mc_key, plan_week)
                    _yd_week_color_setups[_wcs_key] = _yd_week_color_setups.get(_wcs_key, 0) + 1
                    _yd_sub_color_setup = 1  # หัก 1 วันต่อครั้งที่เปลี่ยนสี (ไม่ใช้ cumulative เพราะ remaining cap หักไปแล้วจากครั้งก่อน)

            def _forward_sim(carry, new, q_left):
                q = q_left
                wk = plan_week
                first = True
                _rw = rev_weight or 0
                # เมื่อ plan_week เลย rdd แล้ว (past_rdd) → ไม่ตัด sim ที่ rdd_idx
                # เพราะต้องจำลองต่อเพื่อเช็คว่า carry-only เพียงพอผลิตให้เสร็จหรือไม่
                # ป้องกัน setup เครื่องใหม่ที่ใช้แค่ 1 week แล้วหยุด
                _is_past_rdd = (
                    rdd_idx is not None
                    and current_week_idx is not None
                    and current_week_idx > rdd_idx
                )
                _sim_count = 0
                while wk is not None and q > _rw:
                    if _sim_count > 300:  # SAFETY: กัน infinite loop ข้ามปลายปี
                        break
                    w_idx_sim = week_index(wk)
                    if w_idx_sim is None:
                        break
                    # ยังไม่เลย rdd → หยุดจำลองที่ rdd เหมือนเดิม
                    if not _is_past_rdd and rdd_idx is not None and w_idx_sim > rdd_idx:
                        break
                    # เลย rdd แล้ว → จำกัดจำลอง 2 สัปดาห์
                    # setup เครื่องใหม่ที่ใช้ < 2 สัปดาห์ไม่คุ้ม setup days
                    # ถ้า carry-only ผลิตครบภายใน 2 สัปดาห์ → ไม่ต้อง setup เพิ่ม
                    if _is_past_rdd and _sim_count >= 2:
                        break
                    # หักวันหยุดตามปฏิทินให้แล้วใน WorkDay.get_working_days() — ห้าม min ซ้ำ
                    wd = get_working_days_by_factory(mc_group, carry + new, week=wk, gauge=_sel_gauge)
                    if first:
                        # YD-ORDERS: SUB_COLOR เปลี่ยน → carryover ต้องเสีย 1 วัน setup
                        _carry_wd = max(0, wd - _yd_sub_color_setup) if carry > 0 else wd
                        c_prod = carry * _carry_wd * daily_capacity
                        n_prod = new * max(0, wd - item_setup_days) * daily_capacity
                        first = False
                        total = fold_round_mixed(
                            c_prod + n_prod, rev_weight, mc_group,
                            _carry_wd * daily_capacity, carry,
                            max(0, wd - item_setup_days) * daily_capacity, new,
                        )
                    else:
                        c_prod = (carry + new) * wd * daily_capacity
                        n_prod = 0
                        total = fold_round(
                            c_prod, rev_weight, mc_group,
                            per_mc_cap=wd * daily_capacity, mc_count=carry + new,
                        )
                    q -= total
                    wk = next_week(wk)
                    _sim_count += 1
                return q

            def _carry_span_blocked(carry, new):
                """หาสัปดาห์แรกที่งานนี้จะ carry ต่อไม่ได้เพราะพูลไม่มีเครื่องเหลือ

                จำลองการผลิตด้วยเครื่องชุดนี้ไปข้างหน้า สัปดาห์ไหนที่ยังผลิตไม่จบต้อง carry ต่อ
                → สัปดาห์นั้นต้องมีเครื่องในพูลพอ ถ้า booking จองเกินพูลแล้ว (remain ติดลบ)
                เครื่องที่แผนถืออยู่จะถูกยึด → setup ที่สัปดาห์นี้ไม่มีประโยชน์ (เสียวัน setup เปล่า)

                คืน (สัปดาห์ที่ติด, เครื่องเหลือจริง, เครื่องที่ต้องใช้) หรือ None ถ้าวิ่งได้ครบ
                """
                total_mc = carry + new
                if total_mc <= 0 or not daily_capacity:
                    return None
                q = qty_left
                wk = plan_week
                first = True
                _rw = rev_weight or 0
                _guard = 0
                while wk is not None and q > _rw and _guard < 26:
                    _guard += 1
                    # หักวันหยุดตามปฏิทินให้แล้วใน WorkDay.get_working_days() — ห้าม min ซ้ำ
                    wd = get_working_days_by_factory(mc_group, total_mc, week=wk, gauge=_sel_gauge)
                    if first:
                        prod = (
                            carry * wd * daily_capacity
                            + new * max(0, wd - item_setup_days) * daily_capacity
                        )
                        first = False
                    else:
                        prod = total_mc * wd * daily_capacity
                    if _rw > 0 and prod > 0:
                        prod = (prod // _rw) * _rw
                    q -= prod
                    if q <= _rw:
                        return None  # ผลิตจบในสัปดาห์นี้ → ไม่ต้อง carry ต่อ
                    wk = next_week(wk)
                    while wk is not None and wk in SKIP_WEEKS:
                        wk = next_week(wk)
                    if wk is None:
                        return None  # หมด horizon → ปล่อยให้ logic เดิมจัดการ
                    # สัปดาห์ถัดไปต้อง carry ต่อ: เครื่อง booking ของไอเทมนี้ถูกหักจาก remain
                    # ไปแล้ว จึงนับเป็นเครื่องที่ใช้ได้ (ไม่ต้องขอจากพูลซ้ำ)
                    _bk_next = int(
                        booking_mc_by_week.get(_ck(item, mc_group, _sel_gauge), {}).get(
                            week_index(wk), 0
                        )
                        or 0
                    )
                    _need = max(0, total_mc - _bk_next)
                    if _need <= 0:
                        continue
                    _rem_next = get_actual_mc_remain(
                        mc_group, wk, gauge=_sel_gauge, item_code=item, allow_negative=True
                    )
                    if _rem_next < _need:
                        return (wk, _rem_next, _need)
                return None

            if rdd_idx is not None and not rts_local_force and new_mc > 0 and not _s9_active:
                _rw_tol = 0.0
                # ตรวจก่อน: carry เพียงพอจบ FG ปัจจุบันหรือไม่ (qty_left FG นี้เท่านั้น)
                # ป้องกันการเพิ่มเครื่องเพราะ total demand สูง ทั้งที่ carryover จบ FG นี้ได้พอ
                if carryover_mc > 0 and _forward_sim(carryover_mc, 0, qty_left) <= _rw_tol:
                    new_mc = max(0, _cyl_pending_cnt)  # 🔧 FIX: รักษาเครื่องที่ committed cylinder change
                    if new_mc == 0:
                        print(f"   [NO NEW MC] Carryover {carryover_mc} mc เพียงพอจบ FG ปัจจุบัน (qty={qty_left:.0f}) → ไม่เพิ่มเครื่องใหม่")
                    else:
                        print(f"   [NO NEW MC KEEP CYL] Carryover {carryover_mc} mc เพียงพอ แต่รักษา {new_mc} cyl machine(s) → new_mc={new_mc}")
                else:
                    # ใช้ demand ระดับ item (ทุก FG ของ item) เพื่อคงเครื่องไว้เมื่อยังมี FG เหลือของ item
                    _fwd_qty = _qty_for_machine_calc if _qty_for_machine_calc > qty_left else qty_left
                    if carryover_mc > 0 and _forward_sim(carryover_mc, 0, _fwd_qty) <= _rw_tol:
                        # carryover เพียงพอทัน → ไม่ต้อง setup เพิ่ม
                        new_mc = max(0, _cyl_pending_cnt)  # 🔧 FIX: รักษาเครื่องที่ committed cylinder change
                    else:
                        # หา new_mc น้อยสุดที่ทัน (จาก carryover + new)
                        found_n = new_mc  # fallback = ทั้งหมด
                        for try_n in range(1, new_mc + 1):
                            if _forward_sim(carryover_mc, try_n, _fwd_qty) <= _rw_tol:
                                found_n = try_n
                                break
                        new_mc = found_n
                available_machines = carryover_mc + new_mc
            # จำนวน new_mc ที่ _forward_sim หาได้ = น้อยสุดที่ทัน → เป็น lower bound
            _forward_min_new = new_mc

            # ===== Dynamic setup limit ตาม urgency RDD =====
            # S9: ข้าม limit ทั้งหมด — ใช้เต็ม pool ที่เหลือ
            if not _s9_active:
                _remaining_slots = get_remaining_job_slots(plan_week, mc_group)
                _req_mc_dyn = required_machines_info[2] if required_machines_info else new_mc
                # ถ้า simulate แล้วไม่ทัน → เปิดเต็ม slots (ไม่ cap ที่ required_mc)
                if not _req_feasible:
                    _dyn_limit = _remaining_slots
                else:
                    _dyn_limit = _dynamic_setup_limit(
                        plan_week, rdd_idx, _req_mc_dyn, _remaining_slots
                    )
                # ห้าม cap ต่ำกว่า _forward_min_new — ไม่งั้น forward sim ไม่มีความหมาย
                _dyn_limit = max(_dyn_limit, _forward_min_new)
                if new_mc > _dyn_limit:
                    new_mc = _dyn_limit
                # ===== Hard cap: เครื่องใหม่ (new setup) ≤ MAX_NEW_SETUP_MC เสมอ =====
                if new_mc > MAX_NEW_SETUP_MC:
                    new_mc = MAX_NEW_SETUP_MC
                available_machines = carryover_mc + new_mc
            # ===== Setup payback gate: เครื่องใหม่ต้องได้วิ่งเต็ม >= 2 สัปดาห์ =====
            # เครื่อง setup ใหม่จ่าย setup days ในสัปดาห์แรก ถ้าสัปดาห์ถัดไปงานเหลือน้อย
            # จน carry เดิมทำคนเดียวได้ เครื่องนั้นจะถูกปลดทันที → เสีย setup ฟรี
            # และแผนจบช้ากว่าการไม่ setup เลย
            def _new_mc_full_weeks(carry, new, q_start):
                """นับสัปดาห์ (หลังสัปดาห์ setup) ที่งานยังเหลือมากพอให้เครื่องใหม่วิ่งเต็ม"""
                if new <= 0 or not daily_capacity:
                    return 0
                _rw_pb = rev_weight or 0

                def _wd_pb(wk_, mc_):
                    # หักวันหยุดตามปฏิทินให้แล้วใน WorkDay.get_working_days() — ห้าม min ซ้ำ
                    return get_working_days_by_factory(mc_group, mc_, week=wk_, gauge=_sel_gauge)

                def _round_pb(v):
                    if _rw_pb > 0 and v > 0:
                        return (v // _rw_pb) * _rw_pb
                    return v

                q_pb = float(q_start)
                wk_pb = plan_week
                _wd0 = _wd_pb(wk_pb, carry + new)
                # วันผลิตของเครื่องใหม่ในสัปดาห์ setup — สูตรเดียวกับ prod_days_new ด้านล่าง
                if _wd0 < item_setup_days:
                    _pdn0 = max(0, _wd0 - 0.5)
                else:
                    _pdn0 = max(0.5, _wd0 - item_setup_days)
                q_pb -= _round_pb(
                    carry * _wd0 * daily_capacity + new * _pdn0 * daily_capacity
                )
                _full_pb = 0
                _guard_pb = 0
                while q_pb > _rw_pb and _guard_pb < 26:
                    _guard_pb += 1
                    wk_pb = next_week(wk_pb)
                    while wk_pb is not None and wk_pb in SKIP_WEEKS:
                        wk_pb = next_week(wk_pb)
                    if wk_pb is None:
                        break
                    _wd_n = _wd_pb(wk_pb, carry + new)
                    _carry_only_pb = _round_pb(carry * _wd_n * daily_capacity)
                    if q_pb <= _carry_only_pb:
                        break  # carry เดิมพอแล้ว → เครื่องใหม่ไม่ได้วิ่งเต็มสัปดาห์นี้
                    _full_pb += 1
                    q_pb -= _round_pb((carry + new) * _wd_n * daily_capacity)
                return _full_pb

            # (ด่าน payback ถูกบังคับใช้หลัง GRADUAL BOOST — จุดสุดท้ายที่ยังเพิ่มเครื่องได้)

            # ใช้ actual_working_days (หักวันหยุดแล้ว) แทน factory_working_days แบบ static
            prod_days_old = actual_working_days  # เครื่อง carry-over ผลิตตามวันเปิดจริง
            # YD-ORDERS: ถ้า SUB_COLOR เปลี่ยน → carryover machines เสีย 1 วัน setup จากที่เหลือ
            if _yd_sub_color_setup > 0:
                prod_days_old = max(0, prod_days_old - 1)
                _total_color_setups = _yd_week_color_setups.get((mc_key, plan_week), 0)
                print(f"[YD SUB_COLOR] {item} W{plan_week}: SUB_COLOR เปลี่ยนครั้งที่ {_total_color_setups} → prod_days_old={prod_days_old} (หัก 1 วันจากที่เหลือ)")
            # เครื่องใหม่ (new_mc) ต้อง setup เสมอ แม้ item จะ warm บนเครื่องเดิม (setup_needed=False)
            # setup_needed=False หมายถึงเครื่องที่วิ่งอยู่แล้ว ไม่ใช่เครื่องที่เพิ่งเพิ่มมา
            # 🔧 FIX: ถ้าวันทำงานน้อยกว่า setup days → ใช้ 0.5 วัน setup แทน (เพื่อให้มีวันผลิตบ้าง)
            # S9: ไม่หัก setup — ผลิตเต็มวัน
            if _s9_active:
                prod_days_new = actual_working_days
            elif new_mc > 0 and actual_working_days < item_setup_days:
                # วันทำงานน้อยเกินไป → ลด setup เป็น 0.5 วัน
                prod_days_new = max(0, actual_working_days - 0.5)
            elif new_mc > 0:
                # ปกติ แต่ถ้า actual_working_days == item_setup_days → ผลิต 0 วัน → บังคับ 0.5 วันขั้นต่ำ
                prod_days_new = max(0.5, actual_working_days - item_setup_days)
            else:
                prod_days_new = max(0, actual_working_days - item_setup_days)

            # ===== NEW: Multi-FG min-mc — ลดเครื่องตาม demand รวมของทุก FG ที่เหลือ =====
            # ถ้า total item demand ผลิตครบได้ภายใน latest FG target ด้วยเครื่องน้อยกว่า
            # ปัจจุบัน → ลดลง เพื่อ free เครื่องให้ item อื่น
            # (ทำนอกบล็อก _sim_produce > qty_left เพราะ FG ปัจจุบันอาจใหญ่กว่า 1 week cap)
            if (
                carryover_mc + new_mc > 1
                and _qty_for_machine_calc > qty_left + 1.0
                and daily_capacity
            ):
                _cur_fg_int_m = 0
                try:
                    _cur_fg_int_m = int(fg_week or 0)
                except Exception:
                    pass
                _latest_tgt_idx_m = rdd_idx if rdd_idx is not None else current_week_idx
                for _oi_m, _or_m in orders_sorted.iterrows():
                    if str(_or_m.get("Item Code", "")).strip() != str(item).strip():
                        continue
                    _ofg_m = _or_m.get("FG Week")
                    if not pd.notna(_ofg_m):
                        continue
                    try:
                        _ofgi_m = int(_ofg_m)
                    except Exception:
                        continue
                    if _ofgi_m < _cur_fg_int_m:
                        continue
                    _ofs_m = str(_ofgi_m)
                    try:
                        if len(_ofs_m) >= 5:
                            _oy_m = int(_ofs_m[:4])
                            _owk_m = int(_ofs_m[4:])
                        elif len(_ofs_m) <= 2:
                            _oy_m = TODAY.year
                            _owk_m = int(_ofs_m)
                        else:
                            _oy_m = TODAY.year
                            _owk_m = int(_ofs_m)
                        _orow_m = calendar_week[
                            (calendar_week["YEAR"] == _oy_m)
                            & (calendar_week["WEEK"] == _owk_m)
                        ]
                        if _orow_m.empty:
                            continue
                        _ofb_m = str(_or_m.get("FOB_TYPE", "")).strip()
                        if _ofb_m in ["PILOT_RUN", "Salesman", "Salesman-PO", "Sample"]:
                            _ooff_m = 1
                        else:
                            _ooff_m = 4
                        _ordd_m = max(0, int(_orow_m.index[0]) - _ooff_m)
                        if _latest_tgt_idx_m is None or _ordd_m > _latest_tgt_idx_m:
                            _latest_tgt_idx_m = _ordd_m
                    except Exception:
                        continue


                def _sim_total_fit_multi(total_mc, qty_override=None):
                    if total_mc <= 0 or daily_capacity is None:
                        return False
                    _q = float(qty_override if qty_override is not None else _qty_for_machine_calc)
                    _wk = plan_week
                    _first = True
                    # ถ้า plan_week เลย latest target แล้ว (past RDD) → deadline พลาดไปแล้ว
                    # ห้ามตัด sim ที่ target (จะ break loop แรกทันทีจน _q ไม่ลด → คืน False เสมอ
                    # → GRADUAL BOOST เพิ่มเครื่องผิด ทั้งที่เครื่องเดิม carry ก็จบทัน)
                    # จำลองไปข้างหน้าแบบจำกัด 2 สัปดาห์แทน (สอดคล้อง _forward_sim:
                    # เครื่อง setup ใหม่ที่รัน < 2 สัปดาห์ไม่คุ้ม setup days)
                    _pw_idx_fit = week_index(plan_week)
                    _is_past_tgt_fit = (
                        _latest_tgt_idx_m is not None
                        and _pw_idx_fit is not None
                        and _pw_idx_fit > _latest_tgt_idx_m
                    )
                    _sim_cnt_fit = 0
                    while _wk is not None and _q > 0:
                        if _sim_cnt_fit > 300:  # SAFETY: กัน infinite loop ข้ามปลายปี
                            break
                        _wi = week_index(_wk)
                        if _wi is None:
                            break
                        # ยังไม่เลย target → ตัด sim ที่ target เหมือนเดิม
                        if not _is_past_tgt_fit and _latest_tgt_idx_m is not None and _wi > _latest_tgt_idx_m:
                            break
                        # เลย target แล้ว → จำลองได้สูงสุด 2 สัปดาห์
                        if _is_past_tgt_fit and _sim_cnt_fit >= 2:
                            break
                        # หักวันหยุดตามปฏิทินให้แล้วใน WorkDay.get_working_days() — ห้าม min ซ้ำ
                        _wd = get_working_days_by_factory(mc_group, total_mc, week=_wk, gauge=_sel_gauge)
                        if _first:
                            _c_cnt = min(total_mc, carryover_mc)
                            _n_cnt = max(0, total_mc - _c_cnt)
                            _c_prod = _c_cnt * _wd * daily_capacity
                            _n_prod = _n_cnt * max(0, _wd - item_setup_days) * daily_capacity
                            _first = False
                            _prod_wk = fold_round_mixed(
                                _c_prod + _n_prod, rev_weight, mc_group,
                                _wd * daily_capacity, _c_cnt,
                                max(0, _wd - item_setup_days) * daily_capacity, _n_cnt,
                            )
                        else:
                            _prod_wk = fold_round(
                                total_mc * _wd * daily_capacity, rev_weight, mc_group,
                                per_mc_cap=_wd * daily_capacity, mc_count=total_mc,
                            )
                        _q -= _prod_wk
                        _wk = next_week(_wk)
                        _sim_cnt_fit += 1
                    return _q <= 0
                # demand ที่เหลือจริง = qty_left ของ order นี้ + future orders
                # (แก้ _qty_for_machine_calc ที่ stale เพราะตั้งตอนเริ่ม order ด้วย original qty)
                _adjusted_qty_m = qty_left + max(0, _qty_for_machine_calc - _initial_qty_for_mc_calc)
                _cur_tot_m = carryover_mc + new_mc
                _min_tot_m = _cur_tot_m
                if not _s9_active:
                    for _tt in range(1, _cur_tot_m):
                        if _sim_total_fit_multi(_tt, qty_override=_adjusted_qty_m):
                            _min_tot_m = _tt
                            break
                if not _s9_active and _min_tot_m < _cur_tot_m:
                    _new_carry = min(carryover_mc, _min_tot_m)
                    _new_new = max(0, _min_tot_m - _new_carry)
                    print(
                        f"[MIN MC FIT] {item} W{plan_week}: ลดเครื่อง "
                        f"{_cur_tot_m}→{_min_tot_m} (carry {carryover_mc}→{_new_carry}, "
                        f"new {new_mc}→{_new_new}, latest_tgt_idx={_latest_tgt_idx_m}, "
                        f"total_demand={_qty_for_machine_calc:.0f})"
                    )
                    carryover_mc = _new_carry
                    new_mc = _new_new
                    available_machines = carryover_mc + new_mc
                    # recompute prod_days_new ตามจำนวนเครื่องใหม่
                    if new_mc > 0 and actual_working_days < item_setup_days:
                        prod_days_new = max(0, actual_working_days - 0.5)
                    elif new_mc > 0:
                        prod_days_new = max(0.5, actual_working_days - item_setup_days)
                    else:
                        prod_days_new = max(0, actual_working_days - item_setup_days)
                # ===== GRADUAL MACHINE BOOST =====
                # หลัง reduction: ถ้าเครื่องยังไม่พอจบทัน latest target
                # → เพิ่ม new_mc ได้อีก <= MAX_NEW_SETUP_MC ต่อ week (gradual build-up)
                if not _s9_active and _latest_tgt_idx_m is not None:
                    _post_red_mc = carryover_mc + new_mc
                    if not _sim_total_fit_multi(_post_red_mc, qty_override=_adjusted_qty_m):
                        _avail_pool = get_actual_mc_remain(
                            mc_group, plan_week, gauge=_sel_gauge, item_code=item
                        )
                        _boost_cap = min(
                            _post_red_mc + MAX_NEW_SETUP_MC,
                            _avail_pool,
                            _post_red_mc + _remaining_slots,
                        )
                        _boosted = False
                        _bt = _post_red_mc + 1
                        while _bt <= int(_boost_cap):
                            if _sim_total_fit_multi(_bt, qty_override=_adjusted_qty_m):
                                new_mc = _bt - carryover_mc
                                available_machines = carryover_mc + new_mc
                                _boosted = True
                                break
                            _bt += 1
                        if not _boosted and int(_boost_cap) > _post_red_mc:
                            new_mc = int(_boost_cap) - carryover_mc
                            available_machines = carryover_mc + new_mc
                            _boosted = True
                        if _boosted and available_machines > _post_red_mc:
                            print(
                                f"[GRADUAL BOOST] {item} W{plan_week}: "
                                f"{_post_red_mc}->{available_machines} MC "
                                f"(target W{_latest_tgt_idx_m})"
                            )
                            if new_mc > 0 and actual_working_days < item_setup_days:
                                prod_days_new = max(0, actual_working_days - 0.5)
                            elif new_mc > 0:
                                prod_days_new = max(0.5, actual_working_days - item_setup_days)
                            else:
                                prod_days_new = max(0, actual_working_days - item_setup_days)
            # Hard cap จริง: ห้าม new_mc เกิน MAX_NEW_SETUP_MC แม้หลัง boost
            if not _s9_active and new_mc > MAX_NEW_SETUP_MC:
                print(f"🔒 Post-boost hard cap: {item} W{plan_week} new_mc {new_mc} → {MAX_NEW_SETUP_MC}")
                new_mc = MAX_NEW_SETUP_MC
                available_machines = carryover_mc + new_mc
                if new_mc > 0 and actual_working_days < item_setup_days:
                    prod_days_new = max(0, actual_working_days - 0.5)
                elif new_mc > 0:
                    prod_days_new = max(0.5, actual_working_days - item_setup_days)

            # ===== Setup payback gate: เครื่องใหม่ต้องได้วิ่งเต็ม >= 2 สัปดาห์ =====
            # วางไว้ท้ายสุดของฝั่ง "เพิ่มเครื่อง" (หลัง GRADUAL BOOST + hard cap)
            # เพราะ boost สามารถเพิ่ม new_mc กลับมาได้ — โค้ดหลังจุดนี้มีแต่ลดเครื่อง
            if new_mc > 0 and not _s9_active and not rts_local_force:
                _pb_floor = max(0, _cyl_pending_cnt)  # เครื่อง cylinder change ที่ user อนุมัติ
                if carryover_mc <= 0:
                    _pb_floor = max(_pb_floor, 1)  # ไม่มี carry → ต้องเหลือเครื่องผลิตอย่างน้อย 1
                # demand ที่เหลือจริง = qty_left ของ FG นี้ + demand ของ FG อนาคต
                # (_qty_for_machine_calc ตั้งตอนเริ่ม order ด้วย original qty → stale)
                _pb_qty = qty_left + max(
                    0.0, _qty_for_machine_calc - _initial_qty_for_mc_calc
                )
                while new_mc > _pb_floor:
                    _pb_weeks = _new_mc_full_weeks(carryover_mc, new_mc, _pb_qty)
                    if _pb_weeks >= 2:
                        break
                    print(
                        f"   [SETUP PAYBACK] {item} W{plan_week}: new {new_mc}→{new_mc - 1} "
                        f"(เครื่องใหม่ได้วิ่งเต็มแค่ {_pb_weeks} สัปดาห์ < 2 → setup ไม่คุ้ม, "
                        f"carry={carryover_mc}, qty={_pb_qty:.0f})"
                    )
                    new_mc -= 1
                available_machines = carryover_mc + new_mc
                if new_mc > 0 and actual_working_days < item_setup_days:
                    prod_days_new = max(0, actual_working_days - 0.5)
                elif new_mc > 0:
                    prod_days_new = max(0.5, actual_working_days - item_setup_days)
                else:
                    prod_days_new = max(0, actual_working_days - item_setup_days)

            # ===== Optimize: ลดเครื่องให้น้อยสุดที่ยังผลิตพอครอบคลุม qty_left =====
            # เช่น week15 carry=3 แต่ qty_left น้อย → ใช้แค่ 1 เครื่องก็เสร็จใน week นี้
            # ใช้การจำลองผลิตจริง (รวม rev_weight rounding) เพื่อความแม่นยำ
            def _sim_produce(c_mc, n_mc):
                if daily_capacity is None:
                    return 0
                c_cap = daily_capacity * prod_days_old * c_mc
                n_cap = daily_capacity * prod_days_new * n_mc
                total_cap = c_cap + n_cap
                if rev_weight and rev_weight > 0 and total_cap > 0:
                    return fold_round_mixed(
                        total_cap, rev_weight, mc_group,
                        daily_capacity * prod_days_old, c_mc,
                        daily_capacity * prod_days_new, n_mc,
                    )
                return total_cap

            if carryover_mc + new_mc > 0 and _sim_produce(carryover_mc, new_mc) > qty_left:
                # ใช้ demand ระดับ item เพื่อไม่ลดเครื่องถ้ายังมี FG ของ item เดียวกันรอผลิต
                # 🔧 FIX: คำนวณ demand ที่ยังไม่ได้วางแผน = total demand - ที่ plan ไปแล้ว - qty_left ปัจจุบัน
                # ครอบคลุมทั้ง 2 กรณี:
                #   1. คนละ SC แต่ item เดียวกัน (SC อื่น plan ไปแล้ว → ไม่ hold เครื่อง)
                #   2. SC เดียวกัน คนละ FG (FG ก่อนหน้า plan ไปแล้ว → ไม่ hold เครื่อง)
                # _qty_for_machine_calc ครอบเฉพาะ FG ที่ยังไม่ plan อยู่แล้ว
                # หักแค่ order_qty ของ FG ปัจจุบัน = demand ของ FG ถัดไป
                next_fg_qty = max(0, _qty_for_machine_calc - order_qty)

                # ถ้ามี FG ถัดไป ให้คำนึงถึง ORDER_QTY รวมของ item ในการตัดสินใจลดเครื่อง
                total_qty_to_consider = _qty_for_machine_calc  # total ของ item (current + future FG)
                opt_carry, opt_new = carryover_mc, new_mc  # fallback = ไม่ลด

                # ถ้ามี FG ถัดไปใน SC เดียวกัน → ไม่ลดเครื่อง เพื่อ carry ต่อในสัปดาห์เดียวกัน/สัปดาห์ถัดไป
                # 🔧 FIX: ใช้ threshold > 1.0 แทน > 0 เพื่อป้องกัน floating-point residual
                # จากการหัก qty ที่มีทศนิยม (เช่น 2530 - 687.89 - 1841.77 = 0.34)
                if next_fg_qty > 1.0:
                    # 🔧 NEW: หา min mc ที่ยังผลิต total item demand ทัน latest target
                    # ของ FG ที่เหลือทั้งหมด เพื่อ free เครื่องให้ item อื่น
                    _cur_fg_int_opt = 0
                    try:
                        _cur_fg_int_opt = int(fg_week or 0)
                    except Exception:
                        pass
                    _latest_tgt_idx_opt = rdd_idx if rdd_idx is not None else current_week_idx
                    for _o_idx_c, _o_row_c in orders_sorted.iterrows():
                        if str(_o_row_c.get("Item Code", "")).strip() != str(item).strip():
                            continue
                        _o_fg_raw_c = _o_row_c.get("FG Week")
                        if not pd.notna(_o_fg_raw_c):
                            continue
                        try:
                            _o_fg_i_c = int(_o_fg_raw_c)
                        except Exception:
                            continue
                        if _o_fg_i_c < _cur_fg_int_opt:
                            continue
                        _ofs_c = str(_o_fg_i_c)
                        try:
                            if len(_ofs_c) >= 5:
                                _oy_c = int(_ofs_c[:4])
                                _owk_c = int(_ofs_c[4:])
                            elif len(_ofs_c) <= 2:
                                _oy_c = TODAY.year
                                _owk_c = int(_ofs_c)
                            else:
                                _oy_c = TODAY.year
                                _owk_c = int(_ofs_c)
                            _orow_c = calendar_week[
                                (calendar_week["YEAR"] == _oy_c)
                                & (calendar_week["WEEK"] == _owk_c)
                            ]
                            if _orow_c.empty:
                                continue
                            _ofb_c = str(_o_row_c.get("FOB_TYPE", "")).strip()
                            if _ofb_c in ["PILOT_RUN", "Salesman", "Salesman-PO", "Sample"]:
                                _ooff_c = 1
                            else:
                                _ooff_c = 4
                            _o_rdd_c = max(0, int(_orow_c.index[0]) - _ooff_c)
                            if _latest_tgt_idx_opt is None or _o_rdd_c > _latest_tgt_idx_opt:
                                _latest_tgt_idx_opt = _o_rdd_c
                        except Exception:
                            continue

                    def _sim_total_fit(total_mc):
                        if total_mc <= 0 or daily_capacity is None:
                            return False
                        _q = float(_qty_for_machine_calc)
                        _wk = plan_week
                        _first = True
                        _stf_guard = 0
                        while _wk is not None and _q > 0:
                            _stf_guard += 1  # SAFETY: กัน infinite loop ข้ามปลายปี
                            if _stf_guard > 300:
                                break
                            _wi = week_index(_wk)
                            if _wi is None:
                                break
                            if _latest_tgt_idx_opt is not None and _wi > _latest_tgt_idx_opt:
                                break
                            # หักวันหยุดตามปฏิทินให้แล้วใน WorkDay.get_working_days() — ห้าม min ซ้ำ
                            _wd = get_working_days_by_factory(mc_group, total_mc, week=_wk, gauge=_sel_gauge)
                            if _first:
                                _c_cnt = min(total_mc, carryover_mc)
                                _n_cnt = max(0, total_mc - _c_cnt)
                                _c_prod = _c_cnt * _wd * daily_capacity
                                _n_prod = _n_cnt * max(0, _wd - item_setup_days) * daily_capacity
                                _first = False
                                _prod_wk = fold_round_mixed(
                                    _c_prod + _n_prod, rev_weight, mc_group,
                                    _wd * daily_capacity, _c_cnt,
                                    max(0, _wd - item_setup_days) * daily_capacity, _n_cnt,
                                )
                            else:
                                _prod_wk = fold_round(
                                    total_mc * _wd * daily_capacity, rev_weight, mc_group,
                                    per_mc_cap=_wd * daily_capacity, mc_count=total_mc,
                                )
                            _q -= _prod_wk
                            _wk = next_week(_wk)
                        return _q <= 0

                    _current_total_opt = carryover_mc + new_mc
                    _min_total_opt = _current_total_opt
                    for _try_t in range(1, _current_total_opt):
                        if _sim_total_fit(_try_t):
                            _min_total_opt = _try_t
                            break

                    if _min_total_opt < _current_total_opt:
                        opt_carry = min(carryover_mc, _min_total_opt)
                        opt_new = max(0, _min_total_opt - opt_carry)
                        print(
                            f"[MIN MC FIT] {item} W{plan_week}: ลดเครื่อง "
                            f"{_current_total_opt}→{_min_total_opt} "
                            f"(latest_tgt_idx={_latest_tgt_idx_opt}, demand={_qty_for_machine_calc:.0f})"
                        )
                    else:
                        opt_carry = carryover_mc
                        opt_new = new_mc
                else:
                    # ไม่มี FG ถัดไป หรือ capacity เพียงพอสำหรับทั้ง FG ปัจจุบันและถัดไป
                    # ให้ลดเครื่องตามปกติ
                    # ขั้นที่ 1: ลอง carry-only (ไม่ต้อง setup เพิ่ม) หาน้อยสุดที่ produce ≥ qty_left
                    found = False
                    for try_c in range(1, carryover_mc + 1):
                        if _sim_produce(try_c, 0) >= qty_left:
                            opt_carry = try_c
                            opt_new = 0
                            found = True
                            break
                    if not found and new_mc > 0:
                        # ขั้นที่ 2: ต้องมี new ด้วย → ลด new ให้น้อยสุด
                        # ถ้า carryover_mc=0 ต้องเริ่มจาก 1 เพราะต้องมีเครื่องอย่างน้อย 1 เครื่อง
                        _try_n_start = 0 if carryover_mc > 0 else 1
                        for try_n in range(_try_n_start, new_mc + 1):
                            if _sim_produce(carryover_mc, try_n) >= qty_left:
                                opt_carry = carryover_mc
                                opt_new = try_n
                                break

                if opt_carry + opt_new < available_machines:
                    carryover_mc = opt_carry
                    new_mc = opt_new
                    available_machines = opt_carry + opt_new

            # ===== Hard-cap: enforce job cap ก่อนคำนวณ produce เสมอ =====
            # S9: ไม่หัก job cap และไม่มี limit เครื่อง (ใช้เต็ม pool ที่มี)
            if not _s9_active:
                _type_used_now = get_type_used_jobs(plan_week, mc_group)
                # carryover ไม่ควรนับเป็น new setup jobs เสมอ
                _committed_carryover = carryover_mc
                # check_job_capacity_limit รับ "จำนวนเครื่องรวม" (carryover+new) แล้วคืนจำนวนรวมที่อนุญาต
                # ต้องส่ง carryover_mc + new_mc ไม่ใช่ new_mc เดี่ยว ไม่งั้น carryover จะถูกลบซ้ำ → cap ไม่ทำงาน
                _total_allowed = check_job_capacity_limit(mc_group, carryover_mc + new_mc, False, _type_used_now, committed_carryover=_committed_carryover)
                _allowed_new = max(0, _total_allowed - carryover_mc)
                if _allowed_new < new_mc:
                    new_mc = _allowed_new
                available_machines = carryover_mc + new_mc
            if new_mc == 0 and carryover_mc == 0:
                # ไม่มีเครื่องเลย ข้ามไป week ถัดไป
                if _s9_active:
                    _s9_no_cap_weeks.add(plan_week)  # บันทึก week ที่ S9 ลองแล้วได้ 0 เครื่อง
                plan_week = next_week(plan_week)
                continue
            # YD-ORDERS: ล็อกจำนวนเครื่องภายใน week เดียวกัน
            # SO แรกกำหนดจำนวนเครื่อง → SO ถัดไปในสัปดาห์เดียวกันใช้เท่าเดิม
            _yd_lock_key = (item, mc_group, plan_week)
            if order_type == "YD-ORDERS" and _yd_lock_key in _yd_week_locked_mc:
                _locked = _yd_week_locked_mc[_yd_lock_key]
                available_machines = _locked[0]
                carryover_mc = _locked[1]
                new_mc = 0  # เครื่อง setup ไปแล้วจาก SO แรก ไม่ต้อง setup ซ้ำ
                print(f"[YD LOCK] {item} W{plan_week} SO={sc_so_no}: ใช้เครื่องเท่าเดิม actual={available_machines} carry={carryover_mc}")
            # 🔧 FIX: ถ้า qty_left < rev_weight × จำนวนเครื่อง → ลดเครื่องลง
            # เพราะไม่สามารถแบ่ง qty น้อยกว่า rev_weight ลงหลายเครื่องได้
            _already_planned_rev = _item_cumulative_planned.get(item, 0)
            _next_fg_qty_rev = max(0, _qty_for_machine_calc - order_qty)
            if (
                rev_weight and rev_weight > 0
                and available_machines > 1
                and qty_left < rev_weight * available_machines
                and _next_fg_qty_rev <= 1.0
            ):
                _max_mc_for_qty = max(1, int(qty_left // rev_weight))
                if _max_mc_for_qty < available_machines:
                    print(f"[REV_WEIGHT CAP] {item} W{plan_week}: qty_left={qty_left:.0f} < rev_weight({rev_weight})×mc({available_machines})={rev_weight*available_machines:.0f} → ลดเครื่องจาก {available_machines} เป็น {_max_mc_for_qty}")
                    # ลด available_machines (total) โดยตรง แล้วคำนวณ new_mc/carryover_mc ใหม่
                    available_machines = _max_mc_for_qty
                    # คำนวณ new_mc และ carryover_mc ใหม่จาก available_machines ที่ลดลงแล้ว
                    # พยายามรักษา carryover_mc ให้มากที่สุด (ลด new_mc ก่อน)
                    new_mc = max(0, min(new_mc, available_machines - carryover_mc))
                    carryover_mc = available_machines - new_mc

            # setup_days_used สำหรับ log — เครื่องใหม่ต้อง setup เสมอ (คูณจำนวนเครื่องใหม่)
            setup_days_used = item_setup_days * new_mc if new_mc > 0 else 0
            # available_days สำหรับ log (ใช้เครื่องใหม่เป็นหลักถ้ามี)
            available_days = prod_days_new if new_mc > 0 else prod_days_old

            # === Same-week remaining cap: item+machine เดียวกัน ต่อจาก FG ก่อนหน้า ===
            # ใช้ remaining_week_cap ตาม item+machine ในสัปดาห์เดียวกัน (ไม่ผูก SC)
            # S9: track cap แยกจาก normal โดยใช้ COMKN เป็น key
            _same_week_key = (plan_week, item, "COMKN" if _s9_active else mc_group)
            _same_week_rem_cap = remaining_week_cap.get(_same_week_key, None)
            _same_week_total_cap = None
            # DEBUG: FD3BASPK34B0 Week 26-27
            if item == "FD3BASPK34B0" and plan_week in [26, 27]:
                print(f"[DEBUG W{plan_week}] _same_week_rem_cap = {_same_week_rem_cap}, carryover_mc = {carryover_mc}, new_mc = {new_mc}, prod_days_old = {prod_days_old}, daily_capacity = {daily_capacity}, qty_left = {qty_left}")
            # 🔧 FIX: remaining_week_cap override logic
            # - ถ้า remaining มาจาก new plan (owner != None): remaining คำนวณถูกต้องจาก production จริง
            #   ใน week เดียวกัน → ต้องเคารพ (ห้าม override) เพราะเครื่องที่ setup ไปแล้ว
            #   ไม่มี productive days เหลือใน week นั้น
            # - ถ้า remaining มาจาก booking (owner = None): booking production เกิดขึ้นจริงแล้ว
            #   ต้องเคารพ remaining ที่เหลือเช่นกัน
            _rem_cap_owner = remaining_week_cap_owner.get(_same_week_key, None)
            _rem_from_new_plan = _rem_cap_owner is not None  # owner=None → จาก booking seed
            # คำนวณ PRODUCE_QTY ตามสูตรที่แม่นยำ
            if _same_week_rem_cap is not None:
                # ผลิตด้วย cap ที่เหลือจาก FG ก่อนหน้า + cap incremental จากเครื่องใหม่ในสัปดาห์เดียวกัน
                _same_week_base_cap = max(0.0, float(_same_week_rem_cap))
                # YD-ORDERS: ถ้า SUB_COLOR เปลี่ยน → หัก 1 วัน × เครื่อง carry × daily_cap จาก remaining cap
                if _yd_sub_color_setup > 0 and carryover_mc > 0 and daily_capacity:
                    _color_cap_deduct = carryover_mc * 1 * daily_capacity  # หัก 1 วันเต็ม ไม่ round ลง rev_weight
                    _same_week_base_cap = max(0.0, _same_week_base_cap - _color_cap_deduct)
                    _total_cs = _yd_week_color_setups.get((mc_key, plan_week), 0)
                    print(f"[YD SUB_COLOR SAME-WEEK] {item} W{plan_week}: หัก {_color_cap_deduct:.0f} units จาก remaining cap (เปลี่ยนสีครั้งที่ {_total_cs}, remaining={_same_week_base_cap:.0f})")
                _same_week_new_cap = max(0.0, float(daily_capacity * prod_days_new * new_mc))
                # FIX: ห้ามเพิ่ม carryover capacity เมื่อมี remaining_week_cap
                # เพราะ remaining_week_cap คำนวณจากวันทำงานของสัปดาห์ปัจจุบันแล้ว
                # การเพิ่ม carryover capacity จะทำให้นับซ้ำ (double counting)
                _same_week_carry_cap = 0.0
                # ไม่เพิ่ม carry capacity เมื่อมี remaining cap อยู่แล้ว
                _same_week_total_cap = _same_week_base_cap + _same_week_new_cap + _same_week_carry_cap

                if rev_weight and rev_weight > 0:
                    _swt_mc = max(1, int(available_machines or 1))
                    produce = min(
                        qty_left,
                        fold_round(
                            _same_week_total_cap, rev_weight, mc_group,
                            per_mc_cap=_same_week_total_cap / _swt_mc, mc_count=_swt_mc,
                        ),
                    )
                else:
                    produce = min(qty_left, _same_week_total_cap)
            elif rev_weight is not None and rev_weight > 0:
                cap_old = daily_capacity * prod_days_old * carryover_mc
                cap_new = daily_capacity * prod_days_new * new_mc
                produce = min(
                    qty_left,
                    fold_round_mixed(
                        cap_old + cap_new, rev_weight, mc_group,
                        daily_capacity * prod_days_old, carryover_mc,
                        daily_capacity * prod_days_new, new_mc,
                    ),
                )
            else:
                cap_old = daily_capacity * prod_days_old * carryover_mc
                cap_new = daily_capacity * prod_days_new * new_mc
                produce = min(qty_left, cap_old + cap_new)

            # Hard clamp: ห้ามยอดรวม item+week+mc_group เกิน capacity ของสัปดาห์นี้
            # ป้องกันกรณี merge ข้าม FG แล้วถูกบวก produce ซ้ำจนเกินเพดาน
            _already_planned_qty = 0.0
            for _p in plans:
                if (
                    _p.get("ITEM_CODE") == item
                    and _p.get("PLAN_WEEK") == plan_week
                    and _p.get("MC_GROUP") == ("COMKN" if _s9_active else mc_group)
                ):
                    _already_planned_qty += float(_p.get("PRODUCE_QTY", 0) or 0)

            if _same_week_rem_cap is not None:
                _max_additional_qty = max(0.0, float(_same_week_total_cap or 0.0))
            else:
                # 🔧 FIX: ตรวจสอบว่า booking ในสัปดาห์เดียวกันมีการ setup เพิ่มเครื่องหรือไม่
                # ถ้ามี → true_week_cap ต้องหัก setup days จากเครื่องใหม่ใน booking ออกก่อน
                # ป้องกัน new order ผลิตเกินกว่าที่เครื่องสามารถทำได้จริงใน week ที่ old booking ต้อง setup
                _bk_cur_w_idx = week_index(plan_week)
                _bk_prev_w_idx_cap = (_bk_cur_w_idx - 1) if _bk_cur_w_idx is not None else None
                _bk_mc_cur = booking_mc_by_week.get(mc_key, {}).get(_bk_cur_w_idx, 0)
                _bk_mc_prev = (
                    booking_mc_by_week.get(mc_key, {}).get(_bk_prev_w_idx_cap, 0)
                    if _bk_prev_w_idx_cap is not None
                    else 0
                )
                _bk_new_mc_cnt = max(0, _bk_mc_cur - _bk_mc_prev)   # เครื่องใหม่ที่ setup ใน booking week นี้
                _bk_carry_mc_cnt = _bk_mc_cur - _bk_new_mc_cnt        # เครื่องที่ carry-over จาก booking week ก่อนหน้า
                if _bk_new_mc_cnt > 0 and daily_capacity is not None and daily_capacity > 0:
                    # คำนวณ true_week_cap โดยหัก setup days จากเครื่องใหม่ใน old booking
                    _true_week_cap = (
                        _bk_carry_mc_cnt * actual_working_days
                        + _bk_new_mc_cnt * max(0, actual_working_days - item_setup_days)
                    ) * daily_capacity
                    _week_capacity = max(0.0, _true_week_cap)
                    print(
                        f"[OLD SETUP CLAMP] {item} W{plan_week} {mc_group}: "
                        f"bk_new_mc={_bk_new_mc_cnt:.0f} bk_carry_mc={_bk_carry_mc_cnt:.0f} "
                        f"→ true_cap={_week_capacity:.2f} (naive={float(cap_old + cap_new):.2f})"
                    )
                else:
                    _week_capacity = max(0.0, float(cap_old + cap_new))
                _max_additional_qty = max(0.0, _week_capacity - _already_planned_qty)
            if produce > _max_additional_qty:
                print(
                    f"[CAP CLAMP] {item} W{plan_week} {mc_group}: "
                    f"produce {produce:.2f} -> {_max_additional_qty:.2f} "
                    f"(already={_already_planned_qty:.2f})"
                )
                _clamped_qty = max(0.0, float(_max_additional_qty))
                if rev_weight and rev_weight > 0:
                    _clamp_mc = max(1, int(available_machines or 1))
                    _rounded_cap = fold_round(
                        _clamped_qty, rev_weight, mc_group,
                        per_mc_cap=_clamped_qty / _clamp_mc, mc_count=_clamp_mc,
                    )
                    if _rounded_cap > 0:
                        _clamped_qty = _rounded_cap
                    elif _clamped_qty < rev_weight:
                        _clamped_qty = 0
                produce = _clamped_qty

            # ไม่เพิ่มแถวถ้าไม่มีการผลิต — แต่ถ้า setup ไปแล้ว (new_mc>0) ให้บันทึกว่าเครื่อง setup เสร็จ
            # เพื่อให้ week ถัดไปนับเป็น carryover (ไม่ต้อง setup ซ้ำ) → ผลิตได้เต็มจำนวนวัน
            if _s9_active and item == "FD5GNTMX102/02A0":
                print(f"[S9 DEBUG] {item} W{plan_week}: produce={produce:.1f}, avail={available_machines}, cap={daily_capacity}, prod_days_new={prod_days_new}, new_mc={new_mc}, carryover_mc={carryover_mc}, qty_left={qty_left:.1f}")
            if produce <= 0:
                if _s9_active:
                    print(f"[S9 FULL] {item} W{plan_week} MC={mc_group}: S9 pool เต็ม (produce=0) → next week")
                    plan_week = next_week(plan_week)
                    if plan_week is None:
                        break
                    continue
                if new_mc > 0:
                    _plan_ck_setup = _resolve_carry_key(item, mc_group, _sel_gauge)
                    machines_in_use[_plan_ck_setup] = available_machines
                    last_sc_machines[(item, mc_group, _sel_gauge, sc_so_no)] = available_machines
                    last_sc_week[(item, mc_group, _sel_gauge, sc_so_no)] = week_index(plan_week)
                    # 🔧 FIX: SETUP ONLY (ไม่มีการผลิต) ควรบันทึก last_production
                    # เพื่อให้ week ถัดไปรู้ว่าเครื่อง setup แล้ว (ไม่ต้อง setup ซ้ำ)
                    # แต่ต้องบันทึกเฉพาะกรณีที่ week นี้ >= week ที่บันทึกไว้
                    _cur_lp = last_production.get(_plan_ck_setup)
                    if _cur_lp is None or week_index(plan_week) >= _cur_lp:
                        last_production[_plan_ck_setup] = week_index(plan_week)
                    new_plan_started_items.add(_plan_ck_setup)
                    # บันทึก job usage สำหรับ week นี้ (เครื่อง setup แล้วแม้ยังไม่ผลิต)
                    if plan_week not in weekly_job_usage:
                        weekly_job_usage[plan_week] = {}
                    weekly_job_usage[plan_week][mc_group] = (
                        weekly_job_usage[plan_week].get(mc_group, 0) + new_mc
                    )
                    _week_mc_key_setup = (week_index(plan_week), mc_group)
                    weekly_mc_usage[_week_mc_key_setup] = weekly_mc_usage.get(_week_mc_key_setup, 0) + available_machines
                    last_sc_so_no[_plan_ck_setup] = sc_so_no
                    if order_type == "YD-ORDERS" and sub_color:
                        last_sub_color[_plan_ck_setup] = sub_color
                    # 🔧 FIX: บันทึก SETUP ONLY ลง _existing_item_week_mc
                    # เพื่อป้องกัน SETUP EARLY PRIORITY ดึงกลับมา week เดิมซ้ำ
                    # ซึ่งทำให้ double-count วันทำงาน (setup กิน 5 วัน แล้วมาผลิตอีก 5 วัน)
                    _existing_item_week_mc.add((item, plan_week, "COMKN" if _s9_active else mc_group))
                    print(f"[SETUP ONLY] {item} W{plan_week} MC={mc_group}: setup {new_mc} machines (0 produce, setup ate all {actual_working_days} working days)")

                plan_week = next_week(plan_week)
                if plan_week is None:
                    break
                continue

            # จำนวนเครื่องที่วางแผนไว้ (จาก calculate_required_machines)
            # แสดงจำนวนเครื่องจริงตาม required_mc หรือ available_machines
            prev_week_mc = machines_in_use.get(
                _resolve_carry_key(item, mc_group, _sel_gauge), available_machines
            )
            planned_mc = (
                required_machines_info[2]
                if required_machines_info
                else available_machines
            )
            # 🔧 FIX: ตรวจสอบว่า Item+Week+MC_GROUP นี้มีการ plan ไปแล้วหรือไม่
            # เพื่อป้องกันการซ้ำซ้อนของ orders ที่มี Item+MC_GROUP เดียวกันใน week เดียวกัน
            # โดยไม่สนใจว่าจะเป็น SC หรือ TARGET_KNIT อะไร (เพราะอาจมีการ merge SC)
            # S9: ใช้ COMKN เป็น key เพื่อไม่ merge เข้า normal SKP row
            _current_item_week_mc = (item, plan_week, "COMKN" if _s9_active else mc_group)
            _is_fg_split = False  # Flag: สร้าง row ใหม่สำหรับ FG_WEEK ที่ต่างกัน (ไม่ merge)

            # 🔧 FIX: Skip weeks where OLD bookings already exist for same item/MC/Gauge
            # to prevent exceeding capacity that old bookings already use
            _ck_key = _ck(item, mc_group, _sel_gauge)
            _pw_idx = week_index(plan_week)
            if not _s9_active and _ck_key and _pw_idx is not None and _ck_key in booking_mc_by_week:
                if booking_mc_by_week[_ck_key].get(_pw_idx, 0) > 0:
                    # ถ้า cylinder change เพิ่งทำสำหรับ item นี้ใน plan_week นี้ → ไม่ skip
                    # เพราะ cylinder เพิ่มเครื่องใหม่ที่ไม่ใช่ old booking เดิม
                    _cyl_done_key = (int(plan_week) - 1, item.strip().upper(), mc_group.strip().upper() if mc_group else "")
                    # ตรวจว่า new plan มี carry machine จาก week ก่อนหน้าติดกัน (gap=1)
                    # ถ้าใช่ → ไม่ skip เพราะเครื่องนั้นเป็นคนละชุดกับ old booking (setup ต่างรอบ)
                    # เช่น W23 setup ใหม่ 1 เครื่อง → W24 carry ต่อ แม้ old booking W24 เป็น SETUP (N เครื่อง)
                    _carry_ck_skip = _resolve_carry_key(item, mc_group, _sel_gauge)
                    _prev_lp_skip = last_production.get(_carry_ck_skip)
                    _is_new_plan_carry = (
                        _carry_ck_skip in new_plan_started_items
                        and _prev_lp_skip is not None
                        and _pw_idx - _prev_lp_skip == 1
                        and machines_in_use.get(_carry_ck_skip, 0) > 0
                    )
                    if _cyl_done_key not in _cylinder_change_for_item and not _is_new_plan_carry:
                        print(f"[SKIP OLD BOOKING IN MAIN LOOP] {item}+{plan_week}+{mc_group}: OLD booking using {booking_mc_by_week[_ck_key].get(_pw_idx, 0)} machines, skip to next week")
                        plan_week = next_week(plan_week)
                        if plan_week is None:
                            break
                        continue
                    if _is_new_plan_carry:
                        # รวม carry จาก new plan (W23) + old booking (W24) เป็น machines_in_use เดียวกัน
                        # เพื่อให้ planning loop เห็น cap รวม = N+1 เครื่อง แทนที่จะเห็นแค่ 1
                        # ⚠️ ต้องหักเครื่อง booking ของสัปดาห์ก่อนออกจาก carry ก่อน มิฉะนั้นเครื่อง booking
                        # ตัวเดิมจะถูกบวกซ้ำทุกสัปดาห์ (เครื่องโตขึ้นเรื่อยๆ ทั้งที่เป็นเครื่องตัวเดียวกัน)
                        _new_carry_mc = machines_in_use.get(_carry_ck_skip, 0)
                        _combined_carry, _skip_own_mc, _skip_bk_prev, _old_bk_mc_here = _combine_carry_with_booking(
                            _ck_key, _new_carry_mc, _prev_lp_skip, _pw_idx
                        )
                        if _old_bk_mc_here > 0:
                            machines_in_use[_carry_ck_skip] = _combined_carry
                        print(f"[CARRY THROUGH OLD BOOKING] {item}+{plan_week}+{mc_group}: plan-own={_skip_own_mc} (carry {_new_carry_mc} − booking W ก่อน {_skip_bk_prev}) + old booking W{plan_week}={_old_bk_mc_here} → combined={_combined_carry} mc")
            # YD-ORDERS: ห้าม merge — แต่ละ SO ต้องแยกเป็น row ของตัวเอง
            if order_type == "YD-ORDERS":
                _is_fg_split = True  # บังคับสร้าง row ใหม่เสมอ
            elif _current_item_week_mc in _existing_item_week_mc:
                # Merge ลงแถวเดิมของ Item+Week+MC_GROUP เฉพาะ FG_WEEK + SC เดียวกันเท่านั้น
                # ⚠️ ห้าม merge ข้าม SC — มิฉะนั้น production ของ SC อื่นจะถูกนับรวม
                # ทำให้ PRODUCE_QTY ของแถวเกิน ORDER_QTY ของ SC นั้น
                _merged = False
                for _p in plans:
                    if (
                        _p.get("ITEM_CODE") == item
                        and _p.get("PLAN_WEEK") == plan_week
                        and _p.get("MC_GROUP") == ("COMKN" if _s9_active else mc_group)
                        and _p.get("FG_WEEK") == fg_week
                        and _p.get("SC_SO_NO") == sc_so_no

                    ):
                        _p["PRODUCE_QTY"] = float(_p.get("PRODUCE_QTY", 0) or 0) + float(produce)
                        _p["PLAN_QTY"] = max(0, float(qty_left - produce))
                        _merged = True
                        break

                if _merged:
                    print(
                        f"[DEBUG DUPLICATE] MERGED: {item}+{plan_week}+{mc_group} FG={fg_week} +{produce:.1f}"
                    )


                    qty_left -= produce
                    if qty_left <= 0:
                        qty_left = 0

                    # บันทึก week ที่ผลิตจริงก่อนเปลี่ยน plan_week
                    _produced_week = plan_week

                    # อัปเดต remaining cap แบบเดียวกับเส้นทาง append ใหม่
                    if _same_week_rem_cap is not None:
                        _new_rem = max(0, float(_same_week_total_cap or 0.0) - produce)
                        if _new_rem > 0:
                            remaining_week_cap[(_produced_week, item, mc_group)] = _new_rem
                            remaining_week_cap_owner[(_produced_week, item, mc_group)] = sc_so_no
                        else:
                            remaining_week_cap.pop((_produced_week, item, mc_group), None)
                            remaining_week_cap_owner.pop((_produced_week, item, mc_group), None)
                    else:
                        _full_week_cap = cap_old + cap_new
                        _rem = max(0, _full_week_cap - produce)
                        remaining_week_cap[(_produced_week, item, mc_group)] = _rem
                        remaining_week_cap_owner[(_produced_week, item, mc_group)] = sc_so_no

                    _plan_ck_merge = _resolve_carry_key(item, mc_group, _sel_gauge)
                    last_production[_plan_ck_merge] = week_index(plan_week)
                    machines_in_use[_plan_ck_merge] = available_machines
                    last_sc_machines[(item, mc_group, _sel_gauge, sc_so_no)] = available_machines
                    last_sc_week[(item, mc_group, _sel_gauge, sc_so_no)] = week_index(plan_week)
                    # FIX: ถ้า MERGE iteration นำเครื่องใหม่เข้ามา (new_mc > 0)
                    # ต้องอัปเดต row + tracking ให้สะท้อนจำนวนเครื่องจริง
                    # เช่น cross-SC fill สร้าง row carry=2,new=0 → MERGE เพิ่ม new=2 →
                    # row ต้องแสดง actual=4 เพื่อไม่ให้ overcapacity ในผลลัพธ์
                    _existing_actual_mc = int(_p.get("ACTUAL_MC", 0) or 0)
                    _existing_new_mc = int(_p.get("NEW_MC", 0) or 0)
                    if new_mc > 0 and available_machines > _existing_actual_mc:
                        _merged_actual = available_machines
                        _merged_new = _existing_new_mc + new_mc
                        _p["ACTUAL_MC"] = _merged_actual
                        _p["NEW_MC"] = _merged_new
                        _p["REQUIRED_MC"] = _merged_actual
                        # สะสม SETUP_DAYS ตาม NEW_MC ที่เพิ่มเข้ามา
                        _p["SETUP_DAYS"] = int(_p.get("SETUP_DAYS", 0) or 0) + setup_days_used
                        _p["AVAILABLE_DAYS"] = available_days
                        # อัพเดต tracking สำหรับเครื่องใหม่ที่ MERGE เพิ่มเข้ามา
                        _week_mc_key_m = (week_index(plan_week), mc_group)
                        weekly_mc_usage[_week_mc_key_m] = weekly_mc_usage.get(_week_mc_key_m, 0) + new_mc
                        if plan_week not in weekly_job_usage:
                            weekly_job_usage[plan_week] = {}
                        weekly_job_usage[plan_week][mc_group] = (
                            weekly_job_usage[plan_week].get(mc_group, 0) + new_mc
                        )
                        print(
                            f"[MERGE MC UPDATE] {item} W{plan_week}: "
                            f"actual {_existing_actual_mc}→{_merged_actual} "
                            f"(carry={_p.get('CARRYOVER_MC')}, new={_merged_new})"
                        )
                    # FIX: ถ้า new_mc = 0 (ใช้ remaining cap) ไม่ต้องอัปเดต machines_in_use
                    # เพราะถูกตั้งค่าไว้อย่างถูกต้องแล้วที่ line 5492 = available_machines
                    # ห้าม overwrite ด้วย _existing_actual_mc เพราะจะทำให้ week ถัดไปคิดว่าไม่มี carryover
                    else:
                        # machines_in_use already set correctly at line 5492
                        # Only update SC-specific tracking if needed
                        if _existing_actual_mc > 0:
                            last_sc_machines[(item, mc_group, _sel_gauge, sc_so_no)] = _existing_actual_mc
                        last_sc_week[(item, mc_group, _sel_gauge, sc_so_no)] = week_index(plan_week)
                    # FIX: ถ้า available_machines > 0 แต่ _existing_actual_mc = 0
                    # หมายถึง MERGE ครั้งแรกที่ใช้ remaining cap จาก booking → ต้องบันทึก available_machines
                    if available_machines > 0 and _existing_actual_mc == 0:
                        machines_in_use[_plan_ck_merge] = available_machines
                        last_sc_machines[(item, mc_group, _sel_gauge, sc_so_no)] = available_machines
                        last_sc_week[(item, mc_group, _sel_gauge, sc_so_no)] = week_index(plan_week)
                        print(
                            f"[MERGE FIRST TIME] {item} W{plan_week}: "
                            f"set machines_in_use={available_machines} (from booking remaining cap)"
                        )
                    last_sc_so_no[_plan_ck_merge] = sc_so_no  # บันทึก SC/SO NO ล่าสุดที่ผลิต
                    if order_type == "YD-ORDERS" and sub_color:
                        last_sub_color[_plan_ck_merge] = sub_color
                    new_plan_started_items.add(_plan_ck_merge)
                    # บันทึก total machines per (week, mc_group) สำหรับ gradual increase
                    # 🔧 FG SPLIT: ถ้าเป็น FG split row → เครื่องนับไปแล้วจาก FG แรก ห้ามนับซ้ำ
                    if not _is_fg_split:
                        _week_mc_key = (week_index(plan_week), mc_group)
                        weekly_mc_usage[_week_mc_key] = weekly_mc_usage.get(_week_mc_key, 0) + available_machines
                    # อัพเดท job usage สำหรับสัปดาห์นี้ (นับเฉพาะ new_mc = machines ที่ setup ใหม่)
                    # เครื่อง carry ต่อ (carryover_mc) ไม่ต้องนับเป็น job ใหม่เพราะไม่ต้อง setup
                    # ⚠️ ต้องหักนอกเงื่อนไข _is_fg_split — FG split / YD-ORDERS ก็ตั้งเครื่องใหม่จริง
                    # (SETUP_TRACKING นับทุกแถวที่ NEW_MC>0 ถ้าไม่หักตรงนี้ job/week cap จะรั่ว → แผนเกินโควตา)
                    # ⚠️ ต้อง setdefault เสมอ (แม้ new_mc=0) — โค้ดตรวจ overload ท้ายลูปอ่าน
                    # weekly_job_usage[_produced_week] ตรงๆ ถ้าไม่มี key จะ KeyError
                    weekly_job_usage.setdefault(plan_week, {})
                    if new_mc > 0:
                        weekly_job_usage[plan_week][mc_group] = (
                            weekly_job_usage[plan_week].get(mc_group, 0) + new_mc
                        )
                    # อัพเดต new plan usage (นับทั้ง carryover+new สำหรับ get_actual_mc_remain)
                    # key ต้องเป็น (mc_group, gauge_str) เสมอ — ห้าม pool ข้าม GUAGE
                    # ถ้าเป็น same-week continuation หรือ FG split → เครื่องนับไปแล้วจาก FG ก่อนหน้า ห้ามนับซ้ำ
                    if _same_week_rem_cap is None and not _is_fg_split:
                        if plan_week not in weekly_new_plan_usage:
                            weekly_new_plan_usage[plan_week] = {}
                        _wpu_cap = item_cap_data[
                            (item_cap_data["ITEM_CODE"] == item)
                            & (item_cap_data["MC_GROUP"] == mc_group)
                        ]
                        _wpu_gauge_raw = _wpu_cap.iloc[0]["GUAGE"] if not _wpu_cap.empty else None
                        _wpu_gauge_str = (
                            str(_wpu_gauge_raw).strip()
                            if _wpu_gauge_raw is not None
                            and not (isinstance(_wpu_gauge_raw, float) and pd.isna(_wpu_gauge_raw))
                            else ""
                        )
                        # ถ้า mc_group+gauge อยู่ใน MC_GROUP_REDIRECT → หักเครื่องจาก target แทน
                        # เช่น SKP 20 → หักจาก FA 20 เสมอ
                        _wpu_mc_r, _wpu_gauge_r = _apply_mc_redirect(mc_group, _wpu_gauge_str)
                        _wpu_key = (_mc_to_type1(_wpu_mc_r, _wpu_gauge_r), _wpu_gauge_r)
                        _wpu_added = max(0, available_machines - _booking_week_mc)
                        # MC Special: item ที่มี sub-pool (COTTON/POLY) กินเฉพาะ sub-pool ของตัวเอง
                        # ไม่ไปกินพูลทั่วไปด้วย (ไม่งั้นหักซ้ำ 2 พูล)
                        _wpu_sp_type = _get_subgroup_by_item_prefix(mc_group, _wpu_gauge_str, item)
                        if _wpu_sp_type:
                            if _wpu_added > 0:
                                _wpu_sp_f = _mc_to_factory(str(mc_group).strip().upper(), _wpu_gauge_str)
                                _wpu_sp_cat = _mc_to_type1(str(mc_group).strip().upper(), _wpu_gauge_str)
                                _wpu_sp_key = (_wpu_sp_f, _wpu_sp_cat, _wpu_gauge_r, plan_week, _wpu_sp_type)
                                _mc_special_weekly_usage[_wpu_sp_key] = _mc_special_weekly_usage.get(_wpu_sp_key, 0) + _wpu_added
                        else:
                            weekly_new_plan_usage[plan_week][_wpu_key] = (
                                weekly_new_plan_usage[plan_week].get(_wpu_key, 0) + _wpu_added
                            )
                            # พูลแยก (SKP/SKPTA·SKPLE): นับต่อ pool label ให้หักถูกพูล
                            _wpu_pool = _mc_to_pool_label(_wpu_mc_r, _wpu_gauge_r)
                            if _wpu_pool:
                                weekly_new_plan_pool_usage.setdefault(plan_week, {})[_wpu_pool] = (
                                    weekly_new_plan_pool_usage.setdefault(plan_week, {}).get(_wpu_pool, 0) + _wpu_added
                                )
                        # TYPE_SPECIAL quota tracking
                        if _wpu_added > 0 and _TYPE_DESC_RULES_PLAN:
                            _ts_mc_u3 = str(mc_group).strip().upper()
                            _ts_fac3  = _mc_to_factory(_ts_mc_u3, _wpu_gauge_str)
                            _ts_typ3  = _mc_to_type_raw_plan.get((_ts_mc_u3, _wpu_gauge_str), "").strip().upper()
                            _ts_rk3   = (_ts_fac3.upper(), _ts_typ3)
                            if _ts_rk3 in _TYPE_DESC_RULES_PLAN:
                                _ts_rule3  = _TYPE_DESC_RULES_PLAN[_ts_rk3]
                                _ts_mcat3  = _ts_rule3.get("mc_cat", "")
                                _ts_t13    = _mc_to_type1(_ts_mc_u3, _wpu_gauge_str)
                                if not ((_ts_mcat3 and _ts_t13 != _ts_mcat3) or _wpu_gauge_str == "20"):
                                    _ts_desc3 = _item_desc_map_plan.get(str(item).strip().upper(), "")
                                    if _is_description_special_type_plan(_ts_desc3, _ts_rule3["keywords"]):
                                        _ts_uk3 = (_ts_fac3, _ts_typ3, plan_week)
                                        _type_special_weekly_usage[_ts_uk3] = _type_special_weekly_usage.get(_ts_uk3, 0) + _wpu_added
                    # ก้าวไป week ถัดไปเสมอหลัง produce (ห้าม plan item เดิมใน week เดิมซ้ำ)
                    _produced_week = plan_week
                    plan_week = next_week(plan_week)
                    continue

                # ตรวจว่ามี row ของ FG อื่นอยู่หรือไม่ → ถ้ามี = FG split (สร้าง row ใหม่แทน merge)
                _has_other_fg = any(
                    _p.get("ITEM_CODE") == item
                    and _p.get("PLAN_WEEK") == plan_week
                    and _p.get("MC_GROUP") == mc_group
                    for _p in plans
                )
                if _has_other_fg:
                    # FG_WEEK ต่างกัน → สร้าง row ใหม่ (fall through ไปที่ append ด้านล่าง)
                    _is_fg_split = True
                    print(
                        f"[DEBUG FG SPLIT] {item}+{plan_week}+{mc_group}: existing FG differs → new row for FG {fg_week}"
                    )
                else:
                    print(
                        f"[DEBUG DUPLICATE] FALLBACK SKIP: {item}+{plan_week}+{mc_group} already planned but merge target not found"
                    )
                    # ข้ามไป week ถัดไป
                    plan_week = next_week(plan_week)
                    if plan_week is None:
                        break
                    continue

            # carry ถูกตัดจนไม่เหลือเครื่อง ([CARRY POOL FULL]) → วางสัปดาห์นี้ไม่ได้
            # ห้ามสร้างแถวที่ใช้เครื่อง 0 (ผลิตไม่ได้ แต่กินที่ในแผน)
            if available_machines <= 0 and not _s9_active:
                print(f"[NO MC] {item} SC {sc_so_no} W{plan_week}: ไม่มีเครื่องให้ใช้ → ข้ามไปสัปดาห์ถัดไป")
                plan_week = next_week(plan_week)
                while plan_week is not None and plan_week in SKIP_WEEKS:
                    plan_week = next_week(plan_week)
                if plan_week is None:
                    break
                continue

            # ===== Look-ahead: setup ใหม่ต้องมีเครื่องพอ "ทุกสัปดาห์" ที่งานจะวิ่ง =====
            # ถ้างานต้องวิ่งข้ามไปสัปดาห์ที่ booking จองเกินพูลแล้ว → carry ต่อไม่ได้
            # การ setup ที่สัปดาห์นี้จะเสียวัน setup เปล่าแล้วค้างกลางคัน → ไม่วางตั้งแต่แรก
            # (เลื่อนไปลองสัปดาห์ถัดไป; ถ้าไม่มีสัปดาห์ไหนวิ่งได้เลย งานจะไม่ถูกวาง → รายงานท้ายรัน)
            if new_mc > 0 and not _s9_active and not rts_local_force:
                _span_block = _carry_span_blocked(carryover_mc, new_mc)
                if _span_block is not None:
                    _blk_wk, _blk_rem, _blk_need = _span_block
                    print(
                        f"[NO CARRY ROOM] {item} SC {sc_so_no}: setup W{plan_week} "
                        f"({carryover_mc}+{new_mc} เครื่อง) แต่ W{_blk_wk} วิ่งต่อไม่ได้ "
                        f"(พูลเหลือ {_blk_rem} ต้องการ {_blk_need}) → ไม่วางที่ W{plan_week}"
                    )
                    no_mc_skips.append({
                        "item": item, "sc": sc_so_no, "mc_group": mc_group,
                        "gauge": _sel_gauge, "setup_week": plan_week,
                        "blocked_week": _blk_wk, "remain": _blk_rem, "need": _blk_need,
                        "order_qty": order_qty, "pending_plan": pending_plan, "fg_week": fg_week,
                    })
                    plan_week = next_week(plan_week)
                    while plan_week is not None and plan_week in SKIP_WEEKS:
                        plan_week = next_week(plan_week)
                    if plan_week is None:
                        break
                    continue

            item_gauge = str(order.get("MC_GUAGE", "")).strip()
            print(f"[DEBUG DUPLICATE] Adding plan for {item}+{plan_week}+{mc_group}")

            # เครื่องที่ booking ถักไอเทมนี้อยู่แล้วในสัปดาห์นี้ (carry มาจาก booking ไม่ใช่ setup ใหม่)
            _row_booking_mc = (
                0
                if _s9_active
                else int(
                    booking_mc_by_week.get(_ck(item, mc_group, item_gauge), {}).get(
                        week_index(plan_week), 0
                    )
                    or 0
                )
            )

            plans.append(
                {
                    "ITEM_CODE": item,
                    "SC_SO_NO": str(order.get("SO_NO", order.get("SC/SO NO", ""))).strip(),
                    "MC_GROUP": "COMKN" if _s9_active else mc_group,
                    "MC_GUAGE": "C0" if _s9_active else order["MC_GUAGE"],
                    "FACTORY_TYPE": "OUTSOURCE" if _s9_active else FACTORY_TYPE_MAP.get(mc_group, "UNKNOWN"),
                    "PLAN_WEEK": plan_week,
                    # เก็บ row index บน timeline ไว้ด้วย — PLAN_WEEK เป็น bare week
                    # ที่ระบุปีไม่ได้เมื่อแผนคร่อมปี (ใช้เติม PLAN_YEAR ตอน export)
                    "PLAN_IDX": week_index(plan_week),
                    "PRODUCE_QTY": produce,
                    "SETUP_DAYS": setup_days_used,
                    "REQUIRED_MC": planned_mc,  # เครื่องที่คำนวณไว้ล่วงหน้า (RDD target) หรือ "Maxmc" ถ้าไม่ทัน RDD
                    "ACTUAL_MC": available_machines,  # เครื่องที่ใช้จริง week นี้
                    "CARRYOVER_MC": carryover_mc,  # เครื่องที่ carry-over จาก week ก่อน
                    "NEW_MC": new_mc,  # เครื่องใหม่ที่ setup week นี้
                    "MC_SHARED": 0,  # แถวนี้ประกาศเครื่องเอง (ไม่ได้ใช้เครื่องร่วมกับแถวอื่น)
                    "MC_BOOKING": _row_booking_mc,  # เครื่องที่ booking ถักไอเทมนี้อยู่แล้วในสัปดาห์นี้
                    "REMARK": _build_plan_remark(
                        "COMKN" if _s9_active else mc_group,
                        carryover_mc,
                        new_mc,
                        setup_days_used,
                        booking_mc=_row_booking_mc,
                    ),
                    "FACTORY_WORKING_DAYS": get_working_days_by_factory(
                        mc_group, available_machines, week=plan_week, gauge=item_gauge
                    ),
                    # วันทำงานฐานของกลุ่ม (ค่ามาตรฐาน ไม่อิงสัปดาห์) — ให้หน้าเว็บใช้ fallback
                    # เมื่อลากงานข้ามสัปดาห์ (ค่ารายสัปดาห์จริงหน้าเว็บดึงจาก /api/workday)
                    "WD_BASE": get_working_days_by_factory(
                        mc_group, available_machines, week=None, item_code=item, gauge=item_gauge
                    ),
                    "CALENDAR_WORKING_DAYS": len(get_working_days_in_week(plan_week)),
                    "ACTUAL_WORKING_DAYS": get_working_days_by_factory(mc_group, available_machines, week=plan_week, gauge=item_gauge),
                    "DAILY_CAPACITY": daily_capacity,
                    "REVOLUTION_WEIGHT": rev_weight if rev_weight is not None else 0,
                    "AVAILABLE_DAYS": available_days,
                    "ORDERS_QTY": order_qty,
                    "PENDING_PLAN": pending_plan,
                    "PLAN_QTY": qty_left - produce,
                    "ORDER_TYPE": order_type,
                    "ORDER_DATE": order["Date"],
                    "FG_WEEK": fg_week,
                    "TARGET_KNIT": fg_week_int,
                    "MATERIAL_CONTENT": str(order.get("MATERIAL_CONTENT", "")).strip(),
                    "IS_CORE_ITEM": "CORE ITEM" if is_core_item else "",
                    "CUSTOMER": str(order.get("Customer", "")).strip(),
                    "PLAN_SOURCE": "NEW",
                    "LT_YARN": order.get("DYE_END_DATE") if order_type == "YD-ORDERS" else get_yarn_lt_days(item),
                    "YARN_USED": _yarn_used_lookup.get(str(item).strip().upper(), ""),
                    "DATE_IN": order.get("DATE_IN"),
                    "EARLIEST_PLAN_WEEK": order.get("DYE_END_DATE") if order_type == "YD-ORDERS" else get_yarn_lt_earliest_week(item, date_in=order.get("DATE_IN")),
                    "SUB_COLOR": sub_color,
                    "PO_NO": str(order.get("PO_NO", "")).strip(),
                    "YARNPO": str(order.get("YARNPO", "") or "").strip(),
                    "RDD_WEEK": fg_week,
                    "SC_LINE_ID": str(order.get("SC_LINE_ID", "")).strip(),
                }
            )
            plans[-1]["NAY_COLOR"] = str(order.get("NAY_COLOR", "")).strip()
            plans[-1]["COLOR_DESC"] = str(order.get("COLOR_DESC", "")).strip()
            plans[-1]["LINE_REMARK"] = str(order.get("LINE_REMARK", "")).strip()
            plans[-1]["S9_ROUTING"] = _s9_active
            # OUTSOURCE = ก้อนที่ user สั่งจ้างทอเอง (แยกจาก S9 Only ที่ MasterMC บังคับ)
            plans[-1]["OUTSOURCE"] = "Yes" if _outsource_force else ""
            qty_left -= produce
            # S9 weekly usage tracking — อัปเดตหลัง plans.append ทุกครั้งที่ S9 active
            if _s9_active and mc_group:
                _s9_mc_upper_t = str(mc_group).strip().upper()
                _s9_g_norm_t = _normalize_gauge(_sel_gauge) if _sel_gauge is not None else ""
                _s9_wk = plan_week
                _s9_weekly_usage[(_s9_wk, _s9_mc_upper_t, _s9_g_norm_t)] = (
                    _s9_weekly_usage.get((_s9_wk, _s9_mc_upper_t, _s9_g_norm_t), 0) + available_machines
                )
                print(f"[S9 USAGE] W{_s9_wk} MC={_s9_mc_upper_t} gauge={_s9_g_norm_t}: total_used={_s9_weekly_usage[(_s9_wk, _s9_mc_upper_t, _s9_g_norm_t)]}")
            if qty_left <= 0:
                qty_left = 0  # ป้องกันค่าติดลบ


            # 🔧 FIX: บันทึก Item+Week+MC_GROUP ที่ได้วางแผนไปแล้ว
            _existing_item_week_mc.add((item, plan_week, "COMKN" if _s9_active else mc_group))
            print(f"[DEBUG DUPLICATE] Added to existing set: {item}+{plan_week}+{'COMKN' if _s9_active else mc_group}")
            # บันทึก/อัปเดต remaining cap สำหรับ FG ถัดไปของ item+machine เดียวกัน
            _cap_key = _same_week_key  # ใช้ same_week_key ที่มี COMKN สำหรับ S9 แล้ว
            if _same_week_rem_cap is not None:
                # อัปเดต remaining cap หลังใช้งาน
                _new_rem = max(0, float(_same_week_total_cap or 0.0) - produce)
                if _new_rem > 0:
                    remaining_week_cap[_cap_key] = _new_rem
                    remaining_week_cap_owner[_cap_key] = sc_so_no
                else:
                    remaining_week_cap.pop(_cap_key, None)
                    remaining_week_cap_owner.pop(_cap_key, None)
            else:
                # บันทึกส่วนที่เหลือสำหรับ FG ถัดไป (ไม่ว่า order จะจบหรือไม่)
                # เพื่อให้ SC/SO อื่นสามารถ carry ต่อใน week เดียวกันได้
                _full_week_cap = cap_old + cap_new
                _rem = max(0, _full_week_cap - produce)
                # NOTE: ห้ามเพิ่ม freed capacity จากเครื่องที่ลดลง เพราะเครื่องที่ "ปล่อย"
                # ไม่ได้วิ่งอยู่ใน week นี้แล้ว → capacity ต้องอิงจากเครื่องที่ใช้จริงเท่านั้น
                # เครื่องที่ปล่อยจะว่างให้ item อื่นใช้ผ่าน machine pool แทน
                remaining_week_cap[_cap_key] = _rem
                remaining_week_cap_owner[_cap_key] = sc_so_no
            # S9: ใช้ COMKN เป็น key ใน last_production/machines_in_use เพื่อแยก pool จาก normal
            _plan_ck = _resolve_carry_key(item, "COMKN" if _s9_active else mc_group, _sel_gauge)
            last_production[_plan_ck] = week_index(plan_week)
            machines_in_use[_plan_ck] = available_machines  # บันทึกจำนวนเครื่องที่ใช้จริง
            last_sc_machines[(item, "COMKN" if _s9_active else mc_group, _sel_gauge, sc_so_no)] = available_machines
            last_sc_week[(item, "COMKN" if _s9_active else mc_group, _sel_gauge, sc_so_no)] = week_index(plan_week)
            last_sc_so_no[_plan_ck] = sc_so_no  # บันทึก SC/SO NO ล่าสุดที่ผลิต
            if order_type == "YD-ORDERS" and sub_color:

                last_sub_color[_plan_ck] = sub_color
            # YD-ORDERS: ล็อกจำนวนเครื่องสำหรับ SO ถัดไปใน week เดียวกัน
            if order_type == "YD-ORDERS":
                _yd_lock_key = (item, mc_group, plan_week)
                if _yd_lock_key not in _yd_week_locked_mc:
                    _yd_week_locked_mc[_yd_lock_key] = (available_machines, carryover_mc, new_mc)
            new_plan_started_items.add(_plan_ck)  # บันทึกว่า item นี้เริ่ม new plan แล้ว

            # บันทึก total machines per (week, mc_group) สำหรับ gradual increase
            # 🔧 FG SPLIT: ถ้าเป็น FG split row → เครื่องนับไปแล้วจาก FG แรก ห้ามนับซ้ำ
            if not _is_fg_split:
                _week_mc_key = (week_index(plan_week), mc_group)
                weekly_mc_usage[_week_mc_key] = weekly_mc_usage.get(_week_mc_key, 0) + available_machines
            # อัพเดท job usage สำหรับสัปดาห์นี้ (นับเฉพาะ new_mc = machines ที่ setup ใหม่)
            # เครื่อง carry ต่อ (carryover_mc) ไม่ต้องนับเป็น job ใหม่เพราะไม่ต้อง setup
            # ⚠️ ต้องหักนอกเงื่อนไข _is_fg_split — FG split / YD-ORDERS ก็ตั้งเครื่องใหม่จริง
            # (SETUP_TRACKING นับทุกแถวที่ NEW_MC>0 ถ้าไม่หักตรงนี้ job/week cap จะรั่ว → แผนเกินโควตา)
            # ⚠️ ต้อง setdefault เสมอ (แม้ new_mc=0) — โค้ดตรวจ overload ท้ายลูปอ่าน
            # weekly_job_usage[_produced_week] ตรงๆ ถ้าไม่มี key จะ KeyError
            weekly_job_usage.setdefault(plan_week, {})
            if new_mc > 0:
                weekly_job_usage[plan_week][mc_group] = (
                    weekly_job_usage[plan_week].get(mc_group, 0) + new_mc
                )
            # อัพเดท new plan usage (นับทั้ง carryover+new สำหรับ get_actual_mc_remain)
            # key ต้องเป็น (mc_group, gauge_str) เสมอ — ห้าม pool ข้าม GUAGE
            # ถ้าเป็น same-week continuation หรือ FG split → เครื่องนับไปแล้วจาก FG ก่อนหน้า ห้ามนับซ้ำ
            if _same_week_rem_cap is None and not _is_fg_split:
                if plan_week not in weekly_new_plan_usage:
                    weekly_new_plan_usage[plan_week] = {}
                _wpu_cap = item_cap_data[
                    (item_cap_data["ITEM_CODE"] == item)
                    & (item_cap_data["MC_GROUP"] == mc_group)
                ]
                _wpu_gauge_raw = _wpu_cap.iloc[0]["GUAGE"] if not _wpu_cap.empty else None
                _wpu_gauge_str = (
                    str(_wpu_gauge_raw).strip()
                    if _wpu_gauge_raw is not None
                    and not (isinstance(_wpu_gauge_raw, float) and pd.isna(_wpu_gauge_raw))
                    else ""
                )
                # ถ้า mc_group+gauge อยู่ใน MC_GROUP_REDIRECT → หักเครื่องจาก target แทน
                # เช่น SKP 20 → หักจาก FA 20 เสมอ
                _wpu_mc_r, _wpu_gauge_r = _apply_mc_redirect(mc_group, _wpu_gauge_str)
                _wpu_key = (_mc_to_type1(_wpu_mc_r, _wpu_gauge_r), _wpu_gauge_r)
                _wpu_added = max(0, available_machines - _booking_week_mc)
                # MC Special: item ที่มี sub-pool (COTTON/POLY) กินเฉพาะ sub-pool ของตัวเอง
                # ไม่ไปกินพูลทั่วไปด้วย (ไม่งั้นหักซ้ำ 2 พูล)
                _wpu_sp_type = _get_subgroup_by_item_prefix(mc_group, _wpu_gauge_str, item)
                if _wpu_sp_type:
                    if _wpu_added > 0:
                        _wpu_sp_f = _mc_to_factory(str(mc_group).strip().upper(), _wpu_gauge_str)
                        _wpu_sp_cat = _mc_to_type1(str(mc_group).strip().upper(), _wpu_gauge_str)
                        _wpu_sp_key = (_wpu_sp_f, _wpu_sp_cat, _wpu_gauge_r, plan_week, _wpu_sp_type)
                        _mc_special_weekly_usage[_wpu_sp_key] = _mc_special_weekly_usage.get(_wpu_sp_key, 0) + _wpu_added
                else:
                    weekly_new_plan_usage[plan_week][_wpu_key] = (
                        weekly_new_plan_usage[plan_week].get(_wpu_key, 0) + _wpu_added
                    )
                    # พูลแยก (SKP/SKPTA·SKPLE): นับต่อ pool label ให้หักถูกพูล
                    _wpu_pool = _mc_to_pool_label(_wpu_mc_r, _wpu_gauge_r)
                    if _wpu_pool:
                        weekly_new_plan_pool_usage.setdefault(plan_week, {})[_wpu_pool] = (
                            weekly_new_plan_pool_usage.setdefault(plan_week, {}).get(_wpu_pool, 0) + _wpu_added
                        )
                # TYPE_SPECIAL quota tracking
                if _wpu_added > 0 and _TYPE_DESC_RULES_PLAN:
                    _ts_mc_u2 = str(mc_group).strip().upper()
                    _ts_fac2  = _mc_to_factory(_ts_mc_u2, _wpu_gauge_str)
                    _ts_typ2  = _mc_to_type_raw_plan.get((_ts_mc_u2, _wpu_gauge_str), "").strip().upper()
                    _ts_rk2   = (_ts_fac2.upper(), _ts_typ2)
                    if _ts_rk2 in _TYPE_DESC_RULES_PLAN:
                        _ts_rule2  = _TYPE_DESC_RULES_PLAN[_ts_rk2]
                        _ts_mcat2  = _ts_rule2.get("mc_cat", "")
                        _ts_t12    = _mc_to_type1(_ts_mc_u2, _wpu_gauge_str)
                        if not ((_ts_mcat2 and _ts_t12 != _ts_mcat2) or _wpu_gauge_str == "20"):
                            _ts_desc2 = _item_desc_map_plan.get(str(item).strip().upper(), "")
                            if _is_description_special_type_plan(_ts_desc2, _ts_rule2["keywords"]):
                                _ts_uk2 = (_ts_fac2, _ts_typ2, plan_week)
                                _type_special_weekly_usage[_ts_uk2] = _type_special_weekly_usage.get(_ts_uk2, 0) + _wpu_added
            # ก้าวไป week ถัดไปเสมอหลัง produce (ห้าม plan item เดิมใน week เดิมซ้ำ)
            _produced_week = plan_week
            plan_week = next_week(plan_week)
            # ตรวจสอบว่าแผนใหม่ไม่เกิน capacity ที่เหลือหลังจากแผนเก่า
            if not old_plan_df.empty and "PLAN_WEEK" in old_plan_df.columns:
                # หาจำนวน jobs ของแผนเก่าในสัปดาห์นี้
                old_week_jobs = old_plan_df[old_plan_df["PLAN_WEEK"] == _produced_week]
                if not old_week_jobs.empty:
                    mc_col = (
                        "REQUIRED_MC"
                        if "REQUIRED_MC" in old_week_jobs.columns
                        else "AVAILABLE_MACHINES"
                    )
                    # Ensure the column is numeric before summing to avoid string concatenation
                    old_jobs_by_mc = (
                        old_week_jobs.assign(
                            _mc_num=pd.to_numeric(
                                old_week_jobs[mc_col], errors="coerce"
                            ).fillna(0)
                        )
                        .groupby("MC_GROUP")["_mc_num"]
                        .sum()
                        .astype(int)
                        .to_dict()
                    )
                    # ตรวจสอบแต่ละ MC_GROUP ว่าเกินแผนเก่าหรือไม่
                    for used_mc_group, new_jobs in weekly_job_usage.get(_produced_week, {}).items():
                        old_jobs = old_jobs_by_mc.get(used_mc_group, 0)
                        if new_jobs > old_jobs:
                            print(
                                f"⚠️  OVER OLD PLAN: Week {_produced_week} {used_mc_group} ใช้ {new_jobs} jobs (เกินแผนเก่า {old_jobs} jobs)"
                            )
                    # ตรวจสอบ capacity ที่เหลือหลังจากแผนเก่า
                    remaining_capacity_by_type = {}
                    for used_mc_group in weekly_job_usage.get(_produced_week, {}):
                        mc_info = master_mc[master_mc["MC"] == used_mc_group]
                        if not mc_info.empty:
                            factory = str(mc_info.iloc[0]["Factory"]).strip().upper()
                            _raw_t = mc_info.iloc[0].get("Type", "")
                            mc_type = "" if pd.isna(_raw_t) else str(_raw_t).strip().upper()
                            type_key = _make_type_key(factory, mc_type)
                            # หาจำนวน jobs ของแผนเก่าใน type เดียวกัน
                            old_type_jobs = 0
                            for old_mc_group, old_jobs in old_jobs_by_mc.items():
                                old_mc_info = master_mc[master_mc["MC"] == old_mc_group]
                                if not old_mc_info.empty:
                                    old_factory = (
                                        str(old_mc_info.iloc[0]["Factory"]).strip().upper()
                                    )
                                    _raw_ot = old_mc_info.iloc[0].get("Type", "")
                                    old_mc_type = (
                                        ""
                                        if pd.isna(_raw_ot)
                                        else str(_raw_ot).strip().upper()
                                    )
                                    old_type_key = _make_type_key(old_factory, old_mc_type)
                                    if old_type_key == type_key:
                                        old_type_jobs += old_jobs
                            # คำนวณ capacity ที่เหลือ
                            if type_key not in remaining_capacity_by_type:
                                if factory == "PHET":
                                    if mc_type == "DOUBLE":
                                        max_capacity = 33
                                    elif mc_type == "SINGLE":
                                        max_capacity = 44
                                    else:
                                        max_capacity = 9999
                                elif factory in ("OM", "OMNOI"):
                                    max_capacity = 13
                                else:
                                    max_capacity = 9999
                                remaining_capacity_by_type[type_key] = (
                                    max_capacity - old_type_jobs
                                )
                    # ตรวจสอบว่าเกิน capacity ที่เหลือหรือไม่
                    for type_key, remaining_capacity in remaining_capacity_by_type.items():
                        # หาจำนวน jobs ใหม่ใน type เดียวกัน
                        new_type_jobs = 0
                        for used_mc_group, new_jobs in weekly_job_usage.get(
                            _produced_week, {}
                        ).items():
                            mc_info = master_mc[master_mc["MC"] == used_mc_group]
                            if not mc_info.empty:
                                factory = str(mc_info.iloc[0]["Factory"]).strip().upper()
                                _raw_t2 = mc_info.iloc[0].get("Type", "")
                                mc_type = (
                                    "" if pd.isna(_raw_t2) else str(_raw_t2).strip().upper()
                                )
                                current_type_key = _make_type_key(factory, mc_type)
                                if current_type_key == type_key:
                                    new_type_jobs += new_jobs
                        if new_type_jobs > remaining_capacity:
                            print(
                                f"⚠️  OVER REMAINING CAPACITY: Week {_produced_week} {type_key} ใช้ {new_type_jobs} jobs (เกิน capacity ที่เหลือ {remaining_capacity} jobs, แผนเก่าใช้ไปแล้ว)"
                            )
            # ตรวจสอบและแสดงผลการทับซ้อนในสัปดาห์นี้ (ตาม factory type)
            total_jobs_by_type = {}
            max_capacity_by_type = {}
            # คำนวณ jobs และ capacity ตาม factory type
            for used_mc_group in weekly_job_usage.get(_produced_week, {}):
                mc_info = master_mc[master_mc["MC"] == used_mc_group]
                if not mc_info.empty:
                    factory = str(mc_info.iloc[0]["Factory"]).strip().upper()
                    _raw_t = mc_info.iloc[0].get("Type", "")
                    mc_type = "" if pd.isna(_raw_t) else str(_raw_t).strip().upper()
                    type_key = _make_type_key(factory, mc_type)
                    # บวก jobs ที่ใช้
                    if type_key not in total_jobs_by_type:
                        total_jobs_by_type[type_key] = 0
                    total_jobs_by_type[type_key] += weekly_job_usage[_produced_week][
                        used_mc_group
                    ]
                    # กำหนด capacity ตาม type
                    if type_key not in max_capacity_by_type:
                        if factory == "PHET":
                            if mc_type == "DOUBLE":
                                max_capacity_by_type[type_key] = 33
                            elif mc_type == "SINGLE":
                                max_capacity_by_type[type_key] = 44
                        elif factory in ("OM", "OMNOI"):
                            max_capacity_by_type[type_key] = 13
                        else:
                            max_capacity_by_type[type_key] = (
                                9999  # ไม่รู้จัก type นี้ ให้ใช้ค่า max เพื่อไม่ overload
                            )
            # ตรวจสอบและแสดงผลเฉพาะตอนที่เกิน
            for type_key, jobs_used in total_jobs_by_type.items():
                capacity = max_capacity_by_type.get(type_key, 9999)
                if jobs_used > capacity:
                    print(
                        f"⚠️  OVERLOAD: Week {_produced_week} {type_key} ใช้ {jobs_used} jobs (เกิน capacity {capacity} jobs)"
                    )
        # ออร์เดอร์ที่เข้า loop วางแผนแล้วแต่วางไม่ครบ (เครื่องไม่พอ/หมดสัปดาห์ในปี/ชน safety guard)
        # หมายเหตุ: กรณี CORE ITEM (batch production) qty_left ที่จับได้คือของ batch ปัจจุบันเท่านั้น
        # ถ้ายังมี batch ถัดไปที่ยังไม่ถึงคิวค้างอยู่ ตัวเลขนี้จะนับยอดคงเหลือต่ำกว่าความจริง
        if qty_left > 0.5:
            _skip_qty_short.append(_unplanned_row(
                order,
                "วางแผนไม่ครบ (เครื่องไม่พอ/หมดสัปดาห์ที่วางได้/ชน safety guard)",
                qty=qty_left,
            ))
        # 🔧 FIX: อัปเดต cumulative planned qty สำหรับ item นี้
        # ใช้ order_qty (ไม่ใช่ qty_left) เพราะ order ถูก process แล้วไม่ว่าผลิตครบหรือไม่
        _item_cumulative_planned[item] = _item_cumulative_planned.get(item, 0) + order_qty
        # บันทึก last plan week สำหรับ SC/SO+Item → บังคับ FG_WEEK ถัดไปเริ่มหลังนี้
        if _produced_week is not None:
            _lpw_idx = week_index(_produced_week)
            if _lpw_idx is not None:
                _prev = _last_fg_plan_idx.get((sc_so_no, item))
                if _prev is None or _lpw_idx > _prev:
                    _last_fg_plan_idx[(sc_so_no, item)] = _lpw_idx

    # =========================
    # CAPACITY OPTIMIZATION
    # =========================
    # Optimize machine utilization by filling unused capacity with same-item orders from different SCs
    plans = detect_and_fill_unused_capacity(plans, orders, summary_mc)
    return plans

# =========================
# รัน planning รอบเดียว (S9 เหลือเฉพาะ routing ของ item ใน "S9 Only")
# =========================
plans = _run_planning_loop()

# =========================
# สรุปงานที่วางไม่ได้เพราะพูลเครื่องเต็ม — ไม่งั้นงานหายจากแผนเงียบๆ
# ❌ = ไม่ได้วางเลยทั้ง SC  |  ↪️ = เลื่อนไปวางสัปดาห์อื่นได้
# =========================
if no_mc_skips:
    _planned_keys = {
        (str(_p.get("ITEM_CODE", "")), str(_p.get("SC_SO_NO", ""))) for _p in plans
    }
    _skips_by_key = {}
    for _s in no_mc_skips:
        _skips_by_key.setdefault((str(_s["item"]), str(_s["sc"])), []).append(_s)
    print("\n" + "=" * 78)
    print("⚠️  งานที่เครื่องไม่พอ (booking จองเกินเครื่องจริง → carry ต่อไม่ได้)")
    print("=" * 78)
    for (_it, _sc), _rows in sorted(_skips_by_key.items()):
        _never_planned = (_it, _sc) not in _planned_keys
        _mark = "❌ ไม่ได้วางเลย" if _never_planned else "↪️  เลื่อนสัปดาห์"
        print(f"  {_mark} | {_it} SC {_sc} | {_rows[0]['mc_group']} เกจ {_rows[0]['gauge']}")
        for _r in _rows[:4]:
            print(
                f"      setup W{_r['setup_week']} → ติดที่ W{_r['blocked_week']} "
                f"(พูลเหลือ {_r['remain']} ต้องใช้ {_r['need']})"
            )
        if len(_rows) > 4:
            print(f"      ... และอีก {len(_rows) - 4} สัปดาห์ที่ลองแล้วไม่ได้")
    print("=" * 78 + "\n")

# EXPORT
# =========================
plan_df = pd.DataFrame(plans)
DATA_PLAN_DIR.mkdir(exist_ok=True)

# สร้าง week -> year lookup จาก calendar_week (รองรับข้ามปี)
_week_year_lookup = {}
for _, row in calendar_week.iterrows():
    w = int(row["WEEK"])
    y = int(row["YEAR"])
    if w not in _week_year_lookup:
        _week_year_lookup[w] = y
    else:
        # ถ้า week ซ้ำข้ามปี → ใช้ปีที่ใกล้ TODAY.year ที่สุด
        if abs(y - TODAY.year) < abs(_week_year_lookup[w] - TODAY.year):
            _week_year_lookup[w] = y



# เพิ่ม PLAN_YEAR ให้ new plan_df
# ใช้ PLAN_IDX (row index บน timeline ที่เก็บไว้ตอนวางแผน) เป็นหลัก เพราะรู้ปีแน่นอน
# _week_year_lookup เป็น fallback เท่านั้น — มันเดาปีจาก "ใกล้ TODAY.year ที่สุด"
# ซึ่งผิดกับแผนที่คร่อมปี (เช่น แผนเดิน W46→W53 แล้วต่อ W1 = ปีหน้า
# แต่ lookup จะตอบปีปัจจุบัน → Gantt/รายงานย้อนอดีต 1 ปี)
if not plan_df.empty and "PLAN_WEEK" in plan_df.columns:
    _fallback_year = plan_df["PLAN_WEEK"].map(_week_year_lookup).fillna(TODAY.year)
    if "PLAN_IDX" in plan_df.columns:
        _idx_year = plan_df["PLAN_IDX"].map(
            lambda i: int(calendar_week.iloc[int(i)]["YEAR"])
            if pd.notna(i) and 0 <= int(i) < len(calendar_week)
            else None
        )
        plan_df["PLAN_YEAR"] = _idx_year.fillna(_fallback_year).astype(int)
    else:
        plan_df["PLAN_YEAR"] = _fallback_year.astype(int)


# เพิ่ม CAT column (= Type_1 ของ MC_GROUP จาก MasterMC)
if not plan_df.empty and "MC_GROUP" in plan_df.columns:
    plan_df["CAT"] = plan_df.apply(
        lambda r: _mc_to_type1(str(r.get("MC_GROUP", "")), r.get("MC_GUAGE")), axis=1
    )
    # S9 routing: override CAT เป็น COMKN เสมอ
    if "S9_ROUTING" in plan_df.columns:
        plan_df.loc[plan_df["S9_ROUTING"] == True, "CAT"] = "COMKN"

# เพิ่ม CYLINDER_CHANGE column — mark เฉพาะ item ที่ trigger การเปลี่ยนจริง
plan_df["CYLINDER_CHANGE"] = ""
_cyl_trigger_keys = set()
for (_ciw, _cii, _cimg) in _cylinder_change_for_item:
    _cyl_trigger_keys.add((int(_ciw) + 1, _cii, str(_cimg).strip().upper()))
for _ci, _cr in plan_df.iterrows():
    _ck_week = int(_cr.get("PLAN_WEEK") or 0)
    _ck_item = str(_cr.get("ITEM_CODE", "")).strip().upper()
    _ck_mc = str(_cr.get("MC_GROUP", "")).strip().upper()
    if (_ck_week, _ck_item, _ck_mc) in _cyl_trigger_keys:
        plan_df.at[_ci, "CYLINDER_CHANGE"] = "Yes"

# =========================
# LOAD BALANCING - Apply to final plan
# =========================
# ปิดการใช้ Load Balancing ตอนสรุปผลลัพธ์เพื่อรักษาความสามารถในการผลิตตาม ORDERS_QTY
# Load Balancing ทำงานเฉพาะตอนเลือกเครื่องเพื่อกระจายงานอย่างสมดุล
if False and USE_LOAD_BALANCING and not plan_df.empty:
    # Group plans by week for load balancing
    balanced_plans = []
    for week in sorted(plan_df['PLAN_WEEK'].unique()):
        week_plans = plan_df[plan_df['PLAN_WEEK'] == week].to_dict('records')

    
        # Apply load balancing to this week's plans
        balanced_week_plans = _analyze_and_balance_load(week, week_plans)
        balanced_plans.extend(balanced_week_plans)
    # Replace original plans with balanced plans
    plan_df = pd.DataFrame(balanced_plans)
    print(f"✅ Load balancing completed. Total plans after balancing: {len(balanced_plans)}")

# สรุปแผนใหม่
if not plan_df.empty:
    print("📊 สรุปแผนการผลิตใหม่:")
    # Coerce `REQUIRED_MC` to numeric to avoid string concatenation when summing
    new_summary = (
        plan_df.assign(
            _req_mc=pd.to_numeric(plan_df["REQUIRED_MC"], errors="coerce")
            .fillna(0)
            .astype(int)
        )
        .groupby(["MC_GROUP", "PLAN_WEEK"])["_req_mc"]
        .sum()
        .reset_index()
        .rename(columns={"_req_mc": "REQUIRED_MC"})
    )
    for week in sorted(new_summary["PLAN_WEEK"].unique()):
        week_data = new_summary[new_summary["PLAN_WEEK"] == week]
        week_sum = week_data["REQUIRED_MC"].sum()
        if week == 19:  # Debug Week 19
            print(f"🔍 Week {week} details:")
            for _, row in week_data.iterrows():
                print(f"   {row['MC_GROUP']}: {row['REQUIRED_MC']} machines")
            print(f"   Total: {week_sum}")
        print(f"  Week {week}: REQUIRED_MC = {week_sum}")
else:
    print("⚠️ ไม่มีแผนการผลิตที่สามารถสร้างได้")
    print()

# =========================
# REMAINING JOBS PER WEEK
# =========================
_CAPACITY_MAP = {
    "PHET_DOUBLE": 33,
    "PHET_SINGLE": 44,
    "OM": 13,
}

def _sum_by_type(job_dict_by_week, week):
    """รวม jobs ตาม factory_type สำหรับ week ที่ระบุ (factory-wide รวมทุก MC_GROUP ใน type)"""
    result: dict = {}
    for _mc_group, _jobs in job_dict_by_week.get(week, {}).items():
        _type_key = _get_type_key_for_mc(_mc_group)
        result[_type_key] = result.get(_type_key, 0) + _jobs
    return result

# 🔧 FIX: NEW_JOBS = setup จริงจากแผนสุดท้าย (plan_df.NEW_MC) เพื่อให้ตรงกับ SETUP_TRACKING
# เดิมคิดจาก weekly_job_usage (ตัวนับระหว่างวาง) ซึ่งไม่สะท้อน balancing ท้ายสุด → นับ job เกินจริง
def _build_remaining_df(plan_frame):
    """สร้าง REMAINING_JOBS จากแผนที่ระบุ — NEW จาก NEW_MC ของ frame, OLD จาก booking baseline
    (weekly_job_usage_old เป็น booking baseline ตัวเดียว ใช้ร่วมได้ทั้ง with-S9 และ no-S9)"""
    _plan_new_by_wt: dict = {}
    if not plan_frame.empty and "NEW_MC" in plan_frame.columns:
        _pnm_series = pd.to_numeric(plan_frame["NEW_MC"], errors="coerce").fillna(0).astype(int)
        for _i, _pr in plan_frame.iterrows():
            _nm = int(_pnm_series.loc[_i])
            if _nm <= 0:
                continue
            _w = int(pd.to_numeric(_pr.get("PLAN_WEEK"), errors="coerce") or 0)
            _tk = _get_type_key_for_mc(str(_pr.get("MC_GROUP", "")).strip().upper())
            _plan_new_by_wt[(_w, _tk)] = _plan_new_by_wt.get((_w, _tk), 0) + _nm

    # OLD จาก booking baseline (weekly_job_usage_old) — ตรงกับ SETUP_TRACKING OLD
    _all_weeks = sorted(
        set(list(weekly_job_usage_old.keys()) + [_w for (_w, _t) in _plan_new_by_wt])
    )
    _remaining_rows = []
    for _week in _all_weeks:
        _old_by_type = _sum_by_type(weekly_job_usage_old, _week)
        _new_by_type = {}
        for (_w, _tk), _v in _plan_new_by_wt.items():
            if _w == _week:
                _new_by_type[_tk] = _new_by_type.get(_tk, 0) + _v
        _all_types = set(list(_old_by_type.keys()) + list(_new_by_type.keys()))
        for _type_key in sorted(_all_types):
            _old_used = _old_by_type.get(_type_key, 0)
            _new_used = _new_by_type.get(_type_key, 0)
            _total_used = _old_used + _new_used
            _cap = _CAPACITY_MAP.get(_type_key, None)
            _rem = (_cap - _total_used) if _cap is not None else None
            _remaining_rows.append(
                {
                    "WEEK": _week,
                    "TYPE": _type_key,
                    "OLD_JOBS": _old_used,
                    "NEW_JOBS": _new_used,
                    "TOTAL_JOBS": _total_used,
                    "CAPACITY": _cap,
                    "REMAINING_JOBS": _rem,
                }
            )
    return (
        pd.DataFrame(_remaining_rows)
        if _remaining_rows
        else pd.DataFrame(
            columns=[
                "WEEK",
                "TYPE",
                "OLD_JOBS",
                "NEW_JOBS",
                "TOTAL_JOBS",
                "CAPACITY",
                "REMAINING_JOBS",
            ]
        )
    )

remaining_df = _build_remaining_df(plan_df)
print("📋 สรุป Remaining Jobs ต่อ Week (factory-wide ต่อ Type, OLD + NEW):")
if not remaining_df.empty:
    for _week in sorted(remaining_df["WEEK"].unique()):
        _wdf = remaining_df[remaining_df["WEEK"] == _week]
        print(f"   Week {_week}:")
        for _, _row in _wdf.iterrows():
            _cap_s = str(int(_row["CAPACITY"])) if pd.notna(_row["CAPACITY"]) else "-"
            _rem_v = _row["REMAINING_JOBS"] if pd.notna(_row["REMAINING_JOBS"]) else 1
            _rem_s = str(int(_rem_v)) if pd.notna(_row["REMAINING_JOBS"]) else "-"
            _icon = "🔴" if _rem_v < 0 else ("🟡" if _rem_v <= 5 else "🟢")
            print(
                f"     {_icon} {_row['TYPE']}: old={int(_row['OLD_JOBS'])} + new={int(_row['NEW_JOBS'])} = {int(_row['TOTAL_JOBS'])}/{_cap_s}  เหลือ {_rem_s} jobs"
            )
    print()
else:
    print("   (ไม่มีข้อมูล)")
    print()
# แสดง item ที่ไม่มี CAP
if _skip_no_cap:
    print(f"\n⚠️  Items ที่ไม่พบ CAP data ({len(_skip_no_cap)} รายการ) → ไม่ได้วางแผน:")
    for _s in _skip_no_cap:
        print(f"   - {_s['ITEM_CODE']} (SC/SO:{_s['SC_SO_NO']})")
    print(f"   กรุณาเพิ่มใน item_cap2025.xlsx")
    print()

# แสดง orders ที่ไม่มี MC GROUP
if _skip_no_mc_group:
    print(f"\n⚠️  Orders ที่ไม่มี MC GROUP ({len(_skip_no_mc_group)} รายการ) → ไม่ได้วางแผน:")
    for _s in _skip_no_mc_group:
        print(f"   - {_s['SC_SO_NO']} | {_s['ITEM_CODE']}")
    print()

# แสดง orders ที่ MC GROUP ไม่มี FACTORY_TYPE
if _skip_no_factory:
    print(f"\n⚠️  Orders ที่ MC GROUP ไม่มี FACTORY_TYPE ({len(_skip_no_factory)} รายการ) → ไม่ได้วางแผน:")
    for _s in _skip_no_factory:
        print(f"   - {_s['SC_SO_NO']} | {_s['ITEM_CODE']} | MC={_s['MC_GROUP']}")

# แสดง orders ที่ MC+Gauge ไม่มีใน MasterMC
if _skip_not_in_master:
    print(f"\n⚠️  Orders ที่ MC+Gauge ไม่มีใน MasterMC ({len(_skip_not_in_master)} รายการ) → ไม่ได้วางแผน:")
    for _s in _skip_not_in_master:
        print(f"   - {_s['SC_SO_NO']} | {_s['ITEM_CODE']} | MC={_s['MC_GROUP']} G{_s['GAUGE']}")
    print()

# แสดง orders ที่เข้า loop วางแผนแล้วแต่วางไม่ครบ (เครื่องไม่พอ/หมดสัปดาห์/ชน safety guard)
if _skip_qty_short:
    print(f"\n⚠️  Orders ที่วางแผนไม่ครบ ({len(_skip_qty_short)} รายการ) → เหลือ qty ที่ยังไม่ได้วาง:")
    for _s in _skip_qty_short:
        print(f"   - {_s['SC_SO_NO']} | {_s['ITEM_CODE']} | เหลือ {_s['UNPLANNED_QTY']:.0f} | {_s['REASON']}")
    print()

# แถวที่ถูกกรองออกตั้งแต่ Order.py (ไม่เคยเข้า planning loop) — เอาเฉพาะที่ยังมีของค้าง
# กติกาเดียวกับ order ปกติ: Pending Plan <= 0 = วางแผนครบแล้ว ไม่ถือว่าค้าง
_excluded_rows = []
if not _excluded_orders.empty and "Pending Plan" in _excluded_orders.columns:
    _exc_pend = pd.to_numeric(_excluded_orders["Pending Plan"], errors="coerce").fillna(0)
    for _, _er in _excluded_orders[_exc_pend > 0].iterrows():
        _excluded_rows.append(
            _unplanned_row(_er, _er.get("REASON", "ถูกกรองออกก่อนวางแผน"))
        )
    if _excluded_rows:
        print(f"\n⚠️  Orders ที่ถูกกรองออกก่อนวางแผน ({len(_excluded_rows)} รายการ) → เข้าชีท UNPLANNED:")
        for _s in _excluded_rows:
            print(f"   - {_s['SC_SO_NO']} | {_s['ITEM_CODE']} | เหลือ {_s['UNPLANNED_QTY']:.0f} | {_s['REASON']}")
        print()

# รวม unplanned orders เป็น DataFrame — คอลัมน์ตาม UNPLANNED_COLS (ชุดเดิม + DATE_IN)
_unplanned_rows = (
    _skip_no_mc_group + _skip_no_factory + _skip_not_in_master
    + _skip_no_cap + _skip_qty_short + _excluded_rows
)
_unplanned_df = pd.DataFrame(_unplanned_rows, columns=UNPLANNED_COLS)
print("Weekly production planning completed")
print(f"Output: {OUTPUT_FILE}")
print(f"Total rows: {len(plan_df)}")
# =========================
# EXPORT COMBINED (OLD + NEW)
# =========================
COMBINED_FILE = DATA_PLAN_DIR / "weekly_production_plan_combined_filtered.xlsx"
new_df = plan_df.copy()
if "PLAN_SOURCE" not in new_df.columns:
    new_df.insert(0, "PLAN_SOURCE", "NEW")
# สร้าง OLD rows จาก detail_mc โดยเทียบ ITEM_CODE + MC_GROUP กับ items ที่อยู่ใน new plan
# (ไม่ต้อง match SO_NO เพราะต้องการดู old plan ของ item นั้นทั้งหมดเพื่อเทียบ CARRYOVER_MC)
new_item_mc_keys = set(
    zip(
        plan_df["ITEM_CODE"].astype(str).str.strip().str.upper(),
        plan_df["MC_GROUP"].astype(str).str.strip().str.upper(),
    )
)
old_booking_df = pd.DataFrame()
if not detail_mc.empty:
    # โหลดเฉพาะ item ที่มีออเดอร์ใหม่ (วางแผนอยู่) เท่านั้น
    if "ITEM_CODE" in detail_mc.columns:
        old_booking_df = detail_mc[detail_mc["ITEM_CODE"].astype(str).str.strip().str.upper().isin(_plan_items)].copy()
    else:
        old_booking_df = detail_mc.copy()
    # Rename detail_mc columns → ชื่อเดียวกับ new plan
    # แต่ยังไม่ rename FIBER_TYPE เพื่อใช้ตรวจสอบ setup days
    old_booking_df = old_booking_df.rename(
        columns={
            "GUAGE": "MC_GUAGE",
            "WEEK": "PLAN_WEEK",
            "KP_WEIGHT": "PRODUCE_QTY",
            "MC_USE_CEIL": "REQUIRED_MC",
            "MC_USE": "ACTUAL_MC",
            "CAP ทอ": "DAILY_CAPACITY",
            "REVOLUTION/WEIGHT": "REVOLUTION_WEIGHT",
            "SO_NO": "SC_SO_NO",
        }
    )
    # เติม MATERIAL_CONTENT จาก FIBER_TYPE ของ booking ก่อน (เพื่อให้ setup days ถูกต้อง)
    # ถ้าไม่มี FIBER_TYPE ค่อย lookup จาก orders
    _mat_item_lookup = {}
    if "MATERIAL_CONTENT" in orders.columns:
        for _, _or in orders.iterrows():
            _ic = str(_or.get("Item Code", "")).strip().upper()
            _mat = str(_or.get("MATERIAL_CONTENT", "")).strip()
            if _ic and _mat and _mat.upper() not in ("NAN", ""):
                _mat_item_lookup[_ic] = _mat
    if "ITEM_CODE" in old_booking_df.columns:
        def _get_mat_from_booking(row):
            _ic = str(row.get("ITEM_CODE", "")).strip().upper()
            # 1. ใช้ MATERIAL_CONTENT จาก booking โดยตรง (detail_mc มีคอลัมน์นี้)
            _mat = str(row.get("MATERIAL_CONTENT", "")).strip()
            if _mat and _mat.upper() not in ("NAN", ""):
                return _mat
            # 2. ใช้ FIBER_TYPE จาก booking (legacy)
            _fiber = str(row.get("FIBER_TYPE", "")).strip()
            if _fiber and _fiber.upper() not in ("NAN", ""):
                return _fiber
            # 3. fallback: lookup จาก orders
            _v = _mat_item_lookup.get(_ic)
            if _v:
                return _v
            return ""
        old_booking_df["MATERIAL_CONTENT"] = old_booking_df.apply(_get_mat_from_booking, axis=1)
    # แปลง DAILY_CAPACITY จากฐาน 24 ชั่วโมง → ตาม Item Special working hour (หรือ 20 ชั่วโมง default)
    if "DAILY_CAPACITY" in old_booking_df.columns:
        def _adj_old_cap(row):
            cap = pd.to_numeric(row["DAILY_CAPACITY"], errors="coerce")
            if pd.isna(cap):
                return cap
            _is = get_item_special(
                str(row.get("ITEM_CODE", "")).strip().upper(),
                str(row.get("MC_GROUP", "")).strip().upper(),
                str(row.get("MC_GUAGE", "")).strip(),
            )
            if _is is not None:
                return cap * (_is[1] / 24)  # Item Special working_hour override (ชนะเสมอ)
            # ไม่มีใน Item Special → ชั่วโมงของกลุ่มจากแผง Work Day (ไม่ใช้ MasterMC)
            return cap * (_base_work_hours(row.get("MC_GROUP"), row.get("MC_GUAGE")) / 24)
        old_booking_df["DAILY_CAPACITY"] = old_booking_df.apply(_adj_old_cap, axis=1)
    # ตัด S นำหน้า SC_SO_NO (เช่น "S717455" → "717455")
    if "SC_SO_NO" in old_booking_df.columns:
        old_booking_df["SC_SO_NO"] = (
            old_booking_df["SC_SO_NO"].astype(str).str.lstrip("Ss")
        )
    old_booking_df.insert(0, "PLAN_SOURCE", "OLD")
    # แปลง PLAN_WEEK เป็นตัวเลขถ้ามี
    if "PLAN_WEEK" in old_booking_df.columns:
        old_booking_df["PLAN_WEEK"] = pd.to_numeric(
            old_booking_df["PLAN_WEEK"], errors="coerce"
        )
    # 🔧 FIX: Recalculate ACTUAL_MC / REQUIRED_MC สำหรับ OLD rows ตาม calendar จริง
    # (booking เดิมคำนวณด้วย factory days 6 วัน แต่บาง week เช่น W19 มี 5 วันทำงานจริง)
    if all(
        c in old_booking_df.columns
        for c in [
            "PRODUCE_QTY",
            "DAILY_CAPACITY",
            "PLAN_WEEK",
            "MC_GROUP",
            "ACTUAL_MC",
            "REQUIRED_MC",
        ]
    ):
        # 🔧 FIX: หา week แรกของแต่ละ (item, mc_group, gauge) เพื่อคำนวณ setup เฉพาะ week แรก
        _first_week_map = {}
        if "ITEM_CODE" in old_booking_df.columns and "MC_GUAGE" in old_booking_df.columns:
            for _, grp in old_booking_df.groupby(["ITEM_CODE", "MC_GROUP", "MC_GUAGE"]):
                weeks = sorted(grp["PLAN_WEEK"].dropna().unique())
                if weeks:
                    _first_week_map[(grp.iloc[0]["ITEM_CODE"], grp.iloc[0]["MC_GROUP"], grp.iloc[0]["MC_GUAGE"])] = min(weeks)

        def _recalc_old_mc(row):
            try:
                qty = float(row.get("PRODUCE_QTY", 0) or 0)
                cap = float(row.get("DAILY_CAPACITY", 0) or 0)
                week_val = row.get("PLAN_WEEK")
                mc_group = str(row.get("MC_GROUP", "")).strip()
                item_code = str(row.get("ITEM_CODE", "")).strip().upper()
                item_material = str(row.get("MATERIAL_CONTENT", "")).strip()
                item_gauge = str(row.get("MC_GUAGE", "")).strip()
                if qty <= 0 or cap <= 0 or pd.isna(week_val):
                    return row.get("ACTUAL_MC", 0), row.get("REQUIRED_MC", 0)
                week = int(week_val)
                # วันทำงานจากชีท Work Day เป็นค่าสุดท้าย (ไม่ cap/หักวันหยุด)
                actual_wd = get_working_days_by_factory(mc_group, 1, week=week, item_code=item_code, gauge=item_gauge)
                if actual_wd <= 0:
                    return row.get("ACTUAL_MC", 0), row.get("REQUIRED_MC", 0)
                # 🔧 FIX: คำนวณ setup days เฉพาะ week แรกของ item เท่านั้น
                # Week ถัดไป = carryover ไม่ต้อง setup (ใช้ capacity เต็ม)
                _key = (item_code, mc_group, item_gauge)
                _first_week = _first_week_map.get(_key)
                is_first_week = (_first_week is not None and week == _first_week)
                if is_first_week:
                    _row_yarn = str(row.get("YARN-USED", "") or row.get("YARN_USED", "") or _yarn_used_lookup.get(item_code, "")).strip()
                    _row_mat = item_material or _material_content_lookup.get(item_code, "")
                    setup_days = get_setup_days_for_item(_row_mat, _row_yarn, mc_group=mc_group, item_code=item_code)
                    productive_wd = max(0, actual_wd - setup_days)
                else:
                    # Week ถัดไป = carryover ไม่ต้อง setup
                    productive_wd = actual_wd
                if productive_wd <= 0:
                    # ถ้า setup days มากกว่า working days → ไม่สามารถผลิตได้
                    return row.get("ACTUAL_MC", 0), row.get("REQUIRED_MC", 0)
                needed = qty / (productive_wd * cap)
                mc = max(
                    1, int(needed) + (1 if (needed - int(needed)) > 1e-9 else 0)
                )
                return mc, mc
            except Exception:
                return row.get("ACTUAL_MC", 0), row.get("REQUIRED_MC", 0)
        old_booking_df[["ACTUAL_MC", "REQUIRED_MC"]] = old_booking_df.apply(
            lambda r: pd.Series(_recalc_old_mc(r)), axis=1
        )

    # เก็บแถว OLD ทั้งหมด (ไม่จำกัดสัปดาห์)
    old_booking_df = old_booking_df.sort_values(
        ["ITEM_CODE", "MC_GROUP", "PLAN_WEEK"], na_position="last"
    ).reset_index(drop=True)
    # เพิ่ม PLAN_YEAR ให้ old booking rows
    if "YEAR" in old_booking_df.columns:
        old_booking_df["PLAN_YEAR"] = pd.to_numeric(old_booking_df["YEAR"], errors="coerce").fillna(TODAY.year).astype(int)
    elif "PLAN_WEEK" in old_booking_df.columns:
        old_booking_df["PLAN_YEAR"] = old_booking_df["PLAN_WEEK"].map(_week_year_lookup).fillna(TODAY.year).astype(int)
    print(
        f"📦 OLD rows จาก booking_final_ready25 (ทั้งหมด): {len(old_booking_df)} rows"
    )
else:
    print("⚠️ ไม่พบข้อมูลใน detail_mc → ข้าม OLD")
# รวม OLD + NEW โดยใช้ common columns เรียงตาม ITEM_CODE, MC_GROUP, PLAN_WEEK
# เพื่อให้เห็น OLD vs NEW week-by-week ของ item เดียวกันติดกัน
if not old_booking_df.empty:
    common_cols = ["PLAN_SOURCE"] + [
        c for c in new_df.columns if c in old_booking_df.columns and c != "PLAN_SOURCE"
    ]
    # บังคับให้ ITEM_CODE, MC_GROUP, PLAN_WEEK อยู่ใน common_cols เสมอ (ถ้ามีใน old)
    for _must in ["ITEM_CODE", "MC_GROUP", "PLAN_WEEK"]:
        if _must not in common_cols and _must in old_booking_df.columns:
            common_cols.append(_must)
    if "TARGET_KNIT" not in common_cols and "TARGET_KNIT" in new_df.columns:
        common_cols.append("TARGET_KNIT")
    for _cyl_col in ["CYLINDER_CHANGE"]:
        if _cyl_col not in common_cols and _cyl_col in new_df.columns:
            common_cols.append(_cyl_col)
        # คำนวณ TARGET_KNIT สำหรับ OLD rows โดย lookup จาก orders (FG Week → TARGET_KNIT week)
        def _lookup_target_knit_old(row):
            _so = str(row.get("SC_SO_NO", "")).strip().lstrip("Ss")
            _item = str(row.get("ITEM_CODE", "")).strip().upper()
            _match = orders[orders["SC/SO NO"].astype(str).str.strip() == _so]
            if _match.empty:
                return None
            _r = _match.iloc[0]
            _fg = _r.get("FG Week")
            _otype = str(_r.get("Orders Type", "")).strip().upper()
            if pd.isna(_fg):
                return None
            try:
                _fg_str = str(int(_fg))
                if len(_fg_str) >= 5:
                    _yr, _wk = int(_fg_str[:4]), int(_fg_str[4:])
                else:
                    _yr, _wk = TODAY.year, int(_fg_str)
                _crow = calendar_week[
                    (calendar_week["YEAR"] == _yr) & (calendar_week["WEEK"] == _wk)
                ]
                if _crow.empty:
                    return None
                _raw_idx = _crow.index[0]
                if _otype == "LAB-DIP":
                    _rdd_idx = min(len(calendar_week) - 1, TODAY_IDX + 2)
                else:
                    _rdd_idx = max(0, _raw_idx - target_offset_weeks_for_row(_r))
                return int(calendar_week.iloc[_rdd_idx]["WEEK"])
            except Exception:
                return None
        old_booking_df["TARGET_KNIT"] = old_booking_df.apply(
            _lookup_target_knit_old, axis=1
        )
    if "IS_CORE_ITEM" not in common_cols and "IS_CORE_ITEM" in new_df.columns:
        common_cols.append("IS_CORE_ITEM")
        old_booking_df["IS_CORE_ITEM"] = ""
    if "CUSTOMER" not in common_cols and "CUSTOMER" in new_df.columns:
        common_cols.append("CUSTOMER")

        def _lookup_customer_old(row):
            _so = str(row.get("SC_SO_NO", "")).strip().lstrip("Ss")
            _match = orders[orders["SC/SO NO"].astype(str).str.strip() == _so]
            if _match.empty:
                return ""
            return str(_match.iloc[0].get("Customer", "")).strip()
        old_booking_df["CUSTOMER"] = old_booking_df.apply(
            _lookup_customer_old, axis=1
        )
    if "ORDER_TYPE" not in common_cols and "ORDER_TYPE" in new_df.columns:
        common_cols.append("ORDER_TYPE")

        def _lookup_order_type_old(row):
            _so = str(row.get("SC_SO_NO", "")).strip().lstrip("Ss")
            _match = orders[orders["SC/SO NO"].astype(str).str.strip() == _so]
            if _match.empty:
                return ""
            return str(_match.iloc[0].get("Orders Type", "")).strip()

        old_booking_df["ORDER_TYPE"] = old_booking_df.apply(
            _lookup_order_type_old, axis=1
        )
    if "FG_WEEK" not in common_cols and "FG_WEEK" in new_df.columns:
        common_cols.append("FG_WEEK")

        def _lookup_fg_week_old(row):
            _so = str(row.get("SC_SO_NO", "")).strip().lstrip("Ss")
            _match = orders[orders["SC/SO NO"].astype(str).str.strip() == _so]
            if _match.empty:
                return None

            _fg = _match.iloc[0].get("FG Week")
            return None if pd.isna(_fg) else _fg

        old_booking_df["FG_WEEK"] = old_booking_df.apply(_lookup_fg_week_old, axis=1)
    if "ORDERS_QTY" not in common_cols and "ORDERS_QTY" in new_df.columns:
        common_cols.append("ORDERS_QTY")

        def _lookup_orders_qty_old(row):
            _so = str(row.get("SC_SO_NO", "")).strip().lstrip("Ss")
            _match = orders[orders["SC/SO NO"].astype(str).str.strip() == _so]
            if _match.empty:
                return None
            return pd.to_numeric(_match.iloc[0].get("Orders.Qty"), errors="coerce")

        old_booking_df["ORDERS_QTY"] = old_booking_df.apply(
            _lookup_orders_qty_old, axis=1
        )
    if "PO_NO" not in common_cols and "PO_NO" in new_df.columns:
        common_cols.append("PO_NO")
        def _lookup_po_no_old(row):
            _so = str(row.get("SC_SO_NO", "")).strip().lstrip("Ss")
            _match = orders[orders["SO_NO"].astype(str).str.strip() == _so]
            if _match.empty:
                _match = orders[orders["SC/SO NO"].astype(str).str.strip() == _so]
            if _match.empty:
                return ""
            return str(_match.iloc[0].get("PO_NO", "")).strip()

        old_booking_df["PO_NO"] = old_booking_df.apply(_lookup_po_no_old, axis=1)
    if "RDD_WEEK" not in common_cols and "RDD_WEEK" in new_df.columns:
        common_cols.append("RDD_WEEK")

        def _lookup_rdd_week_old(row):
            _so = str(row.get("SC_SO_NO", "")).strip().lstrip("Ss")
            _match = orders[orders["SO_NO"].astype(str).str.strip() == _so]
            if _match.empty:
                _match = orders[orders["SC/SO NO"].astype(str).str.strip() == _so]
            if _match.empty:
                return None
            _fg = _match.iloc[0].get("FG Week")
            return None if pd.isna(_fg) else _fg
        old_booking_df["RDD_WEEK"] = old_booking_df.apply(_lookup_rdd_week_old, axis=1)
    if "SC_LINE_ID" not in common_cols and "SC_LINE_ID" in new_df.columns:
        common_cols.append("SC_LINE_ID")

        def _lookup_sc_line_id_old(row):
            _so = str(row.get("SC_SO_NO", "")).strip().lstrip("Ss")
            _match = orders[orders["SO_NO"].astype(str).str.strip() == _so]
            if _match.empty:
                _match = orders[orders["SC/SO NO"].astype(str).str.strip() == _so]
            if _match.empty:
                return ""
            return str(_match.iloc[0].get("SC/SO NO", "")).strip()

        old_booking_df["SC_LINE_ID"] = old_booking_df.apply(_lookup_sc_line_id_old, axis=1)
    combined_df = pd.concat(
        [
            old_booking_df[[c for c in common_cols if c in old_booking_df.columns]],
            new_df[[c for c in common_cols if c in new_df.columns]],
        ],
        ignore_index=True,
    )
    _sort_cols = [
        c
        for c in ["ITEM_CODE", "MC_GROUP", "PLAN_WEEK", "PLAN_SOURCE"]
        if c in combined_df.columns
    ]
    combined_df = combined_df.sort_values(_sort_cols, ignore_index=True)
else:
    combined_df = new_df
# =========================
# คอลัมน์ TARGET_STATUS: ทัน / ไม่ทัน ตาม TARGET_KNIT
# ดูว่า **ผลิตจนครบ order qty** ทันหรือไม่ (ใช้ PLAN_WEEK สุดท้ายของ order)
# ไม่ใช่ดูแค่แต่ละ row — ต้องดูว่า row สุดท้ายของ order ทันตาม target ถึงจะ "ทัน"
# =========================
_grp_cols = ["SC_SO_NO", "ITEM_CODE", "FG_WEEK"]
_grp_cols_available = [c for c in _grp_cols if c in combined_df.columns]

# หา PLAN_WEEK สุดท้าย (ผลิตเสร็จ) ของแต่ละ order group
# ใช้ NEW rows เป็นหลัก (plan ใหม่) — ถ้าไม่มี NEW ถึงใช้ OLD
_last_pw_map = {}  # {normalized_str_key: (last_plan_week, last_plan_week_index)}
if _grp_cols_available:
    # Step 1: คำนวณจาก NEW rows ก่อน (source of truth)
    if "PLAN_SOURCE" in combined_df.columns:
        _new_only = combined_df[combined_df["PLAN_SOURCE"].astype(str).str.upper() == "NEW"]
    else:
        _new_only = combined_df
    if not _new_only.empty:
        for _key, _grp in _new_only.groupby(_grp_cols_available):
            _pws = _grp["PLAN_WEEK"].dropna()
            if _pws.empty:
                continue
            try:
                _last_pw = int(_pws.max())
                _last_idx = week_index(_last_pw)
                if _last_idx is not None:
                    _norm_key = tuple(str(v).strip() for v in (_key if isinstance(_key, tuple) else (_key,)))
                    _last_pw_map[_norm_key] = (_last_pw, _last_idx)
            except Exception:
                continue
    # Step 2: เติม order ที่ไม่มี NEW rows → ใช้ OLD rows
    if "PLAN_SOURCE" in combined_df.columns:
        _old_only = combined_df[combined_df["PLAN_SOURCE"].astype(str).str.upper() != "NEW"]
        if not _old_only.empty:
            for _key, _grp in _old_only.groupby(_grp_cols_available):
                _norm_key = tuple(str(v).strip() for v in (_key if isinstance(_key, tuple) else (_key,)))
                if _norm_key in _last_pw_map:
                    continue  # มี NEW แล้ว ไม่ต้อง overwrite
                _pws = _grp["PLAN_WEEK"].dropna()
                if _pws.empty:
                    continue
                try:
                    _last_pw = int(_pws.max())
                    _last_idx = week_index(_last_pw)
                    if _last_idx is not None:
                        _last_pw_map[_norm_key] = (_last_pw, _last_idx)
                except Exception:
                    continue

def _target_status(row) -> str:
    # CORE ITEM ไม่ต้องบอกทัน/ไม่ทัน
    if str(row.get("IS_CORE_ITEM", "")).strip().upper() == "CORE ITEM":
        return ""

    _tk = row.get("TARGET_KNIT")
    _pw = row.get("PLAN_WEEK")
    if pd.isna(_pw) or pd.isna(_tk):
        return "-"

    try:
        _tk_idx = week_index(int(_tk))
        if _tk_idx is None:
            return "-"

        # สร้าง key (str) เพื่อ lookup PLAN_WEEK สุดท้ายของ order นี้
        _key_vals = tuple(str(row.get(c, "")).strip() for c in _grp_cols_available)
        _finish = _last_pw_map.get(_key_vals)
        if _finish is not None:
            _last_pw_val, _last_pw_idx = _finish
            if _last_pw_idx <= _tk_idx:
                return "ทัน"
            else:
                return f"ไม่ทัน (+{_last_pw_idx - _tk_idx} wk)"

        # fallback: ถ้าไม่มี group → เทียบ row เดียว
        _pw_idx = week_index(int(_pw))
        if _pw_idx is None:
            return "-"
        if _pw_idx <= _tk_idx:
            return "ทัน"
        else:
            return f"ไม่ทัน (+{_pw_idx - _tk_idx} wk)"
    except Exception:
        return "-"

combined_df["TARGET_STATUS"] = combined_df.apply(_target_status, axis=1)
# Reorder columns: move PLAN_YEAR to column AL (index 37) and COLOR_DESC after it
print(f"DEBUG: Total columns before reordering: {len(combined_df.columns)}")
print(f"DEBUG: Columns before reordering: {list(combined_df.columns)}")
if "PLAN_YEAR" in combined_df.columns:
    cols = list(combined_df.columns)
    plan_year_idx = cols.index("PLAN_YEAR")
    print(f"DEBUG: PLAN_YEAR is currently at index {plan_year_idx}")
    # Remove PLAN_YEAR from current position
    cols.remove("PLAN_YEAR")
    print(f"DEBUG: After removing PLAN_YEAR, total columns: {len(cols)}")
    # Insert PLAN_YEAR at index 37 (column AL)
    # If we have fewer than 38 columns, add padding columns first
    if len(cols) < 38:
        print(f"DEBUG: Only {len(cols)} columns, adding padding to reach 38")
        for i in range(38 - len(cols)):
            cols.append(f"_PAD_{i}")
    target_idx = 37
    cols.insert(target_idx, "PLAN_YEAR")
    # Move COLOR_DESC after PLAN_YEAR if it exists
    if "COLOR_DESC" in cols:
        cols.remove("COLOR_DESC")
        cols.insert(target_idx + 1, "COLOR_DESC")

    # Remove padding columns
    cols = [c for c in cols if not c.startswith("_PAD_")]
    combined_df = combined_df[cols]
    print(f"DEBUG: Columns after reordering: {list(combined_df.columns)}")
    print(f"DEBUG: PLAN_YEAR is now at index {cols.index('PLAN_YEAR')} (column {chr(65 + cols.index('PLAN_YEAR'))})")
else:
    print(f"DEBUG: PLAN_YEAR not found in columns")
# =========================
# Write COMBINED_FILE
# =========================
# ❌ DISABLED: auto-backup ผลรอบก่อน → committed_plan.xlsx (feedback loop)
# เดิมโค้ดตรงนี้ก็อป COMBINED_FILE (= ผลลัพธ์รอบก่อน) มาเป็น committed_plan.xlsx
# แล้วรอบถัดไปจะอ่าน committed_plan.xlsx กลับเข้ามาเป็น old-plan carryover
# → เกิด feedback loop: แผนรอบใหม่ขึ้นกับแผนรอบก่อน ทำให้ผลลัพธ์ "ไม่นิ่ง"
#   (booking/order เดิมเป๊ะ ๆ แต่รันคนละครั้งได้คนละสัปดาห์ เช่น W34 บ้าง W36 บ้าง)
# ตัดออกเพื่อให้ทุกครั้งคำนวณสดจาก booking + order → deterministic
# ⚠️ ถ้าจะทำฟีเจอร์ "commit แผน" จริง ให้เขียน committed_plan.xlsx เฉพาะตอน user กด commit เท่านั้น
# ไม่ใช่ auto ทุกครั้งที่รัน Planning
with pd.ExcelWriter(COMBINED_FILE, engine="openpyxl") as writer:
    combined_df.to_excel(writer, sheet_name="PLAN", index=False)
    if not _no_cap_df.empty:
        _no_cap_df.to_excel(writer, sheet_name="NO_CAP", index=False)
    if not _multi_cap_df.empty:
        _multi_cap_df.to_excel(writer, sheet_name="MULTI_CAP", index=False)
print(f"Combined (OLD+NEW): {COMBINED_FILE}")
print(f"  OLD rows: {len(old_booking_df) if not old_booking_df.empty else 0}")
print(f"  NEW rows: {len(plan_df)}")
if not _no_cap_df.empty:
    print(f"  NO_CAP items: {_no_cap_df['Item Code'].nunique()} items, {len(_no_cap_df)} orders → sheet 'NO_CAP'")
if not _multi_cap_df.empty:
        print(f"  MULTI_CAP items: {_multi_cap_df['Item Code'].nunique()} items, {len(_multi_cap_df)} orders → sheet 'MULTI_CAP' (SINGLE MC_TYPE หรือ OM/OMNOI)")



# =========================
# SETUP_TRACKING sheet — แสดงเฉพาะ ITEM ที่หัก Job พร้อม Week, จำนวน Job และจำนวน MC ที่ Setup
# เกณฑ์: หัก Job เมื่อ NEW_MC > 0 (เครื่องใหม่ที่ต้อง setup = 1 job ต่อ 1 เครื่อง)
# =========================
# --- NEW PLAN: หัก Job เมื่อ NEW_MC > 0 (แยกเป็นฟังก์ชันเพื่อสร้างทั้ง with-S9 และ no-S9) ---
def _new_setup_rows(plan_frame):
    """สร้าง setup rows ฝั่ง NEW จากแผนที่ระบุ (NEW_MC > 0)"""
    _rows = []
    if not plan_frame.empty and "NEW_MC" in plan_frame.columns:
        _new_setup = plan_frame[pd.to_numeric(plan_frame["NEW_MC"], errors="coerce").fillna(0) > 0].copy()
        for _, row in _new_setup.iterrows():
            _new_mc_val = int(pd.to_numeric(row["NEW_MC"], errors="coerce") or 0)
            _carryover_val = int(pd.to_numeric(row.get("CARRYOVER_MC", 0), errors="coerce") or 0)
            _type_key_new = _get_type_key_for_mc(str(row["MC_GROUP"]).strip().upper())
            _rows.append({
                "PLAN_SOURCE": "NEW",
                "PLAN_WEEK": row["PLAN_WEEK"],
                "PLAN_YEAR": row.get("PLAN_YEAR"),
                "ITEM_CODE": row["ITEM_CODE"],
                "SC_SO_NO": row["SC_SO_NO"],
                "MC_GROUP": row["MC_GROUP"],
                "MC_GUAGE": row.get("MC_GUAGE", ""),
                "TYPE": _type_key_new,
                "JOBS_DEDUCTED": _new_mc_val,
                "SETUP_MC": _new_mc_val,
                "CARRYOVER_MC": _carryover_val,
                "MC_THIS_WEEK": _new_mc_val + _carryover_val,
                "MC_PREV_WEEK": _carryover_val,
                "SETUP_DAYS": row.get("SETUP_DAYS", 0),
                "KP_WEIGHT": pd.to_numeric(row.get("PRODUCE_QTY", 0), errors="coerce") or 0,
                "DAILY_CAPACITY": row.get("DAILY_CAPACITY", 0),
            })
    return _rows

# --- OLD PLAN (BOOKING): ใช้ _IS_NEW_SETUP / _MC_INCREASE จาก AVA_MC.py ---
# OLD มาจาก booking baseline (detail_mc) เหมือนกันทั้ง 2 pass → สร้างครั้งเดียว ใช้ร่วม
_old_setup_rows = []
# ตรงกับ logic ของ weekly_job_usage_old เพื่อให้ตัวเลขสอดคล้องกับ REMAINING_JOBS
if (
    not detail_mc.empty
    and "WEEK" in detail_mc.columns
    and "ITEM_CODE" in detail_mc.columns
    and "MC_USE_CEIL" in detail_mc.columns
    and "MC_GROUP" in detail_mc.columns
):
    _det_trk = detail_mc.copy()
    _det_trk["WEEK"] = pd.to_numeric(_det_trk["WEEK"], errors="coerce")
    _det_trk["MC_USE_CEIL"] = (
        pd.to_numeric(_det_trk["MC_USE_CEIL"], errors="coerce").fillna(0).astype(int)
    )
    _det_trk = _det_trk.dropna(subset=["WEEK", "ITEM_CODE", "MC_GROUP"])
    _det_trk["WEEK"] = _det_trk["WEEK"].astype(int)
    _det_trk["ITEM_CODE"] = _det_trk["ITEM_CODE"].astype(str).str.strip().str.upper()
    _det_trk["MC_GROUP"] = _det_trk["MC_GROUP"].astype(str).str.strip().str.upper()
    _det_trk_active = _det_trk[_det_trk["MC_USE_CEIL"] > 0].copy()
    # ตัด IMPORT ออกด้วยเกณฑ์เดียวกับ weekly_job_usage — SETUP_TRACKING ต้องนับตรงกับตัวคุมโควตา
    _det_trk_active = _det_trk_active[~_det_trk_active["MC_GROUP"].isin(_JOB_EXCLUDE_MC)]

    _has_ava_flags_trk = "_IS_NEW_SETUP" in _det_trk.columns and "_MC_INCREASE" in _det_trk.columns

    if _has_ava_flags_trk:
        _det_trk_active["_IS_NEW_SETUP"] = _det_trk_active["_IS_NEW_SETUP"].fillna(True).astype(bool)
        _det_trk_active["_MC_INCREASE"] = pd.to_numeric(_det_trk_active["_MC_INCREASE"], errors="coerce").fillna(0).astype(int)

        for _, _drow_t in _det_trk_active.iterrows():
            _is_new_t = bool(_drow_t["_IS_NEW_SETUP"])
            _mc_ceil_t = int(_drow_t["MC_USE_CEIL"])
            _mc_inc_t = int(_drow_t["_MC_INCREASE"])
            _new_jobs_t = _mc_ceil_t if _is_new_t else _mc_inc_t
            if _new_jobs_t <= 0:
                continue

            _carryover_t = 0 if _is_new_t else (_mc_ceil_t - _mc_inc_t)
            _wk_t = int(_drow_t["WEEK"])
            _mc_grp_t = str(_drow_t["MC_GROUP"]).strip().upper()
            _item_t = str(_drow_t["ITEM_CODE"]).strip().upper()

            _gauge_t = str(_drow_t.get("GUAGE", "")).strip() if "GUAGE" in _drow_t.index else ""
            _so_t = ""
            for _so_col in ("SC_SO_NO", "SO_NO"):
                if _so_col in _drow_t.index:
                    _raw_so_t = str(_drow_t.get(_so_col, "")).strip()
                    _so_t = _raw_so_t.lstrip("Ss")
                    break

            _cust_t = ""
            if _so_t:
                _ord_t = orders[orders["SC/SO NO"].astype(str).str.strip() == _so_t]
                if not _ord_t.empty:
                    _cust_t = str(_ord_t.iloc[0].get("Customer", "")).strip()

            _raw_prev = pd.to_numeric(_drow_t.get("_PREV_MC_USE_CEIL", _carryover_t), errors="coerce")
            _prev_mc_t = 0 if pd.isna(_raw_prev) else int(_raw_prev)
            _kp_weight_t = pd.to_numeric(_drow_t.get("KP_WEIGHT", 0), errors="coerce") or 0
            _type_key_old = _get_type_key_for_mc(_mc_grp_t)
            _old_setup_rows.append({
                "PLAN_SOURCE": "OLD",
                "PLAN_WEEK": _wk_t,
                "PLAN_YEAR": _week_year_lookup.get(_wk_t, TODAY.year),
                "ITEM_CODE": _item_t,
                "SC_SO_NO": _so_t,
                "MC_GROUP": _mc_grp_t,
                "MC_GUAGE": _gauge_t,
                "TYPE": _type_key_old,
                "JOBS_DEDUCTED": _new_jobs_t,
                "SETUP_MC": _new_jobs_t,
                "CARRYOVER_MC": _carryover_t,
                "MC_THIS_WEEK": _mc_ceil_t,
                "MC_PREV_WEEK": _prev_mc_t,
                "SETUP_DAYS": 0,
                "KP_WEIGHT": _kp_weight_t,
                "DAILY_CAPACITY": 0,
            })
    else:
        # Fallback: week-over-week comparison (กรณีไม่มี AVA_MC.py flags)
        for _mc_grp, _grp_df in _det_trk_active.groupby("MC_GROUP"):
            _all_weeks_trk = sorted(_grp_df["WEEK"].unique())
            _week_item_mc_trk: dict = {}
            for _wk_t in _all_weeks_trk:
                _wk_rows_t = _grp_df[_grp_df["WEEK"] == _wk_t]
                _week_item_mc_trk[_wk_t] = (
                    _wk_rows_t.groupby("ITEM_CODE")["MC_USE_CEIL"].sum().to_dict()
                )
            for _i_t, _wk_t in enumerate(_all_weeks_trk):
                _curr_items_t = _week_item_mc_trk[_wk_t]
                _prev_items_t = (
                    _week_item_mc_trk.get(_all_weeks_trk[_i_t - 1], {}) if _i_t > 0 else {}
                )
                for _item_t, _mc_t in _curr_items_t.items():
                    _prev_mc_t = _prev_items_t.get(_item_t, 0)
                    if _prev_mc_t == 0:
                        _new_jobs_t = _mc_t
                    elif _mc_t > _prev_mc_t:
                        _new_jobs_t = _mc_t - _prev_mc_t
                    else:
                        continue
                    _item_rows_t = _grp_df[
                        (_grp_df["WEEK"] == _wk_t) & (_grp_df["ITEM_CODE"] == _item_t)
                    ]
                    _gauge_t = ""
                    _so_t = ""
                    if not _item_rows_t.empty:
                        if "GUAGE" in _item_rows_t.columns:
                            _gauge_t = str(_item_rows_t.iloc[0].get("GUAGE", "")).strip()
                        for _so_col in ("SC_SO_NO", "SO_NO"):
                            if _so_col in _item_rows_t.columns:
                                _raw_so_t = str(_item_rows_t.iloc[0].get(_so_col, "")).strip()
                                _so_t = _raw_so_t.lstrip("Ss")
                                break
                    _cust_t = ""
                    if _so_t:
                        _ord_t = orders[orders["SC/SO NO"].astype(str).str.strip() == _so_t]
                        if not _ord_t.empty:
                            _cust_t = str(_ord_t.iloc[0].get("Customer", "")).strip()
                    _kp_weight_fb = pd.to_numeric(_item_rows_t.iloc[0].get("KP_WEIGHT", 0), errors="coerce") if not _item_rows_t.empty and "KP_WEIGHT" in _item_rows_t.columns else 0
                    _type_key_old = _get_type_key_for_mc(str(_mc_grp).strip().upper())
                    _old_setup_rows.append({
                        "PLAN_SOURCE": "OLD",
                        "PLAN_WEEK": _wk_t,
                        "PLAN_YEAR": _week_year_lookup.get(int(_wk_t), TODAY.year),
                        "ITEM_CODE": _item_t,
                        "SC_SO_NO": _so_t,
                        "MC_GROUP": str(_mc_grp).strip().upper(),
                        "MC_GUAGE": _gauge_t,
                        "TYPE": _type_key_old,
                        "JOBS_DEDUCTED": _new_jobs_t,
                        "SETUP_MC": _new_jobs_t,
                        "CARRYOVER_MC": _prev_mc_t,
                        "MC_THIS_WEEK": _mc_t,
                        "MC_PREV_WEEK": _prev_mc_t,
                        "SETUP_DAYS": 0,
                        "KP_WEIGHT": _kp_weight_fb or 0,
                        "DAILY_CAPACITY": 0,
                    })



def _finalize_setup_df(rows):
    """ประกอบ setup_tracking_df จาก rows (NEW + OLD): + PLAN_YEAR, sort, เติม DAILY_CAPACITY"""
    _df = (
        pd.DataFrame(rows)
        if rows
        else pd.DataFrame(columns=[
            "PLAN_SOURCE", "PLAN_WEEK", "PLAN_YEAR", "ITEM_CODE", "SC_SO_NO",
            "MC_GROUP", "MC_GUAGE", "TYPE", "JOBS_DEDUCTED", "SETUP_MC", "CARRYOVER_MC",
            "MC_THIS_WEEK", "MC_PREV_WEEK", "SETUP_DAYS", "KP_WEIGHT", "DAILY_CAPACITY",
        ])
    )
    if not _df.empty and "PLAN_WEEK" in _df.columns:
        _df["PLAN_YEAR"] = _df["PLAN_WEEK"].map(_week_year_lookup).fillna(TODAY.year).astype(int)
    if not _df.empty:
        _df = _df.sort_values(["PLAN_WEEK", "PLAN_SOURCE", "ITEM_CODE"], ignore_index=True)
    # เติม DAILY_CAPACITY สำหรับ OLD PLAN (จาก detail_mc) ที่ยังไม่มีค่า
    if not _df.empty and "DAILY_CAPACITY" in _df.columns:
        _df["DAILY_CAPACITY"] = _df["DAILY_CAPACITY"].fillna(0)
    return _df

setup_tracking_df = _finalize_setup_df(_new_setup_rows(plan_df) + _old_setup_rows)


if not setup_tracking_df.empty:
    print(f"\nDEBUG: setup_tracking_df shape = {setup_tracking_df.shape}")

print("📋 สรุป SETUP_TRACKING (ITEM ที่หัก Job) ต่อ Week:")
if not setup_tracking_df.empty:
    for _wk_s in sorted(setup_tracking_df["PLAN_WEEK"].unique()):
        _wdf_s = setup_tracking_df[setup_tracking_df["PLAN_WEEK"] == _wk_s]
        _total_jobs_s = int(_wdf_s["JOBS_DEDUCTED"].sum())
        _total_mc_s = int(_wdf_s["SETUP_MC"].sum())
        print(f"   Week {_wk_s}: {len(_wdf_s)} items — หัก {_total_jobs_s} Jobs, Setup {_total_mc_s} MC")
        for _, _r_s in _wdf_s.iterrows():
            _src_icon = "🆕" if _r_s["PLAN_SOURCE"] == "NEW" else "📋"
            print(
                f"     {_src_icon} [{_r_s['PLAN_SOURCE']}] {_r_s['ITEM_CODE']}"
                f" ({_r_s.get('SC_SO_NO', '')}) - {_r_s['MC_GROUP']}"
                f" — หัก {_r_s['JOBS_DEDUCTED']} Job, Setup {_r_s['SETUP_MC']} MC"
            )
    print()
else:
    print("   (ไม่มีการหัก Job ในแผนนี้)")
    print()

# =========================
# สรุป CYLINDER CHANGE — แสดงการเปลี่ยน gauge ของเครื่องที่ว่าง (factory-level)
# =========================
# สร้าง reverse lookup: (week, factory, mc_cat, tgt_g) → [item_codes]
_cyl_item_lookup: dict = {}
_cyl_src_lookup: dict = {}  # (week, factory, mc_cat, tgt_g) → src_g
for (_ciw, _cii, _cimg), (_cisrc, _citgt) in _cylinder_change_for_item.items():
    _cif = _mc_to_factory(_cimg, _citgt)
    _cicat = _mc_to_type1(_cimg, _citgt)
    _lk = (int(_ciw), _cif, _cicat, _citgt)
    _cyl_item_lookup.setdefault(_lk, []).append(_cii)
    _cyl_src_lookup[_lk] = _cisrc

print("🔄 สรุป CYLINDER CHANGE (เปลี่ยนเครื่องที่ว่าง):")
if _cyl_item_lookup:
    for _lk in sorted(_cyl_item_lookup.keys()):
        _ciw2, _cif2, _cicat2, _citgt2 = _lk
        _cisrc2 = _cyl_src_lookup.get(_lk, "?")
        _n_mc = len(_cyl_item_lookup[_lk])
        _items_for = ", ".join(_cyl_item_lookup[_lk])
        print(f"   W{_ciw2}  Factory: {_cif2}  MC_CAT: {_cicat2}  Gauge {_cisrc2} → Gauge {_citgt2}  ({_n_mc} เครื่อง)  → {_items_for}")
else:
    print("   ✅ ไม่มี CYLINDER CHANGE ในแผนนี้")
print()

# บันทึกไฟล์ใหม่ (รวม SETUP_TRACKING)
if not plan_df.empty:
    _dbg_item = "FD6PRTPG99A0"
    _dbg_rows = plan_df[
        plan_df["ITEM_CODE"].astype(str).str.strip().str.upper() == _dbg_item
    ]
    if not _dbg_rows.empty:
        _dbg_view = _dbg_rows[[
            "SC_SO_NO",
            "PLAN_WEEK",
            "MC_GROUP",
            "PRODUCE_QTY",
            "SETUP_DAYS",
            "CARRYOVER_MC",
            "NEW_MC",
            "ACTUAL_MC",
        ]].sort_values(["PLAN_WEEK", "SC_SO_NO"])
        print(_dbg_view.to_string(index=False))



# สร้าง DataFrame สำหรับ CYLINDER_CHANGE sheet (factory-level) — per trigger week
_cyl_rows = []
for _lk in sorted(_cyl_item_lookup.keys()):
    _ciw2, _cif2, _cicat2, _citgt2 = _lk
    _cisrc2 = _cyl_src_lookup.get(_lk, "?")
    _cyl_rows.append({
        "WEEK": _ciw2,
        "FACTORY": _cif2,
        "MC_CAT": _cicat2,
        "GAUGE_FROM": _cisrc2,
        "GAUGE_TO": _citgt2,
        "MC_CHANGED": len(_cyl_item_lookup[_lk]),
        "ITEM_CODE": ", ".join(_cyl_item_lookup[_lk]),
    })
_cylinder_change_df = pd.DataFrame(_cyl_rows) if _cyl_rows else pd.DataFrame(
    columns=["WEEK", "FACTORY", "MC_CAT", "GAUGE_FROM", "GAUGE_TO", "MC_CHANGED", "ITEM_CODE"]
)

# ===== FOLD CHECK: ยอดที่ลูกค้าเปิดมา แบ่งเป็นพับแล้วเป็นพับคู่ (หาร 2 ลงตัว) หรือไม่ =====
# เช็คที่ "ยอดรวมทั้ง order" (ITEM_CODE + SC_SO_NO) — ไม่ใช่รายสัปดาห์
# → SC ไหนหารไม่ลงตัว ทุกสัปดาห์ของ SC นั้นติดธงเตือนเหมือนกันหมด
# เฉพาะกลุ่ม IRMT/SJT เท่านั้น (กลุ่มอื่นไม่มีกฎพับคู่)
# แผนวางเต็มตามยอดจริงเสมอ ค่าพวกนี้เป็นแค่ธงเตือนให้ planner ไปแก้ที่ยอดเปิด order
if not plan_df.empty and {"ORDERS_QTY", "REVOLUTION_WEIGHT", "MC_GROUP"} <= set(plan_df.columns):
    _fold_cache = {}
    for _fo_key, _fo_grp in plan_df.groupby(["ITEM_CODE", "SC_SO_NO"], dropna=False):
        _fold_cache[_fo_key] = fold_check_order(
            _fo_grp["ORDERS_QTY"].max(),
            _fo_grp["REVOLUTION_WEIGHT"].max(),
            _fo_grp["MC_GROUP"].unique(),
        )
    _fold_vals = [
        _fold_cache.get((_r_item, _r_sc), (0.0, 0.0, False))
        for _r_item, _r_sc in zip(plan_df["ITEM_CODE"], plan_df["SC_SO_NO"])
    ]
    plan_df["FOLD_QTY"] = [_v[0] for _v in _fold_vals]
    plan_df["FOLD_REMAINDER"] = [_v[1] for _v in _fold_vals]
    plan_df["FOLD_WARN"] = [1 if _v[2] else 0 for _v in _fold_vals]

    _fold_bad = sorted(k for k, v in _fold_cache.items() if v[2])
    _fold_scope = sum(
        1 for _k, _g in plan_df.groupby(["ITEM_CODE", "SC_SO_NO"], dropna=False)
        if any(str(_m).strip().upper() in FOLD_ROUND_MC_GROUPS for _m in _g["MC_GROUP"].unique())
    )
    print(f"\n📏 FOLD CHECK ({'/'.join(sorted(FOLD_ROUND_MC_GROUPS))}): "
          f"order ที่แบ่งพับแล้วหาร {FOLD_MULTIPLE} ไม่ลงตัว "
          f"{len(_fold_bad)}/{_fold_scope} order")
    for _fb_item, _fb_sc in _fold_bad:
        _fb_folds, _fb_rem, _ = _fold_cache[(_fb_item, _fb_sc)]
        print(f"   ⚠️  {_fb_item} SC {_fb_sc}: {_fb_folds:g} พับ → เหลือเศษ {_fb_rem:g} พับ")
    print()

# ทีมของ item (จากชีท Program ใน MasterMC) — ไม่มีในชีท = ว่าง
# หน้าเว็บใช้คู่ ITEM_CODE+TEAM ตัดสินว่าจะทำตัวหนังสือเป็นสีเหลืองไหม (ต้องตรงทั้งคู่)
for _pdf in (plan_df, setup_tracking_df, _unplanned_df):
    if _pdf is not None and not _pdf.empty and "ITEM_CODE" in _pdf.columns:
        _pdf["TEAM"] = _pdf["ITEM_CODE"].map(_lookup_item_team)
print(f"🏷️  เติมคอลัมน์ TEAM จากชีท Program: "
      f"{int((plan_df['TEAM'] != '').sum()) if 'TEAM' in plan_df.columns else 0}/{len(plan_df)} แถวในชีท PLAN")

# ย้ายคอลัมน์เสริมที่เพิ่มเข้ามา (วันทำงาน + ประเภทเครื่อง + REMARK) + S9_ROUTING
# ไปไว้ท้ายสุด เพื่อไม่ให้แทรกกลางคอลัมน์หลักที่คนอ่านแผนใช้ประจำ
_tail_cols = [
    c
    for c in ["WD_BASE", "WD_W32", "MC_SHARED", "MC_BOOKING", "REMARK", "LINE_REMARK", "S9_ROUTING", "OUTSOURCE", "YARNPO",
              "FOLD_QTY", "FOLD_REMAINDER", "FOLD_WARN", "TEAM"]
    if c in plan_df.columns
]
if _tail_cols:
    _other_cols = [c for c in plan_df.columns if c not in _tail_cols]
    plan_df = plan_df[_other_cols + _tail_cols]

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as _writer:
    plan_df.to_excel(_writer, sheet_name="PLAN", index=False)
    remaining_df.to_excel(_writer, sheet_name="REMAINING_JOBS", index=False)
    setup_tracking_df.to_excel(_writer, sheet_name="SETUP_TRACKING", index=False)
    _unplanned_df.to_excel(_writer, sheet_name="UNPLANNED", index=False)
    _cylinder_change_df.to_excel(_writer, sheet_name="CYLINDER_CHANGE", index=False)

# ใส่สีให้ PLAN sheet: เหลือง=CYLINDER_CHANGE, แดง=S9_ROUTING
_need_color = _cyl_trigger_keys or (
    "S9_ROUTING" in plan_df.columns and plan_df["S9_ROUTING"].any()
)
if _need_color:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    _yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    _red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    _wb = load_workbook(OUTPUT_FILE)
    _ws = _wb["PLAN"]
    _hdr = {cell.value: cell.column for cell in _ws[1]}
    _cyl_col_idx = _hdr.get("CYLINDER_CHANGE")
    _s9_col_idx = _hdr.get("S9_ROUTING")
    for _row in _ws.iter_rows(min_row=2, max_row=_ws.max_row):
        _is_s9_row = _s9_col_idx and str(_row[_s9_col_idx - 1].value).strip() in ("True", "TRUE", "1")
        _is_cyl_row = _cyl_col_idx and str(_row[_cyl_col_idx - 1].value).strip() == "Yes"
        if _is_s9_row:
            for _cell in _row:
                _cell.fill = _red_fill
        elif _is_cyl_row:
            for _cell in _row:
                _cell.fill = _yellow
    _wb.save(OUTPUT_FILE)


# =========================
# PIVOT_PLAN sheet — Excel PivotTable จริง (ผ่าน win32com)
# =========================
import time
if not plan_df.empty:
    try:
        import win32com.client
        import pythoncom
        # Add delay to ensure file is fully released by openpyxl
        time.sleep(1)
        pythoncom.CoInitialize()
        _excel_app = None
        try:
            _excel_app = win32com.client.Dispatch("Excel.Application")
            _excel_app.Visible = False
            _excel_app.DisplayAlerts = False
            _abs_path = str(OUTPUT_FILE.resolve())
            _wb_com = _excel_app.Workbooks.Open(_abs_path)
            # กำหนด data range บน PLAN sheet
            _plan_ws_com = _wb_com.Sheets("PLAN")
            _n_rows = plan_df.shape[0] + 1  # รวม header
            _n_cols = plan_df.shape[1]
            _data_range = _plan_ws_com.Range(
                _plan_ws_com.Cells(1, 1),
                _plan_ws_com.Cells(_n_rows, _n_cols),
            )
            # สร้าง PivotCache จาก PLAN sheet
            _pivot_cache = _wb_com.PivotCaches().Create(
                SourceType=1,  # xlDatabase
                SourceData=_data_range,
            )
            # ลบ PIVOT_PLAN sheet เดิม (ถ้ามี) แล้วสร้างใหม่
            for _sh in list(_wb_com.Sheets):
                if _sh.Name == "PIVOT_PLAN":
                    _sh.Delete()
                    break
            _pivot_ws_com = _wb_com.Sheets.Add(
                After=_wb_com.Sheets(_wb_com.Sheets.Count)
            )
            _pivot_ws_com.Name = "PIVOT_PLAN"
            # สร้าง PivotTable บนเซลล์ A3
            _pt = _pivot_cache.CreatePivotTable(
                TableDestination=_pivot_ws_com.Range("A3"),
                TableName="PivotPlan",
            )
            # Filter field: PLAN_SOURCE
            _pt.PivotFields("PLAN_SOURCE").Orientation = 3  # xlPageField
            _pt.PivotFields("PLAN_SOURCE").Position = 1
            # Row fields: ITEM_CODE → MC_GROUP → MC_GUAGE → SC_SO_NO → FG_WEEK → DAILY_CAPACITY → ORDERS_QTY
            for _pos, _fname in enumerate([
                "ITEM_CODE", "MC_GROUP", "MC_GUAGE", "SC_SO_NO", "FG_WEEK",
                "DAILY_CAPACITY", "ORDERS_QTY",
            ], start=1):
                _pt.PivotFields(_fname).Orientation = 1  # xlRowField
                _pt.PivotFields(_fname).Position = _pos
                # Remove subtotals for each row field
                _pt.PivotFields(_fname).Subtotals = [False] * 12  # xlNone for all subtotal types
            # Column field: PLAN_WEEK
            _pt.PivotFields("PLAN_WEEK").Orientation = 2  # xlColumnField
            _pt.PivotFields("PLAN_WEEK").Position = 1
            # Value field: Sum of PRODUCE_QTY
            _pt.AddDataField(
                _pt.PivotFields("PRODUCE_QTY"),
                "Sum of PRODUCE_QTY",
                -4157,  # xlSum
            )

            # Set Grand Totals for Rows Only

            _pt.ColumnGrand = False
            _pt.RowGrand = True
            _wb_com.Save()
            _wb_com.Close(False)
            print("✅ สร้าง PIVOT_PLAN (Excel PivotTable) สำเร็จ")
        finally:
            if _excel_app:
                _excel_app.Quit()
            pythoncom.CoUninitialize()
    except ImportError:
        print("⚠️ ไม่พบ win32com → ข้าม PivotTable")
    except Exception as _e:
        print(f"⚠️ สร้าง PivotTable ไม่สำเร็จ: {_e}")
        import traceback
        traceback.print_exc()