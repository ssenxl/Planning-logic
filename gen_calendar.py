"""
gen_calendar.py — เติมปฏิทินปีถัดไปลงชีท Sheet1 ของ Calendar.xlsx

จุดประสงค์:
    ให้ Calendar.xlsx มีข้อมูลรายวันของปีข้างหน้าเตรียมไว้ล่วงหน้า (default = วันทำงานปกติ)
    แล้ว user แค่มาแก้เองภายหลังว่า
      - วันไหนหยุด (แก้คอลัมน์ status เป็น 0)
      - จะรวม week ไหน (เช่น สงกรานต์ / ปีใหม่)

ค่า default ที่ generator ใส่ให้:
    - WEEK / YEAR  : คำนวณแบบ "ศุกร์-พฤหัส" (บวก 3 วัน -> ISO week) ตรงกับสูตรใน Calendar.load_calendar
    - status       : อาทิตย์ = 0 (หยุด), จันทร์-เสาร์ = 1 (ทำงาน)
    - MONTH/YW/YM/Day : เติมให้ครบตามรูปแบบไฟล์เดิม

ความปลอดภัย:
    - สำรองไฟล์เป็น Calendar.backup-YYYYMMDD-HHMM.xlsx ก่อนเขียนทุกครั้ง
    - ต่อท้ายเฉพาะ Sheet1 (ไม่แตะแถวเดิม / ชีท Work Day / Week Merge / format เดิม)
    - เคลียร์แถวขยะท้ายชีท (แถวที่ไม่มี DATE) ทิ้งก่อนเติม
    - --dry-run : แสดงว่าจะเติมอะไรบ้างโดยไม่เขียนไฟล์

การใช้งาน:
    python gen_calendar.py                 # เติมถึงสิ้นปี DEFAULT_END_YEAR แล้วเขียนไฟล์ (มี backup)
    python gen_calendar.py --dry-run       # ดูผลก่อน ไม่เขียน
    python gen_calendar.py --end-year 2032 # กำหนดปีปลายทางเอง
"""

from __future__ import annotations

import argparse
import datetime as dt
import shutil
import sys
from pathlib import Path

import openpyxl

SHEET = "Sheet1"
DEFAULT_END_YEAR = 2030  # เติมปฏิทินถึงสิ้นปีนี้ (แก้ได้ด้วย --end-year)
DEFAULT_YEARS_AHEAD = 3  # auto-extend: ให้ปฏิทินมีล่วงหน้าอย่างน้อยกี่ปีจากปีปัจจุบัน

# ลำดับคอลัมน์ใน Sheet1 (1-based): YEAR, MONTH, WEEK, DATE, YW, YM, status, Day
COL_YEAR, COL_MONTH, COL_WEEK, COL_DATE, COL_YW, COL_YM, COL_STATUS, COL_DAY = range(1, 9)


def week_year(date: dt.date) -> tuple[int, int]:
    """คืน (YEAR, WEEK) แบบศุกร์-พฤหัส: บวก 3 วันก่อนใช้ ISO week
    (ตรงกับ fallback ใน Calendar.load_calendar บรรทัด 298-302)"""
    iso = (date + dt.timedelta(days=3)).isocalendar()
    return int(iso[0]), int(iso[1])


def _find_last_dated_row(ws) -> tuple[int, dt.date | None]:
    """หาแถวสุดท้ายที่มี DATE จริง (ข้ามแถวขยะท้ายไฟล์)"""
    last_row = 1  # header
    last_date: dt.date | None = None
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=COL_DATE).value
        if isinstance(v, dt.datetime):
            last_date, last_row = v.date(), r
        elif isinstance(v, dt.date):
            last_date, last_row = v, r
    return last_row, last_date


def build_rows(start: dt.date, end: dt.date) -> list[list]:
    """สร้างข้อมูลรายวัน [YEAR, MONTH, WEEK, DATE, YW, YM, status, Day] จาก start ถึง end (รวมปลาย)"""
    rows: list[list] = []
    d = start
    one = dt.timedelta(days=1)
    while d <= end:
        year, week = week_year(d)
        status = 0 if d.weekday() == 6 else 1  # 6 = Sunday -> หยุด, ที่เหลือ = ทำงาน
        yw = int(f"{year}{week:02d}")
        ym = int(f"{year}{d.month:02d}")
        day = d.strftime("%a")  # Mon, Tue, ...
        rows.append([
            year, d.month, week,
            dt.datetime(d.year, d.month, d.day),  # DATE เป็น datetime ให้ตรงชนิดกับข้อมูลเดิม
            yw, ym, status, day,
        ])
        d += one
    return rows


def ensure_calendar_extended(path, end_year: int | None = None,
                             sheet: str = SHEET, do_backup: bool = True) -> dict:
    """เติมปฏิทินปีถัดไปลง Sheet1 ให้ครอบถึงสิ้นปี end_year (idempotent)

    ใช้เรียกอัตโนมัติ (เช่นตอน webapp server start / ก่อนรัน pipeline) — ถ้าปฏิทิน
    มีถึงปีเป้าหมายอยู่แล้วจะไม่ทำอะไร (คืน added=0). ปลอดภัยเรียกซ้ำได้.

    end_year=None → ปีปัจจุบัน + DEFAULT_YEARS_AHEAD (rolling horizon)
    คืน dict สรุปผล: {"added": n, "from": ..., "to": ..., "backup": ...} หรือ {"added": 0, "reason": ...}
    """
    path = Path(path)
    if end_year is None:
        end_year = dt.date.today().year + DEFAULT_YEARS_AHEAD
    if not path.exists():
        return {"added": 0, "reason": f"ไม่พบไฟล์ {path}", "ok": False}
    try:
        wb = openpyxl.load_workbook(path)
    except PermissionError:
        return {"added": 0, "reason": "ไฟล์ถูกเปิดค้าง (PermissionError)", "ok": False}
    if sheet not in wb.sheetnames:
        return {"added": 0, "reason": f"ไม่พบชีท {sheet}", "ok": False}
    ws = wb[sheet]

    last_row, last_date = _find_last_dated_row(ws)
    if last_date is None:
        return {"added": 0, "reason": "ไม่พบแถวที่มี DATE", "ok": False}

    end = dt.date(end_year, 12, 31)
    start = last_date + dt.timedelta(days=1)
    if start > end:
        return {"added": 0, "reason": "already extended", "last_date": str(last_date),
                "end_year": end_year, "ok": True}

    new_rows = build_rows(start, end)
    junk_count = ws.max_row - last_row

    backup_path = None
    if do_backup:
        stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
        backup_path = path.with_name(f"{path.stem}.backup-{stamp}{path.suffix}")
        shutil.copy2(path, backup_path)

    if junk_count > 0:
        ws.delete_rows(last_row + 1, junk_count)
    write_row = last_row + 1
    for row in new_rows:
        for c, val in enumerate(row, start=1):
            ws.cell(row=write_row, column=c, value=val)
        write_row += 1

    try:
        wb.save(path)
    except PermissionError:
        return {"added": 0, "reason": "บันทึกไม่ได้ (ไฟล์ถูกเปิดค้าง)", "ok": False,
                "backup": str(backup_path) if backup_path else None}

    return {"added": len(new_rows), "from": str(start), "to": str(end),
            "end_year": end_year, "backup": str(backup_path) if backup_path else None, "ok": True}


def main() -> int:
    ap = argparse.ArgumentParser(description="เติมปฏิทินปีถัดไปลง Calendar.xlsx")
    ap.add_argument("--end-year", type=int, default=DEFAULT_END_YEAR,
                    help=f"เติมถึงสิ้นปีนี้ (default {DEFAULT_END_YEAR})")
    ap.add_argument("--dry-run", action="store_true", help="แสดงผลโดยไม่เขียนไฟล์")
    args = ap.parse_args()

    from Calendar import _CALENDAR_LOCAL_PATH  # lazy: ให้ import gen_calendar ได้โดยไม่ลาก pandas/Calendar
    path = Path(_CALENDAR_LOCAL_PATH)
    if not path.exists():
        print(f"[X] ไม่พบไฟล์ Calendar: {path}")
        return 1

    try:
        wb = openpyxl.load_workbook(path)
    except PermissionError:
        print(f"[X] เปิดไฟล์ไม่ได้ (อาจกำลังเปิดใน Excel อยู่) — ปิดไฟล์ก่อนแล้วรันใหม่:\n   {path}")
        return 1

    if SHEET not in wb.sheetnames:
        print(f"[X] ไม่พบชีท '{SHEET}' ในไฟล์")
        return 1
    ws = wb[SHEET]

    last_row, last_date = _find_last_dated_row(ws)
    if last_date is None:
        print("[X] ไม่พบแถวที่มี DATE ใน Sheet1")
        return 1

    end = dt.date(args.end_year, 12, 31)
    start = last_date + dt.timedelta(days=1)
    if start > end:
        print(f"[OK] ปฏิทินมีถึง {last_date} อยู่แล้ว (>= สิ้นปี {args.end_year}) — ไม่ต้องเติม")
        return 0

    new_rows = build_rows(start, end)

    # แถวขยะท้ายไฟล์ = แถวหลัง last_row ที่ไม่มี DATE
    junk_count = ws.max_row - last_row

    print(f"[FILE] ไฟล์: {path}")
    print(f"   ข้อมูลเดิมถึง : {last_date}  (แถว {last_row})")
    print(f"   แถวขยะท้ายไฟล์: {junk_count} แถว (จะเคลียร์)")
    print(f"   จะเติม        : {start} -> {end}  = {len(new_rows)} แถว")
    # สรุปจำนวนสัปดาห์ต่อปีที่จะเติม
    by_year: dict[int, set[int]] = {}
    for row in new_rows:
        by_year.setdefault(row[0], set()).add(row[2])
    for y in sorted(by_year):
        wk = sorted(by_year[y])
        print(f"     ปี {y}: {len(wk)} สัปดาห์ (W{wk[0]}-W{wk[-1]})")
    print("   ตัวอย่าง 3 แถวแรก:")
    for row in new_rows[:3]:
        print("     ", row)

    if args.dry_run:
        print("\n[DRY-RUN] ไม่เขียนไฟล์")
        return 0

    # สำรองไฟล์ก่อนเขียน
    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M")
    backup = path.with_name(f"{path.stem}.backup-{stamp}{path.suffix}")
    shutil.copy2(path, backup)
    print(f"\n[BACKUP] สำรองไฟล์เดิม -> {backup}")

    # เคลียร์แถวขยะ แล้วต่อท้ายข้อมูลใหม่
    if junk_count > 0:
        ws.delete_rows(last_row + 1, junk_count)
    write_row = last_row + 1
    for row in new_rows:
        for c, val in enumerate(row, start=1):
            ws.cell(row=write_row, column=c, value=val)
        write_row += 1

    try:
        wb.save(path)
    except PermissionError:
        print(f"[X] บันทึกไม่ได้ (ไฟล์ถูกเปิดอยู่) — ไฟล์เดิมยังปลอดภัย, backup อยู่ที่ {backup}")
        return 1

    print(f"[OK] เติมปฏิทินเสร็จ: เพิ่ม {len(new_rows)} แถว (ถึงสิ้นปี {args.end_year})")
    print("   -> ขั้นถัดไป: เปิด Calendar.xlsx แก้วันหยุด (status->0) และรวม week ที่ต้องการได้เลย")
    return 0


if __name__ == "__main__":
    sys.exit(main())
