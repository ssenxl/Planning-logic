# -*- coding: utf-8 -*-
"""สร้างไฟล์ Excel แบบสอบถาม (questionnaire) สำหรับถาม user ว่าต้องการ logic 'ถอดแผน' แบบไหน"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = Path(__file__).parent / "ถอดแผน_Logic_Survey.xlsx"

# ---------- ชุดคำถาม ----------
# แต่ละข้อ: (หัวข้อ, คำถาม, [(ตัวเลือก, คำอธิบาย), ...])
QUESTIONS = [
    ("ระดับการทำงาน", "ถอดแผนให้ทำงานที่ระดับไหน?", [
        ("Booking (ก่อนวางแผน)", "จัด/สลับที่ระดับ booking แล้วค่อยป้อนเข้า AVA_MC → Planning ตาม pipeline เดิม"),
        ("Final plan (หลังวางแผน)", "แก้ที่ไฟล์ production_plan_*.xlsx ที่วางแผนเสร็จแล้ว"),
        ("ทั้งสองระดับ", "ทำได้ทั้ง booking และ final plan"),
    ]),
    ("แหล่งข้อมูล 'ไม่มีสี'", "ดูว่า item ไหน 'ไม่มีสี' จากอะไร?", [
        ("COLOR / NAY_COLOR ว่าง", "ดูจาก booking ตรง ๆ ถ้าช่องสีว่าง/null = ไม่มีสี"),
        ("join กับ MPS_STATUS", "เอา booking ไป join กับ BI_DATA_MINING (ITEM+WEEK) แล้วดู MPS_STATUS"),
        ("field อื่น (ระบุ)", "มีคอลัมน์/ไฟล์เฉพาะที่บอกสถานะสี — โปรดระบุในช่องหมายเหตุ"),
    ]),
    ("ค่าสถานะสี (กรอกค่าจริง)", "ค่า MPS_STATUS (หรือ field ที่เลือก) ที่ถือว่าอะไรคืออะไร?", [
        ("'ไม่มีสี' = ____________", "กรอกค่าที่หมายถึงไม่มีสี เช่น 8.NO_Fab (หรือค่าอื่น)"),
        ("'มีสี/พร้อมผลิต' = ____________", "กรอกค่าที่หมายถึงมีสีพร้อมผลิต"),
    ]),
    ("ขอบเขตการถอด", "ถอด item ที่ไม่มีสีออกแบบไหน?", [
        ("ถอดทั้ง item / order", "เอาออกทั้ง lot ของ SO/SC ที่ไม่มีสี"),
        ("ถอดเฉพาะส่วน pending", "ส่วนที่ทอไปแล้วเก็บไว้ ถอดเฉพาะส่วนที่ยังไม่ได้ทอ"),
    ]),
    ("ตัวไม่มีสีหลังถอด", "item ไม่มีสีที่ถอดออกมา เอาไปทำอะไร?", [
        ("เลื่อนไปวางตาม cap (ไม่ทิ้ง)", "หาสัปดาห์ถัดไปที่ยังมีเครื่องว่าง (MC_GROUP+GAUGE เดียวกัน) แล้ววางลง"),
        ("ลบออกจากแผนไปเลย", "เอาออกจากการวางแผนรอบนี้"),
        ("พักไว้ (ยังไม่ assign)", "พักไว้ก่อน ไม่ลงเครื่องจนกว่าจะมีสี"),
    ]),
    ("ใครเลือกตัวแทน", "item ที่มีสี (ตัวที่มาแทน) ใครเป็นคนเลือก?", [
        ("ระบบเลือกอัตโนมัติ", "ระบบหา item มีสีที่เร่งด่วน + ลงเครื่องเดียวกันได้ มาแทนเอง"),
        ("user เลือกเอง", "ระบบแสดง candidate ให้ user ตัดสินใจเลือก"),
        ("ระบบเสนอ + user ยืนยัน", "ระบบเสนอตัวเลือกที่ดีที่สุด แล้ว user กดยืนยัน"),
    ]),
    ("เกณฑ์จับคู่เครื่อง", "ตัวแทนต้องลงเครื่องชนิดเดียวกับตัวที่ถอด แค่ไหน?", [
        ("MC_GROUP + GAUGE ตรงเป๊ะ", "แทนได้เฉพาะชนิดเครื่อง+เบอร์เข็มเดียวกันเท่านั้น"),
        ("ยอม redirect ได้", "ใช้ตาราง MC_GROUP_REDIRECT ที่มีอยู่ ข้ามกลุ่มเครื่องที่รับแทนกันได้"),
        ("เฉพาะ MC_GROUP (GAUGE ไม่ต้องตรง)", "ขอแค่กลุ่มเครื่องเดียวกัน เบอร์เข็มต่างได้"),
    ]),
    ("ลำดับความสำคัญตัวแทน", "ถ้ามี item มีสีหลายตัว จะดึงตัวไหนขึ้นมาก่อน?", [
        ("RDD เร่งด่วนสุดก่อน", "เรียงตามกำหนดส่ง (RDD_WEEK) ด่วนสุดได้ก่อน"),
        ("ปริมาณมากสุดก่อน", "ตัวที่ปริมาณ (KP_WEIGHT) เยอะกว่าได้ก่อน"),
        ("ลูกค้า/ทีมสำคัญก่อน", "จัดลำดับตามความสำคัญของลูกค้า/ทีมขาย"),
    ]),
    ("เพดานการเลื่อนตัวไม่มีสี", "ตอนเลื่อนตัวไม่มีสีไปวางตาม cap ยอมเกินกำหนดส่ง (RDD) ไหม?", [
        ("ห้ามเกิน RDD_WEEK", "เลื่อนได้แค่ภายในกำหนดส่ง ถ้า cap ไม่พอให้หยุด/แจ้งเตือน"),
        ("เลย RDD ได้ถ้าจำเป็น", "ถ้า cap ไม่พอ ยอมเลื่อนเลยกำหนดส่งได้ + ติด flag เตือน 'ส่งช้า'"),
    ]),
    ("ผลลัพธ์ / รายงาน", "ต้องการ output แบบไหน?", [
        ("ไฟล์ booking จัดใหม่ + คอลัมน์ flag", "เพิ่มคอลัมน์ COLOR_STATUS / MOVED_FROM_WEEK / LATE_VS_RDD ในไฟล์เดิม"),
        ("รายงานสรุปการสลับ (ก่อน/หลัง)", "ไฟล์สรุปแยกว่าตัวไหนถูกถอด ตัวไหนมาแทน เลื่อนไปสัปดาห์ไหน"),
        ("ทั้งสองอย่าง", "ได้ทั้งไฟล์จัดใหม่และรายงานสรุป"),
    ]),
]

# ---------- สไตล์ ----------
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FILL = PatternFill("solid", fgColor="2E75B6")
Q_FILL = PatternFill("solid", fgColor="DDEBF7")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="center")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")

wb = Workbook()
ws = wb.active
ws.title = "ถอดแผน Logic"
ws.sheet_view.showGridLines = False

# ความกว้างคอลัมน์
widths = {"A": 5, "B": 22, "C": 34, "D": 52, "E": 10, "F": 26}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Title
ws.merge_cells("A1:F1")
c = ws["A1"]
c.value = "แบบสอบถามออกแบบ Logic 'ถอดแผน' (booking ไม่มีสี → สลับ item มีสี → วางตัวไม่มีสีตาม cap)"
c.fill = TITLE_FILL
c.font = Font(color="FFFFFF", bold=True, size=13)
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 38

ws.merge_cells("A2:F2")
c = ws["A2"]
c.value = "วิธีตอบ: ใส่เครื่องหมาย ✓ ในคอลัมน์ 'เลือก' ของตัวเลือกที่ต้องการ (เลือกได้ข้อละ 1) / ข้อที่ให้กรอกค่า ให้พิมพ์ในช่องหมายเหตุ"
c.font = Font(italic=True, color="595959")
c.alignment = Alignment(vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 28

# Header
headers = ["ข้อ", "หัวข้อ", "ตัวเลือก", "คำอธิบาย", "เลือก (✓)", "หมายเหตุ"]
hr = 3
for i, h in enumerate(headers):
    cell = ws.cell(row=hr, column=i + 1, value=h)
    cell.fill = HEAD_FILL
    cell.font = WHITE
    cell.alignment = CENTER
    cell.border = BORDER
ws.row_dimensions[hr].height = 22

# เนื้อหา
r = hr + 1
for qi, (topic, question, opts) in enumerate(QUESTIONS, 1):
    start = r
    for oi, (opt, desc) in enumerate(opts):
        ws.cell(row=r, column=3, value=opt).alignment = WRAP_TOP
        ws.cell(row=r, column=4, value=desc).alignment = WRAP_TOP
        ws.cell(row=r, column=5).alignment = CENTER
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = BORDER
        ws.cell(row=r, column=3).font = BOLD
        r += 1
    # merge ข้อ + หัวข้อ(คำถาม) ครอบทุกตัวเลือก
    ws.merge_cells(start_row=start, start_column=1, end_row=r - 1, end_column=1)
    ws.merge_cells(start_row=start, start_column=2, end_row=r - 1, end_column=2)
    ws.merge_cells(start_row=start, start_column=6, end_row=r - 1, end_column=6)
    a = ws.cell(row=start, column=1, value=qi)
    a.alignment = CENTER
    a.font = BOLD
    b = ws.cell(row=start, column=2, value=f"{topic}\n\n{question}")
    b.alignment = WRAP_TOP
    b.font = BOLD
    b.fill = Q_FILL
    # แต่งสีอ่อนให้แถวของแต่ละคำถามสลับ
    for rr in range(start, r):
        ws.cell(row=rr, column=2).fill = Q_FILL

ws.freeze_panes = "A4"

wb.save(OUT)
print(f"[OK] saved -> {OUT}")
