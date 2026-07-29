import os
import re
import pandas as pd
import numpy as np
import sys
import configparser as _cp
from pathlib import Path
from Calendar import load_calendar, calendar_week_map

# Load calendar from SharePoint URL (auto-sync)
CALENDAR_FILE = "https://nanyangtextilegroup.sharepoint.com/:x:/s/SCM_Cloud/IQCXP4jH73zhQozDNvw1XF8OAY5m4p-UFv35Tcpza6v8mJo?e=43ffCc"
_calendar_df = load_calendar(CALENDAR_FILE, sheet_name="Sheet1")
_calendar_df.columns = _calendar_df.columns.str.strip()
_calendar_df["DATE"] = pd.to_datetime(_calendar_df["DATE"], errors="coerce")
_calendar_df = _calendar_df[_calendar_df["DATE"].notna()].copy()
_calendar_df["is_working_day"] = _calendar_df["status"].map({1: 1, 0: 0}).fillna(0)
# สรุปวันเปิด "ต่อสัปดาห์" แยกรายปี (ปฏิทินมีหลายปี → WEEK เดียวกันซ้ำหลายแถว)
_calendar_week = calendar_week_map(_calendar_df)


def get_working_days_in_week(week):
    """จำนวนวันที่ปฏิทินเปิดของ "หนึ่งสัปดาห์" (ไม่รวมข้ามปี) — ตรงกับ Planning
    ปฏิทินมีหลายปี (2026–2031) จึงห้าม sum ทุกแถว WEEK==week (จะได้ ~30 วัน)
    ใช้แถวแรก (ปีแรก = ปีที่กำลังวางแผน) ของ calendar_week_map"""
    wd = _calendar_week[_calendar_week["week"] == week]
    if wd.empty:
        return 6  # fallback
    return int(wd.iloc[0]["working_days"])

# =========================
# CONFIG
# =========================
SETUP_DAYS = 5  # default setup days
BASE_DIR = Path(sys.executable).parent if getattr(sys, 'frozen', False) else Path(__file__).parent
_cfg = _cp.ConfigParser(interpolation=None)  # ปิด interpolation ให้ใช้ %USERPROFILE% ใน path ได้
if not _cfg.read(BASE_DIR / "config.ini", encoding="utf-8"):
    raise FileNotFoundError(f"ไม่พบ config.ini ที่ {BASE_DIR / 'config.ini'} — กรุณาสร้างไฟล์ config.ini ก่อนรัน")

def _get_item_cotton_poly(item_code: str) -> str:
    """Return 'COTTON' if FD5/F5 prefix, 'POLY' if FD4/F4 prefix, '' otherwise."""
    item = str(item_code).strip().upper()
    if item.startswith("FD5") or item.startswith("F5"):
        return "COTTON"
    if item.startswith("FD4") or item.startswith("F4"):
        return "POLY"
    return ""


_mc_setup_time_map: dict = {}  # populated after _master_mc_df is loaded below

def get_setup_days_for_item(material_content: str, yarn_used: str, mc_group: str = "", item_code: str = "") -> int:
    """
    คำนวณ setup days

    Logic:
    - Type SINGLE (เช็คจาก _MC_GROUP_TO_TYPE[mc_group]) → ใช้ prefix ของ ITEM code จับคู่แบบ
      longest-prefix match จากชีท "Item Setup" (ไม่ดูตรรกะ COTTON/POLY/DTY เดิม)
      ถ้าไม่ match prefix ใดเลย (รวม prefix ที่เว้นว่าง เช่น F9) → default 3 วัน
    - ไม่ใช่ SINGLE → ตรรกะเดิม:
      0. ถ้า mc_group มีค่าใน _mc_setup_time_map → ใช้ค่านั้น
      1. ถ้า MATERIAL_CONTENT เป็น COTTON → 3 วัน (ไม่สนใจ YARN_USED)
      2. ถ้า MATERIAL_CONTENT เป็น POLY → 5 วัน
      3. ถ้า MATERIAL_CONTENT เป็นอื่นๆ (CD, TC, CVC, CT, ฯลฯ) → เช็ค YARN_USED: DTY → 5 วัน
      4. ถ้าไม่มีทั้งสองอย่าง → default 3 วัน
    """
    _mc_u = str(mc_group).strip().upper() if mc_group else ""
    # เฉพาะ Type SINGLE → ใช้ prefix ของ ITEM (ไม่ดูตรรกะเดิม)
    # ครอบคลุม "SINGLE", "SINGLE (TUBE)" และเผื่อสะกด "SINGEL"
    _mc_type = _MC_GROUP_TO_TYPE.get(_mc_u, "").replace("SINGEL", "SINGLE")
    if _mc_u and _mc_type.startswith("SINGLE"):
        _pfx_days = _lookup_item_prefix_setup(item_code)
        return _pfx_days if _pfx_days is not None else 3
    if mc_group:
        if _mc_u in _mc_setup_time_map:
            return _mc_setup_time_map[_mc_u]
    mat = str(material_content).strip().upper() if not pd.isna(material_content) else ""
    yarn = str(yarn_used).strip().upper() if not pd.isna(yarn_used) else ""

    if mat == "COTTON":
        if "DTY" in yarn:
            return 5
        return 3
    if mat == "POLY":
        return 5
    if mat:  # CD, TC, CVC, CT, ฯลฯ
        if "DTY" in yarn:
            return 5
        return 3
    return 3
BOOKING_DIR = BASE_DIR / "Booking"
MASTER_MC_FILE = Path(os.path.expandvars(_cfg["paths"]["master_mc"]))
if not MASTER_MC_FILE.exists():
    raise FileNotFoundError(
        f"ไม่พบ MasterMC.xlsx ที่ {MASTER_MC_FILE}\n"
        f"กรุณาแก้ path ใน config.ini หัวข้อ [paths] master_mc ให้ตรงกับเครื่องของคุณ"
    )
OUTPUT_DIR = BASE_DIR / "data_plan"
from datetime import datetime as _datetime_now
_d = _datetime_now.now()
OUTPUT_FILE = OUTPUT_DIR / f"booking_final_ready_{_d.day}-{_d.month}-{_d.year+543}_{_d.hour:02d}{_d.minute:02d}.xlsx"

EXCLUDE_MC_GROUP = [
    "CL-NP",
    "CL-OM",
    "COMKN",
    "F-CL",
    "CL",
    "F-TSD",
]

# FQC: ไม่มีเครื่องใน MasterMC → วางแผนไม่ได้ แต่ต้องการโชว์ยอดใช้เครื่องในชีท AVA_MC
# จึงปล่อยให้ไหลผ่านการคำนวณ แล้วตัดทิ้งก่อนเขียนชีท DETAIL (Planning.py อ่านชีทนี้)
FQC_MC_GROUP = [
    "FQCCL-NP",
    "FQCCL-OM",
    "FQC-Omnoi",
    "FQC-Phet",
    "FQC",
]
FQC_MC_CAT = {"FQC", "FQC-CL"}

KEEP_COLUMNS = [
    "CAT",
    "MC_GROUP",
    "GUAGE",
    "ITEM_CODE",
    "SO_NO",
    "TEAM_NAME",
    "PO_IN_DATE",
    "CAP_KNIT",
    "KP_WEIGHT",
    "WEEK",
    "TYPE",
    "YARN-USED",
    "YARN_USED",
    "STRUCTURE",
    "MATERIAL_CONTENT",
    "DESCRIPTION",
]


# =========================
# LOAD CAPABILITY GROUP
# =========================
def load_capability_groups(file_path: str) -> pd.DataFrame:
    with pd.ExcelFile(file_path) as xls:
        records = []
        for sheet in xls.sheet_names:
            df_sheet = pd.read_excel(xls, sheet_name=sheet)
            df_sheet.columns = df_sheet.columns.str.strip()

            # Rename columns to match expected format
            if "MC" in df_sheet.columns and "Guage" in df_sheet.columns and "Group" in df_sheet.columns:
                df_sheet = df_sheet.rename(columns={"MC": "MC_GROUP", "Guage": "GUAGE", "Group": "Capability Group"})
                records.append(
                    df_sheet[["MC_GROUP", "GUAGE", "Capability Group"]]
                    .dropna()
                    .drop_duplicates()
                )

    if not records:
        return pd.DataFrame(columns=["MC_GROUP", "GUAGE", "Capability Group"])

    master = pd.concat(records, ignore_index=True)
    master["MC_GROUP"] = master["MC_GROUP"].astype(str).str.strip()
    master["GUAGE"] = master["GUAGE"].astype(str).str.strip()
    return master.drop_duplicates()


# =========================
# LOAD MASTER MC FROM FILE
# =========================
_MASTER_MC_PATH = MASTER_MC_FILE  # path จาก config.ini [paths] master_mc

def _load_master_mc(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path)
    df.columns = df.columns.str.strip()
    # forward-fill merged cells (Excel merged → NaN in non-first rows)
    for _col in ["Factory", "Type", "MC_CAT", "MC"]:
        if _col in df.columns:
            df[_col] = df[_col].ffill()
    df["MC"] = df["MC"].astype(str).str.strip().str.upper()
    df["Guage"] = df["Guage"].astype(str).str.strip()
    return df

try:
    _master_mc_df = _load_master_mc(_MASTER_MC_PATH)
    print(f"✅ โหลด MasterMC สำเร็จ: {len(_master_mc_df)} แถว จาก {_MASTER_MC_PATH}")
except Exception as _e_mmc:
    print(f"⚠️ โหลด MasterMC ไม่ได้ ({_e_mmc}) — ใช้ค่า default")
    _master_mc_df = pd.DataFrame(columns=["MC", "Guage", "Total MC", "Working Hours."])

# วันทำงานรายสัปดาห์ (แหล่งเดียว: ชีท Work Day / Week Merge ใน Calendar.xlsx)
import WorkDay
WorkDay.init(_master_mc_df, get_working_days_in_week)

# MC → Setup time map: mc_upper → setup_days (จาก MasterMC column "Set up time")
for _, _mrow in _master_mc_df.iterrows():
    _mmc = str(_mrow.get("MC", "")).strip().upper()
    _st = _mrow.get("Set up time", None)
    if _mmc and _st is not None and str(_st).strip() not in ("", "nan", "NAN", "NaT", "None"):
        try:
            _mc_setup_time_map[_mmc] = int(float(_st))
        except (ValueError, TypeError):
            pass
print(f"✅ MC→SetupTime map: {len(_mc_setup_time_map)} entries")

# MC group → Type map: mc_upper → Type (SINGLE/DOUBLE/...) — ใช้เช็คว่าเป็น Type SINGLE
_MC_GROUP_TO_TYPE: dict = {}
for _, _mrow in _master_mc_df.iterrows():
    _mmc = str(_mrow.get("MC", "")).strip().upper()
    if _mmc:
        _MC_GROUP_TO_TYPE[_mmc] = str(_mrow.get("Type", "")).strip().upper()

# Item prefix → Setup days map (จากชีท "Item Setup" ใน MasterMC) — ใช้เฉพาะ Type SINGLE
# format: prefix_upper → setup_days (เช่น "FD1" → 2) จับคู่แบบ longest-prefix match
# ช่อง ITEM อาจมีหลาย prefix คั่นด้วย comma (เช่น "F1 , FD1") → แยกเป็นหลาย key
_item_prefix_setup_map: dict = {}
try:
    _isetup_df = pd.read_excel(_MASTER_MC_PATH, sheet_name="Item Setup")
    _isetup_df.columns = _isetup_df.columns.str.strip()
    for _, _isrow in _isetup_df.iterrows():
        _iprefix_raw = str(_isrow.get("ITEM", "")).strip()
        _iday = _isrow.get("Set up (day)", None)
        if not _iprefix_raw or _iprefix_raw.upper() in ("", "NAN", "NONE"):
            continue
        if _iday is None or str(_iday).strip() in ("", "nan", "NAN", "NaT", "None"):
            continue  # prefix ว่าง (เช่น F9) → ข้าม ให้ตกไปใช้ default
        try:
            _iday_int = int(float(_iday))
        except (ValueError, TypeError):
            continue
        for _pfx in _iprefix_raw.split(","):
            _pfx = _pfx.strip().upper()
            if _pfx:
                _item_prefix_setup_map[_pfx] = _iday_int
    print(f"✅ Item prefix→Setup map: {len(_item_prefix_setup_map)} entries")
except Exception as _e_isetup:
    print(f"⚠️ ไม่สามารถโหลดชีท Item Setup: {_e_isetup}")


def _lookup_item_prefix_setup(item_code: str):
    """คืน setup days จาก prefix ของ item (longest-prefix match) หรือ None ถ้าไม่ match"""
    _ic = str(item_code).strip().upper()
    if not _ic or not _item_prefix_setup_map:
        return None
    for _pfx in sorted(_item_prefix_setup_map, key=len, reverse=True):
        if _ic.startswith(_pfx):
            return _item_prefix_setup_map[_pfx]
    return None

def _norm_gauge_ava(gauge) -> str:
    """Normalize gauge: 22.0 → 22"""
    if gauge is None or (isinstance(gauge, float) and pd.isna(gauge)):
        return ""
    s = str(gauge).strip()
    if not s or s.lower() == "nan":
        return ""
    if re.fullmatch(r"\d+\.0+", s):
        return s.split(".", 1)[0]
    return s

# =========================
# วัน/ชั่วโมงทำงาน: อ่านจากชีท Work Day (Calendar.xlsx) ผ่าน WorkDay เท่านั้น
# — ไม่ใช้ Working Day / Working Hours. ของ MasterMC และไม่มีกฎพิเศษ Week 17/32 อีกต่อไป
# — วันทำงาน = min(ค่าที่ตั้งในแผง, วันที่ปฏิทินเปิดของสัปดาห์นั้น) คิดให้แล้วใน WorkDay
# (working_hour ของ Item Special ยังชนะเสมอ — ดู _apply_cap_adj)
# =========================

# =========================
# ITEM SPECIAL: per-(Item, MC, Guage) override for Working day and Working hour
# Source: MasterMC.xlsx sheet "Item Special"
# =========================
_ITEM_SPECIAL_LOOKUP_AVA: dict = {}  # key=(item_upper, mc_upper, gauge_str), value=(working_day, working_hour)


def _get_item_special_ava(item_code, mc_group, gauge=None):
    """Return (working_day, working_hour) from Item Special for (item, MC, gauge), or None."""
    if not item_code or not mc_group:
        return None
    item_u = str(item_code).strip().upper()
    mc_u = str(mc_group).strip().upper()
    g_u = _norm_gauge_ava(gauge)
    result = _ITEM_SPECIAL_LOOKUP_AVA.get((item_u, mc_u, g_u))
    if result is None and g_u:
        result = _ITEM_SPECIAL_LOOKUP_AVA.get((item_u, mc_u, ""))
    return result


try:
    _is_df_ava = pd.read_excel(MASTER_MC_FILE, sheet_name="Item Special")
    _is_df_ava.columns = _is_df_ava.columns.str.strip()
    for _, _is_row in _is_df_ava.iterrows():
        _is_mc = str(_is_row.get("MC", "")).strip().upper()
        _is_guage = _norm_gauge_ava(_is_row.get("Guage", ""))
        _is_item = str(_is_row.get("Item", "")).strip().upper()
        _is_wd = int(_is_row.get("Working day", 6) or 6)
        _is_wh = int(_is_row.get("Working hour", 20) or 20)
        if _is_item and _is_mc:
            _ITEM_SPECIAL_LOOKUP_AVA[(_is_item, _is_mc, _is_guage)] = (_is_wd, _is_wh)
    print(f"Item Special (AVA): {len(_ITEM_SPECIAL_LOOKUP_AVA)} entries loaded")
except Exception as _e_is_ava:
    print(f"Cannot load Item Special ({_e_is_ava}) -- using MasterMC defaults")
    _ITEM_SPECIAL_LOOKUP_AVA = {}

# =========================
# LOCK_MC: เครื่องที่ user "กันไว้ให้ item" ในสัปดาห์ที่ item ไม่ได้ผลิต
# Source: MasterMC.xlsx sheet "Lock_MC" (คอลัมน์ ITEM ไม่บังคับ)
#
# ความหมาย: กันกระบอกไว้ N เครื่อง → สัปดาห์นั้นเครื่องว่าง (Planning หักออกจาก pool)
# แต่พอ item กลับมาผลิต ถือว่าเครื่อง N ตัวนั้น "ยังอุ่น" → ไม่ต้อง setup ใหม่
# ⚠️ จำนวนที่กันไว้ = จำนวนเครื่องที่อุ่นได้ (กัน 3 → อุ่น 3 ตัว ตัวที่ 4 เป็นเครื่องใหม่ต้อง setup)
#
# แถวที่ไม่มี ITEM = กันเครื่องเฉย ๆ (Planning หัก pool) ไม่เกี่ยวกับความอุ่น → ข้ามที่นี่
# คอลัมน์ MC (FA/SKP) ถ้าระบุ = อุ่นเฉพาะกลุ่มเครื่องนั้น, เว้นว่าง = อุ่นทุกกลุ่มใน MC_CAT
# =========================
_LOCK_WARM: dict = {}  # key=(item_upper, mc_cat_upper, gauge_norm, mc_upper|"") → {week: locked_count}

try:
    _lk_df = pd.read_excel(MASTER_MC_FILE, sheet_name="Lock_MC")
    _lk_df.columns = _lk_df.columns.str.strip()

    def _lk_int(v):
        # ทนค่าว่าง/NaN/ทศนิยม/สตริง — คืน None ถ้าแปลงไม่ได้ (กันทั้งชีทพังเพราะแถวเดียว)
        try:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return None
            s = str(v).strip()
            if not s or s.lower() == "nan":
                return None
            return int(float(s))
        except (ValueError, TypeError):
            return None

    for _, _lk_row in _lk_df.iterrows():
        _lk_item = str(_lk_row.get("ITEM", _lk_row.get("ITEM_CODE", "")) or "").strip().upper()
        if not _lk_item or _lk_item == "NAN":
            continue  # ไม่ระบุ item = กันเครื่องเฉย ๆ ไม่มีผลกับความอุ่น
        _lk_week = _lk_int(_lk_row.get("Week"))
        _lk_cnt = _lk_int(_lk_row.get("MC Lock")) or 0
        if _lk_week is None or _lk_cnt <= 0:
            continue
        _lk_key = (
            _lk_item,
            str(_lk_row.get("MC_CAT", "")).strip().upper(),
            _norm_gauge_ava(_lk_row.get("Guage", "")),
            str(_lk_row.get("MC", "") or "").strip().upper().replace("NAN", ""),
        )
        _lk_slot = _LOCK_WARM.setdefault(_lk_key, {})
        # คีย์ซ้ำ (สัปดาห์เดียวกัน) → บวกรวม ให้ตรงกับที่ Planning หัก pool และที่การ์ดเตือนไว้
        _lk_slot[_lk_week] = _lk_slot.get(_lk_week, 0) + _lk_cnt

    if _LOCK_WARM:
        print(f"✅ Lock_MC warm map: {_LOCK_WARM}")
    else:
        print("ℹ️ Lock_MC: ไม่มีแถวที่ระบุ ITEM → ไม่มีเครื่องกันไว้แบบ 'คงความอุ่น'")
except Exception as _e_lk:
    print(f"⚠️ ไม่สามารถโหลด Lock_MC sheet: {_e_lk}")


def _lock_warm_weeks(item, mc_cat, gauge, mc_group) -> dict:
    """{week: จำนวนเครื่องที่ lock ไว้ให้ item นี้} — รวม lock ที่ระบุ MC ตรงกับที่เว้นว่าง

    เว้นว่าง = ครอบคลุมทุกกลุ่มเครื่องใน MC_CAT นั้น → ถ้ามีทั้งสองแบบใช้ค่าที่มากกว่า
    (ไม่บวกกัน เพราะเป็นการอ้างถึงเครื่องชุดเดียวกัน)
    """
    if not _LOCK_WARM:
        return {}
    _item = str(item).strip().upper()
    _cat = str(mc_cat).strip().upper()
    _g = _norm_gauge_ava(gauge)
    _mc = str(mc_group).strip().upper()
    out: dict = {}
    for _k in ((_item, _cat, _g, _mc), (_item, _cat, _g, "")):
        for _w, _c in _LOCK_WARM.get(_k, {}).items():
            out[_w] = max(out.get(_w, 0), _c)
    return out


# =========================
# MC SPECIAL: per-(Factory, MC_CAT, MC, Gauge) → POLY/COTTON machine count
# Source: MasterMC.xlsx sheet "MC Special"
# ถ้า ITEM_CODE ขึ้นต้น FD5/F5 → ใช้ pool COTTON, FD4/F4 → ใช้ pool POLY
# =========================
_MC_SPECIAL_LOOKUP: dict = {}  # key=(factory_upper, mc_cat_upper, mc_upper, gauge_str) → {"POLY": int, "COTTON": int}

try:
    _mcs_df = pd.read_excel(MASTER_MC_FILE, sheet_name="MC Special")
    _mcs_df.columns = _mcs_df.columns.str.strip()
    for _, _mcs_row in _mcs_df.iterrows():
        _mcs_fac = str(_mcs_row.get("Factory", "")).strip().upper()
        _mcs_cat = str(_mcs_row.get("MC_CAT", "")).strip().upper()
        _mcs_mc_raw = str(_mcs_row.get("MC", "")).strip().upper()
        _mcs_mc = "" if _mcs_mc_raw in ("NAN", "NONE", "") else _mcs_mc_raw
        _mcs_g = _norm_gauge_ava(_mcs_row.get("Guage", ""))
        if not _mcs_fac or not _mcs_cat or not _mcs_g:
            continue
        _mcs_poly_raw = _mcs_row.get("POLY")
        _mcs_cotton_raw = _mcs_row.get("COTTON")
        _mcs_poly = int(_mcs_poly_raw) if pd.notna(_mcs_poly_raw) and str(_mcs_poly_raw).strip() not in ("", "-") else 0
        _mcs_cotton = int(_mcs_cotton_raw) if pd.notna(_mcs_cotton_raw) and str(_mcs_cotton_raw).strip() not in ("", "-") else 0
        if _mcs_poly > 0 or _mcs_cotton > 0:
            _MC_SPECIAL_LOOKUP[(_mcs_fac, _mcs_cat, _mcs_mc, _mcs_g)] = {"POLY": _mcs_poly, "COTTON": _mcs_cotton}
    print(f"MC Special: {len(_MC_SPECIAL_LOOKUP)} entries loaded")
    for _k, _v in _MC_SPECIAL_LOOKUP.items():
        print(f"  {_k} → {_v}")
except Exception as _e_mcs:
    print(f"Cannot load MC Special ({_e_mcs}) -- no Cotton/Poly split")
    _MC_SPECIAL_LOOKUP = {}

# --- Type-based description rules (Factory+Type rows ใน MC Special) ---
# key=(factory_upper, type_upper) → {max_mc: int, keywords: list[str]}
_TYPE_DESC_RULES: dict = {}

def _is_description_special_type(desc: str, keywords: list) -> bool:
    """Return True ถ้า DESCRIPTION match keyword ใดๆ แต่ถ้ามี FRENCH TERRY → False เสมอ"""
    d = str(desc).strip().upper()
    if "FRENCH TERRY" in d:
        return False
    base_kws = [kw.lstrip("$") for kw in keywords]
    return any(kw in d for kw in base_kws)

try:
    _mcs_df2 = pd.read_excel(MASTER_MC_FILE, sheet_name="MC Special")
    _mcs_df2.columns = _mcs_df2.columns.str.strip()
    _kw_col = "Unnamed: 8" if "Unnamed: 8" in _mcs_df2.columns else None
    _cur_key = None
    _cur_kws: list = []
    _cur_max = 0
    _cur_cat = ""
    for _, _mcs2_row in _mcs_df2.iterrows():
        _mcs2_fac  = str(_mcs2_row.get("Factory", "")).strip().upper()
        _mcs2_type = str(_mcs2_row.get("Type",    "")).strip().upper()
        _mcs2_cat  = str(_mcs2_row.get("MC_CAT",  "")).strip().upper()
        if _mcs2_cat == "NAN":
            _mcs2_cat = ""
        _mcs2_desc = _mcs2_row.get("DESCRIPTION")
        _mcs2_kw   = str(_mcs2_row.get(_kw_col, "") if _kw_col else "").strip().upper()
        _mcs2_guage_raw = _mcs2_row.get("Guage")
        _mcs2_has_guage = pd.notna(_mcs2_guage_raw) and str(_mcs2_guage_raw).strip() not in ("", "-", "NAN")
        # TYPE_SPECIAL row: มี Factory+Type แต่ไม่มี Guage
        if (_mcs2_fac not in ("", "NAN")
                and _mcs2_type not in ("", "NAN")
                and not _mcs2_has_guage):
            if _cur_key and _cur_kws:
                _TYPE_DESC_RULES[_cur_key] = {"max_mc": _cur_max, "keywords": _cur_kws[:], "mc_cat": _cur_cat}
            _cur_key  = (_mcs2_fac, _mcs2_type)
            _cur_cat  = _mcs2_cat  # เก็บ MC_CAT จาก rule (เช่น SINGLE-32)
            _cur_max  = int(_mcs2_desc) if pd.notna(_mcs2_desc) and str(_mcs2_desc).strip() not in ("", "-") else 0
            _cur_kws  = [_mcs2_kw] if _mcs2_kw and _mcs2_kw != "NAN" else []
        elif _mcs2_fac in ("", "NAN") and _mcs2_type in ("", "NAN") and _mcs2_kw not in ("", "NAN"):
            if _cur_key is not None:
                _cur_kws.append(_mcs2_kw)
    if _cur_key and _cur_kws:
        _TYPE_DESC_RULES[_cur_key] = {"max_mc": _cur_max, "keywords": _cur_kws[:], "mc_cat": _cur_cat}
    print(f"Type DESC rules: {len(_TYPE_DESC_RULES)} entries")
    for _k, _v in _TYPE_DESC_RULES.items():
        print(f"  {_k} → mc_cat={_v['mc_cat']!r}, max_mc={_v['max_mc']}, keywords={_v['keywords']}")
except Exception as _e_tdr:
    print(f"Cannot parse Type DESC rules ({_e_tdr})")
    _TYPE_DESC_RULES = {}

# =========================
# TOTAL MC MASTER  (โหลดจาก MasterMC.xlsx)
# =========================
TOTAL_MC_MAP = {
    (row["MC"], row["Guage"]): int(row["Total MC"])
    for _, row in _master_mc_df.iterrows()
    if pd.notna(row.get("Total MC")) and str(row.get("Total MC", "")).strip() not in ("", "-")
}

# TOTAL_MC รวมตาม Type_1 + Guage (จาก MasterMC)
# แก้บั๊ก sum ซ้ำซ้อน กรณี MC_CAT/GUAGE เดียวกันมีหลายแถวในไฟล์ และบังคับแปลง Total MC เป็นตัวเลข
_mc_cat_g_sum = (
    _master_mc_df.dropna(subset=["MC_CAT", "Guage", "Total MC"])
    .copy()
)
_mc_cat_g_sum["Total MC"] = pd.to_numeric(_mc_cat_g_sum["Total MC"], errors='coerce')
_mc_cat_g_sum = (
    _mc_cat_g_sum.dropna(subset=["Total MC"])
    .groupby(["MC_CAT", "Guage"], as_index=False)["Total MC"]
    .sum()
)
TOTAL_MC_MAP_TYPE1 = {
    (str(row["MC_CAT"]).strip(), str(row["Guage"]).strip()): int(row["Total MC"])
    for _, row in _mc_cat_g_sum.iterrows()
}



# =========================
# SHARED POOL — รวมเครื่องตาม Type ใน MasterMC
# แยกตาม Factory: เครื่องคนละ Factory ไม่นำมารวมกัน
# pool_key format: "Factory|Type_1:Guage" หรือ "Factory|Group"
# =========================
_MC_TO_POOL: dict = {}      # (MC_upper, norm_guage) → pool_key
_MC_TO_FACTORY: dict = {}   # (MC_upper, norm_guage) → factory_str
_MC_TO_TYPE: dict = {}      # (MC_upper, norm_guage) → type_str (SINGLE/DOUBLE/...)
_TOTAL_MC_BY_TYPE: dict = {}   # pool_key → total machines

# รวบรวมข้อมูล MasterMC
_t1g_total: dict = {}    # (factory, Type_1_upper, norm_guage) → total
_group_t1gs: dict = {}   # group_str → set of (factory, Type_1_upper, norm_guage)
_mc_to_group: dict = {}  # (MC_upper, norm_guage) → group_str
_mc_to_t1g: dict = {}    # (MC_upper, norm_guage) → (Type_1_upper, norm_guage)

for _, _row in _master_mc_df.iterrows():
    _mc  = str(_row.get("MC",      "")).strip().upper()
    _t1  = str(_row.get("MC_CAT",  "")).strip().upper()
    _g   = _norm_gauge_ava(_row.get("Guage", ""))
    _grp = str(_row.get("Group",   "")).strip()
    _fac = str(_row.get("Factory", "")).strip()
    _tot = _row.get("Total MC")
    if not _mc or not _t1 or not _g:
        continue
    _tot_int = 0
    if pd.notna(_tot) and str(_tot).strip() not in ("", "-"):
        try:
            _tot_int = int(_tot)
        except (ValueError, TypeError):
            pass
    if _tot_int > 0:
        _t1g_total[(_fac, _t1, _g)] = _t1g_total.get((_fac, _t1, _g), 0) + _tot_int
    if _grp:
        _group_t1gs.setdefault(_grp, set()).add((_fac, _t1, _g))
        _mc_to_group[(_mc, _g)] = _grp
    _mc_to_t1g[(_mc, _g)] = (_t1, _g)
    _MC_TO_FACTORY[(_mc, _g)] = _fac
    _MC_TO_TYPE[(_mc, _g)] = str(_row.get("Type", "")).strip().upper()

# Groups ที่ span ข้าม Type_1+Guage ภายใน Factory เดียวกัน (cross-pool)
_cross_groups = set()
for _grp, _fac_t1gs in _group_t1gs.items():
    _per_fac: dict = {}
    for (_fac, _t1, _g) in _fac_t1gs:
        _per_fac.setdefault(_fac, set()).add((_t1, _g))
    if any(len(v) > 1 for v in _per_fac.values()):
        _cross_groups.add(_grp)

# Build pool key ต่อ MC (รวม Factory เสมอ):
# - SINGLE-32   → "Factory|Group"  (แยกพูลตามคอลัมน์ Group เช่น SKP vs SKPTA/SKPLE)
# - cross-pool  → "Factory|Group name"
# - T1G pool    → "Factory|Type_1:Guage"
for (_mc, _g), (_t1, _gt) in _mc_to_t1g.items():
    _grp = _mc_to_group.get((_mc, _g), "")
    _fac = _MC_TO_FACTORY.get((_mc, _g), "")
    _fac_pfx = f"{_fac}|" if _fac else ""
    # SINGLE-32: แยกพูลตามคอลัมน์ Group ใน MasterMC (เช่น SKP vs SKPTA/SKPLE)
    # เครื่องที่อยู่ Group เดียวกัน = พูลเดียวกัน (ใช้แทนกันได้); ต่าง Group = คนละพูล
    # (แทน special-case เดิมที่แยกเฉพาะ MC=="SKP" ซึ่งไม่ครอบคลุมกรณี Group รวม SKP+SKPTA)
    if _t1 == "SINGLE-32" and _grp:
        _MC_TO_POOL[(_mc, _g)] = f"{_fac_pfx}{_grp}"
    else:
        _MC_TO_POOL[(_mc, _g)] = (
            f"{_fac_pfx}{_grp}" if (_grp and _grp in _cross_groups)
            else f"{_fac_pfx}{_t1}:{_gt}"
        )

# Build _TOTAL_MC_BY_TYPE
for (_fac, _t1, _g), _tot in _t1g_total.items():
    _fac_pfx = f"{_fac}|" if _fac else ""
    _TOTAL_MC_BY_TYPE[f"{_fac_pfx}{_t1}:{_g}"] = _tot    # T1G pool

for _grp in _cross_groups:
    _per_fac_tots: dict = {}
    for (_fac, _t1, _g) in _group_t1gs[_grp]:
        _fac_pfx = f"{_fac}|" if _fac else ""
        _pk = f"{_fac_pfx}{_grp}"
        _per_fac_tots[_pk] = _per_fac_tots.get(_pk, 0) + _t1g_total.get((_fac, _t1, _g), 0)
    _TOTAL_MC_BY_TYPE.update(_per_fac_tots)

# === SINGLE-32: total ต่อพูล "Group" (แทน T1G รวม + SKP special เดิม) ===
# แต่ละ Group (เช่น "SKP 26G", "SKPLE , SKPTA 26G") เป็นพูลแยก total = ผลรวม Total MC
# ของเครื่องสมาชิกกลุ่มนั้น → SKP กับ SKPTA/SKPLE จึงคิดเครื่องแยกกันจริง
_single32_group_tot: dict = {}
for _, _s32 in _master_mc_df.iterrows():
    if str(_s32.get("MC_CAT", "")).strip().upper() != "SINGLE-32":
        continue
    _s_grp = str(_s32.get("Group", "")).strip()
    if not _s_grp:
        continue
    _s_fac = str(_s32.get("Factory", "")).strip()
    _s_tot_raw = _s32.get("Total MC")
    _s_tot = int(_s_tot_raw) if pd.notna(_s_tot_raw) and str(_s_tot_raw).strip() not in ("", "-") else 0
    _s_pfx = f"{_s_fac}|" if _s_fac else ""
    _s_pk = f"{_s_pfx}{_s_grp}"
    _single32_group_tot[_s_pk] = _single32_group_tot.get(_s_pk, 0) + _s_tot
_TOTAL_MC_BY_TYPE.update(_single32_group_tot)

# === Type-based DESC special pools ===
for (_tdr_fac, _tdr_type), _tdr_val in _TYPE_DESC_RULES.items():
    _tdr_pfx = f"{_tdr_fac}|" if _tdr_fac else ""
    _TOTAL_MC_BY_TYPE[f"{_tdr_pfx}{_tdr_type}:TYPE_SPECIAL"] = _tdr_val["max_mc"]

# === MC Special sub-pools: สร้าง COTTON/POLY sub-pool และลด normal pool ===
for (_ms_fac, _ms_cat, _ms_mc, _ms_g), _ms_vals in _MC_SPECIAL_LOOKUP.items():
    _ms_fac_pfx = f"{_ms_fac}|" if _ms_fac else ""
    # หักเครื่องกันไว้จาก "พูล Group" ของเครื่องที่ระบุใน MC Special (เช่น "SKPTA,SKPLE"
    # → group "SKPLE , SKPTA 26G") ไม่ใช่พูลระดับ CAT รวม → POLY จึงกันจากพูล SKPTA/SKPLE
    # เท่านั้น ไม่ไปแตะพูล SKP. Fallback เป็น CAT:Guage ถ้าหา Group ไม่เจอ
    _ms_mc_first = _ms_mc.split(",")[0].strip() if _ms_mc else ""
    _ms_grp = _mc_to_group.get((_ms_mc_first, _ms_g), "") if _ms_mc_first else ""
    if _ms_grp:
        _ms_normal_pk = f"{_ms_fac_pfx}{_ms_grp}"
    else:
        _ms_normal_pk = f"{_ms_fac_pfx}{_ms_cat}:{_ms_g}"
    _ms_sub_base = _ms_normal_pk
    if _ms_vals["COTTON"] > 0:
        _TOTAL_MC_BY_TYPE[f"{_ms_sub_base}:COTTON"] = _ms_vals["COTTON"]
        if _ms_normal_pk in _TOTAL_MC_BY_TYPE:
            _TOTAL_MC_BY_TYPE[_ms_normal_pk] = max(0, _TOTAL_MC_BY_TYPE[_ms_normal_pk] - _ms_vals["COTTON"])
    if _ms_vals["POLY"] > 0:
        _TOTAL_MC_BY_TYPE[f"{_ms_sub_base}:POLY"] = _ms_vals["POLY"]
        if _ms_normal_pk in _TOTAL_MC_BY_TYPE:
            _TOTAL_MC_BY_TYPE[_ms_normal_pk] = max(0, _TOTAL_MC_BY_TYPE[_ms_normal_pk] - _ms_vals["POLY"])

print(f"📊 Pools by Factory (sample): { dict(list(_TOTAL_MC_BY_TYPE.items())[:10]) }")
if not _master_mc_df.empty and "Group" in _master_mc_df.columns:
    _show_cols = [c for c in ["Factory", "Type", "MC_CAT", "MC", "Guage", "Total MC", "Group"] if c in _master_mc_df.columns]
    print(_master_mc_df[_show_cols].head(15).to_string())

import io


def fix_thai(s):
    """แก้ double-encoding: latin-1 → cp874"""
    try:
        return s.encode("latin-1").decode("cp874")
    except Exception:
        return s


def load_booking_file(file: Path) -> pd.DataFrame:
    raw_bytes = file.read_bytes()
    is_zip = raw_bytes[:2] == b"PK"
    is_biff = raw_bytes[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"

    _df = None
    if is_zip:
        _tmp = pd.read_excel(io.BytesIO(raw_bytes), engine="openpyxl")
        if "\t" in str(_tmp.columns[0]):
            col_names = _tmp.columns[0].split("\t")
            data = _tmp.iloc[:, 0].astype(str).str.split("\t", expand=True)
            n = min(len(col_names), data.shape[1])
            data = data.iloc[:, :n]
            data.columns = [fix_thai(c) for c in col_names[:n]]
            data = data.map(lambda x: fix_thai(x) if isinstance(x, str) else x)
            _df = data
        else:
            _df = _tmp
    elif is_biff:
        _df = pd.read_excel(io.BytesIO(raw_bytes), engine="xlrd")
    else:
        for enc in ("cp874", "tis-620", "utf-8-sig", "latin-1"):
            try:
                raw_text = raw_bytes.decode(enc, errors="replace")
                _tmp = pd.read_csv(io.StringIO(raw_text), sep="\t", on_bad_lines="skip")
                if _tmp.shape[1] > 3:
                    _df = _tmp
                    break
            except Exception:
                continue

    if _df is None:
        raise ValueError(f"❌ ไม่สามารถอ่านไฟล์: {file.name}")
    return _df


all_files = [f for f in BOOKING_DIR.iterdir() if f.suffix.lower() in (".xlsx", ".xls")]
if not all_files:
    raise FileNotFoundError(f"❌ ไม่พบไฟล์ใน {BOOKING_DIR}")

df_list = []
for file in all_files:
    print(f"📄 Loading: {file.name}")
    _df = load_booking_file(file)
    _df["SOURCE_FILE"] = file.name
    df_list.append(_df)
    print(f"   ✅ rows={len(_df)}, columns={_df.shape[1]}")

df = pd.concat(df_list, ignore_index=True)
df.columns = df.columns.str.strip().str.upper()
print("📋 Columns:", df.columns.tolist())
if "CAT" in df.columns:
    print(f"📦 CAT unique values:\n{df['CAT'].dropna().unique().tolist()}")
if "MC_GROUP" in df.columns:
    print(f"🔧 MC_GROUP unique values:\n{df['MC_GROUP'].dropna().unique().tolist()}")

# =========================
# CLEAN (❌ COLLAR ถูกลบถาวร)
# =========================
if "MC_GROUP" not in df.columns:
    raise KeyError(f"❌ ไม่พบ column 'MC_GROUP' - columns ที่มี: {df.columns.tolist()}")

df = df[~df["MC_GROUP"].isin(EXCLUDE_MC_GROUP)]

df["TYPE"] = df["TYPE"].astype(str).str.strip().str.upper()
df = df[df["TYPE"] != "COLLAR"]

# สอน WorkDay ว่าเครื่องในบุ๊คกิ้งอยู่กลุ่มไหนของแผง (ใช้ FACTORY+CAT+เกจ ซึ่งเป็นคีย์ของแผง)
# ต้องทำก่อนตัดคอลัมน์ FACTORY ทิ้ง — ครอบคลุมเครื่องที่ไม่มีใน MasterMC เช่น SKPHF/SKPSL/SKPAO
_wd_alias = WorkDay.register_aliases(df[[c for c in ("FACTORY", "CAT", "MC_GROUP", "GUAGE") if c in df.columns]])
if _wd_alias:
    print(f"✅ จับคู่เครื่อง→กลุ่มในแผงวันทำงานเพิ่ม {_wd_alias} คู่ (จาก FACTORY+CAT+เกจ)")

df = df[[c for c in KEEP_COLUMNS if c in df.columns]]

# ใช้ CAT จาก booking โดยตรง (ตรงกับ Type_1 ใน MasterMC)
df["MC_CAT"] = df["CAT"].astype(str).str.strip().str.upper() if "CAT" in df.columns else df["MC_GROUP"].astype(str).str.strip().str.upper()

df["GUAGE"] = df["GUAGE"].astype(str).str.strip()
df["CAP_KNIT"] = pd.to_numeric(df["CAP_KNIT"], errors="coerce")
df["KP_WEIGHT"] = pd.to_numeric(df["KP_WEIGHT"], errors="coerce")

# =========================
# APPLY 20/24 (เฉพาะ MC group ที่มี Working Hours. == 20 ใน MasterMC.xlsx)
# =========================
def _apply_cap_adj(r):
    _is = _get_item_special_ava(r["ITEM_CODE"], r["MC_GROUP"], r["GUAGE"])
    if _is is not None:
        return r["CAP_KNIT"] * (_is[1] / 24)  # Item Special working_hour override (ชนะเสมอ)
    # ชั่วโมงของกลุ่มจากชีท Work Day (หน้า WorkDayPanel) — แหล่งเดียว, ไม่ตั้ง = 24 ชม.
    # (ไม่ใช้ Working Hours. ของ MasterMC แล้ว)
    return r["CAP_KNIT"] * (WorkDay.get_working_hours(r["MC_GROUP"], r["GUAGE"]) / 24)

df["_CAP_ADJ"] = df.apply(_apply_cap_adj, axis=1)

# =========================
# GROUP ITEM
# =========================
agg_dict = {
    "KP_WEIGHT": "sum",
    "CAP_KNIT": "first",
    "_CAP_ADJ": "first",
    "MC_CAT": "first",
}
if "CAT" in df.columns:
    agg_dict["CAT"] = "first"
if "SO_NO" in df.columns:
    agg_dict["SO_NO"] = lambda x: ",".join(x.dropna().astype(str).unique())
if "TEAM_NAME" in df.columns:
    agg_dict["TEAM_NAME"] = lambda x: ",".join(x.dropna().astype(str).unique())
if "PO_IN_DATE" in df.columns:
    # ค่าในไฟล์ booking เป็นวันที่ต่อกันด้วย comma อยู่แล้ว → แตกแล้วรวม unique ไม่ให้วันซ้ำ
    agg_dict["PO_IN_DATE"] = lambda x: ",".join(dict.fromkeys(
        d.strip() for v in x.dropna().astype(str) for d in v.split(",") if d.strip()))
for col in ["YARN-USED", "YARN_USED", "STRUCTURE", "MATERIAL_CONTENT", "DESCRIPTION"]:
    if col in df.columns:
        agg_dict[col] = "first"

df = df.groupby(["MC_GROUP", "GUAGE", "ITEM_CODE", "WEEK"], as_index=False).agg(
    agg_dict
)

# =========================
# WORKING DAY
# =========================
def _get_working_day_for_row(r):
    # แหล่งเดียว: ชีท Work Day (Factory/MC_CAT/Guage [+WEEK]) — ใช้ค่าตรง ๆ ไม่หักวันหยุด
    # รวม logic ยุบสัปดาห์ + สัปดาห์ที่ปฏิทินหยุดทั้งก้อน = 0 ไว้ใน WorkDay.get_working_days แล้ว
    return WorkDay.get_working_days(r["MC_GROUP"], gauge=r["GUAGE"], week=int(r["WEEK"]))

df["WORKING_DAY"] = df.apply(_get_working_day_for_row, axis=1, result_type="reduce")

# เตือนคู่ (เครื่อง, เกจ) ที่จับเข้ากลุ่มในแผง "วันทำงานตามกลุ่มเครื่อง" ไม่ได้
# → ใช้ค่า fallback 6 วัน ไม่ใช่ค่าที่ตั้งไว้ ต้องให้เห็นเพื่อไปเพิ่มใน MasterMC/แผง
_wd_missing = sorted({
    (str(r["MC_GROUP"]).strip(), str(r["GUAGE"]).strip(), str(r["MC_CAT"]).strip())
    for _, r in df.iterrows()
    if not WorkDay.has_group(r["MC_GROUP"], r["GUAGE"])
})
if _wd_missing:
    print(f"⚠️  {len(_wd_missing)} คู่ (เครื่อง,เกจ) จับกลุ่มในแผง 'วันทำงานตามกลุ่มเครื่อง' ไม่ได้ "
          f"→ ใช้ค่า fallback {WorkDay.DEFAULT_WORK_DAYS:g} วัน:")
    for _mc, _g, _cat in _wd_missing[:30]:
        print(f"     - {_mc} เกจ {_g} (CAT={_cat})")
    if len(_wd_missing) > 30:
        print(f"     ... และอีก {len(_wd_missing) - 30} คู่")

# =========================
# SETUP DETECTION (เหมือน Planning.py: SETUP_GAP_WEEK = 3)
# =========================
SETUP_GAP_WEEK = 3
df = df.sort_values(["MC_GROUP", "GUAGE", "ITEM_CODE", "WEEK"])

# สร้าง key สำหรับเช็ค carryover
df["_carry_key"] = df["MC_GROUP"] + "_" + df["GUAGE"].astype(str) + "_" + df["ITEM_CODE"]

# คำนวณ setup days สำหรับแต่ละ item ตาม MATERIAL_CONTENT และ YARN-USED
df["_setup_days"] = df.apply(
    lambda r: get_setup_days_for_item(
        r.get("MATERIAL_CONTENT", ""),
        r.get("YARN-USED", "") or r.get("YARN_USED", ""),
        mc_group=r.get("MC_GROUP", ""),
        item_code=r.get("ITEM_CODE", "")
    ), axis=1
)

# =========================
# SETUP DETECTION — สัปดาห์ก่อนหน้า + Lock_MC ที่ user กันเครื่องไว้ให้ item
#
# เดินเป็นทอด ๆ จากสัปดาห์ผลิตล่าสุดผ่านสัปดาห์ที่ lock ไว้ โดยแต่ละช่วงห้ามเกิน
# SETUP_GAP_WEEK (เครื่องเย็นระหว่างทางไม่ได้) → ถ้าต่อถึงสัปดาห์นี้ได้ = เครื่องยังอุ่น
# _lock_warm_mc = จำนวนที่กันไว้ (แคบสุดตลอดเส้นทาง) = เพดานเครื่องที่อุ่นได้
# =========================
def _detect_setup_group(g):
    """หา _prev_week / _week_gap / _is_new_setup / _lock_warm_mc ของ _carry_key หนึ่งชุด"""
    _item = str(g["ITEM_CODE"].iloc[0]).strip().upper()
    _cat = str(g["MC_CAT"].iloc[0]).strip().upper() if "MC_CAT" in g.columns else ""
    _locks = _lock_warm_weeks(_item, _cat, g["GUAGE"].iloc[0], g["MC_GROUP"].iloc[0])

    out = []
    prev_week = None
    for wk in g["WEEK"].tolist():
        wk = int(wk)
        warm_cap = None
        eff_prev = prev_week
        if prev_week is not None and _locks:
            for lw in sorted(_locks):
                if eff_prev < lw < wk and (lw - eff_prev) <= SETUP_GAP_WEEK:
                    eff_prev = lw
                    warm_cap = _locks[lw] if warm_cap is None else min(warm_cap, _locks[lw])
        if prev_week is None:
            gap, is_new = np.nan, True
        else:
            gap = wk - eff_prev
            is_new = bool(gap > SETUP_GAP_WEEK)
        out.append({
            "_prev_week": prev_week,
            "_week_gap": gap,
            "_is_new_setup": is_new,
            "_lock_warm_mc": warm_cap if warm_cap is not None else np.nan,
        })
        prev_week = wk
    return pd.DataFrame(out, index=g.index)


df = df.join(
    df.groupby("_carry_key", group_keys=False)[
        ["WEEK", "ITEM_CODE", "MC_GROUP", "GUAGE", "MC_CAT"]
    ].apply(_detect_setup_group)
)

# =========================
# MC USE — INITIAL (คำนวณด้วย WORKING_DAY เต็ม)
# =========================
# กัน WORKING_DAY = 0 (กลุ่มไม่ได้ตั้งค่าในแผง / สัปดาห์ถูกยุบ / ปฏิทินปิดทั้งสัปดาห์)
# ไม่งั้นหารด้วยศูนย์ → inf → astype(int) ระเบิด
_wd_ok = df["WORKING_DAY"].fillna(0) > 0
df["MC_USE"] = np.where(
    (df["_CAP_ADJ"] > 0) & _wd_ok,
    df["KP_WEIGHT"] / (df["_CAP_ADJ"] * df["WORKING_DAY"].where(_wd_ok, 1)),
    0
)
df["MC_USE_CEIL"] = np.ceil(df["MC_USE"]).replace([np.inf, -np.inf], 0).fillna(0).astype(int)

# =========================
# PASS 1: eff_days หัก setup เฉพาะ _is_new_setup
# → ได้ FINAL MC_USE_CEIL ที่สะท้อนเครื่องจริงที่รันใน setup week
# =========================
_eff_pass1 = np.where(
    df["_is_new_setup"],
    np.maximum(df["WORKING_DAY"] - df["_setup_days"], 0.5),
    df["WORKING_DAY"]
)
df["_mc_use_ceil_pass1"] = np.ceil(
    np.where(
        (df["_CAP_ADJ"] > 0) & (_eff_pass1 > 0),
        df["KP_WEIGHT"] / (df["_CAP_ADJ"] * _eff_pass1),
        0
    )
).astype(float)
df["_mc_use_ceil_pass1"] = df["_mc_use_ceil_pass1"].fillna(0).astype(int)

# =========================
# PASS 2: mc_increase เทียบกับ FINAL ของ week ก่อนหน้า (ไม่ใช่ INITIAL)
# carry machines = เครื่องที่รันจริงใน prev week (FINAL)
# new machines   = INITIAL ปัจจุบัน - FINAL prev (ส่วนที่เพิ่มขึ้นจริง)
# =========================
# carry-over machines = เครื่องจาก "สัปดาห์ล่าสุดที่ยังรันจริง (pass1 > 0)" ที่ยังอยู่ใน SETUP_GAP_WEEK
# เดิมใช้ shift(1) มองแค่สัปดาห์ติดกัน → ถ้ามีสัปดาห์ KP=0 คั่นกลาง เครื่องที่กลับมารัน (rerun)
# จะถูกนับเป็น "เครื่องใหม่" ผิด ทำให้โดน setup penalty ซ้ำและพองจำนวนเครื่องเกินจริง
# (เช่น item รัน week 25 → ว่าง 26,27 → กลับมา week 28 ควร carry 2 เครื่องเดิม ไม่ใช่ setup ใหม่)
#
# Lock_MC: สัปดาห์ที่ user กันเครื่องไว้ให้ item ใช้เชื่อมช่วงที่ item ไม่ได้ผลิต (เครื่องยังอุ่น)
# และจำนวนที่กันไว้เป็นเพดานของเครื่องที่ carry ได้ (กัน 3 → carry 3, ตัวที่ 4 ต้อง setup)
def _carry_prev_active_ceil(g):
    prev_vals = []
    last_week = None
    last_ceil = 0
    for wk, c, _cap in zip(
        g["WEEK"].tolist(),
        g["_mc_use_ceil_pass1"].tolist(),
        g["_lock_warm_mc"].tolist(),
    ):
        # _cap ไม่ใช่ NaN = สัปดาห์นี้ต่อความอุ่นมาได้เพราะ Lock_MC → ข้ามเกณฑ์ gap ปกติ
        _lock_used = not pd.isna(_cap)
        if last_week is not None and ((wk - last_week) <= SETUP_GAP_WEEK or _lock_used):
            _carry = last_ceil
            if _lock_used:
                _carry = min(_carry, int(_cap))   # กันไว้กี่เครื่อง = carry ได้แค่นั้น
            prev_vals.append(_carry)
        else:
            prev_vals.append(np.nan)      # นอก gap / ครั้งแรก → ไม่มี carry (= setup ใหม่)
        if c > 0:                          # อัปเดต active week เฉพาะสัปดาห์ที่มีเครื่องรันจริง
            last_week = wk
            last_ceil = c
    return pd.Series(prev_vals, index=g.index)

df["_prev_mc_use_ceil"] = df.groupby("_carry_key", group_keys=False).apply(_carry_prev_active_ceil)
df["_mc_increase"] = np.maximum(
    df["MC_USE_CEIL"] - df["_prev_mc_use_ceil"].fillna(0), 0
).astype(int)
df["_has_mc_increase"] = (df["_mc_increase"] > 0) & (~df["_is_new_setup"])

# =========================
# คำนวณ effective working days
# - new_setup : ทุกเครื่องต้อง setup → WORKING_DAY - setup_days
# - mc_increase: เฉพาะเครื่องที่เพิ่ม setup → weighted average
# - carry       : วันเต็ม
# =========================
df["_effective_working_days"] = np.where(
    df["_is_new_setup"],
    np.maximum(df["WORKING_DAY"] - df["_setup_days"], 0.5),
    np.where(
        df["_has_mc_increase"] & (df["MC_USE_CEIL"] > 0),
        np.maximum(df["WORKING_DAY"] - (df["_setup_days"] * df["_mc_increase"] / df["MC_USE_CEIL"]), 0.5),
        df["WORKING_DAY"]
    )
)

df["MC_USE"] = np.where(
    (df["_CAP_ADJ"] > 0) & (df["_effective_working_days"] > 0),
    df["KP_WEIGHT"] / (df["_CAP_ADJ"] * df["_effective_working_days"]),
    0
)
df["MC_USE_CEIL"] = np.ceil(df["MC_USE"]).fillna(0).astype(int)

# =========================
# แถวที่ Lock_MC เชื่อมความอุ่นมา: วนหาจุดลงตัวให้ carry/setup ตรงกับที่ user กันไว้จริง
#
# สูตร 2-pass ด้านบนคิด _mc_increase จาก ceil ตอน INITIAL แต่ ceil สุดท้ายพองขึ้น
# (เพราะโดนหักวัน setup) → ส่วนต่างตกไปเป็น carryover ที่ไม่มีตัวตน เช่น กันไว้ 3 เครื่อง
# แต่รายงาน carry 4 → เกินจำนวนที่กันไว้ ซึ่งขัดกับความหมายของ Lock_MC โดยตรง
#
# จึงวนซ้ำเฉพาะแถวพวกนี้จน n = carry + setup ลงตัว (n เพิ่มทางเดียว มีเพดาน eff ≥ 0.5 → ลู่เข้าเสมอ)
# ⚠️ แถวอื่นปล่อยตามสูตรเดิม — บั๊กเครื่องผีของเดิมยังอยู่ ตั้งใจไม่แตะเพื่อไม่ให้ตัวเลขทั้งไฟล์ขยับ
# =========================
_lk_fix = (
    df["_lock_warm_mc"].notna()
    & (df["_CAP_ADJ"] > 0)
    & (df["KP_WEIGHT"] > 0)
    & (df["WORKING_DAY"] > 0)
)
for _i in df.index[_lk_fix]:
    _kp = float(df.at[_i, "KP_WEIGHT"])
    _cap = float(df.at[_i, "_CAP_ADJ"])
    _wd = float(df.at[_i, "WORKING_DAY"])
    _sd = float(df.at[_i, "_setup_days"])
    _warm = int(pd.to_numeric(df.at[_i, "_prev_mc_use_ceil"], errors="coerce") or 0)
    _n = int(np.ceil(_kp / (_cap * _wd)))
    _eff = _wd
    for _ in range(50):
        _inc = max(_n - _warm, 0)
        _eff = max(_wd - (_sd * _inc / _n), 0.5) if (_n > 0 and _inc > 0) else _wd
        _n2 = int(np.ceil(_kp / (_cap * _eff)))
        if _n2 == _n:
            break
        _n = _n2
    _use = _kp / (_cap * _eff)
    _n = int(np.ceil(_use))
    _inc = max(_n - _warm, 0)
    df.at[_i, "_effective_working_days"] = _eff
    df.at[_i, "MC_USE"] = _use
    df.at[_i, "MC_USE_CEIL"] = _n
    df.at[_i, "_mc_increase"] = _inc
    df.at[_i, "_has_mc_increase"] = bool(_inc > 0)
    print(
        f"[LOCK WARM] {df.at[_i, 'ITEM_CODE']} {df.at[_i, 'MC_GROUP']} W{int(df.at[_i, 'WEEK'])}:"
        f" {_n} เครื่อง = carry {_n - _inc} (กันไว้ {int(df.at[_i, '_lock_warm_mc'])}) + setup {_inc}"
    )

# เก็บ CAP หลังปรับ 20/24 ไว้ก่อน drop — ใช้เป็นคอลัมน์ CAP ในชีท AVA_MC_ITEM
_CAP_ADJ_SNAP = (
    df.groupby(["MC_GROUP", "GUAGE", "ITEM_CODE"], as_index=False)["_CAP_ADJ"]
    .max()
    .rename(columns={"_CAP_ADJ": "CAP"})
)

# drop temp columns
df = df.drop(columns=["_CAP_ADJ", "_mc_use_ceil_pass1"])

# =========================
# TOTAL MC
# =========================
# POOL_TYPE: ดูจาก Type ใน MasterMC ว่า (MC_GROUP, GUAGE) อยู่ใน pool ไหน
# ถ้า MC อยู่ใน MC Special → แยก sub-pool ตาม ITEM_CODE prefix (FD5/F5=COTTON, FD4/F4=POLY)
def _get_pool_type_for_row(r):
    mc_u = str(r["MC_GROUP"]).strip().upper()
    g_norm = _norm_gauge_ava(r["GUAGE"])
    normal_pool = _MC_TO_POOL.get((mc_u, g_norm), "")
    if not normal_pool:
        return normal_pool
    fac_u = _MC_TO_FACTORY.get((mc_u, g_norm), "").strip().upper()
    t1, _ = _mc_to_t1g.get((mc_u, g_norm), ("", ""))
    fac_pfx = f"{fac_u}|" if fac_u else ""

    # 1. MC Special: POLY/COTTON split ตาม ITEM_CODE prefix (priority สูงสุด)
    # จับคู่ MC แบบ membership: ช่อง MC ใน MC Special อาจมีหลายค่าคั่น comma (เช่น "SKPTA,SKPLE")
    # ถ้าเว้นว่าง = ใช้กับทุก MC ใน (Factory, MC_CAT, Gauge) นั้น
    # sub-pool ต่อยอดจาก "พูล Group" ของเครื่องตัวนั้น (normal_pool) → ตรงกับ sub-pool total ด้านบน
    ms_entry = None
    ms_entry_general = None
    for (_k_fac, _k_cat, _k_mc, _k_g), _k_val in _MC_SPECIAL_LOOKUP.items():
        if _k_fac != fac_u or _k_cat != t1 or _k_g != g_norm:
            continue
        _members = [m.strip() for m in str(_k_mc).split(",") if m.strip()]
        if _members:
            if mc_u in _members:
                ms_entry = _k_val
                break
        else:
            ms_entry_general = _k_val
    if ms_entry is None:
        ms_entry = ms_entry_general
    sub_base = normal_pool   # พูล Group ของเครื่องตัวนี้ (เช่น "PHET|SKPLE , SKPTA 26G")
    if ms_entry:
        item_type = _get_item_cotton_poly(r.get("ITEM_CODE", ""))
        if item_type == "COTTON" and ms_entry["COTTON"] > 0:
            return f"{sub_base}:COTTON"
        if item_type == "POLY" and ms_entry["POLY"] > 0:
            return f"{sub_base}:POLY"

    # 2. Type-based description rules (Factory+Type ใน MC Special)
    mc_type = _MC_TO_TYPE.get((mc_u, g_norm), "").strip().upper()
    tdr_key = (fac_u, mc_type)
    if tdr_key in _TYPE_DESC_RULES:
        rule = _TYPE_DESC_RULES[tdr_key]
        rule_mc_cat = rule.get("mc_cat", "")
        # ตรวจ MC_CAT ตรงกับ rule และ ยกเว้น Gauge 20
        if (rule_mc_cat and t1 != rule_mc_cat) or g_norm == "20":
            return normal_pool
        booking_desc = str(r.get("DESCRIPTION", ""))
        if _is_description_special_type(booking_desc, rule["keywords"]):
            return f"{fac_pfx}{mc_type}:TYPE_SPECIAL"

    return normal_pool

df["POOL_TYPE"] = df.apply(_get_pool_type_for_row, axis=1)

# DESC_POOL_TYPE: secondary pool สำหรับ description-based rules
# กำหนดให้ทุก row ที่ match TYPE_SPECIAL description (ไม่ว่า POOL_TYPE หลักจะเป็นอะไร)
def _get_desc_pool_for_row(r):
    mc_u = str(r["MC_GROUP"]).strip().upper()
    g_norm = _norm_gauge_ava(r["GUAGE"])
    fac_u = _MC_TO_FACTORY.get((mc_u, g_norm), "").strip().upper()
    mc_type = _MC_TO_TYPE.get((mc_u, g_norm), "").strip().upper()
    tdr_key = (fac_u, mc_type)
    if tdr_key in _TYPE_DESC_RULES:
        rule = _TYPE_DESC_RULES[tdr_key]
        rule_mc_cat = rule.get("mc_cat", "")
        t1, _ = _mc_to_t1g.get((mc_u, g_norm), ("", ""))
        if (rule_mc_cat and t1 != rule_mc_cat) or g_norm == "20":
            return ""
        if _is_description_special_type(str(r.get("DESCRIPTION", "")), rule["keywords"]):
            fac_pfx = f"{fac_u}|" if fac_u else ""
            return f"{fac_pfx}{mc_type}:TYPE_SPECIAL"
    return ""

df["DESC_POOL_TYPE"] = df.apply(_get_desc_pool_for_row, axis=1)

df["FACTORY"] = df.apply(
    lambda r: _MC_TO_FACTORY.get((str(r["MC_GROUP"]).strip().upper(), _norm_gauge_ava(r["GUAGE"])), ""), axis=1
)
print(f"🔍 POOL_TYPE found: {(df['POOL_TYPE'] != '').sum()} rows / {len(df)} total")
print(f"🔍 Sample FACTORY+MC_CAT+GUAGE → POOL_TYPE:\n{df[['FACTORY','MC_CAT','GUAGE','POOL_TYPE']].drop_duplicates().head(15).to_string()}")

# TOTAL_MC: pool → ใช้ pool total; ไม่มี pool → ใช้ per-(MC_GROUP, GUAGE)
df["TOTAL_MC"] = df.apply(
    lambda r: _TOTAL_MC_BY_TYPE[r["POOL_TYPE"]]
              if r["POOL_TYPE"] and r["POOL_TYPE"] in _TOTAL_MC_BY_TYPE
              else TOTAL_MC_MAP.get((r["MC_GROUP"], r["GUAGE"]), 0),
    axis=1
)

# =========================
# MC CUMULATIVE
# =========================
df = df.sort_values(["MC_GROUP", "GUAGE", "WEEK"])
df["MC_USE_CUM"] = df.groupby(["MC_GROUP", "GUAGE", "WEEK"])["MC_USE_CEIL"].cumsum()

# Pool-level usage per WEEK: รวม MC_USE_CEIL ทุก MC_GROUP ใน pool เดียวกัน
_pool_use = (
    df[df["POOL_TYPE"] != ""]
    .groupby(["POOL_TYPE", "WEEK"])["MC_USE_CEIL"]
    .sum()
    .reset_index()
    .rename(columns={"MC_USE_CEIL": "_POOL_USE_TOTAL"})
)
df = df.merge(_pool_use, on=["POOL_TYPE", "WEEK"], how="left")

# DESC pool usage per WEEK: รวม MC_USE_CEIL จากทุก row ที่ match TYPE_SPECIAL
# (รวมทั้ง POLY primary + TYPE_SPECIAL primary เพื่อหักเครื่องทั้ง 2 pool พร้อมกัน)
_desc_pool_use = (
    df[df["DESC_POOL_TYPE"] != ""]
    .groupby(["DESC_POOL_TYPE", "WEEK"])["MC_USE_CEIL"]
    .sum()
    .reset_index()
    .rename(columns={"MC_USE_CEIL": "_DESC_POOL_USE_TOTAL", "DESC_POOL_TYPE": "_dpk"})
)
df = df.merge(
    _desc_pool_use.rename(columns={"_dpk": "DESC_POOL_TYPE"}),
    on=["DESC_POOL_TYPE", "WEEK"],
    how="left"
)

# Non-pool rows: total MC_USE_CEIL per MC_CAT+GUAGE+WEEK (เหมือน pool rows)
_nonpool_use = (
    df[df["POOL_TYPE"] == ""]
    .groupby(["MC_CAT", "GUAGE", "WEEK"])["MC_USE_CEIL"]
    .sum()
    .reset_index()
    .rename(columns={"MC_USE_CEIL": "_NONPOOL_USE_TOTAL"})
)
df = df.merge(_nonpool_use, on=["MC_CAT", "GUAGE", "WEEK"], how="left")

# TOTAL_MC_REMAIN:
# - TYPE_SPECIAL primary rows → ใช้ desc_pool_use (รวม POLY+TYPE_SPECIAL ทั้งหมด)
# - pool rows ที่ match TYPE_SPECIAL description (เช่น FD4/FD5 ใน POLY pool) → min(pool remain, TYPE_SPECIAL remain)
# - pool rows อื่น → pool total - _pool_use
# - non-pool rows → TOTAL_MC - MC_USE_CUM
_desc_pool_total_s = df["DESC_POOL_TYPE"].map(
    lambda pk: _TOTAL_MC_BY_TYPE.get(pk, 0) if pk else 0
)
_type_special_remain = _desc_pool_total_s - df["_DESC_POOL_USE_TOTAL"].fillna(0)

df["POOL_TYPE"] = df["POOL_TYPE"].fillna("").astype(str)
df["DESC_POOL_TYPE"] = df["DESC_POOL_TYPE"].fillna("").astype(str)

df["TOTAL_MC_REMAIN"] = np.where(
    df["POOL_TYPE"].str.endswith(":TYPE_SPECIAL", na=False),
    df["TOTAL_MC"] - df["_DESC_POOL_USE_TOTAL"].fillna(0),
    np.where(
        (df["POOL_TYPE"] != "") & (df["DESC_POOL_TYPE"] != ""),
        np.minimum(
            df["TOTAL_MC"] - df["_POOL_USE_TOTAL"].fillna(0),
            _type_special_remain
        ),
        np.where(
            df["POOL_TYPE"] != "",
            df["TOTAL_MC"] - df["_POOL_USE_TOTAL"].fillna(0),
            df["TOTAL_MC"] - df["_NONPOOL_USE_TOTAL"].fillna(0)
        )
    )
)
df = df.drop(columns=["_POOL_USE_TOTAL", "_NONPOOL_USE_TOTAL"])

# DESC_MC_REMAIN: เครื่องคงเหลือใน TYPE_SPECIAL pool สำหรับ row ที่ match description
# (แสดงเฉพาะ row ที่มี DESC_POOL_TYPE — ทั้ง POLY primary และ TYPE_SPECIAL primary)
df["DESC_POOL_TOTAL"] = df["DESC_POOL_TYPE"].map(
    lambda pk: _TOTAL_MC_BY_TYPE.get(pk, 0) if pk else 0
)
df["DESC_MC_REMAIN"] = np.where(
    df["DESC_POOL_TYPE"] != "",
    df["DESC_POOL_TOTAL"] - df["_DESC_POOL_USE_TOTAL"].fillna(0),
    np.nan
)
df = df.drop(columns=["_DESC_POOL_USE_TOTAL", "DESC_POOL_TOTAL"])

# =========================
# SUMMARY
# - Pool groups  → group by POOL_TYPE (Type ใน MasterMC), TOTAL_MC = pool total
# - Non-pool groups → group by MC_CAT + GUAGE เหมือนเดิม
# =========================
weeks = sorted(df["WEEK"].unique())
summary_parts = []

# --- Pool summary: แสดง MC_CAT+GUAGE ทุกตัวใน MasterMC (รวมที่ไม่มี booking) ---
pool_df = df[df["POOL_TYPE"] != ""]
# Pool usage from booking
if not pool_df.empty:
    pool_use_week = pool_df.groupby(["POOL_TYPE", "WEEK"], as_index=False)["MC_USE_CEIL"].sum()
else:
    pool_use_week = pd.DataFrame(columns=["POOL_TYPE", "WEEK", "MC_USE_CEIL"])
# Pool total per POOL_TYPE
pool_total_df = pd.DataFrame(list(_TOTAL_MC_BY_TYPE.items()), columns=["POOL_TYPE", "TOTAL_MC"])
# Build pool combos จาก MasterMC ทั้งหมด (ใช้ Type_1 raw = CAT ตรงกับ booking)
_pool_base_records = []
for _, _mmc_row in _master_mc_df.iterrows():
    _mc_u = str(_mmc_row.get("MC", "")).strip().upper()
    _g_norm = _norm_gauge_ava(_mmc_row.get("Guage", ""))
    _t1_raw = str(_mmc_row.get("MC_CAT", "")).strip()
    _g_raw = str(_mmc_row.get("Guage", "")).strip()
    _fac_raw = str(_mmc_row.get("Factory", "")).strip()
    if not _mc_u or not _t1_raw or not _g_norm:
        continue
    _pk = _MC_TO_POOL.get((_mc_u, _g_norm), "")
    if _pk:
        _pool_base_records.append({"FACTORY": _fac_raw, "MC_CAT": _t1_raw, "GUAGE": _g_raw, "POOL_TYPE": _pk})
pool_combos = (
    pd.DataFrame(_pool_base_records).drop_duplicates()
    if _pool_base_records
    else pd.DataFrame(columns=["FACTORY", "MC_CAT", "GUAGE", "POOL_TYPE"])
)
pool_combos = pool_combos.merge(pool_total_df, on="POOL_TYPE", how="left").dropna(subset=["TOTAL_MC"])
if not pool_combos.empty:
    week_df = pd.DataFrame({"WEEK": weeks})
    pool_combos["key"] = 1
    week_df["key"] = 1
    pool_base = pool_combos.merge(week_df, on="key").drop(columns="key")
    pool_sum = pool_base.merge(pool_use_week, on=["POOL_TYPE", "WEEK"], how="left")
    pool_sum["MC_USE_CEIL"] = pool_sum["MC_USE_CEIL"].fillna(0).astype(int)
    pool_sum["TOTAL_MC_REMAIN"] = pool_sum["TOTAL_MC"] - pool_sum["MC_USE_CEIL"]
    pool_sum["MC_POOL"] = pool_sum["POOL_TYPE"]   # ป้ายพูล = POOL_TYPE (แยก SKP vs SKPTA/SKPLE)
    summary_parts.append(pool_sum[["FACTORY", "MC_CAT", "GUAGE", "MC_POOL", "TOTAL_MC", "WEEK", "MC_USE_CEIL", "TOTAL_MC_REMAIN"]])

# --- Non-pool summary: เฉพาะ MC_CAT+GUAGE ที่มี data จริงใน nonpool_df เท่านั้น ---
nonpool_df = df[df["POOL_TYPE"] == ""]
if not nonpool_df.empty:
    mc_use_week = nonpool_df.groupby(["MC_CAT", "GUAGE", "WEEK"], as_index=False)["MC_USE_CEIL"].sum()
    # ดึง TOTAL_MC เฉพาะ combo ที่มีใน nonpool_df จริง
    _np_combos = set(zip(nonpool_df["MC_CAT"].str.strip(), nonpool_df["GUAGE"].astype(str).str.strip()))
    mc_master_np = pd.DataFrame(
        [(k[0], k[1], v) for k, v in TOTAL_MC_MAP_TYPE1.items()
         if (k[0], k[1]) in _np_combos],
        columns=["MC_CAT", "GUAGE", "TOTAL_MC"],
    )
    if not mc_master_np.empty:
        week_df2 = pd.DataFrame({"WEEK": weeks})
        mc_master_np["key"] = 1
        week_df2["key"] = 1
        summary_base = mc_master_np.merge(week_df2, on="key").drop(columns="key")
        summary_np = summary_base.merge(mc_use_week, on=["MC_CAT", "GUAGE", "WEEK"], how="left")
        summary_np["MC_USE_CEIL"] = summary_np["MC_USE_CEIL"].fillna(0).astype(int)
        summary_np["TOTAL_MC_REMAIN"] = summary_np["TOTAL_MC"] - summary_np["MC_USE_CEIL"]
        summary_np["FACTORY"] = ""
        summary_np["MC_POOL"] = ""   # non-pool ไม่มีป้ายพูล (consumer ใช้ cat|gauge)
        summary_parts.append(summary_np[["FACTORY", "MC_CAT", "GUAGE", "MC_POOL", "TOTAL_MC", "WEEK", "MC_USE_CEIL", "TOTAL_MC_REMAIN"]])

summary = (
    pd.concat(summary_parts, ignore_index=True)
    if summary_parts
    else pd.DataFrame(columns=["FACTORY", "MC_CAT", "GUAGE", "MC_POOL", "TOTAL_MC", "WEEK", "MC_USE_CEIL", "TOTAL_MC_REMAIN"])
)

# 🔧 กันนับเครื่องซ้ำ: แถว non-pool (FACTORY="") ดึง TOTAL_MC = grand-total ต่อ MC_CAT+GUAGE
#    ซึ่งเป็นเครื่องชุดเดียวกับที่นับใน pool row ไปแล้ว → ถ้ามี pool row ของ (MC_CAT,GUAGE,WEEK)
#    เดียวกันอยู่แล้ว ให้ตัดแถว non-pool ทิ้ง (ไม่งั้น consumer ที่บวกรวม เช่น get_actual_mc_remain
#    และหน้าเว็บ จะได้เครื่องว่างเกินจริง)
if not summary.empty and "FACTORY" in summary.columns:
    def _cg_key(r):  # normalize เพื่อกัน format เกจ int/str/ช่องว่างไม่ตรงกัน
        return (str(r["MC_CAT"]).strip().upper(), _norm_gauge_ava(r["GUAGE"]), str(r["WEEK"]).strip())
    _facn = summary["FACTORY"].astype(str).str.strip()
    _pool_keys = set(summary.loc[_facn != ""].apply(_cg_key, axis=1)) if (_facn != "").any() else set()
    _dup_np = summary.apply(
        lambda r: str(r["FACTORY"]).strip() == "" and _cg_key(r) in _pool_keys,
        axis=1,
    )
    if _dup_np.any():
        print(f"🔧 ตัดแถว non-pool ซ้ำซ้อน {int(_dup_np.sum())} แถว (มี pool row แล้ว)")
        summary = summary[~_dup_np].reset_index(drop=True)

# เครื่องที่กันไว้ (POLY/COTTON) ต่อ MC_CAT+GUAGE — ให้หน้าเว็บโชว์แยกว่าเครื่องไม่ได้หาย
# (เครื่องกลุ่มนี้อยู่ใน sub-pool ใช้แทนงานปกติไม่ได้ จึงไม่รวมใน TOTAL_MC ปกติ)
_rsv_acc: dict = {}
for (_rf, _rcat, _rmc, _rg), _rv in _MC_SPECIAL_LOOKUP.items():
    # ผูกเครื่องกันไว้เข้ากับ "พูล Group" ของเครื่องที่ระบุ (เช่น SKPTA/SKPLE) ไม่ใช่ทั้ง cat+gauge
    # → SKP จะไม่ถูกโชว์ว่ามี POLY กัน
    _r_first = _rmc.split(",")[0].strip() if _rmc else ""
    _r_grp = _mc_to_group.get((_r_first, _rg), "") if _r_first else ""
    _r_pfx = f"{_rf}|" if _rf else ""
    _r_pool = f"{_r_pfx}{_r_grp}" if _r_grp else f"{_r_pfx}{_rcat}:{_rg}"
    _slot = _rsv_acc.setdefault((_rf, _rcat, _rg, _r_pool), {"POLY": 0, "COTTON": 0})
    _slot["POLY"] += int(_rv.get("POLY", 0) or 0)
    _slot["COTTON"] += int(_rv.get("COTTON", 0) or 0)
reserved_summary = (
    pd.DataFrame(
        [{"FACTORY": _k[0], "MC_CAT": _k[1], "GUAGE": _k[2], "MC_POOL": _k[3],
          "POLY": _v["POLY"], "COTTON": _v["COTTON"]}
         for _k, _v in _rsv_acc.items()]
    )
    if _rsv_acc
    else pd.DataFrame(columns=["FACTORY", "MC_CAT", "GUAGE", "MC_POOL", "POLY", "COTTON"])
)

# เครื่อง POLY/COTTON ที่ booking ใช้ไปแล้ว "ต่อสัปดาห์" (จาก sub-pool rows ที่ POOL_TYPE
# ลงท้าย :POLY / :COTTON) → ให้เว็บ/Planning คำนวณเครื่องกันไว้ที่เหลือจริง = กันไว้ − booking
# (เดิม MC_RESERVED บอกแค่ยอดกันไว้ ไม่รู้ว่า booking กินไปเท่าไหร่ → ชิปโชว์เหมือนยังว่าง)
def _reserved_weekly_usage(_suffix):
    # MC_POOL = POOL_TYPE ตัด suffix :POLY/:COTTON ออก → ป้ายพูล Group เดียวกับ SUMMARY
    _cols = ["MC_CAT", "GUAGE", "MC_POOL", "WEEK", "_U"]
    if "POOL_TYPE" not in df.columns:
        return pd.DataFrame(columns=_cols)
    _sub = df[df["POOL_TYPE"].astype(str).str.endswith(_suffix)].copy()
    if _sub.empty:
        return pd.DataFrame(columns=_cols)
    _sub["MC_POOL"] = _sub["POOL_TYPE"].astype(str).str[:-len(_suffix)]
    return (
        _sub.groupby(["MC_CAT", "GUAGE", "MC_POOL", "WEEK"], as_index=False)["MC_USE_CEIL"].sum()
        .rename(columns={"MC_USE_CEIL": "_U"})
    )

_rw_poly = _reserved_weekly_usage(":POLY").rename(columns={"_U": "POLY_USED"})
_rw_cotton = _reserved_weekly_usage(":COTTON").rename(columns={"_U": "COTTON_USED"})
reserved_weekly_summary = _rw_poly.merge(_rw_cotton, on=["MC_CAT", "GUAGE", "MC_POOL", "WEEK"], how="outer")
if not reserved_weekly_summary.empty:
    reserved_weekly_summary["MC_CAT"] = reserved_weekly_summary["MC_CAT"].astype(str).str.strip()
    reserved_weekly_summary["GUAGE"] = reserved_weekly_summary["GUAGE"].astype(str).str.strip()
    reserved_weekly_summary["MC_POOL"] = reserved_weekly_summary["MC_POOL"].astype(str).str.strip()
    reserved_weekly_summary["WEEK"] = reserved_weekly_summary["WEEK"].astype(int)
    for _c in ("POLY_USED", "COTTON_USED"):
        reserved_weekly_summary[_c] = reserved_weekly_summary[_c].fillna(0).astype(int)
    reserved_weekly_summary = reserved_weekly_summary.sort_values(["MC_CAT", "GUAGE", "MC_POOL", "WEEK"])
else:
    reserved_weekly_summary = pd.DataFrame(columns=["MC_CAT", "GUAGE", "MC_POOL", "WEEK", "POLY_USED", "COTTON_USED"])

# =========================
# AVAILABILITY SUMMARY: WEEK 25-35 (แยกรายสัปดาห์ เรียงไปทางขวา)
# =========================
from datetime import date as _date, timedelta as _td
# สัปดาห์ปัจจุบัน (ศุกร์–พฤหัส = ISO week ของ วันที่+3) → ย้อนหลัง 5 สัปดาห์ ไปข้างหน้าจนหมด
_AVA_CUR_WEEK = (_date.today() + _td(days=3)).isocalendar()[1]
_ava_all_weeks = set(summary["WEEK"].unique())

def _ava_week_delta(_w):
    # ระยะห่าง(สัปดาห์)จากสัปดาห์ปัจจุบัน รองรับข้ามปี (ISO week วน 1..52)
    _d = int(_w) - _AVA_CUR_WEEK
    if _d > 26:
        _d -= 52
    elif _d < -26:
        _d += 52
    return _d

# ย้อนหลัง 5 สัปดาห์ (รวมข้ามปี) → ไปข้างหน้าจนหมด, ตัด week 99
ava_weeks = sorted(
    (w for w in _ava_all_weeks if w != 99 and _ava_week_delta(w) >= -5),
    key=_ava_week_delta,
)
AVA_WK_START = ava_weeks[0]  if ava_weeks else _AVA_CUR_WEEK
AVA_WK_END   = ava_weeks[-1] if ava_weeks else _AVA_CUR_WEEK
AVA_SHEET = "AVA_MC"

# map หมวดที่ต้องรวมแสดง: SYN-30 / SYN-28 → DOUBLE-30
_AVA_CAT_MAP = {"SYN-30": "DOUBLE-30", "SYN-28": "DOUBLE-30"}
def _ava_cat_disp(_s):
    # MC_CAT ที่ใช้แสดง/จัดกลุ่ม (upper + รวม SYN เข้า DOUBLE-30)
    return _s.astype(str).str.strip().str.upper().replace(_AVA_CAT_MAP)
def _ava_catkey(_s):
    # key match ระหว่าง df (SINGLE) กับ summary/MasterMC (SINGEL)
    return _s.str.replace("SINGEL", "SINGLE", regex=False)

# AVA ใช้ Total mc ตรงจาก MasterMC โดยตรง
# เพื่อไม่ให้โดนบวกซ้ำจาก SUMMARY ที่มีทั้ง normal pool / special pool / non-pool พร้อมกัน
_ava_master = _master_mc_df[["MC_CAT", "Guage", "Total MC"]].copy()
_ava_master["MC_CAT"] = _ava_cat_disp(_ava_master["MC_CAT"])
_ava_master["GUAGE"] = _ava_master["Guage"].map(_norm_gauge_ava)
_ava_master["Total mc"] = pd.to_numeric(_ava_master["Total MC"], errors="coerce").fillna(0)
_ava_master_tot = (
    _ava_master.groupby(["MC_CAT", "GUAGE"], as_index=False)["Total mc"]
    .sum()
)
_ava_master_tot["_CATKEY"] = _ava_catkey(_ava_master_tot["MC_CAT"])
_ava_master_tot["_GKEY"] = _ava_master_tot["GUAGE"].astype(str).str.strip()

# MC_USE_Cat = ผลรวมจำนวนเครื่องที่ใช้จริง ต่อ MC_CAT(display) + GUAGE ต่อ WEEK (จาก DETAIL)
_udf = df[["MC_CAT", "GUAGE", "WEEK", "MC_USE_CEIL"]].copy()
_udf["MC_CAT"] = _ava_cat_disp(_udf["MC_CAT"])
_udf["GUAGE"] = _udf["GUAGE"].map(_norm_gauge_ava)
_use_all = (
    _udf.groupby(["MC_CAT", "GUAGE", "WEEK"], as_index=False)["MC_USE_CEIL"]
    .sum()
    .rename(columns={"MC_USE_CEIL": "MC_USE_Cat"})
)
_use_all["_CATKEY"] = _ava_catkey(_use_all["MC_CAT"])
_use_all["_GKEY"] = _use_all["GUAGE"].astype(str).str.strip()

# KP_Cat = ผลรวม KP_WEIGHT ต่อ MC_CAT(display) + GUAGE ต่อ WEEK (จาก DETAIL)
_kpdf = df[["MC_CAT", "GUAGE", "WEEK", "KP_WEIGHT"]].copy()
_kpdf["MC_CAT"] = _ava_cat_disp(_kpdf["MC_CAT"])
_kpdf["GUAGE"] = _kpdf["GUAGE"].map(_norm_gauge_ava)
_kp_all = (
    _kpdf.groupby(["MC_CAT", "GUAGE", "WEEK"], as_index=False)["KP_WEIGHT"].sum()
    .rename(columns={"KP_WEIGHT": "KP_Cat"})
)
_kp_all["_CATKEY"] = _ava_catkey(_kp_all["MC_CAT"])
_kp_all["_GKEY"] = _kp_all["GUAGE"].astype(str).str.strip()

# row index ร่วม = union ของ MasterMC และข้อมูลจริงใน booking
_ava_idx_df = pd.concat(
    [
        _ava_master_tot[["MC_CAT", "GUAGE"]],
        _use_all[["MC_CAT", "GUAGE"]],
        _kp_all[["MC_CAT", "GUAGE"]],
    ],
    ignore_index=True,
).drop_duplicates()
# IMPORT = ผ้านำเข้า ไม่ได้ถักเอง → ไม่ต้องโชว์ในชีท AVA (ยังคงอยู่ใน DETAIL เหมือนเดิม)
_AVA_HIDE_CAT = {"IMPORT"}
# FQC ไม่ใช่เครื่องถักใน MasterMC → ดันไปไว้ล่างสุดของชีท ไม่ให้แทรกกลางกลุ่ม DOUBLE/SINGLE
_AVA_BOTTOM_CAT = set(FQC_MC_CAT)
_ava_idx_df = _ava_idx_df[
    ~_ava_idx_df["MC_CAT"].astype(str).str.strip().str.upper().isin(_AVA_HIDE_CAT)
]
_ava_idx_df["_ORD"] = _ava_idx_df["MC_CAT"].astype(str).str.strip().str.upper().isin(_AVA_BOTTOM_CAT).astype(int)
_ava_idx_df = (
    _ava_idx_df.sort_values(["_ORD", "MC_CAT", "GUAGE"])
    .drop(columns="_ORD")
)
_ava_idx = pd.MultiIndex.from_frame(_ava_idx_df)

_ava_base = _ava_idx_df.copy()
_ava_base["_CATKEY"] = _ava_catkey(_ava_base["MC_CAT"])
_ava_base["_GKEY"] = _ava_base["GUAGE"].astype(str).str.strip()
_ava_base = _ava_base.merge(
    _ava_master_tot[["_CATKEY", "_GKEY", "Total mc"]],
    on=["_CATKEY", "_GKEY"],
    how="left"
)
_ava_base["Total mc"] = _ava_base["Total mc"].fillna(0)

_ava_blocks = []
for _wk in ava_weeks:
    _usew = _use_all[_use_all["WEEK"] == _wk][["_CATKEY", "_GKEY", "MC_USE_Cat"]]
    _kpw = _kp_all[_kp_all["WEEK"] == _wk][["_CATKEY", "_GKEY", "KP_Cat"]]
    _m = _ava_base.merge(_usew, on=["_CATKEY", "_GKEY"], how="left")
    _m = _m.merge(_kpw, on=["_CATKEY", "_GKEY"], how="left")
    _m["MC_USE_Cat"] = _m["MC_USE_Cat"].fillna(0)
    _m["KP_Cat"] = _m["KP_Cat"].fillna(0)
    # Dif/%ava คิดราย gauge: Total mc − MC_USE_Cat
    _m["Dif"] = _m["Total mc"] - _m["MC_USE_Cat"]
    _m["%ava"] = (_m["Dif"] / _m["Total mc"]).where(_m["Total mc"] != 0)
    _blk = (
        _m.set_index(["MC_CAT", "GUAGE"])[["Total mc", "KP_Cat", "MC_USE_Cat", "Dif", "%ava"]]
        .reindex(_ava_idx)
    )
    _blk.columns = pd.MultiIndex.from_product([[f"WEEK {_wk}"], _blk.columns])
    _ava_blocks.append(_blk)

ava_data = (
    pd.concat(_ava_blocks, axis=1)
    if _ava_blocks
    else pd.DataFrame(index=_ava_idx)
)

# ---------- แทรกแถว subtotal แยกกลุ่ม DOUBLE / SINGEL ----------
def _ava_group(_cat):
    _c = str(_cat).strip().upper().replace("SINGEL", "SINGLE")
    if _c.startswith("DOUBLE"):
        return "DOUBLE"
    if _c.startswith("SINGLE"):
        return "SINGLE"
    return "OTHER"

def _ava_subtotal(_members):
    _sub = ava_data.loc[_members]
    _vals = []
    for _col in ava_data.columns:
        _w, _metric = _col
        if _metric == "%ava":
            _st = _sub[(_w, "Total mc")].sum()
            _d = _sub[(_w, "Dif")].sum()
            _vals.append((_d / _st) if _st else None)
        else:
            _vals.append(_sub[_col].sum())
    return pd.Series(_vals, index=ava_data.columns)

_ord_idx, _ord_rows, _sub_pos = [], [], []
_cur_g, _cur_mem = None, []
for _idx in ava_data.index:
    _g2 = _ava_group(_idx[0])
    if _cur_g is None:
        _cur_g = _g2
    if _g2 != _cur_g:
        _ord_idx.append((f"{_cur_g} Total", ""))
        _ord_rows.append(_ava_subtotal(_cur_mem))
        _sub_pos.append(len(_ord_idx) - 1)
        _cur_g, _cur_mem = _g2, []
    _ord_idx.append(_idx)
    _ord_rows.append(ava_data.loc[_idx])
    _cur_mem.append(_idx)
if _cur_mem:
    _ord_idx.append((f"{_cur_g} Total", ""))
    _ord_rows.append(_ava_subtotal(_cur_mem))
    _sub_pos.append(len(_ord_idx) - 1)

if _ord_rows:
    ava_wide = pd.DataFrame(_ord_rows)
    ava_wide.index = pd.MultiIndex.from_tuples(_ord_idx, names=["MC_CAT", "GUAGE"])
    ava_wide.columns = ava_data.columns
else:
    ava_wide = ava_data
# แถว subtotal (1-based excel row): ข้อมูลเริ่มแถว 4
_ava_sub_excel_rows = [4 + _p for _p in _sub_pos]

# =========================
# AVA_MC_ITEM — แตกรายละเอียดระดับ ITEM
# index: MC_CAT | MC | GUAGE | ITEM | TYPE | CAP
#   TYPE = FQC ถ้ามาจากเครื่องกลุ่ม FQC, ไม่งั้น NORMAL
#   CAP  = CAP_KNIT หลังปรับ 20/24 (_CAP_ADJ)
# บล็อกรายสัปดาห์: Total mc / Dif / %ava = ค่าระดับ MC_CAT+GUAGE (เท่ากับชีท AVA_MC)
#                  KP / MC_USE          = ค่าของ item แถวนั้น
# =========================
ITEM_SHEET = "AVA_MC_ITEM"

_it = df[["MC_CAT", "MC_GROUP", "GUAGE", "ITEM_CODE", "WEEK", "KP_WEIGHT", "MC_USE_CEIL"]].copy()
_it = _it[~_it["MC_CAT"].astype(str).str.strip().str.upper().isin(_AVA_HIDE_CAT)]
_it = _it.merge(_CAP_ADJ_SNAP, on=["MC_GROUP", "GUAGE", "ITEM_CODE"], how="left")
_it["TYPE"] = np.where(_it["MC_GROUP"].isin(FQC_MC_GROUP), "FQC", "NORMAL")
_it["MC_CAT"] = _ava_cat_disp(_it["MC_CAT"])
_it["GUAGE"] = _it["GUAGE"].map(_norm_gauge_ava)
_it["MC"] = _it["MC_GROUP"].astype(str).str.strip().str.upper()
_it["ITEM"] = _it["ITEM_CODE"].astype(str).str.strip()
_it["CAP"] = pd.to_numeric(_it["CAP"], errors="coerce").fillna(0).round(2)

_IT_KEYS = ["MC_CAT", "MC", "GUAGE", "ITEM", "TYPE", "CAP"]
_it_wk = (
    _it.groupby(_IT_KEYS + ["WEEK"], as_index=False)
    .agg(KP=("KP_WEIGHT", "sum"), MC_USE=("MC_USE_CEIL", "sum"))
)

_it_idx_df = _it_wk[_IT_KEYS].drop_duplicates()
# FQC / IMPORT ไปล่างสุดเหมือนชีท AVA_MC
_it_idx_df["_ORD"] = _it_idx_df["MC_CAT"].astype(str).str.strip().str.upper().isin(_AVA_BOTTOM_CAT).astype(int)
_it_idx_df = _it_idx_df.sort_values(["_ORD", "MC_CAT", "GUAGE", "MC", "ITEM"]).drop(columns="_ORD")
_it_idx = pd.MultiIndex.from_frame(_it_idx_df)

# Total mc ระดับ MC_CAT+GUAGE (ชุดเดียวกับชีท AVA_MC)
_it_base = _it_idx_df.copy()
_it_base["_CATKEY"] = _ava_catkey(_it_base["MC_CAT"].astype(str).str.strip().str.upper())
_it_base["_GKEY"] = _it_base["GUAGE"].astype(str).str.strip()
_it_base = _it_base.merge(
    _ava_master_tot[["_CATKEY", "_GKEY", "Total mc"]], on=["_CATKEY", "_GKEY"], how="left"
)
_it_base["Total mc"] = _it_base["Total mc"].fillna(0)

_it_blocks = []
for _wk in ava_weeks:
    _iw = _it_wk[_it_wk["WEEK"] == _wk][_IT_KEYS + ["KP", "MC_USE"]]
    _gw = _use_all[_use_all["WEEK"] == _wk][["_CATKEY", "_GKEY", "MC_USE_Cat"]]
    _m = _it_base.merge(_iw, on=_IT_KEYS, how="left").merge(_gw, on=["_CATKEY", "_GKEY"], how="left")
    _m["KP"] = _m["KP"].fillna(0)
    _m["MC_USE"] = _m["MC_USE"].fillna(0)
    _m["MC_USE_Cat"] = _m["MC_USE_Cat"].fillna(0)
    # Dif/%ava = ระดับ MC_CAT+GUAGE (เครื่องว่างของกลุ่ม ไม่ใช่ของ item)
    _m["Dif"] = _m["Total mc"] - _m["MC_USE_Cat"]
    _m["%ava"] = (_m["Dif"] / _m["Total mc"]).where(_m["Total mc"] != 0)
    _blk = (
        _m.set_index(_IT_KEYS)[["Total mc", "KP", "MC_USE", "Dif", "%ava"]]
        .reindex(_it_idx)
    )
    _blk.columns = pd.MultiIndex.from_product([[f"WEEK {_wk}"], _blk.columns])
    _it_blocks.append(_blk)

item_wide = (
    pd.concat(_it_blocks, axis=1)
    if _it_blocks
    else pd.DataFrame(index=_it_idx)
)

# =========================
# SAVE
# =========================
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# FQC โชว์เฉพาะชีท AVA_MC (สร้างเสร็จแล้วด้านบน) — ชีทที่ Planning.py/webapp อ่านต้องไม่มี FQC
_detail_out = df[~df["MC_GROUP"].isin(FQC_MC_GROUP)]
print(f"🧹 ตัด FQC ออกจากชีท DETAIL: {len(df) - len(_detail_out)} แถว (เหลือ {len(_detail_out)})")

def _drop_fqc_cat(_d):
    if _d.empty or "MC_CAT" not in _d.columns:
        return _d
    return _d[~_d["MC_CAT"].astype(str).str.strip().str.upper().isin(FQC_MC_CAT)]

# --- MC_POOL_MAP: (MC_CAT, GUAGE, MC_GROUP) → MC_POOL (ป้ายพูล Group) ---
# ให้ plan_view/หน้าเว็บ map เครื่องแต่ละตัว (เช่น SKP/SKPTA/SKPLE) ไปพูลที่ถูกต้อง
_pool_map_records = []
for _, _pm_row in _master_mc_df.iterrows():
    _pm_mc = str(_pm_row.get("MC", "")).strip().upper()
    _pm_g_norm = _norm_gauge_ava(_pm_row.get("Guage", ""))
    _pm_cat = str(_pm_row.get("MC_CAT", "")).strip()
    _pm_g_raw = str(_pm_row.get("Guage", "")).strip()
    if not _pm_mc or not _pm_cat or not _pm_g_norm:
        continue
    _pm_pk = _MC_TO_POOL.get((_pm_mc, _pm_g_norm), "")
    if _pm_pk:
        _pool_map_records.append({"MC_CAT": _pm_cat, "GUAGE": _pm_g_raw, "MC_GROUP": _pm_mc, "MC_POOL": _pm_pk})
pool_map_df = (
    pd.DataFrame(_pool_map_records).drop_duplicates()
    if _pool_map_records
    else pd.DataFrame(columns=["MC_CAT", "GUAGE", "MC_GROUP", "MC_POOL"])
)

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    _detail_out.to_excel(writer, sheet_name="DETAIL", index=False)
    _drop_fqc_cat(summary).to_excel(writer, sheet_name="SUMMARY_MC_REMAIN", index=False)
    _drop_fqc_cat(reserved_summary).to_excel(writer, sheet_name="MC_RESERVED", index=False)
    _drop_fqc_cat(reserved_weekly_summary).to_excel(writer, sheet_name="MC_RESERVED_WEEKLY", index=False)
    _drop_fqc_cat(pool_map_df).to_excel(writer, sheet_name="MC_POOL_MAP", index=False)
    ava_wide.to_excel(writer, sheet_name=AVA_SHEET)
    # freeze คอลัมน์ A-B (MC_CAT, GUAGE) + แถวหัวตาราง ให้ค้างเวลาเลื่อนดู week ทางขวา
    _ws = writer.sheets[AVA_SHEET]
    _ws.freeze_panes = "C4"

    # ---------- จัดรูปแบบชีท AVA ----------
    from openpyxl.styles import Font, Border, Side, PatternFill
    _n_rows = len(ava_wide)
    _n_weeks = len(ava_weeks)
    _first_row = 4                       # แถวข้อมูลแรก (1-2 หัว, 3 ชื่อ index)
    _last_row = 3 + _n_rows
    _NUM_FMT = "#,##0"                   # ไม่มีทศนิยม + ลูกน้ำ
    _PCT_FMT = "0%"                      # % ไม่มีทศนิยม
    _bold = Font(bold=True)
    _rb = Side(style="medium")
    _sub_fill = PatternFill("solid", fgColor="FFD9E1F2")
    _sub_set = set(_ava_sub_excel_rows)

    # 4) number format + 3) %ava เป็น % และ <=20% สีแดง + bold แถว subtotal
    for _r in range(_first_row, _last_row + 1):
        _is_sub = _r in _sub_set
        for _wi in range(_n_weeks):
            _base = 3 + _wi * 5
            for _c in range(_base, _base + 4):       # Total mc, KP_Cat, MC_USE_Cat, Dif
                _cell = _ws.cell(row=_r, column=_c)
                _cell.number_format = _NUM_FMT
                if _is_sub:
                    _cell.font = _bold
            _pa = _ws.cell(row=_r, column=_base + 4)  # %ava
            _pa.number_format = _PCT_FMT
            _low = isinstance(_pa.value, (int, float)) and _pa.value <= 0.20
            if _low or _is_sub:
                _pa.font = Font(bold=_is_sub, color=("FFFF0000" if _low else "FF000000"))

    # แถว subtotal: ตัวหนา + พื้นสี
    for _r in _sub_set:
        for _c in range(1, 3 + _n_weeks * 5):
            _cell = _ws.cell(row=_r, column=_c)
            _cell.fill = _sub_fill
            if _c <= 2:
                _cell.font = _bold

    # 2) เส้นขอบขวาปิดท้ายทุก week (คอลัมน์ %ava) ตลอดทั้งตาราง
    for _wi in range(_n_weeks):
        _bcol = 3 + _wi * 5 + 4
        for _r in range(1, _last_row + 1):
            _cell = _ws.cell(row=_r, column=_bcol)
            _e = _cell.border
            _cell.border = Border(left=_e.left, top=_e.top, bottom=_e.bottom, right=_rb)

    # ---------- ชีท AVA_MC_ITEM (ระดับ item) ----------
    item_wide.to_excel(writer, sheet_name=ITEM_SHEET)
    _wsi = writer.sheets[ITEM_SHEET]
    _wsi.freeze_panes = "G4"                 # ค้าง 6 คอลัมน์ index (MC_CAT..CAP)
    _i_first, _i_last = 4, 3 + len(item_wide)
    _I_BASE = 7                              # คอลัมน์แรกของบล็อก week (A-F = index)
    for _r in range(_i_first, _i_last + 1):
        for _wi in range(_n_weeks):
            _b = _I_BASE + _wi * 5
            _wsi.cell(row=_r, column=_b).number_format = _NUM_FMT       # Total mc
            _wsi.cell(row=_r, column=_b + 1).number_format = "#,##0.00" # KP
            _wsi.cell(row=_r, column=_b + 2).number_format = _NUM_FMT   # MC_USE
            _wsi.cell(row=_r, column=_b + 3).number_format = _NUM_FMT   # Dif
            _pa = _wsi.cell(row=_r, column=_b + 4)                       # %ava
            _pa.number_format = _PCT_FMT
            if isinstance(_pa.value, (int, float)) and _pa.value <= 0.20:
                _pa.font = Font(color="FFFF0000")
        _wsi.cell(row=_r, column=6).number_format = "#,##0.00"           # CAP
    for _wi in range(_n_weeks):
        _bcol = _I_BASE + _wi * 5 + 4
        for _r in range(1, _i_last + 1):
            _cell = _wsi.cell(row=_r, column=_bcol)
            _e = _cell.border
            _cell.border = Border(left=_e.left, top=_e.top, bottom=_e.bottom, right=_rb)

print("✅ AVA MC FINAL COMPLETE (COLLAR REMOVED)")
print("Saved:", OUTPUT_FILE)
print("DETAIL rows:", len(df))
print("SUMMARY rows:", len(summary))
print(f"{AVA_SHEET} rows:", len(ava_wide), "weeks:", ava_weeks)
print(f"{ITEM_SHEET} rows:", len(item_wide),
      "| FQC:", int((_it_idx_df['TYPE'] == 'FQC').sum()),
      "| NORMAL:", int((_it_idx_df['TYPE'] == 'NORMAL').sum()))
