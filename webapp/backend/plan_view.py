"""
plan_view.py — อ่าน/เขียนไฟล์แผนผลิต "ล่าสุด" (production_plan_*.xlsx) สำหรับหน้า "แผนผลิต"
- หาไฟล์ production_plan_*.xlsx ที่ mtime ใหม่สุดใน OUTPUT_DIR
- อ่านรายชีท → คืน columns + rows "ครบทุกแถว" (ไม่ตัด เพราะต้องบันทึกกลับได้)
- เขียนกลับ: สำรองไฟล์เดิมเป็น .bak ก่อน แล้วแก้เฉพาะชีทเป้าหมาย (ชีทอื่นคงเดิม)
"""
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

import config

_INT_RE = re.compile(r"^-?(0|[1-9][0-9]*)$")
_FLOAT_RE = re.compile(r"^-?(0|[1-9][0-9]*|0?)\.[0-9]+$")


# ชีทแผนหลักในไฟล์ production_plan_*.xlsx
# (เดิมมี PLAN_NO_S9 เป็นชีทซ้ำจาก pass ที่ปิด S9 — เลิกใช้ S9 fallback แล้วจึงเหลือชีทเดียว)
_MAIN_SHEET = "PLAN"


def _display_sheets(sheets) -> list:
    """รายชื่อชีทสำหรับ dropdown — เอา PLAN ขึ้นก่อน"""
    out = [s for s in sheets if s == _MAIN_SHEET]
    out += [s for s in sheets if s != _MAIN_SHEET]
    return out


def latest_path() -> Path | None:
    """ไฟล์ production_plan_*.xlsx ที่แก้ไขล่าสุด (None = ยังไม่มี)"""
    d = config.OUTPUT_DIR
    if not d.exists():
        return None
    files = sorted(d.glob("production_plan_*.xlsx"),
                   key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0] if files else None


def latest_meta() -> dict:
    """ข้อมูลไฟล์แผนล่าสุด + รายชื่อชีท (ให้ frontend โชว์ header)"""
    p = latest_path()
    if p is None:
        return {"exists": False, "name": None, "size": 0, "mtime": None, "sheets": []}
    st = p.stat()
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
        sheets = _display_sheets(wb.sheetnames)
        wb.close()
    except Exception as e:
        return {"exists": True, "name": p.name, "size": st.st_size,
                "mtime": st.st_mtime, "sheets": [], "error": str(e)}
    return {"exists": True, "name": p.name, "size": st.st_size,
            "mtime": st.st_mtime, "sheets": sheets}


def _round2(v):
    """ปัดลง (ตัดทศนิยมทิ้ง) เหลือ 2 ตำแหน่งสำหรับแสดงผล (จำนวนเต็ม/ชนิดอื่นคงเดิม)
    ใช้ Decimal จาก str เพื่อเลี่ยง error จาก float เช่น 1.13*100 = 112.999..."""
    if isinstance(v, float) and not v.is_integer():
        from decimal import Decimal, ROUND_DOWN
        return float(Decimal(str(v)).quantize(Decimal("0.01"), rounding=ROUND_DOWN))
    return v


def read_grid(sheet: str = None) -> dict:
    """อ่านชีทของไฟล์แผนล่าสุด → columns + rows (แถวแรก = header)
    data_only=True: คืนค่าที่คำนวณแล้วของเซลล์สูตร"""
    p = latest_path()
    if p is None:
        raise FileNotFoundError("ยังไม่มีไฟล์แผนผลิต — กรุณากดรันแผนก่อน")
    wb = load_workbook(p, read_only=True, data_only=True)
    sheets = _display_sheets(wb.sheetnames)
    mtime = p.stat().st_mtime
    if not sheets:
        wb.close()
        return {"name": p.name, "mtime": mtime, "sheet": None,
                "sheets": [], "columns": [], "rows": []}
    # ค่าเริ่มต้น = ชีทแรกใน dropdown (= 'PLAN')
    if sheet is None or sheet not in sheets:
        sheet = sheets[0]
    ws = wb[sheet]
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append(["" if v is None else _round2(v) for v in r])
    wb.close()

    if not rows:
        return {"name": p.name, "mtime": mtime, "sheet": sheet,
                "sheets": sheets, "columns": [], "rows": []}

    header = [str(c) if c != "" else f"Col{i+1}" for i, c in enumerate(rows[0])]
    width = len(header)
    data = [list(row)[:width] + [""] * (width - len(row)) for row in rows[1:]]
    return {"name": p.name, "mtime": mtime, "sheet": sheet,
            "sheets": sheets, "columns": header, "rows": data}


def _num(v):
    try:
        f = float(v)
        return int(f) if f.is_integer() else f
    except (TypeError, ValueError):
        return None


def _wk_str(v):
    """แปลงค่า WEEK → string เลขสัปดาห์ (29.0 → '29')"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            return str(int(v))
        except (ValueError, OverflowError):
            return None
    s = str(v).strip()
    return s or None


def _classify_type(factory, cat) -> str:
    """จำแนกแถว PLAN เป็นประเภท job (ตรงกับ REMAINING_JOBS): OM / PHET_DOUBLE / PHET_SINGLE
    None = ไม่นับ (เช่น OUTSOURCE) — ต้องตรงกับ classifyType() ใน PlanGantt.jsx"""
    f = str(factory).strip().upper()
    if f.startswith("OM"):
        return "OM"
    if f == "PHET":
        return "PHET_DOUBLE" if "DOUBLE" in str(cat).upper() else "PHET_SINGLE"
    return None


def _plan_new_mc_by_type() -> dict:
    """ผลรวม NEW_MC (job = setup 1 เครื่อง) ต่อ (สัปดาห์×ประเภท) จากชีท PLAN
    → { (week, TYPE): sum_new_mc }  ใช้แยกส่วน "job จาก booking" ออกจาก NEW_JOBS"""
    p = latest_path()
    if p is None:
        return {}
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
    except Exception:
        return {}
    _sheet = _MAIN_SHEET
    if _sheet not in wb.sheetnames:
        wb.close()
        return {}
    rows = list(wb[_sheet].iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {}
    header = [str(c).strip().upper() if c is not None else "" for c in rows[0]]

    def col(name):
        return header.index(name) if name in header else -1
    wi, fi, cati, nmi = col("PLAN_WEEK"), col("FACTORY_TYPE"), col("CAT"), col("NEW_MC")
    if wi < 0 or fi < 0 or cati < 0 or nmi < 0:
        return {}

    out: dict = {}
    for r in rows[1:]:
        if max(wi, fi, cati, nmi) >= len(r):
            continue
        wk = _wk_str(r[wi])
        typ = _classify_type(r[fi], r[cati])
        if wk is None or typ is None:
            continue
        nm = _num(r[nmi]) or 0
        out[(wk, typ)] = out.get((wk, typ), 0) + int(nm)
    return out


_LOAD_TYPES = {"OM", "PHET_DOUBLE", "PHET_SINGLE"}


def load_by_week() -> dict:
    """โหลดงานแยกตามประเภท (OM/PHET_DOUBLE/PHET_SINGLE) ต่อสัปดาห์
    → { week(str): { TYPE: {"old": int, "cap": number|None, "bookingNew": int} } }
    - old / new = "setup จริง" จากชีท SETUP_TRACKING (แยก PLAN_SOURCE OLD/NEW) — ตรงกับแผนจริง
      (ไม่ใช้ REMAINING_JOBS เพราะนับ job เกินจริง ไม่ตรงกับ SETUP_TRACKING)
    - cap มาจาก REMAINING_JOBS.CAPACITY (คงที่ 13/33/44)
    - bookingNew = new(SETUP_TRACKING) − ผลรวม NEW_MC ในแผน  → frontend คิด live:
      ใหม่ = bookingNew + ผลรวม NEW_MC ของแถวที่วางจริง (ตอนโหลดตรง setup จริง)"""
    p = latest_path()
    if p is None:
        return {}
    try:
        import pandas as pd
        xls = pd.ExcelFile(p)
    except Exception:
        return {}
    sheets = xls.sheet_names
    _rj_sheet = "REMAINING_JOBS"
    _st_sheet = "SETUP_TRACKING"

    # capacity (คงที่) จาก REMAINING_JOBS
    cap: dict = {}
    if _rj_sheet in sheets:
        try:
            rj = xls.parse(_rj_sheet)
        except Exception:
            rj = None
        if rj is not None and {"WEEK", "TYPE", "CAPACITY"} <= set(rj.columns):
            for _, r in rj.iterrows():
                wk = _wk_str(r["WEEK"])
                t = str(r["TYPE"]).strip().upper()
                if wk is None or t not in _LOAD_TYPES:
                    continue
                cap[(wk, t)] = _num(r["CAPACITY"])

    # old + new = setup จริงจาก SETUP_TRACKING (แยก PLAN_SOURCE)
    old: dict = {}
    new_base: dict = {}
    if _st_sheet in sheets:
        try:
            st = xls.parse(_st_sheet)
        except Exception:
            st = None
        need = {"PLAN_WEEK", "TYPE", "PLAN_SOURCE", "JOBS_DEDUCTED"}
        if st is not None and need <= set(st.columns):
            for _, r in st.iterrows():
                wk = _wk_str(r["PLAN_WEEK"])
                t = str(r["TYPE"]).strip().upper()
                if wk is None or t not in _LOAD_TYPES:
                    continue
                j = int(_num(r["JOBS_DEDUCTED"]) or 0)
                src = str(r["PLAN_SOURCE"]).strip().upper()
                if src == "OLD":
                    old[(wk, t)] = old.get((wk, t), 0) + j
                elif src == "NEW":
                    new_base[(wk, t)] = new_base.get((wk, t), 0) + j

    plan_new = _plan_new_mc_by_type()

    out: dict = {}
    for key in set(cap) | set(old) | set(new_base):
        wk, t = key
        booking_new = int(new_base.get(key, 0)) - int(plan_new.get(key, 0))
        out.setdefault(wk, {})[t] = {
            "old": int(old.get(key, 0)),
            "cap": cap.get(key),
            "bookingNew": booking_new,
        }
    return out


def setup_jobs_by_week() -> list:
    """รายการ job ที่ตั้งเครื่องใหม่ ต่อสัปดาห์ (ชีท SETUP_TRACKING ของไฟล์แผนล่าสุด)
    → [ {week, type, source, item, so, mcg, gauge, jobs, setup, carry, mc, days, qty} ]

    ใช้เปิดการ์ดรายละเอียดเมื่อคลิกแถบโหลด (Set up Job/Week) บน Gantt — แถบบอกแค่ยอดรวม
    ว่าใช้ไปกี่ job จากโควตา การ์ดต้องบอกได้ว่า job เหล่านั้นเป็น item ไหนบ้าง
    source = OLD (แผนเดิม/booking — ย้ายไม่ได้) / NEW (แผนรอบนี้)
    ยอดรวมของ source=OLD ตรงกับ old ใน load_by_week() เพราะอ่านชีทเดียวกัน"""
    p = latest_path()
    if p is None:
        return []
    try:
        import pandas as pd
        st = pd.read_excel(p, sheet_name="SETUP_TRACKING")
    except Exception:
        return []
    need = {"PLAN_WEEK", "TYPE", "PLAN_SOURCE", "ITEM_CODE", "JOBS_DEDUCTED"}
    if not need <= set(st.columns):
        return []

    def _c(name):
        return name if name in st.columns else None
    c_mcg, c_g = _c("MC_GROUP"), _c("MC_GUAGE")
    c_so, c_qty, c_days = _c("SC_SO_NO"), _c("KP_WEIGHT"), _c("SETUP_DAYS")
    c_setup, c_carry, c_mc = _c("SETUP_MC"), _c("CARRYOVER_MC"), _c("MC_THIS_WEEK")

    out = []
    for _, r in st.iterrows():
        wk = _wk_str(r["PLAN_WEEK"])
        typ = str(r["TYPE"]).strip().upper()
        if wk is None or typ not in _LOAD_TYPES:
            continue
        out.append({
            "week": wk,
            "type": typ,
            "source": str(r["PLAN_SOURCE"]).strip().upper(),
            "item": str(r["ITEM_CODE"]).strip(),
            "so": str(r[c_so]).strip() if c_so else "",
            "mcg": str(r[c_mcg]).strip() if c_mcg else "",
            "gauge": _gkey(r[c_g]) if c_g else "",
            "jobs": int(_num(r["JOBS_DEDUCTED"]) or 0),
            "setup": int(_num(r[c_setup]) or 0) if c_setup else 0,
            "carry": int(_num(r[c_carry]) or 0) if c_carry else 0,
            "mc": int(_num(r[c_mc]) or 0) if c_mc else 0,
            "days": _num(r[c_days]) or 0 if c_days else 0,
            "qty": round(float(_num(r[c_qty]) or 0), 2) if c_qty else 0.0,
        })
    return out


def _booking_ok(p: Path) -> bool:
    """ไฟล์ booking_final ใช้ได้ = เปิดเป็น xlsx ได้ + มีชีท SUMMARY_MC_REMAIN
    บาง run ของ AVA_MC ล้มเหลว เขียนไฟล์เปล่า ~2KB (ไม่ใช่ zip ด้วยซ้ำ) → ต้องข้าม
    ไม่งั้น _latest คว้าไฟล์เสียมา แล้ว ava/reserved ว่างทั้งที่ไฟล์ก่อนหน้าดีอยู่
    → เครื่องว่าง + คอลัมน์ 🔒 หายไปทั้งหน้า"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(p, read_only=True)
        try:
            return "SUMMARY_MC_REMAIN" in wb.sheetnames
        finally:
            wb.close()
    except Exception:
        return False


def _latest_booking_path() -> Path | None:
    """ไฟล์ booking_final_ready_*.xlsx ที่ "ใช้ได้" ใหม่สุด (ผลจาก AVA_MC — มีชีท SUMMARY_MC_REMAIN)
    ไล่จากใหม่→เก่า ข้ามไฟล์เสีย/เขียนไม่ครบ แล้วถอยไปไฟล์ดีที่ใหม่รองลงมา"""
    d = config.OUTPUT_DIR
    if not d.exists():
        return None
    files = sorted(d.glob("booking_final_ready*.xlsx"),
                   key=lambda p: p.stat().st_mtime, reverse=True)
    for p in files:
        if _booking_ok(p):
            return p
    return None


def _gkey(g) -> str:
    """normalize เกจ: 22.0 / '22 ' → '22'"""
    try:
        return str(int(g))
    except (TypeError, ValueError):
        return str(g).strip()


def booking_mc_by_item_week() -> dict:
    """เครื่องที่ booking ถักไอเทมนั้นอยู่แล้ว ต่อ (สัปดาห์ × ITEM|MC_GROUP|GUAGE)
    → { week(str): { "ITEM|MC_GROUP|GAUGE": int } }

    ใช้ให้ Gantt แยกออกว่าเครื่องของแถวแผนเป็น "เครื่อง booking" (อยู่กับสัปดาห์นั้น
    ไม่ย้ายตามงานที่ลาก และถูกหักจาก TOTAL_MC_REMAIN ไปแล้ว) หรือเป็นเครื่องที่แผน
    จองเพิ่มจากพูล (ต้องหักออกจากเครื่องว่าง และย้ายตามงาน)"""
    p = _latest_booking_path()
    if p is None:
        return {}
    try:
        import pandas as pd
        df = pd.read_excel(p, sheet_name="DETAIL")
    except Exception:
        return {}
    if not {"ITEM_CODE", "MC_GROUP", "GUAGE", "WEEK", "MC_USE_CEIL"} <= set(df.columns):
        return {}
    out: dict = {}
    for _, r in df.iterrows():
        wk = _wk_str(r["WEEK"])
        if wk is None:
            continue
        mc = int(_num(r["MC_USE_CEIL"]) or 0)
        if mc <= 0:
            continue
        key = "{}|{}|{}".format(
            str(r["ITEM_CODE"]).strip().upper(),
            str(r["MC_GROUP"]).strip().upper(),
            _gkey(r["GUAGE"]),
        )
        slot = out.setdefault(wk, {})
        slot[key] = slot.get(key, 0) + mc
    return out


def booking_items_by_week() -> list:
    """item ทั้งหมดจาก booking (ชีท DETAIL ของ booking_final ล่าสุด) — "แผนเก่า" ที่ commit แล้ว
    → [ {week, cat, gauge, mc_group, item, mc, setup, carry, qty, so, material} ]

    setup (_mc_increase) = เครื่องที่ต้องตั้งใหม่, carry (_prev_mc_use_ceil) = เครื่องอุ่นที่ยกมา
    ทั้งคู่มาจากลูกโซ่ carry/setup ของ AVA_MC และรวมกัน = mc (MC_USE_CEIL) เสมอ

    หน้าเว็บ overlay รายการนี้เป็นบล็อกอ่านอย่างเดียวบน Gantt (map เข้าแถว CAT|เกจ|เครื่อง
    และช่องสัปดาห์เดียวกับแผน) เพื่อให้ planner เห็นว่าเครื่องไหน/สัปดาห์ไหนมีงาน booking อยู่แล้ว
    รวมยอดต่อ (สัปดาห์×CAT×เกจ×เครื่อง×item) — DETAIL แตกได้หลายแถวต่อ item เดียวกัน"""
    p = _latest_booking_path()
    if p is None:
        return []
    try:
        import pandas as pd
        df = pd.read_excel(p, sheet_name="DETAIL")
    except Exception:
        return []
    need = {"MC_GROUP", "GUAGE", "ITEM_CODE", "WEEK", "CAT", "MC_USE_CEIL"}
    if not need <= set(df.columns):
        return []
    has_qty = "KP_WEIGHT" in df.columns
    has_so = "SO_NO" in df.columns
    has_mat = "MATERIAL_CONTENT" in df.columns
    # ไฟล์ booking เก่าอาจไม่มีคอลัมน์ลูกโซ่ carry/setup → ส่ง 0 ไป หน้าเว็บซ่อนคอลัมน์เอง
    has_setup = "_mc_increase" in df.columns
    has_carry = "_prev_mc_use_ceil" in df.columns

    agg: dict = {}
    for _, r in df.iterrows():
        wk = _wk_str(r["WEEK"])
        if wk is None:
            continue
        cat = str(r["CAT"]).strip()
        gauge = _gkey(r["GUAGE"])
        mcg = str(r["MC_GROUP"]).strip()
        item = str(r["ITEM_CODE"]).strip()
        key = (wk, cat, gauge, mcg, item)
        slot = agg.get(key)
        if slot is None:
            slot = {"week": wk, "cat": cat, "gauge": gauge, "mc_group": mcg,
                    "item": item, "mc": 0, "setup": 0, "carry": 0, "qty": 0.0,
                    "so": str(r["SO_NO"]).strip() if has_so else "",
                    "material": str(r["MATERIAL_CONTENT"]).strip() if has_mat else ""}
            agg[key] = slot
        slot["mc"] += int(_num(r["MC_USE_CEIL"]) or 0)
        if has_setup:
            slot["setup"] += int(_num(r["_mc_increase"]) or 0)
        if has_carry:
            slot["carry"] += int(_num(r["_prev_mc_use_ceil"]) or 0)
        if has_qty:
            slot["qty"] += float(_num(r["KP_WEIGHT"]) or 0)

    out = []
    for slot in agg.values():
        slot["qty"] = round(slot["qty"], 2)
        out.append(slot)
    return out


# หมายเหตุ: week_days() (อ่านชีท WEEK_DAYS จากไฟล์แผน) ถูกลบแล้ว —
# หน้าเว็บอ่านวันที่ปฏิทินเปิดสดจาก Calendar.xlsx ผ่าน /api/workday (cal_days) แทน
# เพื่อให้แก้ปฏิทินบน server แล้วเห็นผลทันทีโดยไม่ต้องรันแผนใหม่


def ava_by_week() -> dict:
    """เครื่องว่าง (AVA) ต่อ (CAT×เกจ×สัปดาห์) จากชีท SUMMARY_MC_REMAIN ของ booking_final ล่าสุด
    → { week(str): { "CAT|GUAGE": {"remain": int, "total": int, "used": int} } }
    - remain = TOTAL_MC_REMAIN = เครื่องจริง − เครื่องที่ booking จองไว้ (ยังไม่หักเครื่องที่แผนจอง)
    - frontend หักเครื่องที่แผนจองเพิ่มเอง = ACTUAL_MC − เครื่อง booking ของ item นั้น
      (ดู booking_mc_by_item_week) จึงไม่ต้องมี planBase อีก"""
    p = _latest_booking_path()
    if p is None:
        return {}
    try:
        import pandas as pd
        df = pd.read_excel(p, sheet_name="SUMMARY_MC_REMAIN")
    except Exception:
        return {}
    need = {"MC_CAT", "GUAGE", "WEEK", "TOTAL_MC", "MC_USE_CEIL", "TOTAL_MC_REMAIN"}
    if not need <= set(df.columns):
        return {}

    def gkey(g):
        try:
            return str(int(g))
        except (TypeError, ValueError):
            return str(g).strip()

    # แยกแถว pool (FACTORY มีค่า = เครื่องจริงต่อ pool ไม่ซ้ำ) ออกจากแถว non-pool
    # (FACTORY ว่าง = grand-total ที่เป็นเครื่องชุดเดียวกับ pool → นับซ้ำ)
    # ถ้ามี pool row ให้ใช้ pool อย่างเดียว; ไม่มีค่อยใช้ non-pool
    has_fac = "FACTORY" in df.columns
    has_pool = "MC_POOL" in df.columns
    split = _split_pool_set()   # (cat|gauge) ที่แยกพูล → key ต่อพูล (เช่น SKP vs SKPTA/SKPLE)
    pool_acc: dict = {}
    np_acc: dict = {}
    for _, r in df.iterrows():
        wk = _wk_str(r["WEEK"])
        if wk is None:
            continue
        cat = str(r["MC_CAT"]).strip()
        cg = f"{cat}|{gkey(r['GUAGE'])}"
        # กลุ่มที่แยกพูล → key = MC_POOL (เช่น "PHET|SKPLE , SKPTA 26G"); อื่นๆ = cat|gauge เดิม
        mc_pool = str(r.get("MC_POOL", "")).strip() if has_pool else ""
        key = mc_pool if (cg in split and mc_pool) else cg
        fac = str(r["FACTORY"]).strip() if has_fac else ""
        is_np = fac == "" or fac.lower() == "nan"
        acc = np_acc if is_np else pool_acc
        slot = acc.setdefault(wk, {}).setdefault(key, {"remain": 0, "total": 0, "used": 0})
        slot["remain"] += int(_num(r["TOTAL_MC_REMAIN"]) or 0)
        slot["total"] += int(_num(r["TOTAL_MC"]) or 0)
        slot["used"] += int(_num(r["MC_USE_CEIL"]) or 0)

    out: dict = {}
    for wk in set(pool_acc) | set(np_acc):
        for key in set(pool_acc.get(wk, {})) | set(np_acc.get(wk, {})):
            src = pool_acc.get(wk, {}).get(key) or np_acc.get(wk, {}).get(key)
            out.setdefault(wk, {})[key] = dict(src)

    # แนบเครื่องที่กันไว้ (POLY/COTTON) — ยอดกันไว้ไม่ขึ้นกับสัปดาห์ แต่ "ที่ booking ใช้ไป"
    # ขึ้นกับสัปดาห์ → แนบทั้ง poly/cotton (กันไว้) และ poly_used/cotton_used (booking ใช้ต่อสัปดาห์)
    # เพื่อให้ frontend คำนวณเครื่องกันไว้ที่เหลือจริง = กันไว้ − booking − แผน
    reserved = _reserved_by_cat()
    reserved_used = _reserved_used_by_cat_week()
    if reserved:
        for wk, keys in out.items():
            for key, slot in keys.items():
                rv = reserved.get(key)
                if rv and (rv["poly"] or rv["cotton"]):
                    used = reserved_used.get(wk, {}).get(key) or {}
                    slot["reserved"] = {
                        "poly": rv["poly"], "cotton": rv["cotton"],
                        "poly_used": int(used.get("poly_used", 0)),
                        "cotton_used": int(used.get("cotton_used", 0)),
                    }

    return out


def _pool_gkey(g):
    try:
        return str(int(g))
    except (TypeError, ValueError):
        return str(g).strip()


def _split_pool_set() -> set:
    """(cat|gauge) ที่ถูกแยกเป็นหลายพูล (เช่น SINGLE-32|26 = SKP vs SKPTA/SKPLE)
    เฉพาะกลุ่มนี้ที่ consumer ต้อง key ต่อพูล (MC_POOL); กลุ่มอื่น key = cat|gauge เหมือนเดิม
    → backward-compatible: cat/gauge ที่ไม่ได้แยกพูลทำงานเหมือนเดิมทุกอย่าง"""
    p = _latest_booking_path()
    if p is None:
        return set()
    try:
        import pandas as pd
        df = pd.read_excel(p, sheet_name="SUMMARY_MC_REMAIN")
    except Exception:
        return set()
    if "MC_POOL" not in df.columns:
        return set()
    pools_by_cg: dict = {}
    for _, r in df.iterrows():
        cat = str(r["MC_CAT"]).strip().upper()
        # จำกัดการแยกพูลไว้ที่ SINGLE-32 (กลุ่ม SKP vs SKPTA/SKPLE) เท่านั้น
        # cat อื่น (เช่น DOUBLE-34 ที่มีพูลข้ามโรงงาน) คงพฤติกรรมเดิม = รวม cat|gauge
        if cat != "SINGLE-32":
            continue
        pool = str(r.get("MC_POOL", "")).strip()
        if not pool:
            continue
        cg = f"{str(r['MC_CAT']).strip()}|{_pool_gkey(r['GUAGE'])}"
        pools_by_cg.setdefault(cg, set()).add(pool)
    return {cg for cg, pools in pools_by_cg.items() if len(pools) > 1}


def pool_map() -> dict:
    """{ "CAT|GUAGE|MC_GROUP": pool_key } เฉพาะ (cat,gauge) ที่แยกพูล
    ให้หน้าเว็บ map เครื่องแต่ละตัว (SKP/SKPTA/SKPLE) → key ของ ava/reserved ที่ถูกพูล
    (เครื่องที่ไม่มี entry = ใช้ key เดิม cat|gauge)"""
    p = _latest_booking_path()
    if p is None:
        return {}
    try:
        import pandas as pd
        df = pd.read_excel(p, sheet_name="MC_POOL_MAP")
    except Exception:
        return {}
    if not {"MC_CAT", "GUAGE", "MC_GROUP", "MC_POOL"} <= set(df.columns):
        return {}
    split = _split_pool_set()
    out: dict = {}
    for _, r in df.iterrows():
        cg = f"{str(r['MC_CAT']).strip()}|{_pool_gkey(r['GUAGE'])}"
        if cg not in split:
            continue
        mg = str(r["MC_GROUP"]).strip().upper()
        pool = str(r["MC_POOL"]).strip()
        if mg and pool:
            out[f"{cg}|{mg}"] = pool
    return out


def _reserved_by_cat() -> dict:
    """เครื่องที่กันไว้ให้ POLY/COTTON ต่อ key พูล จากชีท MC_RESERVED ของ booking_final ล่าสุด
    → { key: {"poly": int, "cotton": int} }  key = MC_POOL ถ้าแยกพูล ไม่งั้น CAT|GUAGE
    (เครื่องกลุ่มนี้ใช้แทนงานปกติไม่ได้)"""
    p = _latest_booking_path()
    if p is None:
        return {}
    try:
        import pandas as pd
        df = pd.read_excel(p, sheet_name="MC_RESERVED")
    except Exception:
        return {}
    if not {"MC_CAT", "GUAGE", "POLY", "COTTON"} <= set(df.columns):
        return {}

    gkey = _pool_gkey
    has_pool = "MC_POOL" in df.columns
    split = _split_pool_set()
    out: dict = {}
    for _, r in df.iterrows():
        cg = f"{str(r['MC_CAT']).strip()}|{gkey(r['GUAGE'])}"
        mc_pool = str(r.get("MC_POOL", "")).strip() if has_pool else ""
        key = mc_pool if (cg in split and mc_pool) else cg
        slot = out.setdefault(key, {"poly": 0, "cotton": 0})
        slot["poly"] += int(_num(r["POLY"]) or 0)
        slot["cotton"] += int(_num(r["COTTON"]) or 0)
    return out


def _reserved_used_by_cat_week() -> dict:
    """เครื่อง POLY/COTTON ที่ booking ใช้ไปแล้ว 'ต่อสัปดาห์' จากชีท MC_RESERVED_WEEKLY
    → { week(str): { "CAT|GUAGE": {"poly_used": int, "cotton_used": int} } }
    ใช้คู่กับ _reserved_by_cat() เพื่อรู้เครื่องกันไว้ที่เหลือจริง = กันไว้ − booking"""
    p = _latest_booking_path()
    if p is None:
        return {}
    try:
        import pandas as pd
        df = pd.read_excel(p, sheet_name="MC_RESERVED_WEEKLY")
    except Exception:
        return {}
    if not {"MC_CAT", "GUAGE", "WEEK", "POLY_USED", "COTTON_USED"} <= set(df.columns):
        return {}

    gkey = _pool_gkey
    has_pool = "MC_POOL" in df.columns
    split = _split_pool_set()
    out: dict = {}
    for _, r in df.iterrows():
        wk = _wk_str(r["WEEK"])
        if wk is None:
            continue
        cg = f"{str(r['MC_CAT']).strip()}|{gkey(r['GUAGE'])}"
        mc_pool = str(r.get("MC_POOL", "")).strip() if has_pool else ""
        key = mc_pool if (cg in split and mc_pool) else cg
        slot = out.setdefault(wk, {}).setdefault(key, {"poly_used": 0, "cotton_used": 0})
        slot["poly_used"] += int(_num(r["POLY_USED"]) or 0)
        slot["cotton_used"] += int(_num(r["COTTON_USED"]) or 0)
    return out


def _coerce(v):
    """แปลงค่าจาก grid (string) → ชนิดที่เหมาะ; เก็บรหัสที่มี leading zero เป็น string"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if s == "":
        return None
    if _INT_RE.match(s):
        try:
            return int(s)
        except ValueError:
            return s
    if _FLOAT_RE.match(s):
        try:
            return float(s)
        except ValueError:
            return s
    return s


def program_map() -> dict:
    """ชีท "Program" ใน MasterMC (user กำหนดเอง) → { ITEM_CODE(upper): [TEAM, ...] }
    หน้าแผนใช้คู่กับคอลัมน์ TEAM ของแถว: ตรงทั้ง item และทีม → ทำตัวหนังสือเป็นสีเหลือง
    อ่านสดทุกครั้ง (ไม่ cache) เพื่อให้แก้ Master แล้วเห็นผลโดยไม่ต้องรันแผนใหม่"""
    p = config.master_files().get("MasterMC")
    if p is None or not p.exists():
        return {}
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
    except Exception:
        return {}
    if "Program" not in wb.sheetnames:
        wb.close()
        return {}
    rows = list(wb["Program"].iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {}
    header = [str(c).strip().upper() if c is not None else "" for c in rows[0]]
    ii = header.index("ITEM_CODE") if "ITEM_CODE" in header else -1
    ti = header.index("TEAM_NAME") if "TEAM_NAME" in header else -1
    if ii < 0 or ti < 0:
        return {}
    out: dict = {}
    for r in rows[1:]:
        item = str(r[ii]).strip().upper() if ii < len(r) and r[ii] is not None else ""
        team = str(r[ti]).strip() if ti < len(r) and r[ti] is not None else ""
        if not item or not team:
            continue
        teams = out.setdefault(item, [])
        if team not in teams:
            teams.append(team)
    return out


def write_grid(sheet: str, columns: list, rows: list) -> dict:
    """เขียนชีทกลับไฟล์แผนล่าสุด (แก้เฉพาะชีทนี้ ชีทอื่นคงเดิม) + สำรอง .bak"""
    p = latest_path()
    if p is None:
        raise FileNotFoundError("ไม่พบไฟล์แผนผลิตให้บันทึก")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    bak = p.with_suffix(p.suffix + f".{ts}.bak")
    shutil.copy2(p, bak)

    wb = load_workbook(p, data_only=False)
    if sheet not in wb.sheetnames:
        wb.close()
        raise KeyError(f"ไม่พบชีท '{sheet}' ในไฟล์แผน")
    ws = wb[sheet]

    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    ws.append([str(c) for c in columns])
    for row in rows:
        ws.append([_coerce(v) for v in row])

    wb.save(p)
    wb.close()
    return {"ok": True, "backup": bak.name, "rows": len(rows), "name": p.name}
