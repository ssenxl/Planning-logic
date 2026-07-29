"""
workday.py — API สำหรับ "วันทำงานตามกลุ่มเครื่อง" (Factory/MC_CAT/Guage)

แหล่งเดียวของความจริง = Calendar.xlsx
  ชีท "Work Day"   : Factory | MC_CAT | Guage | WEEK | WORK_DAY
                     WEEK ว่าง = ค่ามาตรฐานของกลุ่ม · WEEK มีเลข = เฉพาะสัปดาห์นั้น
  ชีท "Week Merge" : WEEK | MERGE_TO  (ยุบสัปดาห์ ใช้ทั้งระบบ)

ครั้งแรกที่ยังไม่มีชีท Work Day → seed จากคอลัมน์ Working Day ของ MasterMC ให้อัตโนมัติ
เพื่อให้แผนไม่เปลี่ยนทันทีตอนย้ายมาใช้ระบบใหม่
"""
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

import config
from config import master_files

WORKDAY_SHEET = "Work Day"
WEEK_MERGE_SHEET = "Week Merge"
WORKDAY_HEADER = ["Factory", "MC_CAT", "Guage", "WEEK", "WORK_DAY", "WORK_HOUR"]
WEEK_MERGE_HEADER = ["WEEK", "MERGE_TO"]
DEFAULT_WORK_DAYS = 6.0     # หากลุ่มไม่เจอ → 6 วัน (ต้องตรงกับ Calendar.DEFAULT_WORK_DAYS)
DEFAULT_WORK_HOURS = 24.0   # ไม่ตั้ง = 24 ชม. (ไม่ลดกำลังผลิต)
SEED_WORK_DAYS = DEFAULT_WORK_DAYS   # ใช้ตอน seed จาก MasterMC (แถวที่ MasterMC ไม่มีค่า)


def _norm(v) -> str:
    s = "" if v is None else str(v).strip()
    if s.lower() in ("nan", "none"):
        return ""
    if re.fullmatch(r"\d+\.0+", s):  # 12.0 → 12 (Excel เก็บ Guage เป็น float)
        return s.split(".", 1)[0]
    return s


def _num(v):
    try:
        return float(str(v).strip())
    except (TypeError, ValueError):
        return None


def _path(key: str) -> Path:
    p = master_files().get(key)
    if p is None or not p.exists():
        raise FileNotFoundError(f"ไม่พบไฟล์ {key}")
    return p


def _rows_of(path: Path, sheet: str) -> list:
    """คืนแถวข้อมูล (ไม่รวม header) ของชีท — ไม่มีชีท = []"""
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return []
    out = [list(r) for r in wb[sheet].iter_rows(values_only=True)][1:]
    wb.close()
    return out


def mc_groups() -> list:
    """กลุ่มเครื่องทั้งหมดจาก MasterMC → [{factory, mc_cat, guage, mc, master_work_day}]
    ใช้เป็นรายการให้ user เลือกใน UI (unique ตาม Factory+MC_CAT+Guage)"""
    path = _path("MasterMC")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Master MC"] if "Master MC" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        return []

    hdr = {str(c).strip().upper(): i for i, c in enumerate(rows[0]) if c is not None}
    fi, ci, gi = hdr.get("FACTORY"), hdr.get("MC_CAT"), hdr.get("GUAGE")
    mi, wi = hdr.get("MC"), hdr.get("WORKING DAY")
    if fi is None or ci is None or gi is None:
        return []

    out, seen = [], set()
    for r in rows[1:]:
        def get(i):
            return r[i] if i is not None and i < len(r) else None
        key = (_norm(get(fi)).upper(), _norm(get(ci)).upper(), _norm(get(gi)).upper())
        if not any(key) or key in seen:
            continue
        seen.add(key)
        out.append({
            "factory": _norm(get(fi)), "mc_cat": _norm(get(ci)), "guage": _norm(get(gi)),
            "mc": _norm(get(mi)),
            "master_work_day": _num(get(wi)),   # ค่าเดิมใน MasterMC — ใช้ seed เท่านั้น
        })
    return out


def mc_group_map() -> dict:
    """ทุกเครื่องใน MasterMC → กลุ่มในแผง : {"MC|GUAGE": "FACTORY|MC_CAT|GUAGE"}

    ต่างจาก mc_groups() ที่ยุบเหลือกลุ่มละ 1 เครื่อง (ไว้ทำ tree) — อันนี้เก็บ "ทุกเครื่อง"
    ให้ Gantt แปลง (เครื่อง, เกจ) → กลุ่มได้ครบ เช่น TSB/26 → PHET|SINGEL-18|26
    ไม่งั้นเครื่องที่ไม่ใช่ตัวแทนกลุ่มจะหาไม่เจอ → ตกไปใช้ fallback 6 ผิด"""
    path = _path("MasterMC")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Master MC"] if "Master MC" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        return {}
    hdr = {str(c).strip().upper(): i for i, c in enumerate(rows[0]) if c is not None}
    fi, ci, gi, mi = hdr.get("FACTORY"), hdr.get("MC_CAT"), hdr.get("GUAGE"), hdr.get("MC")
    if None in (fi, ci, gi, mi):
        return {}
    out = {}
    for r in rows[1:]:
        def g(i):
            return _norm(r[i]).upper() if i < len(r) else ""
        mc, guage = g(mi), g(gi)
        if not mc:
            continue
        out.setdefault(f"{mc}|{guage}", f"{g(fi)}|{g(ci)}|{guage}")
    return out


def calendar_week_days() -> dict:
    """วันที่ปฏิทินเปิดของแต่ละสัปดาห์ — อ่านสดจาก Calendar.xlsx บน server → {week(str): วัน}

    ใช้เป็นเพดานเดียวกับฝั่ง pipeline (WorkDay.get_working_days) เพื่อให้หน้าเว็บกับแผน
    เห็นวันหยุดชุดเดียวกัน — แก้ปฏิทินบน server แล้วหน้าเว็บเห็นทันทีโดยไม่ต้องรันแผนใหม่
    อ่านไม่ได้ → {} (หน้าเว็บจะไม่ cap ด้วยปฏิทิน)"""
    try:
        if str(config.REPO_DIR) not in sys.path:
            sys.path.insert(0, str(config.REPO_DIR))
        from Calendar import load_calendar
        df = load_calendar(_path("Calendar"), sheet_name="Sheet1")
        return {str(int(w)): int(n) for w, n in df.groupby("WEEK")["is_working_day"].sum().items()}
    except Exception:
        return {}


def _aliases(known: set) -> dict:
    """คู่ (เครื่อง|เกจ) → กลุ่มในแผง จาก booking_final ล่าสุด → {"MC|GUAGE": "FAC|CAT|GUAGE"}

    ต้องมีเพราะ booking มีชื่อเครื่องที่ไม่มีใน MasterMC (SKPHF/SKPSL/SKPAO) หรือคู่
    (เครื่อง, เกจ) ไม่ตรง — แต่ CAT+เกจ ตรงกับกลุ่มในแผงพอดี (ตรรกะเดียวกับ
    WorkDay.register_aliases ฝั่ง pipeline) · หาไฟล์ไม่ได้ → {}"""
    if not known:
        return {}
    try:
        import pandas as pd
        import plan_view
        p = plan_view._latest_booking_path()
        if p is None:
            return {}
        df = pd.read_excel(p, sheet_name="DETAIL")
    except Exception:
        return {}

    by_cat: dict = {}
    for k in known:
        parts = k.split("|")
        if len(parts) == 3:
            by_cat.setdefault((parts[1], parts[2]), set()).add(k)

    cols = {c: c for c in df.columns}
    out: dict = {}
    for r in df.to_dict("records"):
        mc = _norm(r.get("MC_GROUP")).upper()
        g = _norm(r.get("GUAGE")).upper()
        if not mc or f"{mc}|{g}" in out:
            continue
        cat = _norm(r.get("MC_CAT")).upper() or _norm(r.get("CAT")).upper()
        fac = _norm(r.get("FACTORY")).upper()
        key = None
        if fac and f"{fac}|{cat}|{g}" in known:
            key = f"{fac}|{cat}|{g}"
        else:
            cands = by_cat.get((cat, g), set())
            if len(cands) == 1:
                key = next(iter(cands))
        if key:
            out[f"{mc}|{g}"] = key
    return out


def get_workday() -> dict:
    """คืนค่าที่ตั้งไว้ + รายการกลุ่ม + การยุบสัปดาห์
    {groups: [...], defaults: {"FAC|CAT|G": วัน}, weeks: {"FAC|CAT|G|W": วัน},
     hours: {"FAC|CAT|G": ชั่วโมง}, merges: {week: to}}"""
    cal = _path("Calendar")
    groups = mc_groups()

    defaults, weeks, hours = {}, {}, {}
    for r in _rows_of(cal, WORKDAY_SHEET):
        r = list(r) + [None] * (6 - len(r))
        key = "|".join(_norm(x).upper() for x in r[:3])
        if key == "||":
            continue
        wk, days, hr = _num(r[3]), _num(r[4]), _num(r[5])
        if wk is None:
            # ค่ามาตรฐานของกลุ่ม: วัน + ชั่วโมง (ชั่วโมงเป็น override ต่อกลุ่มอย่างเดียว)
            if days is not None:
                defaults[key] = days
            if hr is not None:
                hours[key] = hr
        elif days is not None:
            weeks[f"{key}|{int(wk)}"] = days

    merges = {}
    for r in _rows_of(cal, WEEK_MERGE_SHEET):
        r = list(r) + [None] * (2 - len(r))
        src, dst = _num(r[0]), _num(r[1])
        if src is None or dst is None or int(src) == int(dst):
            continue
        merges[int(src)] = int(dst)

    known = set(defaults) | set(hours) | {"|".join(k.split("|")[:3]) for k in weeks}
    return {"groups": groups, "defaults": defaults, "weeks": weeks, "hours": hours,
            "merges": merges, "fallback_days": DEFAULT_WORK_DAYS,
            "fallback_hours": DEFAULT_WORK_HOURS,
            # วันที่ปฏิทินเปิดต่อสัปดาห์ (สดจาก server) — เพดานของวันทำงาน
            "cal_days": calendar_week_days(),
            # ทุกเครื่องใน MasterMC → กลุ่มในแผง (ให้ Gantt แปลงเครื่องไม่ใช่ตัวแทนได้ถูก)
            "mc_map": mc_group_map(),
            # เครื่องที่ไม่มีใน MasterMC → กลุ่มในแผง (ให้ Gantt คำนวณตรงกับ pipeline)
            "aliases": _aliases(known)}


def _write_sheet(wb, sheet: str, header: list, rows: list) -> None:
    ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(sheet)
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    ws.append(header)
    for r in rows:
        ws.append(r)


def save_workday(defaults: dict, weeks: dict, hours: dict = None, merges: dict = None) -> dict:
    """เขียนชีท Work Day + Week Merge กลับ Calendar.xlsx (สำรอง .bak ก่อน)
    defaults: {"FAC|CAT|G": วัน} · weeks: {"FAC|CAT|G|W": วัน}
    hours: {"FAC|CAT|G": ชั่วโมง} (ค่ามาตรฐานต่อกลุ่ม) · merges: {"31": 32}"""
    cal = _path("Calendar")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    bak = cal.with_suffix(cal.suffix + f".{ts}.bak")
    shutil.copy2(cal, bak)

    defaults, weeks, hours, merges = defaults or {}, weeks or {}, hours or {}, merges or {}

    wd_rows = []
    # แถวค่ามาตรฐานของกลุ่ม (WEEK ว่าง) = union ของกลุ่มที่ตั้งวัน และ/หรือ ชั่วโมง
    for key in sorted(set(defaults) | set(hours)):
        fac, cat, g = (key.split("|") + ["", "", ""])[:3]
        d, h = _num(defaults.get(key)), _num(hours.get(key))
        if d is not None or h is not None:
            wd_rows.append([fac, cat, g, None, d, h])
    for key, days in sorted(weeks.items()):
        parts = key.split("|")
        if len(parts) != 4:
            continue
        d, wk = _num(days), _num(parts[3])
        if d is not None and wk is not None:
            wd_rows.append([parts[0], parts[1], parts[2], int(wk), d, None])

    mg_rows = []
    for src, dst in sorted((merges or {}).items(), key=lambda kv: int(kv[0])):
        s, t = _num(src), _num(dst)
        if s is not None and t is not None and int(s) != int(t):
            mg_rows.append([int(s), int(t)])

    wb = load_workbook(cal, data_only=False)
    _write_sheet(wb, WORKDAY_SHEET, WORKDAY_HEADER, wd_rows)
    _write_sheet(wb, WEEK_MERGE_SHEET, WEEK_MERGE_HEADER, mg_rows)
    wb.save(cal)
    wb.close()
    return {"ok": True, "backup": bak.name, "rows": len(wd_rows), "merges": len(mg_rows)}


def seed_from_mastermc() -> dict:
    """สร้างชีท Work Day ครั้งแรกจากคอลัมน์ Working Day ของ MasterMC (ไม่ทับของเดิมที่ตั้งไว้)"""
    cur = get_workday()
    defaults = dict(cur["defaults"])
    added = 0
    for g in cur["groups"]:
        key = "|".join(x.upper() for x in (g["factory"], g["mc_cat"], g["guage"]))
        if key in defaults:
            continue
        defaults[key] = g["master_work_day"] if g["master_work_day"] is not None else SEED_WORK_DAYS
        added += 1
    if not added:
        return {"ok": True, "added": 0}
    r = save_workday(defaults, cur["weeks"], cur.get("hours"), cur["merges"])
    r["added"] = added
    return r
