"""
substitute_view.py — "Item ทดแทน" จาก Master_Item ชีท "Master_Item V2"

รหัสเต็มที่อยู่ใน ITEM_LIST เดียวกัน = spec เครื่อง/ด้ายเหมือนกัน → ใช้แทนกันได้
  - ดู/ค้นหากลุ่มทดแทน พร้อม stock คงเหลือของแต่ละรหัส (QA ว่างเท่านั้น)
  - import: อัปโหลด .xlsx มาทับ "เฉพาะชีท Master_Item V2" (ชีทอื่นคงเดิม + สำรอง .bak)
"""
import io
import re
import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

import config

SHEET_V2 = "Master_Item V2"
COL_ITEM_LIST = "ITEM_LIST"
COL_SPEC_KEY = "SPEC_KEY"
COL_SUFFIX = "ITEM_SUFFIX"
SPEC_COLS = ["MC_GROUP", "KNIT_MC_CAT", "MC_GAUGE", "YARN_ITEM", "MC_NEEDLE", "YARN_SL"]
# คอลัมน์ที่ไฟล์ import ต้องมีอย่างน้อย (ไม่งั้นถือว่าเลือกไฟล์ผิด)
REQUIRED_COLS = [COL_ITEM_LIST, COL_SPEC_KEY]

DEFAULT_LIMIT = 200
MAX_LIMIT = 1000

_STOCK_FILE = config.REPO_DIR / "Stock" / "view_stock.xlsx"

_index_cache = None      # (mtime, index)
_stock_cache = None      # (mtime, {code: kg})


class SheetMissing(KeyError):
    """ไฟล์ Master_Item ยังไม่มีชีท V2 — สืบทอด KeyError ให้ server คืน 404 ตามเดิม
    แต่ __str__ ไม่ครอบ quote เหมือน KeyError มาตรฐาน (ข้อความไปโชว์บนหน้าเว็บตรง ๆ)"""

    def __str__(self) -> str:
        return str(self.args[0]) if self.args else ""


def _master_path() -> Path:
    p = config.master_files().get("Master_Item")
    if not p:
        raise FileNotFoundError(
            "ยังไม่ได้ตั้ง path ของ Master_Item ใน config.ini ([paths] master_item)"
        )
    return Path(p)


def _fmt(v) -> str:
    """ค่าจาก Excel -> ข้อความ (เลขจำนวนเต็มไม่ให้ติด .0)"""
    if v is None or (not isinstance(v, str) and pd.isna(v)):
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _split_codes(cell) -> list:
    if pd.isna(cell):
        return []
    return [c.strip() for c in str(cell).split(",") if c.strip() and c.strip().lower() != "nan"]


def _derive_item(code: str, suffix: str = "") -> str:
    """รหัสเต็ม -> ITEM Color (ตัด suffix A0/B0 ท้ายออก)"""
    c = str(code).strip()
    suf = str(suffix).strip().upper()
    if suf and suf != "NAN" and c.upper().endswith(suf):
        return c[: -len(suf)]
    return re.sub(r"[A-Z]\d$", "", c)


# ---------------- stock ----------------
def _load_stock() -> dict:
    """รหัสเต็ม -> BALANCE_KG รวม (เฉพาะแถวที่ QA_REASON ว่าง); ไม่มีไฟล์ = {}"""
    global _stock_cache
    if not _STOCK_FILE.exists():
        return {}
    mtime = _STOCK_FILE.stat().st_mtime
    if _stock_cache and _stock_cache[0] == mtime:
        return _stock_cache[1]

    try:
        st = pd.read_excel(_STOCK_FILE, usecols=["ITEM_CODE", "BALANCE_KG", "QA_REASON"])
    except Exception:
        return {}
    st["ITEM_CODE"] = st["ITEM_CODE"].astype(str).str.strip()
    st["BALANCE_KG"] = pd.to_numeric(st["BALANCE_KG"], errors="coerce").fillna(0)
    qa = st["QA_REASON"].astype(str).str.strip()
    blank = st["QA_REASON"].isna() | qa.eq("") | qa.str.lower().eq("nan")
    bal = st[blank].groupby("ITEM_CODE")["BALANCE_KG"].sum().to_dict()
    _stock_cache = (mtime, bal)
    return bal


# ---------------- index ----------------
def _build_index(path: Path) -> dict:
    """อ่านชีท V2 -> {"groups": [...], "codes": {code: gi}, "items": {item: [gi]}}"""
    xls = pd.ExcelFile(path)
    if SHEET_V2 not in xls.sheet_names:
        # ไฟล์เก่ายังไม่มีชีทนี้ — บอกทางแก้ให้ user แทนที่จะเป็น error ดิบ ๆ
        raise SheetMissing(
            f"ไฟล์ {path.name} ยังไม่มีชีท '{SHEET_V2}' "
            f"(มีชีท: {', '.join(xls.sheet_names)}) — ให้ import ไฟล์เข้ามาก่อน"
        )
    df = xls.parse(SHEET_V2)
    missing = [c for c in REQUIRED_COLS if c not in df.columns]
    if missing:
        raise SheetMissing(f"ชีท '{SHEET_V2}' ไม่มีคอลัมน์ {', '.join(missing)}")

    groups, by_code, by_item = [], {}, {}
    for rec in df.to_dict("records"):
        codes = _split_codes(rec.get(COL_ITEM_LIST))
        if not codes:
            continue
        suffix = _fmt(rec.get(COL_SUFFIX))
        items = []
        for c in codes:
            it = _derive_item(c, suffix)
            if it not in items:
                items.append(it)
        gi = len(groups)
        groups.append({
            "spec_key": _fmt(rec.get(COL_SPEC_KEY)),
            "suffix": suffix,
            "codes": codes,
            "items": items,
            "count": len(codes),
            **{c.lower(): _fmt(rec.get(c)) for c in SPEC_COLS},
        })
        for c in codes:
            by_code[c.upper()] = gi
        for it in items:
            by_item.setdefault(it.upper(), []).append(gi)

    return {"groups": groups, "codes": by_code, "items": by_item}


def load_index(force: bool = False) -> dict:
    """index กลุ่มทดแทน (cache ตาม mtime ของไฟล์ Master_Item)"""
    global _index_cache
    path = _master_path()
    if not path.exists():
        raise FileNotFoundError(f"ไม่พบไฟล์ {path}")
    mtime = path.stat().st_mtime
    if _index_cache and _index_cache[0] == mtime and not force:
        return _index_cache[1]
    idx = _build_index(path)
    _index_cache = (mtime, idx)
    return idx


def _group_payload(g: dict, stock: dict) -> dict:
    """กลุ่ม + stock ของแต่ละรหัส (เรียงรหัสที่มีของก่อน)"""
    codes = [
        {"code": c, "item": _derive_item(c, g["suffix"]), "stock": round(float(stock.get(c, 0.0)), 2)}
        for c in g["codes"]
    ]
    codes.sort(key=lambda x: (-x["stock"], x["code"]))
    return {
        **{k: g[k] for k in ("spec_key", "suffix", "count")},
        **{c.lower(): g[c.lower()] for c in SPEC_COLS},
        "codes": codes,
        "stock_total": round(sum(c["stock"] for c in codes), 2),
    }


def summary() -> dict:
    """สถานะไฟล์ + สถิติกลุ่มทดแทน (ใช้โชว์หัวหน้าเว็บ)"""
    try:
        path = _master_path()
    except FileNotFoundError as e:
        return {"exists": False, "error": str(e)}
    if not path.exists():
        return {"exists": False, "path": str(path), "error": f"ไม่พบไฟล์ {path.name}"}

    out = {
        "exists": True,
        "path": str(path),
        "name": path.name,
        "sheet": SHEET_V2,
        "mtime": path.stat().st_mtime,
        "size": path.stat().st_size,
        "has_stock": bool(_load_stock()),
    }
    try:
        idx = load_index()
    except Exception as e:
        out.update({"ok": False, "error": str(e)})
        return out

    groups = idx["groups"]
    multi = [g for g in groups if g["count"] > 1]
    out.update({
        "ok": True,
        "groups": len(groups),
        "codes": len(idx["codes"]),
        "items": len(idx["items"]),
        "multi_groups": len(multi),
        "multi_codes": sum(g["count"] for g in multi),
        "max_group": max((g["count"] for g in groups), default=0),
    })
    return out


def search(q: str = "", only_multi: bool = True, limit: int = DEFAULT_LIMIT) -> dict:
    """ค้นหากลุ่มทดแทนด้วยรหัสเต็ม / ITEM Color / SPEC_KEY / กลุ่มเครื่อง

    q ว่าง = คืนกลุ่มที่ใหญ่ที่สุดก่อน (ไว้ browse ดูเฉย ๆ)
    """
    idx = load_index()
    stock = _load_stock()
    groups = idx["groups"]
    limit = max(1, min(int(limit or DEFAULT_LIMIT), MAX_LIMIT))

    key = str(q or "").strip().upper()
    if key:
        hits = []
        gi = idx["codes"].get(key)
        if gi is not None:
            hits.append(gi)
        for g in idx["items"].get(key, []):
            if g not in hits:
                hits.append(g)
        if not hits:      # ไม่ตรงเป๊ะ -> ค้นแบบ substring
            for i, g in enumerate(groups):
                if any(key in c.upper() for c in g["codes"]) \
                        or key in g["spec_key"].upper() \
                        or key in str(g.get("mc_group", "")).upper():
                    hits.append(i)
        cand = [groups[i] for i in hits]
    else:
        cand = sorted(groups, key=lambda g: -g["count"])

    if only_multi:
        cand = [g for g in cand if g["count"] > 1]

    total = len(cand)
    rows = [_group_payload(g, stock) for g in cand[:limit]]
    return {"total": total, "shown": len(rows), "truncated": total > len(rows), "rows": rows}


# ---------------- import ----------------
def preview_import(data: bytes) -> dict:
    """อ่านไฟล์ที่อัปโหลด -> สรุปว่าจะได้อะไร (ยังไม่เขียนทับ)"""
    df, sheet = _read_upload(data)
    codes = df[COL_ITEM_LIST].apply(_split_codes)
    n_codes = sum(len(c) for c in codes)
    return {
        "sheet": sheet,
        "rows": len(df),
        "columns": [str(c) for c in df.columns],
        "codes": n_codes,
        "multi_groups": int((codes.apply(len) > 1).sum()),
    }


def _read_upload(data: bytes) -> tuple:
    """อ่าน .xlsx ที่อัปโหลด -> (DataFrame, ชื่อชีทที่ใช้) + ตรวจคอลัมน์ที่จำเป็น"""
    try:
        xls = pd.ExcelFile(io.BytesIO(data))
    except Exception as e:
        raise ValueError(f"อ่านไฟล์ Excel ไม่ได้: {e}")
    sheet = SHEET_V2 if SHEET_V2 in xls.sheet_names else xls.sheet_names[0]
    df = xls.parse(sheet)
    missing = [c for c in REQUIRED_COLS if c not in df.columns]
    if missing:
        raise ValueError(
            f"ชีท '{sheet}' ไม่มีคอลัมน์ที่จำเป็น: {', '.join(missing)} "
            f"(ต้องเป็นไฟล์รูปแบบเดียวกับชีท '{SHEET_V2}')"
        )
    if df.empty:
        raise ValueError("ไฟล์ที่อัปโหลดไม่มีข้อมูล")
    return df, sheet


def import_v2(data: bytes, filename: str = "") -> dict:
    """เขียนทับเฉพาะชีท 'Master_Item V2' ใน Master_Item.xlsx (ชีทอื่นคงเดิม)

    สำรองไฟล์เดิมเป็น .<timestamp>.bak ก่อนเสมอ
    """
    df, src_sheet = _read_upload(data)
    path = _master_path()
    if not path.exists():
        raise FileNotFoundError(f"ไม่พบไฟล์ {path}")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    bak = path.with_suffix(path.suffix + f".{ts}.bak")
    shutil.copy2(path, bak)

    wb = load_workbook(path)
    pos = wb.sheetnames.index(SHEET_V2) if SHEET_V2 in wb.sheetnames else len(wb.sheetnames)
    if SHEET_V2 in wb.sheetnames:
        wb.remove(wb[SHEET_V2])
    ws = wb.create_sheet(SHEET_V2, pos)

    ws.append([str(c) for c in df.columns])
    for row in df.itertuples(index=False, name=None):
        ws.append([None if (not isinstance(v, str) and pd.isna(v)) else v for v in row])

    wb.save(path)
    wb.close()

    global _index_cache
    _index_cache = None      # บังคับให้อ่าน index ใหม่รอบหน้า
    idx = load_index(force=True)
    multi = sum(1 for g in idx["groups"] if g["count"] > 1)
    return {
        "ok": True,
        "file": filename or "(ไม่ระบุชื่อไฟล์)",
        "source_sheet": src_sheet,
        "rows": len(df),
        "codes": len(idx["codes"]),
        "multi_groups": multi,
        "backup": bak.name,
    }
