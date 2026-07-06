"""
order_color_view.py — อ่าน/เขียนไฟล์ "Order Color" (datamining_booking_mapped.xlsx)
สำหรับหน้า "ดึง Order Color"
- ไฟล์นี้เป็นผลลัพธ์ Datamining → Booking (ระดับ ITEM Color) สร้างจาก MapItem.py
- อ่านชีท → คืน columns + rows ครบทุกแถว (เพื่อบันทึกกลับได้)
- เขียนกลับ: สำรองไฟล์เดิมเป็น .bak ก่อน แล้วแก้เฉพาะชีทเป้าหมาย (ชีทอื่นคงเดิม)
โครงเดียวกับ plan_view.py
"""
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

import config

# ชื่อไฟล์ผลลัพธ์ Datamining → Booking (คงที่ ไม่มี timestamp)
FILE_NAME = "datamining_booking_mapped.xlsx"

_INT_RE = re.compile(r"^-?(0|[1-9][0-9]*)$")
_FLOAT_RE = re.compile(r"^-?(0|[1-9][0-9]*|0?)\.[0-9]+$")


def latest_path() -> Path | None:
    """path ของไฟล์ Order Color (None = ยังไม่มี — ต้องกดดึงข้อมูลก่อน)"""
    p = config.OUTPUT_DIR / FILE_NAME
    return p if p.exists() and p.is_file() else None


def latest_meta() -> dict:
    """ข้อมูลไฟล์ + รายชื่อชีท (ให้ frontend โชว์ header)"""
    p = latest_path()
    if p is None:
        return {"exists": False, "name": FILE_NAME, "size": 0, "mtime": None, "sheets": []}
    st = p.stat()
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
        sheets = list(wb.sheetnames)
        wb.close()
    except Exception as e:
        return {"exists": True, "name": p.name, "size": st.st_size,
                "mtime": st.st_mtime, "sheets": [], "error": str(e)}
    return {"exists": True, "name": p.name, "size": st.st_size,
            "mtime": st.st_mtime, "sheets": sheets}


def read_grid(sheet: str = None) -> dict:
    """อ่านชีทของไฟล์ Order Color → columns + rows (แถวแรก = header)"""
    p = latest_path()
    if p is None:
        raise FileNotFoundError("ยังไม่มีไฟล์ Order Color — กรุณากดปุ่มดึงข้อมูลก่อน")
    wb = load_workbook(p, read_only=True, data_only=True)
    sheets = wb.sheetnames
    mtime = p.stat().st_mtime
    if not sheets:
        wb.close()
        return {"name": p.name, "mtime": mtime, "sheet": None,
                "sheets": [], "columns": [], "rows": []}
    if sheet is None or sheet not in sheets:
        sheet = sheets[0]
    ws = wb[sheet]
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append(["" if v is None else v for v in r])
    wb.close()

    if not rows:
        return {"name": p.name, "mtime": mtime, "sheet": sheet,
                "sheets": sheets, "columns": [], "rows": []}

    header = [str(c) if c != "" else f"Col{i+1}" for i, c in enumerate(rows[0])]
    width = len(header)
    data = [list(row)[:width] + [""] * (width - len(row)) for row in rows[1:]]
    return {"name": p.name, "mtime": mtime, "sheet": sheet,
            "sheets": sheets, "columns": header, "rows": data}


def _coerce(v):
    """แปลงค่าจาก grid (string) → ชนิดที่เหมาะ; เก็บรหัสที่มี leading zero เป็น string"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if s == "":
        return None
    if _INT_RE.match(s):
        try:
            return int(s)
        except ValueError:
            return s
    if _FLOAT_RE.match(s):
        try:
            return float(s)
        except ValueError:
            return s
    return s


def write_grid(sheet: str, columns: list, rows: list) -> dict:
    """เขียนชีทกลับไฟล์ Order Color (แก้เฉพาะชีทนี้ ชีทอื่นคงเดิม) + สำรอง .bak"""
    p = latest_path()
    if p is None:
        raise FileNotFoundError("ไม่พบไฟล์ Order Color ให้บันทึก")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    bak = p.with_suffix(p.suffix + f".{ts}.bak")
    shutil.copy2(p, bak)

    wb = load_workbook(p, data_only=False)
    if sheet not in wb.sheetnames:
        wb.close()
        raise KeyError(f"ไม่พบชีท '{sheet}' ในไฟล์ Order Color")
    ws = wb[sheet]

    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    ws.append([str(c) for c in columns])
    for row in rows:
        ws.append([_coerce(v) for v in row])

    wb.save(p)
    wb.close()
    return {"ok": True, "backup": bak.name, "rows": len(rows), "name": p.name}
