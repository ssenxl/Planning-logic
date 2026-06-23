"""
masters.py — อ่าน/เขียนไฟล์ Master (.xlsx) รายชีท สำหรับ grid editor
- อ่าน: คืน columns + rows เป็น JSON
- เขียน: สำรองไฟล์เดิมเป็น .bak ก่อน, แก้เฉพาะชีทเป้าหมาย (ชีทอื่นคงเดิม)
"""
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from config import master_files

_INT_RE = re.compile(r"^-?(0|[1-9][0-9]*)$")
_FLOAT_RE = re.compile(r"^-?(0|[1-9][0-9]*|0?)\.[0-9]+$")


def list_masters() -> list:
    """คืนรายการไฟล์ Master + ชีท + สถานะมีไฟล์จริงไหม"""
    out = []
    for name, path in master_files().items():
        info = {"name": name, "path": str(path), "exists": path.exists(), "sheets": []}
        if path.exists():
            try:
                wb = load_workbook(path, read_only=True, data_only=False)
                info["sheets"] = wb.sheetnames
                wb.close()
            except Exception as e:
                info["error"] = str(e)
        out.append(info)
    return out


def _resolve(name: str) -> Path:
    reg = master_files()
    if name not in reg:
        raise KeyError(f"ไม่รู้จัก Master '{name}'")
    p = reg[name]
    if not p.exists():
        raise FileNotFoundError(f"ไม่พบไฟล์ {p}")
    return p


def read_sheet(name: str, sheet: str) -> dict:
    """อ่านชีท → {columns:[...], rows:[[...]]}  (แถวแรก = header)
    data_only=True: คืน "ค่าที่คำนวณแล้ว" ของเซลล์สูตร (ไม่ใช่ข้อความสูตร =...)
    เพื่อให้ grid โชว์ตัวเลขจริง และบน server ไม่พึ่ง Excel"""
    path = _resolve(name)
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise KeyError(f"ไม่พบชีท '{sheet}' ใน {name}")
    ws = wb[sheet]
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append(["" if v is None else v for v in r])
    wb.close()

    if not rows:
        return {"columns": [], "rows": []}

    header = [str(c) if c != "" else f"Col{i+1}" for i, c in enumerate(rows[0])]
    data = rows[1:]
    # pad ให้ทุกแถวยาวเท่า header
    width = len(header)
    norm = [list(row) + [""] * (width - len(row)) for row in data]
    norm = [row[:width] for row in norm]
    return {"columns": header, "rows": norm}


def _coerce(v):
    """แปลงค่าจาก grid (string) → ชนิดที่เหมาะ; เก็บรหัสที่มี leading zero เป็น string"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if s == "":
        return None
    if _INT_RE.match(s):
        try:
            return int(s)
        except ValueError:
            return s
    if _FLOAT_RE.match(s):
        try:
            return float(s)
        except ValueError:
            return s
    return s


def _norm_hdr(s) -> str:
    return re.sub(r"\s+", " ", str(s).strip()).upper()


def _to_num(v):
    """แปลงเป็น float ถ้าได้ ไม่งั้นคืน None"""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip())
    except ValueError:
        return None


def _clean_num(x):
    """ทำให้เลขจำนวนเต็มเป็น int (กัน 2500.0)"""
    if isinstance(x, float) and x.is_integer():
        return int(x)
    return x


def _recompute_target_stock(columns: list, rows: list) -> list:
    """คำนวณคอลัมน์สูตรของ Target_Stock ใหม่จากคอลัมน์ดิบ (จับด้วย "ชื่อหัวคอลัมน์")
    เพื่อให้ไฟล์เก็บ "ค่าตัวเลขจริง" ไม่พึ่ง Excel cache → pandas บน server อ่านได้ถูกต้อง
      Match core   = Item code & Team Name
      TARGET SCM   = TARGET/MONTH / 4 * 2
      STOCK MIN    = TARGET/MONTH / 4
      STOCK MAX    = IF Team=RTS: (Type1=CG ? TARGET/MONTH : TARGET SCM*1.5) else TARGET SCM*1.5
      Stock 5 Week = STOCK MIN * 5
    คำนวณเฉพาะคอลัมน์ที่มีอยู่ (รองรับการเพิ่ม/ลบ/เปลี่ยนชื่อคอลัมน์)"""
    idx = {_norm_hdr(c): i for i, c in enumerate(columns)}
    item_i = idx.get("ITEM CODE")
    team_i = idx.get("TEAM NAME")
    type1_i = idx.get("TYPE 1")
    target_i = idx.get("TARGET/MONTH")
    match_i = idx.get("MATCH CORE")
    scm_i = idx.get("TARGET SCM")
    min_i = idx.get("STOCK MIN")
    max_i = idx.get("STOCK MAX")
    week_i = idx.get("STOCK 5 WEEK")

    def get(r, i):
        return r[i] if (i is not None and i < len(r)) else None

    out = []
    for row in rows:
        r = list(row)
        target = _to_num(get(r, target_i))

        # Match core = Item code & Team Name (ต่อข้อความ)
        if match_i is not None and (item_i is not None or team_i is not None):
            a = "" if get(r, item_i) is None else str(get(r, item_i))
            b = "" if get(r, team_i) is None else str(get(r, team_i))
            r[match_i] = a + b

        # TARGET SCM = TARGET/MONTH / 4 * 2
        scm = None
        if scm_i is not None:
            if target is not None:
                scm = target / 4 * 2
                r[scm_i] = _clean_num(scm)
            else:
                scm = _to_num(get(r, scm_i))
        else:
            scm = (target / 4 * 2) if target is not None else None

        # STOCK MIN = TARGET/MONTH / 4
        if min_i is not None and target is not None:
            r[min_i] = _clean_num(target / 4)

        # STOCK MAX
        if max_i is not None:
            team = str(get(r, team_i)).strip().upper() if get(r, team_i) is not None else ""
            t1 = str(get(r, type1_i)).strip().upper() if get(r, type1_i) is not None else ""
            if team == "RTS" and t1 == "CG":
                mx = target
            else:
                mx = (scm * 1.5) if scm is not None else None
            if mx is not None:
                r[max_i] = _clean_num(mx)

        # Stock 5 Week = STOCK MIN * 5
        if week_i is not None:
            mn = (target / 4) if target is not None else _to_num(get(r, min_i))
            if mn is not None:
                r[week_i] = _clean_num(mn * 5)

        out.append(r)
    return out


def write_sheet(name: str, sheet: str, columns: list, rows: list) -> dict:
    """เขียนชีทกลับ (แก้เฉพาะชีทนี้ ชีทอื่นคงเดิม) + สำรอง .bak"""
    path = _resolve(name)

    # Target_Stock: คำนวณคอลัมน์สูตรใหม่ก่อนเขียน (เก็บค่าจริง ใช้บน server ได้)
    if name == "Target_Stock":
        rows = _recompute_target_stock(columns, rows)

    # สำรองไฟล์เดิม
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    bak = path.with_suffix(path.suffix + f".{ts}.bak")
    shutil.copy2(path, bak)

    wb = load_workbook(path, data_only=False)
    if sheet not in wb.sheetnames:
        wb.close()
        raise KeyError(f"ไม่พบชีท '{sheet}' ใน {name}")
    ws = wb[sheet]

    # ล้างค่าทุกเซลล์เดิมในชีท แล้วเขียนใหม่ (header + rows)
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)

    ws.append([str(c) for c in columns])
    for row in rows:
        ws.append([_coerce(v) for v in row])

    wb.save(path)
    wb.close()
    return {"ok": True, "backup": bak.name, "rows": len(rows)}
