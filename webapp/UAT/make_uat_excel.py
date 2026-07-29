"""
make_uat_excel.py — สร้างเอกสาร UAT (User Acceptance Test) ของ Knit Plan Web Server

ครอบคลุมทุกความสามารถของ server (webapp/backend): 52 API endpoints ใน 12 กลุ่มงาน + หน้าเว็บ/static
output: UAT_KnitPlan_Server.xlsx (ชีท ภาพรวม / UAT_TestCases / รายการ_API / Defect_Log / คู่มือใช้งาน)

รัน:  python webapp/UAT/make_uat_excel.py
"""
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

OUT = Path(__file__).resolve().parent / "UAT_KnitPlan_Server.xlsx"

# ---------------- สไตล์ ----------------
C_HEAD = "1F4E79"        # หัวตาราง น้ำเงินเข้ม
C_SUB = "DDEBF7"         # แถบหัวข้อย่อย
C_MOD = "FFF2CC"         # แถบคั่นโมดูล
FONT_HEAD = Font(name="Tahoma", size=10, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Tahoma", size=10)
FONT_BOLD = Font(name="Tahoma", size=10, bold=True)
FONT_TITLE = Font(name="Tahoma", size=16, bold=True, color="1F4E79")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header(ws, row: int, ncol: int):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_HEAD
        cell.fill = PatternFill("solid", fgColor=C_HEAD)
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[row].height = 34


def set_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# =====================================================================
# 1) รายการ API ทั้งหมดของ server
#    (module, method, endpoint, สิทธิ์, คำอธิบาย, พารามิเตอร์/body)
# =====================================================================
APIS = [
    ("เข้าสู่ระบบ (Auth)", "POST", "/api/login", "เปิด (ไม่ต้อง login)",
     "ตรวจรหัสผ่าน (PBKDF2-SHA256) แล้วออก token (HMAC, อายุ 12 ชม.)", "body: {username, password}"),
    ("เข้าสู่ระบบ (Auth)", "GET", "/api/me", "ทุก user ที่ login",
     "คืน username + role จาก token (ใช้เช็คสิทธิ์หลังรีเฟรชหน้า)", "header: Authorization: Bearer <token>"),
    ("เข้าสู่ระบบ (Auth)", "GET", "/api/health", "เปิด (ไม่ต้อง login)",
     "health check ของ server", "-"),

    ("สั่งรัน Pipeline (Run)", "POST", "/api/run", "ทุก user ที่ login",
     "สั่งรัน pipeline เป็น subprocess (ล็อกกันรันซ้อน 1 งานต่อครั้ง)",
     "body: {mode} = full|db|plan|map-item|stock|booking|sc|datamining"),
    ("สั่งรัน Pipeline (Run)", "POST", "/api/run/stop", "ทุก user ที่ login",
     "หยุดงานที่รันอยู่ — kill ทั้ง process tree (taskkill /T บน Windows)", "-"),
    ("สั่งรัน Pipeline (Run)", "GET", "/api/run/status", "ทุก user ที่ login",
     "สถานะรอบรันปัจจุบัน: running, mode, label, step_num/total_steps, progress, returncode", "-"),
    ("สั่งรัน Pipeline (Run)", "GET", "/api/run/logs", "ทุก user ที่ login",
     "log สดแบบ incremental (คืน lines + next_offset; reset=true เมื่อรอบใหม่เริ่ม)", "query: offset (int)"),

    ("ตั้งเวลาอัตโนมัติ (Schedule)", "GET", "/api/schedule", "ทุก user ที่ login",
     "เวลาที่ตั้งไว้ + next_run ของ job (APScheduler, timezone Asia/Bangkok)", "-"),
    ("ตั้งเวลาอัตโนมัติ (Schedule)", "PUT", "/api/schedule", "admin เท่านั้น",
     "บันทึกเวลาลง settings.json แล้วตั้ง job ใหม่", "body: {schedule: {full: {enabled, hour, minute}}}"),

    ("Master Data", "GET", "/api/masters", "ทุก user ที่ login",
     "รายการไฟล์ Master (MasterMC / Calendar / Target_Stock / Master_Item) + ชีท + สถานะไฟล์", "-"),
    ("Master Data", "GET", "/api/masters/{name}/{sheet}", "ทุก user ที่ login",
     "อ่านชีท Master → columns + rows (เซลล์สูตรคืนค่าที่คำนวณแล้ว)", "path: name, sheet"),
    ("Master Data", "PUT", "/api/masters/{name}/{sheet}", "ทุก user ที่ login",
     "บันทึกชีทกลับไฟล์ (สำรอง .bak ก่อน, ชีทอื่นคงเดิม; Target_Stock คำนวณคอลัมน์สูตรใหม่)",
     "path: name, sheet · body: {columns, rows}"),

    ("วันทำงาน (Work Day)", "GET", "/api/workday", "ทุก user ที่ login",
     "กลุ่มเครื่อง + วัน/ชั่วโมงมาตรฐาน + ค่าเฉพาะสัปดาห์ + การยุบสัปดาห์ + วันปฏิทิน + mc_map/aliases", "-"),
    ("วันทำงาน (Work Day)", "PUT", "/api/workday", "ทุก user ที่ login",
     "บันทึกชีท Work Day + Week Merge ลง Calendar.xlsx (สำรอง .bak ก่อน)",
     "body: {defaults, weeks, hours, merges}"),
    ("วันทำงาน (Work Day)", "POST", "/api/workday/seed", "ทุก user ที่ login",
     "สร้างชีท Work Day ครั้งแรกจากคอลัมน์ Working Day ของ MasterMC (ไม่ทับค่าที่ตั้งไว้แล้ว)", "-"),

    ("ไฟล์ผลลัพธ์ (History)", "GET", "/api/outputs", "ทุก user ที่ login",
     "รายการไฟล์แผนผลิต production_plan_*.xlsx (ใหม่→เก่า)", "-"),
    ("ไฟล์ผลลัพธ์ (History)", "GET", "/api/outputs/booking", "ทุก user ที่ login",
     "รายการไฟล์ booking_final_ready_*.xlsx", "-"),
    ("ไฟล์ผลลัพธ์ (History)", "GET", "/api/outputs/sc", "ทุก user ที่ login",
     "รายการไฟล์ view_sc_*.xlsx (ข้อมูล SC ต่อรอบรัน)", "-"),
    ("ไฟล์ผลลัพธ์ (History)", "GET", "/api/outputs/{fname}", "เปิด (ลิงก์ดาวน์โหลดตรง)",
     "ดาวน์โหลดไฟล์ใน data_plan (ส่ง header no-store กันแคช)", "path: fname"),
    ("ไฟล์ผลลัพธ์ (History)", "DELETE", "/api/outputs/{fname}", "ทุก user ที่ login",
     "ลบไฟล์ผลลัพธ์", "path: fname"),

    ("ฐานข้อมูล (DATA)", "GET", "/api/database", "ทุก user ที่ login",
     "รายการไฟล์ต้นทางแยกกลุ่ม: Stock / Booking / SC / Datamining", "-"),
    ("ฐานข้อมูล (DATA)", "GET", "/api/database/sheet", "ทุก user ที่ login",
     "อ่านไฟล์เป็นตาราง (cap 5,000 แถว, truncated=true ถ้าเกิน)", "query: file=<group>/<name>, sheet"),
    ("ฐานข้อมูล (DATA)", "GET", "/api/database/download", "เปิด (ลิงก์ดาวน์โหลดตรง)",
     "ดาวน์โหลดไฟล์ต้นทางเต็มไฟล์", "query: file=<group>/<name>"),

    ("Map Item", "GET", "/api/map-item", "ทุก user ที่ login",
     "รายการไฟล์ผล map: datamining_mapped.xlsx, datamining_booking_mapped.xlsx", "-"),
    ("Map Item", "GET", "/api/map-item/sheet", "ทุก user ที่ login",
     "อ่านไฟล์ map เป็นตาราง (cap 5,000 แถว)", "query: file, sheet"),
    ("Map Item", "GET", "/api/map-item/download", "เปิด (ลิงก์ดาวน์โหลดตรง)",
     "ดาวน์โหลดไฟล์ map", "query: file"),

    ("แผนผลิต (Plan)", "GET", "/api/plan", "ทุก user ที่ login",
     "ข้อมูลไฟล์แผนล่าสุด: exists, name, size, mtime, รายชื่อชีท", "-"),
    ("แผนผลิต (Plan)", "GET", "/api/plan/sheet", "ทุก user ที่ login",
     "อ่านชีทของไฟล์แผน (ไม่ระบุ = ชีทแรก)", "query: sheet"),
    ("แผนผลิต (Plan)", "PUT", "/api/plan/sheet", "ทุก user ที่ login",
     "บันทึกชีทกลับไฟล์แผน (สำรอง .bak, ชีทอื่นคงเดิม)", "body: {sheet, columns, rows}"),
    ("แผนผลิต (Plan)", "GET", "/api/plan/load", "ทุก user ที่ login",
     "โหลดงานรายสัปดาห์ (ใช้วาดแถบโหลดบน Gantt)", "-"),
    ("แผนผลิต (Plan)", "GET", "/api/plan/ava", "ทุก user ที่ login",
     "เครื่องว่าง/ที่ใช้ ต่อสัปดาห์ ต่อ CAT|GUAGE (remain, used, planBase)", "-"),
    ("แผนผลิต (Plan)", "GET", "/api/plan/pool-map", "ทุก user ที่ login",
     "แผนที่กลุ่มที่แยกพูลเครื่อง (เช่น SKP vs SKPTA/SKPLE) → {CAT|GUAGE|MC_GROUP: pool}", "-"),
    ("แผนผลิต (Plan)", "GET", "/api/plan/booking-mc", "ทุก user ที่ login",
     "เครื่องที่ booking ถักไอเทมนั้นอยู่แล้ว ต่อ (สัปดาห์ × ITEM|MC_GROUP|GUAGE)", "-"),
    ("แผนผลิต (Plan)", "GET", "/api/plan/booking-items", "ทุก user ที่ login",
     "item ทั้งหมดจาก booking (แผนเก่า) — overlay อ่านอย่างเดียวบน Gantt", "-"),
    ("แผนผลิต (Plan)", "GET", "/api/plan/download", "เปิด (ลิงก์ดาวน์โหลดตรง)",
     "ดาวน์โหลดไฟล์แผนล่าสุด (no-store; ใส่ ?t=mtime กันแคช)", "query: t (cache-buster)"),

    ("Order Color", "GET", "/api/order-color", "ทุก user ที่ login",
     "ข้อมูลไฟล์ datamining_booking_mapped.xlsx + รายชื่อชีท", "-"),
    ("Order Color", "GET", "/api/order-color/sheet", "ทุก user ที่ login",
     "อ่านชีทของไฟล์ Order Color", "query: sheet"),
    ("Order Color", "PUT", "/api/order-color/sheet", "ทุก user ที่ login",
     "บันทึกชีทกลับไฟล์ Order Color (สำรอง .bak)", "body: {sheet, columns, rows}"),
    ("Order Color", "GET", "/api/order-color/download", "เปิด (ลิงก์ดาวน์โหลดตรง)",
     "ดาวน์โหลดไฟล์ Order Color", "query: t (cache-buster)"),
    ("Order Color", "POST", "/api/order-color/advise", "ทุก user ที่ login",
     "วิเคราะห์ item ที่ต้องย้อมสี → คำแนะนำต่อ item + summary", "-"),
    ("Order Color", "GET", "/api/order-color/cat-history", "ทุก user ที่ login",
     "booking รายสัปดาห์จัดกลุ่มตาม CAT (เฉพาะ CAT ที่มี item สี) — default view ของหน้า", "-"),
    ("Order Color", "GET", "/api/order-color/plan", "ทุก user ที่ login",
     "grid แผน what-if = แผนจริง + แทรกแถว item สีที่ยังไม่มีในแผน (+ color_idx)", "-"),
    ("Order Color", "GET", "/api/order-color/booking-gantt", "ทุก user ที่ login",
     "grid จาก booking DETAIL ทุก item ต่อสัปดาห์ + mark งานสี (ตัด item ที่ stock พอออก)", "-"),
    ("Order Color", "GET", "/api/order-color/booking-load", "ทุก user ที่ login",
     "โควตา setup job ต่อสัปดาห์ (live จาก NEW_MC + plan-new) + cap จาก REMAINING_JOBS", "-"),
    ("Order Color", "GET", "/api/order-color/booking-ava", "ทุก user ที่ login",
     "เครื่องว่างต่อสัปดาห์แบบสอดคล้อง booking (planBase = used)", "-"),
    ("Order Color", "POST", "/api/order-color/advise-moves", "ทุก user ที่ login",
     "AI จัดอันดับ + อธิบายผลกระทบการดึงงานสีเข้าแทนงานไม่มีสี (มี fallback ถ้า AI ไม่พร้อม)",
     "body: {cat, gauge, items[], setup_load{}}"),
    ("Order Color", "POST", "/api/order-color/plan/export", "ทุก user ที่ login",
     "แปลง grid what-if → ไฟล์ Excel order_color_plan.xlsx (ชีท PLAN_ORDER_COLOR)", "body: {columns, rows}"),

    ("จ้างทอ (Outsource)", "POST", "/api/outsource/advise", "ทุก user ที่ login",
     "AI แนะนำ item ที่ควรส่งจ้างทอ (shortlist + เหตุผล)", "-"),
    ("จ้างทอ (Outsource)", "GET", "/api/outsource/split", "ทุก user ที่ login",
     "การแบ่งจ้างทอที่ใช้อยู่ + สัปดาห์ในแผน (Planning.py อ่านไปใช้รอบถัดไป)", "-"),
    ("จ้างทอ (Outsource)", "POST", "/api/outsource/split", "ทุก user ที่ login",
     "บันทึกการแบ่งจ้างทอ (qty=0 = ยกเลิก); ตรวจไม่ให้เกินของค้าง", "body: {item_code, outsource_qty, start_week}"),
    ("จ้างทอ (Outsource)", "DELETE", "/api/outsource/split/{item_code}", "ทุก user ที่ login",
     "ลบการแบ่งจ้างทอของ item", "path: item_code (รองรับ / ในรหัส)"),

    ("Change Cylinder", "POST", "/api/cylinder/advise", "ทุก user ที่ login",
     "AI แนะนำการเปลี่ยน cylinder เพื่อปลดงานที่ติดเครื่อง (shortlist + เหตุผล)", "-"),

    ("หน้าเว็บ / ระบบ", "GET", "/", "เปิด",
     "เสิร์ฟ React SPA (index.html, no-cache)", "-"),
    ("หน้าเว็บ / ระบบ", "GET", "/assets/*", "เปิด",
     "ไฟล์ static ของ frontend build", "-"),
    ("หน้าเว็บ / ระบบ", "GET", "/{full_path}", "เปิด",
     "SPA routing — path ที่ไม่ใช่ไฟล์จริงคืน index.html", "-"),
]

# =====================================================================
# 2) Test Case
#    (module, feature, method, endpoint, role, ptype, risk, pre, steps, data, expected)
# =====================================================================
P, N, S, F = "Positive", "Negative", "Security", "Non-functional"
R_READ, R_WRITE, R_HEAVY = "อ่านอย่างเดียว", "เขียน/แก้ไขไฟล์", "รันงานหนัก"

TESTS = [
    # ---------------- Auth ----------------
    ("เข้าสู่ระบบ (Auth)", "เข้าสู่ระบบด้วยบัญชีที่ถูกต้อง", "POST", "/api/login", "-", P, R_READ,
     "server ทำงานอยู่ · มีบัญชีใน users.json",
     "1. เปิดเว็บ → หน้า Login\n2. กรอกชื่อผู้ใช้/รหัสผ่านที่ถูกต้อง\n3. กดเข้าสู่ระบบ",
     "username/password ของบัญชีทดสอบ",
     "HTTP 200 · ได้ token + username + role · เข้าหน้าหลักได้ · token ถูกเก็บใน localStorage (knitplan_token)"),
    ("เข้าสู่ระบบ (Auth)", "รหัสผ่านผิด", "POST", "/api/login", "-", N, R_READ,
     "server ทำงานอยู่",
     "1. กรอกชื่อผู้ใช้ถูก รหัสผ่านผิด\n2. กดเข้าสู่ระบบ",
     "username ถูก + password ผิด",
     "HTTP 401 · ข้อความ \"ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง\" · ไม่เข้าระบบ · ไม่มี token ถูกเก็บ"),
    ("เข้าสู่ระบบ (Auth)", "ชื่อผู้ใช้ไม่มีในระบบ", "POST", "/api/login", "-", N, R_READ,
     "server ทำงานอยู่",
     "1. กรอกชื่อผู้ใช้ที่ไม่มีอยู่จริง\n2. กดเข้าสู่ระบบ",
     "username = 'no_such_user'",
     "HTTP 401 · ข้อความเดียวกับรหัสผ่านผิด (ไม่บอกว่า user มี/ไม่มี)"),
    ("เข้าสู่ระบบ (Auth)", "ตรวจสิทธิ์ผู้ใช้ปัจจุบัน", "GET", "/api/me", "ทุก user", P, R_READ,
     "login แล้ว",
     "1. login\n2. รีเฟรชหน้าเว็บ (F5)",
     "-",
     "HTTP 200 · คืน username + role ตรงกับบัญชีที่ login · ยังอยู่ในระบบ ไม่ถูกเด้งออก"),
    ("เข้าสู่ระบบ (Auth)", "เรียก API โดยไม่มี token", "GET", "/api/plan", "-", S, R_READ,
     "-",
     "1. เรียก /api/plan โดยไม่ส่ง header Authorization",
     "ไม่ส่ง Authorization",
     "HTTP 401 · ข้อความ \"ยังไม่ได้เข้าสู่ระบบ หรือเซสชันหมดอายุ\""),
    ("เข้าสู่ระบบ (Auth)", "token ถูกแก้ไข/ปลอม", "GET", "/api/plan", "-", S, R_READ,
     "มี token จริงมาก่อน",
     "1. คัดลอก token แล้วแก้อักขระท้าย 1 ตัว\n2. เรียก API ด้วย token นั้น",
     "token ที่ signature ไม่ตรง",
     "HTTP 401 (ตรวจ HMAC ไม่ผ่าน) · ไม่มีข้อมูลรั่วออกมา"),
    ("เข้าสู่ระบบ (Auth)", "token หมดอายุ (12 ชม.)", "GET", "/api/me", "-", S, R_READ,
     "ตั้ง KNITPLAN_TOKEN_TTL=60 แล้ว restart เพื่อทดสอบให้เร็ว",
     "1. login\n2. รอให้เลย TTL\n3. เรียก API ใดๆ",
     "token ที่ exp ผ่านไปแล้ว",
     "HTTP 401 · frontend เคลียร์ token แล้วเด้งกลับหน้า Login อัตโนมัติ"),
    ("เข้าสู่ระบบ (Auth)", "สิทธิ์ admin เห็นเมนูครบ", "GET", "/api/me", "admin", P, R_READ,
     "มีบัญชี admin (role=admin ใน users.json)",
     "1. login ด้วยบัญชี admin\n2. ดูแถบเมนูด้านบน",
     "-",
     "เห็นแท็บ \"หน้าหลัก\" (Dashboard) ครบทุกเมนู · role ที่คืนจาก /api/me = admin"),
    ("เข้าสู่ระบบ (Auth)", "user ธรรมดาไม่เห็นหน้าหลัก", "GET", "/api/me", "user", P, R_READ,
     "มีบัญชี user ธรรมดา",
     "1. login ด้วยบัญชี user ธรรมดา\n2. ดูแถบเมนู",
     "-",
     "ไม่มีแท็บ \"หน้าหลัก\" · ระบบเปิดหน้า \"แผนผลิต\" ให้เป็นหน้าแรก"),
    ("เข้าสู่ระบบ (Auth)", "user ธรรมดาห้ามตั้งเวลาอัตโนมัติ", "PUT", "/api/schedule", "user", S, R_READ,
     "login ด้วย user ธรรมดา",
     "1. เรียก PUT /api/schedule ด้วย token ของ user ธรรมดา",
     "{schedule:{full:{enabled:true,hour:6,minute:0}}}",
     "HTTP 403 · ข้อความ \"ต้องเป็นผู้ดูแลระบบ (admin) เท่านั้น\" · settings.json ไม่ถูกแก้"),
    ("เข้าสู่ระบบ (Auth)", "health check ไม่ต้อง login", "GET", "/api/health", "-", P, R_READ,
     "server ทำงานอยู่",
     "1. เปิด /api/health โดยไม่ login",
     "-",
     "HTTP 200 · {\"ok\": true}"),
    ("เข้าสู่ระบบ (Auth)", "ลิงก์ดาวน์โหลดใช้ได้โดยไม่ต้องมี token (ตามการออกแบบ)", "GET", "/api/plan/download",
     "เปิด", P, R_READ,
     "มีไฟล์แผนใน data_plan",
     "1. เปิด URL /api/plan/download ตรงในเบราว์เซอร์ (ไม่ login)",
     "-",
     "ดาวน์โหลดไฟล์ได้ (ออกแบบให้ <a href> โหลดตรงได้) · ยืนยันกับผู้ใช้ว่ายอมรับพฤติกรรมนี้"),
    ("เข้าสู่ระบบ (Auth)", "ลบไฟล์โดยไม่มี token ต้องไม่ได้", "DELETE", "/api/outputs/{fname}", "-", S, R_READ,
     "มีไฟล์ใน data_plan",
     "1. เรียก DELETE /api/outputs/<ชื่อไฟล์> โดยไม่ส่ง token",
     "ชื่อไฟล์ที่มีจริง",
     "HTTP 401 · ไฟล์ยังอยู่ครบ (เส้นทางดาวน์โหลดเปิดเฉพาะ GET เท่านั้น)"),
    ("เข้าสู่ระบบ (Auth)", "ออกจากระบบ", "-", "-", "ทุก user", P, R_READ,
     "login แล้ว",
     "1. กดออกจากระบบ\n2. รีเฟรชหน้า",
     "-",
     "token ถูกลบจาก localStorage · กลับไปหน้า Login · เรียก API ต่อไม่ได้ (401)"),

    # ---------------- Run pipeline ----------------
    ("สั่งรัน Pipeline", "รันทั้ง Pipeline", "POST", "/api/run", "ทุก user", P, R_HEAVY,
     "ไม่มีงานรันค้างอยู่ · DB/ไฟล์ต้นทางพร้อม",
     "1. เข้าหน้าหลัก\n2. กด \"รันทั้ง Pipeline\"\n3. ดูแถบความคืบหน้าและ log",
     "mode = full",
     "ok=true · สถานะเปลี่ยนเป็นกำลังรัน · log ไหลต่อเนื่อง · จบด้วย returncode=0 · progress=100 · มีไฟล์ production_plan_DD-MM-YYYY.xlsx ใหม่"),
    ("สั่งรัน Pipeline", "ดึงข้อมูล DB อย่างเดียว", "POST", "/api/run", "ทุก user", P, R_HEAVY,
     "เชื่อม Oracle 172.16.7.55 ได้",
     "1. กด \"ดึงข้อมูล DB\"\n2. รอจนจบ",
     "mode = db",
     "รันเฉพาะขั้นดึงข้อมูล (ข้าม Calendar/Stock/AVA_MC/Order/Planning) · ไฟล์ใน Booking/, Stock/, Order/ อัปเดต mtime ใหม่"),
    ("สั่งรัน Pipeline", "รันแผนผลิตอย่างเดียว", "POST", "/api/run", "ทุก user", P, R_HEAVY,
     "มีข้อมูล booking/stock ล่าสุดแล้ว",
     "1. กด \"รันแผน\"\n2. รอจนจบ",
     "mode = plan",
     "รันตั้งแต่ AVA_MC → Order → Planning · ได้ไฟล์แผนใหม่ · ไม่แตะขั้นดึง DB"),
    ("สั่งรัน Pipeline", "ดึง Stock", "POST", "/api/run", "ทุก user", P, R_HEAVY,
     "เชื่อม Oracle ได้",
     "1. สั่งรันโหมด stock\n2. ดู log",
     "mode = stock",
     "รัน View_Stock.py เดี่ยว · ได้ไฟล์ Stock/view_stock.xlsx ใหม่"),
    ("สั่งรัน Pipeline", "ดึง Booking", "POST", "/api/run", "ทุก user", P, R_HEAVY,
     "เชื่อม Oracle ได้",
     "1. สั่งรันโหมด booking",
     "mode = booking",
     "รัน View_Booking.py เดี่ยว · ได้ไฟล์ใน Booking/ ใหม่"),
    ("สั่งรัน Pipeline", "ดึง SC", "POST", "/api/run", "ทุก user", P, R_HEAVY,
     "เชื่อม Oracle ได้",
     "1. สั่งรันโหมด sc",
     "mode = sc",
     "รัน View_SC.py เดี่ยว · ได้ไฟล์ view_sc_<timestamp>.xlsx เพิ่มในรายการ SC"),
    ("สั่งรัน Pipeline", "ดึง Datamining", "POST", "/api/run", "ทุก user", P, R_HEAVY,
     "เชื่อม Oracle ได้",
     "1. สั่งรันโหมด datamining",
     "mode = datamining",
     "รัน View_Datamining.py เดี่ยว · ไฟล์ในโฟลเดอร์ Datamining/ อัปเดต"),
    ("สั่งรัน Pipeline", "รัน Map Item", "POST", "/api/run", "ทุก user", P, R_HEAVY,
     "มีไฟล์ Datamining + Booking แล้ว",
     "1. สั่งรันโหมด map-item\n2. เปิดหน้า Map Item ตรวจไฟล์",
     "mode = map-item",
     "รัน MapItem.py · ได้ datamining_mapped.xlsx และ datamining_booking_mapped.xlsx ใหม่"),
    ("สั่งรัน Pipeline", "โหมดที่ไม่รู้จัก", "POST", "/api/run", "ทุก user", N, R_READ,
     "-",
     "1. เรียก POST /api/run ด้วย mode ที่ไม่มีในระบบ",
     "mode = 'xxx'",
     "ok=false · ข้อความ \"โหมดไม่ถูกต้อง: xxx\" · ไม่มี process ถูกสร้าง"),
    ("สั่งรัน Pipeline", "กันรันซ้อน", "POST", "/api/run", "ทุก user", N, R_HEAVY,
     "มีงานกำลังรันอยู่",
     "1. สั่งรันงานหนึ่ง\n2. ระหว่างที่ยังรัน สั่งรันอีกงาน (คนละเครื่อง/คนละ user ก็ได้)",
     "mode ใดก็ได้",
     "ok=false · ข้อความ \"มีงานกำลังรันอยู่ — รอให้เสร็จก่อน\" · งานเดิมยังรันต่อไม่สะดุด"),
    ("สั่งรัน Pipeline", "ติดตามสถานะระหว่างรัน", "GET", "/api/run/status", "ทุก user", P, R_READ,
     "มีงานกำลังรัน",
     "1. สั่งรัน\n2. ดูแถบสถานะ/ความคืบหน้าบนหน้าเว็บ",
     "-",
     "running=true · label ตรงกับโหมด · step_num/total_steps และ progress เพิ่มขึ้นตามขั้นตอนจริง · started_at ถูกต้อง"),
    ("สั่งรัน Pipeline", "อ่าน log แบบต่อเนื่อง", "GET", "/api/run/logs", "ทุก user", P, R_READ,
     "มีงานกำลังรัน",
     "1. เรียก /api/run/logs?offset=0\n2. เรียกซ้ำด้วย next_offset ที่ได้",
     "offset = 0 แล้วตามด้วย next_offset",
     "ได้เฉพาะบรรทัดใหม่ทุกครั้ง · ไม่ซ้ำ/ไม่ตกหล่น · next_offset เพิ่มขึ้นเรื่อยๆ"),
    ("สั่งรัน Pipeline", "log resync เมื่อรอบใหม่เริ่ม", "GET", "/api/run/logs", "ทุก user", N, R_READ,
     "รันรอบหนึ่งจบแล้ว เริ่มรอบใหม่",
     "1. จำ offset ของรอบเก่าไว้\n2. เริ่มรอบใหม่\n3. เรียก logs ด้วย offset เก่า (ค่ามากกว่าจำนวนบรรทัดปัจจุบัน)",
     "offset สูงกว่าจำนวน log จริง",
     "reset=true · ส่ง log ของรอบใหม่ทั้งหมดกลับมา · หน้าเว็บไม่ค้างว่างเปล่า"),
    ("สั่งรัน Pipeline", "สั่งหยุดงานที่กำลังรัน", "POST", "/api/run/stop", "ทุก user", P, R_HEAVY,
     "มีงานกำลังรัน",
     "1. สั่งรัน full\n2. กดปุ่มหยุด\n3. ดู log + Task Manager",
     "-",
     "ok=true · log มีบรรทัด \"⛔ ผู้ใช้สั่งหยุด\" และ \"⛔ ยกเลิกโดยผู้ใช้\" · process python ลูก (run_all + step) ถูกปิดหมด · running=false"),
    ("สั่งรัน Pipeline", "สั่งหยุดตอนไม่มีงานรัน", "POST", "/api/run/stop", "ทุก user", N, R_READ,
     "ไม่มีงานรันอยู่",
     "1. กดปุ่มหยุด",
     "-",
     "ok=false · ข้อความ \"ไม่มีงานกำลังรันอยู่\" · ไม่มี error/exception"),
    ("สั่งรัน Pipeline", "ไฟล์ log ต่อรอบถูกเก็บไว้", "-", "-", "ทุก user", P, R_WRITE,
     "รันงาน 1 รอบ",
     "1. รันงาน\n2. เปิดโฟลเดอร์ webapp/logs",
     "-",
     "มีไฟล์ run_<mode>_<YYYYMMDD_HHMMSS>.log · เนื้อหาตรงกับ log บนหน้าเว็บ (ภาษาไทยอ่านออก ไม่เป็นตัวประหลาด)"),
    ("สั่งรัน Pipeline", "user ธรรมดาสั่งรันได้", "POST", "/api/run", "user", P, R_HEAVY,
     "login ด้วย user ธรรมดา",
     "1. login user ธรรมดา\n2. สั่งรันแผน",
     "mode = plan",
     "รันได้ปกติ (ไม่ใช่สิทธิ์ admin only) · ok=true"),

    # ---------------- Schedule ----------------
    ("ตั้งเวลาอัตโนมัติ", "ดูเวลาที่ตั้งไว้", "GET", "/api/schedule", "ทุก user", P, R_READ,
     "server ทำงานอยู่",
     "1. เปิดหน้าตั้งเวลา",
     "-",
     "แสดง enabled/hour/minute ของ job full · next_run เป็นเวลาไทย (Asia/Bangkok) ที่ถูกต้อง"),
    ("ตั้งเวลาอัตโนมัติ", "แก้เวลารันอัตโนมัติ", "PUT", "/api/schedule", "admin", P, R_WRITE,
     "login ด้วย admin",
     "1. เปลี่ยนเวลาเป็นเวลาที่ต้องการ\n2. กดบันทึก\n3. เรียก GET /api/schedule ซ้ำ",
     "{full:{enabled:true, hour:6, minute:30}}",
     "บันทึกลง settings.json · next_run เปลี่ยนตามเวลาใหม่ทันที"),
    ("ตั้งเวลาอัตโนมัติ", "ปิดการรันอัตโนมัติ", "PUT", "/api/schedule", "admin", P, R_WRITE,
     "login ด้วย admin",
     "1. ตั้ง enabled=false\n2. บันทึก",
     "{full:{enabled:false, hour:6, minute:0}}",
     "job ถูกถอดออก · next_run = null · ไม่มีการรันอัตโนมัติเกิดขึ้น"),
    ("ตั้งเวลาอัตโนมัติ", "ตั้งเวลาแล้วรันจริงตามเวลา", "-", "-", "admin", P, R_HEAVY,
     "ตั้งเวลาเป็นอีก 2-3 นาทีข้างหน้า",
     "1. ตั้งเวลาใกล้ๆ\n2. รอถึงเวลา\n3. ดูสถานะ/log",
     "hour/minute = เวลาปัจจุบัน + 2 นาที",
     "งานเริ่มรันเองตามเวลา · สถานะ trigger = \"schedule\" · ได้ผลลัพธ์เหมือนสั่งรันด้วยมือ"),
    ("ตั้งเวลาอัตโนมัติ", "ค่าที่ตั้งคงอยู่หลัง restart", "-", "-", "admin", P, R_WRITE,
     "ตั้งเวลาไว้แล้ว",
     "1. restart server\n2. เปิดหน้าตั้งเวลา",
     "-",
     "เวลาที่ตั้งไว้ยังอยู่ (อ่านจาก settings.json) · job ถูกตั้งใหม่อัตโนมัติตอน startup"),
    ("ตั้งเวลาอัตโนมัติ", "ลบไฟล์เก่าอัตโนมัติ (cleanup)", "-", "-", "-", F, R_WRITE,
     "มีไฟล์ log/แผน อายุเกิน 14 วันใน webapp/logs และ data_plan",
     "1. restart server (cleanup รันทันที 1 ครั้ง)\n2. ดู log บรรทัด [cleanup]\n3. ตรวจโฟลเดอร์",
     "-",
     "ลบเฉพาะ *.log, production_plan_*.xlsx, booking_final_ready_*.xlsx ที่เกิน 14 วัน · ไฟล์ทำงานอื่น (เช่น datamining_booking_mapped.xlsx) ไม่ถูกลบ"),

    # ---------------- Masters ----------------
    ("Master Data", "ดูรายการไฟล์ Master", "GET", "/api/masters", "ทุก user", P, R_READ,
     "config.ini ชี้ path Master ถูกต้อง",
     "1. เปิดหน้า Master Data",
     "-",
     "แสดง MasterMC / Calendar / Target_Stock / Master_Item · exists=true ทุกไฟล์ · รายชื่อชีทครบตรงกับไฟล์จริง"),
    ("Master Data", "เปิดดูชีท MasterMC", "GET", "/api/masters/MasterMC/Master MC", "ทุก user", P, R_READ,
     "มีไฟล์ MasterMC.xlsx",
     "1. เลือกไฟล์ MasterMC → ชีท Master MC",
     "-",
     "ตารางแสดงครบทุกคอลัมน์/แถว · เซลล์สูตรแสดงเป็นค่าที่คำนวณแล้ว ไม่ใช่ข้อความ =..."),
    ("Master Data", "เปิดชีทที่ไม่มีอยู่จริง", "GET", "/api/masters/{name}/{sheet}", "ทุก user", N, R_READ,
     "-",
     "1. เรียก /api/masters/MasterMC/ชีทมั่ว",
     "sheet = 'NOT_EXIST'",
     "HTTP 404 · ข้อความ \"ไม่พบชีท ... ใน MasterMC\""),
    ("Master Data", "เปิดไฟล์ Master ที่ไม่รู้จัก", "GET", "/api/masters/{name}/{sheet}", "ทุก user", N, R_READ,
     "-",
     "1. เรียก /api/masters/XXX/Sheet1",
     "name = 'XXX'",
     "HTTP 404 · ข้อความ \"ไม่รู้จัก Master 'XXX'\""),
    ("Master Data", "แก้ไขและบันทึกชีท Master", "PUT", "/api/masters/{name}/{sheet}", "ทุก user", P, R_WRITE,
     "สำรองไฟล์ Master ไว้ก่อน · ไฟล์ต้องไม่ถูกเปิดค้างใน Excel",
     "1. แก้ค่าในตาราง\n2. กดบันทึก\n3. เปิดไฟล์จริงตรวจสอบ",
     "แก้ 1 เซลล์ที่ไม่กระทบแผน",
     "ok=true · คืนชื่อไฟล์ .bak ที่สำรอง · ค่าที่แก้ถูกเขียนลงไฟล์ · ชีทอื่นในไฟล์เดียวกันไม่เปลี่ยน"),
    ("Master Data", "ไฟล์สำรอง .bak ถูกสร้างทุกครั้ง", "PUT", "/api/masters/{name}/{sheet}", "ทุก user", P, R_WRITE,
     "-",
     "1. บันทึก Master 2 ครั้ง\n2. ดูโฟลเดอร์ที่เก็บไฟล์ Master",
     "-",
     "มีไฟล์ <ชื่อไฟล์>.xlsx.<YYYYMMDD_HHMMSS>.bak เพิ่มทุกครั้งที่บันทึก · เปิดไฟล์ .bak ได้และเป็นข้อมูลก่อนแก้"),
    ("Master Data", "Target_Stock คำนวณคอลัมน์สูตรใหม่", "PUT", "/api/masters/Target_Stock/{sheet}", "ทุก user", P, R_WRITE,
     "มีไฟล์ Target_Stock.xlsx",
     "1. แก้ค่า TARGET/MONTH ของ 1 แถว\n2. บันทึก\n3. เปิดไฟล์ตรวจคอลัมน์สูตร",
     "TARGET/MONTH = 4000",
     "TARGET SCM = 2000 · STOCK MIN = 1000 · STOCK MAX = 3000 · Stock 5 Week = 5000 · Match core = Item code + Team Name · ค่าที่เขียนเป็นตัวเลขจริง (ไม่ใช่สูตร)"),
    ("Master Data", "Target_Stock กติกา RTS + CG", "PUT", "/api/masters/Target_Stock/{sheet}", "ทุก user", P, R_WRITE,
     "มีแถวที่ Team Name = RTS และ Type 1 = CG",
     "1. แก้ TARGET/MONTH ของแถวนั้น\n2. บันทึก\n3. ตรวจ STOCK MAX",
     "Team=RTS, Type1=CG, TARGET/MONTH = 4000",
     "STOCK MAX = 4000 (เท่ากับ TARGET/MONTH ไม่ใช่ TARGET SCM × 1.5)"),
    ("Master Data", "ค่าวันที่ถูกเก็บเป็นวันที่จริง", "PUT", "/api/masters/{name}/{sheet}", "ทุก user", P, R_WRITE,
     "ชีทมีคอลัมน์วันที่",
     "1. แก้ค่าเซลล์วันที่\n2. บันทึก\n3. เปิดไฟล์ใน Excel",
     "2026-08-15",
     "Excel แสดงเป็นวันที่ (จัดชิดขวา/ฟอร์แมตวันที่) ไม่ใช่ข้อความ · pipeline อ่านค่าไปใช้ได้"),
    ("Master Data", "รหัสที่มีเลข 0 นำหน้าไม่ถูกแปลง", "PUT", "/api/masters/{name}/{sheet}", "ทุก user", N, R_WRITE,
     "-",
     "1. กรอกรหัสที่ขึ้นต้นด้วย 0\n2. บันทึก\n3. เปิดไฟล์ตรวจ",
     "007",
     "ค่าที่เก็บยังเป็น \"007\" (ข้อความ) ไม่กลายเป็น 7"),
    ("Master Data", "บันทึกขณะไฟล์ถูกเปิดค้างใน Excel", "PUT", "/api/masters/{name}/{sheet}", "ทุก user", N, R_WRITE,
     "เปิดไฟล์ Master ค้างไว้ใน Excel บนเครื่อง server",
     "1. เปิดไฟล์ค้างไว้\n2. กดบันทึกจากเว็บ",
     "-",
     "แสดงข้อความผิดพลาดชัดเจน (บันทึกไม่ได้/ไฟล์ถูกใช้งาน) · ไฟล์ต้นฉบับไม่เสียหาย · ไม่มีข้อมูลหาย"),
    ("Master Data", "แก้ Master แล้วมีผลกับแผนรอบถัดไป", "-", "-", "ทุก user", P, R_HEAVY,
     "แก้ค่าใน MasterMC (เช่น Lock_MC / POLY / COTTON)",
     "1. แก้ค่าและบันทึก\n2. สั่งรันแผน (mode=plan)\n3. เปิดแผนใหม่เทียบ",
     "-",
     "แผนที่ได้สะท้อนค่าที่แก้ (จำนวนเครื่อง/การจองเปลี่ยนตามที่คาด)"),

    # ---------------- Work Day ----------------
    ("วันทำงาน (Work Day)", "เปิดแผงวันทำงาน", "GET", "/api/workday", "ทุก user", P, R_READ,
     "มีไฟล์ Calendar.xlsx และ MasterMC.xlsx",
     "1. เปิดแผงวันทำงานตามกลุ่มเครื่อง",
     "-",
     "แสดงกลุ่มเครื่อง (Factory|MC_CAT|Guage) ครบ · มีค่าวันมาตรฐาน/ชั่วโมง/ค่าเฉพาะสัปดาห์/การยุบสัปดาห์ · ค่ามาตรฐาน fallback = 6 วัน 24 ชม."),
    ("วันทำงาน (Work Day)", "ตั้งวันทำงานมาตรฐานต่อกลุ่ม", "PUT", "/api/workday", "ทุก user", P, R_WRITE,
     "สำรอง Calendar.xlsx ไว้",
     "1. แก้วันทำงานของ 1 กลุ่ม\n2. บันทึก\n3. เปิด Calendar.xlsx ชีท \"Work Day\"",
     "กลุ่มใดกลุ่มหนึ่ง = 5 วัน",
     "ok=true + ชื่อไฟล์ .bak · ชีท Work Day มีแถวของกลุ่มนั้น (WEEK ว่าง, WORK_DAY=5) · โหลดหน้าใหม่ยังเห็นค่าเดิม"),
    ("วันทำงาน (Work Day)", "ตั้งวันทำงานเฉพาะสัปดาห์", "PUT", "/api/workday", "ทุก user", P, R_WRITE,
     "-",
     "1. ตั้งค่าเฉพาะสัปดาห์ให้กลุ่มหนึ่ง\n2. บันทึก\n3. ตรวจชีท Work Day",
     "กลุ่ม X สัปดาห์ 33 = 4 วัน",
     "มีแถวที่ WEEK=33, WORK_DAY=4 · ค่ามาตรฐานของกลุ่มยังอยู่แยกแถว · สัปดาห์อื่นไม่กระทบ"),
    ("วันทำงาน (Work Day)", "ตั้งชั่วโมงทำงานต่อกลุ่ม", "PUT", "/api/workday", "ทุก user", P, R_WRITE,
     "-",
     "1. ตั้งชั่วโมงของกลุ่มเป็น 20\n2. บันทึก\n3. ตรวจคอลัมน์ WORK_HOUR",
     "hours: กลุ่ม X = 20",
     "ชีท Work Day คอลัมน์ WORK_HOUR = 20 ในแถวค่ามาตรฐาน · กำลังผลิตของกลุ่มถูกคูณ 20/24 ในแผนรอบถัดไป"),
    ("วันทำงาน (Work Day)", "ยุบสัปดาห์ (Week Merge)", "PUT", "/api/workday", "ทุก user", P, R_WRITE,
     "-",
     "1. ตั้งยุบสัปดาห์ 31 → 32\n2. บันทึก\n3. ตรวจชีท \"Week Merge\"",
     "merges: {31: 32}",
     "ชีท Week Merge มีแถว 31|32 · งานของสัปดาห์ 31 ถูกยุบไปสัปดาห์ 32 ในการคำนวณ"),
    ("วันทำงาน (Work Day)", "ยุบสัปดาห์เข้าตัวเอง (ค่าไม่ถูกต้อง)", "PUT", "/api/workday", "ทุก user", N, R_WRITE,
     "-",
     "1. ตั้ง merge 31 → 31\n2. บันทึก",
     "merges: {31: 31}",
     "แถวนี้ถูกตัดทิ้ง ไม่ถูกเขียนลงไฟล์ · ไม่เกิด loop/ข้อผิดพลาด"),
    ("วันทำงาน (Work Day)", "สร้างค่าเริ่มต้นจาก MasterMC (seed)", "POST", "/api/workday/seed", "ทุก user", P, R_WRITE,
     "Calendar.xlsx ยังไม่มีชีท Work Day (หรือมีไม่ครบทุกกลุ่ม)",
     "1. กดปุ่ม seed\n2. เปิดชีท Work Day",
     "-",
     "added > 0 · ทุกกลุ่มได้ค่าจากคอลัมน์ Working Day ของ MasterMC (ไม่มีค่า = 6 วัน) · ค่าที่ผู้ใช้ตั้งไว้ก่อนหน้าไม่ถูกทับ"),
    ("วันทำงาน (Work Day)", "กด seed ซ้ำ", "POST", "/api/workday/seed", "ทุก user", N, R_WRITE,
     "seed ไปแล้วครั้งหนึ่ง",
     "1. กด seed อีกครั้ง",
     "-",
     "added = 0 · ไฟล์ Calendar.xlsx ไม่ถูกแก้/ไม่มี .bak ใหม่ · ค่าที่ตั้งไว้คงเดิม"),
    ("วันทำงาน (Work Day)", "วันทำงานถูก cap ด้วยปฏิทิน", "GET", "/api/workday", "ทุก user", P, R_READ,
     "สัปดาห์ที่เลือกมีวันหยุดในปฏิทิน (เปิดทำงาน 5 วัน)",
     "1. ตั้งวันทำงานกลุ่มเป็น 7\n2. ดูวันทำงานที่ระบบใช้จริงในสัปดาห์นั้น",
     "ตั้ง 7 วัน แต่ปฏิทินเปิด 5 วัน",
     "ระบบใช้ 5 วัน (min ของแผงวันทำงานกับปฏิทิน) · ตรงกับที่ pipeline คำนวณ"),
    ("วันทำงาน (Work Day)", "ไม่มีไฟล์ Calendar", "GET", "/api/workday", "ทุก user", N, R_READ,
     "เปลี่ยนชื่อ/ย้าย Calendar.xlsx ชั่วคราว",
     "1. เปิดแผงวันทำงาน",
     "-",
     "HTTP 404 พร้อมข้อความบอกว่าไม่พบไฟล์ · server ไม่ crash"),
    ("วันทำงาน (Work Day)", "เติมปฏิทินปีถัดไปอัตโนมัติตอน start", "-", "-", "-", F, R_WRITE,
     "Calendar.xlsx ยังไม่มีวันของปีถัดไป",
     "1. restart server\n2. ดู console log\n3. เปิด Calendar.xlsx",
     "-",
     "log แสดง [CALENDAR AUTO-EXTEND] เติม N วัน + ชื่อไฟล์ backup · ปฏิทินมีวันครบถึงสิ้นปีถัดไป · start ซ้ำแล้วไม่เติมซ้ำ"),

    # ---------------- Outputs ----------------
    ("ไฟล์ผลลัพธ์ (History)", "ดูรายการไฟล์แผนผลิต", "GET", "/api/outputs", "ทุก user", P, R_READ,
     "มีไฟล์ production_plan_*.xlsx",
     "1. เปิดหน้า History",
     "-",
     "รายการเรียงจากใหม่ → เก่า · แสดงชื่อ/ขนาด/เวลาแก้ไขถูกต้อง"),
    ("ไฟล์ผลลัพธ์ (History)", "ดูรายการไฟล์ booking_final", "GET", "/api/outputs/booking", "ทุก user", P, R_READ,
     "มีไฟล์ booking_final_ready_*.xlsx",
     "1. เปิดแท็บ Booking ในหน้า History",
     "-",
     "แสดงเฉพาะไฟล์ booking_final_ready_*.xlsx เรียงใหม่→เก่า"),
    ("ไฟล์ผลลัพธ์ (History)", "ดูรายการไฟล์ SC", "GET", "/api/outputs/sc", "ทุก user", P, R_READ,
     "มีไฟล์ view_sc_*.xlsx",
     "1. เปิดแท็บ SC",
     "-",
     "แสดงเฉพาะไฟล์ view_sc_*.xlsx เรียงใหม่→เก่า"),
    ("ไฟล์ผลลัพธ์ (History)", "ดาวน์โหลดไฟล์ผลลัพธ์", "GET", "/api/outputs/{fname}", "เปิด", P, R_READ,
     "มีไฟล์ในรายการ",
     "1. กดดาวน์โหลดไฟล์\n2. เปิดไฟล์ด้วย Excel",
     "-",
     "ได้ไฟล์ .xlsx ชื่อเดิม เปิดได้ไม่เสียหาย · header Cache-Control: no-store"),
    ("ไฟล์ผลลัพธ์ (History)", "ดาวน์โหลดไฟล์ที่ไม่มีอยู่", "GET", "/api/outputs/{fname}", "เปิด", N, R_READ,
     "-",
     "1. เรียก /api/outputs/ไฟล์ที่ไม่มีจริง.xlsx",
     "ชื่อไฟล์มั่ว",
     "HTTP 404 · ข้อความ \"ไม่พบไฟล์\""),
    ("ไฟล์ผลลัพธ์ (History)", "กันเข้าถึงไฟล์นอกโฟลเดอร์ (path traversal)", "GET", "/api/outputs/{fname}", "เปิด", S, R_READ,
     "-",
     "1. เรียก /api/outputs/..%2F..%2Fconfig.ini",
     "../../config.ini",
     "ไม่ได้ไฟล์นอกโฟลเดอร์ data_plan (404 หรือ error) · ไม่มีเนื้อหาไฟล์ระบบหลุดออกมา"),
    ("ไฟล์ผลลัพธ์ (History)", "ลบไฟล์ผลลัพธ์", "DELETE", "/api/outputs/{fname}", "ทุก user", P, R_WRITE,
     "มีไฟล์ทดสอบที่ลบได้ (คัดลอกไว้ก่อน)",
     "1. กดลบไฟล์\n2. รีเฟรชรายการ",
     "ชื่อไฟล์ทดสอบ",
     "ok=true, deleted=<ชื่อไฟล์> · ไฟล์หายจากโฟลเดอร์และรายการ · ไฟล์อื่นไม่ถูกลบ"),
    ("ไฟล์ผลลัพธ์ (History)", "ลบไฟล์ที่ไม่มีอยู่", "DELETE", "/api/outputs/{fname}", "ทุก user", N, R_READ,
     "-",
     "1. ลบไฟล์ที่ไม่มีจริง",
     "ชื่อไฟล์มั่ว",
     "HTTP 404 · ข้อความ \"ไม่พบไฟล์\""),

    # ---------------- Database (DATA) ----------------
    ("ฐานข้อมูล (DATA)", "ดูรายการไฟล์ข้อมูลต้นทาง", "GET", "/api/database", "ทุก user", P, R_READ,
     "มีไฟล์ในโฟลเดอร์ Stock/Booking/Order/Datamining",
     "1. เปิดหน้า DATA",
     "-",
     "เห็น 4 กลุ่ม: Stock, Booking, SC, Datamining · แต่ละกลุ่มแสดงไฟล์เรียงใหม่→เก่า พร้อมขนาด/เวลา"),
    ("ฐานข้อมูล (DATA)", "เปิดดูข้อมูลในไฟล์", "GET", "/api/database/sheet", "ทุก user", P, R_READ,
     "-",
     "1. เลือกไฟล์ 1 ไฟล์\n2. ดูตาราง",
     "file = stock/view_stock.xlsx",
     "แสดงชื่อชีททั้งหมด + คอลัมน์/ข้อมูลถูกต้อง · ค่าว่างแสดงเป็นช่องว่าง ไม่ใช่ NaN"),
    ("ฐานข้อมูล (DATA)", "สลับชีทในไฟล์", "GET", "/api/database/sheet", "ทุก user", P, R_READ,
     "ไฟล์มีมากกว่า 1 ชีท",
     "1. เลือกชีทอื่นจากรายการ",
     "sheet = ชีทที่สอง",
     "ตารางเปลี่ยนตามชีทที่เลือก · ระบุชื่อชีทผิด → แสดงชีทแรกแทน (ไม่ error)"),
    ("ฐานข้อมูล (DATA)", "ไฟล์ที่มีแถวเกิน 5,000", "GET", "/api/database/sheet", "ทุก user", P, R_READ,
     "มีไฟล์ที่แถวเกิน 5,000",
     "1. เปิดไฟล์นั้น\n2. ดูจำนวนแถวที่แสดงและข้อความเตือน",
     "-",
     "truncated=true · แสดง 5,000 แถวแรก · total = จำนวนแถวจริง · มีข้อความบอกให้ดาวน์โหลดไฟล์เต็ม"),
    ("ฐานข้อมูล (DATA)", "ดาวน์โหลดไฟล์ต้นทาง", "GET", "/api/database/download", "เปิด", P, R_READ,
     "-",
     "1. กดดาวน์โหลดไฟล์\n2. เปิดตรวจ",
     "file = booking/<ชื่อไฟล์>",
     "ได้ไฟล์เต็ม (ทุกแถว ไม่ถูก cap) เปิดได้ปกติ"),
    ("ฐานข้อมูล (DATA)", "ระบุกลุ่ม/ไฟล์ไม่ถูกต้อง", "GET", "/api/database/sheet", "ทุก user", N, R_READ,
     "-",
     "1. เรียกด้วย file=xxx/yyy.xlsx",
     "กลุ่มไม่มีจริง",
     "HTTP 404 · ข้อความ \"ไม่รู้จักกลุ่ม 'xxx'\" หรือ \"ไม่พบไฟล์\""),
    ("ฐานข้อมูล (DATA)", "กัน path traversal", "GET", "/api/database/download", "เปิด", S, R_READ,
     "-",
     "1. เรียก file=stock/../../config.ini",
     "stock/../../config.ini",
     "ระบบตัดเหลือชื่อไฟล์อย่างเดียว → 404 · ไม่มีไฟล์นอกโฟลเดอร์หลุดออกมา"),
    ("ฐานข้อมูล (DATA)", "ไม่แสดงไฟล์ชั่วคราวของ Excel", "GET", "/api/database", "ทุก user", N, R_READ,
     "เปิดไฟล์ Excel ค้างไว้ (เกิดไฟล์ ~$xxx.xlsx)",
     "1. เปิดหน้า DATA",
     "-",
     "ไฟล์ที่ขึ้นต้นด้วย ~$ ไม่ปรากฏในรายการ"),

    # ---------------- Map Item ----------------
    ("Map Item", "ดูรายการไฟล์ผล Map Item", "GET", "/api/map-item", "ทุก user", P, R_READ,
     "รัน MapItem แล้ว",
     "1. เปิดหน้า Map Item",
     "-",
     "เห็น 2 รายการ: Datamining → ORA Item และ Datamining → Booking · exists=true พร้อมขนาด/เวลา"),
    ("Map Item", "เปิดตาราง Datamining → ORA Item", "GET", "/api/map-item/sheet", "ทุก user", P, R_READ,
     "มีไฟล์ datamining_mapped.xlsx",
     "1. เลือกไฟล์แรก",
     "file = datamining_mapped.xlsx",
     "ตารางแสดง ITEM จาก datamining ที่ map เป็น ORA_ITEM_CODE ถูกต้อง"),
    ("Map Item", "เปิดตาราง Datamining → Booking", "GET", "/api/map-item/sheet", "ทุก user", P, R_READ,
     "มีไฟล์ datamining_booking_mapped.xlsx",
     "1. เลือกไฟล์ที่สอง",
     "file = datamining_booking_mapped.xlsx",
     "ตารางแสดง item ที่เชื่อมกับ booking รายสัปดาห์ครบ"),
    ("Map Item", "ดาวน์โหลดไฟล์ map", "GET", "/api/map-item/download", "เปิด", P, R_READ,
     "-",
     "1. กดดาวน์โหลด",
     "file = datamining_booking_mapped.xlsx",
     "ได้ไฟล์ .xlsx ล่าสุด (ไม่ใช่ไฟล์เก่าจากแคช เพราะ URL มี ?t=mtime + header no-store)"),
    ("Map Item", "ยังไม่เคยรัน Map Item", "GET", "/api/map-item/sheet", "ทุก user", N, R_READ,
     "ยังไม่มีไฟล์ผลลัพธ์ใน data_plan",
     "1. เปิดหน้า Map Item",
     "-",
     "รายการแสดง exists=false · เปิดตารางแล้วได้ HTTP 404 พร้อมข้อความบอกให้กดดึงข้อมูลก่อน · ไม่ค้าง/ไม่ crash"),
    ("Map Item", "ไฟล์อัปเดตหลังรัน map-item", "-", "-", "ทุก user", P, R_HEAVY,
     "-",
     "1. จำเวลาแก้ไขไฟล์เดิม\n2. สั่งรันโหมด map-item\n3. รีเฟรชหน้า",
     "-",
     "เวลาแก้ไข (mtime) ของทั้ง 2 ไฟล์ใหม่กว่าเดิม · ข้อมูลในตารางเปลี่ยนตามข้อมูลล่าสุด"),

    # ---------------- Plan ----------------
    ("แผนผลิต (Plan)", "เปิดหน้าแผนผลิตล่าสุด", "GET", "/api/plan", "ทุก user", P, R_READ,
     "มีไฟล์ production_plan_*.xlsx",
     "1. เปิดหน้าแผนผลิต",
     "-",
     "แสดงชื่อไฟล์แผนล่าสุด + เวลาแก้ไข + รายชื่อชีทครบ (PLAN, SUMMARY_MC_REMAIN, REMAINING_JOBS ฯลฯ)"),
    ("แผนผลิต (Plan)", "อ่านชีท PLAN", "GET", "/api/plan/sheet", "ทุก user", P, R_READ,
     "-",
     "1. เลือกชีท PLAN",
     "-",
     "ตารางแสดงครบทุกแถว/คอลัมน์ ตรงกับไฟล์ Excel · ตัวเลขไม่เพี้ยน"),
    ("แผนผลิต (Plan)", "สลับดูชีทอื่นของแผน", "GET", "/api/plan/sheet", "ทุก user", P, R_READ,
     "-",
     "1. เลือกชีท SUMMARY_MC_REMAIN\n2. เลือกชีทอื่นๆ ทีละชีท",
     "sheet = แต่ละชีทในไฟล์",
     "ทุกชีทเปิดได้ ข้อมูลตรงกับไฟล์จริง · ชีทที่ไม่มีจริง → HTTP 404"),
    ("แผนผลิต (Plan)", "แถบโหลดงานรายสัปดาห์", "GET", "/api/plan/load", "ทุก user", P, R_READ,
     "-",
     "1. เปิด Gantt แผนผลิต\n2. ดูแถบโหลดของแต่ละสัปดาห์",
     "-",
     "สัดส่วนโหลดต่อสัปดาห์ตรงกับข้อมูลในไฟล์แผน · สัปดาห์ที่ยุบ (merge) แสดงรวมกันถูกต้อง"),
    ("แผนผลิต (Plan)", "เครื่องว่าง/ที่ใช้ ต่อสัปดาห์", "GET", "/api/plan/ava", "ทุก user", P, R_READ,
     "-",
     "1. เปิด Gantt\n2. ดูตัวเลขเครื่องว่างของแต่ละ CAT|เกจ",
     "-",
     "remain/used ตรงกับชีท SUMMARY_MC_REMAIN · เครื่องที่กันไว้ POLY/COTTON ถูกหักตาม booking ที่ใช้ไปแล้ว"),
    ("แผนผลิต (Plan)", "การแยกพูลเครื่อง", "GET", "/api/plan/pool-map", "ทุก user", P, R_READ,
     "มีกลุ่มที่แยกพูล (เช่น SKP กับ SKPTA/SKPLE)",
     "1. เรียก /api/plan/pool-map\n2. เทียบกับ MasterMC",
     "-",
     "คืนเฉพาะกลุ่มที่แยกพูลจริง · Gantt ไม่นับเครื่องข้ามพูลให้กัน"),
    ("แผนผลิต (Plan)", "เครื่องจาก booking ต่อ item/สัปดาห์", "GET", "/api/plan/booking-mc", "ทุก user", P, R_READ,
     "มีไฟล์ booking_final_ready ล่าสุด",
     "1. เปิด Gantt\n2. ดูงานที่มาจาก booking",
     "-",
     "เครื่องที่ booking ถักอยู่แล้วถูกแยกจากเครื่องที่แผนจองจากพูล · ไม่ถูกนับซ้ำ"),
    ("แผนผลิต (Plan)", "overlay งานเดิมจาก booking", "GET", "/api/plan/booking-items", "ทุก user", P, R_READ,
     "-",
     "1. เปิด Gantt\n2. เปิด/ปิดการแสดง overlay booking",
     "-",
     "แสดงบล็อกงานเดิมเป็นแบบอ่านอย่างเดียว (ลากไม่ได้) · ไม่ปนกับงานของแผนใหม่"),
    ("แผนผลิต (Plan)", "แก้ไขแผนแล้วบันทึก", "PUT", "/api/plan/sheet", "ทุก user", P, R_WRITE,
     "สำรองไฟล์แผนไว้ก่อน",
     "1. แก้ค่าในตารางแผน\n2. กดบันทึก\n3. ดาวน์โหลดไฟล์มาเปิดตรวจ",
     "แก้ 1 แถว",
     "ok=true + ชื่อ .bak · ค่าใหม่อยู่ในไฟล์ · ชีทอื่นไม่เปลี่ยน"),
    ("แผนผลิต (Plan)", "ลากย้ายงานบน Gantt", "-", "-", "ทุก user", P, R_READ,
     "เปิดหน้า Gantt",
     "1. ลากงานจากสัปดาห์หนึ่งไปอีกสัปดาห์\n2. ดูตัวเลขเครื่องว่าง/โหลด",
     "-",
     "จำนวนเครื่องว่างและแถบโหลดคำนวณใหม่ทันทีทั้งต้นทางและปลายทาง · เกินเครื่องที่มี → มีการเตือน"),
    ("แผนผลิต (Plan)", "กรอง/แสดงผลบน Gantt", "-", "-", "ทุก user", P, R_READ,
     "เปิดหน้า Gantt",
     "1. กรองตามประเภทงาน / CAT / เกจ\n2. ดูสีของบล็อกงาน",
     "-",
     "ตัวกรองทำงานถูกต้อง (เหลือเฉพาะที่เลือก) · สีบล็อกแยกตาม CAT อ่านง่าย"),
    ("แผนผลิต (Plan)", "ดาวน์โหลดไฟล์แผน", "GET", "/api/plan/download", "เปิด", P, R_READ,
     "มีไฟล์แผน",
     "1. กดดาวน์โหลดแผน\n2. เปิดไฟล์",
     "-",
     "ได้ไฟล์แผนล่าสุด เปิดได้ปกติ"),
    ("แผนผลิต (Plan)", "ดาวน์โหลดหลังบันทึกต้องได้ไฟล์ใหม่", "GET", "/api/plan/download", "เปิด", N, R_WRITE,
     "-",
     "1. บันทึกการแก้ไขแผน\n2. กดดาวน์โหลดทันที\n3. เปิดไฟล์ตรวจค่าที่แก้",
     "-",
     "ไฟล์ที่ได้มีค่าที่เพิ่งแก้ (ไม่ใช่ไฟล์เก่าจากแคชเบราว์เซอร์)"),
    ("แผนผลิต (Plan)", "ยังไม่มีไฟล์แผน", "GET", "/api/plan/download", "เปิด", N, R_READ,
     "ย้ายไฟล์ production_plan_*.xlsx ออกชั่วคราว",
     "1. เปิดหน้าแผนผลิต\n2. กดดาวน์โหลด",
     "-",
     "หน้าเว็บแจ้งว่ายังไม่มีแผน · ดาวน์โหลด → HTTP 404 \"ยังไม่มีไฟล์แผนผลิต\" · ไม่ crash"),

    # ---------------- Order Color ----------------
    ("Order Color", "เปิดหน้า Order Color", "GET", "/api/order-color", "ทุก user", P, R_READ,
     "มีไฟล์ datamining_booking_mapped.xlsx",
     "1. เปิดเมนู Order Color",
     "-",
     "แสดงชื่อไฟล์ + เวลาแก้ไข + รายชื่อชีท"),
    ("Order Color", "มุมมองหลัก: booking ตาม CAT", "GET", "/api/order-color/cat-history", "ทุก user", P, R_READ,
     "-",
     "1. เปิดหน้า Order Color (มุมมองเริ่มต้น)",
     "-",
     "แสดงเฉพาะ CAT ที่มีงานสี (LOAD_DYE) · ในแต่ละ CAT มีทุก item · ค่าที่โชว์ = BK_KP_WEIGHT · เริ่มจากสัปดาห์ปัจจุบันไปข้างหน้า"),
    ("Order Color", "อ่าน/แก้ตารางไฟล์ Order Color", "GET", "/api/order-color/sheet", "ทุก user", P, R_READ,
     "-",
     "1. เปิดตารางข้อมูลดิบ",
     "-",
     "แสดงคอลัมน์/แถวครบตรงกับไฟล์"),
    ("Order Color", "บันทึกการแก้ไข Order Color", "PUT", "/api/order-color/sheet", "ทุก user", P, R_WRITE,
     "สำรองไฟล์ไว้ก่อน",
     "1. แก้ค่าในตาราง\n2. บันทึก\n3. เปิดไฟล์ตรวจ",
     "แก้ 1 เซลล์",
     "ok=true + ชื่อ .bak · ค่าใหม่อยู่ในไฟล์ · ชีทอื่นคงเดิม"),
    ("Order Color", "ดาวน์โหลดไฟล์ Order Color", "GET", "/api/order-color/download", "เปิด", P, R_READ,
     "-",
     "1. กดดาวน์โหลด",
     "-",
     "ได้ไฟล์ datamining_booking_mapped.xlsx ล่าสุด (สะท้อนการแก้ไขล่าสุด)"),
    ("Order Color", "วิเคราะห์งานสี", "POST", "/api/order-color/advise", "ทุก user", P, R_READ,
     "-",
     "1. กดปุ่มวิเคราะห์",
     "-",
     "คืนรายการ item ที่มีสี พร้อมสรุป (summary) · ระบุ lead time ย้อม (LEAD_WEEKS) · ไฟล์ว่าง → มี note บอกเหตุผล ไม่ error"),
    ("Order Color", "Gantt จาก booking", "GET", "/api/order-color/booking-gantt", "ทุก user", P, R_READ,
     "มีไฟล์ booking_final_ready + Order Color",
     "1. เปิดมุมมอง Gantt ของ Order Color",
     "-",
     "แสดงทุก item ต่อสัปดาห์จาก booking DETAIL · งานสีถูกไฮไลต์ · item ที่ stock พออยู่แล้วไม่ถูกนำมาแสดง"),
    ("Order Color", "โควตา setup job ต่อสัปดาห์", "GET", "/api/order-color/booking-load", "ทุก user", P, R_READ,
     "-",
     "1. เปิด Gantt Order Color\n2. ดูแถบจำนวน job ของแต่ละสัปดาห์",
     "-",
     "จำนวน job = งานจาก booking (live) + งานใหม่ที่แผนวาง · เพดานมาจาก REMAINING_JOBS ของแผนล่าสุด (ไม่มีไฟล์แผน → OM 13 / PHET_DOUBLE 33 / PHET_SINGLE 44)"),
    ("Order Color", "เครื่องว่างสอดคล้อง booking", "GET", "/api/order-color/booking-ava", "ทุก user", P, R_READ,
     "-",
     "1. เปิด Gantt Order Color\n2. เทียบตัวเลขเครื่องว่างตอนเพิ่งโหลดหน้า",
     "-",
     "ตอนยังไม่ขยับงาน เครื่องว่าง = remain (planBase = used) · ขยับงานแล้วตัวเลขเปลี่ยนตามจริง"),
    ("Order Color", "แผน what-if แทรกงานสี", "GET", "/api/order-color/plan", "ทุก user", P, R_READ,
     "-",
     "1. เปิดมุมมองเทียบแผน",
     "-",
     "grid = แผนจริง + แถว item สีที่ยังไม่มีในแผน · แถวงานสีถูกไฮไลต์ (color_idx) · ไฟล์แผนจริงไม่ถูกแก้"),
    ("Order Color", "AI แนะนำการดึงงานสีเข้ามา", "POST", "/api/order-color/advise-moves", "ทุก user", P, R_READ,
     "เลือกกลุ่ม CAT × เกจ ที่มีงานสีไม่พอเครื่อง",
     "1. เลือกกลุ่ม\n2. กดให้ AI แนะนำ\n3. อ่านลำดับและเหตุผล",
     "cat/gauge + รายการงานสี",
     "ได้ลำดับความสำคัญ + เหตุผล/ผลกระทบเป็นภาษาไทย · ทุกงานสีที่ไม่พอเครื่องถูกจัดอันดับครบ"),
    ("Order Color", "AI ไม่พร้อมใช้งาน (ไม่มี API key)", "POST", "/api/order-color/advise-moves", "ทุก user", N, R_READ,
     "ไม่ตั้ง OPENAI_API_KEY",
     "1. กดให้ AI แนะนำ",
     "-",
     "ระบบยังตอบผลได้ด้วยตรรกะสำรอง (ai=false) + note บอกเหตุผล · ไม่ค้าง ไม่ error 500"),
    ("Order Color", "กลุ่มที่งานสีพอเครื่องอยู่แล้ว", "POST", "/api/order-color/advise-moves", "ทุก user", N, R_READ,
     "เลือกกลุ่มที่งานสีพอเครื่องทุกตัว",
     "1. กดให้ AI แนะนำ",
     "-",
     "summary = \"งานสีในกลุ่มนี้พอเครื่องทุกตัว — วางได้เลย ไม่ต้องขยับ\" · ranking ว่าง"),
    ("Order Color", "ส่งออกแผน what-if เป็น Excel", "POST", "/api/order-color/plan/export", "ทุก user", P, R_READ,
     "จัดแผน what-if ไว้แล้ว",
     "1. กดส่งออก Excel\n2. เปิดไฟล์ที่ดาวน์โหลด",
     "-",
     "ได้ไฟล์ order_color_plan.xlsx ชีท PLAN_ORDER_COLOR · หัวคอลัมน์และข้อมูลตรงกับที่เห็นบนหน้าจอทุกแถว"),
    ("Order Color", "ยังไม่มีไฟล์ Order Color", "GET", "/api/order-color/sheet", "ทุก user", N, R_READ,
     "ย้ายไฟล์ datamining_booking_mapped.xlsx ออกชั่วคราว",
     "1. เปิดหน้า Order Color",
     "-",
     "HTTP 404 · ข้อความ \"ยังไม่มีไฟล์ Order Color — กรุณากดปุ่มดึงข้อมูลก่อน\" · หน้าเว็บแสดงข้อความแนะนำ ไม่ค้าง"),

    # ---------------- Outsource ----------------
    ("จ้างทอ (Outsource)", "AI แนะนำ item ที่ควรจ้างทอ", "POST", "/api/outsource/advise", "ทุก user", P, R_READ,
     "มีไฟล์แผนล่าสุด",
     "1. เปิดหน้าจ้างทอ\n2. กดให้ AI วิเคราะห์",
     "-",
     "ได้รายการ item ที่คุ้มค่าจะจ้างทอ พร้อมเหตุผล/ตัวเลขประกอบ · อ้างอิงชื่อไฟล์แผนที่ใช้"),
    ("จ้างทอ (Outsource)", "ดูรายการแบ่งจ้างทอที่ใช้อยู่", "GET", "/api/outsource/split", "ทุก user", P, R_READ,
     "-",
     "1. เปิดหน้าจ้างทอ",
     "-",
     "แสดงรายการที่บันทึกไว้ + รายชื่อสัปดาห์ในแผนให้เลือก"),
    ("จ้างทอ (Outsource)", "บันทึกการแบ่งจ้างทอ", "POST", "/api/outsource/split", "ทุก user", P, R_WRITE,
     "รู้จำนวนของค้างของ item ที่จะทดสอบ",
     "1. เลือก item\n2. ใส่จำนวน กก. และสัปดาห์\n3. บันทึก",
     "item จริง, qty น้อยกว่าของค้าง, สัปดาห์ในแผน",
     "ok=true, saved=<item> · รายการแสดงค่าที่บันทึก · ค่าคงอยู่หลังรีเฟรชหน้า"),
    ("จ้างทอ (Outsource)", "ไม่ระบุ item", "POST", "/api/outsource/split", "ทุก user", N, R_READ,
     "-",
     "1. ส่งคำขอโดย item_code ว่าง",
     "item_code = ''",
     "HTTP 400 · ข้อความ \"ไม่ได้ระบุ item\""),
    ("จ้างทอ (Outsource)", "ไม่ระบุสัปดาห์", "POST", "/api/outsource/split", "ทุก user", N, R_READ,
     "-",
     "1. ใส่จำนวนแต่ไม่เลือกสัปดาห์\n2. บันทึก",
     "start_week = null, qty > 0",
     "HTTP 400 · ข้อความ \"ต้องระบุสัปดาห์ที่จ้างทอ\" · ไม่บันทึก"),
    ("จ้างทอ (Outsource)", "จำนวนเกินของค้าง", "POST", "/api/outsource/split", "ทุก user", N, R_READ,
     "รู้ยอดค้างของ item",
     "1. ใส่จำนวนมากกว่ายอดค้าง\n2. บันทึก",
     "qty = ยอดค้าง + 1000",
     "HTTP 400 · ข้อความ \"จ้างทอ ... เกินของค้างของ ...\" พร้อมตัวเลขจริง · ไม่บันทึก"),
    ("จ้างทอ (Outsource)", "ยกเลิกด้วยการใส่จำนวน 0", "POST", "/api/outsource/split", "ทุก user", P, R_WRITE,
     "มีรายการที่บันทึกไว้แล้ว",
     "1. แก้จำนวนเป็น 0\n2. บันทึก",
     "outsource_qty = 0",
     "รายการนั้นถูกยกเลิก/ถอดออกจากการแบ่ง · รายการอื่นไม่กระทบ"),
    ("จ้างทอ (Outsource)", "ลบรายการแบ่งจ้างทอ", "DELETE", "/api/outsource/split/{item_code}", "ทุก user", P, R_WRITE,
     "มีรายการที่บันทึกไว้",
     "1. กดลบรายการ\n2. รีเฟรช",
     "item_code ที่มีอยู่",
     "ok=true · รายการหายไป · รายการอื่นยังอยู่ครบ"),
    ("จ้างทอ (Outsource)", "รหัส item ที่มีเครื่องหมาย /", "DELETE", "/api/outsource/split/{item_code}", "ทุก user", N, R_WRITE,
     "มี item ที่รหัสมี / อยู่ในรายการ",
     "1. กดลบรายการนั้น",
     "item_code = 'AB/CD-01'",
     "ลบได้ถูกต้อง (path param รองรับ /) · ไม่เกิด 404"),
    ("จ้างทอ (Outsource)", "การแบ่งจ้างทอถูกใช้ในการรันแผน", "-", "-", "ทุก user", P, R_HEAVY,
     "บันทึกการแบ่งจ้างทอไว้",
     "1. บันทึกการแบ่ง\n2. สั่งรันแผน\n3. เปิดแผนใหม่ดู item นั้น",
     "-",
     "แผนรอบใหม่หักปริมาณที่จ้างทอออกตามสัปดาห์ที่กำหนด · การแบ่งยังคงอยู่หลังรัน (ไม่หายเอง)"),

    # ---------------- Cylinder ----------------
    ("Change Cylinder", "AI แนะนำการเปลี่ยน cylinder", "POST", "/api/cylinder/advise", "ทุก user", P, R_READ,
     "มีไฟล์แผนล่าสุด",
     "1. เปิดหน้า Change Cylinder\n2. กดวิเคราะห์",
     "-",
     "ได้รายการ item/เครื่องที่ควรเปลี่ยน cylinder พร้อมเหตุผลและสัปดาห์คอขวด · อ้างอิงชื่อไฟล์แผนที่ใช้"),
    ("Change Cylinder", "ยังไม่มีไฟล์แผน", "POST", "/api/cylinder/advise", "ทุก user", N, R_READ,
     "ย้ายไฟล์แผนออกชั่วคราว",
     "1. กดวิเคราะห์",
     "-",
     "HTTP 404 พร้อมข้อความอธิบาย · ไม่ crash"),
    ("Change Cylinder", "AI ไม่พร้อม (ไม่มี API key)", "POST", "/api/cylinder/advise", "ทุก user", N, R_READ,
     "ไม่ตั้ง OPENAI_API_KEY",
     "1. กดวิเคราะห์",
     "-",
     "ยังได้ shortlist จากตรรกะสำรอง + note บอกว่า AI ไม่พร้อม · ไม่ error 500"),
    ("Change Cylinder", "การอนุมัติเปลี่ยน cylinder โดยผู้ใช้", "-", "-", "ทุก user", P, R_WRITE,
     "มีคำแนะนำที่รออนุมัติบน Gantt",
     "1. กดอนุมัติรายการหนึ่ง\n2. ตรวจไฟล์แผน/ชีทที่บันทึกงาน",
     "-",
     "การเปลี่ยน cylinder เกิดขึ้นเมื่อผู้ใช้อนุมัติเท่านั้น (ไม่ทำอัตโนมัติ) · มีการบันทึกรายการพร้อมเวลา"),

    # ---------------- ระบบ / Non-functional ----------------
    ("ระบบ (System)", "เปิดหน้าเว็บหลัก", "GET", "/", "เปิด", P, R_READ,
     "server ทำงานอยู่ · มี frontend/dist",
     "1. เปิด URL ของระบบในเบราว์เซอร์",
     "-",
     "หน้าเว็บโหลดครบ · ไม่มี error ใน console · header ของ index.html เป็น no-cache (ได้ bundle ล่าสุดเสมอ)"),
    ("ระบบ (System)", "SPA routing", "GET", "/{path}", "เปิด", P, R_READ,
     "-",
     "1. เปิด URL ที่ไม่ใช่ไฟล์จริง เช่น /plan\n2. กด F5 ที่หน้านั้น",
     "-",
     "ได้หน้าเว็บเดิม (index.html) ไม่ขึ้น 404 ของ server"),
    ("ระบบ (System)", "ไฟล์ static โหลดได้", "GET", "/assets/*", "เปิด", P, R_READ,
     "-",
     "1. เปิด DevTools → Network\n2. รีโหลดหน้า",
     "-",
     "ไฟล์ js/css ใน /assets โหลดสำเร็จทั้งหมด (200)"),
    ("ระบบ (System)", "หลาย user ใช้งานพร้อมกัน", "-", "-", "ทุก user", F, R_HEAVY,
     "มี 2 บัญชีขึ้นไป",
     "1. เปิด 2 เบราว์เซอร์ login คนละบัญชี\n2. คนแรกสั่งรัน\n3. คนที่สองดูสถานะ/สั่งรัน",
     "-",
     "คนที่สองเห็นสถานะและ log ของงานที่รันอยู่ · สั่งรันซ้อนไม่ได้ (มีข้อความแจ้ง) · ข้อมูลไม่ปนกัน"),
    ("ระบบ (System)", "เวลาตอบสนองหน้าหลักๆ", "-", "-", "ทุก user", F, R_READ,
     "ข้อมูลจริงขนาดใช้งานจริง",
     "1. จับเวลาโหลดหน้าแผนผลิต / Order Color / DATA",
     "-",
     "แต่ละหน้าแสดงผลภายในเวลาที่ผู้ใช้ยอมรับได้ (แนะนำ ≤ 10 วินาที) · ระหว่างรอมีสถานะกำลังโหลด"),
    ("ระบบ (System)", "restart server แล้วใช้งานต่อได้", "-", "-", "admin", F, R_HEAVY,
     "-",
     "1. restart server\n2. login\n3. เปิดทุกหน้าหลัก",
     "-",
     "ทุกหน้าใช้งานได้ · ตารางเวลาอัตโนมัติถูกตั้งใหม่เอง · ไฟล์ข้อมูล/ผลลัพธ์ยังครบ · ต้อง login ใหม่ (token เดิมยังไม่หมดอายุก็ยังใช้ได้)"),
    ("ระบบ (System)", "ข้อความผิดพลาดเป็นภาษาไทยเข้าใจง่าย", "-", "-", "ทุก user", F, R_READ,
     "-",
     "1. ทำให้เกิด error ต่างๆ (ไฟล์หาย, ไม่มีสิทธิ์, token หมดอายุ)",
     "-",
     "ทุกกรณีแสดงข้อความภาษาไทยที่ผู้ใช้เข้าใจได้ ไม่โชว์ stack trace/ข้อความระบบดิบ"),
]


# =====================================================================
# 3) สถานการณ์ใช้งานจริง (Scenario / End-to-End) — ผู้ใช้ทำงานบนหน้าจอจริง
#    (กลุ่ม, สถานการณ์, หน้าจอ, ผู้ใช้, เงื่อนไขก่อนเริ่ม, ขั้นตอน, จุดตรวจ, ผลที่คาดหวัง)
# =====================================================================
SCENARIOS = [
    # ---------- A. ปรับแผนบน Gantt ----------
    ("ปรับแผนผลิต", "ลากงานย้ายสัปดาห์", "แผนผลิต → Gantt", "ผู้วางแผน",
     "มีไฟล์แผนล่าสุด · เปิดชีท PLAN · เปิด Gantt",
     "1. หางานที่ต้องการเลื่อน\n2. ลากบล็อกงานไปคอลัมน์สัปดาห์อื่น (เครื่องเดิม)\n3. ปล่อยเมาส์\n4. ดูตารางด้านล่างแถวเดียวกัน",
     "• PLAN_WEEK เปลี่ยนเป็นสัปดาห์ปลายทาง\n• CALENDAR_WORKING_DAYS / ACTUAL_WORKING_DAYS / AVAILABLE_DAYS คำนวณใหม่ตามสัปดาห์ปลายทาง\n• แถบโหลดของสัปดาห์ต้นทางลดลง ปลายทางเพิ่มขึ้น\n• เครื่องว่าง (มุมขวาบนของช่อง) เปลี่ยนทั้ง 2 สัปดาห์\n• หัวข้อ \"แผนผลิต\" ขึ้นจุด ● (ยังไม่บันทึก)",
     "งานย้ายไปสัปดาห์ใหม่ ตัวเลขวันทำงาน/โหลด/เครื่องว่างถูกต้องทั้งสองฝั่ง และระบบเตือนว่ายังไม่บันทึก"),
    ("ปรับแผนผลิต", "ย้ายงานข้ามเครื่อง (cross MC)", "แผนผลิต → Gantt", "ผู้วางแผน",
     "-",
     "1. ลากบล็อกงานไปวางในแถวเครื่องอื่น (เช่น FA → SKP)\n2. อ่านกล่องยืนยัน\n3. ทดสอบทั้งกด \"ตกลง\" และ \"ยกเลิก\"",
     "• ต้องมีกล่องยืนยัน \"ยืนยันย้ายงาน <ITEM> จากเครื่อง <A> → <B> (สัปดาห์ X)?\"\n• กดยกเลิก = งานอยู่ที่เดิมทุกค่า\n• กดตกลง = MC_GROUP และ PLAN_WEEK เปลี่ยนพร้อมกัน",
     "ระบบไม่ย้ายข้ามเครื่องเองโดยไม่ถาม · ยกเลิกแล้วไม่มีอะไรเปลี่ยน"),
    ("ปรับแผนผลิต", "ย้ายไปสัปดาห์ที่วันทำงานน้อยกว่า", "แผนผลิต → Gantt", "ผู้วางแผน",
     "มีสัปดาห์ที่วันทำงานน้อย (วันหยุดยาว/ยุบสัปดาห์)",
     "1. ลากงานก้อนใหญ่ไปสัปดาห์ที่วันทำงานน้อย\n2. อ่านข้อความเตือน",
     "• เตือนว่า \"เครื่องที่ถืออยู่ (carry X + ใหม่ Y) ผลิตได้แค่ N กก. แต่ต้องผลิต M กก.\"\n• บอกให้รันแผนใหม่เพื่อคำนวณเครื่อง\n• ตัวเลขในข้อความตรงกับข้อมูลจริงของแถวนั้น",
     "ผู้ใช้ได้รับคำเตือนก่อนตัดสินใจ ไม่ปล่อยให้วางงานเกินกำลังผลิตแบบเงียบๆ"),
    ("ปรับแผนผลิต", "แก้จำนวนผลิตบนบล็อก", "แผนผลิต → Gantt", "ผู้วางแผน",
     "-",
     "1. double click ที่ตัวเลขบนบล็อกงาน\n2. พิมพ์จำนวนใหม่ (น้อยกว่ายอดสั่ง)\n3. ยืนยัน\n4. ดูตารางด้านล่าง",
     "• PRODUCE_QTY ในตารางเปลี่ยนตาม\n• ขนาด/ข้อความบนบล็อกอัปเดต\n• แถบโหลดของสัปดาห์นั้นเปลี่ยนตามสัดส่วน",
     "แก้จำนวนได้จากบล็อกโดยตรง และค่าไปตรงกับตารางข้อมูล"),
    ("ปรับแผนผลิต", "แก้จำนวนเกินยอดสั่ง", "แผนผลิต → Gantt", "ผู้วางแผน",
     "รู้ค่า ORDERS_QTY ของออร์เดอร์ที่จะทดสอบ",
     "1. double click แก้จำนวนให้มากกว่ายอดสั่งที่เหลือ\n2. อ่านข้อความ",
     "• ขึ้นข้อความ \"วางเกินยอดสั่ง (ORDERS_QTY) ไม่ได้ — ออร์เดอร์นี้วางแถวนี้ได้สูงสุด N กก.\"\n• ค่าไม่ถูกบันทึก (หรือถูกปรับลงมาเท่าที่วางได้พร้อมแจ้งเตือน)\n• ยอดสูงสุด N = ORDERS_QTY − ยอดที่วางไว้ในแถวอื่นของออร์เดอร์เดียวกัน",
     "ระบบกันไม่ให้วางเกินยอดสั่งของลูกค้า"),
    ("ปรับแผนผลิต", "แบ่งงานออกเป็น 2 สัปดาห์", "แผนผลิต → Gantt", "ผู้วางแผน",
     "เลือกงานที่จำนวนมากพอจะแบ่ง",
     "1. คลิกบล็อกงาน → เปิดกล่องรายละเอียด\n2. ใส่จำนวนที่ต้องการแบ่ง + เลือกสัปดาห์ปลายทาง\n3. กด \"แบ่ง\"\n4. อ่านคำเตือนแล้วยืนยัน",
     "• เตือนว่า \"จำนวนเครื่อง/setup จะ copy มาตามเดิม ไม่คำนวณใหม่ — ต้องรันแผนใหม่\"\n• เกิดแถวใหม่ต่อจากแถวเดิม\n• จำนวนแถวเดิม + แถวใหม่ = จำนวนเดิมก่อนแบ่ง (ไม่หาย ไม่งอก)\n• แถวใหม่อยู่สัปดาห์ปลายทาง และคอลัมน์วันทำงานคำนวณใหม่ถูกต้อง",
     "แบ่งงานได้ถูกต้อง ยอดรวมไม่เพี้ยน และผู้ใช้รู้ว่าต้องรันแผนใหม่เพื่อได้เลขเครื่องจริง"),
    ("ปรับแผนผลิต", "แบ่งงานด้วยจำนวนที่ไม่ถูกต้อง", "แผนผลิต → Gantt", "ผู้วางแผน",
     "-",
     "1. ลองแบ่งด้วยจำนวน 0\n2. ลองแบ่งด้วยจำนวนเท่ากับ/มากกว่าก้อนเดิม",
     "• ขึ้นข้อความ \"แบ่งไม่ได้: จำนวนที่แบ่งต้องมากกว่า 0 และน้อยกว่าจำนวนก้อนเดิม\"\n• ไม่มีแถวใหม่เกิดขึ้น",
     "ระบบปฏิเสธการแบ่งที่ไม่สมเหตุสมผล"),
    ("ปรับแผนผลิต", "ลบงานออกจากแผน", "แผนผลิต → Gantt", "ผู้วางแผน",
     "เตรียมชื่อ item ที่จะลบไว้ตรวจสอบ",
     "1. กดปุ่มลบบนบล็อกงาน\n2. อ่านกล่องยืนยัน\n3. ยืนยัน\n4. ตรวจตารางด้านล่าง",
     "• กล่องยืนยันระบุ item และสัปดาห์ที่จะลบ + บอกว่า \"ลบก้อนนี้ก้อนเดียว\"\n• ลบแล้วเฉพาะก้อนนั้นหาย ก้อนอื่นของ item เดียวกันยังอยู่\n• ต้องกดบันทึกจึงจะมีผลถาวร (ยังไม่กดบันทึก → รีเฟรชแล้วงานกลับมา)",
     "ลบได้ตรงก้อน และยังไม่กระทบไฟล์จริงจนกว่าจะบันทึก"),
    ("ปรับแผนผลิต", "ดูรายละเอียดงาน", "แผนผลิต → Gantt", "ผู้วางแผน",
     "-",
     "1. คลิกบล็อกงาน 1 ครั้ง\n2. อ่านข้อมูลในกล่อง\n3. กด Esc หรือ ✕ ปิด",
     "• แสดงค่าครบทุกคอลัมน์ของแถวนั้น (ITEM_CODE, SC_SO_NO, ORDERS_QTY, PRODUCE_QTY, NEW_MC, CARRYOVER_MC, SETUP_DAYS ฯลฯ)\n• ค่าที่แสดงตรงกับตารางด้านล่างแถวเดียวกัน",
     "ผู้ใช้ตรวจข้อมูลงานได้ครบโดยไม่ต้องเลื่อนหาในตาราง"),
    ("ปรับแผนผลิต", "บันทึกทั้งที่ยอดไม่ตรงยอดสั่ง", "แผนผลิต", "ผู้วางแผน",
     "แก้จำนวนงานให้รวมแล้วไม่เท่า ORDERS_QTY",
     "1. ลดจำนวนของงานหนึ่งให้ยอดรวมของ SC นั้นขาด\n2. กดบันทึก\n3. อ่านกล่องเตือน",
     "• เตือน \"⚠ มี N ออร์เดอร์ที่วางไม่ตรงยอดสั่ง (ORDERS_QTY)\" พร้อมรายการ SC / สั่ง / วาง / ขาด-เกิน\n• กดยกเลิก = ไม่บันทึก\n• กดยืนยัน = บันทึกได้ (ให้ผู้ใช้ตัดสินใจเอง)",
     "ระบบเตือนยอดไม่ครบก่อนบันทึกเสมอ พร้อมบอกว่า SC ไหนขาดเท่าไร"),
    ("ปรับแผนผลิต", "บันทึกแผนสำเร็จและตรวจไฟล์", "แผนผลิต", "ผู้วางแผน",
     "สำรองไฟล์แผนไว้ก่อน",
     "1. ปรับแผน 2-3 จุด\n2. กดบันทึก\n3. อ่านข้อความผลลัพธ์\n4. กดดาวน์โหลดไฟล์แผน แล้วเปิดด้วย Excel",
     "• ข้อความ \"บันทึกแล้ว (N แถว) — สำรองไฟล์เดิมเป็น <ชื่อ .bak>\"\n• จุด ● หายไป\n• ไฟล์ที่ดาวน์โหลดมีค่าที่เพิ่งแก้ทุกจุด (ไม่ใช่ไฟล์เก่า)\n• ชีทอื่นในไฟล์ยังครบและไม่เปลี่ยน",
     "การแก้ไขจากหน้าเว็บลงไฟล์จริงครบถ้วน และมีไฟล์สำรองให้ย้อนได้"),
    ("ปรับแผนผลิต", "เปลี่ยนชีทขณะยังไม่บันทึก", "แผนผลิต", "ผู้วางแผน",
     "แก้ค่าค้างไว้โดยยังไม่บันทึก",
     "1. แก้ค่า 1 จุด\n2. เลือกชีทอื่นจาก dropdown\n3. อ่านกล่องเตือน",
     "• เตือน \"มีการแก้ไขที่ยังไม่บันทึก จะทิ้งแล้วเปลี่ยนชีทไหม?\"\n• กดยกเลิก = อยู่ชีทเดิม ค่าที่แก้ยังอยู่",
     "ระบบกันไม่ให้ผู้ใช้ทำงานที่แก้ไว้หายโดยไม่รู้ตัว"),
    ("ปรับแผนผลิต", "ค้นหาและกรองงานในแผน", "แผนผลิต", "ผู้วางแผน",
     "-",
     "1. พิมพ์รหัส item ในช่องค้นหา\n2. กรองด้วย CAT / Gauge / ประเภทงาน\n3. กด ▾ ที่หัวคอลัมน์เพื่อกรองเฉพาะค่า\n4. กด \"ล้างตัวกรองทั้งหมด\"",
     "• ตารางและ Gantt แสดงเฉพาะแถวที่ตรงเงื่อนไข\n• จำนวนแถวที่แสดงถูกต้อง\n• ล้างตัวกรองแล้วกลับมาครบเหมือนเดิม",
     "ผู้ใช้หางานที่ต้องการเจอเร็ว และตัวกรองไม่ทำให้ข้อมูลหาย"),
    ("ปรับแผนผลิต", "เทียบกับแผนเดิมจาก booking (overlay)", "แผนผลิต → Gantt", "ผู้วางแผน",
     "-",
     "1. เปิดโหมด overlay \"Item (ทั้งหมด)\"\n2. สลับเป็น \"Item (วางแผนวันนี้)\"\n3. เลือกเฉพาะบาง item จากรายการ\n4. ลองลากบล็อกของแผนเดิม",
     "• บล็อกแผนเดิมแสดงแยกสีจากแผนใหม่ (ดูคำอธิบายสีในกล่องช่วยเหลือ)\n• ลาก/แก้บล็อกแผนเดิมไม่ได้ (อ่านอย่างเดียว)\n• เลือก/ล้าง item ได้ตามต้องการ",
     "ผู้ใช้เทียบแผนใหม่กับของเดิมได้ โดยไม่เผลอแก้ข้อมูลเก่า"),
    ("ปรับแผนผลิต", "อ่านเครื่องว่างและเครื่องที่กันไว้", "แผนผลิต → Gantt", "ผู้วางแผน",
     "-",
     "1. ดูตัวเลขเครื่องว่างมุมขวาบนของแต่ละช่อง\n2. ชี้เมาส์ที่ตัวเลขเครื่องกันไว้ POLY/COTTON\n3. เทียบกับชีท SUMMARY_MC_REMAIN",
     "• เครื่องว่างตรงกับ SUMMARY_MC_REMAIN ของ CAT|เกจ|สัปดาห์นั้น\n• tooltip อธิบายว่าเป็นเครื่องกันไว้ให้งาน POLY/COTTON ใช้แทนงานปกติไม่ได้\n• ช่องที่เต็มแสดงสถานะ \"เกิน N\" ชัดเจน",
     "ตัวเลขเครื่องบนหน้าจอเชื่อถือได้ ตรงกับข้อมูลในไฟล์แผน"),
    ("ปรับแผนผลิต", "สั่งรันแผนใหม่จากหน้าแผน", "แผนผลิต", "ผู้วางแผน",
     "ไม่มีงานอื่นกำลังรัน",
     "1. กด \"▶ รันแผนใหม่\"\n2. ดูความคืบหน้า %\n3. รอจนจบ",
     "• ปุ่มถูกปิดระหว่างรัน (กดซ้ำไม่ได้)\n• แสดงความคืบหน้าเป็น %\n• เมื่อจบระบบโหลดแผนใหม่ให้อัตโนมัติ + เวลา \"แผนล่าสุด\" อัปเดต",
     "ผู้ใช้รันแผนใหม่ได้จากหน้าเดียว และเห็นผลทันทีโดยไม่ต้องรีเฟรชเอง"),

    # ---------- B. ปรับสัปดาห์ / วันทำงาน ----------
    ("ปรับสัปดาห์ / วันทำงาน", "ตั้งวันทำงานมาตรฐานของกลุ่มเดียว", "แผนผลิต → แผงวันทำงาน", "ผู้วางแผน",
     "สำรอง Calendar.xlsx ไว้",
     "1. เลือกกลุ่มจาก tree ด้านซ้าย (Factory → MC_CAT → เกจ)\n2. กรอกช่อง \"ค่ามาตรฐานของกลุ่มนี้\" เป็น 5\n3. กด \"บันทึกวันทำงาน\"",
     "• ชื่อกลุ่มด้านขวาแสดงถูกต้อง (Factory · MC_CAT · Gเกจ)\n• ขึ้นแถบ \"มีการเปลี่ยนแปลงที่ยังไม่บันทึก\" ก่อนกดบันทึก\n• ข้อความ \"บันทึกแล้ว (N แถว, ยุบ M สัปดาห์) — สำรองไฟล์เดิมเป็น ...\"\n• รายการฝั่งซ้ายของกลุ่มนั้นแสดง \"5 วัน\"",
     "ตั้งวันทำงานรายกลุ่มได้ และค่าคงอยู่หลังโหลดหน้าใหม่"),
    ("ปรับสัปดาห์ / วันทำงาน", "ตั้งค่าหลายกลุ่มพร้อมกัน", "แผนผลิต → แผงวันทำงาน", "ผู้วางแผน",
     "-",
     "1. ติ๊ก checkbox ที่ระดับ Factory หรือ MC_CAT (เลือกทั้งกลุ่มย่อย)\n2. สังเกตหัวข้อ \"เลือก N กลุ่ม\"\n3. กรอกค่ามาตรฐาน 1 ครั้ง\n4. บันทึก แล้วไล่ตรวจทีละกลุ่ม",
     "• checkbox แสดง 3 สถานะถูกต้อง (เลือกครบ / บางส่วน / ไม่เลือก)\n• กรอกครั้งเดียวค่าไปลงทุกกลุ่มที่เลือก\n• กลุ่มที่ไม่ได้เลือกไม่เปลี่ยน\n• ถ้าค่าเดิมไม่ตรงกันช่องกรอกแสดง \"หลายค่า\"",
     "แก้ค่าทีละหลายกลุ่มได้ถูกต้อง ไม่กระทบกลุ่มที่ไม่ได้เลือก"),
    ("ปรับสัปดาห์ / วันทำงาน", "ค้นหากลุ่มเครื่อง", "แผนผลิต → แผงวันทำงาน", "ผู้วางแผน",
     "-",
     "1. พิมพ์คำค้นแบบหลายคำ เช่น \"phet double-30\"\n2. ดูรายการที่เหลือ\n3. ล้างคำค้น",
     "• แสดงเฉพาะกลุ่มที่ตรงทุกคำที่พิมพ์\n• ระหว่างค้นหา หัวข้อทั้งหมดถูกกางอัตโนมัติ\n• ไม่พบ → แสดง \"ไม่พบกลุ่ม\"",
     "ค้นหากลุ่มเครื่องได้เร็วแม้มีหลายร้อยกลุ่ม"),
    ("ปรับสัปดาห์ / วันทำงาน", "ตั้งวันทำงานเฉพาะสัปดาห์", "แผนผลิต → แผงวันทำงาน", "ผู้วางแผน",
     "-",
     "1. เลือกกลุ่ม\n2. ในตารางสัปดาห์ พิมพ์ค่าเฉพาะที่ช่อง W33 เป็น 4\n3. บันทึก แล้วโหลดหน้าใหม่",
     "• ช่อง W33 เปลี่ยนสีเป็นสถานะ \"แก้แล้ว\"\n• สัปดาห์อื่นยังใช้ค่ามาตรฐาน (แสดงเป็น placeholder)\n• หลังบันทึกและโหลดใหม่ ค่า 4 ยังอยู่ที่ W33\n• รายการกลุ่มฝั่งซ้ายมีสัญลักษณ์ ⚙ บอกว่ามีค่ารายสัปดาห์",
     "ตั้งวันทำงานเฉพาะสัปดาห์ได้โดยไม่กระทบสัปดาห์อื่น"),
    ("ปรับสัปดาห์ / วันทำงาน", "ตั้งชั่วโมงทำงานต่อวัน", "แผนผลิต → แผงวันทำงาน", "ผู้วางแผน",
     "-",
     "1. เลือกกลุ่ม\n2. ตั้งช่อง \"ชั่วโมง/วัน\" เป็น 20\n3. บันทึก\n4. รันแผนใหม่ แล้วเทียบกำลังผลิตของกลุ่มนั้น",
     "• รายการฝั่งซ้ายแสดง \"... วัน · 20 ชม.\"\n• หลังรันแผน กำลังผลิตของกลุ่มลดลงประมาณ 20/24 เท่าของเดิม\n• กลุ่มที่ไม่ได้ตั้งยังใช้ 24 ชม. (ไม่ลดกำลังผลิต)",
     "การตั้งชั่วโมงมีผลจริงกับกำลังผลิตตามกฎ 20/24"),
    ("ปรับสัปดาห์ / วันทำงาน", "ยุบสัปดาห์ที่วันหยุดเยอะ", "แผนผลิต → แผงวันทำงาน", "ผู้วางแผน",
     "รู้ว่าสัปดาห์ไหนวันหยุดเยอะ (เช่น สงกรานต์/ปีใหม่)",
     "1. ที่ปฏิทินรวมด้านบน กดปุ่ม ▶ ที่ W31 เพื่อยุบเข้า W32\n2. ดูสถานะของทั้งสองช่อง\n3. เลือกกลุ่มแล้วดูตารางสัปดาห์\n4. บันทึก",
     "• W31 แสดงป้าย \"→W32\" · W32 แสดงป้าย \"◀รวม\" พร้อมยอดวันรวม\n• ช่องกรอกของ W31 ถูกปิด (กรอกไม่ได้)\n• ค่ารายสัปดาห์เดิมของ W31 ถูกล้างอัตโนมัติ\n• ยอดของ W32 = วันของ W32 + วันของ W31",
     "ยุบสัปดาห์ได้ครบทั้งระบบ และวันทำงานไปรวมที่สัปดาห์ปลายทางถูกต้อง"),
    ("ปรับสัปดาห์ / วันทำงาน", "แก้ยอดรวมของสัปดาห์ที่รับยุบ", "แผนผลิต → แผงวันทำงาน", "ผู้วางแผน",
     "ยุบ W31 → W32 ไว้แล้ว",
     "1. เลือกกลุ่ม\n2. พิมพ์ตัวเลขในช่อง W32 (เช่น 9)\n3. ลบค่าที่พิมพ์ออก",
     "• พิมพ์ค่า = ระบบใช้ค่าที่พิมพ์แทนผลรวมอัตโนมัติ\n• ลบค่าออก = กลับไปใช้ผลรวมอัตโนมัติ (แสดงเป็น placeholder)",
     "ผู้ใช้ควบคุมยอดวันทำงานของก้อนที่ยุบได้เอง"),
    ("ปรับสัปดาห์ / วันทำงาน", "ยกเลิกการยุบสัปดาห์", "แผนผลิต → แผงวันทำงาน", "ผู้วางแผน",
     "มีการยุบสัปดาห์อยู่",
     "1. กดปุ่มทิศทางเดิมซ้ำอีกครั้งที่สัปดาห์ที่ยุบไว้\n2. บันทึก",
     "• ป้าย →W และ ◀รวม หายไป\n• ช่องกรอกของสัปดาห์นั้นใช้งานได้อีกครั้ง\n• ชีท Week Merge ในไฟล์ไม่มีแถวนั้นแล้ว",
     "ยกเลิกการยุบได้ กลับสู่สภาพเดิมสมบูรณ์"),
    ("ปรับสัปดาห์ / วันทำงาน", "ยกเลิกการแก้ไขที่ยังไม่บันทึก", "แผนผลิต → แผงวันทำงาน", "ผู้วางแผน",
     "-",
     "1. แก้ค่าหลายจุด (วัน/ชั่วโมง/ยุบสัปดาห์)\n2. กดปุ่ม \"ยกเลิก\"",
     "• ทุกค่ากลับไปเป็นค่าที่บันทึกไว้ล่าสุด\n• แถบ \"มีการเปลี่ยนแปลงที่ยังไม่บันทึก\" หายไป\n• ไฟล์ Calendar.xlsx ไม่ถูกแตะ (ไม่มี .bak ใหม่)",
     "ผู้ใช้ทิ้งการแก้ไขได้ทันทีโดยไม่กระทบไฟล์"),
    ("ปรับสัปดาห์ / วันทำงาน", "ตั้งค่าเริ่มต้นจาก MasterMC", "แผนผลิต → แผงวันทำงาน", "ผู้วางแผน",
     "-",
     "1. กด \"↻ ตั้งค่าเริ่มต้นจาก MasterMC\"\n2. อ่านกล่องยืนยันแล้วตกลง\n3. กดซ้ำอีกครั้ง",
     "• กล่องยืนยันระบุว่า \"ไม่ทับค่าที่ตั้งไว้แล้ว\"\n• ครั้งแรก: ข้อความ \"เพิ่มค่าเริ่มต้น N กลุ่มจาก MasterMC\"\n• ครั้งที่สอง: \"มีค่าครบทุกกลุ่มแล้ว\"\n• ค่าที่ผู้ใช้เคยตั้งเองไม่ถูกทับ",
     "seed ใช้ได้ปลอดภัย กดซ้ำไม่ทำให้ค่าที่ตั้งไว้หาย"),
    ("ปรับสัปดาห์ / วันทำงาน", "ปรับสัปดาห์แล้วกลับไปดูผลที่แผน", "แผนผลิต", "ผู้วางแผน",
     "ปรับวันทำงาน/ยุบสัปดาห์ไว้แล้ว",
     "1. บันทึกแผงวันทำงาน\n2. กลับไปหน้าแผนผลิต รีเฟรช\n3. ลากงานเข้าสัปดาห์ที่ปรับ\n4. ดูคอลัมน์ ACTUAL_WORKING_DAYS",
     "• ค่าวันทำงานที่หน้าแผนใช้ = ค่าที่เพิ่งตั้ง (ไม่ใช่ค่าเก่า)\n• สัปดาห์ที่ยุบไม่มีงานวางได้ / งานไปรวมที่ปลายทาง\n• ตรงกับที่ backend คำนวณตอนรันแผน",
     "หน้าจอแผนกับการคำนวณจริงใช้วันทำงานชุดเดียวกัน ไม่ขัดกัน"),
    ("ปรับสัปดาห์ / วันทำงาน", "วันทำงานต้องไม่เกินวันที่ปฏิทินเปิด", "แผนผลิต → แผงวันทำงาน", "ผู้วางแผน",
     "รู้ว่าสัปดาห์ทดสอบปฏิทินเปิดกี่วัน (เช่น 5 วัน)",
     "1. ตั้งวันทำงานของกลุ่มเป็น 7\n2. ดูวันทำงานที่ระบบใช้จริงในสัปดาห์นั้น\n3. รันแผนแล้วเทียบ ACTUAL_WORKING_DAYS",
     "• ระบบใช้ 5 วัน (ค่าน้อยกว่าระหว่างแผงวันทำงานกับปฏิทิน)\n• ไม่ผลิตเกินวันที่โรงงานเปิดจริง",
     "ปฏิทินวันหยุดเป็นเพดานเสมอ ป้องกันการวางแผนเกินวันทำงานจริง"),

    # ---------- C. ตรวจสอบความถูกต้องของข้อมูล ----------
    ("ตรวจสอบข้อมูล", "เทียบ Booking บนเว็บกับไฟล์จริง", "DATA / History", "ผู้วางแผน",
     "รันดึงข้อมูลรอบล่าสุดแล้ว",
     "1. เปิดหน้า DATA → กลุ่ม Booking → เลือกไฟล์ล่าสุด\n2. สุ่ม 5 item จดยอดไว้\n3. ดาวน์โหลดไฟล์เดียวกันมาเปิดใน Excel\n4. เทียบทีละ item",
     "• ยอดบนหน้าเว็บตรงกับไฟล์ทุกตัว\n• เวลาโหลดล่าสุดที่แสดงตรงกับเวลาไฟล์จริง\n• ไฟล์เกิน 5,000 แถวมีข้อความบอกให้ดาวน์โหลดดูครบ",
     "ข้อมูลที่ผู้ใช้เห็นบนเว็บเชื่อถือได้ ตรงกับไฟล์ที่ระบบใช้คำนวณ"),
    ("ตรวจสอบข้อมูล", "เทียบ Stock กับต้นทาง", "DATA", "ผู้วางแผน",
     "รัน \"ดึงข้อมูลล่าสุด\" แล้ว",
     "1. เปิดหน้า DATA → กลุ่ม Stock\n2. สุ่มตรวจ 5 item เทียบกับระบบ Oracle/รายงานที่ใช้อยู่เดิม",
     "• ยอดคงเหลือตรงกับต้นทาง\n• เวลาที่ดึงข้อมูลตรงกับรอบที่รัน",
     "ยอด Stock ที่ใช้วางแผนตรงกับความจริงหน้างาน"),
    ("ตรวจสอบข้อมูล", "ตรวจจำนวนเครื่องไม่ถูกนับซ้ำ", "แผนผลิต → SUMMARY_MC_REMAIN", "ผู้วางแผน",
     "มีแผนล่าสุด",
     "1. เปิดชีท SUMMARY_MC_REMAIN\n2. เลือก CAT|เกจ 1 กลุ่ม 1 สัปดาห์\n3. บวก เครื่องที่ใช้ + เครื่องคงเหลือ\n4. เทียบกับจำนวนเครื่องทั้งหมดของกลุ่มใน MasterMC",
     "• ผลบวกเท่ากับจำนวนเครื่องจริงของกลุ่ม\n• เครื่องในพูลที่แยกกัน (เช่น SKP กับ SKPTA/SKPLE) ไม่ถูกนับข้ามพูล\n• เครื่อง booking ไม่ถูกนับซ้ำกับเครื่องที่แผนจอง",
     "การนับเครื่องถูกต้อง ไม่เกินจำนวนเครื่องที่มีจริง"),
    ("ตรวจสอบข้อมูล", "ตรวจเครื่องที่กันไว้ POLY / COTTON", "แผนผลิต → Gantt", "ผู้วางแผน",
     "MasterMC ตั้งจำนวน POLY/COTTON ไว้",
     "1. ดูจำนวนเครื่องที่กันไว้ของ CAT|เกจ ที่ทดสอบ\n2. เทียบกับค่าใน MasterMC\n3. ตรวจสัปดาห์ที่ booking ใช้เครื่องกันไว้ไปแล้ว",
     "• เครื่องกันไว้ที่เหลือ = ค่าตั้งต้น − ที่ booking ใช้ไปแล้วในสัปดาห์นั้น\n• งานปกติเอาเครื่องกันไว้ไปใช้ไม่ได้",
     "เครื่องที่กันไว้ถูกหักตามการใช้งานจริงรายสัปดาห์"),
    ("ตรวจสอบข้อมูล", "ตรวจจำนวนวัน setup", "แผนผลิต → ชีท PLAN", "ผู้วางแผน",
     "มีแผนล่าสุด",
     "1. หา item ที่เป็น COTTON บริสุทธิ์ → ดู SETUP_DAYS\n2. หา item ที่เป็น POLY / CD / TC หรือมีหลายเส้น (มี +) → ดู SETUP_DAYS",
     "• COTTON บริสุทธิ์ = 3 วัน\n• POLY / CD / TC / ผสมหลายเส้น = 5 วัน",
     "จำนวนวัน setup ตรงตามกฎที่โรงงานใช้จริง"),
    ("ตรวจสอบข้อมูล", "ตรวจงานที่รันต่อเนื่องไม่ setup ซ้ำ", "แผนผลิต → ชีท PLAN", "ผู้วางแผน",
     "มี item ที่ถักต่อเนื่องหลายสัปดาห์บนเครื่องเดิม",
     "1. หา item เดียวกัน เครื่องเดียวกัน ที่ถักติดกัน (ห่างไม่เกิน 3 สัปดาห์)\n2. ดู CARRYOVER_MC และ SETUP_DAYS ของสัปดาห์ถัดมา",
     "• สัปดาห์ต่อเนื่องมี CARRYOVER_MC > 0 และไม่คิด setup ซ้ำ\n• ถ้าเว้นเกิน 3 สัปดาห์ ต้องคิด setup ใหม่",
     "ระบบไม่คิดเวลา setup เกินจริง ทำให้กำลังผลิตไม่ต่ำกว่าความเป็นจริง"),
    ("ตรวจสอบข้อมูล", "ตรวจนิยามสัปดาห์ (ศุกร์–พฤหัส)", "Master Data → ปฏิทิน", "ผู้วางแผน",
     "-",
     "1. เปิดชีทปฏิทินในหน้า Master Data\n2. เลือกสัปดาห์หนึ่ง ดูวันแรกและวันสุดท้าย",
     "• สัปดาห์เริ่มวันศุกร์ สิ้นสุดวันพฤหัสบดี\n• เลขสัปดาห์ตรงกับที่แสดงในหน้าแผนและ Gantt",
     "การอ้างอิงสัปดาห์ทั้งระบบตรงกับปฏิทินโรงงาน"),
    ("ตรวจสอบข้อมูล", "ตรวจยอดผลิตรวมเท่ากับยอดสั่ง", "แผนผลิต → ชีท PLAN", "ผู้วางแผน",
     "-",
     "1. เลือก SC 1 ใบที่มีหลายแถว\n2. รวม PRODUCE_QTY ทุกแถวของ SC นั้น\n3. เทียบกับ ORDERS_QTY",
     "• ผลรวมเท่ากับยอดสั่ง (คลาดเคลื่อนไม่เกินค่าที่ระบบยอมรับ)\n• ถ้าไม่เท่า ระบบต้องเตือนตอนกดบันทึก",
     "ทุกออร์เดอร์ถูกวางแผนครบตามยอดที่ลูกค้าสั่ง"),
    ("ตรวจสอบข้อมูล", "ตรวจงานสีถักทันย้อม", "Order Color", "ผู้วางแผน",
     "มีข้อมูล Order Color ล่าสุด",
     "1. เปิดผลวิเคราะห์แผนสี\n2. สุ่ม item ที่สถานะ \"ทันย้อม\" 1 ตัว\n3. เทียบสัปดาห์ที่ถักในแผน กับสัปดาห์ย้อม − lead time",
     "• สัปดาห์ที่ถักเสร็จ ≤ สัปดาห์ย้อม − lead weeks ที่ระบบระบุ\n• item ที่สถานะ \"ช้าเกิน\" ต้องผิดเงื่อนไขนี้จริง",
     "การจัดกลุ่มสถานะงานสีถูกต้อง ใช้ตัดสินใจได้"),
    ("ตรวจสอบข้อมูล", "ตรวจการ map item (Datamining ↔ Booking)", "Map Item", "ผู้วางแผน",
     "รัน Map Item แล้ว",
     "1. เปิด \"Datamining → ORA Item\"\n2. สุ่ม 5 แถว เทียบกับ Master_Item\n3. เปิด \"Datamining → Booking\" ตรวจว่าเชื่อมสัปดาห์ถูก",
     "• รหัสที่ map ตรงกับ Master_Item\n• item ที่ map ไม่ได้ต้องแสดงให้เห็น ไม่ถูกซ่อนเงียบๆ",
     "การเชื่อมข้อมูลระหว่างระบบถูกต้อง ไม่มี item ตกหล่นแบบไม่รู้ตัว"),
    ("ตรวจสอบข้อมูล", "ตรวจการจ้างทอถูกหักออกจากแผน", "แผนผลิต / จ้างทอ", "ผู้วางแผน",
     "บันทึกการแบ่งจ้างทอไว้แล้ว",
     "1. จดจำนวนที่ส่งจ้างทอ + สัปดาห์\n2. สั่งรันแผนใหม่\n3. เปิดแผนดู item นั้นในสัปดาห์ที่กำหนด",
     "• ยอดที่ถักในโรงงานลดลงเท่ากับจำนวนที่จ้างทอ\n• การแบ่งจ้างทอยังคงอยู่หลังรัน (ไม่หายเอง)",
     "การตัดสินใจจ้างทอของผู้ใช้มีผลจริงกับแผนรอบถัดไป"),
    ("ตรวจสอบข้อมูล", "ตรวจค่าคำนวณของ Target Stock", "Master Data → Target_Stock", "ผู้วางแผน",
     "-",
     "1. แก้ TARGET/MONTH ของ 1 แถวเป็น 4000\n2. บันทึก\n3. เปิดไฟล์จริงดูคอลัมน์ที่คำนวณ",
     "• TARGET SCM = 2000 · STOCK MIN = 1000 · STOCK MAX = 3000 · Stock 5 Week = 5000\n• แถว Team=RTS + Type1=CG → STOCK MAX = 4000\n• ค่าที่เก็บเป็นตัวเลขจริง ไม่ใช่สูตร",
     "สูตร Target Stock คำนวณถูกต้องและ pipeline อ่านค่าไปใช้ได้"),
    ("ตรวจสอบข้อมูล", "ตรวจไฟล์แผนที่ดาวน์โหลดครบถ้วน", "แผนผลิต / History", "ผู้วางแผน",
     "-",
     "1. ดาวน์โหลดไฟล์แผนล่าสุด\n2. เปิดดูรายชื่อชีททั้งหมด\n3. สุ่มเทียบตัวเลข 10 จุดกับหน้าจอ",
     "• ชีทครบตามที่หน้าเว็บแสดง\n• ตัวเลขตรงกันทุกจุดที่สุ่ม\n• ไฟล์เปิดได้ไม่เสียหาย",
     "ไฟล์ที่ส่งต่อให้ฝ่ายผลิตตรงกับสิ่งที่ผู้วางแผนเห็นบนหน้าจอ"),
    ("ตรวจสอบข้อมูล", "เทียบแผนใหม่กับแผนรอบก่อน", "History", "ผู้วางแผน",
     "มีไฟล์แผนอย่างน้อย 2 รอบ",
     "1. ดาวน์โหลดแผนรอบก่อนจากหน้า History\n2. เทียบกับแผนล่าสุด\n3. หาจุดที่ต่างกันแล้วหาสาเหตุ",
     "• ความต่างอธิบายได้ด้วยข้อมูลที่เปลี่ยน (booking ใหม่ / stock / วันทำงาน / จ้างทอ)\n• ไม่มีความต่างที่หาสาเหตุไม่ได้",
     "แผนเปลี่ยนแปลงอย่างมีเหตุผล ตรวจสอบย้อนหลังได้"),
    ("ตรวจสอบข้อมูล", "รันซ้ำด้วยข้อมูลเดิมต้องได้ผลเหมือนเดิม", "หน้าหลัก / แผนผลิต", "ผู้วางแผน",
     "ไม่มีการเปลี่ยนข้อมูลต้นทางระหว่างสองรอบ",
     "1. รันแผน (mode=plan) เก็บไฟล์ผลลัพธ์ไว้\n2. รันแผนซ้ำอีกครั้งทันที\n3. เทียบไฟล์สองรอบ",
     "• ผลลัพธ์เหมือนกัน (item เดียวกันอยู่สัปดาห์เดียวกัน)\n• ถ้าต่าง ต้องอธิบายสาเหตุได้ชัดเจน",
     "แผนคาดเดาได้ ไม่เปลี่ยนเองเมื่อข้อมูลไม่เปลี่ยน"),

    # ---------- D. งานสี (Order Color) ----------
    ("งานสี (Order Color)", "ดูภาพรวมงานสีตาม CAT", "Order Color", "ผู้วางแผน",
     "มีไฟล์ Order Color ล่าสุด",
     "1. เปิดเมนู Order Color\n2. ดูกลุ่ม CAT × เกจ ที่แสดง",
     "• แสดงเฉพาะ CAT ที่มีงานสี (LOAD_DYE)\n• เริ่มนับจากสัปดาห์ปัจจุบันไปข้างหน้า\n• งานสีมีเครื่องหมาย ★ พร้อมสัปดาห์ย้อมและกำหนดถัก",
     "ผู้ใช้เห็นทันทีว่ากลุ่มไหนมีงานสีที่ต้องจัดการ"),
    ("งานสี (Order Color)", "ดูงานสีที่ต้องดึงเข้ามาให้เร็วขึ้น", "Order Color → Gantt", "ผู้วางแผน",
     "เลือก CAT × เกจ ที่มีงานสี",
     "1. เลือกกลุ่ม\n2. อ่านหัวข้อสรุปจำนวนงานสี",
     "• หัวข้อระบุ \"งานสีใน <CAT> — N ตัว (ดึงเข้ามาเร็วขึ้นได้ X, จัดแล้ว Y)\"\n• จำนวนตรงกับรายการที่แสดงด้านล่าง",
     "ผู้ใช้รู้ปริมาณงานที่ต้องจัดในกลุ่มนั้นก่อนลงมือ"),
    ("งานสี (Order Color)", "ย้ายงานสีเข้ามาให้เร็วขึ้น (เครื่องว่าง)", "Order Color → Gantt", "ผู้วางแผน",
     "มีสัปดาห์ที่เครื่องว่าง",
     "1. เลือกงานสีที่ทออยู่สัปดาห์ไกล\n2. ย้ายเข้าสัปดาห์ที่มีเครื่องว่าง\n3. ดูตัวเลขเครื่องว่าง/จำนวน job",
     "• เครื่องว่างของสัปดาห์ปลายทางลดลง ต้นทางเพิ่มขึ้น\n• จำนวน job ไม่เกินโควตา setup ของสัปดาห์นั้น\n• สถานะงานสีเปลี่ยนเป็น \"จัดแล้ว\"",
     "ย้ายงานสีเข้ามาได้เมื่อมีเครื่องว่าง โดยระบบคุมโควตา job ให้"),
    ("งานสี (Order Color)", "ถอดงานไม่มีสีเพื่อให้งานสีเข้าได้", "Order Color → Gantt", "ผู้วางแผน",
     "เลือกกลุ่มที่เครื่อง/job ไม่พอ",
     "1. เลือกงานสีที่ยังไม่พอเครื่อง\n2. ดูรายการงานไม่มีสีที่ระบบแนะนำให้ถอด (⭐)\n3. ถอดตัวที่แนะนำ",
     "• ระบบแนะนำเฉพาะงานไม่มีสีที่ถอดแล้วยังหาที่วางใหม่ได้\n• เมื่อถอดแล้ว งานสีวางได้จริง\n• งานที่ถอดถูกวางแผนใหม่ให้อัตโนมัติ",
     "ระบบช่วยหาทางให้งานสีได้ทอทันย้อม โดยไม่ทิ้งงานที่ถอดออกไป"),
    ("งานสี (Order Color)", "ขอคำแนะนำจาก AI", "Order Color", "ผู้วางแผน",
     "เลือกกลุ่มที่มีงานสีไม่พอเครื่อง",
     "1. กดปุ่มให้ AI วิเคราะห์\n2. รอผล\n3. อ่านลำดับและเหตุผล\n4. กด \"ยกเลิก\" แล้วลองใหม่และกดใช้",
     "• ระหว่างรอมีข้อความ \"กำลังให้ AI วิเคราะห์...\"\n• ได้ลำดับควรทำก่อน-หลัง พร้อมเหตุผลภาษาไทยที่เข้าใจได้\n• กดยกเลิก = แผนไม่เปลี่ยน\n• กดใช้ = แผนปรับตามที่แนะนำ",
     "AI ช่วยจัดลำดับได้ และผู้ใช้ยังตัดสินใจเองได้ทุกขั้น"),
    ("งานสี (Order Color)", "ดูตารางเทียบแผนเดิม → แผนใหม่", "Order Color", "ผู้วางแผน",
     "ปรับงานสีไว้แล้ว",
     "1. เปิดตารางเทียบแผน\n2. ดูแถว \"เดิม\" และ \"ใหม่\" ของ item ที่ปรับ",
     "• แสดงเฉพาะ item ที่มีการปรับ\n• แถว \"เดิม\" = ตำแหน่งที่ย้ายออก · \"ใหม่\" = ตำแหน่งใหม่\n• ยอดรวมของแต่ละ item เท่าเดิมก่อน-หลังปรับ",
     "ผู้ใช้ตรวจได้ว่าปรับอะไรไปบ้าง ก่อนส่งต่อ"),
    ("งานสี (Order Color)", "ส่งออกแผนงานสีเป็น Excel", "Order Color", "ผู้วางแผน",
     "จัดแผนไว้แล้ว",
     "1. กดส่งออก Excel\n2. เปิดไฟล์ order_color_plan.xlsx\n3. เทียบกับหน้าจอ",
     "• ไฟล์มีชีท PLAN_ORDER_COLOR\n• หัวคอลัมน์และข้อมูลตรงกับหน้าจอทุกแถว",
     "ผู้ใช้เอาแผน what-if ไปใช้ต่อ/ส่งให้ทีมอื่นได้"),
    ("งานสี (Order Color)", "แผน what-if ต้องไม่แตะแผนจริง", "Order Color / แผนผลิต", "ผู้วางแผน",
     "-",
     "1. ปรับงานสีบนหน้า Order Color หลายจุด\n2. ไปเปิดหน้าแผนผลิต\n3. ดาวน์โหลดไฟล์แผนจริงมาตรวจ",
     "• แผนจริงและไฟล์แผนไม่เปลี่ยนตามการปรับใน Order Color\n• เวลาแก้ไขไฟล์แผนไม่เปลี่ยน",
     "การทดลองจัดแผนงานสีไม่ทำให้แผนจริงเสียหาย"),

    # ---------- E. Master Data ----------
    ("Master Data", "แก้ Lock_MC (เครื่องที่ล็อกไว้)", "Master Data → Lock_MC", "ผู้วางแผน",
     "สำรอง MasterMC.xlsx",
     "1. เปิดชีท Lock_MC (สลับเป็นมุมมองเฉพาะ)\n2. ล็อกเครื่องหนึ่งให้ item หนึ่ง\n3. บันทึก\n4. รันแผนใหม่ แล้วดูเครื่องนั้น",
     "• มุมมองเฉพาะแสดงข้อมูลอ่านง่ายกว่าตารางดิบ\n• หลังรันแผน เครื่องที่ล็อกไม่ถูกเอาไปใช้กับงานอื่น\n• ความอุ่นของเครื่อง (carry) ไม่หายไปเพราะการล็อก",
     "การล็อกเครื่องมีผลจริงกับการจัดแผน"),
    ("Master Data", "แก้ปฏิทินวันหยุด", "Master Data → ปฏิทิน", "ผู้วางแผน",
     "สำรอง Calendar.xlsx",
     "1. เปิดมุมมองปฏิทิน\n2. เปลี่ยนวันหนึ่งจากวันทำงานเป็นวันหยุด\n3. บันทึก\n4. กลับไปหน้าแผน รีเฟรช ดูวันทำงานของสัปดาห์นั้น",
     "• วันทำงานของสัปดาห์นั้นลดลง 1 วันทันที (ไม่ต้องรันแผนใหม่)\n• ค่าที่ Gantt ใช้คำนวณเปลี่ยนตาม",
     "แก้ปฏิทินแล้วทั้งระบบเห็นวันหยุดชุดเดียวกันทันที"),
    ("Master Data", "แก้เครื่อง POLY / COTTON ที่กันไว้", "Master Data → Master MC", "ผู้วางแผน",
     "สำรอง MasterMC.xlsx",
     "1. เปิดมุมมอง Master MC\n2. แก้จำนวนเครื่อง POLY หรือ COTTON ของกลุ่มหนึ่ง\n3. บันทึก แล้วรันแผนใหม่\n4. ดูเครื่องกันไว้บน Gantt",
     "• จำนวนเครื่องกันไว้บนหน้าจอเปลี่ยนตามค่าที่ตั้ง\n• งานปกติยังใช้เครื่องกันไว้ไม่ได้",
     "ผู้ใช้ปรับนโยบายกันเครื่องได้เองผ่านหน้าเว็บ"),
    ("Master Data", "แก้ข้อมูลจ้างทอ S9", "Master Data → S9", "ผู้วางแผน",
     "สำรอง MasterMC.xlsx",
     "1. เปิดมุมมอง S9\n2. เปลี่ยนสถานะ item หนึ่ง (เช่น ให้จ้างทอได้)\n3. บันทึก แล้วรันแผนใหม่",
     "• หลังรันแผน การจัดงานของ item นั้นเปลี่ยนตามที่ตั้ง\n• item ที่กำหนดให้จ้างทออย่างเดียวไม่ถูกวางบนเครื่องในโรงงาน",
     "การตั้งค่า S9 มีผลกับการจัดสรรงานจริง"),
    ("Master Data", "แก้ตาราง แล้วย้อนกลับ (Undo)", "Master Data", "ผู้วางแผน",
     "-",
     "1. แก้ค่าหลายเซลล์\n2. กด Ctrl+Z หรือปุ่ม \"↩ ย้อนกลับ\" หลายครั้ง\n3. กด \"คืนค่าเดิม\" แล้วยืนยัน",
     "• ย้อนกลับได้ทีละขั้นตามที่แก้\n• \"คืนค่าเดิม\" ถามยืนยันก่อน แล้วโหลดค่าจากไฟล์ใหม่ทั้งหมด\n• ไฟล์จริงไม่ถูกแตะจนกว่าจะกดบันทึก",
     "ผู้ใช้แก้ผิดแล้วกู้คืนได้เอง ไม่ต้องเรียกผู้ดูแลระบบ"),
    ("Master Data", "เพิ่ม / ลบ / เปลี่ยนชื่อคอลัมน์", "Master Data", "ผู้วางแผน",
     "สำรองไฟล์ไว้ก่อน (มีผลกับ pipeline)",
     "1. กด ⋮ ที่หัวคอลัมน์ → เปลี่ยนชื่อ\n2. เพิ่มคอลัมน์ใหม่\n3. ลบคอลัมน์ทดสอบ แล้วอ่านคำเตือน\n4. บันทึก",
     "• ลบคอลัมน์มีคำเตือนว่าข้อมูลทั้งคอลัมน์จะหาย และบอกว่ากดย้อนกลับได้\n• หลังบันทึก โครงสร้างในไฟล์เปลี่ยนตาม\n• ตรวจว่าการรันแผนยังทำงานได้ (คอลัมน์ที่ pipeline ใช้ต้องไม่ถูกลบ)",
     "แก้โครงสร้างตารางได้ พร้อมคำเตือนที่ชัดเจนก่อนทำสิ่งที่ย้อนยาก"),
    ("Master Data", "ค้นหา เรียง กรอง ในตาราง Master", "Master Data", "ผู้วางแผน",
     "-",
     "1. พิมพ์คำค้น\n2. คลิกหัวคอลัมน์เพื่อเรียงลำดับ\n3. กด ▾ กรองเฉพาะค่า\n4. ลากขอบคอลัมน์ปรับความกว้าง / double click ให้พอดีเนื้อหา\n5. ล้างทั้งหมด",
     "• จำนวนแถวที่แสดงถูกต้อง (แสดง N / ทั้งหมด M แถว)\n• เรียง/กรองไม่ทำให้ข้อมูลหายเมื่อล้างตัวกรอง\n• การแก้ค่าขณะกรองอยู่ ลงที่แถวที่ถูกต้อง",
     "ทำงานกับตารางใหญ่ได้สะดวกและปลอดภัย"),
    ("Master Data", "บันทึกด้วยคีย์ลัดและตรวจไฟล์สำรอง", "Master Data", "ผู้วางแผน",
     "-",
     "1. แก้ค่า 1 จุด\n2. กด Ctrl+S\n3. เปิดโฟลเดอร์ไฟล์ Master",
     "• บันทึกสำเร็จพร้อมข้อความยืนยัน + ชื่อไฟล์สำรอง\n• มีไฟล์ .bak ที่มี timestamp เพิ่มขึ้น\n• เปิดไฟล์ .bak ได้และเป็นข้อมูลก่อนแก้",
     "ทุกการบันทึกมีจุดย้อนกลับเสมอ"),

    # ---------- F. งานประจำวัน (End-to-End) ----------
    ("งานประจำวัน", "รอบงานเช้า: ดึงข้อมูล → รันแผน → ปรับ → ส่งต่อ", "ทุกหน้าจอ", "ผู้วางแผน",
     "เริ่มต้นวันทำงานปกติ",
     "1. เข้าระบบ ตรวจว่าแผนอัตโนมัติรันสำเร็จ\n2. ถ้ายังไม่รัน กด \"ดึงข้อมูลล่าสุด\" แล้วรันแผน\n3. เปิดหน้าแผน ตรวจงานที่ผิดปกติ\n4. ปรับแผน (ย้าย/แบ่ง/แก้จำนวน)\n5. บันทึก\n6. ดาวน์โหลดไฟล์ส่งให้ฝ่ายผลิต",
     "• ทุกขั้นตอนทำได้ต่อเนื่องไม่ติดขัด\n• เวลารวมทั้งรอบอยู่ในเกณฑ์ที่ยอมรับได้\n• ไฟล์สุดท้ายมีค่าที่ปรับครบทุกจุด",
     "ผู้ใช้ทำงานประจำวันจบได้ในระบบเดียว ตั้งแต่ดึงข้อมูลจนส่งแผนต่อ"),
    ("งานประจำวัน", "ตรวจผลการรันอัตโนมัติเมื่อวาน/เช้านี้", "หน้าหลัก / History", "ผู้ดูแลระบบ",
     "ตั้งเวลารันอัตโนมัติไว้",
     "1. เปิดหน้าหลัก ดูสถานะรอบล่าสุดและเวลารอบถัดไป\n2. เปิด History ดูไฟล์แผนที่ได้\n3. ตรวจ log ของรอบนั้น",
     "• สถานะบอกว่ารันโดยตารางเวลา (ไม่ใช่สั่งเอง)\n• มีไฟล์แผนของรอบนั้นใน History\n• log ไม่มีข้อผิดพลาดค้าง",
     "ทีมมั่นใจได้ว่าระบบทำงานเองได้ทุกวันโดยไม่ต้องเฝ้า"),
    ("งานประจำวัน", "ใช้งานพร้อมกันหลายคน", "ทุกหน้าจอ", "ผู้วางแผน 2 คน",
     "มี 2 บัญชี",
     "1. คนที่ 1 สั่งรันแผน\n2. คนที่ 2 เปิดดูสถานะและ log\n3. คนที่ 2 ลองสั่งรันซ้ำ\n4. หลังรันจบ คนที่ 1 แก้แผนและบันทึก คนที่ 2 รีเฟรช",
     "• คนที่ 2 เห็นสถานะและ log ของรอบที่กำลังรัน\n• สั่งรันซ้อนไม่ได้ พร้อมข้อความบอกเหตุผล\n• คนที่ 2 รีเฟรชแล้วเห็นแผนที่คนที่ 1 บันทึก\n• ไม่มีข้อมูลของใครหายหรือถูกทับแบบไม่รู้ตัว",
     "ทำงานเป็นทีมได้โดยไม่แย่งกันแก้ไฟล์"),
    ("งานประจำวัน", "ตัดสินใจจ้างทอด้วย AI", "แผนผลิต → จ้างทอ (AI)", "ผู้วางแผน",
     "มีแผนล่าสุด",
     "1. กดปุ่ม \"🧵 จ้างทอ (AI)\"\n2. อ่านรายการที่แนะนำพร้อมเหตุผล\n3. เลือก item ใส่จำนวนและสัปดาห์ แล้วบันทึก\n4. รันแผนใหม่ ตรวจผล",
     "• คำแนะนำมีเหตุผลและตัวเลขประกอบ\n• ใส่จำนวนเกินของค้าง → ระบบเตือนพร้อมบอกยอดค้างจริง\n• หลังรันแผน ยอดถักในโรงงานลดลงตามที่จ้างทอ",
     "ผู้ใช้ตัดสินใจจ้างทอโดยมีข้อมูลรองรับ และผลสะท้อนในแผนจริง"),
    ("งานประจำวัน", "อนุมัติการเปลี่ยน Cylinder", "แผนผลิต → Change Cylinder", "ผู้วางแผน",
     "มีแผนล่าสุด",
     "1. กดปุ่ม \"🔩 Change Cylinder\"\n2. อ่านคำแนะนำและสัปดาห์คอขวด\n3. อนุมัติรายการหนึ่ง\n4. ตรวจผลในแผน",
     "• ระบบไม่เปลี่ยน cylinder เองโดยไม่ขออนุมัติ\n• อนุมัติแล้วมีการบันทึกรายการพร้อมเวลา\n• งานที่เคยติดเครื่องถูกปลดตามที่คาด",
     "การเปลี่ยน cylinder อยู่ในการควบคุมของผู้ใช้เสมอ"),
    ("งานประจำวัน", "จัดการไฟล์ย้อนหลัง", "History", "ผู้ดูแลระบบ",
     "มีไฟล์เก่าหลายรอบ",
     "1. เปิด History ดูไฟล์ทั้ง 3 ประเภท (แผน / booking / SC)\n2. ดาวน์โหลดไฟล์ย้อนหลัง 1 ไฟล์\n3. ลบไฟล์ทดสอบ 1 ไฟล์",
     "• ไฟล์เรียงจากใหม่ไปเก่า แสดงขนาดและเวลาถูกต้อง\n• ดาวน์โหลดย้อนหลังได้และเปิดได้\n• ลบแล้วหายจากรายการทันที ไฟล์อื่นไม่กระทบ",
     "ผู้ดูแลจัดการพื้นที่เก็บไฟล์ได้เอง และย้อนดูแผนเก่าได้"),
]


def _summary_block(ws, start_row: int, title: str, sheet_ref: str, modules: list,
                   mod_col: str, status_col: str) -> int:
    """เขียนตารางสรุปผล 1 บล็อก (นับ Pass/Fail จากชีทที่อ้าง) → คืนแถวถัดไปที่ว่าง"""
    r = start_row
    ws.cell(row=r, column=1, value=title).font = FONT_TITLE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

    head = r
    for i, h in enumerate(["กลุ่มงาน", "จำนวนรายการ", "Pass", "Fail", "Blocked", "N/A", "ยังไม่ทดสอบ"], start=1):
        ws.cell(row=head, column=i, value=h)
    style_header(ws, head, 7)

    r = head + 1
    first = r
    q = f"'{sheet_ref}'"
    for m in modules:
        ws.cell(row=r, column=1, value=m)
        ws.cell(row=r, column=2, value=f'=COUNTIF({q}!${mod_col}:${mod_col},$A{r})')
        for col, st in ((3, "Pass"), (4, "Fail"), (5, "Blocked"), (6, "N/A")):
            ws.cell(row=r, column=col,
                    value=f'=COUNTIFS({q}!${mod_col}:${mod_col},$A{r},{q}!${status_col}:${status_col},"{st}")')
        ws.cell(row=r, column=7, value=f"=B{r}-C{r}-D{r}-E{r}-F{r}")
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.font = FONT_BODY
            if c > 1:
                cell.alignment = CENTER
        r += 1

    last = r - 1
    ws.cell(row=r, column=1, value="รวม").font = FONT_BOLD
    for c in range(2, 8):
        col = get_column_letter(c)
        cell = ws.cell(row=r, column=c, value=f"=SUM({col}{first}:{col}{last})")
        cell.font = FONT_BOLD
        cell.alignment = CENTER
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=C_MOD)
    total = r
    r += 1
    ws.cell(row=r, column=1, value="อัตราผ่าน (Pass Rate)").font = FONT_BOLD
    pr = ws.cell(row=r, column=2, value=f"=IF(B{total}=0,0,C{total}/B{total})")
    pr.number_format = "0.0%"
    pr.font = FONT_BOLD
    pr.alignment = CENTER
    return r + 2


def build():
    wb = Workbook()

    # ================= ชีท 1: ภาพรวม =================
    ws = wb.active
    ws.title = "ภาพรวม"
    set_widths(ws, {"A": 26, "B": 46, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16})

    ws["A1"] = "เอกสารทดสอบการยอมรับระบบ (UAT) — Knit Plan Web Server"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 28

    info = [
        ("ระบบที่ทดสอบ", "Knit Plan Web (FastAPI backend + React frontend) — webapp/backend/server.py"),
        ("ขอบเขต", "ทุกความสามารถที่ server ให้บริการ: 52 API endpoints ใน 12 กลุ่มงาน + การเสิร์ฟหน้าเว็บ/ไฟล์ static "
                   "+ การทำงานเบื้องหลัง (ตั้งเวลารันอัตโนมัติ, ลบไฟล์เก่า, เติมปฏิทินอัตโนมัติตอน start)"),
        ("URL ระบบ (กรอก)", ""),
        ("เวอร์ชัน / commit (กรอก)", ""),
        ("ผู้ทดสอบ (กรอก)", ""),
        ("วันที่ทดสอบ (กรอก)", ""),
        ("สร้างเอกสารเมื่อ", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]
    r = 3
    for k, v in info:
        ws.cell(row=r, column=1, value=k).font = FONT_BOLD
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=C_SUB)
        ws.cell(row=r, column=1).border = BORDER
        c = ws.cell(row=r, column=2, value=v)
        c.font = FONT_BODY
        c.alignment = WRAP_TOP
        c.border = BORDER
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        r += 1

    r += 1

    sc_modules = []
    for s in SCENARIOS:
        if s[0] not in sc_modules:
            sc_modules.append(s[0])
    modules = []
    for t in TESTS:
        if t[0] not in modules:
            modules.append(t[0])

    r = _summary_block(ws, r, "สรุปผล ① การใช้งานจริงของผู้ใช้ (ชีท UAT_งานจริง)",
                       "UAT_งานจริง", sc_modules, "B", "J")
    r = _summary_block(ws, r, "สรุปผล ② รายฟังก์ชันของ server (ชีท UAT_TestCases)",
                       "UAT_TestCases", modules, "B", "M")

    ws.cell(row=r, column=1, value="เกณฑ์การยอมรับ").font = FONT_BOLD
    ws.cell(row=r, column=2, value="ไม่มีข้อบกพร่องระดับ Critical/High ค้าง · สถานการณ์ใช้งานจริงผ่านครบทุกข้อ "
                                   "· Test Case ประเภท Positive ผ่านครบ 100%")
    ws.cell(row=r, column=2).font = FONT_BODY
    ws.cell(row=r, column=2).alignment = WRAP_TOP
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)

    # ================= ชีท 2: UAT_งานจริง (สถานการณ์ผู้ใช้) =================
    wsx = wb.create_sheet("UAT_งานจริง")
    hx = ["SC_ID", "กลุ่มงาน", "สถานการณ์ใช้งานจริง", "หน้าจอที่ใช้", "ผู้ใช้",
          "เงื่อนไขก่อนเริ่ม", "ขั้นตอนที่ผู้ใช้ทำ", "จุดที่ต้องตรวจ (Checkpoint)",
          "ผลลัพธ์ที่ยอมรับได้", "สถานะ", "ผลที่ได้จริง", "ผู้ทดสอบ", "วันที่ทดสอบ", "หมายเหตุ / Defect ID"]
    wsx.append(hx)
    style_header(wsx, 1, len(hx))

    sc_prefix = {
        "ปรับแผนผลิต": "PLAN", "ปรับสัปดาห์ / วันทำงาน": "WEEK", "ตรวจสอบข้อมูล": "CHK",
        "งานสี (Order Color)": "COLOR", "Master Data": "MDATA", "งานประจำวัน": "DAY",
    }
    sc_count = {}
    row = 2
    for (grp, name, screen, who, pre, steps, checks, expected) in SCENARIOS:
        pfx = sc_prefix.get(grp, "SC")
        sc_count[pfx] = sc_count.get(pfx, 0) + 1
        wsx.append([f"{pfx}-{sc_count[pfx]:02d}", grp, name, screen, who, pre, steps, checks,
                    expected, "", "", "", "", ""])
        for c in range(1, len(hx) + 1):
            cell = wsx.cell(row=row, column=c)
            cell.font = FONT_BODY
            cell.border = BORDER
            cell.alignment = CENTER if c in (1, 5, 10, 13) else WRAP_TOP
        wsx.row_dimensions[row].height = 96
        row += 1

    sc_last = row - 1
    set_widths(wsx, {"A": 11, "B": 20, "C": 34, "D": 22, "E": 14, "F": 26,
                     "G": 46, "H": 56, "I": 44, "J": 12, "K": 32, "L": 12, "M": 13, "N": 20})
    wsx.freeze_panes = "C2"
    wsx.auto_filter.ref = f"A1:N{sc_last}"

    dvx = DataValidation(type="list", formula1='"Pass,Fail,Blocked,N/A"', allow_blank=True, showDropDown=False)
    dvx.error = "เลือกได้เฉพาะ Pass / Fail / Blocked / N/A"
    wsx.add_data_validation(dvx)
    dvx.add(f"J2:J{sc_last}")
    for st, color in (("Pass", "C6EFCE"), ("Fail", "FFC7CE"), ("Blocked", "FFEB9C"), ("N/A", "E7E6E6")):
        wsx.conditional_formatting.add(
            f"J2:J{sc_last}",
            CellIsRule(operator="equal", formula=[f'"{st}"'], fill=PatternFill("solid", fgColor=color)))

    # ================= ชีท 3: UAT_TestCases =================
    ws2 = wb.create_sheet("UAT_TestCases")
    headers = ["TC_ID", "โมดูล", "ฟีเจอร์ / สิ่งที่ทดสอบ", "Method", "Endpoint", "สิทธิ์ที่ใช้",
               "ประเภท", "ผลกระทบข้อมูล", "เงื่อนไขก่อนทดสอบ", "ขั้นตอนการทดสอบ", "ข้อมูลทดสอบ",
               "ผลลัพธ์ที่คาดหวัง", "สถานะ", "ผลที่ได้จริง", "ผู้ทดสอบ", "วันที่ทดสอบ", "หมายเหตุ / Defect ID"]
    ws2.append(headers)
    style_header(ws2, 1, len(headers))

    prefix = {
        "เข้าสู่ระบบ (Auth)": "AUTH", "สั่งรัน Pipeline": "RUN", "ตั้งเวลาอัตโนมัติ": "SCH",
        "Master Data": "MST", "วันทำงาน (Work Day)": "WD", "ไฟล์ผลลัพธ์ (History)": "OUT",
        "ฐานข้อมูล (DATA)": "DB", "Map Item": "MAP", "แผนผลิต (Plan)": "PLN",
        "Order Color": "OC", "จ้างทอ (Outsource)": "OS", "Change Cylinder": "CYL",
        "ระบบ (System)": "SYS",
    }
    counters = {}
    row = 2
    for (mod, feat, method, endpoint, role, ptype, risk, pre, steps, data, expected) in TESTS:
        pfx = prefix.get(mod, "TC")
        counters[pfx] = counters.get(pfx, 0) + 1
        tc_id = f"{pfx}-{counters[pfx]:02d}"
        ws2.append([tc_id, mod, feat, method, endpoint, role, ptype, risk, pre, steps, data,
                    expected, "", "", "", "", ""])
        for c in range(1, len(headers) + 1):
            cell = ws2.cell(row=row, column=c)
            cell.font = FONT_BODY
            cell.border = BORDER
            cell.alignment = CENTER if c in (1, 4, 6, 7, 8, 13, 16) else WRAP_TOP
        row += 1

    last = row - 1
    set_widths(ws2, {"A": 10, "B": 20, "C": 34, "D": 9, "E": 30, "F": 15, "G": 13, "H": 15,
                     "I": 30, "J": 40, "K": 24, "L": 52, "M": 12, "N": 30, "O": 12, "P": 13, "Q": 22})
    ws2.freeze_panes = "C2"
    ws2.auto_filter.ref = f"A1:Q{last}"

    dv = DataValidation(type="list", formula1='"Pass,Fail,Blocked,N/A"', allow_blank=True, showDropDown=False)
    dv.error = "เลือกได้เฉพาะ Pass / Fail / Blocked / N/A"
    ws2.add_data_validation(dv)
    dv.add(f"M2:M{last}")

    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    grey = PatternFill("solid", fgColor="E7E6E6")
    ws2.conditional_formatting.add(f"M2:M{last}",
                                   CellIsRule(operator="equal", formula=['"Pass"'], fill=green))
    ws2.conditional_formatting.add(f"M2:M{last}",
                                   CellIsRule(operator="equal", formula=['"Fail"'], fill=red))
    ws2.conditional_formatting.add(f"M2:M{last}",
                                   CellIsRule(operator="equal", formula=['"Blocked"'], fill=yellow))
    ws2.conditional_formatting.add(f"M2:M{last}",
                                   CellIsRule(operator="equal", formula=['"N/A"'], fill=grey))

    # ================= ชีท 3: รายการ_API =================
    ws3 = wb.create_sheet("รายการ_API")
    h3 = ["ลำดับ", "กลุ่มงาน", "Method", "Endpoint", "สิทธิ์", "หน้าที่", "พารามิเตอร์ / Body"]
    ws3.append(h3)
    style_header(ws3, 1, len(h3))
    for i, (mod, method, ep, role, desc, params) in enumerate(APIS, start=1):
        ws3.append([i, mod, method, ep, role, desc, params])
        for c in range(1, len(h3) + 1):
            cell = ws3.cell(row=i + 1, column=c)
            cell.font = FONT_BODY
            cell.border = BORDER
            cell.alignment = CENTER if c in (1, 3) else WRAP_TOP
    set_widths(ws3, {"A": 7, "B": 24, "C": 9, "D": 36, "E": 24, "F": 62, "G": 44})
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:G{len(APIS) + 1}"

    # ================= ชีท 4: Defect_Log =================
    ws4 = wb.create_sheet("Defect_Log")
    h4 = ["Defect ID", "TC_ID ที่เกี่ยวข้อง", "โมดูล", "หัวข้อปัญหา", "รายละเอียด / ขั้นตอนทำให้เกิดซ้ำ",
          "ความรุนแรง", "ผู้แจ้ง", "วันที่แจ้ง", "ผู้รับผิดชอบ", "สถานะ", "วันที่แก้เสร็จ", "ผลการทดสอบซ้ำ"]
    ws4.append(h4)
    style_header(ws4, 1, len(h4))
    for r in range(2, 42):
        for c in range(1, len(h4) + 1):
            cell = ws4.cell(row=r, column=c)
            cell.border = BORDER
            cell.font = FONT_BODY
            cell.alignment = WRAP_TOP
    set_widths(ws4, {"A": 12, "B": 16, "C": 20, "D": 34, "E": 52, "F": 13,
                     "G": 14, "H": 13, "I": 14, "J": 13, "K": 13, "L": 22})
    ws4.freeze_panes = "A2"

    dv_sev = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True, showDropDown=False)
    ws4.add_data_validation(dv_sev)
    dv_sev.add("F2:F41")
    dv_st = DataValidation(type="list", formula1='"Open,In Progress,Fixed,Retest,Closed,Rejected"',
                           allow_blank=True, showDropDown=False)
    ws4.add_data_validation(dv_st)
    dv_st.add("J2:J41")

    # ================= ชีท 5: คู่มือใช้งาน =================
    ws5 = wb.create_sheet("คู่มือใช้งาน")
    set_widths(ws5, {"A": 4, "B": 110})
    ws5["B1"] = "วิธีใช้เอกสาร UAT ฉบับนี้"
    ws5["B1"].font = FONT_TITLE
    guide = [
        "",
        "1) เตรียมก่อนทดสอบ",
        "   • ระบุ URL ระบบ / เวอร์ชัน / ผู้ทดสอบ / วันที่ ในชีท \"ภาพรวม\"",
        "   • เตรียมบัญชีทดสอบ 2 บัญชี: 1 บัญชี role = admin และ 1 บัญชี user ธรรมดา (สร้างด้วย webapp/backend/add_user.py)",
        "   • สำรองไฟล์สำคัญก่อนทดสอบกลุ่มที่มีการเขียนไฟล์: MasterMC.xlsx, Calendar.xlsx, Target_Stock.xlsx,",
        "     Master_Item.xlsx และโฟลเดอร์ data_plan (แม้ระบบจะสร้าง .bak ให้อัตโนมัติทุกครั้งที่บันทึกก็ตาม)",
        "   • ปิดไฟล์ Excel ที่เปิดค้างบนเครื่อง server ก่อนทดสอบการบันทึก (ยกเว้น Test Case ที่ตั้งใจทดสอบกรณีนี้)",
        "",
        "2) เอกสารมี 2 ชุดการทดสอบ — ทดสอบชุดแรกก่อน",
        "   • ชีท \"UAT_งานจริง\"    = สถานการณ์ที่ผู้ใช้ทำจริงบนหน้าจอ (ปรับแผน · ปรับสัปดาห์ · ตรวจข้อมูล · งานสี ·",
        "     Master Data · งานประจำวัน) — ชุดหลักที่ผู้ใช้งานจริงต้องทดสอบ ทุกข้อบอกว่าต้องกดอะไรและตรวจอะไร",
        "   • ชีท \"UAT_TestCases\" = ทดสอบรายฟังก์ชันของ server ทีละตัว รวมกรณีผิดพลาดและสิทธิ์การเข้าถึง",
        "     (เหมาะกับ IT / QA ใช้ตรวจให้ครบทุกมุม)",
        "",
        "3) วิธีกรอกผล (ทั้ง 2 ชีท)",
        "   • คอลัมน์ \"สถานะ\" เลือกจาก dropdown: Pass / Fail / Blocked / N/A (สีจะเปลี่ยนอัตโนมัติ)",
        "   • Fail หรือ Blocked ให้กรอก \"ผลที่ได้จริง\" ให้ละเอียด แล้วเปิด Defect ในชีท Defect_Log พร้อมอ้าง ID",
        "   • ชีท \"ภาพรวม\" สรุปจำนวน Pass/Fail และอัตราผ่านของทั้ง 2 ชุดให้อัตโนมัติ ไม่ต้องกรอกเอง",
        "   • ชีท \"UAT_งานจริง\" คอลัมน์ \"จุดที่ต้องตรวจ\" คือรายการที่ต้องดูให้ครบทุกบรรทัดก่อนตัดสินว่า Pass",
        "",
        "4) ความหมายของคอลัมน์ \"ประเภท\" (ชีท UAT_TestCases)",
        "   • Positive       = ใช้งานตามปกติ ต้องทำงานได้ถูกต้อง",
        "   • Negative       = กรณีผิดพลาด/ข้อมูลไม่ครบ ต้องขึ้นข้อความที่ถูกต้องและไม่ทำให้ระบบพัง",
        "   • Security       = สิทธิ์การเข้าถึงและความปลอดภัย (token, สิทธิ์ admin, path traversal)",
        "   • Non-functional = ประสิทธิภาพ ความเสถียร การใช้งานหลายคนพร้อมกัน",
        "",
        "5) ความหมายของคอลัมน์ \"ผลกระทบข้อมูล\"",
        "   • อ่านอย่างเดียว     = ปลอดภัย ไม่แก้ไขข้อมูลใดๆ",
        "   • เขียน/แก้ไขไฟล์  = มีการเขียนทับไฟล์จริง (ระบบสร้าง .bak ให้ก่อนเสมอ) — ควรทำในช่วงที่ไม่มีคนใช้งาน",
        "   • รันงานหนัก       = สั่งรัน pipeline จริง ใช้เวลานานและเชื่อมต่อฐานข้อมูล — สั่งได้ทีละ 1 งานเท่านั้น",
        "",
        "6) ลำดับการทดสอบที่แนะนำ",
        "   ชุดที่ 1 (UAT_งานจริง): CHK ตรวจข้อมูล → PLAN ปรับแผน → WEEK ปรับสัปดาห์ → COLOR งานสี →",
        "     MDATA แก้ Master → DAY งานประจำวันแบบครบวงจร",
        "   ชุดที่ 2 (UAT_TestCases): AUTH → SYS → DB/MAP/PLN/OC (อ่านอย่างเดียว) → MST/WD/OUT/OS (เขียนไฟล์)",
        "     → RUN/SCH (รันจริง)",
        "   เหตุผล: ตรวจว่าข้อมูลถูกต้องก่อน แล้วค่อยทดสอบการปรับแก้ ปิดท้ายด้วยงานที่ใช้เวลานาน",
        "",
        "7) เกณฑ์ยอมรับระบบ (แนะนำ)",
        "   • สถานการณ์ในชีท UAT_งานจริง ผ่านครบทุกข้อ (ผู้ใช้ทำงานประจำวันได้จบในระบบ)",
        "   • Test Case ประเภท Positive ผ่านครบ 100%",
        "   • ไม่มี Defect ระดับ Critical / High ค้างในสถานะ Open",
        "   • Defect ระดับ Medium/Low ที่ยังเหลือ ต้องได้รับการยอมรับจากผู้ใช้งานเป็นลายลักษณ์อักษร",
    ]
    r = 2
    for line in guide:
        cell = ws5.cell(row=r, column=2, value=line)
        cell.font = FONT_BOLD if (line and not line.startswith(" ") and line[0].isdigit()) else FONT_BODY
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    wb.save(OUT)
    return len(SCENARIOS), len(TESTS), len(APIS), OUT


if __name__ == "__main__":
    n_sc, n_tc, n_api, path = build()
    print(f"สร้างเอกสาร UAT เสร็จ: {path}")
    print(f"  - สถานการณ์ใช้งานจริง: {n_sc} รายการ")
    print(f"  - Test Case รายฟังก์ชัน: {n_tc} รายการ")
    print(f"  - API ในทะเบียน: {n_api} รายการ")
